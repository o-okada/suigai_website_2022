#!/usr/bin/env python3
# -*- coding: utf-8 -*-
###############################################################################
### ファイル名：P0200ExcelDownload/views.py
### EXCELダウンロード
###############################################################################

###############################################################################
### 処理名：インポート処理
###############################################################################
import sys
from datetime import date, datetime, timedelta, timezone
from django.contrib.auth.decorators import login_required
from django.db import connection
from django.db import transaction
from django.db.models import Max
from django.http import Http404
from django.http import HttpResponse
from django.http import HttpResponseRedirect
from django.shortcuts import render
from django.template import loader
from django.views import generic
from django.views.generic import FormView
from django.views.generic.base import TemplateView

from django.shortcuts import redirect

import openpyxl
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.writer.excel import save_virtual_workbook
from openpyxl.styles import PatternFill
from openpyxl.formatting.rule import FormulaRule

import hashlib
import os
import time
import glob

from P0000Common.models import BUILDING                ### 1000: 建物区分
from P0000Common.models import KEN                     ### 1010: 都道府県
from P0000Common.models import CITY                    ### 1020: 市区町村
from P0000Common.models import KASEN_KAIGAN            ### 1030: 水害発生地点工種（河川海岸区分）
from P0000Common.models import SUIKEI                  ### 1040: 水系（水系・沿岸）
from P0000Common.models import SUIKEI_TYPE             ### 1050: 水系種別（水系・沿岸種別）
from P0000Common.models import KASEN                   ### 1060: 河川（河川・海岸）
from P0000Common.models import KASEN_TYPE              ### 1070: 河川種別（河川・海岸種別）
from P0000Common.models import CAUSE                   ### 1080: 水害原因
from P0000Common.models import UNDERGROUND             ### 1090: 地上地下区分
from P0000Common.models import USAGE                   ### 1100: 地下空間の利用形態
from P0000Common.models import FLOOD_SEDIMENT          ### 1110: 浸水土砂区分
from P0000Common.models import GRADIENT                ### 1120: 地盤勾配区分
from P0000Common.models import INDUSTRY                ### 1130: 産業分類

from P0000Common.models import HOUSE_ASSET             ### 2000: 家屋評価額
from P0000Common.models import HOUSE_RATE              ### 2010: 家屋被害率
from P0000Common.models import HOUSE_ALT               ### 2020: 家庭応急対策費_代替活動費
from P0000Common.models import HOUSE_CLEAN             ### 2030: 家庭応急対策費_清掃日数、清掃労働単価

from P0000Common.models import HOUSEHOLD_ASSET         ### 3000: 家庭用品自動車以外所有額
from P0000Common.models import HOUSEHOLD_RATE          ### 3010: 家庭用品自動車以外被害率

from P0000Common.models import CAR_ASSET               ### 4000: 家庭用品自動車所有額
from P0000Common.models import CAR_RATE                ### 4010: 家庭用品自動車被害率

from P0000Common.models import OFFICE_ASSET            ### 5000: 事業所資産額
from P0000Common.models import OFFICE_RATE             ### 5010: 事業所被害率
from P0000Common.models import OFFICE_SUSPEND          ### 5020: 事業所営業停止日数
from P0000Common.models import OFFICE_STAGNATE         ### 5030: 事業所営業停滞日数
from P0000Common.models import OFFICE_ALT              ### 5040: 事業所応急対策費_代替活動費

from P0000Common.models import FARMER_FISHER_ASSET     ### 6000: 農漁家資産額
from P0000Common.models import FARMER_FISHER_RATE      ### 6010: 農漁家被害率

from P0000Common.models import AREA                    ### 7000: 入力データ_水害区域
from P0000Common.models import WEATHER                 ### 7010: 入力データ_異常気象
from P0000Common.models import SUIGAI                  ### 7020: 入力データ_ヘッダ部分
from P0000Common.models import IPPAN                   ### 7030: 入力データ_一覧表部分
from P0000Common.models import IPPAN_VIEW              ### 7040: ビューデータ_一覧表部分

from P0000Common.models import IPPAN_SUMMARY           ### 8000: 集計データ_集計結果

from P0000Common.common import get_debug_log
from P0000Common.common import get_error_log
from P0000Common.common import get_info_log
from P0000Common.common import get_warn_log
from P0000Common.common import print_log
from P0000Common.common import reset_log

###############################################################################
### 関数名：index_view(request)
### urlpattern：path('', views.index_view, name='index_view')
### urlpattern：path('data_type/<slug:data_type>', views.index_view, name='index_view')
### template：P0200ExcelDownload/index.html
###############################################################################
@login_required(None, login_url='/P0100Login/')
def index_view(request, data_type):
    try:
        #######################################################################
        ### 引数チェック処理(0000)
        ### ブラウザからのリクエストと引数をチェックする。
        #######################################################################
        ### reset_log()
        print_log('[INFO] P0200ExcelDownload.index_view()関数が開始しました。', 'INFO')
        print_log('[DEBUG] P0200ExcelDownload.index_view()関数 request = {}'.format(request.method), 'DEBUG')
        print_log('[DEBUG] P0200ExcelDownload.index_view()関数 data_type = {}'.format(data_type), 'DEBUG')
        print_log('[DEBUG] P0200ExcelDownload.index_view()関数 STEP 1/4.', 'DEBUG')
        
        #######################################################################
        ### DBアクセス処理(0010)
        ### (1)DBにアクセスして、都道府県データを取得する。
        ### (2)DBにアクセスして、市区町村データを取得する。
        #######################################################################
        print_log('[DEBUG] P0200ExcelDownload.index_view()関数 STEP 2/4.', 'DEBUG')
        ken_list = KEN.objects.raw("""SELECT * FROM KEN ORDER BY CAST(KEN_CODE AS INTEGER)""", [])
        city_list01 = CITY.objects.raw("""SELECT * FROM CITY WHERE KEN_CODE=%s ORDER BY CAST(CITY_CODE AS INTEGER)""", ['01', ])
        city_list02 = CITY.objects.raw("""SELECT * FROM CITY WHERE KEN_CODE=%s ORDER BY CAST(CITY_CODE AS INTEGER)""", ['02', ])
        city_list03 = CITY.objects.raw("""SELECT * FROM CITY WHERE KEN_CODE=%s ORDER BY CAST(CITY_CODE AS INTEGER)""", ['03', ])
        city_list04 = CITY.objects.raw("""SELECT * FROM CITY WHERE KEN_CODE=%s ORDER BY CAST(CITY_CODE AS INTEGER)""", ['04', ])
        city_list05 = CITY.objects.raw("""SELECT * FROM CITY WHERE KEN_CODE=%s ORDER BY CAST(CITY_CODE AS INTEGER)""", ['05', ])
        city_list06 = CITY.objects.raw("""SELECT * FROM CITY WHERE KEN_CODE=%s ORDER BY CAST(CITY_CODE AS INTEGER)""", ['06', ])
        city_list07 = CITY.objects.raw("""SELECT * FROM CITY WHERE KEN_CODE=%s ORDER BY CAST(CITY_CODE AS INTEGER)""", ['07', ])
        city_list08 = CITY.objects.raw("""SELECT * FROM CITY WHERE KEN_CODE=%s ORDER BY CAST(CITY_CODE AS INTEGER)""", ['08', ])
        city_list09 = CITY.objects.raw("""SELECT * FROM CITY WHERE KEN_CODE=%s ORDER BY CAST(CITY_CODE AS INTEGER)""", ['09', ])
        city_list10 = CITY.objects.raw("""SELECT * FROM CITY WHERE KEN_CODE=%s ORDER BY CAST(CITY_CODE AS INTEGER)""", ['10', ])
        city_list11 = CITY.objects.raw("""SELECT * FROM CITY WHERE KEN_CODE=%s ORDER BY CAST(CITY_CODE AS INTEGER)""", ['11', ])
        city_list12 = CITY.objects.raw("""SELECT * FROM CITY WHERE KEN_CODE=%s ORDER BY CAST(CITY_CODE AS INTEGER)""", ['12', ])
        city_list13 = CITY.objects.raw("""SELECT * FROM CITY WHERE KEN_CODE=%s ORDER BY CAST(CITY_CODE AS INTEGER)""", ['13', ])
        city_list14 = CITY.objects.raw("""SELECT * FROM CITY WHERE KEN_CODE=%s ORDER BY CAST(CITY_CODE AS INTEGER)""", ['14', ])
        city_list15 = CITY.objects.raw("""SELECT * FROM CITY WHERE KEN_CODE=%s ORDER BY CAST(CITY_CODE AS INTEGER)""", ['15', ])
        city_list16 = CITY.objects.raw("""SELECT * FROM CITY WHERE KEN_CODE=%s ORDER BY CAST(CITY_CODE AS INTEGER)""", ['16', ])
        city_list17 = CITY.objects.raw("""SELECT * FROM CITY WHERE KEN_CODE=%s ORDER BY CAST(CITY_CODE AS INTEGER)""", ['17', ])
        city_list18 = CITY.objects.raw("""SELECT * FROM CITY WHERE KEN_CODE=%s ORDER BY CAST(CITY_CODE AS INTEGER)""", ['18', ])
        city_list19 = CITY.objects.raw("""SELECT * FROM CITY WHERE KEN_CODE=%s ORDER BY CAST(CITY_CODE AS INTEGER)""", ['19', ])
        city_list20 = CITY.objects.raw("""SELECT * FROM CITY WHERE KEN_CODE=%s ORDER BY CAST(CITY_CODE AS INTEGER)""", ['20', ])
        city_list21 = CITY.objects.raw("""SELECT * FROM CITY WHERE KEN_CODE=%s ORDER BY CAST(CITY_CODE AS INTEGER)""", ['21', ])
        city_list22 = CITY.objects.raw("""SELECT * FROM CITY WHERE KEN_CODE=%s ORDER BY CAST(CITY_CODE AS INTEGER)""", ['22', ])
        city_list23 = CITY.objects.raw("""SELECT * FROM CITY WHERE KEN_CODE=%s ORDER BY CAST(CITY_CODE AS INTEGER)""", ['23', ])
        city_list24 = CITY.objects.raw("""SELECT * FROM CITY WHERE KEN_CODE=%s ORDER BY CAST(CITY_CODE AS INTEGER)""", ['24', ])
        city_list25 = CITY.objects.raw("""SELECT * FROM CITY WHERE KEN_CODE=%s ORDER BY CAST(CITY_CODE AS INTEGER)""", ['25', ])
        city_list26 = CITY.objects.raw("""SELECT * FROM CITY WHERE KEN_CODE=%s ORDER BY CAST(CITY_CODE AS INTEGER)""", ['26', ])
        city_list27 = CITY.objects.raw("""SELECT * FROM CITY WHERE KEN_CODE=%s ORDER BY CAST(CITY_CODE AS INTEGER)""", ['27', ])
        city_list28 = CITY.objects.raw("""SELECT * FROM CITY WHERE KEN_CODE=%s ORDER BY CAST(CITY_CODE AS INTEGER)""", ['28', ])
        city_list29 = CITY.objects.raw("""SELECT * FROM CITY WHERE KEN_CODE=%s ORDER BY CAST(CITY_CODE AS INTEGER)""", ['29', ])
        city_list30 = CITY.objects.raw("""SELECT * FROM CITY WHERE KEN_CODE=%s ORDER BY CAST(CITY_CODE AS INTEGER)""", ['30', ])
        city_list31 = CITY.objects.raw("""SELECT * FROM CITY WHERE KEN_CODE=%s ORDER BY CAST(CITY_CODE AS INTEGER)""", ['31', ])
        city_list32 = CITY.objects.raw("""SELECT * FROM CITY WHERE KEN_CODE=%s ORDER BY CAST(CITY_CODE AS INTEGER)""", ['32', ])
        city_list33 = CITY.objects.raw("""SELECT * FROM CITY WHERE KEN_CODE=%s ORDER BY CAST(CITY_CODE AS INTEGER)""", ['33', ])
        city_list34 = CITY.objects.raw("""SELECT * FROM CITY WHERE KEN_CODE=%s ORDER BY CAST(CITY_CODE AS INTEGER)""", ['34', ])
        city_list35 = CITY.objects.raw("""SELECT * FROM CITY WHERE KEN_CODE=%s ORDER BY CAST(CITY_CODE AS INTEGER)""", ['35', ])
        city_list36 = CITY.objects.raw("""SELECT * FROM CITY WHERE KEN_CODE=%s ORDER BY CAST(CITY_CODE AS INTEGER)""", ['36', ])
        city_list37 = CITY.objects.raw("""SELECT * FROM CITY WHERE KEN_CODE=%s ORDER BY CAST(CITY_CODE AS INTEGER)""", ['37', ])
        city_list38 = CITY.objects.raw("""SELECT * FROM CITY WHERE KEN_CODE=%s ORDER BY CAST(CITY_CODE AS INTEGER)""", ['38', ])
        city_list39 = CITY.objects.raw("""SELECT * FROM CITY WHERE KEN_CODE=%s ORDER BY CAST(CITY_CODE AS INTEGER)""", ['39', ])
        city_list40 = CITY.objects.raw("""SELECT * FROM CITY WHERE KEN_CODE=%s ORDER BY CAST(CITY_CODE AS INTEGER)""", ['40', ])
        city_list41 = CITY.objects.raw("""SELECT * FROM CITY WHERE KEN_CODE=%s ORDER BY CAST(CITY_CODE AS INTEGER)""", ['41', ])
        city_list42 = CITY.objects.raw("""SELECT * FROM CITY WHERE KEN_CODE=%s ORDER BY CAST(CITY_CODE AS INTEGER)""", ['42', ])
        city_list43 = CITY.objects.raw("""SELECT * FROM CITY WHERE KEN_CODE=%s ORDER BY CAST(CITY_CODE AS INTEGER)""", ['43', ])
        city_list44 = CITY.objects.raw("""SELECT * FROM CITY WHERE KEN_CODE=%s ORDER BY CAST(CITY_CODE AS INTEGER)""", ['44', ])
        city_list45 = CITY.objects.raw("""SELECT * FROM CITY WHERE KEN_CODE=%s ORDER BY CAST(CITY_CODE AS INTEGER)""", ['45', ])
        city_list46 = CITY.objects.raw("""SELECT * FROM CITY WHERE KEN_CODE=%s ORDER BY CAST(CITY_CODE AS INTEGER)""", ['46', ])
        city_list47 = CITY.objects.raw("""SELECT * FROM CITY WHERE KEN_CODE=%s ORDER BY CAST(CITY_CODE AS INTEGER)""", ['47', ])

        #######################################################################
        ### レスポンスセット処理(0020)
        ### テンプレートとコンテキストを設定して、レスポンスをブラウザに戻す。
        #######################################################################
        print_log('[DEBUG] P0200ExcelDownload.index_view()関数 STEP 3/4.', 'DEBUG')
        template = loader.get_template('P0200ExcelDownload/index.html')
        context = {
            'ken_list': ken_list,
            'city_list01': city_list01,
            'city_list02': city_list02,
            'city_list03': city_list03,
            'city_list04': city_list04,
            'city_list05': city_list05,
            'city_list06': city_list06,
            'city_list07': city_list07,
            'city_list08': city_list08,
            'city_list09': city_list09,
            'city_list10': city_list10,
            'city_list11': city_list11,
            'city_list12': city_list12,
            'city_list13': city_list13,
            'city_list14': city_list14,
            'city_list15': city_list15,
            'city_list16': city_list16,
            'city_list17': city_list17,
            'city_list18': city_list18,
            'city_list19': city_list19,
            'city_list20': city_list20,
            'city_list21': city_list21,
            'city_list22': city_list22,
            'city_list23': city_list23,
            'city_list24': city_list24,
            'city_list25': city_list25,
            'city_list26': city_list26,
            'city_list27': city_list27,
            'city_list28': city_list28,
            'city_list29': city_list29,
            'city_list30': city_list30,
            'city_list31': city_list31,
            'city_list32': city_list32,
            'city_list33': city_list33,
            'city_list34': city_list34,
            'city_list35': city_list35,
            'city_list36': city_list36,
            'city_list37': city_list37,
            'city_list38': city_list38,
            'city_list39': city_list39,
            'city_list40': city_list40,
            'city_list41': city_list41,
            'city_list42': city_list42,
            'city_list43': city_list43,
            'city_list44': city_list44,
            'city_list45': city_list45,
            'city_list46': city_list46,
            'city_list47': city_list47,
            'data_type': data_type,
        }
        print_log('[INFO] P0200ExcelDownload.index_view()関数が正常終了しました。', 'INFO')
        return HttpResponse(template.render(context, request))
    
    except:
        print_log('[ERROR] P0200ExcelDownload.index_view()関数 {}'.format(sys.exc_info()[0]), 'ERROR')
        print_log('[ERROR] P0200ExcelDownload.index_view()関数でエラーが発生しました。', 'ERROR')
        print_log('[ERROR] P0200ExcelDownload.index_view()関数が異常終了しました。', 'ERROR')
        return render(request, 'error.html')

###############################################################################
### 関数名：building_view(request, lock)
### 1000：建物区分
### urlpattern：path('building/', views.building_view, name='building_view')
### template：
###############################################################################
### @login_required(None, login_url='/P0100Login/')
def building_view(request, lock):
    try:
        #######################################################################
        ### 引数チェック処理(0000)
        ### ブラウザからのリクエストと引数をチェックする。
        #######################################################################
        ### reset_log()
        print_log('[INFO] P0200ExcelDownload.building_view()関数が開始しました。', 'INFO')
        print_log('[DEBUG] P0200ExcelDownload.building_view()関数 request = {}'.format(request.method), 'DEBUG')
        print_log('[DEBUG] P0200ExcelDownload.building_view()関数 lock = {}'.format(lock), 'DEBUG')
        print_log('[DEBUG] P0200ExcelDownload.building_view()関数 STEP 1/4.', 'DEBUG')
        
        #######################################################################
        ### DBアクセス処理(0010)
        ### DBにアクセスして、建物区分データを取得する。
        #######################################################################
        print_log('[DEBUG] P0200ExcelDownload.building_view()関数 STEP 2/4.', 'DEBUG')
        building_list = BUILDING.objects.raw("""SELECT * FROM BUILDING ORDER BY CAST(BUILDING_CODE AS INTEGER)""", [])
        
        #######################################################################
        ### EXCEL入出力処理(0020)
        ### (1)テンプレート用のEXCELファイルを読み込む。
        ### (2)セルにデータをセットして、ダウンロード用のEXCELファイルを保存する。
        #######################################################################
        print_log('[DEBUG] P0200ExcelDownload.building_view()関数 STEP 3/4.', 'DEBUG')
        template_file_path = 'static/template_building.xlsx'
        download_file_path = 'static/download_building.xlsx'
        wb = openpyxl.load_workbook(template_file_path)
        ws = wb.active
        ws.title = '建物区分'
        ws.cell(row=1, column=1).value = '建物区分コード'
        ws.cell(row=1, column=2).value = '建物区分名'
        
        if building_list:
            for i, building in enumerate(building_list):
                ws.cell(row=i+2, column=1).value = building.building_code
                ws.cell(row=i+2, column=2).value = building.building_name
        
        wb.save(download_file_path)
        
        #######################################################################
        ### レスポンスセット処理(0030)
        ### テンプレートとコンテキストを設定して、レスポンスをブラウザに戻す。
        #######################################################################
        print_log('[DEBUG] P0200ExcelDownload.building_view()関数 STEP 4/4.', 'DEBUG')
        print_log('[INFO] P0200ExcelDownload.building_view()関数が正常終了しました。', 'INFO')
        response = HttpResponse(content=save_virtual_workbook(wb), content_type='application/vnd.ms-excel')
        response['Content-Disposition'] = 'attachment; filename="building.xlsx"'
        return response
        
    except:
        print_log('[ERROR] P0200ExcelDownload.building_view()関数 {}'.format(sys.exc_info()[0]), 'ERROR')
        print_log('[ERROR] P0200ExcelDownload.building_view()関数でエラーが発生しました。', 'ERROR')
        print_log('[ERROR] P0200ExcelDownload.building_view()関数が異常終了しました。', 'ERROR')
        return render(request, 'error.html')

###############################################################################
### 関数名：ken_view(request, lock)
### 1010：都道府県
### urlpattern：path('ken/', views.ken_view, name='ken_view')
### template：
###############################################################################
### @login_required(None, login_url='/P0100Login/')
def ken_view(request, lock):
    try:
        #######################################################################
        ### 引数チェック処理(0000)
        ### ブラウザからのリクエストと引数をチェックする。
        #######################################################################
        ### reset_log()
        print_log('[INFO] P0200ExcelDownload.ken_view()関数が開始しました。', 'INFO')
        print_log('[DEBUG] P0200ExcelDownload.ken_view()関数 request = {}'.format(request.method), 'DEBUG')
        print_log('[DEBUG] P0200ExcelDownload.ken_view()関数 lock = {}'.format(lock), 'DEBUG')
        print_log('[DEBUG] P0200ExcelDownload.ken_view()関数 STEP 1/4.', 'DEBUG')
        
        #######################################################################
        ### DBアクセス処理(0010)
        ### DBにアクセスして、都道府県データを取得する。
        #######################################################################
        print_log('[DEBUG] P0200ExcelDownload.ken_view()関数 STEP 2/4.', 'DEBUG')
        ken_list = KEN.objects.raw("""SELECT * FROM KEN ORDER BY CAST(KEN_CODE AS INTEGER)""", [])
        
        #######################################################################
        ### EXCEL入出力処理(0020)
        ### (1)テンプレート用のEXCELファイルを読み込む。
        ### (2)セルにデータをセットして、ダウンロード用のEXCELファイルを保存する。
        #######################################################################
        print_log('[DEBUG] P0200ExcelDownload.ken_view()関数 STEP 3/4.', 'DEBUG')
        template_file_path = 'static/template_ken.xlsx'
        download_file_path = 'static/download_ken.xlsx'
        wb = openpyxl.load_workbook(template_file_path)
        ws = wb.active
        ws.title = '都道府県'
        ws.cell(row=1, column=1).value = '都道府県コード'
        ws.cell(row=1, column=2).value = '都道府県名'
        
        if ken_list:
            for i, ken in enumerate(ken_list):
                ws.cell(row=i+2, column=1).value = ken.ken_code
                ws.cell(row=i+2, column=2).value = ken.ken_name
        
        wb.save(download_file_path)
        
        #######################################################################
        ### レスポンスセット処理(0030)
        ### テンプレートとコンテキストを設定して、レスポンスをブラウザに戻す。
        #######################################################################
        print_log('[DEBUG] P0200ExcelDownload.ken_view()関数 STEP 4/4.', 'DEBUG')
        print_log('[INFO] P0200ExcelDownload.ken_view()関数が正常終了しました。', 'INFO')
        response = HttpResponse(content=save_virtual_workbook(wb), content_type='application/vnd.ms-excel')
        response['Content-Disposition'] = 'attachment; filename="ken.xlsx"'
        return response
        
    except:
        print_log('[ERROR] P0200ExcelDownload.ken_view()関数 {}'.format(sys.exc_info()[0]), 'ERROR')
        print_log('[ERROR] P0200ExcelDownload.ken_view()関数でエラーが発生しました。', 'ERROR')
        print_log('[ERROR] P0200ExcelDownload.ken_view()関数が異常終了しました。', 'ERROR')
        return render(request, 'error.html')

###############################################################################
### 関数名：city_view(request, lock)
### 1020：市区町村
### urlpattern：path('city/', views.city_view, name='city_view')
### template：
###############################################################################
### @login_required(None, login_url='/P0100Login/')
def city_view(request, lock):
    try:
        #######################################################################
        ### 引数チェック処理(0000)
        ### ブラウザからのリクエストと引数をチェックする。
        #######################################################################
        ### reset_log()
        print_log('[INFO] P0200ExcelDownload.city_view()関数が開始しました。', 'INFO')
        print_log('[DEBUG] P0200ExcelDownload.city_view()関数 request = {}'.format(request.method), 'DEBUG')
        print_log('[DEBUG] P0200ExcelDownload.city_view()関数 lock = {}'.format(lock), 'DEBUG')
        print_log('[DEBUG] P0200ExcelDownload.city_view()関数 STEP 1/4.', 'DEBUG')
        
        #######################################################################
        ### DBアクセス処理(0010)
        ### DBにアクセスして、市区町村データを取得する。
        #######################################################################
        print_log('[DEBUG] P0200ExcelDownload.city_view()関数 STEP 2/4.', 'DEBUG')
        city_list = CITY.objects.raw("""
            SELECT 
                CT1.city_code AS city_code, 
                CT1.city_name AS city_name, 
                CT1.ken_code AS ken_code, 
                KE1.ken_name AS ken_name, 
                CT1.city_population AS city_population, 
                CT1.city_area AS city_area 
            FROM CITY CT1 
            LEFT JOIN KEN KE1 ON CT1.ken_code=KE1.ken_code 
            ORDER BY CAST(CT1.city_code AS INTEGER)""", [])
        
        #######################################################################
        ### EXCEL入出力処理(0020)
        ### (1)テンプレート用のEXCELファイルを読み込む。
        ### (2)セルにデータをセットして、ダウンロード用のEXCELファイルを保存する。
        #######################################################################
        print_log('[DEBUG] P0200ExcelDownload.city_view()関数 STEP 3/4.', 'DEBUG')
        template_file_path = 'static/template_city.xlsx'
        download_file_path = 'static/download_city.xlsx'
        wb = openpyxl.load_workbook(template_file_path)
        ws = wb.active
        ws.title = '市区町村'
        ws.cell(row=1, column=1).value = '市区町村コード'
        ws.cell(row=1, column=2).value = '市区町村名'
        ws.cell(row=1, column=3).value = '都道府県コード'
        ws.cell(row=1, column=4).value = '都道府県名'
        ws.cell(row=1, column=5).value = '市区町村人口'
        ws.cell(row=1, column=6).value = '市区町村面積'
        
        if city_list:
            for i, city in enumerate(city_list):
                ws.cell(row=i+2, column=1).value = city.city_code
                ws.cell(row=i+2, column=2).value = city.city_name
                ws.cell(row=i+2, column=3).value = city.ken_code
                ws.cell(row=i+2, column=4).value = city.ken_name
                ws.cell(row=i+2, column=5).value = city.city_population
                ws.cell(row=i+2, column=6).value = city.city_area
        
        wb.save(download_file_path)
        
        #######################################################################
        ### レスポンスセット処理(0030)
        ### テンプレートとコンテキストを設定して、レスポンスをブラウザに戻す。
        #######################################################################
        print_log('[DEBUG] P0200ExcelDownload.city_view()関数 STEP 4/4.', 'DEBUG')
        print_log('[INFO] P0200ExcelDownload.city_view()関数が正常終了しました。', 'INFO')
        response = HttpResponse(content=save_virtual_workbook(wb), content_type='application/vnd.ms-excel')
        response['Content-Disposition'] = 'attachment; filename="city.xlsx"'
        return response
        
    except:
        print_log('[ERROR] P0200ExcelDownload.city_view()関数 {}'.format(sys.exc_info()[0]), 'ERROR')
        print_log('[ERROR] P0200ExcelDownload.city_view()関数でエラーが発生しました。', 'ERROR')
        print_log('[ERROR] P0200ExcelDownload.city_view()関数が異常終了しました。', 'ERROR')
        return render(request, 'error.html')

###############################################################################
### 関数名：kasen_kaigan_view(request, lock)
### 1030：水害発生地点工種（河川海岸区分）
### urlpattern：path('kasen_kaigan/', views.kasen_kaigan_view, name='kasen_kaigan_view')
### template：
###############################################################################
### @login_required(None, login_url='/P0100Login/')
def kasen_kaigan_view(request, lock):
    try:
        #######################################################################
        ### 引数チェック処理(0000)
        ### ブラウザからのリクエストと引数をチェックする。
        #######################################################################
        ### reset_log()
        print_log('[INFO] P0200ExcelDownload.kasen_kaigan_view()関数が開始しました。', 'INFO')
        print_log('[DEBUG] P0200ExcelDownload.kasen_kaigan_view()関数 request = {}'.format(request.method), 'DEBUG')
        print_log('[DEBUG] P0200ExcelDownload.kasen_kaigan_view()関数 lock = {}'.format(lock), 'DEBUG')
        print_log('[DEBUG] P0200ExcelDownload.kasen_kaigan_view()関数 STEP 1/4.', 'DEBUG')
        
        #######################################################################
        ### DBアクセス処理(0010)
        ### DBにアクセスして、水害発生地点工種（河川海岸区分）データを取得する。
        #######################################################################
        print_log('[DEBUG] P0200ExcelDownload.kasen_kaigan_view()関数 STEP 2/4.', 'DEBUG')
        kasen_kaigan_list = KASEN_KAIGAN.objects.raw("""SELECT * FROM KASEN_KAIGAN ORDER BY CAST(KASEN_KAIGAN_CODE AS INTEGER)""", [])
    
        #######################################################################
        ### EXCEL入出力処理(0020)
        ### (1)テンプレート用のEXCELファイルを読み込む。
        ### (2)セルにデータをセットして、ダウンロード用のEXCELファイルを保存する。
        #######################################################################
        print_log('[DEBUG] P0200ExcelDownload.kasen_kaigan_view()関数 STEP 3/4.', 'DEBUG')
        template_file_path = 'static/template_kasen_kaigan.xlsx'
        download_file_path = 'static/download_kasen_kaigan.xlsx'
        wb = openpyxl.load_workbook(template_file_path)
        ws = wb.active
        ws.title = '河川海岸区分'
        ws.cell(row=1, column=1).value = '河川海岸区分コード'
        ws.cell(row=1, column=2).value = '河川海岸区分名'
        
        if kasen_kaigan_list:
            for i, kasen_kaigan in enumerate(kasen_kaigan_list):
                ws.cell(row=i+2, column=1).value = kasen_kaigan.kasen_kaigan_code
                ws.cell(row=i+2, column=2).value = kasen_kaigan.kasen_kaigan_name
        
        wb.save(download_file_path)
        
        #######################################################################
        ### レスポンスセット処理(0030)
        ### テンプレートとコンテキストを設定して、レスポンスをブラウザに戻す。
        #######################################################################
        print_log('[DEBUG] P0200ExcelDownload.kasen_kaigan_view()関数 STEP 4/4.', 'DEBUG')
        print_log('[INFO] P0200ExcelDownload.kasen_kaigan_view()関数が正常終了しました。', 'INFO')
        response = HttpResponse(content=save_virtual_workbook(wb), content_type='application/vnd.ms-excel')
        response['Content-Disposition'] = 'attachment; filename="kasen_kaigan.xlsx"'
        return response
        
    except:
        print_log('[ERROR] P0200ExcelDownload.kasen_kaigan_view()関数 {}'.format(sys.exc_info()[0]), 'ERROR')
        print_log('[ERROR] P0200ExcelDownload.kasen_kaigan_view()関数でエラーが発生しました。', 'ERROR')
        print_log('[ERROR] P0200ExcelDownload.kasen_kaigan_view()関数が異常終了しました。', 'ERROR')
        return render(request, 'error.html')

###############################################################################
### 関数名：suikei_view(request, lock)
### 1040：水系（水系・沿岸）
### urlpattern：path('suikei/', views.suikei_view, name='suikei_view')
### template：
###############################################################################
### @login_required(None, login_url='/P0100Login/')
def suikei_view(request, lock):
    try:
        #######################################################################
        ### 引数チェック処理(0000)
        ### ブラウザからのリクエストと引数をチェックする。
        #######################################################################
        ### reset_log()
        print_log('[INFO] P0200ExcelDownload.suikei_view()関数が開始しました。', 'INFO')
        print_log('[DEBUG] P0200ExcelDownload.suikei_view()関数 request = {}'.format(request.method), 'DEBUG')
        print_log('[DEBUG] P0200ExcelDownload.suikei_view()関数 lock = {}'.format(lock), 'DEBUG')
        print_log('[DEBUG] P0200ExcelDownload.suikei_view()関数 STEP 1/4.', 'DEBUG')
        
        #######################################################################
        ### DBアクセス処理(0010)
        ### DBにアクセスして、水系（水系・沿岸）データを取得する。
        #######################################################################
        print_log('[DEBUG] P0200ExcelDownload.suikei_view()関数 STEP 2/4.', 'DEBUG')
        suikei_list = SUIKEI.objects.raw("""
            SELECT 
                SK1.suikei_code AS suikei_code, 
                SK1.suikei_name AS suikei_name, 
                SK1.suikei_type_code AS suikei_type_code, 
                ST1.suikei_type_name AS suikei_type_name 
            FROM SUIKEI SK1 
            LEFT JOIN SUIKEI_TYPE ST1 ON SK1.suikei_type_code=ST1.suikei_type_code 
            ORDER BY CAST(SK1.suikei_code AS INTEGER)""", [])
    
        #######################################################################
        ### EXCEL入出力処理(0020)
        ### (1)テンプレート用のEXCELファイルを読み込む。
        ### (2)セルにデータをセットして、ダウンロード用のEXCELファイルを保存する。
        #######################################################################
        print_log('[DEBUG] P0200ExcelDownload.suikei_view()関数 STEP 3/4.', 'DEBUG')
        template_file_path = 'static/template_suikei.xlsx'
        download_file_path = 'static/download_suikei.xlsx'
        wb = openpyxl.load_workbook(template_file_path)
        ws = wb.active
        ws.title = '水系'
        ws.cell(row=1, column=1).value = '水系コード'
        ws.cell(row=1, column=2).value = '水系名'
        ws.cell(row=1, column=3).value = '水系種別コード'
        ws.cell(row=1, column=4).value = '水系種別名'
        
        if suikei_list:
            for i, suikei in enumerate(suikei_list):
                ws.cell(row=i+2, column=1).value = suikei.suikei_code
                ws.cell(row=i+2, column=2).value = suikei.suikei_name
                ws.cell(row=i+2, column=3).value = suikei.suikei_type_code
                ws.cell(row=i+2, column=4).value = suikei.suikei_type_name
        
        wb.save(download_file_path)
        
        #######################################################################
        ### レスポンスセット処理(0030)
        ### テンプレートとコンテキストを設定して、レスポンスをブラウザに戻す。
        #######################################################################
        print_log('[DEBUG] P0200ExcelDownload.suikei_view()関数 STEP 4/4.', 'DEBUG')
        print_log('[INFO] P0200ExcelDownload.suikei_view()関数が正常終了しました。', 'INFO')
        response = HttpResponse(content=save_virtual_workbook(wb), content_type='application/vnd.ms-excel')
        response['Content-Disposition'] = 'attachment; filename="suikei.xlsx"'
        return response
        
    except:
        print_log('[ERROR] P0200ExcelDownload.suikei_view()関数 {}'.format(sys.exc_info()[0]), 'ERROR')
        print_log('[ERROR] P0200ExcelDownload.suikei_view()関数でエラーが発生しました。', 'ERROR')
        print_log('[ERROR] P0200ExcelDownload.suikei_view()関数が異常終了しました。', 'ERROR')
        return render(request, 'error.html')

###############################################################################
### 関数名：suikei_type_view(request, lock)
### 1050：水系種別（水系・沿岸種別）
### urlpattern：path('suikei_type/', views.suikei_type_view, name='suikei_type_view')
### template：
###############################################################################
### @login_required(None, login_url='/P0100Login/')
def suikei_type_view(request, lock):
    try:
        #######################################################################
        ### 引数チェック処理(0000)
        ### ブラウザからのリクエストと引数をチェックする。
        #######################################################################
        ### reset_log()
        print_log('[INFO] P0200ExcelDownload.suikei_type_view()関数が開始しました。', 'INFO')
        print_log('[DEBUG] P0200ExcelDownload.suikei_type_view()関数 request = {}'.format(request.method), 'DEBUG')
        print_log('[DEBUG] P0200ExcelDownload.suikei_type_view()関数 lock = {}'.format(lock), 'DEBUG')
        print_log('[DEBUG] P0200ExcelDownload.suikei_type_view()関数 STEP 1/4.', 'DEBUG')
        
        #######################################################################
        ### DBアクセス処理(0010)
        ### DBにアクセスして、水系種別（水系・沿岸種別）データを取得する。
        #######################################################################
        print_log('[DEBUG] P0200ExcelDownload.suikei_type_view()関数 STEP 2/4.', 'DEBUG')
        suikei_type_list = SUIKEI_TYPE.objects.raw("""SELECT * FROM SUIKEI_TYPE ORDER BY CAST(SUIKEI_TYPE_CODE AS INTEGER)""", [])
    
        #######################################################################
        ### EXCEL入出力処理(0020)
        ### (1)テンプレート用のEXCELファイルを読み込む。
        ### (2)セルにデータをセットして、ダウンロード用のEXCELファイルを保存する。
        #######################################################################
        print_log('[DEBUG] P0200ExcelDownload.suikei_type_view()関数 STEP 3/4.', 'DEBUG')
        template_file_path = 'static/template_suikei_type.xlsx'
        download_file_path = 'static/download_suikei_type.xlsx'
        wb = openpyxl.load_workbook(template_file_path)
        ws = wb.active
        ws.title = '水系種別'
        ws.cell(row=1, column=1).value = '水系種別コード'
        ws.cell(row=1, column=2).value = '水系種別名'
        
        if suikei_type_list:
            for i, suikei_type in enumerate(suikei_type_list):
                ws.cell(row=i+2, column=1).value = suikei_type.suikei_type_code
                ws.cell(row=i+2, column=2).value = suikei_type.suikei_type_name
        
        wb.save(download_file_path)
        
        #######################################################################
        ### レスポンスセット処理(0030)
        ### テンプレートとコンテキストを設定して、レスポンスをブラウザに戻す。
        #######################################################################
        print_log('[DEBUG] P0200ExcelDownload.suikei_type_view()関数 STEP 4/4.', 'DEBUG')
        print_log('[INFO] P0200ExcelDownload.suikei_type_view()関数が正常終了しました。', 'INFO')
        response = HttpResponse(content=save_virtual_workbook(wb), content_type='application/vnd.ms-excel')
        response['Content-Disposition'] = 'attachment; filename="suikei_type.xlsx"'
        return response
        
    except:
        print_log('[ERROR] P0200ExcelDownload.suikei_type_view()関数 {}'.format(sys.exc_info()[0]), 'ERROR')
        print_log('[ERROR] P0200ExcelDownload.suikei_type_view()関数でエラーが発生しました。', 'ERROR')
        print_log('[ERROR] P0200ExcelDownload.suikei_type_view()関数が異常終了しました。', 'ERROR')
        return render(request, 'error.html')

###############################################################################
### 関数名：kasen_view(request, lock)
### 1060：河川（河川・海岸）
### urlpattern：path('kasen/', views.kasen_view, name='kasen_view')
### template：
###############################################################################
### @login_required(None, login_url='/P0100Login/')
def kasen_view(request, lock):
    try:
        #######################################################################
        ### 引数チェック処理(0000)
        ### ブラウザからのリクエストと引数をチェックする。
        #######################################################################
        ### reset_log()
        print_log('[INFO] P0200ExcelDownload.kasen_view()関数が開始しました。', 'INFO')
        print_log('[DEBUG] P0200ExcelDownload.kasen_view()関数 request = {}'.format(request.method), 'DEBUG')
        print_log('[DEBUG] P0200ExcelDownload.kasen_view()関数 lock = {}'.format(lock), 'DEBUG')
        print_log('[DEBUG] P0200ExcelDownload.kasen_view()関数 STEP 1/4.', 'DEBUG')
        
        #######################################################################
        ### DBアクセス処理(0010)
        ### DBにアクセスして、河川（河川・海岸）データを取得する。
        #######################################################################
        print_log('[DEBUG] P0200ExcelDownload.kasen_view()関数 STEP 2/4.', 'DEBUG')
        kasen_list = KASEN.objects.raw("""
            SELECT 
                KA1.kasen_code AS kasen_code, 
                KA1.kasen_name AS kasen_name, 
                KA1.kasen_type_code AS kasen_type_code, 
                KT1.kasen_type_name AS kasen_type_name, 
                KA1.suikei_code AS suikei_code, 
                SK1.suikei_name AS suikei_name 
            FROM KASEN KA1 
            LEFT JOIN KASEN_TYPE KT1 ON KA1.kasen_type_code=KT1.kasen_type_code 
            LEFT JOIN SUIKEI SK1 ON KA1.suikei_code=SK1.suikei_code 
            ORDER BY CAST(KA1.kasen_code AS INTEGER)""", [])
    
        #######################################################################
        ### EXCEL入出力処理(0020)
        ### (1)テンプレート用のEXCELファイルを読み込む。
        ### (2)セルにデータをセットして、ダウンロード用のEXCELファイルを保存する。
        #######################################################################
        print_log('[DEBUG] P0200ExcelDownload.kasen_view()関数 STEP 3/4.', 'DEBUG')
        template_file_path = 'static/template_kasen.xlsx'
        download_file_path = 'static/download_kasen.xlsx'
        wb = openpyxl.load_workbook(template_file_path)
        ws = wb.active
        ws.title = '河川'
        ws.cell(row=1, column=1).value = '河川コード'
        ws.cell(row=1, column=2).value = '河川名'
        ws.cell(row=1, column=3).value = '河川種別コード'
        ws.cell(row=1, column=4).value = '河川種別名'
        ws.cell(row=1, column=5).value = '水系コード'
        ws.cell(row=1, column=6).value = '水系名'
        
        if kasen_list:
            for i, kasen in enumerate(kasen_list):
                ws.cell(row=i+2, column=1).value = kasen.kasen_code
                ws.cell(row=i+2, column=2).value = kasen.kasen_name
                ws.cell(row=i+2, column=3).value = kasen.kasen_type_code
                ws.cell(row=i+2, column=4).value = kasen.kasen_type_name
                ws.cell(row=i+2, column=5).value = kasen.suikei_code
                ws.cell(row=i+2, column=6).value = kasen.suikei_name
        
        wb.save(download_file_path)
        
        #######################################################################
        ### レスポンスセット処理(0030)
        ### テンプレートとコンテキストを設定して、レスポンスをブラウザに戻す。
        #######################################################################
        print_log('[DEBUG] P0200ExcelDownload.kasen_view()関数 STEP 4/4.', 'DEBUG')
        print_log('[INFO] P0200ExcelDownload.kasen_view()関数が正常終了しました。', 'INFO')
        response = HttpResponse(content=save_virtual_workbook(wb), content_type='application/vnd.ms-excel')
        response['Content-Disposition'] = 'attachment; filename="kasen.xlsx"'
        return response
        
    except:
        print_log('[ERROR] P0200ExcelDownload.kasen_view()関数 {}'.format(sys.exc_info()[0]), 'ERROR')
        print_log('[ERROR] P0200ExcelDownload.kasen_view()関数でエラーが発生しました。', 'ERROR')
        print_log('[ERROR] P0200ExcelDownload.kasen_view()関数が異常終了しました。', 'ERROR')
        return render(request, 'error.html')

###############################################################################
### 関数名：kasen_type_view(request, lock)
### 1070：河川種別（河川・海岸種別）
### urlpattern：path('kasen_type/', views.kasen_type_view, name='kasen_type_view')
### template：
###############################################################################
### @login_required(None, login_url='/P0100Login/')
def kasen_type_view(request, lock):
    try:
        #######################################################################
        ### 引数チェック処理(0000)
        ### ブラウザからのリクエストと引数をチェックする。
        #######################################################################
        ### reset_log()
        print_log('[INFO] P0200ExcelDownload.kasen_type_view()関数が開始しました。', 'INFO')
        print_log('[DEBUG] P0200ExcelDownload.kasen_type_view()関数 request = {}'.format(request.method), 'DEBUG')
        print_log('[DEBUG] P0200ExcelDownload.kasen_type_view()関数 lock = {}'.format(lock), 'DEBUG')
        print_log('[DEBUG] P0200ExcelDownload.kasen_type_view()関数 STEP 1/4.', 'DEBUG')
        
        #######################################################################
        ### DBアクセス処理(0010)
        ### DBにアクセスして、河川種別（河川・海岸種別）データを取得する。
        #######################################################################
        print_log('[DEBUG] P0200ExcelDownload.kasen_type_view()関数 STEP 2/4.', 'DEBUG')
        kasen_type_list = KASEN_TYPE.objects.raw("""SELECT * FROM KASEN_TYPE ORDER BY CAST(KASEN_TYPE_CODE AS INTEGER)""", [])
    
        #######################################################################
        ### EXCEL入出力処理(0020)
        ### (1)テンプレート用のEXCELファイルを読み込む。
        ### (2)セルにデータをセットして、ダウンロード用のEXCELファイルを保存する。
        #######################################################################
        print_log('[DEBUG] P0200ExcelDownload.kasen_type_view()関数 STEP 3/4.', 'DEBUG')
        template_file_path = 'static/template_kasen_type.xlsx'
        download_file_path = 'static/download_kasen_type.xlsx'
        wb = openpyxl.load_workbook(template_file_path)
        ws = wb.active
        ws.title = '河川種別'
        ws.cell(row=1, column=1).value = '河川種別コード'
        ws.cell(row=1, column=2).value = '河川種別名'
        
        if kasen_type_list:
            for i, kasen_type in enumerate(kasen_type_list):
                ws.cell(row=i+2, column=1).value = kasen_type.kasen_type_code
                ws.cell(row=i+2, column=2).value = kasen_type.kasen_type_name
        
        wb.save(download_file_path)
        
        #######################################################################
        ### レスポンスセット処理(0030)
        ### テンプレートとコンテキストを設定して、レスポンスをブラウザに戻す。
        #######################################################################
        print_log('[DEBUG] P0200ExcelDownload.kasen_type_view()関数 STEP 4/4.', 'DEBUG')
        print_log('[INFO] P0200ExcelDownload.kasen_type_view()関数が正常終了しました。', 'INFO')
        response = HttpResponse(content=save_virtual_workbook(wb), content_type='application/vnd.ms-excel')
        response['Content-Disposition'] = 'attachment; filename="kasen_type.xlsx"'
        return response
        
    except:
        print_log('[ERROR] P0200ExcelDownload.kasen_type_view()関数 {}'.format(sys.exc_info()[0]), 'ERROR')
        print_log('[ERROR] P0200ExcelDownload.kasen_type_view()関数でエラーが発生しました。', 'ERROR')
        print_log('[ERROR] P0200ExcelDownload.kasen_type_view()関数が異常終了しました。', 'ERROR')
        return render(request, 'error.html')

###############################################################################
### 関数名：cause_view(request, lock)
### 1080：水害原因
### urlpattern：path('cause/', views.cause_view, name='cause_view')
### template：
###############################################################################
### @login_required(None, login_url='/P0100Login/')
def cause_view(request, lock):
    try:
        #######################################################################
        ### 引数チェック処理(0000)
        ### ブラウザからのリクエストと引数をチェックする。
        #######################################################################
        ### reset_log()
        print_log('[INFO] P0200ExcelDownload.cause_view()関数が開始しました。', 'INFO')
        print_log('[DEBUG] P0200ExcelDownload.cause_view()関数 request = {}'.format(request.method), 'DEBUG')
        print_log('[DEBUG] P0200ExcelDownload.cause_view()関数 lock = {}'.format(lock), 'DEBUG')
        print_log('[DEBUG] P0200ExcelDownload.cause_view()関数 STEP 1/4.', 'DEBUG')
        
        #######################################################################
        ### DBアクセス処理(0010)
        ### DBにアクセスして、水害原因データを取得する。
        #######################################################################
        print_log('[DEBUG] P0200ExcelDownload.cause_view()関数 STEP 2/4.', 'DEBUG')
        cause_list = CAUSE.objects.raw("""SELECT * FROM CAUSE ORDER BY CAST(CAUSE_CODE AS INTEGER)""", [])
    
        #######################################################################
        ### EXCEL入出力処理(0020)
        ### (1)テンプレート用のEXCELファイルを読み込む。
        ### (2)セルにデータをセットして、ダウンロード用のEXCELファイルを保存する。
        #######################################################################
        print_log('[DEBUG] P0200ExcelDownload.cause_view()関数 STEP 3/4.', 'DEBUG')
        template_file_path = 'static/template_cause.xlsx'
        download_file_path = 'static/download_cause.xlsx'
        wb = openpyxl.load_workbook(template_file_path)
        ws = wb.active
        ws.title = '水害原因'
        ws.cell(row=1, column=1).value = '水害原因コード'
        ws.cell(row=1, column=2).value = '水害原因名'
        
        if cause_list:
            for i, cause in enumerate(cause_list):
                ws.cell(row=i+2, column=1).value = cause.cause_code
                ws.cell(row=i+2, column=2).value = cause.cause_name
        
        wb.save(download_file_path)
        
        #######################################################################
        ### レスポンスセット処理(0030)
        ### テンプレートとコンテキストを設定して、レスポンスをブラウザに戻す。
        #######################################################################
        print_log('[DEBUG] P0200ExcelDownload.cause_view()関数 STEP 4/4.', 'DEBUG')
        print_log('[INFO] P0200ExcelDownload.cause_view()関数が正常終了しました。', 'INFO')
        response = HttpResponse(content=save_virtual_workbook(wb), content_type='application/vnd.ms-excel')
        response['Content-Disposition'] = 'attachment; filename="cause.xlsx"'
        return response
        
    except:
        print_log('[ERROR] P0200ExcelDownload.cause_view()関数 {}'.format(sys.exc_info()[0]), 'ERROR')
        print_log('[ERROR] P0200ExcelDownload.cause_view()関数でエラーが発生しました。', 'ERROR')
        print_log('[ERROR] P0200ExcelDownload.cause_view()関数が異常終了しました。', 'ERROR')
        return render(request, 'error.html')

###############################################################################
### 関数名：underground_view(request, lock)
### 1090：地上地下区分
### urlpattern：path('underground/', views.underground_view, name='underground_view')
### template：
###############################################################################
### @login_required(None, login_url='/P0100Login/')
def underground_view(request, lock):
    try:
        #######################################################################
        ### 引数チェック処理(0000)
        ### ブラウザからのリクエストと引数をチェックする。
        #######################################################################
        ### reset_log()
        print_log('[INFO] P0200ExcelDownload.underground_view()関数が開始しました。', 'INFO')
        print_log('[DEBUG] P0200ExcelDownload.underground_view()関数 request = {}'.format(request.method), 'DEBUG')
        print_log('[DEBUG] P0200ExcelDownload.underground_view()関数 lock = {}'.format(lock), 'DEBUG')
        print_log('[DEBUG] P0200ExcelDownload.underground_view()関数 STEP 1/4.', 'DEBUG')
        
        #######################################################################
        ### DBアクセス処理(0010)
        ### DBにアクセスして、地上地下区分データを取得する。
        #######################################################################
        print_log('[DEBUG] P0200ExcelDownload.underground_view()関数 STEP 2/4.', 'DEBUG')
        underground_list = UNDERGROUND.objects.raw("""SELECT * FROM UNDERGROUND ORDER BY CAST(UNDERGROUND_CODE AS INTEGER)""", [])
    
        #######################################################################
        ### EXCEL入出力処理(0020)
        ### (1)テンプレート用のEXCELファイルを読み込む。
        ### (2)セルにデータをセットして、ダウンロード用のEXCELファイルを保存する。
        #######################################################################
        print_log('[DEBUG] P0200ExcelDownload.underground_view()関数 STEP 3/4.', 'DEBUG')
        template_file_path = 'static/template_underground.xlsx'
        download_file_path = 'static/download_underground.xlsx'
        wb = openpyxl.load_workbook(template_file_path)
        ws = wb.active
        ws.title = '地上地下区分'
        ws.cell(row=1, column=1).value = '地上地下区分コード'
        ws.cell(row=1, column=2).value = '地上地下区分名'
        
        if underground_list:
            for i, underground in enumerate(underground_list):
                ws.cell(row=i+2, column=1).value = underground.underground_code
                ws.cell(row=i+2, column=2).value = underground.underground_name
        
        wb.save(download_file_path)
        
        #######################################################################
        ### レスポンスセット処理(0030)
        ### テンプレートとコンテキストを設定して、レスポンスをブラウザに戻す。
        #######################################################################
        print_log('[DEBUG] P0200ExcelDownload.underground_view()関数 STEP 4/4.', 'DEBUG')
        print_log('[INFO] P0200ExcelDownload.underground_view()関数が正常終了しました。', 'INFO')
        response = HttpResponse(content=save_virtual_workbook(wb), content_type='application/vnd.ms-excel')
        response['Content-Disposition'] = 'attachment; filename="underground.xlsx"'
        return response
        
    except:
        print_log('[ERROR] P0200ExcelDownload.underground_view()関数 {}'.format(sys.exc_info()[0]), 'ERROR')
        print_log('[ERROR] P0200ExcelDownload.underground_view()関数でエラーが発生しました。', 'ERROR')
        print_log('[ERROR] P0200ExcelDownload.underground_view()関数が異常終了しました。', 'ERROR')
        return render(request, 'error.html')

###############################################################################
### 関数名：usage_view(request, lock)
### 1100：地下空間の利用形態
### urlpattern：path('usage/', views.usage_view, name='usage_view')
### template：
###############################################################################
### @login_required(None, login_url='/P0100Login/')
def usage_view(request, lock):
    try:
        #######################################################################
        ### 引数チェック処理(0000)
        ### ブラウザからのリクエストと引数をチェックする。
        #######################################################################
        ### reset_log()
        print_log('[INFO] P0200ExcelDownload.usage_view()関数が開始しました。', 'INFO')
        print_log('[DEBUG] P0200ExcelDownload.usage_view()関数 request = {}'.format(request.method), 'DEBUG')
        print_log('[DEBUG] P0200ExcelDownload.usage_view()関数 lock = {}'.format(lock), 'DEBUG')
        print_log('[DEBUG] P0200ExcelDownload.usage_view()関数 STEP 1/4.', 'DEBUG')
        
        #######################################################################
        ### DBアクセス処理(0010)
        ### DBにアクセスして、地下空間の利用形態データを取得する。
        #######################################################################
        print_log('[DEBUG] P0200ExcelDownload.usage_view()関数 STEP 2/4.', 'DEBUG')
        usage_list = USAGE.objects.raw("""SELECT * FROM USAGE ORDER BY CAST(USAGE_CODE AS INTEGER)""", [])
    
        #######################################################################
        ### EXCEL入出力処理(0020)
        ### (1)テンプレート用のEXCELファイルを読み込む。
        ### (2)セルにデータをセットして、ダウンロード用のEXCELファイルを保存する。
        #######################################################################
        print_log('[DEBUG] P0200ExcelDownload.usage_view()関数 STEP 3/4.', 'DEBUG')
        template_file_path = 'static/template_usage.xlsx'
        download_file_path = 'static/download_usage.xlsx'
        wb = openpyxl.load_workbook(template_file_path)
        ws = wb.active
        ws.title = '地下空間の利用形態'
        ws.cell(row=1, column=1).value = '地下空間の利用形態コード'
        ws.cell(row=1, column=2).value = '地下空間の利用形態名'
        
        if usage_list:
            for i, usage in enumerate(usage_list):
                ws.cell(row=i+2, column=1).value = usage.usage_code
                ws.cell(row=i+2, column=2).value = usage.usage_name
        
        wb.save(download_file_path)
        
        #######################################################################
        ### レスポンスセット処理(0030)
        ### テンプレートとコンテキストを設定して、レスポンスをブラウザに戻す。
        #######################################################################
        print_log('[DEBUG] P0200ExcelDownload.usage_view()関数 STEP 4/4.', 'DEBUG')
        print_log('[INFO] P0200ExcelDownload.usage_view()関数が正常終了しました。', 'INFO')
        response = HttpResponse(content=save_virtual_workbook(wb), content_type='application/vnd.ms-excel')
        response['Content-Disposition'] = 'attachment; filename="usage.xlsx"'
        return response
        
    except:
        print_log('[ERROR] P0200ExcelDownload.usage_view()関数 {}'.format(sys.exc_info()[0]), 'ERROR')
        print_log('[ERROR] P0200ExcelDownload.usage_view()関数でエラーが発生しました。', 'ERROR')
        print_log('[ERROR] P0200ExcelDownload.usage_view()関数が異常終了しました。', 'ERROR')
        return render(request, 'error.html')

###############################################################################
### 関数名：flood_sediment_view(request, lock)
### 1110：浸水土砂区分
### urlpattern：path('flood_sediment/', views.flood_sediment_view, name='flood_sediment_view')
### template：
###############################################################################
### @login_required(None, login_url='/P0100Login/')
def flood_sediment_view(request, lock):
    try:
        #######################################################################
        ### 引数チェック処理(0000)
        ### ブラウザからのリクエストと引数をチェックする。
        #######################################################################
        ### reset_log()
        print_log('[INFO] P0200ExcelDownload.flood_sediment()関数が開始しました。', 'INFO')
        print_log('[DEBUG] P0200ExcelDownload.flood_sediment()関数 request = {}'.format(request.method), 'DEBUG')
        print_log('[DEBUG] P0200ExcelDownload.flood_sediment_view()関数 lock = {}'.format(lock), 'DEBUG')
        print_log('[DEBUG] P0200ExcelDownload.flood_sediment_view()関数 STEP 1/4.', 'DEBUG')
        
        #######################################################################
        ### DBアクセス処理(0010)
        ### DBにアクセスして、浸水土砂区分データを取得する。
        #######################################################################
        print_log('[DEBUG] P0200ExcelDownload.flood_sediment_view()関数 STEP 2/4.', 'DEBUG')
        flood_sediment_list = FLOOD_SEDIMENT.objects.raw("""SELECT * FROM FLOOD_SEDIMENT ORDER BY CAST(FLOOD_SEDIMENT_CODE AS INTEGER)""", [])
    
        #######################################################################
        ### EXCEL入出力処理(0020)
        ### (1)テンプレート用のEXCELファイルを読み込む。
        ### (2)セルにデータをセットして、ダウンロード用のEXCELファイルを保存する。
        #######################################################################
        print_log('[DEBUG] P0200ExcelDownload.flood_sediment_view()関数 STEP 3/4.', 'DEBUG')
        template_file_path = 'static/template_flood_sediment.xlsx'
        download_file_path = 'static/download_flood_sediment.xlsx'
        wb = openpyxl.load_workbook(template_file_path)
        ws = wb.active
        ws.title = '浸水土砂区分'
        ws.cell(row=1, column=1).value = '浸水土砂区分コード'
        ws.cell(row=1, column=2).value = '浸水土砂区分名'
        
        if flood_sediment_list:
            for i, flood_sediment in enumerate(flood_sediment_list):
                ws.cell(row=i+2, column=1).value = flood_sediment.flood_sediment_code
                ws.cell(row=i+2, column=2).value = flood_sediment.flood_sediment_name
        
        wb.save(download_file_path)
        
        #######################################################################
        ### レスポンスセット処理(0030)
        ### テンプレートとコンテキストを設定して、レスポンスをブラウザに戻す。
        #######################################################################
        print_log('[DEBUG] P0200ExcelDownload.flood_sediment_view()関数 STEP 4/4.', 'DEBUG')
        print_log('[INFO] P0200ExcelDownload.flood_sediment_view()関数が正常終了しました。', 'INFO')
        response = HttpResponse(content=save_virtual_workbook(wb), content_type='application/vnd.ms-excel')
        response['Content-Disposition'] = 'attachment; filename="flood_sediment.xlsx"'
        return response
        
    except:
        print_log('[ERROR] P0200ExcelDownload.flood_sediment_view()関数 '.format(sys.exc_info()[0]), 'ERROR')
        print_log('[ERROR] P0200ExcelDownload.flood_sediment_view()関数でエラーが発生しました。', 'ERROR')
        print_log('[ERROR] P0200ExcelDownload.flood_sediment_view()関数が異常終了しました。', 'ERROR')
        return render(request, 'error.html')

###############################################################################
### 関数名：gradient_view(request, lock)
### 1120：地盤勾配区分
### urlpattern：path('gradient/', views.gradient_view, name='gradient_view')
### template：
###############################################################################
### @login_required(None, login_url='/P0100Login/')
def gradient_view(request, lock):
    try:
        #######################################################################
        ### 引数チェック処理(0000)
        ### ブラウザからのリクエストと引数をチェックする。
        #######################################################################
        ### reset_log()
        print_log('[INFO] P0200ExcelDownload.gradient_view()関数が開始しました。', 'INFO')
        print_log('[DEBUG] P0200ExcelDownload.gradient_view()関数 request = {}'.format(request.method), 'DEBUG')
        print_log('[DEBUG] P0200ExcelDownload.gradient_view()関数 lock = {}'.format(lock), 'DEBUG')
        print_log('[DEBUG] P0200ExcelDownload.gradient_view()関数 STEP 1/4.', 'DEBUG')
        
        #######################################################################
        ### DBアクセス処理(0010)
        ### DBにアクセスして、地盤勾配区分データを取得する。
        #######################################################################
        print_log('[DEBUG] P0200ExcelDownload.gradient_view()関数 STEP 2/4.', 'DEBUG')
        gradient_list = GRADIENT.objects.raw("""SELECT * FROM GRADIENT ORDER BY CAST(GRADIENT_CODE AS INTEGER)""", [])
    
        #######################################################################
        ### EXCEL入出力処理(0020)
        ### (1)テンプレート用のEXCELファイルを読み込む。
        ### (2)セルにデータをセットして、ダウンロード用のEXCELファイルを保存する。
        #######################################################################
        print_log('[DEBUG] P0200ExcelDownload.gradient_view()関数 STEP 3/4.', 'DEBUG')
        template_file_path = 'static/template_gradient.xlsx'
        download_file_path = 'static/download_gradient.xlsx'
        wb = openpyxl.load_workbook(template_file_path)
        ws = wb.active
        ws.title = '地盤勾配区分'
        ws.cell(row=1, column=1).value = '地盤勾配区分コード'
        ws.cell(row=1, column=2).value = '地盤勾配区分名'
        
        if gradient_list:
            for i, gradient in enumerate(gradient_list):
                ws.cell(row=i+2, column=1).value = gradient.gradient_code
                ws.cell(row=i+2, column=2).value = gradient.gradient_name
        
        wb.save(download_file_path)
        
        #######################################################################
        ### レスポンスセット処理(0030)
        ### テンプレートとコンテキストを設定して、レスポンスをブラウザに戻す。
        #######################################################################
        print_log('[DEBUG] P0200ExcelDownload.gradient_view()関数 STEP 4/4.', 'DEBUG')
        print_log('[INFO] P0200ExcelDownload.gradient_view()関数が正常終了しました。', 'INFO')
        response = HttpResponse(content=save_virtual_workbook(wb), content_type='application/vnd.ms-excel')
        response['Content-Disposition'] = 'attachment; filename="gradient.xlsx"'
        return response
        
    except:
        print_log('[ERROR] P0200ExcelDownload.gradient_view()関数 {}'.format(sys.exc_info()[0]), 'ERROR')
        print_log('[ERROR] P0200ExcelDownload.gradient_view()関数でエラーが発生しました。', 'ERROR')
        print_log('[ERROR] P0200ExcelDownload.gradient_view()関数が異常終了しました。', 'ERROR')
        return render(request, 'error.html')

###############################################################################
### 関数名：industry_view(request, lock)
### 1130：産業分類
### urlpattern：path('industry/', views.industry_view, name='industry_view')
### template：
###############################################################################
### @login_required(None, login_url='/P0100Login/')
def industry_view(request, lock):
    try:
        #######################################################################
        ### 引数チェック処理(0000)
        ### ブラウザからのリクエストと引数をチェックする。
        #######################################################################
        ### reset_log()
        print_log('[INFO] P0200ExcelDownload.industry_view()関数が開始しました。', 'INFO')
        print_log('[DEBUG] P0200ExcelDownload.industry_view()関数 request = {}'.format(request.method), 'DEBUG')
        print_log('[DEBUG] P0200ExcelDownload.industry_view()関数 lock = {}'.format(lock), 'DEBUG')
        print_log('[DEBUG] P0200ExcelDownload.industry_view()関数 STEP 1/4.', 'DEBUG')
        
        #######################################################################
        ### DBアクセス処理(0010)
        ### DBにアクセスして、産業分類データを取得する。
        #######################################################################
        print_log('[DEBUG] P0200ExcelDownload.industry_view()関数 STEP 2/4.', 'DEBUG')
        industry_list = INDUSTRY.objects.raw("""SELECT * FROM INDUSTRY ORDER BY CAST(INDUSTRY_CODE AS INTEGER)""", [])
    
        #######################################################################
        ### EXCEL入出力処理(0020)
        ### (1)テンプレート用のEXCELファイルを読み込む。
        ### (2)セルにデータをセットして、ダウンロード用のEXCELファイルを保存する。
        #######################################################################
        print_log('[DEBUG] P0200ExcelDownload.industry_view()関数 STEP 3/4.', 'DEBUG')
        template_file_path = 'static/template_industry.xlsx'
        download_file_path = 'static/download_industry.xlsx'
        wb = openpyxl.load_workbook(template_file_path)
        ws = wb.active
        ws.title = '産業分類'
        ws.cell(row=1, column=1).value = '産業分類コード'
        ws.cell(row=1, column=2).value = '産業分類名'
        
        if industry_list:
            for i, industry in enumerate(industry_list):
                ws.cell(row=i+2, column=1).value = industry.industry_code
                ws.cell(row=i+2, column=2).value = industry.industry_name
        
        wb.save(download_file_path)
        
        #######################################################################
        ### レスポンスセット処理(0030)
        ### テンプレートとコンテキストを設定して、レスポンスをブラウザに戻す。
        #######################################################################
        print_log('[DEBUG] P0200ExcelDownload.industry_view()関数 STEP 4/4.', 'DEBUG')
        print_log('[INFO] P0200ExcelDownload.industry_view()関数が正常終了しました。', 'INFO')
        response = HttpResponse(content=save_virtual_workbook(wb), content_type='application/vnd.ms-excel')
        response['Content-Disposition'] = 'attachment; filename="industry.xlsx"'
        return response
        
    except:
        print_log('[ERROR] P0200ExcelDownload.industry_view()関数 {}'.format(sys.exc_info()[0]), 'ERROR')
        print_log('[ERROR] P0200ExcelDownload.industry_view()関数でエラーが発生しました。', 'ERROR')
        print_log('[ERROR] P0200ExcelDownload.industry_view()関数が異常終了しました。', 'ERROR')
        return render(request, 'error.html')

###############################################################################
### 関数名：house_asset_view(request, lock)
### 2000：家屋評価額
### urlpattern：path('house_asset/', views.house_asset_view, name='house_asset_view')
### template：
###############################################################################
### @login_required(None, login_url='/P0100Login/')
def house_asset_view(request, lock):
    try:
        #######################################################################
        ### 引数チェック処理(0000)
        ### ブラウザからのリクエストと引数をチェックする。
        #######################################################################
        ### reset_log()
        print_log('[INFO] P0200ExcelDownload.house_asset_view()関数が開始しました。', 'INFO')
        print_log('[DEBUG] P0200ExcelDownload.house_asset_view()関数 request = {}'.format(request.method), 'DEBUG')
        print_log('[DEBUG] P0200ExcelDownload.house_asset_view()関数 lock = {}'.format(lock), 'DEBUG')
        print_log('[DEBUG] P0200ExcelDownload.house_asset_view()関数 STEP 1/4.', 'DEBUG')
        
        #######################################################################
        ### DBアクセス処理(0010)
        ### DBにアクセスして、県別家屋評価額データを取得する。
        #######################################################################
        print_log('[DEBUG] P0200ExcelDownload.house_asset_view()関数 STEP 2/4.', 'DEBUG')
        house_asset_list = HOUSE_ASSET.objects.raw("""
            SELECT 
                HA1.house_asset_code AS house_asset_code, 
                HA1.ken_code AS ken_code, 
                KE1.ken_name AS ken_name, 
                HA1.house_asset AS house_asset 
            FROM HOUSE_ASSET HA1 
            LEFT JOIN KEN KE1 ON HA1.ken_code=KE1.ken_code 
            ORDER BY CAST(HA1.house_asset_code AS INTEGER)""", [])
    
        #######################################################################
        ### EXCEL入出力処理(0020)
        ### (1)テンプレート用のEXCELファイルを読み込む。
        ### (2)セルにデータをセットして、ダウンロード用のEXCELファイルを保存する。
        #######################################################################
        print_log('[DEBUG] P0200ExcelDownload.house_asset_view()関数 STEP 3/4.', 'DEBUG')
        template_file_path = 'static/template_house_asset.xlsx'
        download_file_path = 'static/download_house_asset.xlsx'
        wb = openpyxl.load_workbook(template_file_path)
        ws = wb.active
        ws.title = '家屋被害'
        ws.cell(row=1, column=1).value = '家屋被害コード'
        ws.cell(row=1, column=2).value = '都道府県コード'
        ws.cell(row=1, column=3).value = '都道府県名'
        ws.cell(row=1, column=4).value = '家屋評価額'
        
        if house_asset_list:
            for i, house_asset in enumerate(house_asset_list):
                ws.cell(row=i+2, column=1).value = house_asset.house_asset_code
                ws.cell(row=i+2, column=2).value = house_asset.ken_code
                ws.cell(row=i+2, column=3).value = house_asset.ken_name
                ws.cell(row=i+2, column=4).value = house_asset.house_asset
        
        wb.save(download_file_path)
        
        #######################################################################
        ### レスポンスセット処理(0030)
        ### テンプレートとコンテキストを設定して、レスポンスをブラウザに戻す。
        #######################################################################
        print_log('[DEBUG] P0200ExcelDownload.house_asset_view()関数 STEP 4/4.', 'DEBUG')
        print_log('[INFO] P0200ExcelDownload.house_asset_view()関数が正常終了しました。', 'INFO')
        response = HttpResponse(content=save_virtual_workbook(wb), content_type='application/vnd.ms-excel')
        response['Content-Disposition'] = 'attachment; filename="house_asset.xlsx"'
        return response
        
    except:
        print_log('[ERROR] P0200ExcelDownload.house_asset_view()関数 {}'.format(sys.exc_info()[0]), 'ERROR')
        print_log('[ERROR] P0200ExcelDownload.house_asset_view()関数でエラーが発生しました。', 'ERROR')
        print_log('[ERROR] P0200ExcelDownload.house_asset_view()関数が異常終了しました。', 'ERROR')
        return render(request, 'error.html')

###############################################################################
### 関数名：house_rate_view(request, lock)
### 2010：家屋被害率
### urlpattern：path('house_rate/', views.house_rate_view, name='house_rate_view'
### template：
###############################################################################
### @login_required(None, login_url='/P0100Login/')
def house_rate_view(request, lock):
    try:
        #######################################################################
        ### 引数チェック処理(0000)
        ### ブラウザからのリクエストと引数をチェックする。
        #######################################################################
        ### reset_log()
        print_log('[INFO] P0200ExcelDownload.house_rate_view()関数が開始しました。', 'INFO')
        print_log('[DEBUG] P0200ExcelDownload.house_rate_view()関数 request = {}'.format(request.method), 'DEBUG')
        print_log('[DEBUG] P0200ExcelDownload.house_rate_view()関数 lock = {}'.format(lock), 'DEBUG')
        print_log('[DEBUG] P0200ExcelDownload.house_rate_view()関数 STEP 1/4.', 'DEBUG')
        
        #######################################################################
        ### DBアクセス処理(0010)
        ### DBにアクセスして、家屋被害率データを取得する。
        #######################################################################
        print_log('[DEBUG] P0200ExcelDownload.house_rate_view()関数 STEP 2/4.', 'DEBUG')
        house_rate_list = HOUSE_RATE.objects.raw("""
            SELECT 
                HR1.house_rate_code AS house_rate_code, 
                HR1.flood_sediment_code AS flood_sediment_code, 
                flood_sediment_name AS flood_sediment_name, 
                HR1.gradient_code AS gradient_code, 
                gradient_name AS gradient_name, 
                HR1.house_rate_lv00 AS house_rate_lv00, 
                HR1.house_rate_lv00_50 AS house_rate_lv00_50, 
                HR1.house_rate_lv50_100 AS house_rate_lv50_100, 
                HR1.house_rate_lv100_200 AS house_rate_lv100_200, 
                HR1.house_rate_lv200_300 AS house_rate_lv200_300 
            FROM HOUSE_RATE HR1 
            LEFT JOIN FLOOD_SEDIMENT FS1 ON HR1.flood_sediment_code=FS1.flood_sediment_code 
            LEFT JOIN GRADIENT GR1 ON HR1.gradient_code=GR1.gradient_code 
            ORDER BY CAST(HR1.house_rate_code AS INTEGER)""", [])
    
        #######################################################################
        ### EXCEL入出力処理(0020)
        ### (1)テンプレート用のEXCELファイルを読み込む。
        ### (2)セルにデータをセットして、ダウンロード用のEXCELファイルを保存する。
        #######################################################################
        print_log('[DEBUG] P0200ExcelDownload.house_rate_view()関数 STEP 3/4.', 'DEBUG')
        template_file_path = 'static/template_house_rate.xlsx'
        download_file_path = 'static/download_house_rate.xlsx'
        wb = openpyxl.load_workbook(template_file_path)
        ws = wb.active
        ws.title = '家屋被害率'
        ws.cell(row=1, column=1).value = '家屋被害率コード'
        ws.cell(row=1, column=2).value = '浸水土砂区分コード'
        ws.cell(row=1, column=3).value = '浸水土砂区分名'
        ws.cell(row=1, column=4).value = '地盤勾配区分コード'
        ws.cell(row=1, column=5).value = '地盤勾配区分名'
        ws.cell(row=1, column=6).value = '家屋被害率_床下'
        ws.cell(row=1, column=7).value = '家屋被害率_0から50cm未満'
        ws.cell(row=1, column=8).value = '家屋被害率_50から100cm未満'
        ws.cell(row=1, column=9).value = '家屋被害率_100から200cm未満'
        ws.cell(row=1, column=10).value = '家屋被害率_200から300cm未満'
        ws.cell(row=1, column=11).value = '家屋被害率_300cm以上'
        
        if house_rate_list:
            for i, house_rate in enumerate(house_rate_list):
                ws.cell(row=i+2, column=1).value = house_rate.house_rate_code
                ws.cell(row=i+2, column=2).value = house_rate.flood_sediment_code
                ws.cell(row=i+2, column=3).value = house_rate.flood_sediment_name
                ws.cell(row=i+2, column=4).value = house_rate.gradient_code
                ws.cell(row=i+2, column=5).value = house_rate.gradient_name
                ws.cell(row=i+2, column=6).value = house_rate.house_rate_lv00
                ws.cell(row=i+2, column=7).value = house_rate.house_rate_lv00_50
                ws.cell(row=i+2, column=8).value = house_rate.house_rate_lv50_100
                ws.cell(row=i+2, column=9).value = house_rate.house_rate_lv100_200
                ws.cell(row=i+2, column=10).value = house_rate.house_rate_lv200_300
                ws.cell(row=i+2, column=11).value = house_rate.house_rate_lv300
        
        wb.save(download_file_path)
        
        #######################################################################
        ### レスポンスセット処理(0030)
        ### テンプレートとコンテキストを設定して、レスポンスをブラウザに戻す。
        #######################################################################
        print_log('[DEBUG] P0200ExcelDownload.house_rate_view()関数 STEP 4/4.', 'DEBUG')
        print_log('[INFO] P0200ExcelDownload.house_rate_view()関数が正常終了しました。', 'INFO')
        response = HttpResponse(content=save_virtual_workbook(wb), content_type='application/vnd.ms-excel')
        response['Content-Disposition'] = 'attachment; filename="house_rate.xlsx"'
        return response
        
    except:
        print_log('[ERROR] P0200ExcelDownload.house_rate_view()関数 {}'.format(sys.exc_info()[0]), 'ERROR')
        print_log('[ERROR] P0200ExcelDownload.house_rate_view()関数でエラーが発生しました。', 'ERROR')
        print_log('[ERROR] P0200ExcelDownload.house_rate_view()関数が異常終了しました。', 'ERROR')
        return render(request, 'error.html')

###############################################################################
### 関数名：house_alt_view(request, lock)
### 2020：家庭応急対策費_代替活動費
### urlpattern：path('house_alt/', views.house_alt_view, name='house_alt_view')
### template：
###############################################################################
### @login_required(None, login_url='/P0100Login/')
def house_alt_view(request, lock):
    try:
        #######################################################################
        ### 引数チェック処理(0000)
        ### ブラウザからのリクエストと引数をチェックする。
        #######################################################################
        ### reset_log()
        print_log('[INFO] P0200ExcelDownload.house_alt_view()関数が開始しました。', 'INFO')
        print_log('[DEBUG] P0200ExcelDownload.house_alt_view()関数 request = {}'.format(request.method), 'DEBUG')
        print_log('[DEBUG] P0200ExcelDownload.house_alt_view()関数 lock = {}'.format(lock), 'DEBUG')
        print_log('[DEBUG] P0200ExcelDownload.house_alt_view()関数 STEP 1/4.', 'DEBUG')
        
        #######################################################################
        ### DBアクセス処理(0010)
        ### DBにアクセスして、家庭応急対策費データを取得する。
        #######################################################################
        print_log('[DEBUG] P0200ExcelDownload.house_alt_view()関数 STEP 2/4.', 'DEBUG')
        house_alt_list = HOUSE_ALT.objects.raw("""SELECT * FROM HOUSE_ALT ORDER BY CAST(HOUSE_ALT_CODE AS INTEGER)""", [])
    
        #######################################################################
        ### EXCEL入出力処理(0020)
        ### (1)テンプレート用のEXCELファイルを読み込む。
        ### (2)セルにデータをセットして、ダウンロード用のEXCELファイルを保存する。
        #######################################################################
        print_log('[DEBUG] P0200ExcelDownload.house_alt_view()関数 STEP 3/4.', 'DEBUG')
        template_file_path = 'static/template_house_alt.xlsx'
        download_file_path = 'static/download_house_alt.xlsx'
        wb = openpyxl.load_workbook(template_file_path)
        ws = wb.active
        ws.title = '家庭応急対策費_代替活動費'
        ws.cell(row=1, column=1).value = '家庭応急対策費_代替活動費コード'
        ws.cell(row=1, column=2).value = '家庭応急対策費_代替活動費_床下'
        ws.cell(row=1, column=3).value = '家庭応急対策費_代替活動費_0から50cm未満'
        ws.cell(row=1, column=4).value = '家庭応急対策費_代替活動費_50から100cm未満'
        ws.cell(row=1, column=5).value = '家庭応急対策費_代替活動費_100から200cm未満'
        ws.cell(row=1, column=6).value = '家庭応急対策費_代替活動費_200から300cm未満'
        ws.cell(row=1, column=7).value = '家庭応急対策費_代替活動費_300cm以上'
        
        if house_alt_list:
            for i, house_alt in enumerate(house_alt_list):
                ws.cell(row=i+2, column=1).value = house_alt.house_alt_code
                ws.cell(row=i+2, column=2).value = house_alt.house_alt_lv00
                ws.cell(row=i+2, column=3).value = house_alt.house_alt_lv00_50
                ws.cell(row=i+2, column=4).value = house_alt.house_alt_lv50_100
                ws.cell(row=i+2, column=5).value = house_alt.house_alt_lv100_200
                ws.cell(row=i+2, column=6).value = house_alt.house_alt_lv200_300
                ws.cell(row=i+2, column=7).value = house_alt.house_alt_lv300
        
        wb.save(download_file_path)
        
        #######################################################################
        ### レスポンスセット処理(0030)
        ### テンプレートとコンテキストを設定して、レスポンスをブラウザに戻す。
        #######################################################################
        print_log('[DEBUG] P0200ExcelDownload.house_alt_view()関数 STEP 4/4.', 'DEBUG')
        print_log('[INFO] P0200ExcelDownload.house_alt_view()関数が正常終了しました。', 'INFO')
        response = HttpResponse(content=save_virtual_workbook(wb), content_type='application/vnd.ms-excel')
        response['Content-Disposition'] = 'attachment; filename="household_alt.xlsx"'
        return response
        
    except:
        print_log('[ERROR] P0200ExcelDownload.house_alt_view()関数 {}'.format(sys.exc_info()[0]), 'ERROR')
        print_log('[ERROR] P0200ExcelDownload.house_alt_view()関数でエラーが発生しました。', 'ERROR')
        print_log('[ERROR] P0200ExcelDownload.house_alt_view()関数が異常終了しました。', 'ERROR')
        return render(request, 'error.html')

###############################################################################
### 関数名：house_clean_view(request, lock)
### 2030：家庭応急対策費_清掃日数、清掃労働単価
### urlpattern：path('house_clean/', views.house_clean_view, name='house_clean_view')
### template：
###############################################################################
### @login_required(None, login_url='/P0100Login/')
def house_clean_view(request, lock):
    try:
        #######################################################################
        ### 引数チェック処理(0000)
        ### ブラウザからのリクエストと引数をチェックする。
        #######################################################################
        ### reset_log()
        print_log('[INFO] P0200ExcelDownload.house_clean_view()関数が開始しました。', 'INFO')
        print_log('[DEBUG] P0200ExcelDownload.house_clean_view()関数 request = {}'.format(request.method), 'DEBUG')
        print_log('[DEBUG] P0200ExcelDownload.house_clean_view()関数 lock = {}'.format(lock), 'DEBUG')
        print_log('[DEBUG] P0200ExcelDownload.house_clean_view()関数 STEP 1/4.', 'DEBUG')
        
        #######################################################################
        ### DBアクセス処理(0010)
        ### DBにアクセスして、家庭応急対策費_清掃日数、清掃労働単価データを取得する。
        #######################################################################
        print_log('[DEBUG] P0200ExcelDownload.house_clean_view()関数 STEP 2/4.', 'DEBUG')
        house_clean_list = HOUSE_CLEAN.objects.raw("""SELECT * FROM HOUSE_CLEAN ORDER BY CAST(HOUSE_CLEAN_CODE AS INTEGER)""", [])
    
        #######################################################################
        ### EXCEL入出力処理(0020)
        ### (1)テンプレート用のEXCELファイルを読み込む。
        ### (2)セルにデータをセットして、ダウンロード用のEXCELファイルを保存する。
        #######################################################################
        print_log('[DEBUG] P0200ExcelDownload.house_clean_view()関数 STEP 3/4.', 'DEBUG')
        template_file_path = 'static/template_house_clean.xlsx'
        download_file_path = 'static/download_house_clean.xlsx'
        wb = openpyxl.load_workbook(template_file_path)
        ws = wb.active
        ws.title = '家庭応急対策費_清掃日数'
        ws.cell(row=1, column=1).value = '家庭応急対策費_清掃日数コード'
        ws.cell(row=1, column=2).value = '家庭応急対策費_清掃日数_床下'
        ws.cell(row=1, column=3).value = '家庭応急対策費_清掃日数_0から50cm未満'
        ws.cell(row=1, column=4).value = '家庭応急対策費_清掃日数_50から100cm未満'
        ws.cell(row=1, column=5).value = '家庭応急対策費_清掃日数_100から200cm未満'
        ws.cell(row=1, column=6).value = '家庭応急対策費_清掃日数_200から300cm未満'
        ws.cell(row=1, column=7).value = '家庭応急対策費_清掃日数_300cm以上'
        ws.cell(row=1, column=8).value = '家庭応急対策費_清掃労働単価'
        
        if house_clean_list:
            for i, house_clean in enumerate(house_clean_list):
                ws.cell(row=i+2, column=1).value = house_clean.house_clean_code
                ws.cell(row=i+2, column=2).value = house_clean.house_clean_days_lv00
                ws.cell(row=i+2, column=3).value = house_clean.house_clean_days_lv00_50
                ws.cell(row=i+2, column=4).value = house_clean.house_clean_days_lv50_100
                ws.cell(row=i+2, column=5).value = house_clean.house_clean_days_lv100_200
                ws.cell(row=i+2, column=6).value = house_clean.house_clean_days_lv200_300
                ws.cell(row=i+2, column=7).value = house_clean.house_clean_days_lv300
                ws.cell(row=i+2, column=8).value = house_clean.house_clean_unit_cost
        
        wb.save(download_file_path)
        
        #######################################################################
        ### レスポンスセット処理(0030)
        ### テンプレートとコンテキストを設定して、レスポンスをブラウザに戻す。
        #######################################################################
        print_log('[DEBUG] P0200ExcelDownload.house_clean_view()関数 STEP 4/4.', 'DEBUG')
        print_log('[INFO] P0200ExcelDownload.house_clean_view()関数が正常終了しました。', 'INFO')
        response = HttpResponse(content=save_virtual_workbook(wb), content_type='application/vnd.ms-excel')
        response['Content-Disposition'] = 'attachment; filename="house_clean.xlsx"'
        return response
        
    except:
        print_log('[ERROR] P0200ExcelDownload.house_clean_view()関数 {}'.format(sys.exc_info()[0]), 'ERROR')
        print_log('[ERROR] P0200ExcelDownload.house_clean_view()関数でエラーが発生しました。', 'ERROR')
        print_log('[ERROR] P0200ExcelDownload.house_clean_view()関数が異常終了しました。', 'ERROR')
        return render(request, 'error.html')

###############################################################################
### 関数名：household_asset_view(request, lock)
### 3000：家庭用品自動車以外所有額
### urlpattern：path('household_asset/', views.household_asset_view, name='household_asset_view')
### template：
###############################################################################
### @login_required(None, login_url='/P0100Login/')
def household_asset_view(request, lock):
    try:
        #######################################################################
        ### 引数チェック処理(0000)
        ### ブラウザからのリクエストと引数をチェックする。
        #######################################################################
        ### reset_log()
        print_log('[INFO] P0200ExcelDownload.household_asset_view()関数が開始しました。', 'INFO')
        print_log('[DEBUG] P0200ExcelDownload.household_asset_view()関数 request = {}'.format(request.method), 'DEBUG')
        print_log('[DEBUG] P0200ExcelDownload.household_asset_view()関数 lock = {}'.format(lock), 'DEBUG')
        print_log('[DEBUG] P0200ExcelDownload.household_asset_view()関数 STEP 1/4.', 'DEBUG')
        
        #######################################################################
        ### DBアクセス処理(0010)
        ### DBにアクセスして、家庭用品自動車以外所有額データを取得する。
        #######################################################################
        print_log('[DEBUG] P0200ExcelDownload.household_asset_view()関数 STEP 2/4.', 'DEBUG')
        household_asset_list = HOUSEHOLD_ASSET.objects.raw("""SELECT * FROM HOUSEHOLD_ASSET ORDER BY CAST(HOUSEHOLD_ASSET_CODE AS INTEGER)""", [])
    
        #######################################################################
        ### EXCEL入出力処理(0020)
        ### (1)テンプレート用のEXCELファイルを読み込む。
        ### (2)セルにデータをセットして、ダウンロード用のEXCELファイルを保存する。
        #######################################################################
        print_log('[DEBUG] P0200ExcelDownload.household_asset_view()関数 STEP 3/4.', 'DEBUG')
        template_file_path = 'static/template_household_asset.xlsx'
        download_file_path = 'static/download_household_asset.xlsx'
        wb = openpyxl.load_workbook(template_file_path)
        ws = wb.active
        ws.title = '家庭用品自動車以外所有額'
        ws.cell(row=1, column=1).value = '家庭用品自動車以外所有額コード'
        ws.cell(row=1, column=2).value = '家庭用品自動車以外所有額'
        
        if household_asset_list:
            for i, household_asset in enumerate(household_asset_list):
                ws.cell(row=i+2, column=1).value = household_asset.household_asset_code
                ws.cell(row=i+2, column=2).value = household_asset.household_asset
        
        wb.save(download_file_path)
        
        #######################################################################
        ### レスポンスセット処理(0030)
        ### テンプレートとコンテキストを設定して、レスポンスをブラウザに戻す。
        #######################################################################
        print_log('[DEBUG] P0200ExcelDownload.household_asset_view()関数 STEP 4/4.', 'DEBUG')
        print_log('[INFO] P0200ExcelDownload.household_asset_view()関数が正常終了しました。', 'INFO')
        response = HttpResponse(content=save_virtual_workbook(wb), content_type='application/vnd.ms-excel')
        response['Content-Disposition'] = 'attachment; filename="household_asset.xlsx"'
        return response
        
    except:
        print_log('[ERROR] P0200ExcelDownload.household_asset_view()関数 {}'.format(sys.exc_info()[0]), 'ERROR')
        print_log('[ERROR] P0200ExcelDownload.household_asset_view()関数でエラーが発生しました。', 'ERROR')
        print_log('[ERROR] P0200ExcelDownload.household_asset_view()関数が異常終了しました。', 'ERROR')
        return render(request, 'error.html')

###############################################################################
### 関数名：household_rate_view(request, lock)
### 3010：家庭用品自動車以外被害率
### urlpattern：path('household_rate/', views.household_rate_view, name='household_rate_view')
### template：
###############################################################################
### @login_required(None, login_url='/P0100Login/')
def household_rate_view(request, lock):
    try:
        #######################################################################
        ### 引数チェック処理(0000)
        ### ブラウザからのリクエストと引数をチェックする。
        #######################################################################
        ### reset_log()
        print_log('[INFO] P0200ExcelDownload.household_rate_view()関数が開始しました。', 'INFO')
        print_log('[DEBUG] P0200ExcelDownload.household_rate_view()関数 request = {}'.format(request.method), 'DEBUG')
        print_log('[DEBUG] P0200ExcelDownload.household_rate_view()関数 lock = {}'.format(lock), 'DEBUG')
        print_log('[DEBUG] P0200ExcelDownload.household_rate_view()関数 STEP 1/4.', 'DEBUG')
        
        #######################################################################
        ### DBアクセス処理(0010)
        ### DBにアクセスして、家庭用品自動車以外被害率データを取得する。
        #######################################################################
        print_log('[DEBUG] P0200ExcelDownload.household_rate_view()関数 STEP 2/4.', 'DEBUG')
        household_rate_list = HOUSEHOLD_RATE.objects.raw("""
            SELECT 
                HR1.househole_rate_code AS household_rate_code, 
                HR1.flood_sediment_code AS flood_sediment_code, 
                FS1.flood_sediment_name AS flood_sediment_name, 
                HR1.household_rate_lv00 AS household_rate_lv00, 
                HR1.household_rate_lv00_50 AS household_rate_lv00_50, 
                HR1.household_rate_lv50_100 AS household_rate_lv50_100, 
                HR1.household_rate_lv100_200 AS household_rate_lv100_200, 
                HR1.household_rate_lv200_300 AS household_rate_lv200_300, 
                HR1.household_rate_lv300 AS household_rate_lv300 
            FROM HOUSEHOLD_RATE HR1 
            LEFT JOIN FLOOD_SEDIMENT FS1 ON HR1.flood_sediment_code=FS1.flood_sediment_code 
            ORDER BY CAST(HR1.household_rate_code AS INTEGER)""", [])
    
        #######################################################################
        ### EXCEL入出力処理(0020)
        ### (1)テンプレート用のEXCELファイルを読み込む。
        ### (2)セルにデータをセットして、ダウンロード用のEXCELファイルを保存する。
        #######################################################################
        print_log('[DEBUG] P0200ExcelDownload.household_rate_view()関数 STEP 3/4.', 'DEBUG')
        template_file_path = 'static/template_household_rate.xlsx'
        download_file_path = 'static/download_household_rate.xlsx'
        wb = openpyxl.load_workbook(template_file_path)
        ws = wb.active
        ws.title = '家庭用品自動車以外被害率'
        ws.cell(row=1, column=1).value = '家庭用品自動車以外被害率コード'
        ws.cell(row=1, column=2).value = '浸水土砂区分コード'
        ws.cell(row=1, column=3).value = '浸水土砂区分名'
        ws.cell(row=1, column=4).value = '家庭用品自動車以外被害率_床下'
        ws.cell(row=1, column=5).value = '家庭用品自動車以外被害率_0から50cm未満'
        ws.cell(row=1, column=6).value = '家庭用品自動車以外被害率_50から100cm未満'
        ws.cell(row=1, column=7).value = '家庭用品自動車以外被害率_100から200cm未満'
        ws.cell(row=1, column=8).value = '家庭用品自動車以外被害率_200から300cm未満'
        ws.cell(row=1, column=9).value = '家庭用品自動車以外被害率_300cm以上'
        
        if household_rate_list:
            for i, household_rate in enumerate(household_rate_list):
                ws.cell(row=i+2, column=1).value = household_rate.household_rate_code
                ws.cell(row=i+2, column=2).value = household_rate.flood_sediment_code
                ws.cell(row=i+2, column=3).value = household_rate.flood_sediment_name
                ws.cell(row=i+2, column=4).value = household_rate.household_rate_lv00
                ws.cell(row=i+2, column=5).value = household_rate.household_rate_lv00_50
                ws.cell(row=i+2, column=6).value = household_rate.household_rate_lv50_100
                ws.cell(row=i+2, column=7).value = household_rate.household_rate_lv100_200
                ws.cell(row=i+2, column=8).value = household_rate.household_rate_lv200_300
                ws.cell(row=i+2, column=9).value = household_rate.household_rate_lv300
        
        wb.save(download_file_path)
        
        #######################################################################
        ### レスポンスセット処理(0030)
        ### テンプレートとコンテキストを設定して、レスポンスをブラウザに戻す。
        #######################################################################
        print_log('[DEBUG] P0200ExcelDownload.household_rate_view()関数 STEP 4/4.', 'DEBUG')
        print_log('[INFO] P0200ExcelDownload.household_rate_view()関数が正常終了しました。', 'INFO')
        response = HttpResponse(content=save_virtual_workbook(wb), content_type='application/vnd.ms-excel')
        response['Content-Disposition'] = 'attachment; filename="household_rate.xlsx"'
        return response
        
    except:
        print_log('[ERROR] P0200ExcelDownload.household_rate_view()関数 {}'.format(sys.exc_info()[0]), 'ERROR')
        print_log('[ERROR] P0200ExcelDownload.household_rate_view()関数でエラーが発生しました。', 'ERROR')
        print_log('[ERROR] P0200ExcelDownload.household_rate_view()関数が異常終了しました。', 'ERROR')
        return render(request, 'error.html')

###############################################################################
### 関数名：car_asset_view(request, lock)
### 4000：家庭用品自動車所有額
### urlpattern：path('car_asset/', views.car_asset_view, name='car_asset_view')
### template：
###############################################################################
### @login_required(None, login_url='/P0100Login/')
def car_asset_view(request, lock):
    try:
        #######################################################################
        ### 引数チェック処理(0000)
        ### ブラウザからのリクエストと引数をチェックする。
        #######################################################################
        ### reset_log()
        print_log('[INFO] P0200ExcelDownload.car_asset_view()関数が開始しました。', 'INFO')
        print_log('[DEBUG] P0200ExcelDownload.car_asset_view()関数 request = {}'.format(request.method), 'DEBUG')
        print_log('[DEBUG] P0200ExcelDownload.car_asset_view()関数 lock = {}'.format(lock), 'DEBUG')
        print_log('[DEBUG] P0200ExcelDownload.car_asset_view()関数 STEP 1/4.', 'DEBUG')
        
        #######################################################################
        ### DBアクセス処理(0010)
        ### DBにアクセスして、家庭用品自動車所有額データを取得する。
        #######################################################################
        print_log('[DEBUG] P0200ExcelDownload.car_asset_view()関数 STEP 2/4.', 'DEBUG')
        car_asset_list = CAR_ASSET.objects.raw("""SELECT * FROM CAR_ASSET ORDER BY CAST(CAR_ASSET_CODE AS INTEGER)""", [])
    
        #######################################################################
        ### EXCEL入出力処理(0020)
        ### (1)テンプレート用のEXCELファイルを読み込む。
        ### (2)セルにデータをセットして、ダウンロード用のEXCELファイルを保存する。
        #######################################################################
        print_log('[DEBUG] P0200ExcelDownload.car_asset_view()関数 STEP 3/4.', 'DEBUG')
        template_file_path = 'static/template_car_asset.xlsx'
        download_file_path = 'static/download_car_asset.xlsx'
        wb = openpyxl.load_workbook(template_file_path)
        ws = wb.active
        ws.title = '家庭用品自動車所有額'
        ws.cell(row=1, column=1).value = '家庭用品自動車所有額コード'
        ws.cell(row=1, column=2).value = '家庭用品自動車所有額'
        
        if car_asset_list:
            for i, car_asset in enumerate(car_asset_list):
                ws.cell(row=i+2, column=1).value = car_asset.car_asset_code
                ws.cell(row=i+2, column=2).value = car_asset.car_asset
        
        wb.save(download_file_path)
        
        #######################################################################
        ### レスポンスセット処理(0000)
        ### (1)テンプレートとコンテキストを設定して、レスポンスをブラウザに戻す。
        #######################################################################
        print_log('[DEBUG] P0200ExcelDownload.car_asset_view()関数 STEP 4/4.', 'DEBUG')
        print_log('[INFO] P0200ExcelDownload.car_asset_view()関数が正常終了しました。', 'INFO')
        response = HttpResponse(content=save_virtual_workbook(wb), content_type='application/vnd.ms-excel')
        response['Content-Disposition'] = 'attachment; filename="car_asset.xlsx"'
        return response
        
    except:
        print_log('[ERROR] P0200ExcelDownload.car_asset_view()関数 {}'.format(sys.exc_info()[0]), 'ERROR')
        print_log('[ERROR] P0200ExcelDownload.car_asset_view()関数でエラーが発生しました。', 'ERROR')
        print_log('[ERROR] P0200ExcelDownload.car_asset_view()関数が異常終了しました。', 'ERROR')
        return render(request, 'error.html')

###############################################################################
### 関数名：car_rate_view(request, lock)
### 4010：家庭用品自動車被害率
### urlpattern：path('car_rate/', views.car_rate_view, name='car_rate_view')
### template：
###############################################################################
### @login_required(None, login_url='/P0100Login/')
def car_rate_view(request, lock):
    try:
        #######################################################################
        ### 引数チェック処理(0000)
        ### ブラウザからのリクエストと引数をチェックする。
        #######################################################################
        ### reset_log()
        print_log('[INFO] P0200ExcelDownload.car_rate_view()関数が開始しました。', 'INFO')
        print_log('[DEBUG] P0200ExcelDownload.car_rate_view()関数 request = {}'.format(request.method), 'DEBUG')
        print_log('[DEBUG] P0200ExcelDownload.car_rate_view()関数 lock = {}'.format(lock), 'DEBUG')
        print_log('[DEBUG] P0200ExcelDownload.car_rate_view()関数 STEP 1/4.', 'DEBUG')
        
        #######################################################################
        ### DBアクセス処理(0010)
        ### DBにアクセスして、事業所営業停止損失データを取得する。
        #######################################################################
        print_log('[DEBUG] P0200ExcelDownload.car_rate_view()関数 STEP 2/4.', 'DEBUG')
        car_rate_list = CAR_RATE.objects.raw("""SELECT * FROM CAR_RATE ORDER BY CAST(CAR_RATE_CODE AS INTEGER)""", [])
    
        #######################################################################
        ### EXCEL入出力処理
        ### (1)テンプレート用のEXCELファイルを読み込む。
        ### (2)セルにデータをセットして、ダウンロード用のEXCELファイルを保存する。
        #######################################################################
        print_log('[DEBUG] P0200ExcelDownload.car_rate_view()関数 STEP 3/4.', 'DEBUG')
        template_file_path = 'static/template_car_rate.xlsx'
        download_file_path = 'static/download_car_rate.xlsx'
        wb = openpyxl.load_workbook(template_file_path)
        ws = wb.active
        ws.title = '家庭用品自動車被害率'
        ws.cell(row=1, column=1).value = '家庭用品自動車被害率コード'
        ws.cell(row=1, column=2).value = '家庭用品自動車被害率_床下'
        ws.cell(row=1, column=3).value = '家庭用品自動車被害率_0から50cm未満'
        ws.cell(row=1, column=4).value = '家庭用品自動車被害率_50から100cm未満'
        ws.cell(row=1, column=5).value = '家庭用品自動車被害率_100から200cm未満'
        ws.cell(row=1, column=6).value = '家庭用品自動車被害率_200から300cm未満'
        ws.cell(row=1, column=7).value = '家庭用品自動車被害率_300cm以上'
        
        if car_rate_list:
            for i, car_rate in enumerate(car_rate_list):
                ws.cell(row=i+2, column=1).value = car_rate.car_rate_code
                ws.cell(row=i+2, column=2).value = car_rate.car_rate_lv00
                ws.cell(row=i+2, column=3).value = car_rate.car_rate_lv00_50
                ws.cell(row=i+2, column=4).value = car_rate.car_rate_lv50_100
                ws.cell(row=i+2, column=5).value = car_rate.car_rate_lv100_200
                ws.cell(row=i+2, column=6).value = car_rate.car_rate_lv200_300
                ws.cell(row=i+2, column=7).value = car_rate.car_rate_lv300
        
        wb.save(download_file_path)
        
        #######################################################################
        ### レスポンスセット処理(0030)
        ### テンプレートとコンテキストを設定して、レスポンスをブラウザに戻す。
        #######################################################################
        print_log('[DEBUG] P0200ExcelDownload.car_rate_view()関数 STEP 4/4.', 'DEBUG')
        print_log('[INFO] P0200ExcelDownload.car_rate_view()関数が正常終了しました。', 'INFO')
        response = HttpResponse(content=save_virtual_workbook(wb), content_type='application/vnd.ms-excel')
        response['Content-Disposition'] = 'attachment; filename="car_rate.xlsx"'
        return response
        
    except:
        print_log('[ERROR] P0200ExcelDownload.car_rate_view()関数 {}'.format(sys.exc_info()[0]), 'ERROR')
        print_log('[ERROR] P0200ExcelDownload.car_rate_view()関数でエラーが発生しました。', 'ERROR')
        print_log('[ERROR] P0200ExcelDownload.car_rate_view()関数が異常終了しました。', 'ERROR')
        return render(request, 'error.html')

###############################################################################
### 関数名：office_asset_view(request, lock)
### 5000：事業所資産額
### urlpattern：path('office_asset/', views.office_asset_view, name='office_asset_view')
### template：
###############################################################################
### @login_required(None, login_url='/P0100Login/')
def office_asset_view(request, lock):
    try:
        #######################################################################
        ### 引数チェック処理(0000)
        ### ブラウザからのリクエストと引数をチェックする。
        #######################################################################
        ### reset_log()
        print_log('[INFO] P0200ExcelDownload.office_asset_view()関数が開始しました。', 'INFO')
        print_log('[DEBUG] P0200ExcelDownload.office_asset_view()関数 request = {}'.format(request.method), 'DEBUG')
        print_log('[DEBUG] P0200ExcelDownload.office_asset_view()関数 lock = {}'.format(lock), 'DEBUG')
        print_log('[DEBUG] P0200ExcelDownload.office_asset_view()関数 STEP 1/4.', 'DEBUG')
        
        #######################################################################
        ### DBアクセス処理(0010)
        ### DBにアクセスして、事業所資産額データを取得する。
        #######################################################################
        print_log('[DEBUG] P0200ExcelDownload.office_asset_view()関数 STEP 2/4.', 'DEBUG')
        office_asset_list = OFFICE_ASSET.objects.raw("""
            SELECT 
                OA1.office_asset_code AS office_asset_code, 
                OA1.industry_code AS industry_code, 
                IN1.industry_name AS industry_name, 
                OA1.office_dep_asset AS office_dep_asset, 
                OA1.office_inv_asset AS office_inv_asset, 
                OA1.office_va_asset AS office_va_asset 
            FROM OFFICE_ASSET OA1 
            LEFT JOIN INDUSTRY IN1 ON OA1.industry_code=IN1.industry_code 
            ORDER BY CAST(OA1.office_asset_code AS INTEGER)""", [])
    
        #######################################################################
        ### EXCEL入出力処理(0020)
        ### (1)テンプレート用のEXCELファイルを読み込む。
        ### (2)セルにデータをセットして、ダウンロード用のEXCELファイルを保存する。
        #######################################################################
        print_log('[DEBUG] P0200ExcelDownload.office_asset_view()関数 STEP 3/4.', 'DEBUG')
        template_file_path = 'static/template_office_asset.xlsx'
        download_file_path = 'static/download_office_asset.xlsx'
        wb = openpyxl.load_workbook(template_file_path)
        ws = wb.active
        ws.title = '事業所資産額'
        ws.cell(row=1, column=1).value = '事業所資産額コード'
        ws.cell(row=1, column=2).value = '産業分類コード'
        ws.cell(row=1, column=3).value = '産業分類名'
        ws.cell(row=1, column=4).value = '事業所資産額_償却資産額'
        ws.cell(row=1, column=5).value = '事業所資産額_在庫資産額'
        ws.cell(row=1, column=6).value = '事業所資産額_付加価値額'
        
        if office_asset_list:
            for i, office_asset in enumerate(office_asset_list):
                ws.cell(row=i+2, column=1).value = office_asset.office_asset_code
                ws.cell(row=i+2, column=2).value = office_asset.industry_code
                ws.cell(row=i+2, column=3).value = office_asset.industry_name
                ws.cell(row=i+2, column=4).value = office_asset.office_dep_asset
                ws.cell(row=i+2, column=5).value = office_asset.office_inv_asset
                ws.cell(row=i+2, column=6).value = office_asset.office_va_asset
        
        wb.save(download_file_path)
        
        #######################################################################
        ### レスポンスセット処理(0030)
        ### テンプレートとコンテキストを設定して、レスポンスをブラウザに戻す。
        #######################################################################
        print_log('[DEBUG] P0200ExcelDownload.office_asset_view()関数 STEP 4/4.', 'DEBUG')
        print_log('[INFO] P0200ExcelDownload.office_asset_view()関数が正常終了しました。', 'INFO')
        response = HttpResponse(content=save_virtual_workbook(wb), content_type='application/vnd.ms-excel')
        response['Content-Disposition'] = 'attachment; filename="office_asset.xlsx"'
        return response
        
    except:
        print_log('[ERROR] P0200ExcelDownload.office_asset_view()関数 {}'.format(sys.exc_info()[0]), 'ERROR')
        print_log('[ERROR] P0200ExcelDownload.office_asset_view()関数でエラーが発生しました。', 'ERROR')
        print_log('[ERROR] P0200ExcelDownload.office_asset_view()関数が異常終了しました。', 'ERROR')
        return render(request, 'error.html')

###############################################################################
### 関数名：office_rate_view(request, lock)
### 5010：事業所被害率
### urlpattern：path('office_rate/', views.office_rate_view, name='office_rate_view')
### template：
###############################################################################
### @login_required(None, login_url='/P0100Login/')
def office_rate_view(request, lock):
    try:
        #######################################################################
        ### 引数チェック処理(0000)
        ### ブラウザからのリクエストと引数をチェックする。
        #######################################################################
        ### reset_log()
        print_log('[INFO] P0200ExcelDownload.office_rate_view()関数が開始しました。', 'INFO')
        print_log('[DEBUG] P0200ExcelDownload.office_rate_view()関数 request = {}'.format(request.method), 'DEBUG')
        print_log('[DEBUG] P0200ExcelDownload.office_rate_view()関数 lock = {}'.format(lock), 'DEBUG')
        print_log('[DEBUG] P0200ExcelDownload.office_rate_view()関数 STEP 1/4.', 'DEBUG')
        
        #######################################################################
        ### DBアクセス処理(0010)
        ### DBにアクセスして、事業所被害率データを取得する。
        #######################################################################
        print_log('[DEBUG] P0200ExcelDownload.office_rate_view()関数 STEP 2/4.', 'DEBUG')
        office_rate_list = OFFICE_RATE.objects.raw("""
            SELECT 
                OR1.office_rate_code AS office_rate_code, 
                OR1.flood_sediment_code AS flood_sediment_code, 
                FS1.flood_sediment_name AS flood_sediment_name, 
                OR1.office_dep_rate_lv00 AS office_dep_rate_lv00, 
                OR1.office_dep_rate_lv00_50 AS office_dep_rate_lv00_50, 
                OR1.office_dep_rate_lv50_100 AS office_dep_rate_lv50_100, 
                OR1.office_dep_rate_lv100_200 AS office_dep_rate_lv100_200, 
                OR1.office_dep_rate_lv200_300 AS office_dep_rate_lv200_300, 
                OR1.office_dep_rate_lv300 AS office_dep_rate_lv300, 
                OR1.office_inv_rate_lv00 AS office_inv_rate_lv00, 
                OR1.office_inv_rate_lv00_50 AS office_inv_rate_lv00_50, 
                OR1.office_inv_rate_lv50_100 AS office_inv_rate_lv50_100, 
                OR1.office_inv_rate_lv100_200 AS office_inv_rate_lv100_200, 
                OR1.office_inv_rate_lv200_300 AS office_inv_rate_lv200_300, 
                OR1.office_inv_rate_lv300 AS office_inv_rate_lv300 
            FROM OFFICE_RATE OR1 
            LEFT JOIN FLOOD_SEDIMENT FS1 ON OR1.flood_sediment_code=FS1.flood_sediment_code 
            ORDER BY CAST(OR1.office_rate_code AS INTEGER)""", [])
    
        #######################################################################
        ### EXCEL入出力処理(0020)
        ### (1)テンプレート用のEXCELファイルを読み込む。
        ### (2)セルにデータをセットして、ダウンロード用のEXCELファイルを保存する。
        #######################################################################
        print_log('[DEBUG] P0200ExcelDownload.office_rate_view()関数 STEP 3/4.', 'DEBUG')
        template_file_path = 'static/template_office_rate.xlsx'
        download_file_path = 'static/download_office_rate.xlsx'
        wb = openpyxl.load_workbook(template_file_path)
        ws = wb.active
        ws.title = '事業所被害率'
        ws.cell(row=1, column=1).value = '事業所被害率コード'
        ws.cell(row=1, column=2).value = '浸水土砂区分コード'
        ws.cell(row=1, column=3).value = '浸水土砂区分名'
        ws.cell(row=1, column=4).value = '事業所被害率_償却資産被害率_床下'
        ws.cell(row=1, column=5).value = '事業所被害率_償却資産被害率_0から50cm未満'
        ws.cell(row=1, column=6).value = '事業所被害率_償却資産被害率_50から100cm未満'
        ws.cell(row=1, column=7).value = '事業所被害率_償却資産被害率_100から200cm未満'
        ws.cell(row=1, column=8).value = '事業所被害率_償却資産被害率_200から300cm未満'
        ws.cell(row=1, column=9).value = '事業所被害率_償却資産被害率数_300cm以上'
        ws.cell(row=1, column=10).value = '事業所被害率_在庫資産被害率_床下'
        ws.cell(row=1, column=11).value = '事業所被害率_在庫資産被害率_0から50cm未満'
        ws.cell(row=1, column=12).value = '事業所被害率_在庫資産被害率_50から100cm未満'
        ws.cell(row=1, column=13).value = '事業所被害率_在庫資産被害率_100から200cm未満'
        ws.cell(row=1, column=14).value = '事業所被害率_在庫資産被害率_200から300cm未満'
        ws.cell(row=1, column=15).value = '事業所被害率_在庫資産被害率数_300cm以上'
        
        if office_rate_list:
            for i, office_rate in enumerate(office_rate_list):
                ws.cell(row=i+2, column=1).value = office_rate.office_rate_code
                ws.cell(row=i+2, column=2).value = office_rate.flood_sediment_code
                ws.cell(row=i+2, column=3).value = office_rate.flood_sediment_name
                ws.cell(row=i+2, column=4).value = office_rate.office_dep_rate_lv00
                ws.cell(row=i+2, column=5).value = office_rate.office_dep_rate_lv00_50
                ws.cell(row=i+2, column=6).value = office_rate.office_dep_rate_lv50_100
                ws.cell(row=i+2, column=7).value = office_rate.office_dep_rate_lv100_200
                ws.cell(row=i+2, column=8).value = office_rate.office_dep_rate_lv200_300
                ws.cell(row=i+2, column=9).value = office_rate.office_dep_rate_lv300
                ws.cell(row=i+2, column=10).value = office_rate.office_inv_rate_lv00
                ws.cell(row=i+2, column=11).value = office_rate.office_inv_rate_lv00_50
                ws.cell(row=i+2, column=12).value = office_rate.office_inv_rate_lv50_100
                ws.cell(row=i+2, column=13).value = office_rate.office_inv_rate_lv100_200
                ws.cell(row=i+2, column=14).value = office_rate.office_inv_rate_lv200_300
                ws.cell(row=i+2, column=15).value = office_rate.office_inv_rate_lv300
        
        wb.save(download_file_path)
        
        #######################################################################
        ### レスポンスセット処理(0030)
        ### テンプレートとコンテキストを設定して、レスポンスをブラウザに戻す。
        #######################################################################
        print_log('[DEBUG] P0200ExcelDownload.office_rate_view()関数 STEP 4/4.', 'DEBUG')
        print_log('[INFO] P0200ExcelDownload.office_rate_view()関数が正常終了しました。', 'INFO')
        response = HttpResponse(content=save_virtual_workbook(wb), content_type='application/vnd.ms-excel')
        response['Content-Disposition'] = 'attachment; filename="office_rate.xlsx"'
        return response
        
    except:
        print_log('[ERROR] P0200ExcelDownload.office_rate_view()関数 {}'.format(sys.exc_info()[0]), 'ERROR')
        print_log('[ERROR] P0200ExcelDownload.office_rate_view()関数でエラーが発生しました。', 'ERROR')
        print_log('[ERROR] P0200ExcelDownload.office_rate_view()関数が異常終了しました。', 'ERROR')
        return render(request, 'error.html')

###############################################################################
### 関数名：office_suspend_view(request, lock)
### 5020：事業所営業停止日数
### urlpattern：path('office_suspend/', views.office_suspend_view, name='office_suspend_view')
### template：
###############################################################################
### @login_required(None, login_url='/P0100Login/')
def office_suspend_view(request, lock):
    try:
        #######################################################################
        ### 引数チェック処理(0000)
        ### ブラウザからのリクエストと引数をチェックする。
        #######################################################################
        ### reset_log()
        print_log('[INFO] P0200ExcelDownload.office_suspend_view()関数が開始しました。', 'INFO')
        print_log('[DEBUG] P0200ExcelDownload.office_suspend_view()関数 request = {}'.format(request.method), 'DEBUG')
        print_log('[DEBUG] P0200ExcelDownload.office_suspend_view()関数 lock = {}'.format(lock), 'DEBUG')
        print_log('[DEBUG] P0200ExcelDownload.office_suspend_view()関数 STEP 1/4.', 'DEBUG')
        
        #######################################################################
        ### DBアクセス処理(0010)
        ### DBにアクセスして、事業所営業停止日数データを取得する。
        #######################################################################
        print_log('[DEBUG] P0200ExcelDownload.office_suspend_view()関数 STEP 2/4.', 'DEBUG')
        office_suspend_list = OFFICE_SUSPEND.objects.raw("""SELECT * FROM OFFICE_SUSPEND ORDER BY CAST(OFFICE_SUS_CODE AS INTEGER)""", [])
    
        #######################################################################
        ### EXCEL入出力処理(0020)
        ### (1)テンプレート用のEXCELファイルを読み込む。
        ### (2)セルにデータをセットして、ダウンロード用のEXCELファイルを保存する。
        #######################################################################
        print_log('[DEBUG] P0200ExcelDownload.office_suspend_view()関数 STEP 3/4.', 'DEBUG')
        template_file_path = 'static/template_office_suspend.xlsx'
        download_file_path = 'static/download_office_suspend.xlsx'
        wb = openpyxl.load_workbook(template_file_path)
        ws = wb.active
        ws.title = '事業所営業停止日数'
        ws.cell(row=1, column=1).value = '事業所営業停止日数コード'
        ws.cell(row=1, column=2).value = '事業所営業停止日数_床下'
        ws.cell(row=1, column=3).value = '事業所営業停止日数_0から50cm未満'
        ws.cell(row=1, column=4).value = '事業所営業停止日数_50から100cm未満'
        ws.cell(row=1, column=5).value = '事業所営業停止日数_100から200cm未満'
        ws.cell(row=1, column=6).value = '事業所営業停止日数_200から300cm未満'
        ws.cell(row=1, column=7).value = '事業所営業停止日数_300cm以上'
        
        if office_suspend_list:
            for i, office_suspend in enumerate(office_suspend_list):
                ws.cell(row=i+2, column=1).value = office_suspend.office_sus_code
                ws.cell(row=i+2, column=2).value = office_suspend.office_sus_days_lv00
                ws.cell(row=i+2, column=3).value = office_suspend.office_sus_days_lv00_50
                ws.cell(row=i+2, column=4).value = office_suspend.office_sus_days_lv50_100
                ws.cell(row=i+2, column=5).value = office_suspend.office_sus_days_lv100_200
                ws.cell(row=i+2, column=6).value = office_suspend.office_sus_days_lv200_300
                ws.cell(row=i+2, column=7).value = office_suspend.office_sus_days_lv300
        
        wb.save(download_file_path)
        
        #######################################################################
        ### レスポンスセット処理(0030)
        ### テンプレートとコンテキストを設定して、レスポンスをブラウザに戻す。
        #######################################################################
        print_log('[DEBUG] P0200ExcelDownload.office_suspend_view()関数 STEP 4/4.', 'DEBUG')
        print_log('[INFO] P0200ExcelDownload.office_suspend_view()関数が正常終了しました。', 'INFO')
        response = HttpResponse(content=save_virtual_workbook(wb), content_type='application/vnd.ms-excel')
        response['Content-Disposition'] = 'attachment; filename="office_suspend.xlsx"'
        return response
        
    except:
        print_log('[ERROR] P0200ExcelDownload.office_suspend_view()関数 {}'.format(sys.exc_info()[0]), 'ERROR')
        print_log('[ERROR] P0200ExcelDownload.office_suspend_view()関数でエラーが発生しました。', 'ERROR')
        print_log('[ERROR] P0200ExcelDownload.office_suspend_view()関数が異常終了しました。', 'ERROR')
        return render(request, 'error.html')

###############################################################################
### 関数名：office_stagnate_view(request, lock)
### 5030：事業所営業停滞日数
### urlpattern：path('office_stagnate/', views.office_stagnate_view, name='office_stagnate_view')
### template：
###############################################################################
### @login_required(None, login_url='/P0100Login/')
def office_stagnate_view(request, lock):
    try:
        #######################################################################
        ### 引数チェック処理(0000)
        ### ブラウザからのリクエストと引数をチェックする。
        #######################################################################
        ### reset_log()
        print_log('[INFO] P0200ExcelDownload.office_stagnate_view()関数が開始しました。', 'INFO')
        print_log('[DEBUG] P0200ExcelDownload.office_stagnate_view()関数 request = {}'.format(request.method), 'DEBUG')
        print_log('[DEBUG] P0200ExcelDownload.office_stagnate_view()関数 lock = {}'.format(lock), 'DEBUG')
        print_log('[DEBUG] P0200ExcelDownload.office_stagnate_view()関数 STEP 1/4.', 'DEBUG')
        
        #######################################################################
        ### DBアクセス処理(0010)
        ### DBにアクセスして、事業所営業停滞日数データを取得する。
        #######################################################################
        print_log('[DEBUG] P0200ExcelDownload.office_stagnate_view()関数 STEP 2/4.', 'DEBUG')
        office_stagnate_list = OFFICE_STAGNATE.objects.raw("""SELECT * FROM OFFICE_STAGNATE ORDER BY CAST(OFFICE_STG_CODE AS INTEGER)""", [])
    
        #######################################################################
        ### EXCEL入出力処理(0020)
        ### (1)テンプレート用のEXCELファイルを読み込む。
        ### (2)セルにデータをセットして、ダウンロード用のEXCELファイルを保存する。
        #######################################################################
        print_log('[DEBUG] P0200ExcelDownload.office_stagnate_view()関数 STEP 3/4.', 'DEBUG')
        template_file_path = 'static/template_office_stagnate.xlsx'
        download_file_path = 'static/download_office_stagnate.xlsx'
        wb = openpyxl.load_workbook(template_file_path)
        ws = wb.active
        ws.title = '事業所営業停滞日数'
        ws.cell(row=1, column=1).value = '事業所営業停滞日数コード'
        ws.cell(row=1, column=2).value = '事業所営業停滞日数_床下'
        ws.cell(row=1, column=3).value = '事業所営業停滞日数_0から50cm未満'
        ws.cell(row=1, column=4).value = '事業所営業停滞日数_50から100cm未満'
        ws.cell(row=1, column=5).value = '事業所営業停滞日数_100から200cm未満'
        ws.cell(row=1, column=6).value = '事業所営業停滞日数_200から300cm未満'
        ws.cell(row=1, column=7).value = '事業所営業停滞日数_300cm以上'
        
        if office_stagnate_list:
            for i, office_stagnate in enumerate(office_stagnate_list):
                ws.cell(row=i+2, column=1).value = office_stagnate.office_stg_code
                ws.cell(row=i+2, column=2).value = office_stagnate.office_stg_days_lv00
                ws.cell(row=i+2, column=3).value = office_stagnate.office_stg_days_lv00_50
                ws.cell(row=i+2, column=4).value = office_stagnate.office_stg_days_lv50_100
                ws.cell(row=i+2, column=5).value = office_stagnate.office_stg_days_lv100_200
                ws.cell(row=i+2, column=6).value = office_stagnate.office_stg_days_lv200_300
                ws.cell(row=i+2, column=7).value = office_stagnate.office_stg_days_lv300
        
        wb.save(download_file_path)
        
        #######################################################################
        ### レスポンスセット処理(0030)
        ### テンプレートとコンテキストを設定して、レスポンスをブラウザに戻す。
        #######################################################################
        print_log('[DEBUG] P0200ExcelDownload.office_stagnate_view()関数 STEP 4/4.', 'DEBUG')
        print_log('[INFO] P0200ExcelDownload.office_stagnate_view()関数が正常終了しました。', 'INFO')
        response = HttpResponse(content=save_virtual_workbook(wb), content_type='application/vnd.ms-excel')
        response['Content-Disposition'] = 'attachment; filename="office_stagnate.xlsx"'
        return response
        
    except:
        print_log('[ERROR] P0200ExcelDownload.office_stagnate_view()関数 {}'.format(sys.exc_info()[0]), 'ERROR')
        print_log('[ERROR] P0200ExcelDownload.office_stagnate_view()関数でエラーが発生しました。', 'ERROR')
        print_log('[ERROR] P0200ExcelDownload.office_stagnate_view()関数が異常終了しました。', 'ERROR')
        return render(request, 'error.html')

###############################################################################
### 関数名：office_alt_view(request, lock)
### 5040：事業所応急対策費_代替活動費
### urlpattern：
### template：
###############################################################################
### @login_required(None, login_url='/P0100Login/')
def office_alt_view(request, lock):
    try:
        #######################################################################
        ### 引数チェック処理(0000)
        ### ブラウザからのリクエストと引数をチェックする。
        #######################################################################
        ### reset_log()
        print_log('[INFO] P0200ExcelDownload.office_alt_view()関数が開始しました。', 'INFO')
        print_log('[DEBUG] P0200ExcelDownload.office_alt_view()関数 request = {}'.format(request.method), 'DEBUG')
        print_log('[DEBUG] P0200ExcelDownload.office_alt_view()関数 lock = {}'.format(lock), 'DEBUG')
        print_log('[DEBUG] P0200ExcelDownload.office_alt_view()関数 STEP 1/4.', 'DEBUG')
        
        #######################################################################
        ### DBアクセス処理(0010)
        ### DBにアクセスして、事業所応急対策費_代替活動費データを取得する。
        #######################################################################
        print_log('[DEBUG] P0200ExcelDownload.office_alt_view()関数 STEP 2/4.', 'DEBUG')
        office_alt_list = OFFICE_ALT.objects.raw("""SELECT * FROM OFFICE_ALT ORDER BY CAST(OFFICE_ALT_CODE AS INTEGER)""", [])
    
        #######################################################################
        ### EXCEL入出力処理(0020)
        ### (1)テンプレート用のEXCELファイルを読み込む。
        ### (2)セルにデータをセットして、ダウンロード用のEXCELファイルを保存する。
        #######################################################################
        print_log('[DEBUG] P0200ExcelDownload.office_alt_view()関数 STEP 3/4.', 'DEBUG')
        template_file_path = 'static/template_office_alt.xlsx'
        download_file_path = 'static/download_office_alt.xlsx'
        wb = openpyxl.load_workbook(template_file_path)
        ws = wb.active
        ws.title = '事業所応急対策費_代替活動費'
        ws.cell(row=1, column=1).value = '事業所応急対策費_代替活動費コード'
        ws.cell(row=1, column=2).value = '事業所応急対策費_代替活動費_床下'
        ws.cell(row=1, column=3).value = '事業所応急対策費_代替活動費_0から50cm未満'
        ws.cell(row=1, column=4).value = '事業所応急対策費_代替活動費_50から100cm未満'
        ws.cell(row=1, column=5).value = '事業所応急対策費_代替活動費_100から200cm未満'
        ws.cell(row=1, column=6).value = '事業所応急対策費_代替活動費_200から300cm未満'
        ws.cell(row=1, column=7).value = '事業所応急対策費_代替活動費_300cm以上'
        
        if office_alt_list:
            for i, office_alt in enumerate(office_alt_list):
                ws.cell(row=i+2, column=1).value = office_alt.office_alt_code
                ws.cell(row=i+2, column=2).value = office_alt.office_alt_lv00
                ws.cell(row=i+2, column=3).value = office_alt.office_alt_lv00_50
                ws.cell(row=i+2, column=4).value = office_alt.office_alt_lv50_100
                ws.cell(row=i+2, column=5).value = office_alt.office_alt_lv100_200
                ws.cell(row=i+2, column=6).value = office_alt.office_alt_lv200_300
                ws.cell(row=i+2, column=7).value = office_alt.office_alt_lv300
        
        wb.save(download_file_path)
        
        #######################################################################
        ### レスポンスセット処理(0030)
        ### テンプレートとコンテキストを設定して、レスポンスをブラウザに戻す。
        #######################################################################
        print_log('[DEBUG] P0200ExcelDownload.office_alt_view()関数 STEP 4/4.', 'DEBUG')
        print_log('[INFO] P0200ExcelDownload.office_alt_view()関数が正常終了しました。', 'INFO')
        response = HttpResponse(content=save_virtual_workbook(wb), content_type='application/vnd.ms-excel')
        response['Content-Disposition'] = 'attachment; filename="office_alt.xlsx"'
        return response
        
    except:
        print_log('[ERROR] P0200ExcelDownload.office_alt_view()関数 {}'.format(sys.exc_info()[0]), 'ERROR')
        print_log('[ERROR] P0200ExcelDownload.office_alt_view()関数でエラーが発生しました。', 'ERROR')
        print_log('[ERROR] P0200ExcelDownload.office_alt_view()関数が異常終了しました。', 'ERROR')
        return render(request, 'error.html')

###############################################################################
### 関数名：farmer_fisher_asset_view(request, lock)
### 6000：農漁家資産額
### urlpattern：path('farmer_fisher_asset/', views.farmer_fisher_asset_view, name='farmer_fisher_asset_view')
### template：
###############################################################################
### @login_required(None, login_url='/P0100Login/')
def farmer_fisher_asset_view(request, lock):
    try:
        #######################################################################
        ### 引数チェック処理(0000)
        ### ブラウザからのリクエストと引数をチェックする。
        #######################################################################
        ### reset_log()
        print_log('[INFO] P0200ExcelDownload.farmer_fisher_asset_view()関数が開始しました。', 'INFO')
        print_log('[DEBUG] P0200ExcelDownload.farmer_fisher_asset_view()関数 request = {}'.format(request.method), 'DEBUG')
        print_log('[DEBUG] P0200ExcelDownload.farmer_fisher_asset_view()関数 lock = {}'.format(lock), 'DEBUG')
        print_log('[DEBUG] P0200ExcelDownload.farmer_fisher_asset_view()関数 STEP 1/4.', 'DEBUG')
        
        #######################################################################
        ### DBアクセス処理(0010)
        ### DBにアクセスして、農漁家資産額データを取得する。
        #######################################################################
        print_log('[DEBUG] P0200ExcelDownload.farmer_fisher_asset_view()関数 STEP 2/4.', 'DEBUG')
        farmer_fisher_asset_list = FARMER_FISHER_ASSET.objects.raw("""SELECT * FROM FARMER_FISHER_ASSET ORDER BY CAST(FARMER_FISHER_ASSET_CODE AS INTEGER)""", [])
    
        #######################################################################
        ### EXCEL入出力処理(0020)
        ### (1)テンプレート用のEXCELファイルを読み込む。
        ### (2)セルにデータをセットして、ダウンロード用のEXCELファイルを保存する。
        #######################################################################
        print_log('[DEBUG] P0200ExcelDownload.farmer_fisher_asset_view()関数 STEP 3/4.', 'DEBUG')
        template_file_path = 'static/template_farmer_fisher_asset.xlsx'
        download_file_path = 'static/download_farmer_fisher_asset.xlsx'
        wb = openpyxl.load_workbook(template_file_path)
        ws = wb.active
        ws.title = '農漁家資産額'
        ws.cell(row=1, column=1).value = '農漁家資産額コード'
        ws.cell(row=1, column=2).value = '農漁家資産額_償却資産額'
        ws.cell(row=1, column=3).value = '農漁家資産額_在庫資産額'
        
        if farmer_fisher_asset_list:
            for i, farmer_fisher_asset in enumerate(farmer_fisher_asset_list):
                ws.cell(row=i+2, column=1).value = farmer_fisher_asset.farmer_fisher_asset_code
                ws.cell(row=i+2, column=2).value = farmer_fisher_asset.farmer_fisher_dep_asset
                ws.cell(row=i+2, column=3).value = farmer_fisher_asset.farmer_fisher_inv_asset
        
        wb.save(download_file_path)
        
        #######################################################################
        ### レスポンスセット処理(0030)
        ### テンプレートとコンテキストを設定して、レスポンスをブラウザに戻す。
        #######################################################################
        print_log('[DEBUG] P0200ExcelDownload.farmer_fisher_asset_view()関数 STEP 4/4.', 'DEBUG')
        print_log('[INFO] P0200ExcelDownload.farmer_fisher_asset_view()関数が正常終了しました。', 'INFO')
        response = HttpResponse(content=save_virtual_workbook(wb), content_type='application/vnd.ms-excel')
        response['Content-Disposition'] = 'attachment; filename="farmer_fisher_asset.xlsx"'
        return response
        
    except:
        print_log('[ERROR] P0200ExcelDownload.farmer_fisher_asset_view()関数 {}'.format(sys.exc_info()[0]), 'ERROR')
        print_log('[ERROR] P0200ExcelDownload.farmer_fisher_asset_view()関数でエラーが発生しました。', 'ERROR')
        print_log('[ERROR] P0200ExcelDownload.farmer_fisher_asset_view()関数が異常終了しました。', 'ERROR')
        return render(request, 'error.html')

###############################################################################
### 関数名：farmer_fisher_rate_view(request, lock)
### 6010：農漁家被害率
### urlpattern：path('farmer_fisher_rate/', views.farmer_fisher_rate_view, name='farmer_fisher_rate_view')
### template：
###############################################################################
### @login_required(None, login_url='/P0100Login/')
def farmer_fisher_rate_view(request, lock):
    try:
        #######################################################################
        ### 引数チェック処理(0000)
        ### ブラウザからのリクエストと引数をチェックする。
        #######################################################################
        ### reset_log()
        print_log('[INFO] P0200ExcelDownload.farmer_fisher_rate_view()関数が開始しました。', 'INFO')
        print_log('[DEBUG] P0200ExcelDownload.farmer_fisher_rate_view()関数 request = {}'.format(request.method), 'DEBUG')
        print_log('[DEBUG] P0200ExcelDownload.farmer_fisher_rate_view()関数 lock = {}'.format(lock), 'DEBUG')
        print_log('[DEBUG] P0200ExcelDownload.farmer_fisher_rate_view()関数 STEP 1/4.', 'DEBUG')
        
        #######################################################################
        ### DBアクセス処理(0010)
        ### DBにアクセスして、農漁家被害率データを取得する。
        #######################################################################
        print_log('[DEBUG] P0200ExcelDownload.farmer_fisher_rate_view()関数 STEP 2/4.', 'DEBUG')
        farmer_fisher_rate_list = FARMER_FISHER_RATE.objects.raw("""
            SELECT 
                FF1.farmer_fisher_rate_code AS farmer_fisher_rate_code, 
                FF1.flood_sediment_code AS flood_sediment_code, 
                FS1.flood_sediment_name AS flood_sediment_name, 
                FF1.farmer_fisher_dep_rate_lv00 AS farmer_fisher_dep_lv00, 
                FF1.farmer_fisher_dep_rate_lv00_50 AS farmer_fisher_dep_lv00_50, 
                FF1.farmer_fisher_dep_rate_lv50_100 AS farmer_fisher_dep_lv50_100, 
                FF1.farmer_fisher_dep_rate_lv100_200 AS farmer_fisher_dep_lv100_200, 
                FF1.farmer_fisher_dep_rate_lv200_300 AS farmer_fisher_dep_lv200_300, 
                FF1.farmer_fisher_dep_rate_lv300 AS farmer_fisher_dep_lv300, 
                FF1.farmer_fisher_inv_rate_lv00 AS farmer_fisher_inv_lv00, 
                FF1.farmer_fisher_inv_rate_lv00_50 AS farmer_fisher_inv_lv00_50, 
                FF1.farmer_fisher_inv_rate_lv50_100 AS farmer_fisher_inv_lv50_100, 
                FF1.farmer_fisher_inv_rate_lv100_200 AS farmer_fisher_inv_lv100_200, 
                FF1.farmer_fisher_inv_rate_lv200_300 AS farmer_fisher_inv_lv200_300, 
                FF1.farmer_fisher_inv_rate_lv300 AS farmer_fisher_inv_lv300 
            FROM FARMER_FISHER_RATE FF1 
            LEFT JOIN FLOOD_SEDIMENT FS1 ON FF1.flood_sediment_code=FS1.flood_sediment_code 
            ORDER BY CAST(FF1.farmer_fisher_rate_code AS INTEGER)""", [])
    
        #######################################################################
        ### EXCEL入出力処理(0020)
        ### (1)テンプレート用のEXCELファイルを読み込む。
        ### (2)セルにデータをセットして、ダウンロード用のEXCELファイルを保存する。
        #######################################################################
        print_log('[DEBUG] P0200ExcelDownload.farmer_fisher_rate_view()関数 STEP 3/4.', 'DEBUG')
        template_file_path = 'static/template_farmer_fisher_rate.xlsx'
        download_file_path = 'static/download_farmer_fisher_rate.xlsx'
        wb = openpyxl.load_workbook(template_file_path)
        ws = wb.active
        ws.title = '農漁家被害率'
        ws.cell(row=1, column=1).value = '農漁家被害率コード'
        ws.cell(row=1, column=2).value = '浸水土砂区分コード'
        ws.cell(row=1, column=3).value = '浸水土砂区分名'
        ws.cell(row=1, column=4).value = '農漁家被害率_償却資産被害率_床下'
        ws.cell(row=1, column=5).value = '農漁家被害率_償却資産被害率_0から50cm未満'
        ws.cell(row=1, column=6).value = '農漁家被害率_償却資産被害率_50から100cm未満'
        ws.cell(row=1, column=7).value = '農漁家被害率_償却資産被害率_100から200cm未満'
        ws.cell(row=1, column=8).value = '農漁家被害率_償却資産被害率_200から300cm未満'
        ws.cell(row=1, column=9).value = '農漁家被害率_償却資産被害率_300cm以上'
        ws.cell(row=1, column=10).value = '農漁家被害率_在庫資産被害率_床下'
        ws.cell(row=1, column=11).value = '農漁家被害率_在庫資産被害率_0から50cm未満'
        ws.cell(row=1, column=12).value = '農漁家被害率_在庫資産被害率_50から100cm未満'
        ws.cell(row=1, column=13).value = '農漁家被害率_在庫資産被害率_100から200cm未満'
        ws.cell(row=1, column=14).value = '農漁家被害率_在庫資産被害率_200から300cm未満'
        ws.cell(row=1, column=15).value = '農漁家被害率_在庫資産被害率_300cm以上'
        
        if farmer_fisher_rate_list:
            for i, farmer_fisher_rate in enumerate(farmer_fisher_rate_list):
                ws.cell(row=i+2, column=1).value = farmer_fisher_rate.farmer_fisher_rate_code
                ws.cell(row=i+2, column=2).value = farmer_fisher_rate.flood_sediment_code
                ws.cell(row=i+2, column=3).value = farmer_fisher_rate.flood_sediment_name
                ws.cell(row=i+2, column=4).value = farmer_fisher_rate.farmer_fisher_dep_rate_lv00
                ws.cell(row=i+2, column=5).value = farmer_fisher_rate.farmer_fisher_dep_rate_lv00_50
                ws.cell(row=i+2, column=6).value = farmer_fisher_rate.farmer_fisher_dep_rate_lv50_100
                ws.cell(row=i+2, column=7).value = farmer_fisher_rate.farmer_fisher_dep_rate_lv100_200
                ws.cell(row=i+2, column=8).value = farmer_fisher_rate.farmer_fisher_dep_rate_lv200_300
                ws.cell(row=i+2, column=9).value = farmer_fisher_rate.farmer_fisher_dep_rate_lv300
                ws.cell(row=i+2, column=10).value = farmer_fisher_rate.farmer_fisher_inv_rate_lv00
                ws.cell(row=i+2, column=11).value = farmer_fisher_rate.farmer_fisher_inv_rate_lv00_50
                ws.cell(row=i+2, column=12).value = farmer_fisher_rate.farmer_fisher_inv_rate_lv50_100
                ws.cell(row=i+2, column=13).value = farmer_fisher_rate.farmer_fisher_inv_rate_lv100_200
                ws.cell(row=i+2, column=14).value = farmer_fisher_rate.farmer_fisher_inv_rate_lv200_300
                ws.cell(row=i+2, column=15).value = farmer_fisher_rate.farmer_fisher_inv_rate_lv300
        
        wb.save(download_file_path)
        
        #######################################################################
        ### レスポンスセット処理(0030)
        ### テンプレートとコンテキストを設定して、レスポンスをブラウザに戻す。
        #######################################################################
        print_log('[DEBUG] P0200ExcelDownload.farmer_fisher_rate_view()関数 STEP 4/4.', 'DEBUG')
        print_log('[INFO] P0200ExcelDownload.farmer_fisher_rate_view()関数が正常終了しました。', 'INFO')
        response = HttpResponse(content=save_virtual_workbook(wb), content_type='application/vnd.ms-excel')
        response['Content-Disposition'] = 'attachment; filename="farmer_fisher_rate.xlsx"'
        return response
        
    except:
        print_log('[ERROR] P0200ExcelDownload.farmer_fisher_rate_view()関数 {}'.format(sys.exc_info()[0]), 'ERROR')
        print_log('[ERROR] P0200ExcelDownload.farmer_fisher_rate_view()関数でエラーが発生しました。', 'ERROR')
        print_log('[ERROR] P0200ExcelDownload.farmer_fisher_rate_view()関数が異常終了しました。', 'ERROR')
        return render(request, 'error.html')

###############################################################################
### 関数名：area_view(request, lock)
### 7000：入力データ_水害区域
### urlpattern：path('area/', views.area_view, name='area_view')
### template：
###############################################################################
### @login_required(None, login_url='/P0100Login/')
def area_view(request, lock):
    try:
        #######################################################################
        ### 引数チェック処理(0000)
        ### ブラウザからのリクエストと引数をチェックする。
        #######################################################################
        ### reset_log()
        print_log('[INFO] P0200ExcelDownload.area_view()関数が開始しました。', 'INFO')
        print_log('[DEBUG] P0200ExcelDownload.area_view()関数 request = {}'.format(request.method), 'DEBUG')
        print_log('[DEBUG] P0200ExcelDownload.area_view()関数 lock = {}'.format(lock), 'DEBUG')
        print_log('[DEBUG] P0200ExcelDownload.area_view()関数 STEP 1/4.', 'DEBUG')
        
        #######################################################################
        ### DBアクセス処理(0010)
        ### DBにアクセスして、一般資産入力データ_水害区域データを取得する。
        #######################################################################
        print_log('[DEBUG] P0200ExcelDownload.area_view()関数 STEP 2/4.', 'DEBUG')
        area_list = AREA.objects.raw("""SELECT * FROM AREA ORDER BY CAST(AREA_ID AS INTEGER)""", [])
    
        #######################################################################
        ### EXCEL入出力処理(0020)
        ### (1)テンプレート用のEXCELファイルを読み込む。
        ### (2)セルにデータをセットして、ダウンロード用のEXCELファイルを保存する。
        #######################################################################
        print_log('[DEBUG] P0200ExcelDownload.area_view()関数 STEP 3/4.', 'DEBUG')
        template_file_path = 'static/template_area.xlsx'
        download_file_path = 'static/download_area.xlsx'
        wb = openpyxl.load_workbook(template_file_path)
        ws = wb.active
        ws.title = '一般資産入力データ_水害区域'
        ws.cell(row=1, column=1).value = '区域ID'
        ws.cell(row=1, column=2).value = '区域名'
        
        if area_list:
            for i, area in enumerate(area_list):
                ws.cell(row=i+2, column=1).value = area.area_id
                ws.cell(row=i+2, column=2).value = area.area_name
        
        wb.save(download_file_path)
        
        #######################################################################
        ### レスポンスセット処理(0030)
        ### テンプレートとコンテキストを設定して、レスポンスをブラウザに戻す。
        #######################################################################
        print_log('[DEBUG] P0200ExcelDownload.area_view()関数 STEP 4/4.', 'DEBUG')
        print_log('[INFO] P0200ExcelDownload.area_view()関数が正常終了しました。', 'INFO')
        response = HttpResponse(content=save_virtual_workbook(wb), content_type='application/vnd.ms-excel')
        response['Content-Disposition'] = 'attachment; filename="area.xlsx"'
        return response
        
    except:
        print_log('[ERROR] P0200ExcelDownload.area_view()関数 {}'.format(sys.exc_info()[0]), 'ERROR')
        print_log('[ERROR] P0200ExcelDownload.area_view()関数でエラーが発生しました。', 'ERROR')
        print_log('[ERROR] P0200ExcelDownload.area_view()関数が異常終了しました。', 'ERROR')
        return render(request, 'error.html')

###############################################################################
### 関数名：weather_view(request, lock)
### 7010：入力データ_異常気象
### urlpattern：path('weather/', views.weather_view, name='weather_view')
### template：
###############################################################################
### @login_required(None, login_url='/P0100Login/')
def weather_view(request, lock):
    try:
        #######################################################################
        ### 引数チェック処理(0000)
        ### ブラウザからのリクエストと引数をチェックする。
        #######################################################################
        ### reset_log()
        print_log('[INFO] P0200ExcelDownload.weather_view()関数が開始しました。', 'INFO')
        print_log('[DEBUG] P0200ExcelDownload.weather_view()関数 request = {}'.format(request.method), 'DEBUG')
        print_log('[DEBUG] P0200ExcelDownload.weather_view()関数 lock = {}'.format(lock), 'DEBUG')
        print_log('[DEBUG] P0200ExcelDownload.weather_view()関数 STEP 1/4.', 'DEBUG')
        
        #######################################################################
        ### DBアクセス処理(0010)
        ### DBにアクセスして、入力データ_異常気象データを取得する。
        #######################################################################
        print_log('[DEBUG] P0200ExcelDownload.weather_view()関数 STEP 2/4.', 'DEBUG')
        weather_list = WEATHER.objects.raw("""
            SELECT 
                WE1.weather_id AS weather_id, 
                WE1.weather_name AS weather_name, 
                TO_CHAR(timezone('JST', WE1.begin_date::timestamptz), 'yyyy/mm/dd') AS begin_date, 
                TO_CHAR(timezone('JST', WE1.end_date::timestamptz), 'yyyy/mm/dd') AS end_date, 
                WE1.ken_code AS ken_code, 
                KE1.ken_name AS ken_name 
            FROM WEATHER WE1 
            LEFT JOIN KEN KE1 ON WE1.ken_code=KE1.ken_code 
            ORDER BY CAST(weather_id AS INTEGER)""", [])
    
        #######################################################################
        ### EXCEL入出力処理(0020)
        ### (1)テンプレート用のEXCELファイルを読み込む。
        ### (2)セルにデータをセットして、ダウンロード用のEXCELファイルを保存する。
        #######################################################################
        print_log('[DEBUG] P0200ExcelDownload.weather_view()関数 STEP 3/4.', 'DEBUG')
        template_file_path = 'static/template_weather.xlsx'
        download_file_path = 'static/download_weather.xlsx'
        wb = openpyxl.load_workbook(template_file_path)
        ws = wb.active
        ws.title = '入力データ_異常気象'
        ws.cell(row=1, column=1).value = '異常気象ID'
        ws.cell(row=1, column=2).value = '異常気象名'
        ws.cell(row=1, column=3).value = '開始日'
        ws.cell(row=1, column=4).value = '終了日'
        ws.cell(row=1, column=5).value = '都道府県コード'
        ws.cell(row=1, column=6).value = '都道府県名'

        ws.cell(row=2, column=1).value = 'weather_id'
        ws.cell(row=2, column=2).value = 'weather_name'
        ws.cell(row=2, column=3).value = 'begin_date'
        ws.cell(row=2, column=4).value = 'end_date'
        ws.cell(row=2, column=5).value = 'ken_code'
        ws.cell(row=2, column=6).value = 'ken_name'
        
        if weather_list:
            for i, weather in enumerate(weather_list):
                ws.cell(row=i+3, column=1).value = weather.weather_id
                ws.cell(row=i+3, column=2).value = weather.weather_name
                ws.cell(row=i+3, column=3).value = weather.begin_date
                ws.cell(row=i+3, column=4).value = weather.end_date
                ws.cell(row=i+3, column=5).value = weather.ken_code
                ws.cell(row=i+3, column=6).value = weather.ken_name
        
        wb.save(download_file_path)
        
        #######################################################################
        ### レスポンスセット処理(0030)
        ### テンプレートとコンテキストを設定して、レスポンスをブラウザに戻す。
        #######################################################################
        print_log('[DEBUG] P0200ExcelDownload.weather_view()関数 STEP 4/4.', 'DEBUG')
        print_log('[INFO] P0200ExcelDownload.weather_view()関数が正常終了しました。', 'INFO')
        response = HttpResponse(content=save_virtual_workbook(wb), content_type='application/vnd.ms-excel')
        response['Content-Disposition'] = 'attachment; filename="weather.xlsx"'
        return response
        
    except:
        print_log('[ERROR] P0200ExcelDownload.weather_view()関数 {}'.format(sys.exc_info()[0]), 'ERROR')
        print_log('[ERROR] P0200ExcelDownload.weather_view()関数でエラーが発生しました。', 'ERROR')
        print_log('[ERROR] P0200ExcelDownload.weather_view()関数が異常終了しました。', 'ERROR')
        return render(request, 'error.html')

###############################################################################
### 関数名：suigai_view(request, lock)
### 7020：入力データ_ヘッダ部分
### urlpattern：path('suigai/', views.suigai_view, name='suigai_view')
### template：
###############################################################################
### @login_required(None, login_url='/P0100Login/')
def suigai_view(request, lock):
    try:
        #######################################################################
        ### 引数チェック処理(0000)
        ### ブラウザからのリクエストと引数をチェックする。
        #######################################################################
        ### reset_log()
        print_log('[INFO] P0200ExcelDownload.suigai_view()関数が開始しました。', 'INFO')
        print_log('[DEBUG] P0200ExcelDownload.suigai_view()関数 request = {}'.format(request.method), 'DEBUG')
        print_log('[DEBUG] P0200ExcelDownload.suigai_view()関数 lock = {}'.format(lock), 'DEBUG')
        print_log('[DEBUG] P0200ExcelDownload.suigai_view()関数 STEP 1/4.', 'DEBUG')
        
        #######################################################################
        ### DBアクセス処理(0010)
        ### DBにアクセスして、水害データを取得する。
        #######################################################################
        print_log('[DEBUG] P0200ExcelDownload.suigai_view()関数 STEP 2/4.', 'DEBUG')
        suigai_list = SUIGAI.objects.raw("""
            SELECT 
                SG1.suigai_id AS suigai_id, 
                SG1.suigai_name AS suigai_name, 
                SG1.ken_code AS ken_code, 
                KE1.ken_name AS ken_name, 
                SG1.city_code AS city_code, 
                CT1.city_name AS city_name, 
                TO_CHAR(timezone('JST', SG1.begin_date::timestamptz), 'yyyy/mm/dd') AS begin_date, 
                TO_CHAR(timezone('JST', SG1.end_date::timestamptz), 'yyyy/mm/dd') AS end_date, 
                SG1.cause_1_code AS cause_1_code, 
                CA1.cause_name AS cause_1_name, 
                SG1.cause_2_code AS cause_2_code, 
                CA2.cause_name AS cause_2_name, 
                SG1.cause_3_code AS cause_3_code, 
                CA3.cause_name AS cause_3_name, 
                SG1.area_id AS area_id, 
                AR1.area_name AS area_name, 
                SG1.suikei_code AS suikei_code, 
                SK1.suikei_name AS suikei_name, 
                SG1.kasen_code AS kasen_code, 
                KA1.kasen_name AS kasen_name, 
                SG1.gradient_code AS gradient_code, 
                GR1.gradient_name AS gradient_name, 
                SG1.residential_area AS residential_area, 
                SG1.agricultural_area AS agricultural_area, 
                SG1.underground_area AS underground_area, 
                SG1.kasen_kaigan_code AS kasen_kaigan_code, 
                KK1.kasen_kaigan_name AS kasen_kaigan_name, 
                SG1.crop_damage AS crop_damage, 
                SG1.weather_id AS weather_id, 
                WE1.weather_name AS weather_name, 
                TO_CHAR(timezone('JST', SG1.committed_at::timestamptz), 'yyyy/mm/dd HH24:MI') AS committed_at, 
                TO_CHAR(timezone('JST', SG1.deleted_at::timestamptz), 'yyyy/mm/dd HH24:MI') AS deleted_at, 
                SG1.upload_file_path AS upload_file_path, 
                SG1.upload_file_name AS upload_file_name, 
                SG1.summary_file_path AS summary_file_path, 
                SG1.summary_file_name AS summary_file_name, 
                SG1.action_code AS action_code, 
                AC1.action_name AS action_name, 
                AC1.action_name_en AS action_name_en, 
                SG1.status_code AS status_code, 
                ST1.status_name AS status_name 
            FROM SUIGAI SG1 
            LEFT JOIN KEN KE1 ON SG1.ken_code=KE1.ken_code 
            LEFT JOIN CITY CT1 ON SG1.city_code=CT1.city_code 
            LEFT JOIN CAUSE CA1 ON SG1.cause_1_code=CA1.cause_code 
            LEFT JOIN CAUSE CA2 ON SG1.cause_2_code=CA2.cause_code 
            LEFT JOIN CAUSE CA3 ON SG1.cause_3_code=CA3.cause_code 
            LEFT JOIN AREA AR1 ON SG1.area_id=AR1.area_id 
            LEFT JOIN SUIKEI SK1 ON SG1.suikei_code=SK1.suikei_code 
            LEFT JOIN KASEN KA1 ON SG1.kasen_code=KA1.kasen_code 
            LEFT JOIN GRADIENT GR1 ON SG1.gradient_code=GR1.gradient_code 
            LEFT JOIN KASEN_KAIGAN KK1 ON SG1.kasen_kaigan_code=KK1.kasen_kaigan_code 
            LEFT JOIN WEATHER WE1 ON SG1.weather_id=WE1.weather_id 
            LEFT JOIN ACTION AC1 ON SG1.action_code=AC1.action_code 
            LEFT JOIN STATUS ST1 ON SG1.status_code=ST1.status_code 
            ORDER BY CAST(SG1.suigai_id AS INTEGER)""", [])
    
        #######################################################################
        ### EXCEL入出力処理(0020)
        ### (1)テンプレート用のEXCELファイルを読み込む。
        ### (2)セルにデータをセットして、ダウンロード用のEXCELファイルを保存する。
        #######################################################################
        print_log('[DEBUG] P0200ExcelDownload.suigai_view()関数 STEP 3/4.', 'DEBUG')
        template_file_path = 'static/template_suigai.xlsx'
        download_file_path = 'static/download_suigai.xlsx'
        wb = openpyxl.load_workbook(template_file_path)
        ws = wb.active
        ws.title = 'シート'
        ws.cell(row=1, column=1).value = 'シートID'
        ws.cell(row=1, column=2).value = 'シート名'
        ws.cell(row=1, column=3).value = '都道府県コード'
        ws.cell(row=1, column=4).value = '都道府県名'
        ws.cell(row=1, column=5).value = '市区町村コード'
        ws.cell(row=1, column=6).value = '市区町村名'
        ws.cell(row=1, column=7).value = '開始日'
        ws.cell(row=1, column=8).value = '終了日'
        ws.cell(row=1, column=9).value = '水害原因_1_コード'
        ws.cell(row=1, column=10).value = '水害原因_1_名'
        ws.cell(row=1, column=11).value = '水害原因_2_コード'
        ws.cell(row=1, column=12).value = '水害原因_2_名'
        ws.cell(row=1, column=13).value = '水害原因_3_コード'
        ws.cell(row=1, column=14).value = '水害原因_3_名'
        ws.cell(row=1, column=15).value = '水害区域ID'
        ws.cell(row=1, column=16).value = '水害区域名'
        ws.cell(row=1, column=17).value = '水系コード'
        ws.cell(row=1, column=18).value = '水系名'
        ws.cell(row=1, column=19).value = '河川コード'
        ws.cell(row=1, column=20).value = '河川名'
        ws.cell(row=1, column=21).value = '地盤勾配区分コード'
        ws.cell(row=1, column=22).value = '地盤勾配区分名'
        ws.cell(row=1, column=23).value = '宅地面積（単位m2）'
        ws.cell(row=1, column=24).value = '農地面積（単位m2）'
        ws.cell(row=1, column=25).value = '地下面積（単位m2）'
        ws.cell(row=1, column=26).value = '河川海岸（工種）コード'
        ws.cell(row=1, column=27).value = '河川海岸（工種）名'
        ws.cell(row=1, column=28).value = '農作物被害額（単位千円）'
        ws.cell(row=1, column=29).value = '異常気象ID'
        ws.cell(row=1, column=30).value = '異常気象名'
        ws.cell(row=1, column=31).value = 'コミット日時'
        ws.cell(row=1, column=32).value = '削除日時'
        ws.cell(row=1, column=33).value = 'アップロードファイルパス'
        ws.cell(row=1, column=34).value = 'アップロードファイル名'
        ws.cell(row=1, column=35).value = '集計結果ファイルパス'
        ws.cell(row=1, column=36).value = '集計結果ファイル名'
        ws.cell(row=1, column=37).value = 'アクションコード'
        ws.cell(row=1, column=38).value = 'アクション名'
        ws.cell(row=1, column=39).value = '状態コード'
        ws.cell(row=1, column=40).value = '状態名'

        ws.cell(row=2, column=1).value = 'suigai_id'
        ws.cell(row=2, column=2).value = 'suigai_name'
        ws.cell(row=2, column=3).value = 'ken_code'
        ws.cell(row=2, column=4).value = 'ken_name'
        ws.cell(row=2, column=5).value = 'city_code'
        ws.cell(row=2, column=6).value = 'city_name'
        ws.cell(row=2, column=7).value = 'begin_date'
        ws.cell(row=2, column=8).value = 'end_date'
        ws.cell(row=2, column=9).value = 'cause_1_code'
        ws.cell(row=2, column=10).value = 'cause_1_name'
        ws.cell(row=2, column=11).value = 'cause_2_code'
        ws.cell(row=2, column=12).value = 'cause_2_name'
        ws.cell(row=2, column=13).value = 'cause_3_code'
        ws.cell(row=2, column=14).value = 'cause_3_name'
        ws.cell(row=2, column=15).value = 'area_id'
        ws.cell(row=2, column=16).value = 'area_name'
        ws.cell(row=2, column=17).value = 'suikei_code'
        ws.cell(row=2, column=18).value = 'suikei_name'
        ws.cell(row=2, column=19).value = 'kasen_code'
        ws.cell(row=2, column=20).value = 'kasen_name'
        ws.cell(row=2, column=21).value = 'gradient_code'
        ws.cell(row=2, column=22).value = 'gradient_name'
        ws.cell(row=2, column=23).value = 'residential_area'
        ws.cell(row=2, column=24).value = 'agricultural_area'
        ws.cell(row=2, column=25).value = 'underground_area'
        ws.cell(row=2, column=26).value = 'kasen_kaigan_code'
        ws.cell(row=2, column=27).value = 'kasen_kaigan_name'
        ws.cell(row=2, column=28).value = 'crop_damage'
        ws.cell(row=2, column=29).value = 'weather_id'
        ws.cell(row=2, column=30).value = 'weather_name'
        ws.cell(row=2, column=31).value = 'committed_at'
        ws.cell(row=2, column=32).value = 'deleted_at'
        ws.cell(row=2, column=33).value = 'upload_file_path'
        ws.cell(row=2, column=34).value = 'upload_file_name'
        ws.cell(row=2, column=35).value = 'summary_file_path'
        ws.cell(row=2, column=36).value = 'summary_file_name'
        ws.cell(row=2, column=37).value = 'action_code'
        ws.cell(row=2, column=38).value = 'action_name'
        ws.cell(row=2, column=39).value = 'status_code'
        ws.cell(row=2, column=40).value = 'status_name'
        
        if suigai_list:
            for i, suigai in enumerate(suigai_list):
                ws.cell(row=i+3, column=1).value = suigai.suigai_id
                ws.cell(row=i+3, column=2).value = suigai.suigai_name
                ws.cell(row=i+3, column=3).value = suigai.ken_code
                ws.cell(row=i+3, column=4).value = suigai.ken_name
                ws.cell(row=i+3, column=5).value = suigai.city_code
                ws.cell(row=i+3, column=6).value = suigai.city_name
                ws.cell(row=i+3, column=7).value = suigai.begin_date
                ws.cell(row=i+3, column=8).value = suigai.end_date
                ws.cell(row=i+3, column=9).value = suigai.cause_1_code
                ws.cell(row=i+3, column=10).value = suigai.cause_1_name
                ws.cell(row=i+3, column=11).value = suigai.cause_2_code
                ws.cell(row=i+3, column=12).value = suigai.cause_2_name
                ws.cell(row=i+3, column=13).value = suigai.cause_3_code
                ws.cell(row=i+3, column=14).value = suigai.cause_3_name
                ws.cell(row=i+3, column=15).value = suigai.area_id
                ws.cell(row=i+3, column=16).value = suigai.area_name
                ws.cell(row=i+3, column=17).value = suigai.suikei_code
                ws.cell(row=i+3, column=18).value = suigai.suikei_name
                ws.cell(row=i+3, column=19).value = suigai.kasen_code
                ws.cell(row=i+3, column=20).value = suigai.kasen_name
                ws.cell(row=i+3, column=21).value = suigai.gradient_code
                ws.cell(row=i+3, column=22).value = suigai.gradient_name
                ws.cell(row=i+3, column=23).value = suigai.residential_area
                ws.cell(row=i+3, column=24).value = suigai.agricultural_area
                ws.cell(row=i+3, column=25).value = suigai.underground_area
                ws.cell(row=i+3, column=26).value = suigai.kasen_kaigan_code
                ws.cell(row=i+3, column=27).value = suigai.kasen_kaigan_name
                ws.cell(row=i+3, column=28).value = suigai.crop_damage
                ws.cell(row=i+3, column=29).value = suigai.weather_id
                ws.cell(row=i+3, column=30).value = suigai.weather_name
                ws.cell(row=i+3, column=31).value = suigai.committed_at
                ws.cell(row=i+3, column=32).value = suigai.deleted_at
                ws.cell(row=i+3, column=33).value = suigai.upload_file_path
                ws.cell(row=i+3, column=34).value = suigai.upload_file_name
                ws.cell(row=i+3, column=35).value = suigai.summary_file_path
                ws.cell(row=i+3, column=36).value = suigai.summary_file_name
                ws.cell(row=i+3, column=37).value = suigai.action_code
                ws.cell(row=i+3, column=38).value = suigai.action_name
                ws.cell(row=i+3, column=39).value = suigai.status_code
                ws.cell(row=i+3, column=40).value = suigai.status_name
        
        wb.save(download_file_path)
        
        #######################################################################
        ### レスポンスセット処理(0030)
        ### テンプレートとコンテキストを設定して、レスポンスをブラウザに戻す。
        #######################################################################
        print_log('[DEBUG] P0200ExcelDownload.suigai_view()関数 STEP 4/4.', 'DEBUG')
        print_log('[INFO] P0200ExcelDownload.suigai_view()関数が正常終了しました。', 'INFO')
        response = HttpResponse(content=save_virtual_workbook(wb), content_type='application/vnd.ms-excel')
        response['Content-Disposition'] = 'attachment; filename="suigai.xlsx"'
        return response
        
    except:
        print_log('[ERROR] P0200ExcelDownload.suigai_view()関数 {}'.format(sys.exc_info()[0]), 'ERROR')
        print_log('[ERROR] P0200ExcelDownload.suigai_view()関数でエラーが発生しました。', 'ERROR')
        print_log('[ERROR] P0200ExcelDownload.suigai_view()関数が異常終了しました。', 'ERROR')
        return render(request, 'error.html')

###############################################################################
### 関数名：ippan_view(request, lock)
### 7030：入力データ_一覧表部分
### urlpattern：path('ippan/', views.ippan_view, name='ippan_view')
### template：
###############################################################################
### @login_required(None, login_url='/P0100Login/')
def ippan_view(request, lock):
    try:
        #######################################################################
        ### 引数チェック処理(0000)
        ### ブラウザからのリクエストと引数をチェックする。
        #######################################################################
        reset_log()
        print_log('[INFO] P0200ExcelDownload.ippan_view()関数が開始しました。', 'INFO')
        print_log('[DEBUG] P0200ExcelDownload.ippan_view()関数 request = {}'.format(request.method), 'DEBUG')
        print_log('[DEBUG] P0200ExcelDownload.ippan_view()関数 lock = {}'.format(lock), 'DEBUG')
        print_log('[DEBUG] P0200ExcelDownload.ippan_view()関数 STEP 1/4.', 'DEBUG')
        
        #######################################################################
        ### DBアクセス処理(0010)
        ### DBにアクセスして、入力データ_一覧表部分データを取得する。
        #######################################################################
        print_log('[DEBUG] P0200ExcelDownload.ippan_view()関数 STEP 2/4.', 'DEBUG')
        ippan_list = IPPAN.objects.raw("""SELECT * FROM IPPAN ORDER BY CAST(IPPAN_ID AS INTEGER)""", [])
    
        #######################################################################
        ### EXCEL入出力処理(0020)
        ### (1)テンプレート用のEXCELファイルを読み込む。
        ### (2)セルにデータをセットして、ダウンロード用のEXCELファイルを保存する。
        #######################################################################
        print_log('[DEBUG] P0200ExcelDownload.ippan_view()関数 STEP 3/4.', 'DEBUG')
        template_file_path = 'static/template_ippan.xlsx'
        download_file_path = 'static/download_ippan.xlsx'
        wb = openpyxl.load_workbook(template_file_path)
        ws = wb.active
        ws.title = '入力データ_一覧表部分'
        ws.cell(row=1, column=1).value = '行ID'
        ws.cell(row=1, column=2).value = '町丁目、大字名'
        ws.cell(row=1, column=3).value = '水害ID'
        ws.cell(row=1, column=4).value = '建物区分コード'
        ws.cell(row=1, column=5).value = '地上地下区分コード'
        ws.cell(row=1, column=6).value = '浸水土砂区分コード'
        ws.cell(row=1, column=7).value = '被害建物棟数_床下'
        ws.cell(row=1, column=8).value = '被害建物棟数_01から49cm'
        ws.cell(row=1, column=9).value = '被害建物棟数_50から99cm'
        ws.cell(row=1, column=10).value = '被害建物棟数_100cm以上'
        ws.cell(row=1, column=11).value = '被害建物棟数_半壊'
        ws.cell(row=1, column=12).value = '被害建物棟数_全壊'
        ws.cell(row=1, column=13).value = '延床面積'
        ws.cell(row=1, column=14).value = '被災世帯数'
        ws.cell(row=1, column=15).value = '被災事業所数'
        ws.cell(row=1, column=16).value = '農漁家戸数_床下'
        ws.cell(row=1, column=17).value = '農漁家戸数_01から49cm'
        ws.cell(row=1, column=18).value = '農漁家戸数_50から99cm'
        ws.cell(row=1, column=19).value = '農漁家戸数_100cm以上'
        ws.cell(row=1, column=20).value = '農漁家戸数_全壊'
        ws.cell(row=1, column=21).value = '被災従業者数_床下'
        ws.cell(row=1, column=22).value = '被災従業者数_01から49cm'
        ws.cell(row=1, column=23).value = '被災従業者数_50から99cm'
        ws.cell(row=1, column=24).value = '被災従業者数_100cm以上'
        ws.cell(row=1, column=25).value = '被災従業者数_全壊'
        ws.cell(row=1, column=26).value = '産業分類コード'
        ws.cell(row=1, column=27).value = '地下空間の利用形態コード'
        ws.cell(row=1, column=28).value = '備考'
        ws.cell(row=1, column=29).value = '削除日時'

        ws.cell(row=2, column=1).value = 'ippan_id'
        ws.cell(row=2, column=2).value = 'ippan_name'
        ws.cell(row=2, column=3).value = 'suigai_id'
        ws.cell(row=2, column=4).value = 'building_code'
        ws.cell(row=2, column=5).value = 'underground_code'
        ws.cell(row=2, column=6).value = 'flood_sediment_code'
        ws.cell(row=2, column=7).value = 'building_lv00'
        ws.cell(row=2, column=8).value = 'building_lv01_49'
        ws.cell(row=2, column=9).value = 'building_lv50_99'
        ws.cell(row=2, column=10).value = 'building_lv100'
        ws.cell(row=2, column=11).value = 'building_half'
        ws.cell(row=2, column=12).value = 'building_full'
        ws.cell(row=2, column=13).value = 'floor_area'
        ws.cell(row=2, column=14).value = 'family_area'
        ws.cell(row=2, column=15).value = 'office_area'
        ws.cell(row=2, column=16).value = 'farmer_fisher_lv00'
        ws.cell(row=2, column=17).value = 'farmer_fisher_lv01_49'
        ws.cell(row=2, column=18).value = 'farmer_fisher_lv50_99'
        ws.cell(row=2, column=19).value = 'farmer_fisher_lv100'
        ws.cell(row=2, column=20).value = 'farmer_fisher_full'
        ws.cell(row=2, column=21).value = 'employee_lv00'
        ws.cell(row=2, column=22).value = 'employee_lv01_49'
        ws.cell(row=2, column=23).value = 'employee_lv50_99'
        ws.cell(row=2, column=24).value = 'employee_lv100'
        ws.cell(row=2, column=25).value = 'employee_full'
        ws.cell(row=2, column=26).value = 'industry_code'
        ws.cell(row=2, column=27).value = 'usage_code'
        ws.cell(row=2, column=28).value = 'comment'
        ws.cell(row=2, column=29).value = 'deleted_at'
        
        if ippan_list:
            for i, ippan in enumerate(ippan_list):
                ws.cell(row=i+3, column=1).value = ippan.ippan_id
                ws.cell(row=i+3, column=2).value = ippan.ippan_name
                ws.cell(row=i+3, column=3).value = ippan.suigai_id
                ws.cell(row=i+3, column=4).value = ippan.building_code
                ws.cell(row=i+3, column=5).value = ippan.underground_code
                ws.cell(row=i+3, column=6).value = ippan.flood_sediment_code
                ws.cell(row=i+3, column=7).value = ippan.building_lv00
                ws.cell(row=i+3, column=8).value = ippan.building_lv01_49
                ws.cell(row=i+3, column=9).value = ippan.building_lv50_99
                ws.cell(row=i+3, column=10).value = ippan.building_lv100
                ws.cell(row=i+3, column=11).value = ippan.building_half
                ws.cell(row=i+3, column=12).value = ippan.building_full
                ws.cell(row=i+3, column=13).value = ippan.floor_area
                ws.cell(row=i+3, column=14).value = ippan.family
                ws.cell(row=i+3, column=15).value = ippan.office
                ws.cell(row=i+3, column=16).value = ippan.farmer_fisher_lv00
                ws.cell(row=i+3, column=17).value = ippan.farmer_fisher_lv01_49
                ws.cell(row=i+3, column=18).value = ippan.farmer_fisher_lv50_99
                ws.cell(row=i+3, column=19).value = ippan.farmer_fisher_lv100
                ws.cell(row=i+3, column=20).value = ippan.farmer_fisher_full
                ws.cell(row=i+3, column=21).value = ippan.employee_lv00
                ws.cell(row=i+3, column=22).value = ippan.employee_lv01_49
                ws.cell(row=i+3, column=23).value = ippan.employee_lv50_99
                ws.cell(row=i+3, column=24).value = ippan.employee_lv100
                ws.cell(row=i+3, column=25).value = ippan.employee_full
                ws.cell(row=i+3, column=26).value = ippan.industry_code
                ws.cell(row=i+3, column=27).value = ippan.usage_code
                ws.cell(row=i+3, column=28).value = ippan.comment
                ws.cell(row=i+3, column=29).value = str(ippan.deleted_at)

        wb.save(download_file_path)
        
        #######################################################################
        ### レスポンスセット処理(0030)
        ### テンプレートとコンテキストを設定して、レスポンスをブラウザに戻す。
        #######################################################################
        print_log('[DEBUG] P0200ExcelDownload.ippan_view()関数 STEP 4/4.', 'DEBUG')
        print_log('[INFO] P0200ExcelDownload.ippan_view()関数が正常終了しました。', 'INFO')
        response = HttpResponse(content=save_virtual_workbook(wb), content_type='application/vnd.ms-excel')
        response['Content-Disposition'] = 'attachment; filename="ippan.xlsx"'
        return response
        
    except:
        print_log('[ERROR] P0200ExcelDownload.ippan_view()関数 {}'.format(sys.exc_info()[0]), 'ERROR')
        print_log('[ERROR] P0200ExcelDownload.ippan_view()関数でエラーが発生しました。', 'ERROR')
        print_log('[ERROR] P0200ExcelDownload.ippan_view()関数が異常終了しました。', 'ERROR')
        return render(request, 'error.html')

###############################################################################
### 関数名：ippan_view_view(request, lock)
### 7040：ビューデータ_一覧表部分
### urlpattern：path('ippan_view/', views.ippan_view_view, name='ippan_view_view')
### template：
###############################################################################
### @login_required(None, login_url='/P0100Login/')
def ippan_view_view(request, lock):
    try:
        #######################################################################
        ### 引数チェック処理(0000)
        ### ブラウザからのリクエストと引数をチェックする。
        #######################################################################
        ### reset_log()
        print_log('[INFO] P0200ExcelDownload.ippan_view_view()関数が開始しました。', 'INFO')
        print_log('[DEBUG] P0200ExcelDownload.ippan_view_view()関数 request = {}'.format(request.method), 'DEBUG')
        print_log('[DEBUG] P0200ExcelDownload.ippan_view_view()関数 lock = {}'.format(lock), 'DEBUG')
        print_log('[DEBUG] P0200ExcelDownload.ippan_view_view()関数 STEP 1/4.', 'DEBUG')
        
        #######################################################################
        ### DBアクセス処理(0010)
        ### DBにアクセスして、一般資産ビューデータ一覧表部分データを取得する。
        #######################################################################
        print_log('[DEBUG] P0200ExcelDownload.ippan_view_view()関数 STEP 2/4.', 'DEBUG')
        ippan_view_list = IPPAN_VIEW.objects.raw("""SELECT * FROM IPPAN_VIEW ORDER BY CAST(IPPAN_ID AS INTEGER)""", [])
    
        #######################################################################
        ### EXCEL入出力処理(0020)
        ### (1)テンプレート用のEXCELファイルを読み込む。
        ### (2)セルにデータをセットして、ダウンロード用のEXCELファイルを保存する。
        #######################################################################
        print_log('[DEBUG] P0200ExcelDownload.ippan_view_view()関数 STEP 3/4.', 'DEBUG')
        template_file_path = 'static/template_ippan_view.xlsx'
        download_file_path = 'static/download_ippan_view.xlsx'
        wb = openpyxl.load_workbook(template_file_path)
        ws = wb.active
        ws.title = 'ビューデータ_一覧表部分'
        ws.cell(row=1, column=1).value = '行ID'
        ws.cell(row=1, column=2).value = '町丁目、大字名'
        ws.cell(row=1, column=3).value = '水害ID'
        ws.cell(row=1, column=4).value = '水害名'
        ws.cell(row=1, column=5).value = '都道府県コード'
        ws.cell(row=1, column=6).value = '都道府県名'
        ws.cell(row=1, column=7).value = '市区町村コード'
        ws.cell(row=1, column=8).value = '市区町村名'
        ws.cell(row=1, column=9).value = '水害原因コード1'
        ws.cell(row=1, column=10).value = '水害原因名1'
        ws.cell(row=1, column=11).value = '水害原因コード2'
        ws.cell(row=1, column=12).value = '水害原因名2'
        ws.cell(row=1, column=13).value = '水害原因コード3'
        ws.cell(row=1, column=14).value = '水害原因名3'
        ws.cell(row=1, column=15).value = '水害区域ID'
        ws.cell(row=1, column=16).value = '水害区域名'
        ws.cell(row=1, column=17).value = '水系コード'
        ws.cell(row=1, column=18).value = '水系名'
        ws.cell(row=1, column=19).value = '河川コード'
        ws.cell(row=1, column=20).value = '河川名'
        ws.cell(row=1, column=21).value = '地盤勾配区分コード'
        ws.cell(row=1, column=22).value = '地盤勾配区分名'
        ws.cell(row=1, column=23).value = '宅地面積'
        ws.cell(row=1, column=24).value = '農地面積'
        ws.cell(row=1, column=25).value = '地下面積'
        ws.cell(row=1, column=26).value = '河川海岸（工種）コード'
        ws.cell(row=1, column=27).value = '河川海岸（工種）名'
        ws.cell(row=1, column=28).value = '農作物被害額'
        ws.cell(row=1, column=29).value = '異常気象ID'
        ws.cell(row=1, column=30).value = '異常気象名'
        ws.cell(row=1, column=31).value = '建物区分コード'
        ws.cell(row=1, column=32).value = '建物区分名'
        ws.cell(row=1, column=33).value = '地上地下区分コード'
        ws.cell(row=1, column=34).value = '地上地下区分名'
        ws.cell(row=1, column=35).value = '浸水土砂区分コード'
        ws.cell(row=1, column=36).value = '浸水土砂区分名'
        ws.cell(row=1, column=37).value = '被害建物棟数_床下'
        ws.cell(row=1, column=38).value = '被害建物棟数_01から49cm'
        ws.cell(row=1, column=39).value = '被害建物棟数_50から99cm'
        ws.cell(row=1, column=40).value = '被害建物棟数_100cm以上'
        ws.cell(row=1, column=41).value = '被害建物棟数_半壊'
        ws.cell(row=1, column=42).value = '被害建物棟数_全壊'
        ws.cell(row=1, column=43).value = '被害建物棟数_合計'
        ws.cell(row=1, column=44).value = '延床面積'
        ws.cell(row=1, column=45).value = '被災世帯数'
        ws.cell(row=1, column=46).value = '被災事業所数'
        ws.cell(row=1, column=47).value = '延床面積_床下'
        ws.cell(row=1, column=48).value = '延床面積_01から49cm'
        ws.cell(row=1, column=49).value = '延床面積_50から99cm'
        ws.cell(row=1, column=50).value = '延床面積_100cm以上'
        ws.cell(row=1, column=51).value = '延床面積_半壊'
        ws.cell(row=1, column=52).value = '延床面積_全壊'
        ws.cell(row=1, column=53).value = '延床面積_合計'
        ws.cell(row=1, column=54).value = '被災世帯数_床下'
        ws.cell(row=1, column=55).value = '被災世帯数_01から49cm'
        ws.cell(row=1, column=56).value = '被災世帯数_50から99cm'
        ws.cell(row=1, column=57).value = '被災世帯数_100cm以上'
        ws.cell(row=1, column=58).value = '被災世帯数_半壊'
        ws.cell(row=1, column=59).value = '被災世帯数_全壊'
        ws.cell(row=1, column=60).value = '被災世帯数_合計'
        ws.cell(row=1, column=61).value = '被災事業所数_床下'
        ws.cell(row=1, column=62).value = '被災事業所数_01から49cm'
        ws.cell(row=1, column=63).value = '被災事業所数_50から99cm'
        ws.cell(row=1, column=64).value = '被災事業所数_100cm以上'
        ws.cell(row=1, column=65).value = '被災事業所数_半壊'
        ws.cell(row=1, column=66).value = '被災事業所数_全壊'
        ws.cell(row=1, column=67).value = '被災事業所数_合計'
        ws.cell(row=1, column=68).value = '農漁家戸数_床下'
        ws.cell(row=1, column=69).value = '農漁家戸数_01から49cm'
        ws.cell(row=1, column=70).value = '農漁家戸数_50から99cm'
        ws.cell(row=1, column=71).value = '農漁家戸数_100cm以上'
        ws.cell(row=1, column=72).value = '農漁家戸数_全壊'
        ws.cell(row=1, column=73).value = '農漁家戸数_合計'
        ws.cell(row=1, column=74).value = '被災従業者数_床下'
        ws.cell(row=1, column=75).value = '被災従業者数_01から49cm'
        ws.cell(row=1, column=76).value = '被災従業者数_50から99cm'
        ws.cell(row=1, column=77).value = '被災従業者数_100cm以上'
        ws.cell(row=1, column=78).value = '被災従業者数_全壊'
        ws.cell(row=1, column=79).value = '被災従業者数_合計'
        ws.cell(row=1, column=80).value = '産業分類コード'
        ws.cell(row=1, column=81).value = '産業分類名'
        ws.cell(row=1, column=82).value = '地下空間の利用形態コード'
        ws.cell(row=1, column=83).value = '地下空間の利用形態名'
        ws.cell(row=1, column=84).value = '備考'
        ws.cell(row=1, column=85).value = '削除日時'

        ws.cell(row=2, column=1).value = 'ippan_id'
        ws.cell(row=2, column=2).value = 'ippan_name'
        ws.cell(row=2, column=3).value = 'suigai_id'
        ws.cell(row=2, column=4).value = 'suigai_name'
        ws.cell(row=2, column=5).value = 'ken_code'
        ws.cell(row=2, column=6).value = 'ken_name'
        ws.cell(row=2, column=7).value = 'city_code'
        ws.cell(row=2, column=8).value = 'city_name'
        ws.cell(row=2, column=9).value = 'cause_1_code'
        ws.cell(row=2, column=10).value = 'cause_1_name'
        ws.cell(row=2, column=11).value = 'cause_2_code'
        ws.cell(row=2, column=12).value = 'cause_2_name'
        ws.cell(row=2, column=13).value = 'cause_3_code'
        ws.cell(row=2, column=14).value = 'cause_3_name'
        ws.cell(row=2, column=15).value = 'area_id'
        ws.cell(row=2, column=16).value = 'area_name'
        ws.cell(row=2, column=17).value = 'suikei_code'
        ws.cell(row=2, column=18).value = 'suikei_name'
        ws.cell(row=2, column=19).value = 'kasen_code'
        ws.cell(row=2, column=20).value = 'kasen_name'
        ws.cell(row=2, column=21).value = 'gradient_code'
        ws.cell(row=2, column=22).value = 'gradient_name'
        ws.cell(row=2, column=23).value = 'residential_area'
        ws.cell(row=2, column=24).value = 'agricultural_area'
        ws.cell(row=2, column=25).value = 'underground_area'
        ws.cell(row=2, column=26).value = 'kasen_kaigan_code'
        ws.cell(row=2, column=27).value = 'kasen_kaigan_name'
        ws.cell(row=2, column=28).value = 'crop_damage'
        ws.cell(row=2, column=29).value = 'weather_id'
        ws.cell(row=2, column=30).value = 'weather_name'
        ws.cell(row=2, column=31).value = 'building_code'
        ws.cell(row=2, column=32).value = 'building_name'
        ws.cell(row=2, column=33).value = 'underground_code'
        ws.cell(row=2, column=34).value = 'underground_name'
        ws.cell(row=2, column=35).value = 'flood_sediment_code'
        ws.cell(row=2, column=36).value = 'flood_sediment_name'
        ws.cell(row=2, column=37).value = 'building_lv00'
        ws.cell(row=2, column=38).value = 'building_lv01_49'
        ws.cell(row=2, column=39).value = 'building_lv50_99'
        ws.cell(row=2, column=40).value = 'building_lv100'
        ws.cell(row=2, column=41).value = 'building_half'
        ws.cell(row=2, column=42).value = 'building_full'
        ws.cell(row=2, column=43).value = 'building_total'
        ws.cell(row=2, column=44).value = 'floor_area'
        ws.cell(row=2, column=45).value = 'family'
        ws.cell(row=2, column=46).value = 'office'
        ws.cell(row=2, column=47).value = 'floor_area_lv00'
        ws.cell(row=2, column=48).value = 'floor_area_lv01_49'
        ws.cell(row=2, column=49).value = 'floor_area_lv50_99'
        ws.cell(row=2, column=50).value = 'floor_area_lv100'
        ws.cell(row=2, column=51).value = 'floor_area_half'
        ws.cell(row=2, column=52).value = 'floor_area_full'
        ws.cell(row=2, column=53).value = 'floor_area_total'
        ws.cell(row=2, column=54).value = 'family_lv00'
        ws.cell(row=2, column=55).value = 'family_lv01_49'
        ws.cell(row=2, column=56).value = 'family_lv50_99'
        ws.cell(row=2, column=57).value = 'family_lv100'
        ws.cell(row=2, column=58).value = 'family_half'
        ws.cell(row=2, column=59).value = 'family_full'
        ws.cell(row=2, column=60).value = 'family_total'
        ws.cell(row=2, column=61).value = 'office_lv00'
        ws.cell(row=2, column=62).value = 'office_lv01_49'
        ws.cell(row=2, column=63).value = 'office_lv50_99'
        ws.cell(row=2, column=64).value = 'office_lv100'
        ws.cell(row=2, column=65).value = 'office_half'
        ws.cell(row=2, column=66).value = 'office_full'
        ws.cell(row=2, column=67).value = 'office_total'
        ws.cell(row=2, column=68).value = 'farmer_fisher_lv00'
        ws.cell(row=2, column=69).value = 'farmer_fisher_lv01_49'
        ws.cell(row=2, column=70).value = 'farmer_fisher_lv50_99'
        ws.cell(row=2, column=71).value = 'farmer_fisher_lv100'
        ws.cell(row=2, column=72).value = 'farmer_fisher_full'
        ws.cell(row=2, column=73).value = 'farmer_fisher_total'
        ws.cell(row=2, column=74).value = 'employee_lv00'
        ws.cell(row=2, column=75).value = 'employee_lv01_49'
        ws.cell(row=2, column=76).value = 'employee_lv50_99'
        ws.cell(row=2, column=77).value = 'employee_lv100'
        ws.cell(row=2, column=78).value = 'employee_full'
        ws.cell(row=2, column=79).value = 'employee_total'
        ws.cell(row=2, column=80).value = 'industry_code'
        ws.cell(row=2, column=81).value = 'industry_name'
        ws.cell(row=2, column=82).value = 'usage_code'
        ws.cell(row=2, column=83).value = 'usage_name'
        ws.cell(row=2, column=84).value = 'comment'
        ws.cell(row=2, column=85).value = 'deleted_at'
        
        if ippan_view_list:
            for i, ippan_view in enumerate(ippan_view_list):
                ws.cell(row=i+3, column=1).value = ippan_view.ippan_id
                ws.cell(row=i+3, column=2).value = ippan_view.ippan_name
                ws.cell(row=i+3, column=3).value = ippan_view.suigai_id
                ws.cell(row=i+3, column=4).value = ippan_view.suigai_name
                ws.cell(row=i+3, column=5).value = ippan_view.ken_code
                ws.cell(row=i+3, column=6).value = ippan_view.ken_name
                ws.cell(row=i+3, column=7).value = ippan_view.city_code
                ws.cell(row=i+3, column=8).value = ippan_view.city_name
                ws.cell(row=i+3, column=9).value = ippan_view.cause_1_code
                ws.cell(row=i+3, column=10).value = ippan_view.cause_1_name
                ws.cell(row=i+3, column=11).value = ippan_view.cause_2_code
                ws.cell(row=i+3, column=12).value = ippan_view.cause_2_name
                ws.cell(row=i+3, column=13).value = ippan_view.cause_3_code
                ws.cell(row=i+3, column=14).value = ippan_view.cause_3_name
                ws.cell(row=i+3, column=15).value = ippan_view.area_id
                ws.cell(row=i+3, column=16).value = ippan_view.area_name
                ws.cell(row=i+3, column=17).value = ippan_view.suikei_code
                ws.cell(row=i+3, column=18).value = ippan_view.suikei_name
                ws.cell(row=i+3, column=19).value = ippan_view.kasen_code
                ws.cell(row=i+3, column=20).value = ippan_view.kasen_name
                ws.cell(row=i+3, column=21).value = ippan_view.gradient_code
                ws.cell(row=i+3, column=22).value = ippan_view.gradient_name
                ws.cell(row=i+3, column=23).value = ippan_view.residential_area
                ws.cell(row=i+3, column=24).value = ippan_view.agricultural_area
                ws.cell(row=i+3, column=25).value = ippan_view.underground_area
                ws.cell(row=i+3, column=26).value = ippan_view.kasen_kaigan_code
                ws.cell(row=i+3, column=27).value = ippan_view.kasen_kaigan_name
                ws.cell(row=i+3, column=28).value = ippan_view.crop_damage
                ws.cell(row=i+3, column=29).value = ippan_view.weather_id
                ws.cell(row=i+3, column=30).value = ippan_view.weather_name
                ws.cell(row=i+3, column=31).value = ippan_view.building_code
                ws.cell(row=i+3, column=32).value = ippan_view.building_name
                ws.cell(row=i+3, column=33).value = ippan_view.underground_code
                ws.cell(row=i+3, column=34).value = ippan_view.underground_name
                ws.cell(row=i+3, column=35).value = ippan_view.flood_sediment_code
                ws.cell(row=i+3, column=36).value = ippan_view.flood_sediment_name
                ws.cell(row=i+3, column=37).value = ippan_view.building_lv00
                ws.cell(row=i+3, column=38).value = ippan_view.building_lv01_49
                ws.cell(row=i+3, column=39).value = ippan_view.building_lv50_99
                ws.cell(row=i+3, column=40).value = ippan_view.building_lv100
                ws.cell(row=i+3, column=41).value = ippan_view.building_half
                ws.cell(row=i+3, column=42).value = ippan_view.building_full
                ws.cell(row=i+3, column=43).value = ippan_view.building_total
                ws.cell(row=i+3, column=44).value = ippan_view.floor_area
                ws.cell(row=i+3, column=45).value = ippan_view.family
                ws.cell(row=i+3, column=46).value = ippan_view.office
                ws.cell(row=i+3, column=47).value = ippan_view.floor_area_lv00
                ws.cell(row=i+3, column=48).value = ippan_view.floor_area_lv01_49
                ws.cell(row=i+3, column=49).value = ippan_view.floor_area_lv50_99

                ws.cell(row=i+3, column=50).value = ippan_view.floor_area_lv100
                ws.cell(row=i+3, column=51).value = ippan_view.floor_area_half
                ws.cell(row=i+3, column=52).value = ippan_view.floor_area_full
                ws.cell(row=i+3, column=53).value = ippan_view.floor_area_total
                ws.cell(row=i+3, column=54).value = ippan_view.family_lv00
                ws.cell(row=i+3, column=55).value = ippan_view.family_lv01_49
                ws.cell(row=i+3, column=56).value = ippan_view.family_lv50_99
                ws.cell(row=i+3, column=57).value = ippan_view.family_lv100
                ws.cell(row=i+3, column=58).value = ippan_view.family_half
                ws.cell(row=i+3, column=59).value = ippan_view.family_full

                ws.cell(row=i+3, column=60).value = ippan_view.family_total
                ws.cell(row=i+3, column=61).value = ippan_view.office_lv00
                ws.cell(row=i+3, column=62).value = ippan_view.office_lv01_49
                ws.cell(row=i+3, column=63).value = ippan_view.office_lv50_99
                ws.cell(row=i+3, column=64).value = ippan_view.office_lv100
                ws.cell(row=i+3, column=65).value = ippan_view.office_half
                ws.cell(row=i+3, column=66).value = ippan_view.office_full
                ws.cell(row=i+3, column=67).value = ippan_view.office_total
                ws.cell(row=i+3, column=68).value = ippan_view.farmer_fisher_lv00
                ws.cell(row=i+3, column=69).value = ippan_view.farmer_fisher_lv01_49

                ws.cell(row=i+3, column=70).value = ippan_view.farmer_fisher_lv50_99
                ws.cell(row=i+3, column=71).value = ippan_view.farmer_fisher_lv100
                ws.cell(row=i+3, column=72).value = ippan_view.farmer_fisher_full
                ws.cell(row=i+3, column=73).value = ippan_view.farmer_fisher_total
                ws.cell(row=i+3, column=74).value = ippan_view.employee_lv00
                ws.cell(row=i+3, column=75).value = ippan_view.employee_lv01_49
                ws.cell(row=i+3, column=76).value = ippan_view.employee_lv50_99
                ws.cell(row=i+3, column=77).value = ippan_view.employee_lv100
                ws.cell(row=i+3, column=78).value = ippan_view.employee_full
                ws.cell(row=i+3, column=79).value = ippan_view.employee_total

                ws.cell(row=i+3, column=80).value = ippan_view.industry_code
                ws.cell(row=i+3, column=81).value = ippan_view.industry_name
                ws.cell(row=i+3, column=82).value = ippan_view.usage_code
                ws.cell(row=i+3, column=83).value = ippan_view.usage_name
                ws.cell(row=i+3, column=84).value = ippan_view.comment
                ws.cell(row=i+3, column=85).value = str(ippan_view.deleted_at)

        wb.save(download_file_path)
        
        #######################################################################
        ### レスポンスセット処理(0030)
        ### テンプレートとコンテキストを設定して、レスポンスをブラウザに戻す。
        #######################################################################
        print_log('[DEBUG] P0200ExcelDownload.ippan_view_view()関数 STEP 4/4.', 'DEBUG')
        print_log('[INFO] P0200ExcelDownload.ippan_view_view()関数が正常終了しました。', 'INFO')
        response = HttpResponse(content=save_virtual_workbook(wb), content_type='application/vnd.ms-excel')
        response['Content-Disposition'] = 'attachment; filename="ippan_view.xlsx"'
        return response
        
    except:
        print_log('[ERROR] P0200ExcelDownload.ippan_view_view()関数 {}'.format(sys.exc_info()[0]), 'ERROR')
        print_log('[ERROR] P0200ExcelDownload.ippan_view_view()関数でエラーが発生しました。', 'ERROR')
        print_log('[ERROR] P0200ExcelDownload.ippan_view_view()関数が異常終了しました。', 'ERROR')
        return render(request, 'error.html')

###############################################################################
### 関数名：ippan_summary_view(request, lock)
### 8000：集計データ_集計結果
### urlpattern：path('ippan_summary/', views.ippan_summary_view, name='ippan_summary_view')
### template：
###############################################################################
### @login_required(None, login_url='/P0100Login/')
def ippan_summary_view(request, lock):
    try:
        #######################################################################
        ### 引数チェック処理(0000)
        ### ブラウザからのリクエストと引数をチェックする。
        #######################################################################
        ### reset_log()
        print_log('[INFO] P0200ExcelDownload.ippan_summary_view()関数が開始しました。', 'INFO')
        print_log('[DEBUG] P0200ExcelDownload.ippan_summary_view()関数 request = {}'.format(request.method), 'DEBUG')
        print_log('[DEBUG] P0200ExcelDownload.ippan_summary_view()関数 lock = {}'.format(lock), 'DEBUG')
        print_log('[DEBUG] P0200ExcelDownload.ippan_summary_view()関数 STEP 1/4.', 'DEBUG')
        
        #######################################################################
        ### DBアクセス処理(0010)
        ### DBにアクセスして、集計データ_集計結果データを取得する。
        #######################################################################
        print_log('[DEBUG] P0200ExcelDownload.ippan_summary_view()関数 STEP 2/4.', 'DEBUG')
        ippan_summary_list = IPPAN_SUMMARY.objects.raw("""SELECT * FROM IPPAN_SUMMARY ORDER BY CAST(IPPAN_ID AS INTEGER)""", [])
    
        #######################################################################
        ### EXCEL入出力処理(0020)
        ### (1)テンプレート用のEXCELファイルを読み込む。
        ### (2)セルにデータをセットして、ダウンロード用のEXCELファイルを保存する。
        #######################################################################
        print_log('[DEBUG] P0200ExcelDownload.ippan_summary_view()関数 STEP 3/4.', 'DEBUG')
        template_file_path = 'static/template_ippan_summary.xlsx'
        download_file_path = 'static/download_ippan_summary.xlsx'
        wb = openpyxl.load_workbook(template_file_path)
        ws = wb.active
        ws.title = '集計データ_集計結果'
        ws.cell(row=1, column=1).value = '行ID'
        ws.cell(row=1, column=2).value = '水害ID'
        ws.cell(row=1, column=3).value = '家屋被害額_床下'
        ws.cell(row=1, column=4).value = '家屋被害額_01から49cm'
        ws.cell(row=1, column=5).value = '家屋被害額_50から99cm'
        ws.cell(row=1, column=6).value = '家屋被害額_100cm以上'
        ws.cell(row=1, column=7).value = '家屋被害額_半壊'
        ws.cell(row=1, column=8).value = '家屋被害額_全壊'
        ws.cell(row=1, column=9).value = '家庭用品自動車以外被害額_床下'
        ws.cell(row=1, column=10).value = '家庭用品自動車以外被害額_01から49cm'
        ws.cell(row=1, column=11).value = '家庭用品自動車以外被害額_50から99cm'
        ws.cell(row=1, column=12).value = '家庭用品自動車以外被害額_100cm以上'
        ws.cell(row=1, column=13).value = '家庭用品自動車以外被害額_半壊'
        ws.cell(row=1, column=14).value = '家庭用品自動車以外被害額_全壊'
        ws.cell(row=1, column=15).value = '家庭用品自動車被害額_床下'
        ws.cell(row=1, column=16).value = '家庭用品自動車被害額_01から49cm'
        ws.cell(row=1, column=17).value = '家庭用品自動車被害額_50から99cm'
        ws.cell(row=1, column=18).value = '家庭用品自動車被害額_100cm以上'
        ws.cell(row=1, column=19).value = '家庭用品自動車被害額_半壊'
        ws.cell(row=1, column=20).value = '家庭用品自動車被害額_全壊'
        ws.cell(row=1, column=21).value = '家庭応急対策費_代替活動費_床下'
        ws.cell(row=1, column=22).value = '家庭応急対策費_代替活動費_01から49cm'
        ws.cell(row=1, column=23).value = '家庭応急対策費_代替活動費_50から99cm'
        ws.cell(row=1, column=24).value = '家庭応急対策費_代替活動費_100cm以上'
        ws.cell(row=1, column=25).value = '家庭応急対策費_代替活動費_半壊'
        ws.cell(row=1, column=26).value = '家庭応急対策費_代替活動費_全壊'
        ws.cell(row=1, column=27).value = '家庭応急対策費_清掃費_床下'
        ws.cell(row=1, column=28).value = '家庭応急対策費_清掃費_01から49cm'
        ws.cell(row=1, column=29).value = '家庭応急対策費_清掃費_50から99cm'
        ws.cell(row=1, column=30).value = '家庭応急対策費_清掃費_100cm以上'
        ws.cell(row=1, column=31).value = '家庭応急対策費_清掃費_半壊'
        ws.cell(row=1, column=32).value = '家庭応急対策費_清掃費_全壊'
        ws.cell(row=1, column=33).value = '事業所被害額_償却資産被害額_床下'
        ws.cell(row=1, column=34).value = '事業所被害額_償却資産被害額_01から49cm'
        ws.cell(row=1, column=35).value = '事業所被害額_償却資産被害額_50から99cm'
        ws.cell(row=1, column=36).value = '事業所被害額_償却資産被害額_100cm以上'
        ws.cell(row=1, column=37).value = '事業所被害額_償却資産被害額_全壊'
        ws.cell(row=1, column=38).value = '事業所被害額_在庫資産被害額_床下'
        ws.cell(row=1, column=39).value = '事業所被害額_在庫資産被害額_01から49cm'
        ws.cell(row=1, column=40).value = '事業所被害額_在庫資産被害額_50から99cm'
        ws.cell(row=1, column=41).value = '事業所被害額_在庫資産被害額_100cm以上'
        ws.cell(row=1, column=42).value = '事業所被害額_在庫資産被害額_全壊'
        ws.cell(row=1, column=43).value = '事業所被害額_営業停止に伴う被害額_床下'
        ws.cell(row=1, column=44).value = '事業所被害額_営業停止に伴う被害額_01から49cm'
        ws.cell(row=1, column=45).value = '事業所被害額_営業停止に伴う被害額_50から99cm'
        ws.cell(row=1, column=46).value = '事業所被害額_営業停止に伴う被害額_100cm以上'
        ws.cell(row=1, column=47).value = '事業所被害額_営業停止に伴う被害額_全壊'
        ws.cell(row=1, column=48).value = '事業所被害額_営業停滞に伴う被害額_床下'
        ws.cell(row=1, column=49).value = '事業所被害額_営業停滞に伴う被害額_01から49cm'
        ws.cell(row=1, column=50).value = '事業所被害額_営業停滞に伴う被害額_50から99cm'
        ws.cell(row=1, column=51).value = '事業所被害額_営業停滞に伴う被害額_100cm以上'
        ws.cell(row=1, column=52).value = '事業所被害額_営業停滞に伴う被害額_全壊'
        ws.cell(row=1, column=53).value = '農漁家被害額_償却資産被害額_床下'
        ws.cell(row=1, column=54).value = '農漁家被害額_償却資産被害額_01から49cm'
        ws.cell(row=1, column=55).value = '農漁家被害額_償却資産被害額_50から99cm'
        ws.cell(row=1, column=56).value = '農漁家被害額_償却資産被害額_100cm以上'
        ws.cell(row=1, column=57).value = '農漁家被害額_償却資産被害額_全壊'
        ws.cell(row=1, column=58).value = '農漁家被害額_在庫資産被害額_床下'
        ws.cell(row=1, column=59).value = '農漁家被害額_在庫資産被害額_01から49cm'
        ws.cell(row=1, column=60).value = '農漁家被害額_在庫資産被害額_50から99cm'
        ws.cell(row=1, column=61).value = '農漁家被害額_在庫資産被害額_100cm以上'
        ws.cell(row=1, column=62).value = '農漁家被害額_在庫資産被害額_全壊'
        ws.cell(row=1, column=63).value = '事業所応急対策費_代替活動費_床下'
        ws.cell(row=1, column=64).value = '事業所応急対策費_代替活動費_01から49cm'
        ws.cell(row=1, column=65).value = '事業所応急対策費_代替活動費_50から99cm'
        ws.cell(row=1, column=66).value = '事業所応急対策費_代替活動費_100cm以上'
        ws.cell(row=1, column=67).value = '事業所応急対策費_代替活動費_半壊'
        ws.cell(row=1, column=68).value = '事業所応急対策費_代替活動費_全壊'
        ws.cell(row=1, column=69).value = '削除日時'

        ws.cell(row=2, column=1).value = 'ippan_id'
        ws.cell(row=2, column=2).value = 'suigai_id'
        ws.cell(row=2, column=3).value = 'house_summary_lv00'
        ws.cell(row=2, column=4).value = 'house_summary_lv01_49'
        ws.cell(row=2, column=5).value = 'house_summary_lv50_99'
        ws.cell(row=2, column=6).value = 'house_summary_lv100'
        ws.cell(row=2, column=7).value = 'house_summary_half'
        ws.cell(row=2, column=8).value = 'house_summary_full'
        ws.cell(row=2, column=9).value = 'household_summary_lv00'
        ws.cell(row=2, column=10).value = 'household_summary_lv01_49'
        ws.cell(row=2, column=11).value = 'household_summary_lv50_99'
        ws.cell(row=2, column=12).value = 'household_summary_lv100'
        ws.cell(row=2, column=13).value = 'household_summary_half'
        ws.cell(row=2, column=14).value = 'household_summary_full'
        ws.cell(row=2, column=15).value = 'car_summary_lv00'
        ws.cell(row=2, column=16).value = 'car_summary_lv01_49'
        ws.cell(row=2, column=17).value = 'car_summary_lv50_99'
        ws.cell(row=2, column=18).value = 'car_summary_lv100'
        ws.cell(row=2, column=19).value = 'car_summary_half'
        ws.cell(row=2, column=20).value = 'car_summary_full'
        ws.cell(row=2, column=21).value = 'house_alt_lv00'
        ws.cell(row=2, column=22).value = 'house_alt_lv01_49'
        ws.cell(row=2, column=23).value = 'house_alt_lv50_99'
        ws.cell(row=2, column=24).value = 'house_alt_lv100'
        ws.cell(row=2, column=25).value = 'house_alt_half'
        ws.cell(row=2, column=26).value = 'house_alt_full'
        ws.cell(row=2, column=27).value = 'house_clean_lv00'
        ws.cell(row=2, column=28).value = 'house_clean_lv01_49'
        ws.cell(row=2, column=29).value = 'house_clean_lv50_99'
        ws.cell(row=2, column=30).value = 'house_clean_lv100'
        ws.cell(row=2, column=31).value = 'house_clean_half'
        ws.cell(row=2, column=32).value = 'house_clean_full'
        ws.cell(row=2, column=33).value = 'office_dep_summary_lv00'
        ws.cell(row=2, column=34).value = 'office_dep_summary_lv01_49'
        ws.cell(row=2, column=35).value = 'office_dep_summary_lv50_99'
        ws.cell(row=2, column=36).value = 'office_dep_summary_lv100'
        ws.cell(row=2, column=37).value = 'office_dep_summary_full'
        ws.cell(row=2, column=38).value = 'office_inv_summary_lv00'
        ws.cell(row=2, column=39).value = 'office_inv_summary_lv01_49'
        ws.cell(row=2, column=40).value = 'office_inv_summary_lv50_99'
        ws.cell(row=2, column=41).value = 'office_inv_summary_lv100'
        ws.cell(row=2, column=42).value = 'office_inv_summary_full'
        ws.cell(row=2, column=43).value = 'office_sus_summary_lv00'
        ws.cell(row=2, column=44).value = 'office_sus_summary_lv01_49'
        ws.cell(row=2, column=45).value = 'office_sus_summary_lv50_99'
        ws.cell(row=2, column=46).value = 'office_sus_summary_lv100'
        ws.cell(row=2, column=47).value = 'office_sus_summary_full'
        ws.cell(row=2, column=48).value = 'office_stg_summary_lv00'
        ws.cell(row=2, column=49).value = 'office_stg_summary_lv01_49'
        ws.cell(row=2, column=50).value = 'office_stg_summary_lv50_99'
        ws.cell(row=2, column=51).value = 'office_stg_summary_lv100'
        ws.cell(row=2, column=52).value = 'office_stg_summary_full'
        ws.cell(row=2, column=53).value = 'farmer_fisher_dep_summary_lv00'
        ws.cell(row=2, column=54).value = 'farmer_fisher_dep_summary_lv01_49'
        ws.cell(row=2, column=55).value = 'farmer_fisher_dep_summary_lv50_99'
        ws.cell(row=2, column=56).value = 'farmer_fisher_dep_summary_lv100'
        ws.cell(row=2, column=57).value = 'farmer_fisher_dep_summary_full'
        ws.cell(row=2, column=58).value = 'farmer_fisher_inv_summary_lv00'
        ws.cell(row=2, column=59).value = 'farmer_fisher_inv_summary_lv01_49'
        ws.cell(row=2, column=60).value = 'farmer_fisher_inv_summary_lv50_99'
        ws.cell(row=2, column=61).value = 'farmer_fisher_inv_summary_lv100'
        ws.cell(row=2, column=62).value = 'farmer_fisher_inv_summary_full'
        ws.cell(row=2, column=63).value = 'office_alt_summary_lv00'
        ws.cell(row=2, column=64).value = 'office_alt_summary_lv01_49'
        ws.cell(row=2, column=65).value = 'office_alt_summary_lv50_99'
        ws.cell(row=2, column=66).value = 'office_alt_summary_lv100'
        ws.cell(row=2, column=67).value = 'office_alt_summary_half'
        ws.cell(row=2, column=68).value = 'office_alt_summary_full'
        ws.cell(row=2, column=69).value = 'deleted_at'
        
        if ippan_summary_list:
            for i, ippan_summary in enumerate(ippan_summary_list):
                ws.cell(row=i+3, column=1).value = ippan_summary.ippan_id
                ws.cell(row=i+3, column=2).value = ippan_summary.suigai_id
                ws.cell(row=i+3, column=3).value = ippan_summary.house_summary_lv00
                ws.cell(row=i+3, column=4).value = ippan_summary.house_summary_lv01_49
                ws.cell(row=i+3, column=5).value = ippan_summary.house_summary_lv50_99
                ws.cell(row=i+3, column=6).value = ippan_summary.house_summary_lv100
                ws.cell(row=i+3, column=7).value = ippan_summary.house_summary_half
                ws.cell(row=i+3, column=8).value = ippan_summary.house_summary_full
                ws.cell(row=i+3, column=9).value = ippan_summary.household_summary_lv00
                ws.cell(row=i+3, column=10).value = ippan_summary.household_summary_lv01_49
                ws.cell(row=i+3, column=11).value = ippan_summary.household_summary_lv50_99
                ws.cell(row=i+3, column=12).value = ippan_summary.household_summary_lv100
                ws.cell(row=i+3, column=13).value = ippan_summary.household_summary_half
                ws.cell(row=i+3, column=14).value = ippan_summary.household_summary_full
                ws.cell(row=i+3, column=15).value = ippan_summary.car_summary_lv00
                ws.cell(row=i+3, column=16).value = ippan_summary.car_summary_lv01_49
                ws.cell(row=i+3, column=17).value = ippan_summary.car_summary_lv50_99
                ws.cell(row=i+3, column=18).value = ippan_summary.car_summary_lv100
                ws.cell(row=i+3, column=19).value = ippan_summary.car_summary_half
                ws.cell(row=i+3, column=20).value = ippan_summary.car_summary_full
                ws.cell(row=i+3, column=21).value = ippan_summary.house_alt_summary_lv00
                ws.cell(row=i+3, column=22).value = ippan_summary.house_alt_summary_lv01_49
                ws.cell(row=i+3, column=23).value = ippan_summary.house_alt_summary_lv50_99
                ws.cell(row=i+3, column=24).value = ippan_summary.house_alt_summary_lv100
                ws.cell(row=i+3, column=25).value = ippan_summary.house_alt_summary_half
                ws.cell(row=i+3, column=26).value = ippan_summary.house_alt_summary_full
                ws.cell(row=i+3, column=27).value = ippan_summary.house_clean_summary_lv00
                ws.cell(row=i+3, column=28).value = ippan_summary.house_clean_summary_lv01_49
                ws.cell(row=i+3, column=29).value = ippan_summary.house_clean_summary_lv50_99
                ws.cell(row=i+3, column=30).value = ippan_summary.house_clean_summary_lv100
                ws.cell(row=i+3, column=31).value = ippan_summary.house_clean_summary_half
                ws.cell(row=i+3, column=32).value = ippan_summary.house_clean_summary_full
                ws.cell(row=i+3, column=33).value = ippan_summary.office_dep_summary_lv00
                ws.cell(row=i+3, column=34).value = ippan_summary.office_dep_summary_lv01_49
                ws.cell(row=i+3, column=35).value = ippan_summary.office_dep_summary_lv50_99
                ws.cell(row=i+3, column=36).value = ippan_summary.office_dep_summary_lv100
                ws.cell(row=i+3, column=37).value = ippan_summary.office_dep_summary_full
                ws.cell(row=i+3, column=38).value = ippan_summary.office_inv_summary_lv00
                ws.cell(row=i+3, column=39).value = ippan_summary.office_inv_summary_lv01_49
                ws.cell(row=i+3, column=40).value = ippan_summary.office_inv_summary_lv50_99
                ws.cell(row=i+3, column=41).value = ippan_summary.office_inv_summary_lv100
                ws.cell(row=i+3, column=42).value = ippan_summary.office_inv_summary_full
                ws.cell(row=i+3, column=43).value = ippan_summary.office_sus_summary_lv00
                ws.cell(row=i+3, column=44).value = ippan_summary.office_sus_summary_lv01_49
                ws.cell(row=i+3, column=45).value = ippan_summary.office_sus_summary_lv50_99
                ws.cell(row=i+3, column=46).value = ippan_summary.office_sus_summary_lv100
                ws.cell(row=i+3, column=47).value = ippan_summary.office_sus_summary_full
                ws.cell(row=i+3, column=48).value = ippan_summary.office_stg_summary_lv00
                ws.cell(row=i+3, column=49).value = ippan_summary.office_stg_summary_lv01_49
                ws.cell(row=i+3, column=50).value = ippan_summary.office_stg_summary_lv50_99
                ws.cell(row=i+3, column=51).value = ippan_summary.office_stg_summary_lv100
                ws.cell(row=i+3, column=52).value = ippan_summary.office_stg_summary_full
                ws.cell(row=i+3, column=53).value = ippan_summary.farmer_fisher_dep_summary_lv00
                ws.cell(row=i+3, column=54).value = ippan_summary.farmer_fisher_dep_summary_lv01_49
                ws.cell(row=i+3, column=55).value = ippan_summary.farmer_fisher_dep_summary_lv50_99
                ws.cell(row=i+3, column=56).value = ippan_summary.farmer_fisher_dep_summary_lv100
                ws.cell(row=i+3, column=57).value = ippan_summary.farmer_fisher_dep_summary_full
                ws.cell(row=i+3, column=58).value = ippan_summary.farmer_fisher_inv_summary_lv00
                ws.cell(row=i+3, column=59).value = ippan_summary.farmer_fisher_inv_summary_lv01_49
                ws.cell(row=i+3, column=60).value = ippan_summary.farmer_fisher_inv_summary_lv50_99
                ws.cell(row=i+3, column=61).value = ippan_summary.farmer_fisher_inv_summary_lv100
                ws.cell(row=i+3, column=62).value = ippan_summary.farmer_fisher_inv_summary_full
                ws.cell(row=i+3, column=63).value = ippan_summary.office_alt_summary_lv00
                ws.cell(row=i+3, column=64).value = ippan_summary.office_alt_summary_lv01_49
                ws.cell(row=i+3, column=65).value = ippan_summary.office_alt_summary_lv50_99
                ws.cell(row=i+3, column=66).value = ippan_summary.office_alt_summary_lv100
                ws.cell(row=i+3, column=67).value = ippan_summary.office_alt_summary_half
                ws.cell(row=i+3, column=68).value = ippan_summary.office_alt_summary_full
                ws.cell(row=i+3, column=69).value = str(ippan_summary.deleted_at)
            
        wb.save(download_file_path)
        
        #######################################################################
        ### レスポンスセット処理(0030)
        ### テンプレートとコンテキストを設定して、レスポンスをブラウザに戻す。
        #######################################################################
        print_log('[DEBUG] P0200ExcelDownload.ippan_summary_view()関数 STEP 4/4.', 'DEBUG')
        print_log('[INFO] P0200ExcelDownload.ippan_summary_view()関数が正常終了しました。', 'INFO')
        response = HttpResponse(content=save_virtual_workbook(wb), content_type='application/vnd.ms-excel')
        response['Content-Disposition'] = 'attachment; filename="ippan_summary.xlsx"'
        return response
        
    except:
        print_log('[ERROR] P0200ExcelDownload.ippan_summary_view()関数 {}'.format(sys.exc_info()[0]), 'ERROR')
        print_log('[ERROR] P0200ExcelDownload.ippan_summary_view()関数でエラーが発生しました。', 'ERROR')
        print_log('[ERROR] P0200ExcelDownload.ippan_summary_view()関数が異常終了しました。', 'ERROR')
        return render(request, 'error.html')

###############################################################################
### 関数名：ippan_group_by_ken_view(request, lock)
### 8010：集計データ_集計結果_都道府県別
### urlpattern：path('ippan_group_by_ken/', views.ippan_group_by_ken_view, name='ippan_group_by_ken_view')
### template：
###############################################################################
### @login_required(None, login_url='/P0100Login/')
def ippan_group_by_ken_view(request, lock):
    try:
        #######################################################################
        ### 引数チェック処理(0000)
        ### ブラウザからのリクエストと引数をチェックする。
        #######################################################################
        ### reset_log()
        print_log('[INFO] P0200ExcelDownload.ippan_group_by_ken_view()関数が開始しました。', 'INFO')
        print_log('[DEBUG] P0200ExcelDownload.ippan_group_by_ken_view()関数 request = {}'.format(request.method), 'DEBUG')
        print_log('[DEBUG] P0200ExcelDownload.ippan_group_by_ken_view()関数 lock = {}'.format(lock), 'DEBUG')
        print_log('[DEBUG] P0200ExcelDownload.ippan_group_by_ken_view()関数 STEP 1/4.', 'DEBUG')
        
        #######################################################################
        ### DBアクセス処理(0010)
        ### DBにアクセスして、集計データ_集計結果_都道府県別データを取得する。
        #######################################################################
        print_log('[DEBUG] P0200ExcelDownload.ippan_group_by_ken_view()関数 STEP 2/4.', 'DEBUG')
        ippan_group_by_ken_list = IPPAN_SUMMARY.objects.raw("""
            SELECT 
                SG1.ken_code AS id, 
                SUM(IS1.house_summary_lv00) AS house_summary_lv00, 
                SUM(IS1.house_summary_lv01_49) AS house_summary_lv01_49, 
                SUM(IS1.house_summary_lv50_99) AS house_summary_lv50_99, 
                SUM(IS1.house_summary_lv100) AS house_summary_lv100, 
                SUM(IS1.house_summary_half) AS house_summary_half, 
                SUM(IS1.house_summary_full) AS house_summary_full, 

                SUM(IS1.household_summary_lv00) AS household_summary_lv00, 
                SUM(IS1.household_summary_lv01_49) AS household_summary_lv01_49, 
                SUM(IS1.household_summary_lv50_99) AS household_summary_lv50_99, 
                SUM(IS1.household_summary_lv100) AS household_summary_lv100, 
                SUM(IS1.household_summary_half) AS household_summary_half, 
                SUM(IS1.household_summary_full) AS household_summary_full, 

                SUM(IS1.car_summary_lv00) AS car_summary_lv00, 
                SUM(IS1.car_summary_lv01_49) AS car_summary_lv01_49, 
                SUM(IS1.car_summary_lv50_99) AS car_summary_lv50_99, 
                SUM(IS1.car_summary_lv100) AS car_summary_lv100, 
                SUM(IS1.car_summary_half) AS car_summary_half, 
                SUM(IS1.car_summary_full) AS car_summary_full, 

                SUM(IS1.house_alt_summary_lv00) AS house_alt_summary_lv00, 
                SUM(IS1.house_alt_summary_lv01_49) AS house_alt_summary_lv01_49, 
                SUM(IS1.house_alt_summary_lv50_99) AS house_alt_summary_lv50_99, 
                SUM(IS1.house_alt_summary_lv100) AS house_alt_summary_lv100, 
                SUM(IS1.house_alt_summary_half) AS house_alt_summary_half, 
                SUM(IS1.house_alt_summary_full) AS house_alt_summary_full, 

                SUM(IS1.house_clean_summary_lv00) AS house_clean_summary_lv00, 
                SUM(IS1.house_clean_summary_lv01_49) AS house_clean_summary_lv01_49, 
                SUM(IS1.house_clean_summary_lv50_99) AS house_clean_summary_lv50_99, 
                SUM(IS1.house_clean_summary_lv100) AS house_clean_summary_lv100, 
                SUM(IS1.house_clean_summary_half) AS house_clean_summary_half, 
                SUM(IS1.house_clean_summary_full) AS house_clean_summary_full, 

                SUM(IS1.office_dep_summary_lv00) AS office_dep_summary_lv00, 
                SUM(IS1.office_dep_summary_lv01_49) AS office_dep_summary_lv01_49, 
                SUM(IS1.office_dep_summary_lv50_99) AS office_dep_summary_lv50_99, 
                SUM(IS1.office_dep_summary_lv100) AS office_dep_summary_lv100, 
                -- SUM(IS1.office_dep_summary_half) AS office_dep_summary_half, 
                SUM(IS1.office_dep_summary_full) AS office_dep_summary_full, 

                SUM(IS1.office_inv_summary_lv00) AS office_inv_summary_lv00, 
                SUM(IS1.office_inv_summary_lv01_49) AS office_inv_summary_lv01_49, 
                SUM(IS1.office_inv_summary_lv50_99) AS office_inv_summary_lv50_99, 
                SUM(IS1.office_inv_summary_lv100) AS office_inv_summary_lv100, 
                -- SUM(IS1.office_inv_summary_half) AS office_inv_summary_half, 
                SUM(IS1.office_inv_summary_full) AS office_inv_summary_full, 

                SUM(IS1.office_sus_summary_lv00) AS office_sus_summary_lv00, 
                SUM(IS1.office_sus_summary_lv01_49) AS office_sus_summary_lv01_49, 
                SUM(IS1.office_sus_summary_lv50_99) AS office_sus_summary_lv50_99, 
                SUM(IS1.office_sus_summary_lv100) AS office_sus_summary_lv100, 
                -- SUM(IS1.office_sus_summary_half) AS office_sus_summary_half, 
                SUM(IS1.office_sus_summary_full) AS office_sus_summary_full, 

                SUM(IS1.office_stg_summary_lv00) AS office_stg_summary_lv00, 
                SUM(IS1.office_stg_summary_lv01_49) AS office_stg_summary_lv01_49, 
                SUM(IS1.office_stg_summary_lv50_99) AS office_stg_summary_lv50_99, 
                SUM(IS1.office_stg_summary_lv100) AS office_stg_summary_lv100, 
                -- SUM(IS1.office_stg_summary_half) AS office_stg_summary_half, 
                SUM(IS1.office_stg_summary_full) AS office_stg_summary_full, 

                SUM(IS1.farmer_fisher_dep_summary_lv00) AS farmer_fisher_dep_summary_lv00, 
                SUM(IS1.farmer_fisher_dep_summary_lv01_49) AS farmer_fisher_dep_summary_lv01_49, 
                SUM(IS1.farmer_fisher_dep_summary_lv50_99) AS farmer_fisher_dep_summary_lv50_99, 
                SUM(IS1.farmer_fisher_dep_summary_lv100) AS farmer_fisher_dep_summary_lv100, 
                -- SUM(IS1.farmer_fisher_dep_summary_half) AS farmer_fisher_dep_summary_half, 
                SUM(IS1.farmer_fisher_dep_summary_full) AS farmer_fisher_dep_summary_full, 

                SUM(IS1.farmer_fisher_inv_summary_lv00) AS farmer_fisher_inv_summary_lv00, 
                SUM(IS1.farmer_fisher_inv_summary_lv01_49) AS farmer_fisher_inv_summary_lv01_49, 
                SUM(IS1.farmer_fisher_inv_summary_lv50_99) AS farmer_fisher_inv_summary_lv50_99, 
                SUM(IS1.farmer_fisher_inv_summary_lv100) AS farmer_fisher_inv_summary_lv100, 
                -- SUM(IS1.farmer_fisher_inv_summary_half) AS farmer_fisher_inv_summary_half, 
                SUM(IS1.farmer_fisher_inv_summary_full) AS farmer_fisher_inv_summary_full, 

                SUM(IS1.office_alt_summary_lv00) AS office_alt_summary_lv00, 
                SUM(IS1.office_alt_summary_lv01_49) AS office_alt_summary_lv01_49, 
                SUM(IS1.office_alt_summary_lv50_99) AS office_alt_summary_lv50_99, 
                SUM(IS1.office_alt_summary_lv100) AS office_alt_summary_lv100, 
                SUM(IS1.office_alt_summary_half) AS office_alt_summary_half, 
                SUM(IS1.office_alt_summary_full) AS office_alt_summary_full 
                
            FROM IPPAN_SUMMARY IS1 
            LEFT JOIN SUIGAI SG1 ON IS1.suigai_id = SG1.suigai_id 
            GROUP BY SG1.ken_code 
            ORDER BY CAST(SG1.KEN_CODE AS INTEGER)
        """, [])
    
        #######################################################################
        ### EXCEL入出力処理(0020)
        ### (1)テンプレート用のEXCELファイルを読み込む。
        ### (2)セルにデータをセットして、ダウンロード用のEXCELファイルを保存する。
        #######################################################################
        print_log('[DEBUG] P0200ExcelDownload.ippan_group_by_ken_view()関数 STEP 3/4.', 'DEBUG')
        template_file_path = 'static/template_ippan_group_by_ken.xlsx'
        download_file_path = 'static/download_ippan_group_by_ken.xlsx'
        wb = openpyxl.load_workbook(template_file_path)
        ws = wb.active
        ws.title = '一般資産集計データ_集計結果_都道府県別'
        ws.cell(row=1, column=1).value = '都道府県コード'
        ### ws.cell(row=1, column=2).value = '一般資産調査票ID'
        ### ws.cell(row=1, column=3).value = '水害ID'
        ws.cell(row=1, column=2).value = '家屋被害額_床下'
        ws.cell(row=1, column=3).value = '家屋被害額_01から49cm'
        ws.cell(row=1, column=4).value = '家屋被害額_50から99cm'
        ws.cell(row=1, column=5).value = '家屋被害額_100cm以上'
        ws.cell(row=1, column=6).value = '家屋被害額_半壊'
        ws.cell(row=1, column=7).value = '家屋被害額_全壊'
        ws.cell(row=1, column=8).value = '家庭用品自動車以外被害額_床下'
        ws.cell(row=1, column=9).value = '家庭用品自動車以外被害額_01から49cm'
        ws.cell(row=1, column=10).value = '家庭用品自動車以外被害額_50から99cm'
        ws.cell(row=1, column=11).value = '家庭用品自動車以外被害額_100cm以上'
        ws.cell(row=1, column=12).value = '家庭用品自動車以外被害額_半壊'
        ws.cell(row=1, column=13).value = '家庭用品自動車以外被害額_全壊'
        ws.cell(row=1, column=14).value = '家庭用品自動車被害額_床下'
        ws.cell(row=1, column=15).value = '家庭用品自動車被害額_01から49cm'
        ws.cell(row=1, column=16).value = '家庭用品自動車被害額_50から99cm'
        ws.cell(row=1, column=17).value = '家庭用品自動車被害額_100cm以上'
        ws.cell(row=1, column=18).value = '家庭用品自動車被害額_半壊'
        ws.cell(row=1, column=19).value = '家庭用品自動車被害額_全壊'
        ws.cell(row=1, column=20).value = '家庭応急対策費_代替活動費_床下'
        ws.cell(row=1, column=21).value = '家庭応急対策費_代替活動費_01から49cm'
        ws.cell(row=1, column=22).value = '家庭応急対策費_代替活動費_50から99cm'
        ws.cell(row=1, column=23).value = '家庭応急対策費_代替活動費_100cm以上'
        ws.cell(row=1, column=24).value = '家庭応急対策費_代替活動費_半壊'
        ws.cell(row=1, column=25).value = '家庭応急対策費_代替活動費_全壊'
        ws.cell(row=1, column=26).value = '家庭応急対策費_清掃費_床下'
        ws.cell(row=1, column=27).value = '家庭応急対策費_清掃費_01から49cm'
        ws.cell(row=1, column=28).value = '家庭応急対策費_清掃費_50から99cm'
        ws.cell(row=1, column=29).value = '家庭応急対策費_清掃費_100cm以上'
        ws.cell(row=1, column=30).value = '家庭応急対策費_清掃費_半壊'
        ws.cell(row=1, column=31).value = '家庭応急対策費_清掃費_全壊'
        ws.cell(row=1, column=32).value = '事業所被害額_償却資産被害額_床下'
        ws.cell(row=1, column=33).value = '事業所被害額_償却資産被害額_01から49cm'
        ws.cell(row=1, column=34).value = '事業所被害額_償却資産被害額_50から99cm'
        ws.cell(row=1, column=35).value = '事業所被害額_償却資産被害額_100cm以上'
        ws.cell(row=1, column=36).value = '事業所被害額_償却資産被害額_全壊'
        ws.cell(row=1, column=37).value = '事業所被害額_在庫資産被害額_床下'
        ws.cell(row=1, column=38).value = '事業所被害額_在庫資産被害額_01から49cm'
        ws.cell(row=1, column=39).value = '事業所被害額_在庫資産被害額_50から99cm'
        ws.cell(row=1, column=40).value = '事業所被害額_在庫資産被害額_100cm以上'
        ws.cell(row=1, column=41).value = '事業所被害額_在庫資産被害額_全壊'
        ws.cell(row=1, column=42).value = '事業所被害額_営業停止に伴う被害額_床下'
        ws.cell(row=1, column=43).value = '事業所被害額_営業停止に伴う被害額_01から49cm'
        ws.cell(row=1, column=44).value = '事業所被害額_営業停止に伴う被害額_50から99cm'
        ws.cell(row=1, column=45).value = '事業所被害額_営業停止に伴う被害額_100cm以上'
        ws.cell(row=1, column=46).value = '事業所被害額_営業停止に伴う被害額_全壊'
        ws.cell(row=1, column=47).value = '事業所被害額_営業停滞に伴う被害額_床下'
        ws.cell(row=1, column=48).value = '事業所被害額_営業停滞に伴う被害額_01から49cm'
        ws.cell(row=1, column=49).value = '事業所被害額_営業停滞に伴う被害額_50から99cm'
        ws.cell(row=1, column=50).value = '事業所被害額_営業停滞に伴う被害額_100cm以上'
        ws.cell(row=1, column=51).value = '事業所被害額_営業停滞に伴う被害額_全壊'
        ws.cell(row=1, column=52).value = '農漁家被害額_償却資産被害額_床下'
        ws.cell(row=1, column=53).value = '農漁家被害額_償却資産被害額_01から49cm'
        ws.cell(row=1, column=54).value = '農漁家被害額_償却資産被害額_50から99cm'
        ws.cell(row=1, column=55).value = '農漁家被害額_償却資産被害額_100cm以上'
        ws.cell(row=1, column=56).value = '農漁家被害額_償却資産被害額_全壊'
        ws.cell(row=1, column=57).value = '農漁家被害額_在庫資産被害額_床下'
        ws.cell(row=1, column=58).value = '農漁家被害額_在庫資産被害額_01から49cm'
        ws.cell(row=1, column=59).value = '農漁家被害額_在庫資産被害額_50から99cm'
        ws.cell(row=1, column=60).value = '農漁家被害額_在庫資産被害額_100cm以上'
        ws.cell(row=1, column=61).value = '農漁家被害額_在庫資産被害額_全壊'
        ws.cell(row=1, column=62).value = '事業所応急対策費_代替活動費_床下'
        ws.cell(row=1, column=63).value = '事業所応急対策費_代替活動費_01から49cm'
        ws.cell(row=1, column=64).value = '事業所応急対策費_代替活動費_50から99cm'
        ws.cell(row=1, column=65).value = '事業所応急対策費_代替活動費_100cm以上'
        ws.cell(row=1, column=66).value = '事業所応急対策費_代替活動費_半壊'
        ws.cell(row=1, column=67).value = '事業所応急対策費_代替活動費_全壊'
        
        if ippan_group_by_ken_list:
            for i, ippan_group_by_ken in enumerate(ippan_group_by_ken_list):
                ws.cell(row=i+2, column=1).value = ippan_group_by_ken.id
                ### ws.cell(row=i+2, column=2).value = ippan_group_by_ken.ippan_id
                ### ws.cell(row=i+2, column=3).value = ippan_group_by_ken.suigai_id
                ws.cell(row=i+2, column=2).value = ippan_group_by_ken.house_summary_lv00
                ws.cell(row=i+2, column=3).value = ippan_group_by_ken.house_summary_lv01_49
                ws.cell(row=i+2, column=4).value = ippan_group_by_ken.house_summary_lv50_99
                ws.cell(row=i+2, column=5).value = ippan_group_by_ken.house_summary_lv100
                ws.cell(row=i+2, column=6).value = ippan_group_by_ken.house_summary_half
                ws.cell(row=i+2, column=7).value = ippan_group_by_ken.house_summary_full
                ws.cell(row=i+2, column=8).value = ippan_group_by_ken.household_summary_lv00
                ws.cell(row=i+2, column=9).value = ippan_group_by_ken.household_summary_lv01_49
                ws.cell(row=i+2, column=10).value = ippan_group_by_ken.household_summary_lv50_99
                ws.cell(row=i+2, column=11).value = ippan_group_by_ken.household_summary_lv100
                ws.cell(row=i+2, column=12).value = ippan_group_by_ken.household_summary_half
                ws.cell(row=i+2, column=13).value = ippan_group_by_ken.household_summary_full
                ws.cell(row=i+2, column=14).value = ippan_group_by_ken.car_summary_lv00
                ws.cell(row=i+2, column=15).value = ippan_group_by_ken.car_summary_lv01_49
                ws.cell(row=i+2, column=16).value = ippan_group_by_ken.car_summary_lv50_99
                ws.cell(row=i+2, column=17).value = ippan_group_by_ken.car_summary_lv100
                ws.cell(row=i+2, column=18).value = ippan_group_by_ken.car_summary_half
                ws.cell(row=i+2, column=19).value = ippan_group_by_ken.car_summary_full
                ws.cell(row=i+2, column=20).value = ippan_group_by_ken.house_alt_summary_lv00
                ws.cell(row=i+2, column=21).value = ippan_group_by_ken.house_alt_summary_lv01_49
                ws.cell(row=i+2, column=22).value = ippan_group_by_ken.house_alt_summary_lv50_99
                ws.cell(row=i+2, column=23).value = ippan_group_by_ken.house_alt_summary_lv100
                ws.cell(row=i+2, column=24).value = ippan_group_by_ken.house_alt_summary_half
                ws.cell(row=i+2, column=25).value = ippan_group_by_ken.house_alt_summary_full
                ws.cell(row=i+2, column=26).value = ippan_group_by_ken.house_clean_summary_lv00
                ws.cell(row=i+2, column=27).value = ippan_group_by_ken.house_clean_summary_lv01_49
                ws.cell(row=i+2, column=28).value = ippan_group_by_ken.house_clean_summary_lv50_99
                ws.cell(row=i+2, column=29).value = ippan_group_by_ken.house_clean_summary_lv100
                ws.cell(row=i+2, column=30).value = ippan_group_by_ken.house_clean_summary_half
                ws.cell(row=i+2, column=31).value = ippan_group_by_ken.house_clean_summary_full
                ws.cell(row=i+2, column=32).value = ippan_group_by_ken.office_dep_summary_lv00
                ws.cell(row=i+2, column=33).value = ippan_group_by_ken.office_dep_summary_lv01_49
                ws.cell(row=i+2, column=34).value = ippan_group_by_ken.office_dep_summary_lv50_99
                ws.cell(row=i+2, column=35).value = ippan_group_by_ken.office_dep_summary_lv100
                ws.cell(row=i+2, column=36).value = ippan_group_by_ken.office_dep_summary_full
                ws.cell(row=i+2, column=37).value = ippan_group_by_ken.office_inv_summary_lv00
                ws.cell(row=i+2, column=38).value = ippan_group_by_ken.office_inv_summary_lv01_49
                ws.cell(row=i+2, column=39).value = ippan_group_by_ken.office_inv_summary_lv50_99
                ws.cell(row=i+2, column=40).value = ippan_group_by_ken.office_inv_summary_lv100
                ws.cell(row=i+2, column=41).value = ippan_group_by_ken.office_inv_summary_full
                ws.cell(row=i+2, column=42).value = ippan_group_by_ken.office_sus_summary_lv00
                ws.cell(row=i+2, column=43).value = ippan_group_by_ken.office_sus_summary_lv01_49
                ws.cell(row=i+2, column=44).value = ippan_group_by_ken.office_sus_summary_lv50_99
                ws.cell(row=i+2, column=45).value = ippan_group_by_ken.office_sus_summary_lv100
                ws.cell(row=i+2, column=46).value = ippan_group_by_ken.office_sus_summary_full
                ws.cell(row=i+2, column=47).value = ippan_group_by_ken.office_stg_summary_lv00
                ws.cell(row=i+2, column=48).value = ippan_group_by_ken.office_stg_summary_lv01_49
                ws.cell(row=i+2, column=49).value = ippan_group_by_ken.office_stg_summary_lv50_99
                ws.cell(row=i+2, column=50).value = ippan_group_by_ken.office_stg_summary_lv100
                ws.cell(row=i+2, column=51).value = ippan_group_by_ken.office_stg_summary_full
                ws.cell(row=i+2, column=52).value = ippan_group_by_ken.farmer_fisher_dep_summary_lv00
                ws.cell(row=i+2, column=53).value = ippan_group_by_ken.farmer_fisher_dep_summary_lv01_49
                ws.cell(row=i+2, column=54).value = ippan_group_by_ken.farmer_fisher_dep_summary_lv50_99
                ws.cell(row=i+2, column=55).value = ippan_group_by_ken.farmer_fisher_dep_summary_lv100
                ws.cell(row=i+2, column=56).value = ippan_group_by_ken.farmer_fisher_dep_summary_full
                ws.cell(row=i+2, column=57).value = ippan_group_by_ken.farmer_fisher_inv_summary_lv00
                ws.cell(row=i+2, column=58).value = ippan_group_by_ken.farmer_fisher_inv_summary_lv01_49
                ws.cell(row=i+2, column=59).value = ippan_group_by_ken.farmer_fisher_inv_summary_lv50_99
                ws.cell(row=i+2, column=60).value = ippan_group_by_ken.farmer_fisher_inv_summary_lv100
                ws.cell(row=i+2, column=61).value = ippan_group_by_ken.farmer_fisher_inv_summary_full
                ws.cell(row=i+2, column=62).value = ippan_group_by_ken.office_alt_summary_lv00
                ws.cell(row=i+2, column=63).value = ippan_group_by_ken.office_alt_summary_lv01_49
                ws.cell(row=i+2, column=64).value = ippan_group_by_ken.office_alt_summary_lv50_99
                ws.cell(row=i+2, column=65).value = ippan_group_by_ken.office_alt_summary_lv100
                ws.cell(row=i+2, column=66).value = ippan_group_by_ken.office_alt_summary_half
                ws.cell(row=i+2, column=67).value = ippan_group_by_ken.office_alt_summary_full
            
        wb.save(download_file_path)
        
        #######################################################################
        ### レスポンスセット処理(0030)
        ### テンプレートとコンテキストを設定して、レスポンスをブラウザに戻す。
        #######################################################################
        print_log('[DEBUG] P0200ExcelDownload.ippan_group_by_ken_view()関数 STEP 4/4.', 'DEBUG')
        print_log('[INFO] P0200ExcelDownload.ippan_group_by_ken_view()関数が正常終了しました。', 'INFO')
        response = HttpResponse(content=save_virtual_workbook(wb), content_type='application/vnd.ms-excel')
        response['Content-Disposition'] = 'attachment; filename="ippan_group_by_ken.xlsx"'
        return response
        
    except:
        print_log('[ERROR] P0200ExcelDownload.ippan_group_by_ken_view()関数 {}'.format(sys.exc_info()[0]), 'ERROR')
        print_log('[ERROR] P0200ExcelDownload.ippan_group_by_ken_view()関数でエラーが発生しました。', 'ERROR')
        print_log('[ERROR] P0200ExcelDownload.ippan_group_by_ken_view()関数が異常終了しました。', 'ERROR')
        return render(request, 'error.html')

###############################################################################
### 関数名：ippan_group_by_suikei_view(request, lock)
### 8020：集計データ_集計結果_水系別
### urlpattern：path('ippan_group_by_suikei/', views.ippan_group_by_suikei_view, name='ippan_group_by_suikei')
### template：
###############################################################################
### @login_required(None, login_url='/P0100Login/')
def ippan_group_by_suikei_view(request, lock):
    try:
        #######################################################################
        ### 引数チェック処理(0000)
        ### ブラウザからのリクエストと引数をチェックする。
        #######################################################################
        ### reset_log()
        print_log('[INFO] P0200ExcelDownload.ippan_group_by_suikei_view()関数が開始しました。', 'INFO')
        print_log('[DEBUG] P0200ExcelDownload.ippan_group_by_suikei_view()関数 request = {}'.format(request.method), 'DEBUG')
        print_log('[DEBUG] P0200ExcelDownload.ippan_group_by_suikei_view()関数 lock = {}'.format(lock), 'DEBUG')
        print_log('[DEBUG] P0200ExcelDownload.ippan_group_by_suikei_view()関数 STEP 1/4.', 'DEBUG')
        
        #######################################################################
        ### DBアクセス処理(0010)
        ### DBにアクセスして、一般資産集計データ_集計結果_水系別データを取得する。
        #######################################################################
        print_log('[DEBUG] P0200ExcelDownload.ippan_group_by_suikei_view()関数 STEP 2/4.', 'DEBUG')
        ippan_group_by_suikei_list = IPPAN_SUMMARY.objects.raw("""
            SELECT 
                SG1.suikei_code AS id, 
                SUM(IS1.house_summary_lv00) AS house_summary_lv00, 
                SUM(IS1.house_summary_lv01_49) AS house_summary_lv01_49, 
                SUM(IS1.house_summary_lv50_99) AS house_summary_lv50_99, 
                SUM(IS1.house_summary_lv100) AS house_summary_lv100, 
                SUM(IS1.house_summary_half) AS house_summary_half, 
                SUM(IS1.house_summary_full) AS house_summary_full, 

                SUM(IS1.household_summary_lv00) AS household_summary_lv00, 
                SUM(IS1.household_summary_lv01_49) AS household_summary_lv01_49, 
                SUM(IS1.household_summary_lv50_99) AS household_summary_lv50_99, 
                SUM(IS1.household_summary_lv100) AS household_summary_lv100, 
                SUM(IS1.household_summary_half) AS household_summary_half, 
                SUM(IS1.household_summary_full) AS household_summary_full, 

                SUM(IS1.car_summary_lv00) AS car_summary_lv00, 
                SUM(IS1.car_summary_lv01_49) AS car_summary_lv01_49, 
                SUM(IS1.car_summary_lv50_99) AS car_summary_lv50_99, 
                SUM(IS1.car_summary_lv100) AS car_summary_lv100, 
                SUM(IS1.car_summary_half) AS car_summary_half, 
                SUM(IS1.car_summary_full) AS car_summary_full, 

                SUM(IS1.house_alt_summary_lv00) AS house_alt_summary_lv00, 
                SUM(IS1.house_alt_summary_lv01_49) AS house_alt_summary_lv01_49, 
                SUM(IS1.house_alt_summary_lv50_99) AS house_alt_summary_lv50_99, 
                SUM(IS1.house_alt_summary_lv100) AS house_alt_summary_lv100, 
                SUM(IS1.house_alt_summary_half) AS house_alt_summary_half, 
                SUM(IS1.house_alt_summary_full) AS house_alt_summary_full, 

                SUM(IS1.house_clean_summary_lv00) AS house_clean_summary_lv00, 
                SUM(IS1.house_clean_summary_lv01_49) AS house_clean_summary_lv01_49, 
                SUM(IS1.house_clean_summary_lv50_99) AS house_clean_summary_lv50_99, 
                SUM(IS1.house_clean_summary_lv100) AS house_clean_summary_lv100, 
                SUM(IS1.house_clean_summary_half) AS house_clean_summary_half, 
                SUM(IS1.house_clean_summary_full) AS house_clean_summary_full, 

                SUM(IS1.office_dep_summary_lv00) AS office_dep_summary_lv00, 
                SUM(IS1.office_dep_summary_lv01_49) AS office_dep_summary_lv01_49, 
                SUM(IS1.office_dep_summary_lv50_99) AS office_dep_summary_lv50_99, 
                SUM(IS1.office_dep_summary_lv100) AS office_dep_summary_lv100, 
                -- SUM(IS1.office_dep_summary_half) AS office_dep_summary_half, 
                SUM(IS1.office_dep_summary_full) AS office_dep_summary_full, 

                SUM(IS1.office_inv_summary_lv00) AS office_inv_summary_lv00, 
                SUM(IS1.office_inv_summary_lv01_49) AS office_inv_summary_lv01_49, 
                SUM(IS1.office_inv_summary_lv50_99) AS office_inv_summary_lv50_99, 
                SUM(IS1.office_inv_summary_lv100) AS office_inv_summary_lv100, 
                -- SUM(IS1.office_inv_summary_half) AS office_inv_summary_half, 
                SUM(IS1.office_inv_summary_full) AS office_inv_summary_full, 

                SUM(IS1.office_sus_summary_lv00) AS office_sus_summary_lv00, 
                SUM(IS1.office_sus_summary_lv01_49) AS office_sus_summary_lv01_49, 
                SUM(IS1.office_sus_summary_lv50_99) AS office_sus_summary_lv50_99, 
                SUM(IS1.office_sus_summary_lv100) AS office_sus_summary_lv100, 
                -- SUM(IS1.office_sus_summary_half) AS office_sus_summary_half, 
                SUM(IS1.office_sus_summary_full) AS office_sus_summary_full, 

                SUM(IS1.office_stg_summary_lv00) AS office_stg_summary_lv00, 
                SUM(IS1.office_stg_summary_lv01_49) AS office_stg_summary_lv01_49, 
                SUM(IS1.office_stg_summary_lv50_99) AS office_stg_summary_lv50_99, 
                SUM(IS1.office_stg_summary_lv100) AS office_stg_summary_lv100, 
                -- SUM(IS1.office_stg_summary_half) AS office_stg_summary_half, 
                SUM(IS1.office_stg_summary_full) AS office_stg_summary_full, 

                SUM(IS1.farmer_fisher_dep_summary_lv00) AS farmer_fisher_dep_summary_lv00, 
                SUM(IS1.farmer_fisher_dep_summary_lv01_49) AS farmer_fisher_dep_summary_lv01_49, 
                SUM(IS1.farmer_fisher_dep_summary_lv50_99) AS farmer_fisher_dep_summary_lv50_99, 
                SUM(IS1.farmer_fisher_dep_summary_lv100) AS farmer_fisher_dep_summary_lv100, 
                -- SUM(IS1.farmer_fisher_dep_summary_half) AS farmer_fisher_dep_summary_half, 
                SUM(IS1.farmer_fisher_dep_summary_full) AS farmer_fisher_dep_summary_full, 

                SUM(IS1.farmer_fisher_inv_summary_lv00) AS farmer_fisher_inv_summary_lv00, 
                SUM(IS1.farmer_fisher_inv_summary_lv01_49) AS farmer_fisher_inv_summary_lv01_49, 
                SUM(IS1.farmer_fisher_inv_summary_lv50_99) AS farmer_fisher_inv_summary_lv50_99, 
                SUM(IS1.farmer_fisher_inv_summary_lv100) AS farmer_fisher_inv_summary_lv100, 
                -- SUM(IS1.farmer_fisher_inv_summary_half) AS farmer_fisher_inv_summary_half, 
                SUM(IS1.farmer_fisher_inv_summary_full) AS farmer_fisher_inv_summary_full, 

                SUM(IS1.office_alt_summary_lv00) AS office_alt_summary_lv00, 
                SUM(IS1.office_alt_summary_lv01_49) AS office_alt_summary_lv01_49, 
                SUM(IS1.office_alt_summary_lv50_99) AS office_alt_summary_lv50_99, 
                SUM(IS1.office_alt_summary_lv100) AS office_alt_summary_lv100, 
                SUM(IS1.office_alt_summary_half) AS office_alt_summary_half, 
                SUM(IS1.office_alt_summary_full) AS office_alt_summary_full 
                
            FROM IPPAN_SUMMARY IS1 
            LEFT JOIN SUIGAI SG1 ON IS1.suigai_id = SG1.suigai_id 
            GROUP BY SG1.suikei_code 
            ORDER BY CAST(SG1.SUIKEI_CODE AS INTEGER)
        """, [])
    
        #######################################################################
        ### EXCEL入出力処理(0020)
        ### (1)テンプレート用のEXCELファイルを読み込む。
        ### (2)セルにデータをセットして、ダウンロード用のEXCELファイルを保存する。
        #######################################################################
        print_log('[DEBUG] P0200ExcelDownload.ippan_group_by_suikei_view()関数 STEP 3/4.', 'DEBUG')
        template_file_path = 'static/template_ippan_group_by_suikei.xlsx'
        download_file_path = 'static/download_ippan_group_by_suikei.xlsx'
        wb = openpyxl.load_workbook(template_file_path)
        ws = wb.active
        ws.title = '一般資産集計データ_集計結果_水系別'
        ws.cell(row=1, column=1).value = '水系コード'
        ### ws.cell(row=1, column=2).value = '一般資産調査票ID'
        ### ws.cell(row=1, column=3).value = '水害ID'
        ws.cell(row=1, column=2).value = '家屋被害額_床下'
        ws.cell(row=1, column=3).value = '家屋被害額_01から49cm'
        ws.cell(row=1, column=4).value = '家屋被害額_50から99cm'
        ws.cell(row=1, column=5).value = '家屋被害額_100cm以上'
        ws.cell(row=1, column=6).value = '家屋被害額_半壊'
        ws.cell(row=1, column=7).value = '家屋被害額_全壊'
        ws.cell(row=1, column=8).value = '家庭用品自動車以外被害額_床下'
        ws.cell(row=1, column=9).value = '家庭用品自動車以外被害額_01から49cm'
        ws.cell(row=1, column=10).value = '家庭用品自動車以外被害額_50から99cm'
        ws.cell(row=1, column=11).value = '家庭用品自動車以外被害額_100cm以上'
        ws.cell(row=1, column=12).value = '家庭用品自動車以外被害額_半壊'
        ws.cell(row=1, column=13).value = '家庭用品自動車以外被害額_全壊'
        ws.cell(row=1, column=14).value = '家庭用品自動車被害額_床下'
        ws.cell(row=1, column=15).value = '家庭用品自動車被害額_01から49cm'
        ws.cell(row=1, column=16).value = '家庭用品自動車被害額_50から99cm'
        ws.cell(row=1, column=17).value = '家庭用品自動車被害額_100cm以上'
        ws.cell(row=1, column=18).value = '家庭用品自動車被害額_半壊'
        ws.cell(row=1, column=19).value = '家庭用品自動車被害額_全壊'
        ws.cell(row=1, column=20).value = '家庭応急対策費_代替活動費_床下'
        ws.cell(row=1, column=21).value = '家庭応急対策費_代替活動費_01から49cm'
        ws.cell(row=1, column=22).value = '家庭応急対策費_代替活動費_50から99cm'
        ws.cell(row=1, column=23).value = '家庭応急対策費_代替活動費_100cm以上'
        ws.cell(row=1, column=24).value = '家庭応急対策費_代替活動費_半壊'
        ws.cell(row=1, column=25).value = '家庭応急対策費_代替活動費_全壊'
        ws.cell(row=1, column=26).value = '家庭応急対策費_清掃費_床下'
        ws.cell(row=1, column=27).value = '家庭応急対策費_清掃費_01から49cm'
        ws.cell(row=1, column=28).value = '家庭応急対策費_清掃費_50から99cm'
        ws.cell(row=1, column=29).value = '家庭応急対策費_清掃費_100cm以上'
        ws.cell(row=1, column=30).value = '家庭応急対策費_清掃費_半壊'
        ws.cell(row=1, column=31).value = '家庭応急対策費_清掃費_全壊'
        ws.cell(row=1, column=32).value = '事業所被害額_償却資産被害額_床下'
        ws.cell(row=1, column=33).value = '事業所被害額_償却資産被害額_01から49cm'
        ws.cell(row=1, column=34).value = '事業所被害額_償却資産被害額_50から99cm'
        ws.cell(row=1, column=35).value = '事業所被害額_償却資産被害額_100cm以上'
        ws.cell(row=1, column=36).value = '事業所被害額_償却資産被害額_全壊'
        ws.cell(row=1, column=37).value = '事業所被害額_在庫資産被害額_床下'
        ws.cell(row=1, column=38).value = '事業所被害額_在庫資産被害額_01から49cm'
        ws.cell(row=1, column=39).value = '事業所被害額_在庫資産被害額_50から99cm'
        ws.cell(row=1, column=40).value = '事業所被害額_在庫資産被害額_100cm以上'
        ws.cell(row=1, column=41).value = '事業所被害額_在庫資産被害額_全壊'
        ws.cell(row=1, column=42).value = '事業所被害額_営業停止に伴う被害額_床下'
        ws.cell(row=1, column=43).value = '事業所被害額_営業停止に伴う被害額_01から49cm'
        ws.cell(row=1, column=44).value = '事業所被害額_営業停止に伴う被害額_50から99cm'
        ws.cell(row=1, column=45).value = '事業所被害額_営業停止に伴う被害額_100cm以上'
        ws.cell(row=1, column=46).value = '事業所被害額_営業停止に伴う被害額_全壊'
        ws.cell(row=1, column=47).value = '事業所被害額_営業停滞に伴う被害額_床下'
        ws.cell(row=1, column=48).value = '事業所被害額_営業停滞に伴う被害額_01から49cm'
        ws.cell(row=1, column=49).value = '事業所被害額_営業停滞に伴う被害額_50から99cm'
        ws.cell(row=1, column=50).value = '事業所被害額_営業停滞に伴う被害額_100cm以上'
        ws.cell(row=1, column=51).value = '事業所被害額_営業停滞に伴う被害額_全壊'
        ws.cell(row=1, column=52).value = '農漁家被害額_償却資産被害額_床下'
        ws.cell(row=1, column=53).value = '農漁家被害額_償却資産被害額_01から49cm'
        ws.cell(row=1, column=54).value = '農漁家被害額_償却資産被害額_50から99cm'
        ws.cell(row=1, column=55).value = '農漁家被害額_償却資産被害額_100cm以上'
        ws.cell(row=1, column=56).value = '農漁家被害額_償却資産被害額_全壊'
        ws.cell(row=1, column=57).value = '農漁家被害額_在庫資産被害額_床下'
        ws.cell(row=1, column=58).value = '農漁家被害額_在庫資産被害額_01から49cm'
        ws.cell(row=1, column=59).value = '農漁家被害額_在庫資産被害額_50から99cm'
        ws.cell(row=1, column=60).value = '農漁家被害額_在庫資産被害額_100cm以上'
        ws.cell(row=1, column=61).value = '農漁家被害額_在庫資産被害額_全壊'
        ws.cell(row=1, column=62).value = '事業所応急対策費_代替活動費_床下'
        ws.cell(row=1, column=63).value = '事業所応急対策費_代替活動費_01から49cm'
        ws.cell(row=1, column=64).value = '事業所応急対策費_代替活動費_50から99cm'
        ws.cell(row=1, column=65).value = '事業所応急対策費_代替活動費_100cm以上'
        ws.cell(row=1, column=66).value = '事業所応急対策費_代替活動費_半壊'
        ws.cell(row=1, column=67).value = '事業所応急対策費_代替活動費_全壊'
        
        if ippan_group_by_suikei_list:
            for i, ippan_group_by_suikei in enumerate(ippan_group_by_suikei_list):
                ws.cell(row=i+2, column=1).value = ippan_group_by_suikei.id
                ### ws.cell(row=i+2, column=2).value = ippan_group_by_suikei.ippan_id
                ### ws.cell(row=i+2, column=3).value = ippan_group_by_suikei.suigai_id
                ws.cell(row=i+2, column=2).value = ippan_group_by_suikei.house_summary_lv00
                ws.cell(row=i+2, column=3).value = ippan_group_by_suikei.house_summary_lv01_49
                ws.cell(row=i+2, column=4).value = ippan_group_by_suikei.house_summary_lv50_99
                ws.cell(row=i+2, column=5).value = ippan_group_by_suikei.house_summary_lv100
                ws.cell(row=i+2, column=6).value = ippan_group_by_suikei.house_summary_half
                ws.cell(row=i+2, column=7).value = ippan_group_by_suikei.house_summary_full
                ws.cell(row=i+2, column=8).value = ippan_group_by_suikei.household_summary_lv00
                ws.cell(row=i+2, column=9).value = ippan_group_by_suikei.household_summary_lv01_49
                ws.cell(row=i+2, column=10).value = ippan_group_by_suikei.household_summary_lv50_99
                ws.cell(row=i+2, column=11).value = ippan_group_by_suikei.household_summary_lv100
                ws.cell(row=i+2, column=12).value = ippan_group_by_suikei.household_summary_half
                ws.cell(row=i+2, column=13).value = ippan_group_by_suikei.household_summary_full
                ws.cell(row=i+2, column=14).value = ippan_group_by_suikei.car_summary_lv00
                ws.cell(row=i+2, column=15).value = ippan_group_by_suikei.car_summary_lv01_49
                ws.cell(row=i+2, column=16).value = ippan_group_by_suikei.car_summary_lv50_99
                ws.cell(row=i+2, column=17).value = ippan_group_by_suikei.car_summary_lv100
                ws.cell(row=i+2, column=18).value = ippan_group_by_suikei.car_summary_half
                ws.cell(row=i+2, column=19).value = ippan_group_by_suikei.car_summary_full
                ws.cell(row=i+2, column=20).value = ippan_group_by_suikei.house_alt_summary_lv00
                ws.cell(row=i+2, column=21).value = ippan_group_by_suikei.house_alt_summary_lv01_49
                ws.cell(row=i+2, column=22).value = ippan_group_by_suikei.house_alt_summary_lv50_99
                ws.cell(row=i+2, column=23).value = ippan_group_by_suikei.house_alt_summary_lv100
                ws.cell(row=i+2, column=24).value = ippan_group_by_suikei.house_alt_summary_half
                ws.cell(row=i+2, column=25).value = ippan_group_by_suikei.house_alt_summary_full
                ws.cell(row=i+2, column=26).value = ippan_group_by_suikei.house_clean_summary_lv00
                ws.cell(row=i+2, column=27).value = ippan_group_by_suikei.house_clean_summary_lv01_49
                ws.cell(row=i+2, column=28).value = ippan_group_by_suikei.house_clean_summary_lv50_99
                ws.cell(row=i+2, column=29).value = ippan_group_by_suikei.house_clean_summary_lv100
                ws.cell(row=i+2, column=30).value = ippan_group_by_suikei.house_clean_summary_half
                ws.cell(row=i+2, column=31).value = ippan_group_by_suikei.house_clean_summary_full
                ws.cell(row=i+2, column=32).value = ippan_group_by_suikei.office_dep_summary_lv00
                ws.cell(row=i+2, column=33).value = ippan_group_by_suikei.office_dep_summary_lv01_49
                ws.cell(row=i+2, column=34).value = ippan_group_by_suikei.office_dep_summary_lv50_99
                ws.cell(row=i+2, column=35).value = ippan_group_by_suikei.office_dep_summary_lv100
                ws.cell(row=i+2, column=36).value = ippan_group_by_suikei.office_dep_summary_full
                ws.cell(row=i+2, column=37).value = ippan_group_by_suikei.office_inv_summary_lv00
                ws.cell(row=i+2, column=38).value = ippan_group_by_suikei.office_inv_summary_lv01_49
                ws.cell(row=i+2, column=39).value = ippan_group_by_suikei.office_inv_summary_lv50_99
                ws.cell(row=i+2, column=40).value = ippan_group_by_suikei.office_inv_summary_lv100
                ws.cell(row=i+2, column=41).value = ippan_group_by_suikei.office_inv_summary_full
                ws.cell(row=i+2, column=42).value = ippan_group_by_suikei.office_sus_summary_lv00
                ws.cell(row=i+2, column=43).value = ippan_group_by_suikei.office_sus_summary_lv01_49
                ws.cell(row=i+2, column=44).value = ippan_group_by_suikei.office_sus_summary_lv50_99
                ws.cell(row=i+2, column=45).value = ippan_group_by_suikei.office_sus_summary_lv100
                ws.cell(row=i+2, column=46).value = ippan_group_by_suikei.office_sus_summary_full
                ws.cell(row=i+2, column=47).value = ippan_group_by_suikei.office_stg_summary_lv00
                ws.cell(row=i+2, column=48).value = ippan_group_by_suikei.office_stg_summary_lv01_49
                ws.cell(row=i+2, column=49).value = ippan_group_by_suikei.office_stg_summary_lv50_99
                ws.cell(row=i+2, column=50).value = ippan_group_by_suikei.office_stg_summary_lv100
                ws.cell(row=i+2, column=51).value = ippan_group_by_suikei.office_stg_summary_full
                ws.cell(row=i+2, column=52).value = ippan_group_by_suikei.farmer_fisher_dep_summary_lv00
                ws.cell(row=i+2, column=53).value = ippan_group_by_suikei.farmer_fisher_dep_summary_lv01_49
                ws.cell(row=i+2, column=54).value = ippan_group_by_suikei.farmer_fisher_dep_summary_lv50_99
                ws.cell(row=i+2, column=55).value = ippan_group_by_suikei.farmer_fisher_dep_summary_lv100
                ws.cell(row=i+2, column=56).value = ippan_group_by_suikei.farmer_fisher_dep_summary_full
                ws.cell(row=i+2, column=57).value = ippan_group_by_suikei.farmer_fisher_inv_summary_lv00
                ws.cell(row=i+2, column=58).value = ippan_group_by_suikei.farmer_fisher_inv_summary_lv01_49
                ws.cell(row=i+2, column=59).value = ippan_group_by_suikei.farmer_fisher_inv_summary_lv50_99
                ws.cell(row=i+2, column=60).value = ippan_group_by_suikei.farmer_fisher_inv_summary_lv100
                ws.cell(row=i+2, column=61).value = ippan_group_by_suikei.farmer_fisher_inv_summary_full
                ws.cell(row=i+2, column=62).value = ippan_group_by_suikei.office_alt_summary_lv00
                ws.cell(row=i+2, column=63).value = ippan_group_by_suikei.office_alt_summary_lv01_49
                ws.cell(row=i+2, column=64).value = ippan_group_by_suikei.office_alt_summary_lv50_99
                ws.cell(row=i+2, column=65).value = ippan_group_by_suikei.office_alt_summary_lv100
                ws.cell(row=i+2, column=66).value = ippan_group_by_suikei.office_alt_summary_half
                ws.cell(row=i+2, column=67).value = ippan_group_by_suikei.office_alt_summary_full
            
        wb.save(download_file_path)
        
        #######################################################################
        ### レスポンスセット処理(0030)
        ### テンプレートとコンテキストを設定して、レスポンスをブラウザに戻す。
        #######################################################################
        print_log('[DEBUG] P0200ExcelDownload.ippan_group_by_suikei_view()関数 STEP 4/4.', 'DEBUG')
        print_log('[INFO] P0200ExcelDownload.ippan_group_by_suikei_view()関数が正常終了しました。', 'INFO')
        response = HttpResponse(content=save_virtual_workbook(wb), content_type='application/vnd.ms-excel')
        response['Content-Disposition'] = 'attachment; filename="ippan_group_by_suikei.xlsx"'
        return response
        
    except:
        print_log('[ERROR] P0200ExcelDownload.ippan_group_by_suikei_view()関数 {}'.format(sys.exc_info()[0]), 'ERROR')
        print_log('[ERROR] P0200ExcelDownload.ippan_group_by_suikei_view()関数でエラーが発生しました。', 'ERROR')
        print_log('[ERROR] P0200ExcelDownload.ippan_group_by_suikei_view()関数が異常終了しました。', 'ERROR')
        return render(request, 'error.html')

###############################################################################
### 関数名：ippan_chosa_view(request, lock)
### 一般資産調査票（調査員用）
### ※複数EXCELファイル、複数EXCELシート対応版
### urlpattern：path('ippan_chosa/', views.ippan_chosa_view, name='ippan_chosa_view')
### template：P0900Action/download.html
###############################################################################
### @login_required(None, login_url='/P0100Login/')
def ippan_chosa_view(request, lock):
    try:            
        #######################################################################
        ### 引数チェック処理(0000)
        ### (1)ブラウザからのリクエストと引数をチェックする。
        ### (2)GETメソッドの場合、関数を抜ける。
        ### (3)POSTリクエストの市区町村数が0件の場合、関数を抜ける。
        #######################################################################
        reset_log()
        print_log('[INFO] P0200ExcelDownload.ippan_chosa_view()関数が開始しました。', 'INFO')
        print_log('[DEBUG] P0200ExcelDownload.ippan_chosa_view()関数 request = {}'.format(request.method), 'DEBUG')
        print_log('[DEBUG] P0200ExcelDownload.ippan_chosa_view()関数 lock = {}'.format(lock), 'DEBUG')
        print_log('[DEBUG] P0200ExcelDownload.ippan_chosa_view()関数 city_code_hidden= {}'.format(request.POST.get('city_code_hidden')), 'DEBUG')
        print_log('[DEBUG] P0200ExcelDownload.ippan_chosa_view()関数 STEP 1/6.', 'DEBUG')

        if request.method == 'GET':
            print_log('[ERROR] P0200ExcelDownload.ippan_chosa_view()関数でエラーが発生しました。', 'ERROR')
            return render(request, 'error.html')

        if request.POST.get('city_code_hidden') is None:
            print_log('[ERROR] P0200ExcelDownload.ippan_chosa_view()関数でエラーが発生しました。', 'ERROR')
            return render(request, 'error.html')
            
        city_code_request = [x.strip() for x in request.POST.get('city_code_hidden').split(',')][:-1]

        if city_code_request is None:
            print_log('[WARN] P0200ExcelDownload.ippan_chosa_view()関数で警告が発生しました。', 'WARN')
            return render(request, 'warn.html')

        if len(city_code_request) == 0:
            print_log('[WARN] P0200ExcelDownload.ippan_chosa_view()関数で警告が発生しました。', 'WARN')
            return render(request, 'warn.html')

        #######################################################################
        ### DBアクセス処理(0010)
        #######################################################################
        print_log('[DEBUG] P0200ExcelDownload.ippan_chosa_view()関数 STEP 2/6.', 'DEBUG')
        ken_code_request = []
        ken_name_request = []
        ### city_code_request = []
        city_name_request = []
        for city_code in city_code_request:
            city_list = CITY.objects.raw("""
                SELECT 
                    CT1.city_code AS city_code, 
                    CT1.city_name AS city_name, 
                    CT1.ken_code AS ken_code, 
                    KE1.ken_name AS ken_name 
                FROM CITY CT1 
                LEFT JOIN KEN KE1 ON CT1.ken_code=KE1.ken_code 
                WHERE 
                    CT1.city_code=%s LIMIT 1""", [city_code, ])
            
            ken_code_request.append([city.ken_code for city in city_list][0])
            ken_name_request.append([city.ken_name for city in city_list][0])
            ### city_code_request.append([city.city_code for city in city_list][0])
            city_name_request.append([city.city_name for city in city_list][0])
            
        print_log('[DEBUG] P0200ExcelDownload.ippan_chosa_view()関数 ken_code_request = {}'.format(ken_code_request), 'DEBUG')
        print_log('[DEBUG] P0200ExcelDownload.ippan_chosa_view()関数 ken_name_request = {}'.format(ken_name_request), 'DEBUG')
        print_log('[DEBUG] P0200ExcelDownload.ippan_chosa_view()関数 city_code_request = {}'.format(city_code_request), 'DEBUG')
        print_log('[DEBUG] P0200ExcelDownload.ippan_chosa_view()関数 city_name_request = {}'.format(city_name_request), 'DEBUG')

        #######################################################################
        ### 局所変数セット処理(0020)
        ### ハッシュコードを生成する。
        ### ※リクエスト固有のディレクトリ名を生成するため等に使用する。
        #######################################################################
        print_log('[DEBUG] P0200ExcelDownload.ippan_chosa_view()関数 STEP 3/6.', 'DEBUG')
        JST = timezone(timedelta(hours=9), 'JST')
        datetime_now_YmdHMS = datetime.now(JST).strftime('%Y%m%d%H%M%S')
        hash_code = hashlib.md5((str(datetime_now_YmdHMS)).encode()).hexdigest()[0:10]
        print_log('[DEBUG] P0200ExcelDownload.ippan_chosa_view()関数 hash_code = {}'.format(hash_code), 'DEBUG')

        #######################################################################
        ### 局所変数セット処理(0030)
        ### ダウンロードファイルパス、ダウンロードファイル名に値をセットする。
        #######################################################################
        print_log('[DEBUG] P0200ExcelDownload.ippan_chosa_view()関数 STEP 4/6.', 'DEBUG')
        download_file_path = []
        download_file_name = []
        for i, city_code in enumerate(city_code_request):
            download_file_path.append('static/' + str(hash_code) + '/ippan_chosa_' + str(city_code_request[i]) + '_' + str(ken_name_request[i]) + '_' + str(city_name_request[i]) + '.xlsx')
            download_file_name.append('ippan_chosa_' + str(city_code_request[i]) + '_' + str(ken_name_request[i]) + '_' + str(city_name_request[i]) + '.xlsx')
            
        new_dir_path = 'static/' + str(hash_code)
        os.makedirs(new_dir_path, exist_ok=True)

        #######################################################################
        ### DBアクセス処理(0040)
        #######################################################################
        print_log('[DEBUG] P0200ExcelDownload.ippan_chosa_view()関数 STEP 5/6.', 'DEBUG')
        connection_cursor = connection.cursor()
        try:
            connection_cursor.execute("""BEGIN""", [])
            
            for i, city_code in enumerate(city_code_request):
                connection_cursor.execute("""
                    INSERT INTO TRIGGER (
                        trigger_id, suigai_id, action_code, status_code, success_count, failure_count, 
                        published_at, consumed_at, deleted_at, integrity_ok, integrity_ng, ken_code, city_code, 
                        download_file_path, download_file_name, upload_file_path, upload_file_name 
                    ) VALUES (
                        (SELECT CASE WHEN (MAX(trigger_id+1)) IS NULL THEN CAST(0 AS INTEGER) ELSE CAST(MAX(trigger_id+1) AS INTEGER) END AS trigger_id FROM TRIGGER), -- trigger_id 
                        %s, -- suigai_id 
                        %s, -- action_code 
                        %s, -- status_code 
                        %s, -- success_count 
                        %s, -- failure_count 
                        CURRENT_TIMESTAMP, -- published_at 
                        %s, -- consumed_at 
                        %s, -- deleted_at 
                        %s, -- integrity_ok 
                        %s, -- integrity_ng 
                        %s, -- ken_code 
                        %s, -- city_code 
                        %s, -- download_file_path 
                        %s, -- download_file_name 
                        %s, -- upload_file_path 
                        %s  -- upload_file_name 
                    )""", [
                        None, ### suigai_id 
                        'A01',  ### action_code 
                        None, ### status_code 
                        None, ### success_count 
                        None, ### failure_count 
                        None, ### consumed_at
                        None, ### deleted_at 
                        None, ### integrity_ok 
                        None, ### integrity_ng 
                        ken_code_request[i],  ### ken_code 
                        city_code_request[i], ### city_code 
                        download_file_path[i], ### download_file_path 
                        download_file_name[i], ### download_file_name 
                        None, ### upload_file_path 
                        None, ### upload_file_name
                    ])
            ### transaction.commit()
            connection_cursor.execute("""COMMIT""", [])        
        except:
            print_log('[ERROR] P0200ExcelDownload.ippan_chosa_view()関数 {}'.format(sys.exc_info()[0]), 'ERROR')
            ### connection_cursor.rollback()
            connection_cursor.execute("""ROLLBACK""", [])
        finally:
            connection_cursor.close()

        #######################################################################
        ### レスポンスセット処理(0050)
        ### テンプレートとコンテキストを設定して、レスポンスをブラウザに戻す。
        ### See https://groups.google.com/g/django-users/c/SLw6SrIC8wI
        ### What you can do is prevent the browser from ever staying on a
        ### POST-generated page: the standard best practice is a technique called 
        ### "Post/Redirect/Get": see http://en.wikipedia.org/wiki/Post/Redirect/Get
        ### for a high-level description.
        #######################################################################
        print_log('[DEBUG] P0200ExcelDownload.ippan_chosa_view()関数 STEP 6/6.', 'DEBUG')
        return redirect('/P0200ExcelDownload/download/' + str(hash_code) + '/' + str(len(city_code_request)) + '/')
        
    except:
        print_log('[ERROR] P0200ExcelDownload.ippan_chosa_view()関数 {}'.format(sys.exc_info()[0]), 'ERROR')
        print_log('[ERROR] P0200ExcelDownload.ippan_chosa_view()関数でエラーが発生しました。', 'ERROR')
        print_log('[ERROR] P0200ExcelDownload.ippan_chosa_view()関数が異常終了しました。', 'ERROR')
        return render(request, 'error.html')

###############################################################################
### 関数名：download_view(request, hash_code)
### urlpattern：path('download/<slug:hash_code>', views.download_view, name='download_view')
### template：P0200ExcelDownload/download.html
###############################################################################
### @login_required(None, login_url='/P0100Login/')
def download_view(request, hash_code, count):
    try:
        #######################################################################
        ### 引数チェック処理(0000)
        ### ブラウザからのリクエストと引数をチェックする。
        #######################################################################
        ### reset_log()
        print_log('[INFO] P0200ExcelDownload.download_view()関数が開始しました。', 'INFO')
        print_log('[DEBUG] P0200ExcelDownload.download_view()関数 request = {}'.format(request.method), 'DEBUG')
        print_log('[DEBUG] P0200ExcelDownload.download_view()関数 hash_code = {}'.format(hash_code), 'DEBUG')
        print_log('[DEBUG] P0200ExcelDownload.download_view()関数 count = {}'.format(count), 'DEBUG')
        print_log('[DEBUG] P0200ExcelDownload.download_view()関数 STEP 1/3.', 'DEBUG')
        
        #######################################################################
        ### レスポンスセット処理(0000)
        ### テンプレートとコンテキストを設定して、レスポンスをブラウザに戻す。
        #######################################################################
        print_log('[DEBUG] P0200ExcelDownload.download_view()関数 STEP 2/3.', 'DEBUG')
        download_file_path = sorted(glob.glob('static/' + str(hash_code) + '/*.xlsx'), key=os.path.getmtime)
        
        #######################################################################
        ### レスポンスセット処理(0000)
        ### テンプレートとコンテキストを設定して、レスポンスをブラウザに戻す。
        #######################################################################
        print_log('[DEBUG] P0200ExcelDownload.download_view()関数 STEP 3/3.', 'DEBUG')
        template = loader.get_template('P0200ExcelDownload/download.html')
        context = {
            'hash_code': hash_code, 
            'download_file_path': download_file_path, 
            'count': count, 
        }
        print_log('[INFO] P0200ExcelDownload.download_view()関数が正常終了しました。', 'INFO')
        return HttpResponse(template.render(context, request))
        
    except:
        print_log('[ERROR] P0200ExcelDownload.download_view()関数 {}'.format(sys.exc_info()[0]), 'ERROR')
        print_log('[ERROR] P0200ExcelDownload.download_view()関数でエラーが発生しました。', 'ERROR')
        print_log('[ERROR] P0200ExcelDownload.download_view()関数が異常終了しました。', 'ERROR')
        return render(request, 'error.html')
    
###############################################################################
### 関数名：ippan_city_view(request, lock)
### 一般資産調査票（市区町村担当者用）
### urlpattern：path('ippan_city/', views.ippan_city_view, name='ippan_city_view')
### template：
###############################################################################
### @login_required(None, login_url='/P0100Login/')
def ippan_city_view(request, lock):
    try:
        #######################################################################
        ### 引数チェック処理(0000)
        ### (1)ブラウザからのリクエストと引数をチェックする。
        ### (2)GETメソッドの場合、関数を抜ける。
        ### (3)POSTリクエストの市区町村数が0件の場合、関数を抜ける。
        #######################################################################
        reset_log()
        print_log('[INFO] P0200ExcelDownload.ippan_city_view()関数が開始しました。', 'INFO')
        print_log('[DEBUG] P0200ExcelDownload.ippan_city_view()関数 request = {}'.format(request.method), 'DEBUG')
        print_log('[DEBUG] P0200ExcelDownload.ippan_city_view()関数 lock = {}'.format(lock), 'DEBUG')
        print_log('[DEBUG] P0200ExcelDownload.ippan_city_view()関数 city_code_hidden= {}'.format(request.POST.get('city_code_hidden')), 'DEBUG')
        print_log('[DEBUG] P0200ExcelDownload.ippan_city_view()関数 STEP 1/6.', 'DEBUG')

        if request.method == 'GET':
            print_log('[ERROR] P0200ExcelDownload.ippan_city_view()関数でエラーが発生しました。', 'ERROR')
            return render(request, 'error.html')

        if request.POST.get('city_code_hidden') is None:
            print_log('[ERROR] P0200ExcelDownload.ippan_city_view()関数でエラーが発生しました。', 'ERROR')
            return render(request, 'error.html')

        city_code_request = [x.strip() for x in request.POST.get('city_code_hidden').split(',')][:-1]

        if city_code_request is None:
            print_log('[WARN] P0200ExcelDownload.ippan_city_view()関数で警告が発生しました。', 'WARN')
            return render(request, 'warning.html')

        if len(city_code_request) == 0:
            print_log('[WARN] P0200ExcelDownload.ippan_city_view()関数で警告が発生しました。', 'WARN')
            return render(request, 'warning.html')

        #######################################################################
        ### DBアクセス処理(0010)
        #######################################################################
        print_log('[DEBUG] P0200ExcelDownload.ippan_city_view()関数 STEP 2/6.', 'DEBUG')
        ken_code_request = []
        ken_name_request = []
        ### city_code_request = []
        city_name_request = []
        
        for city_code in city_code_request:
            city_list = CITY.objects.raw("""
                SELECT 
                    CT1.city_code AS city_code, 
                    CT1.city_name AS city_name, 
                    CT1.ken_code AS ken_code, 
                    KE1.ken_name AS ken_name 
                FROM CITY CT1 
                LEFT JOIN KEN KE1 ON CT1.ken_code=KE1.ken_code 
                WHERE 
                    CT1.city_code=%s LIMIT 1""", [city_code, ])
            
            ken_code_request.append([city.ken_code for city in city_list][0])
            ken_name_request.append([city.ken_name for city in city_list][0])
            ### city_code_request.append([city.city_code for city in city_list][0])
            city_name_request.append([city.city_name for city in city_list][0])
            
        print_log('[DEBUG] P0200ExcelDownload.ippan_city_view()関数 ken_code_request = {}'.format(ken_code_request), 'DEBUG')
        print_log('[DEBUG] P0200ExcelDownload.ippan_city_view()関数 ken_name_request = {}'.format(ken_name_request), 'DEBUG')
        print_log('[DEBUG] P0200ExcelDownload.ippan_city_view()関数 city_code_request = {}'.format(city_code_request), 'DEBUG')
        print_log('[DEBUG] P0200ExcelDownload.ippan_city_view()関数 city_name_request = {}'.format(city_name_request), 'DEBUG')

        #######################################################################
        ### 局所変数セット処理(0020)
        ### ハッシュコードを生成する。
        ### ※リクエスト固有のディレクトリ名を生成するため等に使用する。
        #######################################################################
        print_log('[DEBUG] P0200ExcelDownload.ippan_city_view()関数 STEP 3/6.', 'DEBUG')
        JST = timezone(timedelta(hours=9), 'JST')
        datetime_now_YmdHMS = datetime.now(JST).strftime('%Y%m%d%H%M%S')
        hash_code = hashlib.md5((str(datetime_now_YmdHMS)).encode()).hexdigest()[0:10]
        print_log('[DEBUG] P0200ExcelDownload.ippan_city_view()関数 hash_code = {}'.format(hash_code), 'DEBUG')

        #######################################################################
        ### 局所変数セット処理(0030)
        ### ダウンロードファイルパス、ダウンロードファイル名に値をセットする。
        #######################################################################
        print_log('[DEBUG] P0200ExcelDownload.ippan_city_view()関数 STEP 4/6.', 'DEBUG')
        download_file_path = []
        download_file_name = []
        for i, city_code in enumerate(city_code_request):
            download_file_path.append('static/' + str(hash_code) + '/ippan_city_' + str(city_code_request[i]) + '_' + str(ken_name_request[i]) + '_' + str(city_name_request[i]) + '.xlsx')
            download_file_name.append('ippan_city_' + str(city_code_request[i]) + '_' + str(ken_name_request[i]) + '_' + str(city_name_request[i]) + '.xlsx')
            
        new_dir_path = 'static/' + str(hash_code)
        os.makedirs(new_dir_path, exist_ok=True)

        #######################################################################
        ### DBアクセス処理(0040)
        #######################################################################
        print_log('[DEBUG] P0200ExcelDownload.ippan_city_view()関数 STEP 5/6.', 'DEBUG')
        connection_cursor = connection.cursor()
        try:
            connection_cursor.execute("""BEGIN""", [])
            
            for i, city_code in enumerate(city_code_request):
                connection_cursor.execute("""
                    INSERT INTO TRIGGER (
                        trigger_id, suigai_id, action_code, status_code, success_count, failure_count, 
                        published_at, consumed_at, deleted_at, integrity_ok, integrity_ng, ken_code, city_code, 
                        download_file_path, download_file_name, upload_file_path, upload_file_name 
                    ) VALUES (
                        (SELECT CASE WHEN (MAX(trigger_id+1)) IS NULL THEN CAST(0 AS INTEGER) ELSE CAST(MAX(trigger_id+1) AS INTEGER) END AS trigger_id FROM TRIGGER), -- trigger_id 
                        %s, -- suigai_id 
                        %s, -- action_code 
                        %s, -- status_code 
                        %s, -- success_count 
                        %s, -- failure_count 
                        CURRENT_TIMESTAMP, -- published_at 
                        %s, -- consumed_at 
                        %s, -- deleted_at 
                        %s, -- integrity_ok 
                        %s, -- integrity_ng 
                        %s, -- ken_code 
                        %s, -- city_code 
                        %s, -- download_file_path 
                        %s, -- download_file_name 
                        %s, -- upload_file_path 
                        %s  -- upload_file_name 
                    )""", [
                        None, ### suigai_id 
                        'C01',  ### action_code 
                        None, ### status_code 
                        None, ### success_count 
                        None, ### failure_count 
                        None, ### consumed_at
                        None, ### deleted_at 
                        None, ### integrity_ok 
                        None, ### integrity_ng 
                        ken_code_request[i],  ### ken_code 
                        city_code_request[i], ### city_code 
                        download_file_path[i], ### download_file_path 
                        download_file_name[i], ### download_file_name 
                        None, ### upload_file_path 
                        None, ### upload_file_name
                    ])
            ### transaction.commit()
            connection_cursor.execute("""COMMIT""", [])        
        except:
            print_log('[ERROR] P0200ExcelDownload.ippan_city_view()関数 {}'.format(sys.exc_info()[0]), 'ERROR')
            ### connection_cursor.rollback()
            connection_cursor.execute("""ROLLBACK""", [])
        finally:
            connection_cursor.close()

        #######################################################################
        ### レスポンスセット処理(0050)
        ### テンプレートとコンテキストを設定して、レスポンスをブラウザに戻す。
        ### See https://groups.google.com/g/django-users/c/SLw6SrIC8wI
        ### What you can do is prevent the browser from ever staying on a
        ### POST-generated page: the standard best practice is a technique called 
        ### "Post/Redirect/Get": see http://en.wikipedia.org/wiki/Post/Redirect/Get
        ### for a high-level description.
        #######################################################################
        print_log('[DEBUG] P0200ExcelDownload.ippan_city_view()関数 STEP 6/6.', 'DEBUG')
        return redirect('/P0200ExcelDownload/download/' + str(hash_code) + '/' + str(len(city_code_request)) + '/')
        
    except:
        print_log('[ERROR] P0200ExcelDownload.ippan_city_view()関数 {}'.format(sys.exc_info()[0]), 'ERROR')
        print_log('[ERROR] P0200ExcelDownload.ippan_city_view()関数でエラーが発生しました。', 'ERROR')
        print_log('[ERROR] P0200ExcelDownload.ippan_city_view()関数が異常終了しました。', 'ERROR')
        return render(request, 'error.html')

###############################################################################
### 関数名：ippan_ken_view(request, lock)
### 一般資産調査票（都道府県担当者用）
### ※複数EXCELファイル、複数EXCELシート対応版
### urlpattern：path('ippan_ken/', views.ippan_ken_view, name='ippan_ken_view')
### template：P900Action/download.html
###############################################################################
### @login_required(None, login_url='/P0100Login/')
def ippan_ken_view(request, lock):
    try:
        #######################################################################
        ### 引数チェック処理(0000)
        ### (1)ブラウザからのリクエストと引数をチェックする。
        ### (2)GETメソッドの場合、関数を抜ける。
        ### (3)POSTリクエストの市区町村数が0件の場合、関数を抜ける。
        #######################################################################
        reset_log()
        print_log('[INFO] P0200ExcelDownload.ippan_ken_view()関数が開始しました。', 'INFO')
        print_log('[DEBUG] P0200ExcelDownload.ippan_ken_view()関数 request = {}'.format(request.method), 'DEBUG')
        print_log('[DEBUG] P0200ExcelDownload.ippan_ken_view()関数 lock = {}'.format(lock), 'DEBUG')
        print_log('[DEBUG] P0200ExcelDownload.ippan_ken_view()関数 city_code_hidden= {}'.format(request.POST.get('city_code_hidden')), 'DEBUG')
        print_log('[DEBUG] P0200ExcelDownload.ippan_ken_view()関数 STEP 1/6.', 'DEBUG')

        if request.method == 'GET':
            print_log('[ERROR] P0200ExcelDownload.ippan_ken_view()関数でエラーが発生しました。', 'ERROR')
            return render(request, 'error.html')

        if request.POST.get('city_code_hidden') is None:
            print_log('[ERROR] P0200ExcelDownload.ippan_ken_view()関数でエラーが発生しました。', 'ERROR')
            return render(request, 'error.html')

        city_code_request = [x.strip() for x in request.POST.get('city_code_hidden').split(',')][:-1]

        if city_code_request is None:
            print_log('[WARN] P0200ExcelDownload.ippan_ken_view()関数で警告が発生しました。', 'WARN')
            return render(request, 'warning.html')

        if len(city_code_request) == 0:
            print_log('[WARN] P0200ExcelDownload.ippan_ken_view()関数で警告が発生しました。', 'WARN')
            return render(request, 'warning.html')

        #######################################################################
        ### DBアクセス処理(0010)
        #######################################################################
        print_log('[DEBUG] P0200ExcelDownload.ippan_ken_view()関数 STEP 2/6.', 'DEBUG')
        ken_code_request = []
        ken_name_request = []
        ### city_code_request = []
        city_name_request = []
        for city_code in city_code_request:
            city_list = CITY.objects.raw("""
                SELECT 
                    CT1.city_code AS city_code, 
                    CT1.city_name AS city_name, 
                    CT1.ken_code AS ken_code, 
                    KE1.ken_name AS ken_name 
                FROM CITY CT1 
                LEFT JOIN KEN KE1 ON CT1.ken_code=KE1.ken_code 
                WHERE 
                    CT1.city_code=%s LIMIT 1""", [city_code, ])
            
            ken_code_request.append([city.ken_code for city in city_list][0])
            ken_name_request.append([city.ken_name for city in city_list][0])
            ### city_code_request.append([city.city_code for city in city_list][0])
            city_name_request.append([city.city_name for city in city_list][0])
            
        print_log('[DEBUG] P0200ExcelDownload.ippan_ken_view()関数 ken_code_request = {}'.format(ken_code_request), 'DEBUG')
        print_log('[DEBUG] P0200ExcelDownload.ippan_ken_view()関数 ken_name_request = {}'.format(ken_name_request), 'DEBUG')
        print_log('[DEBUG] P0200ExcelDownload.ippan_ken_view()関数 city_code_request = {}'.format(city_code_request), 'DEBUG')
        print_log('[DEBUG] P0200ExcelDownload.ippan_ken_view()関数 city_name_request = {}'.format(city_name_request), 'DEBUG')

        #######################################################################
        ### 局所変数セット処理(0020)
        ### ハッシュコードを生成する。
        ### ※リクエスト固有のディレクトリ名を生成するため等に使用する。
        #######################################################################
        print_log('[DEBUG] P0200ExcelDownload.ippan_ken_view()関数 STEP 3/6.', 'DEBUG')
        JST = timezone(timedelta(hours=9), 'JST')
        datetime_now_YmdHMS = datetime.now(JST).strftime('%Y%m%d%H%M%S')
        hash_code = hashlib.md5((str(datetime_now_YmdHMS)).encode()).hexdigest()[0:10]
        print_log('[DEBUG] P0200ExcelDownload.ippan_ken_view()関数 hash_code = {}'.format(hash_code), 'DEBUG')

        #######################################################################
        ### 局所変数セット処理(0030)
        ### ダウンロードファイルパス、ダウンロードファイル名に値をセットする。
        #######################################################################
        print_log('[DEBUG] P0200ExcelDownload.ippan_ken_view()関数 STEP 4/6.', 'DEBUG')
        download_file_path = []
        download_file_name = []
        for i, ken_code in enumerate(ken_code_request):
            download_file_path.append('static/' + str(hash_code) + '/ippan_ken_' + str(ken_code_request[i]) + '_' + str(ken_name_request[i]) + '.xlsx')
            download_file_name.append('ippan_ken_' + str(ken_code_request[i]) + '_' + str(ken_name_request[i]) + '.xlsx')
            
        new_dir_path = 'static/' + str(hash_code)
        os.makedirs(new_dir_path, exist_ok=True)
        
        #######################################################################
        ### DBアクセス処理(0040)
        #######################################################################
        print_log('[DEBUG] P0200ExcelDownload.ippan_ken_view()関数 STEP 5/6.', 'DEBUG')
        connection_cursor = connection.cursor()
        try:
            connection_cursor.execute("""BEGIN""", [])
            
            for i, ken_code in enumerate(ken_code_request):
                connection_cursor.execute("""
                    INSERT INTO TRIGGER (
                        trigger_id, suigai_id, action_code, status_code, success_count, failure_count, 
                        published_at, consumed_at, deleted_at, integrity_ok, integrity_ng, ken_code, city_code, 
                        download_file_path, download_file_name, upload_file_path, upload_file_name 
                    ) VALUES (
                        (SELECT CASE WHEN (MAX(trigger_id+1)) IS NULL THEN CAST(0 AS INTEGER) ELSE CAST(MAX(trigger_id+1) AS INTEGER) END AS trigger_id FROM TRIGGER), -- trigger_id 
                        %s, -- suigai_id 
                        %s, -- action_code 
                        %s, -- status_code 
                        %s, -- success_count 
                        %s, -- failure_count 
                        CURRENT_TIMESTAMP, -- published_at 
                        %s, -- consumed_at 
                        %s, -- deleted_at 
                        %s, -- integrity_ok 
                        %s, -- integrity_ng 
                        %s, -- ken_code 
                        %s, -- city_code 
                        %s, -- download_file_path 
                        %s, -- download_file_name 
                        %s, -- upload_file_path 
                        %s  -- upload_file_name 
                    )""", [
                        None, ### suigai_id 
                        'D01',  ### action_code 
                        None, ### status_code 
                        None, ### success_count 
                        None, ### failure_count 
                        None, ### consumed_at
                        None, ### deleted_at 
                        None, ### integrity_ok 
                        None, ### integrity_ng 
                        ken_code_request[i],  ### ken_code 
                        None, ### city_code 
                        download_file_path[i], ### download_file_path 
                        download_file_name[i], ### download_file_name 
                        None, ### upload_file_path 
                        None, ### upload_file_name
                    ])
            ### transaction.commit()
            connection_cursor.execute("""COMMIT""", [])        
        except:
            print_log('[ERROR] P0200ExcelDownload.ippan_ken_view()関数 {}'.format(sys.exc_info()[0]), 'ERROR')
            ### connection_cursor.rollback()
            connection_cursor.execute("""ROLLBACK""", [])
        finally:
            connection_cursor.close()

        #######################################################################
        ### レスポンスセット処理(0050)
        ### テンプレートとコンテキストを設定して、レスポンスをブラウザに戻す。
        ### See https://groups.google.com/g/django-users/c/SLw6SrIC8wI
        ### What you can do is prevent the browser from ever staying on a
        ### POST-generated page: the standard best practice is a technique called 
        ### "Post/Redirect/Get": see http://en.wikipedia.org/wiki/Post/Redirect/Get
        ### for a high-level description.
        #######################################################################
        print_log('[DEBUG] P0200ExcelDownload.ippan_ken_view()関数 STEP 6/6.', 'DEBUG')
        return redirect('/P0200ExcelDownload/download/' + str(hash_code) + '/' + str(len(ken_code_request)) + '/')
        
    except:
        print_log('[ERROR] P0200ExcelDownload.ippan_ken_view()関数 {}'.format(sys.exc_info()[0]), 'ERROR')
        print_log('[ERROR] P0200ExcelDownload.ippan_ken_view()関数でエラーが発生しました。', 'ERROR')
        print_log('[ERROR] P0200ExcelDownload.ippan_ken_view()関数が異常終了しました。', 'ERROR')
        return render(request, 'error.html')
