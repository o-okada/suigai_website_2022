#!/usr/bin/env python3
# -*- coding: utf-8 -*-
###############################################################################
### ファイル名：P0900Action/management/commands/action_N01_download_ippan_chosa.py
### A01：ダウンロード
###############################################################################

###############################################################################
### 処理名：インポート処理
###############################################################################
from django.core.management.base import BaseCommand

import sys
from django.contrib.auth.decorators import login_required
from django.db import connection
from django.db import transaction
from django.http import Http404
from django.http import HttpResponse
from django.http import HttpResponseRedirect
from django.shortcuts import render
from django.template import loader
from django.views import generic
from django.views.generic import FormView
from django.views.generic.base import TemplateView

import openpyxl
from openpyxl.comments import Comment
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import PatternFill
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.writer.excel import save_virtual_workbook

from decimal import Decimal

from P0000Common.models import BUILDING                ### 1000: 建物区分
from P0000Common.models import KEN                     ### 1010: 都道府県
from P0000Common.models import CITY                    ### 1020: 市区町村
from P0000Common.models import KASEN_KAIGAN            ### 1030: 水害発生地点工種（河川海岸区分）
from P0000Common.models import SUIKEI                  ### 1040: 水系（水系・沿岸）
from P0000Common.models import SUIKEI_TYPE             ### 1050: 水系種別（水系・沿岸種別）
from P0000Common.models import KASEN                   ### 1060: 河川（河川・海岸）
from P0000Common.models import KASEN_TYPE              ### 1070: 河川種別（河川・海岸種別）
from P0000Common.models import CAUSE                   ### 1080: 水害原因
from P0000Common.models import UNDERGROUND             ### 1090: 地上地下区分
from P0000Common.models import USAGE                   ### 1100: 地下空間の利用形態
from P0000Common.models import FLOOD_SEDIMENT          ### 1110: 浸水土砂区分
from P0000Common.models import GRADIENT                ### 1120: 地盤勾配区分
from P0000Common.models import INDUSTRY                ### 1130: 産業分類

from P0000Common.models import HOUSE_ASSET             ### 2000: 家屋評価額
from P0000Common.models import HOUSE_RATE              ### 2010: 家屋被害率
from P0000Common.models import HOUSE_ALT               ### 2020: 家庭応急対策費_代替活動費
from P0000Common.models import HOUSE_CLEAN             ### 2030: 家庭応急対策費_清掃日数

from P0000Common.models import HOUSEHOLD_ASSET         ### 3000: 家庭用品自動車以外所有額
from P0000Common.models import HOUSEHOLD_RATE          ### 3010: 家庭用品自動車以外被害率

from P0000Common.models import CAR_ASSET               ### 4000: 家庭用品自動車所有額
from P0000Common.models import CAR_RATE                ### 4010: 家庭用品自動車被害率

from P0000Common.models import OFFICE_ASSET            ### 5000: 事業所資産額
from P0000Common.models import OFFICE_RATE             ### 5010: 事業所被害率
from P0000Common.models import OFFICE_SUSPEND          ### 5020: 事業所営業停止日数
from P0000Common.models import OFFICE_STAGNATE         ### 5030: 事業所営業停滞日数
from P0000Common.models import OFFICE_ALT              ### 5040: 事業所応急対策費_代替活動費

from P0000Common.models import FARMER_FISHER_ASSET     ### 6000: 農漁家資産額
from P0000Common.models import FARMER_FISHER_RATE      ### 6010: 農漁家被害率

from P0000Common.models import AREA                    ### 7000: 入力データ_水害区域
from P0000Common.models import WEATHER                 ### 7010: 入力データ_異常気象
from P0000Common.models import SUIGAI                  ### 7020: 入力データ_ヘッダ部分
from P0000Common.models import IPPAN                   ### 7030: 入力データ_一覧表部分
from P0000Common.models import IPPAN_VIEW              ### 7040: ビューデータ_一覧表部分

from P0000Common.models import IPPAN_SUMMARY           ### 8000: 集計データ

from P0000Common.models import ACTION                  ### 10000: アクション
from P0000Common.models import STATUS                  ### 10010: 状態
from P0000Common.models import TRIGGER                 ### 10020: トリガーメッセージ
from P0000Common.models import APPROVAL                ### 10030: 承認メッセージ
from P0000Common.models import FEEDBACK                ### 10040: フィードバックメッセージ

from P0000Common.common import get_debug_log
from P0000Common.common import get_error_log
from P0000Common.common import get_info_log
from P0000Common.common import get_warn_log
from P0000Common.common import print_log
from P0000Common.common import reset_log

VLOOK_VALUE = [
    'B', 'G', 'L', 'Q', 'V', 'AA', 'AF', 'AK', 'AP', 'AU', 
    'AZ', 'BE', 'BJ', 'BO', 'BT', 'BY', 'CD', 'CI', 'CN', 'CS', 
    'CX', 'DC', 'DH', 'DM', 'DR', 'DW', 'EB', 'EG', 'EL', 'EQ', 
    'EV', 'FA', 'FF', 'FK', 'FP', 'FU', 'FZ', 'GE', 'GJ', 'GO', 
    'GT', 'GY', 'HD', 'HI', 'HN', 'HS', 'HX', 'IC', 'IH', 'IM', 
    'IR', 'IW', 'JB', 'JG', 'JL', 'JQ', 'JV', 'KA', 'KF', 'KK', 
    'KP', 'KU', 'KZ', 'LE', 'LJ', 'LO', 'LT', 'LY', 'MD', 'MI', 
    'MN', 'MS', 'MX', 'NC', 'NH', 'NM', 'NR', 'NW', 'OB', 'OG', 
    'OL', 'OQ', 'OV', 'PA', 'PF', 'PK', 'PP', 'PU', 'PZ', 'QE', 
    'QJ', 'QO', 'QT', 'QY', 'RD', 'RI', 'RN', 'RS', 'RX', 'SC', 
    'SH', 'SM', 'SR', 'SW', 'TB', 'TG', 'TL', 'TQ', 'TV', 'UA', 
    'UF', 'UK', 'UP', 'UU', 'UZ', 'VE', 'VJ', 'VO', 'VT', 'VY', 
    'WD', 'WI', 'WN', 'WS', 'WX', 'XC', 'XH', 'XM', 'XR', 'XW', 
    'YB', 'YG', 'YL', 'YQ', 'YV', 'ZA', 'ZF', 'ZK', 'ZP', 'ZU', 
    'ZZ', 'AAE', 'AAJ', 'AAO', 'AAT', 'AAY', 'ABD', 'ABI', 'ABN', 'ABS', 
    'ABX', 'ACC', 'ACH', 'ACM', 'ACR', 'ACW', 'ADB', 'ADG', 'ADL', 'ADQ', 
    'ADV', 'AEA', 'AEF', 'AEK', 'AEP', 'AEU', 'AEZ', 'AFE', 'AFJ', 'AFO', 
    'AFT', 'AFY', 'AGD', 'AGI', 'AGN', 'AGS', 'AGX', 'AHC', 'AHH', 'AHM', 
    'AHR', 'AHW', 'AIB', 'AIG', 'AIL', 'AIQ', 'AIV', 'AJA', 'AJF', 'AJK', 
    'AJP', 'AJU', 'AJZ', 'AKE', 'AKJ', 'AKO', 'AKT', 'AKY', 'ALD', 'ALI', 
    'ALN', 'ALS', 'ALX', 'AMC', 'AMH', 'AMM', 'AMR', 'AMW', 'ANB', 'ANG', 
    'ANL', 'ANQ', 'ANV', 'AOA', 'AOF', 'AOK', 'AOP', 'AOU', 'AOZ', 'APE', 
    'APJ', 'APO', 'APT', 'APY', 'AQD', 'AQI', 'AQN', 'AQS', 'AQX', 'ARC', 
    'ARH', 'ARM', 'ARR', 'ARW', 'ASB', 'ASG', 'ASL', 'ASQ', 'ASV', 'ATA', 
    'ATF', 'ATK', 'ATP', 'ATU', 'ATZ', 'AUE', 'AUJ', 'AUO', 'AUT', 'AUY', 
    'AVD', 'AVI', 'AVN', 'AVS', 'AVX', 'AWC', 'AWH', 'AWM', 'AWR', 'AWW', 
    'AXB', 'AXG', 'AXL', 'AXQ', 'AXV', 'AYA', 'AYF', 'AYK', 'AYP', 'AYU', 
    'AYZ', 'AZE', 'AZJ', 'AZO', 'AZT', 'AZY'
    ]

###############################################################################
### 関数名：create_ippan_chosa_workbook()
### 一般資産調査票（調査員用）
### ※複数EXCELファイル、複数EXCELシート対応版
###############################################################################
def create_ippan_chosa_workbook(suigai_count, suigai_id_list, suigai_name_list, ken_code, ken_name, city_code, city_name, 
    VLOOK_VALUE, building_list, ken_list, kasen_kaigan_list, suikei_list, suikei_type_list, kasen_type_list, cause_list, 
    underground_list, usage_list, flood_sediment_list, gradient_list, industry_list, area_list, weather_list, suigai_list):
    try:
        #######################################################################
        ### 引数チェック処理(0000)
        #######################################################################
        reset_log()
        print_log('[INFO] P0900Action.action_N01_download_ippan_chosa.create_ippan_chosa_workbook()関数が開始しました。', 'INFO')
        print_log('[DEBUG] P0900Action.action_N01_download_ippan_chosa.create_ippan_chosa_workbook()関数 suigai_count = {}'.format(suigai_count), 'DEBUG')
        print_log('[DEBUG] P0900Action.action_N01_download_ippan_chosa.create_ippan_chosa_workbook()関数 suigai_id_list = {}'.format(suigai_id_list), 'DEBUG')
        print_log('[DEBUG] P0900Action.action_N01_download_ippan_chosa.create_ippan_chosa_workbook()関数 suigai_name_list = {}'.format(suigai_name_list), 'DEBUG')
        print_log('[DEBUG] P0900Action.action_N01_download_ippan_chosa.create_ippan_chosa_workbook()関数 ken_code = {}'.format(ken_code), 'DEBUG')
        print_log('[DEBUG] P0900Action.action_N01_download_ippan_chosa.create_ippan_chosa_workbook()関数 ken_name = {}'.format(ken_name), 'DEBUG')
        print_log('[DEBUG] P0900Action.action_N01_download_ippan_chosa.create_ippan_chosa_workbook()関数 city_code = {}'.format(city_code), 'DEBUG')
        print_log('[DEBUG] P0900Action.action_N01_download_ippan_chosa.create_ippan_chosa_workbook()関数 city_name = {}'.format(city_name), 'DEBUG')
        print_log('[DEBUG] P0900Action.action_N01_download_ippan_chosa.create_ippan_chosa_workbook()関数 VLOOK_VALUE = {}'.format(VLOOK_VALUE), 'DEBUG')
        print_log('[DEBUG] P0900Action.action_N01_download_ippan_chosa.create_ippan_chosa_workbook()関数 building_list = {}'.format(building_list), 'DEBUG')
        print_log('[DEBUG] P0900Action.action_N01_download_ippan_chosa.create_ippan_chosa_workbook()関数 ken_list = {}'.format(ken_list), 'DEBUG')
        print_log('[DEBUG] P0900Action.action_N01_download_ippan_chosa.create_ippan_chosa_workbook()関数 kasen_kaigan_list = {}'.format(kasen_kaigan_list), 'DEBUG')
        print_log('[DEBUG] P0900Action.action_N01_download_ippan_chosa.create_ippan_chosa_workbook()関数 suikei_list = {}'.format(suikei_list), 'DEBUG')
        print_log('[DEBUG] P0900Action.action_N01_download_ippan_chosa.create_ippan_chosa_workbook()関数 suikei_type_list = {}'.format(suikei_type_list), 'DEBUG')
        print_log('[DEBUG] P0900Action.action_N01_download_ippan_chosa.create_ippan_chosa_workbook()関数 kasen_type_list = {}'.format(kasen_type_list), 'DEBUG')
        print_log('[DEBUG] P0900Action.action_N01_download_ippan_chosa.create_ippan_chosa_workbook()関数 cause_list = {}'.format(cause_list), 'DEBUG')
        print_log('[DEBUG] P0900Action.action_N01_download_ippan_chosa.create_ippan_chosa_workbook()関数 underground_list = {}'.format(underground_list), 'DEBUG')
        print_log('[DEBUG] P0900Action.action_N01_download_ippan_chosa.create_ippan_chosa_workbook()関数 usage_list = {}'.format(usage_list), 'DEBUG')
        print_log('[DEBUG] P0900Action.action_N01_download_ippan_chosa.create_ippan_chosa_workbook()関数 flood_sediment_list = {}'.format(flood_sediment_list), 'DEBUG')
        print_log('[DEBUG] P0900Action.action_N01_download_ippan_chosa.create_ippan_chosa_workbook()関数 gradient_list = {}'.format(gradient_list), 'DEBUG')
        print_log('[DEBUG] P0900Action.action_N01_download_ippan_chosa.create_ippan_chosa_workbook()関数 industry_list = {}'.format(industry_list), 'DEBUG')
        print_log('[DEBUG] P0900Action.action_N01_download_ippan_chosa.create_ippan_chosa_workbook()関数 area_list = {}'.format(area_list), 'DEBUG')
        print_log('[DEBUG] P0900Action.action_N01_download_ippan_chosa.create_ippan_chosa_workbook()関数 weather_list = {}'.format(weather_list), 'DEBUG')
        print_log('[DEBUG] P0900Action.action_N01_download_ippan_chosa.create_ippan_chosa_workbook()関数 suigai_list = {}'.format(suigai_list), 'DEBUG')
        print_log('[DEBUG] P0900Action.action_N01_download_ippan_chosa.create_ippan_chosa_workbook()関数 STEP 1/10.', 'DEBUG')
        
        #######################################################################
        ### EXCEL入出力処理(0010)
        ### (1)ダウンロード用ファイルパスをセットする。
        ### (2)ワークシートを局所変数にセットする。
        #######################################################################
        print_log('[DEBUG] P0900Action.action_N01_download_ippan_chosa.create_ippan_chosa_workbook()関数 STEP 2/10.', 'DEBUG')
        template_file_path = 'static/template_ippan_chosa.xlsx'
        download_file_path = 'static/ippan_chosa_' + str(city_code) + '_' + str(ken_name) + '_' + str(city_name) + '.xlsx'
        
        wb = openpyxl.load_workbook(template_file_path, keep_vba=False)
        
        ws_building = wb["BUILDING"]
        ws_ken = wb["KEN"]
        ws_city = wb["CITY"]
        ws_kasen_kaigan = wb["KASEN_KAIGAN"]
        ws_suikei = wb["SUIKEI"]
        ws_suikei_type = wb["SUIKEI_TYPE"]
        ws_kasen = wb["KASEN"]
        ws_kasen_type = wb["KASEN_TYPE"]
        ws_cause = wb["CAUSE"]
        ws_underground = wb["UNDERGROUND"]
        ws_usage = wb["USAGE"]
        ws_flood_sediment = wb["FLOOD_SEDIMENT"]
        ws_gradient = wb["GRADIENT"]
        ws_industry = wb["INDUSTRY"]
        ws_area = wb["AREA"]
        ws_weather = wb["WEATHER"]
        ws_suigai = wb["SUIGAI"]

        ws_city_vlook = wb["CITY_VLOOK"]
        ws_kasen_vlook = wb["KASEN_VLOOK"]
        ws_suikei_type_vlook = wb["SUIKEI_TYPE_VLOOK"]
        ws_kasen_type_vlook = wb["KASEN_TYPE_VLOOK"]
        
        ws_ippan = []
        ws_ippan.append(wb["IPPAN"])
        for i in range(suigai_count + 10):
            ws_ippan.append(wb.copy_worksheet(wb["IPPAN"]))
            ws_ippan[i+1].title = 'IPPAN' + str(i+1)
            
        #######################################################################
        ### EXCEL入出力処理(0020)
        ### 各EXCELシートで枠線を表示しないにセットする。
        #######################################################################
        print_log('[DEBUG] P0900Action.action_N01_download_ippan_chosa.create_ippan_chosa_workbook()関数 STEP 3/10.', 'DEBUG')
        ws_building.sheet_view.showGridLines = False
        ws_ken.sheet_view.showGridLines = False
        ws_city.sheet_view.showGridLines = False
        ws_kasen_kaigan.sheet_view.showGridLines = False
        ws_suikei.sheet_view.showGridLines = False
        ws_suikei_type.sheet_view.showGridLines = False
        ws_kasen.sheet_view.showGridLines = False
        ws_kasen_type.sheet_view.showGridLines = False
        ws_cause.sheet_view.showGridLines = False
        ws_underground.sheet_view.showGridLines = False
        ws_usage.sheet_view.showGridLines = False
        ws_flood_sediment.sheet_view.showGridLines = False
        ws_gradient.sheet_view.showGridLines = False
        ws_industry.sheet_view.showGridLines = False
        ws_area.sheet_view.showGridLines = False
        ws_weather.sheet_view.showGridLines = False
        ws_suigai.sheet_view.showGridLines = False
        
        ws_city_vlook.sheet_view.showGridLines = False
        ws_kasen_vlook.sheet_view.showGridLines = False
        ws_suikei_type_vlook.sheet_view.showGridLines = False
        ws_kasen_type_vlook.sheet_view.showGridLines = False

        for i in range(suigai_count + 10 + 1):
            ws_ippan[i].sheet_view.showGridLines = False

        #######################################################################
        ### EXCEL入出力処理(0030)
        ### マスタ用のシートに値をセットする。
        #######################################################################
        print_log('[DEBUG] P0900Action.action_N01_download_ippan_chosa.create_ippan_chosa_workbook()関数 STEP 4/10.', 'DEBUG')
        ### 1000: 建物区分シート
        print("create_ippan_chosa_workbook_4_1", flush=True)
        if building_list:
            for i, building in enumerate(building_list):
                ws_building.cell(row=i+1, column=1).value = building.building_code
                ws_building.cell(row=i+1, column=2).value = str(building.building_name) + ":" + str(building.building_code)

        ### 1010: 都道府県シート
        print("create_ippan_chosa_workbook_4_2", flush=True)
        if ken_list:
            for i, ken in enumerate(ken_list):
                ws_ken.cell(row=i+1, column=1).value = ken.ken_code
                ws_ken.cell(row=i+1, column=2).value = str(ken.ken_name) + ":" + str(ken.ken_code)

        ### 1020: 市区町村シート
        print("create_ippan_chosa_workbook_4_3", flush=True)
        cities_list = []
        if ken_list:
            for i, ken in enumerate(ken_list):
                cities_list.append(CITY.objects.raw("""
                    SELECT 
                        * 
                    FROM CITY 
                    WHERE ken_code=%s 
                    ORDER BY CAST(CITY_CODE AS INTEGER)""", [ken.ken_code, ]))

        print("create_ippan_chosa_workbook_4_4", flush=True)
        if cities_list:
            for i, cities in enumerate(cities_list):
                if cities:
                    for j, city in enumerate(cities):
                        ws_city.cell(row=j+1, column=i*5+1).value = city.city_code
                        ws_city.cell(row=j+1, column=i*5+2).value = str(city.city_name) + ":" + str(city.city_code)
                        ws_city.cell(row=j+1, column=i*5+3).value = city.ken_code
                        ws_city.cell(row=j+1, column=i*5+4).value = city.city_population
                        ws_city.cell(row=j+1, column=i*5+5).value = city.city_area

        ### 1030: 水害発生地点工種（河川海岸区分）
        print("create_ippan_chosa_workbook_4_5", flush=True)
        if kasen_kaigan_list:
            for i, kasen_kaigan in enumerate(kasen_kaigan_list):
                ws_kasen_kaigan.cell(row=i+1, column=1).value = kasen_kaigan.kasen_kaigan_code
                ws_kasen_kaigan.cell(row=i+1, column=2).value = str(kasen_kaigan.kasen_kaigan_name) + ":" + str(kasen_kaigan.kasen_kaigan_code)

        ### 1040: 水系（水系・沿岸）
        print("create_ippan_chosa_workbook_4_6", flush=True)
        if suikei_list:
            for i, suikei in enumerate(suikei_list):
                ws_suikei.cell(row=i+1, column=1).value = suikei.suikei_code
                ws_suikei.cell(row=i+1, column=2).value = str(suikei.suikei_name) + ":" + str(suikei.suikei_code)
                ws_suikei.cell(row=i+1, column=3).value = suikei.suikei_type_code

        ### 1050: 水系種別（水系・沿岸種別）
        print("create_ippan_chosa_workbook_4_7", flush=True)
        if suikei_type_list:
            for i, suikei_type in enumerate(suikei_type_list):
                ws_suikei_type.cell(row=i+1, column=1).value = suikei_type.suikei_type_code
                ws_suikei_type.cell(row=i+1, column=2).value = str(suikei_type.suikei_type_name) + ":" + str(suikei_type.suikei_type_code)

        ### 1060: 河川（河川・海岸）、連動プルダウン用
        print("create_ippan_chosa_workbook_4_8", flush=True)
        kasens_list = []
        if suikei_list:
            for i, suikei in enumerate(suikei_list):
                kasens_list.append(KASEN.objects.raw("""
                    SELECT 
                        * 
                    FROM KASEN 
                    WHERE suikei_code=%s 
                    ORDER BY CAST(kasen_code AS INTEGER)""", [suikei.suikei_code, ]))

        print("create_ippan_chosa_workbook_4_9", flush=True)
        if kasens_list:
            for i, kasens in enumerate(kasens_list):
                if kasens:
                    for j, kasen in enumerate(kasens):
                        ws_kasen.cell(row=j+1, column=i*5+1).value = kasen.kasen_code
                        ws_kasen.cell(row=j+1, column=i*5+2).value = str(kasen.kasen_name) + ":" + str(kasen.kasen_code)
                        ws_kasen.cell(row=j+1, column=i*5+3).value = kasen.kasen_type_code
                        ws_kasen.cell(row=j+1, column=i*5+4).value = kasen.suikei_code

        ### 1070: 河川種別（河川・海岸種別）
        print("create_ippan_chosa_workbook_4_10", flush=True)
        if kasen_type_list:
            for i, kasen_type in enumerate(kasen_type_list):
                ws_kasen_type.cell(row=i+1, column=1).value = kasen_type.kasen_type_code
                ws_kasen_type.cell(row=i+1, column=2).value = str(kasen_type.kasen_type_name) + ":" + str(kasen_type.kasen_type_code)

        ### 1080: 水害原因
        print("create_ippan_chosa_workbook_4_11", flush=True)
        if cause_list:
            for i, cause in enumerate(cause_list):
                ws_cause.cell(row=i+1, column=1).value = cause.cause_code
                ws_cause.cell(row=i+1, column=2).value = str(cause.cause_name) + ":" + str(cause.cause_code)
        
        ### 1090: 地上地下区分
        print("create_ippan_chosa_workbook_4_12", flush=True)
        if underground_list:
            for i, underground in enumerate(underground_list):
                ws_underground.cell(row=i+1, column=1).value = underground.underground_code
                ws_underground.cell(row=i+1, column=2).value = str(underground.underground_name) + ":" + str(underground.underground_code)

        ### 1100: 地下空間の利用形態
        print("create_ippan_chosa_workbook_4_13", flush=True)
        if usage_list:
            for i, usage in enumerate(usage_list):
                ws_usage.cell(row=i+1, column=1).value = usage.usage_code
                ws_usage.cell(row=i+1, column=2).value = str(usage.usage_name) + ":" + str(usage.usage_code)

        ### 1110: 浸水土砂区分
        print("create_ippan_chosa_workbook_4_14", flush=True)
        if flood_sediment_list:
            for i, flood_sediment in enumerate(flood_sediment_list):
                ws_flood_sediment.cell(row=i+1, column=1).value = flood_sediment.flood_sediment_code
                ws_flood_sediment.cell(row=i+1, column=2).value = str(flood_sediment.flood_sediment_name) + ":" + str(flood_sediment.flood_sediment_code)

        ### 1120: 地盤勾配区分
        print("create_ippan_chosa_workbook_4_15", flush=True)
        if gradient_list:
            for i, gradient in enumerate(gradient_list):
                ws_gradient.cell(row=i+1, column=1).value = gradient.gradient_code
                ws_gradient.cell(row=i+1, column=2).value = str(gradient.gradient_name) + ":" + str(gradient.gradient_code)

        ### 1130: 産業分類
        print("create_ippan_chosa_workbook_4_16", flush=True)
        if industry_list:
            for i, industry in enumerate(industry_list):
                ws_industry.cell(row=i+1, column=1).value = industry.industry_code
                ws_industry.cell(row=i+1, column=2).value = str(industry.industry_name) + ":" + str(industry.industry_code)

        ### 7000: 入力データ_水害区域
        print("create_ippan_chosa_workbook_4_17", flush=True)
        if area_list:
            for i, area in enumerate(area_list):
                ws_area.cell(row=i+1, column=1).value = area.area_id
                ws_area.cell(row=i+1, column=2).value = str(area.area_name) + ":" + str(area.area_id)

        ### 7010: 入力データ_異常気象
        print("create_ippan_chosa_workbook_4_18", flush=True)
        if weather_list:
            for i, weather in enumerate(weather_list):
                ws_weather.cell(row=i+1, column=1).value = weather.weather_id
                ws_weather.cell(row=i+1, column=2).value = str(weather.weather_name) + ":" + str(weather.weather_id)

        ### 7020: 入力データ_ヘッダ部分、水害
        print("create_ippan_chosa_workbook_4_19", flush=True)
        if suigai_list:
            for i, suigai in enumerate(suigai_list):
                ws_suigai.cell(row=i+1, column=1).value = suigai.suigai_id
                ws_suigai.cell(row=i+1, column=2).value = str(suigai.suigai_name) + ":" + str(suigai.suigai_id)
                ws_suigai.cell(row=i+1, column=3).value = suigai.ken_code
                ws_suigai.cell(row=i+1, column=4).value = suigai.city_code
                ws_suigai.cell(row=i+1, column=5).value = suigai.begin_date
                ws_suigai.cell(row=i+1, column=6).value = suigai.end_date
                ws_suigai.cell(row=i+1, column=7).value = suigai.cause_1_code
                ws_suigai.cell(row=i+1, column=8).value = suigai.cause_2_code
                ws_suigai.cell(row=i+1, column=9).value = suigai.cause_3_code
                ws_suigai.cell(row=i+1, column=10).value = suigai.area_id
                ws_suigai.cell(row=i+1, column=11).value = suigai.suikei_code
                ws_suigai.cell(row=i+1, column=12).value = suigai.kasen_code
                ws_suigai.cell(row=i+1, column=13).value = suigai.gradient_code
                ws_suigai.cell(row=i+1, column=14).value = suigai.residential_area
                ws_suigai.cell(row=i+1, column=15).value = suigai.agricultural_area
                ws_suigai.cell(row=i+1, column=16).value = suigai.underground_area
                ws_suigai.cell(row=i+1, column=17).value = suigai.kasen_kaigan_code
                ws_suigai.cell(row=i+1, column=18).value = suigai.crop_damage
                ws_suigai.cell(row=i+1, column=19).value = suigai.weather_id

        #######################################################################
        ### EXCEL入出力処理(0040)
        ### VLOOKUP用のシートに値をセットする。
        #######################################################################
        print_log('[DEBUG] P0900Action.action_N01_download_ippan_chosa.create_ippan_chosa_workbook()関数 STEP 5/10.', 'DEBUG')
        ### 1020: 市区町村VLOOKUP
        print("create_ippan_chosa_workbook_5_1", flush=True)
        if ken_list and cities_list:
            for i, ken in enumerate(ken_list):
                ws_city_vlook.cell(row=i+1, column=1).value = str(ken.ken_name) + ":" + str(ken.ken_code)
    
            for i, cities in enumerate(cities_list):
                ws_city_vlook.cell(row=i+1, column=2).value = 'CITY!$' + VLOOK_VALUE[i] + '$1:$' + VLOOK_VALUE[i] + '$%d' % len(cities)

        ### 1060: 河川（河川・海岸）VLOOKUP
        print("create_ippan_chosa_workbook_5_2", flush=True)
        if suikei_list and kasens_list:
            for i, suikei in enumerate(suikei_list):
                ws_kasen_vlook.cell(row=i+1, column=1).value = str(suikei.suikei_name) + ":" + str(suikei.suikei_code)

            for i, kasens in enumerate(kasens_list):
                ws_kasen_vlook.cell(row=i+1, column=2).value = 'KASEN!$' + VLOOK_VALUE[i] + '$1:$' + VLOOK_VALUE[i] + '$%d' % len(kasens)

        #######################################################################
        ### EXCEL入出力処理(0050)
        ### 入力データ用EXCELシートのキャプションに値をセットする。
        #######################################################################
        print_log('[DEBUG] P0900Action.action_N01_download_ippan_chosa.create_ippan_chosa_workbook()関数 STEP 6/10.', 'DEBUG')
        for i in range(suigai_count + 10 + 1):
            ws_ippan[i].cell(row=5, column=2).value = '都道府県'
            ws_ippan[i].cell(row=5, column=3).value = '市区町村'
            ws_ippan[i].cell(row=5, column=4).value = '水害発生年月日'
            ws_ippan[i].cell(row=5, column=5).value = '水害終了年月日'
            ws_ippan[i].cell(row=5, column=6).value = '水害原因'
            ws_ippan[i].cell(row=5, column=9).value = '水害区域番号'
            ws_ippan[i].cell(row=6, column=6).value = '1'
            ws_ippan[i].cell(row=6, column=7).value = '2'
            ws_ippan[i].cell(row=6, column=8).value = '3'
            ws_ippan[i].cell(row=9, column=2).value = '水系・沿岸名'
            ws_ippan[i].cell(row=9, column=3).value = '水系種別'
            ws_ippan[i].cell(row=9, column=4).value = '河川・海岸名'
            ws_ippan[i].cell(row=9, column=5).value = '河川種別'
            ws_ippan[i].cell(row=9, column=6).value = '地盤勾配区分※1'
            ws_ippan[i].cell(row=12, column=2).value = '水害区域面積（m2）'
            ws_ippan[i].cell(row=12, column=6).value = '工種'
            ws_ippan[i].cell(row=12, column=8).value = '農作物被害額（千円）'
            ws_ippan[i].cell(row=12, column=10).value = '異常気象コード'
            ws_ippan[i].cell(row=16, column=2).value = '町丁名・大字名'
            ws_ippan[i].cell(row=16, column=3).value = '名称'
            ws_ippan[i].cell(row=16, column=4).value = '地上・地下被害の区分※2'
            ws_ippan[i].cell(row=16, column=5).value = '浸水土砂被害の区分※3'
            ws_ippan[i].cell(row=16, column=6).value = '被害建物棟数'
            ws_ippan[i].cell(row=16, column=12).value = '被害建物の延床面積（m2）'
            ws_ippan[i].cell(row=16, column=13).value = '被災世帯数'
            ws_ippan[i].cell(row=16, column=14).value = '被災事業所数'
            ws_ippan[i].cell(row=16, column=15).value = '被害建物内での農業家又は事業所活動'
            ws_ippan[i].cell(row=16, column=25).value = '事業所の産業区分※7'
            ws_ippan[i].cell(row=16, column=26).value = '地下空間の利用形態※8'
            ws_ippan[i].cell(row=16, column=27).value = '備考'
            ws_ippan[i].cell(row=17, column=7).value = '床上浸水・土砂堆積・地下浸水'
            ws_ippan[i].cell(row=17, column=15).value = '農家・漁家戸数※5'
            ws_ippan[i].cell(row=17, column=20).value = '事業所従業者数※6'
            ws_ippan[i].cell(row=18, column=16).value = '床上浸水'
            ws_ippan[i].cell(row=18, column=21).value = '床上浸水'
            ws_ippan[i].cell(row=20, column=7).value = '1cm〜49cm'
            ws_ippan[i].cell(row=20, column=8).value = '50cm〜99cm'
            ws_ippan[i].cell(row=20, column=9).value = '1m以上'
            ws_ippan[i].cell(row=20, column=10).value = '半壊※4'
            ws_ippan[i].cell(row=20, column=11).value = '全壊・流失※4'
            ws_ippan[i].cell(row=20, column=16).value = '1cm〜49cm'
            ws_ippan[i].cell(row=20, column=17).value = '50cm〜99cm'
            ws_ippan[i].cell(row=20, column=18).value = '1m以上半壊'
            ws_ippan[i].cell(row=20, column=19).value = '全壊・流失'
            ws_ippan[i].cell(row=20, column=21).value = '1cm〜49cm'
            ws_ippan[i].cell(row=20, column=22).value = '50cm〜99cm'
            ws_ippan[i].cell(row=20, column=23).value = '1m以上半壊'
            ws_ippan[i].cell(row=20, column=24).value = '全壊・流失'
            ws_ippan[i].cell(row=7, column=2).value = ""
            ws_ippan[i].cell(row=7, column=3).value = ""
            ws_ippan[i].cell(row=7, column=4).value = ""
            ws_ippan[i].cell(row=7, column=5).value = ""
            ws_ippan[i].cell(row=7, column=6).value = ""
            ws_ippan[i].cell(row=7, column=7).value = ""
            ws_ippan[i].cell(row=7, column=8).value = ""
            ws_ippan[i].cell(row=7, column=9).value = ""
            ws_ippan[i].cell(row=10, column=2).value = ""
            ws_ippan[i].cell(row=10, column=3).value = ""
            ws_ippan[i].cell(row=10, column=4).value = ""
            ws_ippan[i].cell(row=10, column=5).value = ""
            ws_ippan[i].cell(row=10, column=6).value = ""
            ws_ippan[i].cell(row=14, column=2).value = ""
            ws_ippan[i].cell(row=14, column=3).value = ""
            ws_ippan[i].cell(row=14, column=4).value = ""
            ws_ippan[i].cell(row=14, column=6).value = ""
            ws_ippan[i].cell(row=14, column=8).value = ""
            ws_ippan[i].cell(row=14, column=10).value = ""
            ws_ippan[i].cell(row=20, column=2).value = ""
            ws_ippan[i].cell(row=20, column=3).value = ""
            ws_ippan[i].cell(row=20, column=4).value = ""
            ws_ippan[i].cell(row=20, column=5).value = ""
            ws_ippan[i].cell(row=20, column=6).value = ""
            ws_ippan[i].cell(row=20, column=7).value = ""
            ws_ippan[i].cell(row=20, column=8).value = ""
            ws_ippan[i].cell(row=20, column=9).value = ""
            ws_ippan[i].cell(row=20, column=10).value = ""
            ws_ippan[i].cell(row=20, column=11).value = ""
            ws_ippan[i].cell(row=20, column=12).value = ""
            ws_ippan[i].cell(row=20, column=13).value = ""
            ws_ippan[i].cell(row=20, column=14).value = ""
            ws_ippan[i].cell(row=20, column=15).value = ""
            ws_ippan[i].cell(row=20, column=16).value = ""
            ws_ippan[i].cell(row=20, column=17).value = ""
            ws_ippan[i].cell(row=20, column=18).value = ""
            ws_ippan[i].cell(row=20, column=19).value = ""
            ws_ippan[i].cell(row=20, column=20).value = ""
            ws_ippan[i].cell(row=20, column=21).value = ""
            ws_ippan[i].cell(row=20, column=22).value = ""
            ws_ippan[i].cell(row=20, column=23).value = ""
            ws_ippan[i].cell(row=20, column=24).value = ""
            ws_ippan[i].cell(row=20, column=25).value = ""
            ws_ippan[i].cell(row=20, column=26).value = ""
            ws_ippan[i].cell(row=20, column=27).value = ""

        #######################################################################
        ### EXCEL入出力処理(0060)
        ### 入力データ用のEXCELシートの背景をセットする。
        #######################################################################
        print_log('[DEBUG] P0900Action.action_N01_download_ippan_chosa.create_ippan_chosa_workbook()関数 STEP 7/10.', 'DEBUG')
        gray_fill = PatternFill(bgColor='C0C0C0', fill_type='solid')
        white_fill = PatternFill(bgColor='FFFFFF', fill_type='solid')

        for i in range(suigai_count + 10 + 1):
            ws_ippan[i].conditional_formatting.add('N20:Y1000', FormulaRule(formula=['$C20="戸建住宅:1"'], fill=gray_fill))
            ws_ippan[i].conditional_formatting.add('N20:Y1000', FormulaRule(formula=['$C20="共同住宅:2"'], fill=gray_fill))
            ws_ippan[i].conditional_formatting.add('N20:Y1000', FormulaRule(formula=['$C20="事業所併用住宅:3"'], fill=white_fill))
            ws_ippan[i].conditional_formatting.add('M20:M1000', FormulaRule(formula=['$C20="事業所:4"'], fill=gray_fill))
            ws_ippan[i].conditional_formatting.add('M20:N1000', FormulaRule(formula=['$C20="その他建物:5"'], fill=gray_fill))
            ws_ippan[i].conditional_formatting.add('T20:Y1000', FormulaRule(formula=['$C20="その他建物:5"'], fill=gray_fill))
            ws_ippan[i].conditional_formatting.add('F20:Z1000', FormulaRule(formula=['$C20="建物以外:6"'], fill=gray_fill))

        #######################################################################
        ### EXCEL入出力処理(0070)
        ### 入力データ用のEXCELシートのプルダウンリストをセットする。
        #######################################################################
        print_log('[DEBUG] P0900Action.action_N01_download_ippan_chosa.create_ippan_chosa_workbook()関数 STEP 8/10.', 'DEBUG')
        for i in range(suigai_count + 10 + 1):
            ### 1000: 建物区分
            if len(building_list) > 0:
                dv_building = DataValidation(type="list", formula1="BUILDING!$B$1:$B$%d" % len(building_list))
                dv_building.ranges = 'C20:C1000'
                ws_ippan[i].add_data_validation(dv_building)
    
            ### 1010: 都道府県
            if len(ken_list) > 0:
                dv_ken = DataValidation(type="list", formula1="KEN!$B$1:$B$%d" % len(ken_list))
                dv_ken.ranges = 'B7:B7'
                ws_ippan[i].add_data_validation(dv_ken)
            
            ### 1020: 市区町村
            dv_city = DataValidation(type="list", formula1="=INDIRECT(AD3)")
            dv_city.ranges = 'C7:C7'
            ws_ippan[i].add_data_validation(dv_city)
            ### ws_ippan.cell(row=3, column=30).value = "=VLOOKUP(B7,CITY_VLOOK.A:B,2,0)" ### FOR LINUX?
            ws_ippan[i].cell(row=3, column=30).value = "=VLOOKUP(B7,CITY_VLOOK!A:B,2,0)" ### FOR WINDOWS
            
            ### 1030: 水害発生地点工種（河川海岸区分）
            if len(kasen_kaigan_list) > 0:
                dv_kasen_kaigan = DataValidation(type="list", formula1="KASEN_KAIGAN!$B$1:$B$%d" % len(kasen_kaigan_list))
                dv_kasen_kaigan.ranges = 'F14:F14'
                ws_ippan[i].add_data_validation(dv_kasen_kaigan)
            
            ### 1040: 水系（水系・沿岸）
            if len(suikei_list) > 0:
                dv_suikei = DataValidation(type="list", formula1="SUIKEI!$B$1:$B$%d" % len(suikei_list))
                dv_suikei.ranges = 'B10:B10'
                ws_ippan[i].add_data_validation(dv_suikei)
            
            ### 1050: 水系種別（水系・沿岸種別）
            if len(suikei_type_list) > 0:
                dv_suikei_type = DataValidation(type="list", formula1="SUIKEI_TYPE!$B$1:$B$%d" % len(suikei_type_list))
                dv_suikei_type.ranges = 'C10:C10'
                ws_ippan[i].add_data_validation(dv_suikei_type)
            
            ### 1060: 河川（河川・海岸）
            dv_kasen = DataValidation(type="list", formula1="=INDIRECT(AD4)")
            dv_kasen.ranges = 'D10:D10'
            ws_ippan[i].add_data_validation(dv_kasen)
            ### ws_ippan.cell(row=4, column=30).value = "=VLOOKUP(B10,KASEN_VLOOK.A:B,2,0)" ### FOR LINUX?
            ws_ippan[i].cell(row=4, column=30).value = "=VLOOKUP(B10,KASEN_VLOOK!A:B,2,0)" ### FOR WINDOWS
            
            ### 1070: 河川種別（河川・海岸種別）
            if len(kasen_type_list) > 0:
                dv_kasen_type = DataValidation(type="list", formula1="KASEN_TYPE!$B$1:$B$%d" % len(kasen_type_list))
                dv_kasen_type.ranges = 'E10:E10'
                ws_ippan[i].add_data_validation(dv_kasen_type)
            
            ### 1080: 水害原因
            if len(cause_list) > 0:
                dv_cause = DataValidation(type="list", formula1="CAUSE!$B$1:$B$%d" % len(cause_list))
                dv_cause.ranges = 'F7:H7'
                ws_ippan[i].add_data_validation(dv_cause)
            
            ### 1090: 地上地下区分
            if len(underground_list) > 0:
                dv_underground = DataValidation(type="list", formula1="UNDERGROUND!$B$1:$B$%d" % len(underground_list))
                dv_underground.ranges = 'D20:D1000'
                ws_ippan[i].add_data_validation(dv_underground)
            
            ### 1100: 地下空間の利用形態
            if len(usage_list) > 0:
                dv_usage = DataValidation(type="list", formula1="USAGE!$B$1:$B$%d" % len(usage_list))
                dv_usage.ranges = 'Z20:Z1000'
                ws_ippan[i].add_data_validation(dv_usage)
            
            ### 1110: 浸水土砂区分
            if len(flood_sediment_list) > 0:
                dv_flood_sediment = DataValidation(type="list", formula1="FLOOD_SEDIMENT!$B$1:$B$%d" % len(flood_sediment_list))
                dv_flood_sediment.ranges = 'E20:E1000'
                ws_ippan[i].add_data_validation(dv_flood_sediment)
            
            ### 1120: 地盤勾配区分
            if len(gradient_list) > 0:
                dv_gradient = DataValidation(type="list", formula1="GRADIENT!$B$1:$B$%d" % len(gradient_list))
                dv_gradient.ranges = 'F10:F10'
                ws_ippan[i].add_data_validation(dv_gradient)
            
            ### 1130: 産業分類
            if len(industry_list) > 0:
                dv_industry = DataValidation(type="list", formula1="INDUSTRY!$B$1:$B$%d" % len(industry_list))
                dv_industry.ranges = 'Y20:Y1000'
                ws_ippan[i].add_data_validation(dv_industry)

            ### 7000: 入力データ_水害区域
            if len(area_list) > 0:
                dv_area = DataValidation(type="list", formula1="AREA!$B$1:$B$%d" % len(area_list))
                dv_area.ranges = 'I7:I7'
                ws_ippan[i].add_data_validation(dv_area)
            
            ### 7010: 入力データ_異常気象
            if len(weather_list) > 0:
                dv_weather = DataValidation(type="list", formula1="WEATHER!$B$1:$B$%d" % len(weather_list))
                dv_weather.ranges = 'J14:J14'
                ws_ippan[i].add_data_validation(dv_weather)

        #######################################################################
        ### EXCEL入出力処理(0080)
        ### 入力データ用のEXCELシートに値をセットする。
        ### ※IPPANシート数分＝SUIGAI_COUNT+10+1分＝SUIGAI_ID_LIST、SUIGAI_NAME_LISTの数+10+1分ループする。
        #######################################################################
        print_log('[DEBUG] P0900Action.action_N01_download_ippan_chosa.create_ippan_chosa_workbook()関数 STEP 9/10.', 'DEBUG')

        ### 7020: 入力データ_ヘッダ部分、水害
        for i in range(suigai_count + 10 + 1):
            ###################################################################
            ### EXCEL入出力処理(0090)
            ###################################################################
            print_log('[DEBUG] P0900Action.action_N01_download_ippan_chosa.create_ippan_chosa_workbook()関数 STEP 9_1/10.', 'DEBUG')
            if len(str(ken_name)) > 0 and len(str(ken_code)) > 0:
                ws_ippan[i].cell(row=7, column=2).value = str(ken_name) + ":" + str(ken_code)
            else:
                ws_ippan[i].cell(row=7, column=2).value = ""
                
            if len(str(city_name)) > 0 and len(str(city_code)) > 0:
                ws_ippan[i].cell(row=7, column=3).value = str(city_name) + ":" + str(city_code)
            else:
                ws_ippan[i].cell(row=7, column=3).value = ""

        ### 7020: 入力データ_ヘッダ部分、水害
        for i in range(suigai_count):
            print_log('[DEBUG] P0900Action.action_N01_download_ippan_chosa.create_ippan_chosa_workbook()関数 STEP 9_2/10.', 'DEBUG')

            ###################################################################
            ### DBアクセス処理(0100)
            ### 入力データ_ヘッダ部分のデータを取得する。
            ###################################################################
            print_log('[DEBUG] P0900Action.action_N01_download_ippan_chosa.create_ippan_chosa_workbook()関数 STEP 9_3/10.', 'DEBUG')
            local_suigai_list = SUIGAI.objects.raw("""
                SELECT 
                    SG1.suigai_id AS suigai_id,
                    SG1.suigai_name AS suigai_name,
                    SG1.ken_code AS ken_code, 
                    KE1.ken_name AS ken_name,
                    SG1.city_code AS city_code,
                    CI1.city_name AS city_name,
                    TO_CHAR(timezone('JST', SG1.begin_date::timestamptz), 'yyyy/mm/dd') AS begin_date,
                    TO_CHAR(timezone('JST', SG1.end_date::timestamptz), 'yyyy/mm/dd') AS end_date,
                    SG1.cause_1_code AS cause_1_code,
                    CA1.cause_name AS cause_1_name,
                    SG1.cause_2_code AS cause_2_code,
                    CA2.cause_name AS cause_2_name,
                    SG1.cause_3_code AS cause_3_code,
                    CA3.cause_name AS cause_3_name,
                    SG1.area_id AS area_id,
                    AR1.area_name AS area_name,
                    
                    SG1.suikei_code AS suikei_code,
                    SK1.suikei_name AS suikei_name,
                    ST1.suikei_type_code AS suikei_type_code,
                    ST1.suikei_type_name AS suikei_type_name,
                    SG1.kasen_code AS kasen_code,
                    KA1.kasen_name AS kasen_name,
                    KT1.kasen_type_code AS kasen_type_code,
                    KT1.kasen_type_name AS kasen_type_name,
                    SG1.gradient_code AS gradient_code,
                    GR1.gradient_name AS gradient_name,
                    
                    CASE WHEN (SG1.residential_area) IS NULL THEN NULL ELSE CAST(SG1.residential_area AS NUMERIC(20,10)) END AS residential_area,
                    CASE WHEN (SG1.agricultural_area) IS NULL THEN NULL ELSE CAST(SG1.agricultural_area AS NUMERIC(20,10)) END AS agricultural_area,
                    CASE WHEN (SG1.underground_area) IS NULL THEN NULL ELSE CAST(SG1.underground_area AS NUMERIC(20,10)) END AS underground_area,
                    SG1.kasen_kaigan_code AS kasen_kaigan_code,
                    KK1.kasen_kaigan_name AS kasen_kaigan_name,
                    CASE WHEN (SG1.crop_damage) IS NULL THEN NULL ELSE CAST(SG1.crop_damage AS NUMERIC(20,10)) END AS crop_damage,
                    SG1.weather_id AS weather_id,
                    WE1.weather_name as weather_name
                    
                FROM SUIGAI SG1 
                LEFT JOIN KEN KE1 ON SG1.ken_code = KE1.ken_code 
                LEFT JOIN CITY CI1 ON SG1.city_code = CI1.city_code 
                LEFT JOIN CAUSE CA1 ON SG1.cause_1_code = CA1.cause_code 
                LEFT JOIN CAUSE CA2 ON SG1.cause_2_code = CA2.cause_code 
                LEFT JOIN CAUSE CA3 ON SG1.cause_3_code = CA3.cause_code 
                LEFT JOIN AREA AR1 ON SG1.area_id = AR1.area_id 
                LEFT JOIN SUIKEI SK1 ON SG1.suikei_code = SK1.suikei_code 
                LEFT JOIN SUIKEI_TYPE ST1 ON SK1.suikei_type_code = ST1.suikei_type_code 
                LEFT JOIN KASEN KA1 ON SG1.kasen_code = KA1.kasen_code 
                LEFT JOIN KASEN_TYPE KT1 ON KA1.kasen_type_code = KT1.kasen_type_code 
                LEFT JOIN GRADIENT GR1 ON SG1.gradient_code = GR1.gradient_code 
                LEFT JOIN KASEN_KAIGAN KK1 ON SG1.kasen_kaigan_code = KK1.kasen_kaigan_code 
                LEFT JOIN WEATHER WE1 ON SG1.weather_id = WE1.weather_id 
                WHERE 
                    SG1.suigai_id = %s LIMIT 1 -- suigai_id
                """, [
                    suigai_id_list[i], ### suigai_id
                ])
    
            ###################################################################
            ### DBアクセス処理(0110)
            ### 入力データ_一覧表部分のデータを取得する。
            ###################################################################
            print_log('[DEBUG] P0900Action.action_N01_download_ippan_chosa.create_ippan_chosa_workbook()関数 STEP 9_4/10.', 'DEBUG')
            local_ippan_list = IPPAN.objects.raw("""
                SELECT 
                    IV1.ippan_id AS ippan_id,
                    IV1.ippan_name AS ippan_name,
                    IV1.building_code AS building_code,
                    IV1.building_name AS building_Name,
                    IV1.underground_code AS underground_code,
                    IV1.underground_name AS underground_name,
                    IV1.flood_sediment_code AS flood_sediment_code,
                    IV1.flood_sediment_name AS flood_sediment_name,
                    
                    CASE WHEN (IV1.building_lv00) IS NULL THEN NULL ELSE CAST(IV1.building_lv00 AS NUMERIC(20,10)) END AS building_lv00,
                    CASE WHEN (IV1.building_lv01_49) IS NULL THEN NULL ELSE CAST(IV1.building_lv01_49 AS NUMERIC(20,10)) END AS building_lv01_49,
                    CASE WHEN (IV1.building_lv50_99) IS NULL THEN NULL ELSE CAST(IV1.building_lv50_99 AS NUMERIC(20,10)) END AS building_lv50_99,
                    CASE WHEN (IV1.building_lv100) IS NULL THEN NULL ELSE CAST(IV1.building_lv100 AS NUMERIC(20,10)) END AS building_lv100,
                    CASE WHEN (IV1.building_half) IS NULL THEN NULL ELSE CAST(IV1.building_half AS NUMERIC(20,10)) END AS building_half,
                    CASE WHEN (IV1.building_full) IS NULL THEN NULL ELSE CAST(IV1.building_full AS NUMERIC(20,10)) END AS building_full,
                    
                    CASE WHEN (IV1.floor_area) IS NULL THEN NULL ELSE CAST(IV1.floor_area AS NUMERIC(20,10)) END AS floor_area,
                    CASE WHEN (IV1.family) IS NULL THEN NULL ELSE CAST(IV1.family AS NUMERIC(20,10)) END AS family,
                    CASE WHEN (IV1.office) IS NULL THEN NULL ELSE CAST(IV1.office AS NUMERIC(20,10)) END AS office,
                    
                    CASE WHEN (IV1.floor_area_lv00) IS NULL THEN NULL ELSE CAST(IV1.floor_area_lv00 AS NUMERIC(20,10)) END AS floor_area_lv00,
                    CASE WHEN (IV1.floor_area_lv01_49) IS NULL THEN NULL ELSE CAST(IV1.floor_area_lv01_49 AS NUMERIC(20,10)) END AS floor_area_lv01_49,
                    CASE WHEN (IV1.floor_area_lv50_99) IS NULL THEN NULL ELSE CAST(IV1.floor_area_lv50_99 AS NUMERIC(20,10)) END AS floor_area_lv50_99,
                    CASE WHEN (IV1.floor_area_lv100) IS NULL THEN NULL ELSE CAST(IV1.floor_area_lv100 AS NUMERIC(20,10)) END AS floor_area_lv100,
                    CASE WHEN (IV1.floor_area_half) IS NULL THEN NULL ELSE CAST(IV1.floor_area_half AS NUMERIC(20,10)) END AS floor_area_half,
                    CASE WHEN (IV1.floor_area_full) IS NULL THEN NULL ELSE CAST(IV1.floor_area_full AS NUMERIC(20,10)) END AS floor_area_full,
                    
                    CASE WHEN (IV1.family_lv00) IS NULL THEN NULL ELSE CAST(IV1.family_lv00 AS NUMERIC(20,10)) END AS family_lv00,
                    CASE WHEN (IV1.family_lv01_49) IS NULL THEN NULL ELSE CAST(IV1.family_lv01_49 AS NUMERIC(20,10)) END AS family_lv01_49,
                    CASE WHEN (IV1.family_lv50_99) IS NULL THEN NULL ELSE CAST(IV1.family_lv50_99 AS NUMERIC(20,10)) END AS family_lv50_99,
                    CASE WHEN (IV1.family_lv100) IS NULL THEN NULL ELSE CAST(IV1.family_lv100 AS NUMERIC(20,10)) END AS family_lv100,
                    CASE WHEN (IV1.family_half) IS NULL THEN NULL ELSE CAST(IV1.family_half AS NUMERIC(20,10)) END AS family_half,
                    CASE WHEN (IV1.family_full) IS NULL THEN NULL ELSE CAST(IV1.family_full AS NUMERIC(20,10)) END AS family_full,
                    
                    CASE WHEN (IV1.office_lv00) IS NULL THEN NULL ELSE CAST(IV1.office_lv00 AS NUMERIC(20,10)) END AS office_lv00,
                    CASE WHEN (IV1.office_lv01_49) IS NULL THEN NULL ELSE CAST(IV1.office_lv01_49 AS NUMERIC(20,10)) END AS office_lv01_49,
                    CASE WHEN (IV1.office_lv50_99) IS NULL THEN NULL ELSE CAST(IV1.office_lv50_99 AS NUMERIC(20,10)) END AS office_lv50_99,
                    CASE WHEN (IV1.office_lv100) IS NULL THEN NULL ELSE CAST(IV1.office_lv100 AS NUMERIC(20,10)) END AS office_lv100,
                    CASE WHEN (IV1.office_half) IS NULL THEN NULL ELSE CAST(IV1.office_half AS NUMERIC(20,10)) END AS office_half,
                    CASE WHEN (IV1.office_full) IS NULL THEN NULL ELSE CAST(IV1.office_full AS NUMERIC(20,10)) END AS office_full,
                    
                    CASE WHEN (IV1.farmer_fisher_lv00) IS NULL THEN NULL ELSE CAST(IV1.farmer_fisher_lv00 AS NUMERIC(20,10)) END AS farmer_fisher_lv00,
                    CASE WHEN (IV1.farmer_fisher_lv01_49) IS NULL THEN NULL ELSE CAST(IV1.farmer_fisher_lv01_49 AS NUMERIC(20,10)) END AS farmer_fisher_lv01_49,
                    CASE WHEN (IV1.farmer_fisher_lv50_99) IS NULL THEN NULL ELSE CAST(IV1.farmer_fisher_lv50_99 AS NUMERIC(20,10)) END AS farmer_fisher_lv50_99,
                    CASE WHEN (IV1.farmer_fisher_lv100) IS NULL THEN NULL ELSE CAST(IV1.farmer_fisher_lv100 AS NUMERIC(20,10)) END AS farmer_fisher_lv100,
                    CASE WHEN (IV1.farmer_fisher_full) IS NULL THEN NULL ELSE CAST(IV1.farmer_fisher_full AS NUMERIC(20,10)) END AS farmer_fisher_full,
                    
                    CASE WHEN (IV1.employee_lv00) IS NULL THEN NULL ELSE CAST(IV1.employee_lv00 AS NUMERIC(20,10)) END AS employee_lv00,
                    CASE WHEN (IV1.employee_lv01_49) IS NULL THEN NULL ELSE CAST(IV1.employee_lv01_49 AS NUMERIC(20,10)) END AS employee_lv01_49,
                    CASE WHEN (IV1.employee_lv50_99) IS NULL THEN NULL ELSE CAST(IV1.employee_lv50_99 AS NUMERIC(20,10)) END AS employee_lv50_99,
                    CASE WHEN (IV1.employee_lv100) IS NULL THEN NULL ELSE CAST(IV1.employee_lv100 AS NUMERIC(20,10)) END AS employee_lv100,
                    CASE WHEN (IV1.employee_full) IS NULL THEN NULL ELSE CAST(IV1.employee_full AS NUMERIC(20,10)) END AS employee_full,
                    
                    IV1.industry_code AS industry_code,
                    IV1.industry_name AS industry_name,
                    IV1.usage_code as usage_code,
                    IV1.usage_name as usage_name,
                    IV1.comment as comment 
                    
                FROM IPPAN_VIEW IV1 
                WHERE 
                    IV1.suigai_id = %s -- suigai_id
                ORDER BY CAST (IV1.ippan_id AS INTEGER)
                """, [
                    suigai_id_list[i], ### suigai_id
                ])

            ###################################################################
            ### EXCEL入出力処理(0120)
            ### EXCELのヘッダ部のセルに、入力データ_ヘッダ部分の値を埋め込む。
            ### ※ k ループの回数＝1
            ###################################################################
            print_log('[DEBUG] P0900Action.action_N01_download_ippan_chosa.create_ippan_chosa_workbook()関数 STEP 9_5/10.', 'DEBUG')
            if local_suigai_list:
                for k, suigai in enumerate(local_suigai_list):
                    if suigai.ken_name is not None and suigai.ken_code is not None and len(str(suigai.ken_name)) > 0 and len(str(suigai.ken_code)) > 0:
                        ws_ippan[i].cell(row=7, column=2).value = str(suigai.ken_name) + ":" + str(suigai.ken_code)
                    
                    if suigai.city_name is not None and suigai.city_code is not None and len(str(suigai.city_name)) > 0 and len(str(suigai.city_code)) > 0:
                        ws_ippan[i].cell(row=7, column=3).value = str(suigai.city_name) + ":" + str(suigai.city_code)

                    if suigai.begin_date is not None and len(str(suigai.begin_date)) > 0:
                        ws_ippan[i].cell(row=7, column=4).value = str(suigai.begin_date)

                    if suigai.end_date is not None and len(str(suigai.end_date)) > 0:
                        ws_ippan[i].cell(row=7, column=5).value = str(suigai.end_date)

                    if suigai.cause_1_name is not None and suigai.cause_1_code is not None and len(str(suigai.cause_1_name)) > 0 and len(str(suigai.cause_1_code)) > 0:
                        ws_ippan[i].cell(row=7, column=6).value = str(suigai.cause_1_name) + ":" + str(suigai.cause_1_code)

                    if suigai.cause_2_name is not None and suigai.cause_2_code is not None and len(str(suigai.cause_2_name)) > 0 and len(str(suigai.cause_2_code)) > 0:
                        ws_ippan[i].cell(row=7, column=7).value = str(suigai.cause_2_name) + ":" + str(suigai.cause_2_code)

                    if suigai.cause_3_name is not None and suigai.cause_3_code is not None and len(str(suigai.cause_3_name)) > 0 and len(str(suigai.cause_3_code)) > 0:
                        ws_ippan[i].cell(row=7, column=8).value = str(suigai.cause_3_name) + ":" + str(suigai.cause_3_code)

                    if suigai.area_name is not None and suigai.area_id is not None and len(str(suigai.area_name)) > 0 and len(str(suigai.area_id)) > 0:
                        ws_ippan[i].cell(row=7, column=9).value = str(suigai.area_name) + ":" + str(suigai.area_id)

                    if suigai.suikei_name is not None and suigai.suikei_code is not None and len(str(suigai.suikei_name)) > 0 and len(str(suigai.suikei_code)) > 0:
                        ws_ippan[i].cell(row=10, column=2).value = str(suigai.suikei_name) + ":" + str(suigai.suikei_code)

                    if suigai.suikei_type_name is not None and suigai.suikei_type_code is not None and len(str(suigai.suikei_type_name)) > 0 and len(str(suigai.suikei_type_code)) > 0:
                        ws_ippan[i].cell(row=10, column=3).value = str(suigai.suikei_type_name) + ":" + str(suigai.suikei_type_code)

                    if suigai.kasen_name is not None and suigai.kasen_code is not None and len(str(suigai.kasen_name)) > 0 and len(str(suigai.kasen_code)) > 0:
                        ws_ippan[i].cell(row=10, column=4).value = str(suigai.kasen_name) + ":" + str(suigai.kasen_code)

                    if suigai.kasen_type_name is not None and suigai.kasen_type_code is not None and len(str(suigai.kasen_type_name)) > 0 and len(str(suigai.kasen_type_code)) > 0:
                        ws_ippan[i].cell(row=10, column=5).value = str(suigai.kasen_type_name) + ":" + str(suigai.kasen_type_code)

                    if suigai.gradient_name is not None and suigai.gradient_code is not None and len(str(suigai.gradient_name)) > 0 and len(str(suigai.gradient_code)) > 0:
                        ws_ippan[i].cell(row=10, column=6).value = str(suigai.gradient_name) + ":" + str(suigai.gradient_code)

                    if suigai.residential_area is not None and len(str(suigai.residential_area)) > 0:
                        ws_ippan[i].cell(row=14, column=2).value = str(suigai.residential_area)

                    if suigai.agricultural_area is not None and len(str(suigai.agricultural_area)) > 0:
                        ws_ippan[i].cell(row=14, column=3).value = str(suigai.agricultural_area)

                    if suigai.underground_area is not None and len(str(suigai.underground_area)) > 0:
                        ws_ippan[i].cell(row=14, column=4).value = str(suigai.underground_area)

                    if suigai.kasen_kaigan_name is not None and suigai.kasen_kaigan_code is not None and len(str(suigai.kasen_kaigan_name)) > 0 and len(str(suigai.kasen_kaigan_code)) > 0:
                        ws_ippan[i].cell(row=14, column=6).value = str(suigai.kasen_kaigan_name) + ":" + str(suigai.kasen_kaigan_code)

                    if suigai.crop_damage is not None and len(str(suigai.crop_damage)) > 0:
                        ws_ippan[i].cell(row=14, column=8).value = str(suigai.crop_damage)

                    if suigai.weather_name is not None and suigai.weather_id is not None and len(str(suigai.weather_name)) > 0 and len(str(suigai.weather_id)) > 0:
                        ws_ippan[i].cell(row=14, column=10).value = str(suigai.weather_name) + ":" + str(suigai.weather_id)

                    if suigai.suigai_id is not None and len(str(suigai.suigai_id)) > 0:
                        ws_ippan[i].cell(row=3, column=28).value = suigai.suigai_id
            else:
                pass
                    
            ###################################################################
            ### EXCEL入出力処理(0130)
            ### EXCELの一覧部のセルに、入力データ_一覧部分の値を埋め込む。
            ### ※ k ループの回数＝行数
            ###################################################################
            print_log('[DEBUG] P0900Action.action_N01_download_ippan_chosa.create_ippan_chosa_workbook()関数 STEP 9_6/10.', 'DEBUG')
            if local_ippan_list:
                for k, ippan in enumerate(local_ippan_list):
                    ### ws_ippan[i].cell(row=k+20, column=2).value = str(ippan.ippan_name) + ":" + str(ippan.ippan_id)
                    if ippan.ippan_name is not None and len(str(ippan.ippan_name)) > 0:
                        ws_ippan[i].cell(row=k+20, column=2).value = str(ippan.ippan_name)
                    
                    if ippan.building_name is not None and ippan.building_code is not None and len(str(ippan.building_name)) > 0 and len(str(ippan.building_code)) > 0:
                        ws_ippan[i].cell(row=k+20, column=3).value = str(ippan.building_name) + ":" + str(ippan.building_code)

                    if ippan.underground_name is not None and ippan.underground_code is not None and len(str(ippan.underground_name)) > 0 and len(str(ippan.underground_code)) > 0:
                        ws_ippan[i].cell(row=k+20, column=4).value = str(ippan.underground_name) + ":" + str(ippan.underground_code)

                    if ippan.flood_sediment_name is not None and ippan.flood_sediment_code is not None and len(str(ippan.flood_sediment_name)) > 0 and len(str(ippan.flood_sediment_code)) > 0:
                        ws_ippan[i].cell(row=k+20, column=5).value = str(ippan.flood_sediment_name) + ":" + str(ippan.flood_sediment_code)

                    if ippan.building_lv00 is not None and len(str(ippan.building_lv00)) > 0:
                        ws_ippan[i].cell(row=k+20, column=6).value = ippan.building_lv00

                    if ippan.building_lv01_49 is not None and len(str(ippan.building_lv01_49)) > 0:
                        ws_ippan[i].cell(row=k+20, column=7).value = ippan.building_lv01_49

                    if ippan.building_lv50_99 is not None and len(str(ippan.building_lv50_99)) > 0:
                        ws_ippan[i].cell(row=k+20, column=8).value = ippan.building_lv50_99

                    if ippan.building_lv100 is not None and len(str(ippan.building_lv100)) > 0:
                        ws_ippan[i].cell(row=k+20, column=9).value = ippan.building_lv100

                    if ippan.building_half is not None and len(str(ippan.building_half)) > 0:
                        ws_ippan[i].cell(row=k+20, column=10).value = ippan.building_half

                    if ippan.building_full is not None and len(str(ippan.building_full)) > 0:
                        ws_ippan[i].cell(row=k+20, column=11).value = ippan.building_full

                    if ippan.floor_area is not None and len(str(ippan.floor_area)) > 0:
                        ws_ippan[i].cell(row=k+20, column=12).value = ippan.floor_area

                    if ippan.family is not None and len(str(ippan.family)) > 0:
                        ws_ippan[i].cell(row=k+20, column=13).value = ippan.family

                    if ippan.office is not None and len(str(ippan.office)) > 0:
                        ws_ippan[i].cell(row=k+20, column=14).value = ippan.office

                    if ippan.farmer_fisher_lv00 is not None and len(str(ippan.farmer_fisher_lv00)) > 0:
                        ws_ippan[i].cell(row=k+20, column=15).value = ippan.farmer_fisher_lv00

                    if ippan.farmer_fisher_lv01_49 is not None and len(str(ippan.farmer_fisher_lv01_49)) > 0:
                        ws_ippan[i].cell(row=k+20, column=16).value = ippan.farmer_fisher_lv01_49

                    if ippan.farmer_fisher_lv50_99 is not None and len(str(ippan.farmer_fisher_lv50_99)) > 0:
                        ws_ippan[i].cell(row=k+20, column=17).value = ippan.farmer_fisher_lv50_99

                    if ippan.farmer_fisher_lv100 is not None and len(str(ippan.farmer_fisher_lv100)) > 0:
                        ws_ippan[i].cell(row=k+20, column=18).value = ippan.farmer_fisher_lv100

                    if ippan.farmer_fisher_full is not None and len(str(ippan.farmer_fisher_full)) > 0:
                        ws_ippan[i].cell(row=k+20, column=19).value = ippan.farmer_fisher_full

                    if ippan.employee_lv00 is not None and len(str(ippan.employee_lv00)) > 0:
                        ws_ippan[i].cell(row=k+20, column=20).value = ippan.employee_lv00

                    if ippan.employee_lv01_49 is not None and len(str(ippan.employee_lv01_49)) > 0:
                        ws_ippan[i].cell(row=k+20, column=21).value = ippan.employee_lv01_49

                    if ippan.employee_lv50_99 is not None and len(str(ippan.employee_lv50_99)) > 0:
                        ws_ippan[i].cell(row=k+20, column=22).value = ippan.employee_lv50_99

                    if ippan.employee_lv100 is not None and len(str(ippan.employee_lv100)) > 0:
                        ws_ippan[i].cell(row=k+20, column=23).value = ippan.employee_lv100

                    if ippan.employee_full is not None and len(str(ippan.employee_full)) > 0:
                        ws_ippan[i].cell(row=k+20, column=24).value = ippan.employee_full

                    if ippan.industry_name is not None and ippan.industry_code is not None and len(str(ippan.industry_name)) > 0 and len(str(ippan.industry_code)):
                        ws_ippan[i].cell(row=k+20, column=25).value = str(ippan.industry_name) + ":" + str(ippan.industry_code)

                    if ippan.usage_name is not None and ippan.usage_code is not None and len(str(ippan.usage_name)) > 0 and len(str(ippan.usage_code)):
                        ws_ippan[i].cell(row=k+20, column=26).value = str(ippan.usage_name) + ":" + str(ippan.usage_code)

                    if ippan.comment is not None and len(str(ippan.comment)) > 0:
                        ws_ippan[i].cell(row=k+20, column=27).value = ippan.comment

                    if ippan.ippan_id is not None and len(str(ippan.ippan_id)) > 0:
                        ws_ippan[i].cell(row=k+20, column=28).value = ippan.ippan_id
            else:
                pass

        #######################################################################
        ### 戻り値セット処理(0140)
        ### 戻り値を戻す。
        #######################################################################
        print_log('[DEBUG] P0900Action.action_N01_download_ippan_chosa.create_ippan_chosa_workbook()関数 STEP 10/10.', 'DEBUG')
        print_log('[INFO] P0900Action.action_N01_download_ippan_chosa.create_ippan_chosa_workbook()関数が正常終了しました。', 'INFO')
        return wb
        
    except:
        print_log(sys.exc_info()[0], 'ERROR')
        print_log('[ERROR] P0900Action.action_N01_download_ippan_chosa.create_ippan_chosa_workbook()関数でエラーが発生しました。', 'ERROR')
        print_log('[ERROR] P0900Action.action_N01_download_ippan_chosa.create_ippan_chosa_workbook()関数が異常終了しました。', 'ERROR')
        return wb

###############################################################################
### クラス名： Command
###############################################################################
class Command(BaseCommand):
    
    ###########################################################################
    ### 関数名： handle
    ### チェックOKの場合、トリガーの状態を成功に更新、消費日時をカレントに更新、新たなトリガーを生成する。
    ### チェックNGの場合、トリガーの状態を失敗に更新、消費日時をカレントに更新する。
    ### チェックNGの場合、当該水害IDについて、以降の処理を止めて、手動？で再実行、または、入力データから再登録するイメージである。
    ### 上記は、このアプリの共通の考え方とする。
    ###########################################################################
    def handle(self, *args, **options):
        try:
            ###################################################################
            ### 引数チェック処理(0000)
            ### コマンドラインからの引数をチェックする。
            ###################################################################
            reset_log()
            print_log('[INFO] P0900Action.action_N01_download_ippan_chosa.handle()関数が開始しました。', 'INFO')
            print_log('[DEBUG] P0900Action.action_N01_download_ippan_chosa.handle()関数 STEP 1/8.', 'DEBUG')

            ###################################################################
            ### DBアクセス処理(0010)
            ### TRIGGERテーブルからSUIGAI_IDのリストを取得する。
            ### トリガーメッセージにアクションが発行されているかを検索する。
            ### 処理対象の水害IDを取得する。
            ### トリガーメッセージにアクションが発行されていなければ、処理を終了する。
            ### ※ネストを浅くするために、処理対象外の場合、処理を終了させる。
            ###################################################################
            print_log('[DEBUG] P0900Action.action_N01_download_ippan_chosa.handle()関数 STEP 2/8.', 'DEBUG')
            trigger_list = TRIGGER.objects.raw("""
                SELECT 
                    * 
                FROM TRIGGER 
                WHERE 
                    action_code='N01' AND 
                    consumed_at IS NULL AND 
                    deleted_at IS NULL 
                ORDER BY CAST(ken_code AS INTEGER), CAST(city_code AS INTEGER) LIMIT 1""", [])
            
            if trigger_list is None:
                print_log('[INFO] P0900Action.action_N01_download_ippan_chosa.handle()関数が正常終了しました。', 'INFO')
                return 0
            
            if len(trigger_list) == 0:
                print_log('[INFO] P0900Action.action_N01_download_ippan_chosa.handle()関数が正常終了しました。', 'INFO')
                return 0
            
            print_log('[DEBUG] P0900Action.action_N01_download_ippan_chosa.handle()関数 trigger_list[0].trigger_id = {}'.format(trigger_list[0].trigger_id), 'DEBUG')
            print_log('[DEBUG] P0900Action.action_N01_download_ippan_chosa.handle()関数 trigger_list[0].suigai_id = {}'.format(trigger_list[0].suigai_id), 'DEBUG')
            print_log('[DEBUG] P0900Action.action_N01_download_ippan_chosa.handle()関数 trigger_list[0].city_code = {}'.format(trigger_list[0].city_code), 'DEBUG')
            print_log('[DEBUG] P0900Action.action_N01_download_ippan_chosa.handle()関数 trigger_list[0].ken_code = {}'.format(trigger_list[0].ken_code), 'DEBUG')
            print_log('[DEBUG] P0900Action.action_N01_download_ippan_chosa.handle()関数 trigger_list[0].download_file_path = {}'.format(trigger_list[0].download_file_path), 'DEBUG')
            print_log('[DEBUG] P0900Action.action_N01_download_ippan_chosa.handle()関数 trigger_list[0].download_file_name = {}'.format(trigger_list[0].download_file_name), 'DEBUG')
            print_log('[DEBUG] P0900Action.action_N01_download_ippan_chosa.handle()関数 trigger_list[0].upload_file_path = {}'.format(trigger_list[0].upload_file_path), 'DEBUG')
            print_log('[DEBUG] P0900Action.action_N01_download_ippan_chosa.handle()関数 trigger_list[0].upload_file_name = {}'.format(trigger_list[0].upload_file_name), 'DEBUG')
    
            ###################################################################
            ### DBアクセス処理(0020)
            ### 市区町村コードを検索キーに、入力データのヘッダ部の件数を取得する。
            ### ※関数のメインの処理の前に、EXCELのシート数を確定するため。
            ###################################################################
            print_log('[DEBUG] P0900Action.action_N01_download_ippan_chosa.handle()関数 STEP 3/8.', 'DEBUG')
            suigai_count_request = 0
            suigai_id_request = []
            
            suigai_list = SUIGAI.objects.raw("""
                SELECT 
                    * 
                FROM SUIGAI 
                WHERE 
                    city_code=%s AND deleted_at IS NULL 
                ORDER BY CAST(suigai_id AS INTEGER)""", [trigger_list[0].city_code, ])
            
            if suigai_list:
                suigai_count_request = len(suigai_list)
            else:
                suigai_count_request = 0

            ### suigai_id_request.append([suigai.suigai_id for suigai in suigai_list])
            suigai_id_request = [suigai.suigai_id for suigai in suigai_list]
                
            print_log('[DEBUG] P0900Action.action_N01_download_ippan_chosa.handle()関数 suigai_count_request = {}'.format(suigai_count_request), 'DEBUG')
            print_log('[DEBUG] P0900Action.action_N01_download_ippan_chosa.handle()関数 suigai_id_request = {}'.format(suigai_id_request), 'DEBUG')
    
            ###################################################################
            ### DBアクセス処理(0030)
            ### 市区町村コードを検索キーに、都道府県コード、都道府県名、市区町村名を取得する。
            ###################################################################
            print_log('[DEBUG] P0900Action.action_N01_download_ippan_chosa.handle()関数 STEP 4/8.', 'DEBUG')
            ken_code_request = trigger_list[0].ken_code
            city_code_request = trigger_list[0].city_code

            city_list = CITY.objects.raw("""
                SELECT 
                    CT1.city_code AS city_code, 
                    CT1.city_name AS city_name, 
                    CT1.ken_code AS ken_code, 
                    KE1.ken_name AS ken_name 
                FROM CITY CT1 
                LEFT JOIN KEN KE1 ON CT1.ken_code=KE1.ken_code 
                WHERE 
                    CT1.city_code=%s LIMIT 1""", [trigger_list[0].city_code, ])
            
            ken_name_request = city_list[0].ken_name
            city_name_request = city_list[0].city_name
                
            print_log('[DEBUG] P0900Action.action_N01_download_ippan_chosa.handle()関数 ken_code_request = {}'.format(ken_code_request), 'DEBUG')
            print_log('[DEBUG] P0900Action.action_N01_download_ippan_chosa.handle()関数 ken_name_request = {}'.format(ken_name_request), 'DEBUG')
            print_log('[DEBUG] P0900Action.action_N01_download_ippan_chosa.handle()関数 city_code_request = {}'.format(city_code_request), 'DEBUG')
            print_log('[DEBUG] P0900Action.action_N01_download_ippan_chosa.handle()関数 city_name_request = {}'.format(city_name_request), 'DEBUG')
    
            ###################################################################
            ### DBアクセス処理(0040)
            ### マスタデータを取得する。
            ###################################################################
            print_log('[DEBUG] P0900Action.action_N01_download_ippan_chosa.handle()関数 STEP 5/8.', 'DEBUG')
            ### 1000: 建物区分シート
            ### 1010: 都道府県シート
            ### 1020: 市区町村シート、連動プルダウン用、VLOOKUP用
            ### 1030: 水害発生地点工種（河川海岸区分）
            ### 1040: 水系（水系・沿岸）
            ### 1050: 水系種別（水系・沿岸種別）
            ### 1060: 河川（河川・海岸）、連動プルダウン用、VLOOKUP用
            ### 1070: 河川種別（河川・海岸種別）
            ### 1080: 水害原因
            ### 1090: 地上地下区分
            ### 1100: 地下空間の利用形態
            ### 1110: 浸水土砂区分
            ### 1120: 地盤勾配区分
            ### 1130: 産業分類
            ### 7000: 入力データ_水害区域
            ### 7010: 入力データ_異常気象
            ### 7020: 入力データ_ヘッダ部分、水害
            building_list = BUILDING.objects.raw("""SELECT * FROM BUILDING ORDER BY CAST(BUILDING_CODE AS INTEGER)""", [])
            
            cities_list = []
            ken_list = KEN.objects.raw("""SELECT * FROM KEN ORDER BY CAST(KEN_CODE AS INTEGER)""", [])
            if ken_list:
                for i, ken in enumerate(ken_list):
                    cities_list.append(CITY.objects.raw("""SELECT * FROM CITY WHERE KEN_CODE=%s ORDER BY CAST(CITY_CODE AS INTEGER)""", [ken.ken_code,]))
    
            kasen_kaigan_list = KASEN_KAIGAN.objects.raw("""SELECT * FROM KASEN_KAIGAN ORDER BY CAST(KASEN_KAIGAN_CODE AS INTEGER)""", [])
            suikei_list = SUIKEI.objects.raw("""SELECT * FROM SUIKEI ORDER BY CAST(SUIKEI_CODE AS INTEGER)""", [])
            suikei_type_list = SUIKEI_TYPE.objects.raw("""SELECT * FROM SUIKEI_TYPE ORDER BY CAST(SUIKEI_TYPE_CODE AS INTEGER)""", [])
    
            kasens_list = []
            suikei_list = SUIKEI.objects.raw("""SELECT * FROM SUIKEI ORDER BY CAST(SUIKEI_CODE AS INTEGER)""", [])
            if suikei_list:
                for i, suikei in enumerate(suikei_list):
                    kasens_list.append(KASEN.objects.raw("""SELECT * FROM KASEN WHERE SUIKEI_CODE=%s ORDER BY CAST(KASEN_CODE AS INTEGER)""", [suikei.suikei_code,]))
    
            kasen_type_list = KASEN_TYPE.objects.raw("""SELECT * FROM KASEN_TYPE ORDER BY CAST(KASEN_TYPE_CODE AS INTEGER)""", [])
            cause_list = CAUSE.objects.raw("""SELECT * FROM CAUSE ORDER BY CAST(CAUSE_CODE AS INTEGER)""", [])
            underground_list = UNDERGROUND.objects.raw("""SELECT * FROM UNDERGROUND ORDER BY CAST(UNDERGROUND_CODE AS INTEGER)""", [])
            usage_list = USAGE.objects.raw("""SELECT * FROM USAGE ORDER BY CAST(USAGE_CODE AS INTEGER)""", [])
            flood_sediment_list = FLOOD_SEDIMENT.objects.raw("""SELECT * FROM FLOOD_SEDIMENT ORDER BY CAST(FLOOD_SEDIMENT_CODE AS INTEGER)""", [])
            gradient_list = GRADIENT.objects.raw("""SELECT * FROM GRADIENT ORDER BY CAST(GRADIENT_CODE AS INTEGER)""", [])
            industry_list = INDUSTRY.objects.raw("""SELECT * FROM INDUSTRY ORDER BY CAST(INDUSTRY_CODE AS INTEGER)""", [])
            area_list = AREA.objects.raw("""SELECT * FROM AREA ORDER BY CAST(AREA_ID AS INTEGER)""", [])
            weather_list = WEATHER.objects.raw("""SELECT * FROM WEATHER ORDER BY CAST(WEATHER_ID AS INTEGER)""", [])
            suigai_list = SUIGAI.objects.raw("""SELECT * FROM SUIGAI ORDER BY CAST(SUIGAI_ID AS INTEGER)""", [])
            
            ###################################################################
            ### EXCEL入出力処理(0050)
            ###################################################################
            print_log('[DEBUG] P0900Action.action_N01_download_ippan_chosa.handle()関数 STEP 6/8.', 'DEBUG')
            template_file_path = 'static/template_ippan_chosa.xlsx'
            ### download_file_path = 'static/ippan_chosa_' + str(ken_name_request) + '_' + str(city_name_request) + '.xlsx'
            ### download_file_name = 'ippan_chosa_' + str(ken_name_request) + '_' + str(city_name_request) + '.xlsx'
            download_file_path = trigger_list[0].download_file_path
            download_file_name = trigger_list[0].download_file_name
            print_log('[DEBUG] P0900Action.action_N01_download_ippan_chosa.handle()関数 download_file_path = {}'.format(download_file_path), 'DEBUG')
            print_log('[DEBUG] P0900Action.action_N01_download_ippan_chosa.handle()関数 download_file_name = {}'.format(download_file_name), 'DEBUG')
            
            wb = create_ippan_chosa_workbook(
                suigai_count=suigai_count_request, 
                suigai_id_list=suigai_id_request, 
                suigai_name_list=suigai_id_request, 
                ken_code=ken_code_request, 
                ken_name=ken_name_request, 
                city_code=city_code_request, 
                city_name=city_name_request, 
                VLOOK_VALUE=VLOOK_VALUE, 
                building_list=building_list, 
                ken_list=ken_list, 
                kasen_kaigan_list=kasen_kaigan_list, 
                suikei_list=suikei_list, 
                suikei_type_list=suikei_type_list, 
                kasen_type_list=kasen_type_list, 
                cause_list=cause_list, 
                underground_list=underground_list, 
                usage_list=usage_list, 
                flood_sediment_list=flood_sediment_list, 
                gradient_list=gradient_list, 
                industry_list=industry_list, 
                area_list=area_list, 
                weather_list=weather_list, 
                suigai_list=suigai_list)
            
            wb.save(download_file_path)

            ###################################################################
            ### DBアクセス処理(0060)
            ### トリガーデータを更新する。
            ###################################################################
            print_log('[DEBUG] P0900Action.action_N01_download_ippan_chosa.handle()関数 STEP 7/8.', 'DEBUG')
            connection_cursor = connection.cursor()
            try:
                connection_cursor.execute("""BEGIN""", [])
                
                if suigai_count_request == 0:
                    print("ippan_chosa_view_7_1", flush=True)
                    connection_cursor.execute("""
                        UPDATE TRIGGER SET 
                            status_code=%s, -- status_code
                            success_count=%s, -- success_count
                            failure_count=%s, -- failure_count
                            consumed_at=CURRENT_TIMESTAMP, 
                            integrity_ok=%s, -- integrity_ok
                            integrity_ng=%s  -- integrity_ng
                        WHERE
                            trigger_id=%s -- trigger_id
                        """, [
                            'SUCCESS', ### status_count
                            1, ### success_count
                            0, ### failure_count
                            '\n'.join(get_info_log()), ### integrity_ok
                            '\n'.join(get_warn_log()), ### integrity_ng
                            trigger_list[0].trigger_id, ### trigger_id
                        ])
                else: 
                    print("ippan_chosa_view_7_2", flush=True)
                    connection_cursor.execute("""
                        UPDATE TRIGGER SET 
                            status_code=%s, -- status_code
                            success_count=%s, -- success_count
                            failure_count=%s, -- failure_count
                            consumed_at=CURRENT_TIMESTAMP, 
                            integrity_ok=%s, -- integrity_ok
                            integrity_ng=%s  -- integrity_ng
                        WHERE 
                            trigger_id=%s -- trigger_id
                        """, [
                            'SUCCESS', ### status_code
                            1, ### success_count
                            0, ### failure_count
                            '\n'.join(get_info_log()), ### integrity_ok
                            '\n'.join(get_warn_log()), ### integrity_ng
                            trigger_list[0].trigger_id, ### trigger_id
                        ])
                ### transaction.commit()    
                connection_cursor.execute("""COMMIT""", [])
            except:
                print_log('[ERROR] P0900Action.action_N01_download_ippan_chosa.handle()関数 {}'.format(sys.exc_info()[0]), 'ERROR')
                ### connection_cursor.rollback()
                connection_cursor.execute("""ROLLBACK""", [])
            finally:
                connection_cursor.close()
                
            ###################################################################
            ### 戻り値セット処理(0070)
            ### 戻り値を戻す。
            ###################################################################
            print_log('[DEBUG] P0900Action.action_N01_download_ippan_chosa.handle()関数 STEP 8/8.', 'DEBUG')
            print_log('[INFO] P0900Action.action_N01_download_ippan_chosa.handle()関数が正常終了しました。', 'INFO')
            return 0
        except:
            print_log('[ERROR] P0900Action.action_N01_download_ippan_chosa.handle()関数 {}'.format(sys.exc_info()[0]), 'ERROR')
            print_log('[ERROR] P0900Action.action_N01_download_ippan_chosa.handle()関数でエラーが発生しました。', 'ERROR')
            print_log('[ERROR] P0900Action.action_N01_download_ippan_chosa.handle()関数が異常終了しました。', 'ERROR')
            return 8
            