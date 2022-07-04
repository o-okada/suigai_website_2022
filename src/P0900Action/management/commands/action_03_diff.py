#!/usr/bin/env python3
# -*- coding: utf-8 -*-
###############################################################################
### ファイル名：P0900Action/management/commands/action_03_diff.py
###############################################################################

###############################################################################
### 処理名：インポート処理
###############################################################################
from django.core.management.base import BaseCommand

import sys
from django.contrib.auth.decorators import login_required
from django.db import connection
from django.db import transaction
from django.http import Http404
from django.http import HttpResponse
from django.http import HttpResponseRedirect
from django.shortcuts import render
from django.template import loader
from django.views import generic
from django.views.generic import FormView
from django.views.generic.base import TemplateView

import openpyxl
from openpyxl.comments import Comment
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import PatternFill
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.writer.excel import save_virtual_workbook

from P0000Common.models import BUILDING                ### 1000: 建物区分
from P0000Common.models import KEN                     ### 1010: 都道府県
from P0000Common.models import CITY                    ### 1020: 市区町村
from P0000Common.models import KASEN_KAIGAN            ### 1030: 水害発生地点工種（河川海岸区分）
from P0000Common.models import SUIKEI                  ### 1040: 水系（水系・沿岸）
from P0000Common.models import SUIKEI_TYPE             ### 1050: 水系種別（水系・沿岸種別）
from P0000Common.models import KASEN                   ### 1060: 河川（河川・海岸）
from P0000Common.models import KASEN_TYPE              ### 1070: 河川種別（河川・海岸種別）
from P0000Common.models import CAUSE                   ### 1080: 水害原因
from P0000Common.models import UNDERGROUND             ### 1090: 地上地下区分
from P0000Common.models import USAGE                   ### 1100: 地下空間の利用形態
from P0000Common.models import FLOOD_SEDIMENT          ### 1110: 浸水土砂区分
from P0000Common.models import GRADIENT                ### 1120: 地盤勾配区分
from P0000Common.models import INDUSTRY                ### 1130: 産業分類

from P0000Common.models import HOUSE_ASSET             ### 2000: 家屋評価額
from P0000Common.models import HOUSE_RATE              ### 2010: 家屋被害率
from P0000Common.models import HOUSE_ALT               ### 2020: 家庭応急対策費_代替活動費
from P0000Common.models import HOUSE_CLEAN             ### 2030: 家庭応急対策費_清掃日数

from P0000Common.models import HOUSEHOLD_ASSET         ### 3000: 家庭用品自動車以外所有額
from P0000Common.models import HOUSEHOLD_RATE          ### 3010: 家庭用品自動車以外被害率

from P0000Common.models import CAR_ASSET               ### 4000: 家庭用品自動車所有額
from P0000Common.models import CAR_RATE                ### 4010: 家庭用品自動車被害率

from P0000Common.models import OFFICE_ASSET            ### 5000: 事業所資産額
from P0000Common.models import OFFICE_RATE             ### 5010: 事業所被害率
from P0000Common.models import OFFICE_SUSPEND          ### 5020: 事業所営業停止日数
from P0000Common.models import OFFICE_STAGNATE         ### 5030: 事業所営業停滞日数
from P0000Common.models import OFFICE_ALT              ### 5040: 事業所応急対策費_代替活動費

from P0000Common.models import FARMER_FISHER_ASSET     ### 6000: 農漁家資産額
from P0000Common.models import FARMER_FISHER_RATE      ### 6010: 農漁家被害率

from P0000Common.models import AREA                    ### 7000: 入力データ_水害区域
from P0000Common.models import WEATHER                 ### 7010: 入力データ_異常気象
from P0000Common.models import SUIGAI                  ### 7020: 入力データ_ヘッダ部分
from P0000Common.models import IPPAN                   ### 7030: 入力データ_一覧表部分
from P0000Common.models import IPPAN_VIEW              ### 7040: ビューデータ_一覧表部分

from P0000Common.models import IPPAN_SUMMARY           ### 8000: 集計データ

from P0000Common.models import ACTION                  ### 10000: アクション
from P0000Common.models import STATUS                  ### 10010: 状態
from P0000Common.models import TRIGGER                 ### 10020: トリガーメッセージ
from P0000Common.models import APPROVAL                ### 10030: 承認メッセージ
from P0000Common.models import FEEDBACK                ### 10040: フィードバックメッセージ
### from P0000Common.models import EXECUTE             ### 10050: 実行管理

from P0000Common.models import REPOSITORY              ### 10100: レポジトリ

from P0000Common.common import print_log

###############################################################################
### クラス名： Command
###############################################################################
class Command(BaseCommand):
    
    ###########################################################################
    ### 関数名： handle
    ### チェックOKの場合、トリガーの状態を成功に更新、消費日時をカレントに更新、新たなトリガーを生成する。
    ### チェックNGの場合、トリガーの状態を失敗に更新、消費日時をカレントに更新する。
    ### チェックNGの場合、当該水害IDについて、以降の処理を止めて、手動？で再実行、または、入力データから再登録するイメージである。
    ### 上記は、このアプリの共通の考え方とする。
    ###########################################################################
    def handle(self, *args, **options):
        connection_cursor = connection.cursor()
        try:
            ###################################################################
            ### 引数チェック処理(0000)
            ### コマンドラインからの引数をチェックする。
            ###################################################################
            print_log('[INFO] ########################################', 'INFO')
            print_log('[INFO] P0900Action.action_03_diff.handle()関数が開始しました。', 'INFO')
            print_log('[INFO] P0900Action.action_03_diff.handle()関数 STEP 1/16.', 'INFO')

            ###################################################################
            ### DBアクセス処理(0000)
            ###################################################################
            print_log('[INFO] P0900Action.action_03_diff.handle()関数 STEP 2/16.', 'INFO')
            trigger_list = TRIGGER.objects.raw("""
                SELECT 
                    TR1.trigger_id AS trigger_id, 
                    TR1.suigai_id AS suigai_id, 
                    SG1.ken_code AS ken_code, 
                    SG1.city_code AS city_code, 
                    TR1.repository_id AS repository_id, 
                    RE1.input_file_path AS input_file_path, 
                    TR1.action_code AS action_code, 
                    TR1.status_code AS status_code, 
                    TR1.published_at AS puslished_at, 
                    TR1.consumed_at AS consumed_at, 
                    TR1.success_count AS success_count, 
                    TR1.failure_count AS failure_count, 
                    TR1.deleted_at AS deleted_at 
                FROM TRIGGER TR1 
                LEFT JOIN REPOSITORY RE1 ON TR1.repository_id=RE1.repository_id 
                LEFT JOIN SUIGAI SG1 ON TR1.suigai_id=SG1.suigai_id 
                WHERE 
                    TR1.action_code='3' AND 
                    TR1.consumed_at IS NULL AND 
                    TR1.deleted_at IS NULL 
                ORDER BY CAST(TR1.trigger_id AS INTEGER) LIMIT 1""", [])
            
            print_log('trigger_list = {}'.format(trigger_list), 'INFO')
            
            if trigger_list is None:
                print_log('[INFO] P0900Action.action_03_diff.handle()関数が正常終了しました。', 'INFO')
                return 0
            
            if len(trigger_list) == 0:
                print_log('[INFO] P0900Action.action_03_diff.handle()関数が正常終了しました。', 'INFO')
                return 0

            ###################################################################
            ### DBアクセス処理(0000)
            ### DBから入力データ_ヘッダ部分のデータを取得する。
            ###################################################################
            print_log('[INFO] P0900Action.action_03_diff.handle()関数 STEP 3/16.', 'INFO')
            suigai_list = SUIGAI.objects.raw("""
                SELECT 
                    SG1.suigai_id AS suigai_id,
                    SG1.suigai_name AS suigai_name,
                    SG1.ken_code AS ken_code, 
                    KE1.ken_name AS ken_name,
                    SG1.city_code AS city_code,
                    CI1.city_name AS city_name,
                    SG1.begin_date AS begin_date,
                    SG1.end_date AS end_date,
                    SG1.cause_1_code AS cause_1_code,
                    CA1.cause_name AS cause_1_name,
                    SG1.cause_2_code AS cause_2_code,
                    CA2.cause_name AS cause_2_name,
                    SG1.cause_3_code AS cause_3_code,
                    CA3.cause_name AS cause_3_name,
                    SG1.area_id AS area_id,
                    AR1.area_name AS area_name,
                    SG1.suikei_code AS suikei_code,
                    SK1.suikei_name AS suikei_name,
                    ST1.suikei_type_code AS suikei_type_code,
                    ST1.suikei_type_name AS suikei_type_name,
                    SG1.kasen_code AS kasen_code,
                    KA1.kasen_name AS kasen_name,
                    KT1.kasen_type_code AS kasen_type_code,
                    KT1.kasen_type_name AS kasen_type_name,
                    SG1.gradient_code AS gradient_code,
                    GR1.gradient_name AS gradient_name,
                    SG1.residential_area AS residential_area,
                    SG1.agricultural_area AS agricultural_area,
                    SG1.underground_area AS underground_area,
                    SG1.kasen_kaigan_code AS kasen_kaigan_code,
                    KK1.kasen_kaigan_name AS kasen_kaigan_name,
                    SG1.crop_damage AS crop_damage,
                    SG1.weather_id AS weather_id,
                    WE1.weather_name as weather_name
                FROM SUIGAI SG1 
                LEFT JOIN KEN KE1 ON SG1.ken_code = KE1.ken_code 
                LEFT JOIN CITY CI1 ON SG1.city_code = CI1.city_code 
                LEFT JOIN CAUSE CA1 ON SG1.cause_1_code = CA1.cause_code 
                LEFT JOIN CAUSE CA2 ON SG1.cause_2_code = CA2.cause_code 
                LEFT JOIN CAUSE CA3 ON SG1.cause_3_code = CA3.cause_code 
                LEFT JOIN AREA AR1 ON SG1.area_id = AR1.area_id 
                LEFT JOIN SUIKEI SK1 ON SG1.suikei_code = SK1.suikei_code 
                LEFT JOIN SUIKEI_TYPE ST1 ON SK1.suikei_type_code = ST1.suikei_type_code 
                LEFT JOIN KASEN KA1 ON SG1.kasen_code = KA1.kasen_code 
                LEFT JOIN KASEN_TYPE KT1 ON KA1.kasen_type_code = KT1.kasen_type_code 
                LEFT JOIN GRADIENT GR1 ON SG1.gradient_code = GR1.gradient_code 
                LEFT JOIN KASEN_KAIGAN KK1 ON SG1.kasen_kaigan_code = KK1.kasen_kaigan_code 
                LEFT JOIN WEATHER WE1 ON SG1.weather_id = WE1.weather_id 
                WHERE SG1.suigai_id = %s 
                """, [trigger_list[0].suigai_id, ])

            ###################################################################
            ### DBアクセス処理(0000)
            ### DBから入力データ_一覧表部分のデータを取得する。
            ###################################################################
            print_log('[INFO] P0900Action.action_03_diff.handle()関数 STEP 4/16.', 'INFO')
            ippan_list = IPPAN.objects.raw("""
                SELECT 
                    IV1.ippan_id AS ippan_id,
                    IV1.ippan_name AS ippan_name,
                    IV1.building_code AS building_code,
                    IV1.building_name AS building_Name,
                    IV1.underground_code AS underground_code,
                    IV1.underground_name AS underground_name,
                    IV1.flood_sediment_code AS flood_sediment_code,
                    IV1.flood_sediment_name AS flood_sediment_name,
                    IV1.building_lv00 AS building_lv00,
                    IV1.building_lv01_49 AS building_lv01_49,
                    IV1.building_lv50_99 AS building_lv50_99,
                    IV1.building_lv100 AS building_lv100,
                    IV1.building_half AS building_half,
                    IV1.building_full AS building_full,
                    IV1.floor_area AS floor_area,
                    IV1.family AS family,
                    IV1.office AS office,
                    IV1.floor_area_lv00 AS floor_area_lv00,
                    IV1.floor_area_lv01_49 AS floor_area_lv01_49,
                    IV1.floor_area_lv50_99 AS floor_area_lv50_99,
                    IV1.floor_area_lv100 AS floor_area_lv100,
                    IV1.floor_area_half AS floor_area_half,
                    IV1.floor_area_full AS floor_area_full,
                    IV1.family_lv00 AS family_lv00,
                    IV1.family_lv01_49 AS family_lv01_49,
                    IV1.family_lv50_99 AS family_lv50_99,
                    IV1.family_lv100 AS family_lv100,
                    IV1.family_half AS family_half,
                    IV1.family_full AS family_full,
                    IV1.office_lv00 AS office_lv00,
                    IV1.office_lv01_49 AS office_lv01_49,
                    IV1.office_lv50_99 AS office_lv50_99,
                    IV1.office_lv100 AS office_lv100,
                    IV1.office_half AS office_half,
                    IV1.office_full AS office_full,
                    IV1.farmer_fisher_lv00 AS farmer_fisher_lv00,
                    IV1.farmer_fisher_lv01_49 AS farmer_fisher_lv01_49,
                    IV1.farmer_fisher_lv50_99 AS farmer_fisher_lv50_99,
                    IV1.farmer_fisher_lv100 AS farmer_fisher_lv100,
                    IV1.farmer_fisher_full AS farmer_fisher_full,
                    IV1.employee_lv00 AS employee_lv00,
                    IV1.employee_lv01_49 AS employee_lv01_49,
                    IV1.employee_lv50_99 AS employee_lv50_99,
                    IV1.employee_lv100 AS employee_lv100,
                    IV1.employee_full AS employee_full,
                    IV1.industry_code AS industry_code,
                    IV1.industry_name AS industry_name,
                    IV1.usage_code as usage_code,
                    IV1.usage_name as usage_name,
                    IV1.comment as comment 
                FROM IPPAN_VIEW IV1 
                WHERE IV1.suigai_id = %s
                ORDER BY CAST (IV1.IPPAN_ID AS INTEGER)            
                """, [trigger_list[0].suigai_id, ])

            ###################################################################
            ### EXCELファイル入出力処理(0000)
            ###################################################################
            print_log('[INFO] P0900Action.action_03_diff.handle()関数 STEP 5/16.', 'INFO')
            input_file_path = trigger_list[0].input_file_path
            output_file_path = trigger_list[0].input_file_path + '_output'
            template_file_path = 'static/template_ippan_chosa.xlsx'
            
            wb_input = openpyxl.load_workbook(input_file_path)
            wb_output = openpyxl.load_workbook(template_file_path, keep_vba=False)

            ###############################################################
            ### EXCEL入出力処理(0000)
            ### EXCELのヘッダ部のセルに、DBから取得した入力データ_ヘッダ部分の値を埋め込む。
            ###############################################################
            print_log('[INFO] P0900Action.action_03_diff.handle()関数 STEP 7/16.', 'INFO')
            if suigai_list:
                for i, suigai in enumerate(suigai_list):
                    if suigai.ken_code is None:
                        wb_output["IPPAN"].cell(row=7, column=2).value = None
                    else:
                        wb_output["IPPAN"].cell(row=7, column=2).value = str(suigai.ken_name) + ":" + str(suigai.ken_code)
                    if suigai.city_code is None:
                        wb_output["IPPAN"].cell(row=7, column=3).value = None
                    else:
                        wb_output["IPPAN"].cell(row=7, column=3).value = str(suigai.city_name) + ":" + str(suigai.city_code)
                    wb_output["IPPAN"].cell(row=7, column=4).value = str(suigai.begin_date)
                    wb_output["IPPAN"].cell(row=7, column=5).value = str(suigai.end_date)
                    if suigai.cause_1_code is None:
                        wb_output["IPPAN"].cell(row=7, column=6).value = None
                    else:
                        wb_output["IPPAN"].cell(row=7, column=6).value = str(suigai.cause_1_name) + ":" + str(suigai.cause_1_code)
                    if suigai.cause_2_code is None:
                        wb_output["IPPAN"].cell(row=7, column=7).value = None
                    else:
                        wb_output["IPPAN"].cell(row=7, column=7).value = str(suigai.cause_2_name) + ":" + str(suigai.cause_2_code)
                    if suigai.cause_3_code is None:
                        wb_output["IPPAN"].cell(row=7, column=8).value = None
                    else:
                        wb_output["IPPAN"].cell(row=7, column=8).value = str(suigai.cause_3_name) + ":" + str(suigai.cause_3_code)
                    if suigai.area_id is None:
                        wb_output["IPPAN"].cell(row=7, column=9).value = None
                    else:
                        wb_output["IPPAN"].cell(row=7, column=9).value = str(suigai.area_name) + ":" + str(suigai.area_id)
                    if suigai.suikei_code is None:
                        wb_output["IPPAN"].cell(row=10, column=2).value = None
                    else:
                        wb_output["IPPAN"].cell(row=10, column=2).value = str(suigai.suikei_name) + ":" + str(suigai.suikei_code)
                    if suigai.suikei_type_code is None:
                        wb_output["IPPAN"].cell(row=10, column=3).value = None
                    else:
                        wb_output["IPPAN"].cell(row=10, column=3).value = str(suigai.suikei_type_name) + ":" + str(suigai.suikei_type_code)
                    if suigai.kasen_code is None:
                        wb_output["IPPAN"].cell(row=10, column=4).value = None
                    else:
                        wb_output["IPPAN"].cell(row=10, column=4).value = str(suigai.kasen_name) + ":" + str(suigai.kasen_code)
                    if suigai.kasen_kaigan_code is None:
                        wb_output["IPPAN"].cell(row=10, column=5).value = None
                    else:
                        wb_output["IPPAN"].cell(row=10, column=5).value = str(suigai.kasen_type_name) + ":" + str(suigai.kasen_type_code)
                    if suigai.gradient_code is None:
                        wb_output["IPPAN"].cell(row=10, column=6).value = None
                    else:
                        wb_output["IPPAN"].cell(row=10, column=6).value = str(suigai.gradient_name) + ":" + str(suigai.gradient_code)
                    wb_output["IPPAN"].cell(row=14, column=2).value = str(suigai.residential_area)
                    wb_output["IPPAN"].cell(row=14, column=3).value = str(suigai.agricultural_area)
                    wb_output["IPPAN"].cell(row=14, column=4).value = str(suigai.underground_area)
                    if suigai.kasen_kaigan_code is None:
                        wb_output["IPPAN"].cell(row=14, column=6).value = None
                    else:
                        wb_output["IPPAN"].cell(row=14, column=6).value = str(suigai.kasen_kaigan_name) + ":" + str(suigai.kasen_kaigan_code)
                    wb_output["IPPAN"].cell(row=14, column=8).value = str(suigai.crop_damage)
                    if suigai.weather_id is None:
                        wb_output["IPPAN"].cell(row=14, column=10).value = None
                    else:
                        wb_output["IPPAN"].cell(row=14, column=10).value = str(suigai.weather_name) + ":" + str(suigai.weather_id)

                    wb_output["IPPAN"].cell(row=3, column=28).value = suigai.suigai_id
                    
            ###############################################################
            ### EXCEL入出力処理(0000)
            ### EXCELの一覧部のセルに、DBから取得した入力データ_一覧表部分の値を埋め込む。
            ###############################################################
            print_log('[INFO] P0900Action.action_03_diff.handle()関数 STEP 8/16.', 'INFO')
            if ippan_list:
                for i, ippan in enumerate(ippan_list):
                    wb_output["IPPAN"].cell(row=i+20, column=2).value = str(ippan.ippan_name)
                    if ippan.building_code is None:
                        wb_output["IPPAN"].cell(row=i+20, column=3).value = None
                    else:
                        wb_output["IPPAN"].cell(row=i+20, column=3).value = str(ippan.building_name) + ":" + str(ippan.building_code)
                    if ippan.underground_code is None:
                        wb_output["IPPAN"].cell(row=i+20, column=4).value = None
                    else:
                        wb_output["IPPAN"].cell(row=i+20, column=4).value = str(ippan.underground_name) + ":" + str(ippan.underground_code)
                    if ippan.flood_sediment_code is None:
                        wb_output["IPPAN"].cell(row=i+20, column=5).value = None
                    else:
                        wb_output["IPPAN"].cell(row=i+20, column=5).value = str(ippan.flood_sediment_name) + ":" + str(ippan.flood_sediment_code)
                    wb_output["IPPAN"].cell(row=i+20, column=6).value = ippan.building_lv00
                    wb_output["IPPAN"].cell(row=i+20, column=7).value = ippan.building_lv01_49
                    wb_output["IPPAN"].cell(row=i+20, column=8).value = ippan.building_lv50_99
                    wb_output["IPPAN"].cell(row=i+20, column=9).value = ippan.building_lv100
                    wb_output["IPPAN"].cell(row=i+20, column=10).value = ippan.building_half
                    wb_output["IPPAN"].cell(row=i+20, column=11).value = ippan.building_full
                    wb_output["IPPAN"].cell(row=i+20, column=12).value = ippan.floor_area
                    wb_output["IPPAN"].cell(row=i+20, column=13).value = ippan.family
                    wb_output["IPPAN"].cell(row=i+20, column=14).value = ippan.office
                    wb_output["IPPAN"].cell(row=i+20, column=15).value = ippan.farmer_fisher_lv00
                    wb_output["IPPAN"].cell(row=i+20, column=16).value = ippan.farmer_fisher_lv01_49
                    wb_output["IPPAN"].cell(row=i+20, column=17).value = ippan.farmer_fisher_lv50_99
                    wb_output["IPPAN"].cell(row=i+20, column=18).value = ippan.farmer_fisher_lv100
                    wb_output["IPPAN"].cell(row=i+20, column=19).value = ippan.farmer_fisher_full
                    wb_output["IPPAN"].cell(row=i+20, column=20).value = ippan.employee_lv00
                    wb_output["IPPAN"].cell(row=i+20, column=21).value = ippan.employee_lv01_49
                    wb_output["IPPAN"].cell(row=i+20, column=22).value = ippan.employee_lv50_99
                    wb_output["IPPAN"].cell(row=i+20, column=23).value = ippan.employee_lv100
                    wb_output["IPPAN"].cell(row=i+20, column=24).value = ippan.employee_full
                    if ippan.industry_code is None:
                        wb_output["IPPAN"].cell(row=i+20, column=25).value = None
                    else:
                        wb_output["IPPAN"].cell(row=i+20, column=25).value = str(ippan.industry_name) + ":" + str(ippan.industry_code)
                    if ippan.usage_code is None:
                        wb_output["IPPAN"].cell(row=i+20, column=26).value = None
                    else:
                        wb_output["IPPAN"].cell(row=i+20, column=26).value = str(ippan.usage_name) + ":" + str(ippan.usage_code)
                    wb_output["IPPAN"].cell(row=i+20, column=27).value = ippan.comment

                    wb_output["IPPAN"].cell(row=i+20, column=28).value = ippan.ippan_id

            wb_output.save(output_file_path)
    
            ###################################################################
            ### EXCELファイル入出力処理(0000)
            ###################################################################
            print_log('[INFO] P0900Action.action_03_diff.handle()関数 STEP 9/16.', 'INFO')
            ws_input = []
            ws_output = []
            ws_max_row_input = []
            ws_max_row_output = []
            for ws_temp in wb_input.worksheets:
                if 'IPPAN' in ws_temp.title:
                    ws_input.append(ws_temp)

            for ws_temp in wb_output.worksheets:
                if 'IPPAN' in ws_temp.title:
                    ws_output.append(ws_temp)

            for i, ws_temp in enumerate(ws_input):
                _max_row_ = 1
                for j in range(ws_temp.max_row + 1, 1, -1):
                    if ws_temp.cell(row=j, column=2).value is None:
                        pass
                    else:
                        _max_row_ = j
                        break
                        
                ws_max_row_input.append(_max_row_)

            for i, ws_temp in enumerate(ws_output):
                _max_row_ = 1
                for j in range(ws_temp.max_row + 1, 1, -1):
                    if ws_temp.cell(row=j, column=2).value is None:
                        pass
                    else:
                        _max_row_ = j
                        break
                        
                ws_max_row_output.append(_max_row_)

            ###################################################################
            ### EXCELファイル入出力処理(0000)
            ###################################################################
            print_log('[INFO] P0900Action.action_03_diff.handle()関数 STEP 12/16.', 'INFO')
            OK_list = []
            NG_list = []
            
            for i in [2, 3, 4, 5, 6, 7, 8, 9]:
                if str(ws_input[0].cell(row=7, column=i).value) != str(ws_output[0].cell(row=7, column=i).value):
                    print_log('ws_input[0].cell.value={}'.format(ws_input[0].cell(row=7, column=i).value), 'INFO')
                    print_log('ws_output[0].cell.value={}'.format(ws_output[0].cell(row=7, column=i).value), 'INFO')
                    NG_list.append([ws_input[0].title, ws_output[0].title, 7, i, ws_input[0].cell(row=7, column=i).value, ws_output[0].cell(row=7, column=i).value])
                else:
                    OK_list.append([ws_input[0].title, ws_output[0].title, 7, i, ws_input[0].cell(row=7, column=i).value, ws_output[0].cell(row=7, column=i).value])

            for i in [2, 3, 4, 5, 6]:
                if str(ws_input[0].cell(row=10, column=i).value) != str(ws_output[0].cell(row=10, column=i).value):
                    print_log('ws_input[0].cell.value={}'.format(ws_input[0].cell(row=10, column=i).value), 'INFO')
                    print_log('ws_output[0].cell.value={}'.format(ws_output[0].cell(row=10, column=i).value), 'INFO')
                    NG_list.append([ws_input[0].title, ws_output[0].title, 10, i, ws_input[0].cell(row=10, column=i).value, ws_output[0].cell(row=10, column=i).value])
                else:
                    OK_list.append([ws_input[0].title, ws_output[0].title, 10, i, ws_input[0].cell(row=10, column=i).value, ws_output[0].cell(row=10, column=i).value])

            for i in [2, 3, 4, 6, 8, 10]:
                if str(ws_input[0].cell(row=14, column=i).value) != str(ws_output[0].cell(row=14, column=i).value):
                    print_log('ws_input[0].cell.value={}'.format(ws_input[0].cell(row=14, column=i).value), 'INFO')
                    print_log('ws_output[0].cell.value={}'.format(ws_output[0].cell(row=14, column=i).value), 'INFO')
                    NG_list.append([ws_input[0].title, ws_output[0].title, 14, i, ws_input[0].cell(row=14, column=i).value, ws_output[0].cell(row=14, column=i).value])
                else:
                    OK_list.append([ws_input[0].title, ws_output[0].title, 14, i, ws_input[0].cell(row=14, column=i).value, ws_output[0].cell(row=14, column=i).value])
            
            for i in range(0, 1):
                for j in range(20, 22):
                    for k in range(1, 28):
                        if str(ws_input[i].cell(row=j, column=k).value) != str(ws_output[i].cell(row=j, column=k).value):
                            print_log('ws_input[0].cell.value={}'.format(ws_input[0].cell(row=j, column=k).value), 'INFO')
                            print_log('ws_output[0].cell.value={}'.format(ws_output[0].cell(row=j, column=k).value), 'INFO')
                            NG_list.append([ws_input[i].title, ws_output[i].title, j, k, ws_input[i].cell(row=j, column=k).value, ws_output[i].cell(row=j, column=k).value])
                        else:
                            OK_list.append([ws_input[i].title, ws_output[i].title, j, k, ws_input[i].cell(row=j, column=k).value, ws_output[i].cell(row=j, column=k).value])
                
            ###################################################################
            ### DBアクセス処理(0000)
            ###################################################################
            print_log('[INFO] P0900Action.action_03_diff.handle()関数 STEP 13/16.', 'INFO')
            ### print_log('len(NG_list) = {}'.format(len(NG_list)), 'INFO')
            ### print_log('len(OK_list) = {}'.format(len(OK_list)), 'INFO')
            ### if NG_list:
            ###     if len(NG_list) > 0:
            ###         for i in range(len(NG_list)):
            ###             print_log('NG_list = {}'.format(str(NG_list[i][0])+','+str(NG_list[i][1])+','+str(NG_list[i][2])+','+str(NG_list[i][3])+','+str(NG_list[i][4])+','+str(NG_list[i][5])), 'INFO')
            ### if OK_list:
            ###     if len(OK_list) > 0:
            ###         for i in range(len(OK_list)):
            ###             print_log('OK_list = {}'.format(str(OK_list[i][0])+','+str(OK_list[i][1])+','+str(OK_list[i][2])+','+str(OK_list[i][3])+','+str(OK_list[i][4])+','+str(OK_list[i][5])), 'INFO')

            NG_str = ''
            if len(NG_list) > 0:
                for i in range(len(NG_list)):
                    NG_str = NG_str+str(NG_list[i][0])+','+str(NG_list[i][1])+','+str(NG_list[i][2])+','+str(NG_list[i][3])+','+str(NG_list[i][4])+','+str(NG_list[i][5])+'\n'

            OK_str = ''
            if len(OK_list) > 0:
                for i in range(len(OK_list)):
                    OK_str = OK_str+str(OK_list[i][0])+','+str(OK_list[i][1])+','+str(OK_list[i][2])+','+str(OK_list[i][3])+','+str(OK_list[i][4])+','+str(OK_list[i][5])+'\n'

            print_log('OK_str={}'.format(OK_str), 'INFO')
            print_log('NG_str={}'.format(NG_str), 'INFO')

            ################################################################### 
            ### DBアクセス処理(0000)
            ### 当該トリガーの実行が終了したため、当該トリガーの状態、成功数、失敗数等を更新する。
            ################################################################### 
            print_log('[INFO] P0900Action.action_03_diff.handle()関数 STEP 14/16.', 'INFO')
            if len(NG_list) == 0:
                connection_cursor.execute("""
                    UPDATE TRIGGER SET 
                        status_code='3', 
                        consumed_at=CURRENT_TIMESTAMP, 
                        success_count=%s, 
                        failure_count=%s, 
                        data_integrity=%s 
                    WHERE 
                        trigger_id=%s""", [
                        len(OK_list), 
                        len(NG_list), 
                        OK_str, 
                        trigger_list[0].trigger_id, ])
            else:
                connection_cursor.execute("""
                    UPDATE TRIGGER SET 
                        status_code='4', 
                        consumed_at=CURRENT_TIMESTAMP, 
                        success_count=%s, 
                        failure_count=%s, 
                        data_integrity=%s 
                    WHERE 
                        trigger_id=%s""", [
                        len(OK_list), 
                        len(NG_list), 
                        OK_str, 
                        trigger_list[0].trigger_id, ])

            ################################################################### 
            ### DBアクセス処理(0000)
            ### 当該トリガーの実行が終了したため、
            ### (1)成功の場合は、次のトリガーを発行する。
            ### (2)失敗の場合は、次のトリガーを発行しない。
            ################################################################### 
            print_log('[INFO] P0900Action.action_03_diff.handle()関数 STEP 15/16.', 'INFO')
            if len(NG_list) == 0:
                connection_cursor.execute("""
                    INSERT INTO TRIGGER (
                        trigger_id, 
                        suigai_id, 
                        repository_id, 
                        action_code, 
                        published_at 
                    ) VALUES (
                        (SELECT MAX(trigger_id) + 1 FROM TRIGGER), -- trigger_id
                        %s,                                        -- suigai_id 
                        %s,                                        -- repository_id
                        '5',                                       -- action_code
                        CURRENT_TIMESTAMP                          -- published_at
                    )""", [
                        trigger_list[0].suigai_id, 
                        trigger_list[0].repository_id, ])

            ###################################################################
            ### 戻り値セット処理(0000)
            ###################################################################
            print_log('[INFO] P0900Action.action_03_diff.handle()関数 STEP 16/16.', 'INFO')
            print_log('[INFO] P0900Action.action_03_diff.handle()関数が正常終了しました。', 'INFO')
            return 0
        
        except:
            print_log(sys.exc_info()[0], 'ERROR')
            print_log('[ERROR] P0900Action.action_03_diff.handle()関数でエラーが発生しました。', 'ERROR')
            print_log('[ERROR] P0900Action.action_03_diff.handle()関数が異常終了しました。', 'ERROR')
            return 8

        finally:
            connection_cursor.close()
            