#!/usr/bin/env python3
# -*- coding: utf-8 -*-
###############################################################################
### ファイル名：P0900Action/management/commands/action_S01_download_koeki.py
### S01：ダウンロード
###############################################################################

###############################################################################
### 処理名：インポート処理
###############################################################################
from django.core.management.base import BaseCommand

import sys
from django.contrib.auth.decorators import login_required
from django.db import connection
from django.db import transaction
from django.http import Http404
from django.http import HttpResponse
from django.http import HttpResponseRedirect
from django.shortcuts import render
from django.template import loader
from django.views import generic
from django.views.generic import FormView
from django.views.generic.base import TemplateView

import openpyxl
from openpyxl.comments import Comment
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import PatternFill
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.writer.excel import save_virtual_workbook

from decimal import Decimal

from P0000Common.models import BUILDING                ### 1000: 建物区分
from P0000Common.models import KEN                     ### 1010: 都道府県
from P0000Common.models import CITY                    ### 1020: 市区町村
from P0000Common.models import KASEN_KAIGAN            ### 1030: 水害発生地点工種（河川海岸区分）
from P0000Common.models import SUIKEI                  ### 1040: 水系（水系・沿岸）
from P0000Common.models import SUIKEI_TYPE             ### 1050: 水系種別（水系・沿岸種別）
from P0000Common.models import KASEN                   ### 1060: 河川（河川・海岸）
from P0000Common.models import KASEN_TYPE              ### 1070: 河川種別（河川・海岸種別）
from P0000Common.models import CAUSE                   ### 1080: 水害原因
from P0000Common.models import UNDERGROUND             ### 1090: 地上地下区分
from P0000Common.models import USAGE                   ### 1100: 地下空間の利用形態
from P0000Common.models import FLOOD_SEDIMENT          ### 1110: 浸水土砂区分
from P0000Common.models import GRADIENT                ### 1120: 地盤勾配区分
from P0000Common.models import INDUSTRY                ### 1130: 産業分類

from P0000Common.models import HOUSE_ASSET             ### 2000: 家屋評価額
from P0000Common.models import HOUSE_RATE              ### 2010: 家屋被害率
from P0000Common.models import HOUSE_ALT               ### 2020: 家庭応急対策費_代替活動費
from P0000Common.models import HOUSE_CLEAN             ### 2030: 家庭応急対策費_清掃日数

from P0000Common.models import HOUSEHOLD_ASSET         ### 3000: 家庭用品自動車以外所有額
from P0000Common.models import HOUSEHOLD_RATE          ### 3010: 家庭用品自動車以外被害率

from P0000Common.models import CAR_ASSET               ### 4000: 家庭用品自動車所有額
from P0000Common.models import CAR_RATE                ### 4010: 家庭用品自動車被害率

from P0000Common.models import OFFICE_ASSET            ### 5000: 事業所資産額
from P0000Common.models import OFFICE_RATE             ### 5010: 事業所被害率
from P0000Common.models import OFFICE_SUSPEND          ### 5020: 事業所営業停止日数
from P0000Common.models import OFFICE_STAGNATE         ### 5030: 事業所営業停滞日数
from P0000Common.models import OFFICE_ALT              ### 5040: 事業所応急対策費_代替活動費

from P0000Common.models import FARMER_FISHER_ASSET     ### 6000: 農漁家資産額
from P0000Common.models import FARMER_FISHER_RATE      ### 6010: 農漁家被害率

from P0000Common.models import AREA                    ### 7000: 入力データ_水害区域
from P0000Common.models import WEATHER                 ### 7010: 入力データ_異常気象
from P0000Common.models import SUIGAI                  ### 7020: 入力データ_ヘッダ部分
from P0000Common.models import IPPAN                   ### 7030: 入力データ_一覧表部分
from P0000Common.models import IPPAN_VIEW              ### 7040: ビューデータ_一覧表部分

from P0000Common.models import IPPAN_SUMMARY           ### 8000: 集計データ

from P0000Common.models import ACTION                  ### 10000: アクション
from P0000Common.models import STATUS                  ### 10010: 状態
from P0000Common.models import TRIGGER                 ### 10020: トリガーメッセージ

from P0000Common.common import get_debug_log
from P0000Common.common import get_error_log
from P0000Common.common import get_info_log
from P0000Common.common import get_warn_log
from P0000Common.common import print_log
from P0000Common.common import reset_log

VLOOK_VALUE = [
    'B', 'G', 'L', 'Q', 'V', 'AA', 'AF', 'AK', 'AP', 'AU', 
    'AZ', 'BE', 'BJ', 'BO', 'BT', 'BY', 'CD', 'CI', 'CN', 'CS', 
    'CX', 'DC', 'DH', 'DM', 'DR', 'DW', 'EB', 'EG', 'EL', 'EQ', 
    'EV', 'FA', 'FF', 'FK', 'FP', 'FU', 'FZ', 'GE', 'GJ', 'GO', 
    'GT', 'GY', 'HD', 'HI', 'HN', 'HS', 'HX', 'IC', 'IH', 'IM', 
    'IR', 'IW', 'JB', 'JG', 'JL', 'JQ', 'JV', 'KA', 'KF', 'KK', 
    'KP', 'KU', 'KZ', 'LE', 'LJ', 'LO', 'LT', 'LY', 'MD', 'MI', 
    'MN', 'MS', 'MX', 'NC', 'NH', 'NM', 'NR', 'NW', 'OB', 'OG', 
    'OL', 'OQ', 'OV', 'PA', 'PF', 'PK', 'PP', 'PU', 'PZ', 'QE', 
    'QJ', 'QO', 'QT', 'QY', 'RD', 'RI', 'RN', 'RS', 'RX', 'SC', 
    'SH', 'SM', 'SR', 'SW', 'TB', 'TG', 'TL', 'TQ', 'TV', 'UA', 
    'UF', 'UK', 'UP', 'UU', 'UZ', 'VE', 'VJ', 'VO', 'VT', 'VY', 
    'WD', 'WI', 'WN', 'WS', 'WX', 'XC', 'XH', 'XM', 'XR', 'XW', 
    'YB', 'YG', 'YL', 'YQ', 'YV', 'ZA', 'ZF', 'ZK', 'ZP', 'ZU', 
    'ZZ', 'AAE', 'AAJ', 'AAO', 'AAT', 'AAY', 'ABD', 'ABI', 'ABN', 'ABS', 
    'ABX', 'ACC', 'ACH', 'ACM', 'ACR', 'ACW', 'ADB', 'ADG', 'ADL', 'ADQ', 
    'ADV', 'AEA', 'AEF', 'AEK', 'AEP', 'AEU', 'AEZ', 'AFE', 'AFJ', 'AFO', 
    'AFT', 'AFY', 'AGD', 'AGI', 'AGN', 'AGS', 'AGX', 'AHC', 'AHH', 'AHM', 
    'AHR', 'AHW', 'AIB', 'AIG', 'AIL', 'AIQ', 'AIV', 'AJA', 'AJF', 'AJK', 
    'AJP', 'AJU', 'AJZ', 'AKE', 'AKJ', 'AKO', 'AKT', 'AKY', 'ALD', 'ALI', 
    'ALN', 'ALS', 'ALX', 'AMC', 'AMH', 'AMM', 'AMR', 'AMW', 'ANB', 'ANG', 
    'ANL', 'ANQ', 'ANV', 'AOA', 'AOF', 'AOK', 'AOP', 'AOU', 'AOZ', 'APE', 
    'APJ', 'APO', 'APT', 'APY', 'AQD', 'AQI', 'AQN', 'AQS', 'AQX', 'ARC', 
    'ARH', 'ARM', 'ARR', 'ARW', 'ASB', 'ASG', 'ASL', 'ASQ', 'ASV', 'ATA', 
    'ATF', 'ATK', 'ATP', 'ATU', 'ATZ', 'AUE', 'AUJ', 'AUO', 'AUT', 'AUY', 
    'AVD', 'AVI', 'AVN', 'AVS', 'AVX', 'AWC', 'AWH', 'AWM', 'AWR', 'AWW', 
    'AXB', 'AXG', 'AXL', 'AXQ', 'AXV', 'AYA', 'AYF', 'AYK', 'AYP', 'AYU', 
    'AYZ', 'AZE', 'AZJ', 'AZO', 'AZT', 'AZY'
    ]

###############################################################################
### クラス名： Command
###############################################################################
class Command(BaseCommand):
    
    ###########################################################################
    ### 関数名： handle
    ### チェックOKの場合、トリガーの状態を成功に更新、消費日時をカレントに更新、新たなトリガーを生成する。
    ### チェックNGの場合、トリガーの状態を失敗に更新、消費日時をカレントに更新する。
    ### チェックNGの場合、当該水害IDについて、以降の処理を止めて、手動？で再実行、または、入力データから再登録するイメージである。
    ### 上記は、このアプリの共通の考え方とする。
    ###########################################################################
    def handle(self, *args, **options):
        try:
            ###################################################################
            ### 引数チェック処理(0000)
            ### コマンドラインからの引数をチェックする。
            ###################################################################
            reset_log()
            print_log('[INFO] P0900Action.action_S01_download_koeki.handle()関数が開始しました。', 'INFO')
            print_log('[DEBUG] P0900Action.action_S01_download_koeki.handle()関数 STEP 1/13.', 'DEBUG')

            ###################################################################
            ### DBアクセス処理(0010)
            ### TRIGGERテーブルからSUIGAI_IDのリストを取得する。
            ### トリガーメッセージにアクションが発行されているかを検索する。
            ### 処理対象の水害IDを取得する。
            ### トリガーメッセージにアクションが発行されていなければ、処理を終了する。
            ### ※ネストを浅くするために、処理対象外の場合、処理を終了させる。
            ###################################################################
            print_log('[DEBUG] P0900Action.action_S01_download_koeki.handle()関数 STEP 2/13.', 'DEBUG')
            trigger_list = TRIGGER.objects.raw("""
                SELECT 
                    * 
                FROM TRIGGER 
                WHERE 
                    action_code='S01' AND 
                    consumed_at IS NULL AND 
                    deleted_at IS NULL 
                ORDER BY CAST(ken_code AS INTEGER), CAST(city_code AS INTEGER) LIMIT 1""", [])
            
            if trigger_list is None:
                print_log('[INFO] P0900Action.action_S01_download_koeki.handle()関数が正常終了しました。', 'INFO')
                return 0
            
            if len(trigger_list) == 0:
                print_log('[INFO] P0900Action.action_S01_download_koeki.handle()関数が正常終了しました。', 'INFO')
                return 0
            
            print_log('[DEBUG] P0900Action.action_S01_download_koeki.handle()関数 trigger_list[0].trigger_id = {}'.format(trigger_list[0].trigger_id), 'DEBUG')
            print_log('[DEBUG] P0900Action.action_S01_download_koeki.handle()関数 trigger_list[0].suigai_id = {}'.format(trigger_list[0].suigai_id), 'DEBUG')
            print_log('[DEBUG] P0900Action.action_S01_download_koeki.handle()関数 trigger_list[0].city_code = {}'.format(trigger_list[0].city_code), 'DEBUG')
            print_log('[DEBUG] P0900Action.action_S01_download_koeki.handle()関数 trigger_list[0].ken_code = {}'.format(trigger_list[0].ken_code), 'DEBUG')
            print_log('[DEBUG] P0900Action.action_S01_download_koeki.handle()関数 trigger_list[0].download_file_path = {}'.format(trigger_list[0].download_file_path), 'DEBUG')
            print_log('[DEBUG] P0900Action.action_S01_download_koeki.handle()関数 trigger_list[0].download_file_name = {}'.format(trigger_list[0].download_file_name), 'DEBUG')
            print_log('[DEBUG] P0900Action.action_S01_download_koeki.handle()関数 trigger_list[0].upload_file_path = {}'.format(trigger_list[0].upload_file_path), 'DEBUG')
            print_log('[DEBUG] P0900Action.action_S01_download_koeki.handle()関数 trigger_list[0].upload_file_name = {}'.format(trigger_list[0].upload_file_name), 'DEBUG')
    
            ###################################################################
            ### DBアクセス処理(0020)
            ### マスタデータを取得する。
            ###################################################################
            print_log('[DEBUG] P0900Action.action_S01_download_koeki.handle()関数 STEP 3/13.', 'DEBUG')
            ### 1010: 都道府県シート
            ### 1020: 市区町村シート、連動プルダウン用、VLOOKUP用
            ### 1030: 水害発生地点工種（河川海岸区分）
            ### 1040: 水系（水系・沿岸）
            ### 1050: 水系種別（水系・沿岸種別）
            ### 1060: 河川（河川・海岸）、連動プルダウン用、VLOOKUP用
            ### 1070: 河川種別（河川・海岸種別）
            ### 1080: 水害原因
            ### 7010: 入力データ_異常気象
            cities_list = []
            ken_list = KEN.objects.raw("""SELECT * FROM KEN ORDER BY CAST(KEN_CODE AS INTEGER)""", [])
            if ken_list:
                for i, ken in enumerate(ken_list):
                    cities_list.append(CITY.objects.raw("""SELECT * FROM CITY WHERE KEN_CODE=%s ORDER BY CAST(CITY_CODE AS INTEGER)""", [ken.ken_code,]))
    
            kasen_kaigan_list = KASEN_KAIGAN.objects.raw("""SELECT * FROM KASEN_KAIGAN ORDER BY CAST(KASEN_KAIGAN_CODE AS INTEGER)""", [])
            suikei_list = SUIKEI.objects.raw("""SELECT * FROM SUIKEI ORDER BY CAST(SUIKEI_CODE AS INTEGER)""", [])
            suikei_type_list = SUIKEI_TYPE.objects.raw("""SELECT * FROM SUIKEI_TYPE ORDER BY CAST(SUIKEI_TYPE_CODE AS INTEGER)""", [])
    
            kasens_list = []
            suikei_list = SUIKEI.objects.raw("""SELECT * FROM SUIKEI ORDER BY CAST(SUIKEI_CODE AS INTEGER)""", [])
            if suikei_list:
                for i, suikei in enumerate(suikei_list):
                    kasens_list.append(KASEN.objects.raw("""SELECT * FROM KASEN WHERE SUIKEI_CODE=%s ORDER BY CAST(KASEN_CODE AS INTEGER)""", [suikei.suikei_code,]))
    
            kasen_type_list = KASEN_TYPE.objects.raw("""SELECT * FROM KASEN_TYPE ORDER BY CAST(KASEN_TYPE_CODE AS INTEGER)""", [])
            cause_list = CAUSE.objects.raw("""SELECT * FROM CAUSE ORDER BY CAST(CAUSE_CODE AS INTEGER)""", [])
            weather_list = WEATHER.objects.raw("""SELECT * FROM WEATHER ORDER BY CAST(WEATHER_ID AS INTEGER)""", [])
            
            ###################################################################
            ### EXCEL入出力処理(0030)
            ### ダウンロード用ファイルパスをセットする。
            ###################################################################
            print_log('[DEBUG] P0900Action.action_S01_download_koeki.handle()関数 STEP 4/13.', 'DEBUG')
            template_file_path = 'static/template_koeki.xlsx'
            download_file_path = trigger_list[0].download_file_path
            download_file_name = trigger_list[0].download_file_name
            print_log('[DEBUG] P0900Action.action_S01_download_koeki.handle()関数 download_file_path = {}'.format(download_file_path), 'DEBUG')
            print_log('[DEBUG] P0900Action.action_S01_download_koeki.handle()関数 download_file_name = {}'.format(download_file_name), 'DEBUG')

            ###################################################################
            ### EXCEL入出力処理(0040)
            ### ワークシートを局所変数にセットする。
            ###################################################################
            print_log('[DEBUG] P0900Action.action_S01_download_koeki.handle()関数 STEP 5/13.', 'DEBUG')
            wb = openpyxl.load_workbook(template_file_path, keep_vba=False)
            
            ws_ken = wb["KEN"]
            ws_city = wb["CITY"]
            ws_kasen_kaigan = wb["KASEN_KAIGAN"]
            ws_suikei = wb["SUIKEI"]
            ws_suikei_type = wb["SUIKEI_TYPE"]
            ws_kasen = wb["KASEN"]
            ws_kasen_type = wb["KASEN_TYPE"]
            ws_cause = wb["CAUSE"]
            ws_weather = wb["WEATHER"]
            
            ws_city_vlook = wb["CITY_VLOOK"]
            ws_kasen_vlook = wb["KASEN_VLOOK"]

            ws_koeki = wb["KOEKI"]

            ###################################################################
            ### EXCEL入出力処理(0050)
            ### 各EXCELシートで枠線を表示しないにセットする。
            ###################################################################
            print_log('[DEBUG] P0900Action.action_S01_download_koeki.handle()関数 STEP 6/13.', 'DEBUG')
            ws_ken.sheet_view.showGridLines = False
            ws_city.sheet_view.showGridLines = False
            ws_kasen_kaigan.sheet_view.showGridLines = False
            ws_suikei.sheet_view.showGridLines = False
            ws_suikei_type.sheet_view.showGridLines = False
            ws_kasen.sheet_view.showGridLines = False
            ws_kasen_type.sheet_view.showGridLines = False
            ws_cause.sheet_view.showGridLines = False
            ws_weather.sheet_view.showGridLines = False
            
            ws_city_vlook.sheet_view.showGridLines = False
            ws_kasen_vlook.sheet_view.showGridLines = False
    
            ws_koeki.sheet_view.showGridLines = False

            ###################################################################
            ### EXCEL入出力処理(0060)
            ### マスタ用のシートに値をセットする。
            ###################################################################
            print_log('[DEBUG] P0900Action.action_S01_download_koeki.handle()関数 STEP 7/13.', 'DEBUG')
            ### 1010: 都道府県シート
            print("handle_7_2", flush=True)
            if ken_list:
                for i, ken in enumerate(ken_list):
                    ws_ken.cell(row=i+1, column=1).value = ken.ken_code
                    ws_ken.cell(row=i+1, column=2).value = str(ken.ken_name) + ":" + str(ken.ken_code)
                    ### ws_city_vlook.cell(row=j+1, column=1).value = str(ken.ken_name) + ":" + str(ken.ken_code)

            ### 1020: 市区町村シート
            print("handle_7_3", flush=True)
            cities_list = []
            if ken_list:
                for i, ken in enumerate(ken_list):
                    cities_list.append(CITY.objects.raw("""
                        SELECT 
                            * 
                        FROM CITY 
                        WHERE ken_code=%s 
                        ORDER BY CAST(city_code AS INTEGER)""", [ken.ken_code, ]))
    
            print("handle_7_4", flush=True)
            if cities_list:
                for i, cities in enumerate(cities_list):
                    if cities:
                        for j, city in enumerate(cities):
                            ws_city.cell(row=j+1, column=i*5+1).value = city.city_code
                            ws_city.cell(row=j+1, column=i*5+2).value = str(city.city_name) + ":" + str(city.city_code)
                            ws_city.cell(row=j+1, column=i*5+3).value = city.ken_code
                            ws_city.cell(row=j+1, column=i*5+4).value = city.city_population
                            ws_city.cell(row=j+1, column=i*5+5).value = city.city_area

            ### 1030: 水害発生地点工種（河川海岸区分）
            print("handle_7_5", flush=True)
            if kasen_kaigan_list:
                for i, kasen_kaigan in enumerate(kasen_kaigan_list):
                    ws_kasen_kaigan.cell(row=i+1, column=1).value = kasen_kaigan.kasen_kaigan_code
                    ws_kasen_kaigan.cell(row=i+1, column=2).value = str(kasen_kaigan.kasen_kaigan_name) + ":" + str(kasen_kaigan.kasen_kaigan_code)

            ### 1040: 水系（水系・沿岸）
            print("handle_7_6", flush=True)
            if suikei_list:
                for i, suikei in enumerate(suikei_list):
                    ws_suikei.cell(row=i+1, column=1).value = suikei.suikei_code
                    ws_suikei.cell(row=i+1, column=2).value = str(suikei.suikei_name) + ":" + str(suikei.suikei_code)
                    ws_suikei.cell(row=i+1, column=3).value = suikei.suikei_type_code

            ### 1050: 水系種別（水系・沿岸種別）
            print("handle_7_7", flush=True)
            if suikei_type_list:
                for i, suikei_type in enumerate(suikei_type_list):
                    ws_suikei_type.cell(row=i+1, column=1).value = suikei_type.suikei_type_code
                    ws_suikei_type.cell(row=i+1, column=2).value = str(suikei_type.suikei_type_name) + ":" + str(suikei_type.suikei_type_code)

            ### 1060: 河川（河川・海岸）、連動プルダウン用
            print("handle_7_8", flush=True)
            kasens_list = []
            if suikei_list:
                for i, suikei in enumerate(suikei_list):
                    kasens_list.append(KASEN.objects.raw("""
                        SELECT 
                            * 
                        FROM KASEN 
                        WHERE suikei_code=%s 
                        ORDER BY CAST(kasen_code AS INTEGER)""", [suikei.suikei_code, ]))
    
            print("handle_7_9", flush=True)
            if kasens_list:
                for i, kasens in enumerate(kasens_list):
                    if kasens:
                        for j, kasen in enumerate(kasens):
                            ws_kasen.cell(row=j+1, column=i*5+1).value = kasen.kasen_code
                            ws_kasen.cell(row=j+1, column=i*5+2).value = str(kasen.kasen_name) + ":" + str(kasen.kasen_code)
                            ws_kasen.cell(row=j+1, column=i*5+3).value = kasen.kasen_type_code
                            ws_kasen.cell(row=j+1, column=i*5+4).value = kasen.suikei_code

            ### 1070: 河川種別（河川・海岸種別）
            print("handle_7_10", flush=True)
            if kasen_type_list:
                for i, kasen_type in enumerate(kasen_type_list):
                    ws_kasen_type.cell(row=i+1, column=1).value = kasen_type.kasen_type_code
                    ws_kasen_type.cell(row=i+1, column=2).value = str(kasen_type.kasen_type_name) + ":" + str(kasen_type.kasen_type_code)
            
            ### 1080: 水害原因
            print("handle_7_11", flush=True)
            if cause_list:
                for i, cause in enumerate(cause_list):
                    ws_cause.cell(row=i+1, column=1).value = cause.cause_code
                    ws_cause.cell(row=i+1, column=2).value = str(cause.cause_name) + ":" + str(cause.cause_code)

            ### 7010: 入力データ_異常気象
            print("handle_7_18", flush=True)
            if weather_list:
                for i, weather in enumerate(weather_list):
                    ws_weather.cell(row=i+1, column=1).value = weather.weather_id
                    ws_weather.cell(row=i+1, column=2).value = str(weather.weather_name) + ":" + str(weather.weather_id)

            ###################################################################
            ### EXCEL入出力処理(0070)
            ### VLOOKUP用のシートに値をセットする。
            ###################################################################
            print_log('[DEBUG] P0900Action.action_S01_download_koeki.handle()関数 STEP 8/13.', 'DEBUG')
            ### 1020: 市区町村VLOOKUP
            print("handle_8_1", flush=True)
            if ken_list and cities_list:
                for i, ken in enumerate(ken_list):
                    ws_city_vlook.cell(row=i+1, column=1).value = str(ken.ken_name) + ":" + str(ken.ken_code)
        
                for i, cities in enumerate(cities_list):
                    ws_city_vlook.cell(row=i+1, column=2).value = 'CITY!$' + VLOOK_VALUE[i] + '$1:$' + VLOOK_VALUE[i] + '$%d' % len(cities)
    
            ### 1060: 河川（河川・海岸）VLOOKUP
            print("handle_8_2", flush=True)
            if suikei_list and kasens_list:
                for i, suikei in enumerate(suikei_list):
                    ws_kasen_vlook.cell(row=i+1, column=1).value = str(suikei.suikei_name) + ":" + str(suikei.suikei_code)
    
                for i, kasens in enumerate(kasens_list):
                    ws_kasen_vlook.cell(row=i+1, column=2).value = 'KASEN!$' + VLOOK_VALUE[i] + '$1:$' + VLOOK_VALUE[i] + '$%d' % len(kasens)

            ###################################################################
            ### EXCEL入出力処理(0080)
            ### 入力データ用EXCELシートのキャプションに値をセットする。
            ###################################################################
            print_log('[DEBUG] P0900Action.action_S01_download_koeki.handle()関数 STEP 9/13.', 'DEBUG')
            ### ws_ippan.cell(row=6, column=1).value = 'NO.'
            ### ws_ippan.cell(row=6, column=2).value = '水害発生年月日'
            ### ws_ippan.cell(row=6, column=3).value = ''
            ### ws_ippan.cell(row=6, column=4).value = ''
            ### ws_ippan.cell(row=6, column=5).value = ''
            ### ws_ippan.cell(row=6, column=6).value = '被害箇所'
            ### ws_ippan.cell(row=6, column=7).value = ''
            ### ws_ippan.cell(row=6, column=8).value = ''
            ### ws_ippan.cell(row=6, column=9).value = ''
            ### ws_ippan.cell(row=6, column=10).value = ''
            ### ws_ippan.cell(row=6, column=11).value = '河川・海岸名・地区名'
            ### ws_ippan.cell(row=6, column=12).value = ''
            ### ws_ippan.cell(row=6, column=13).value = ''
            ### ws_ippan.cell(row=6, column=14).value = ''
            ### ws_ippan.cell(row=6, column=15).value = '工種区分'
            ### ws_ippan.cell(row=6, column=16).value = '水害原因コード'
            ### ws_ippan.cell(row=6, column=17).value = ''
            ### ws_ippan.cell(row=6, column=18).value = ''
            ### ws_ippan.cell(row=6, column=19).value = '異常気象コード'
            ### ws_ippan.cell(row=6, column=20).value = '事業コード'
            ### ws_ippan.cell(row=6, column=21).value = '調査対象機関名称[全角]'
            ### ws_ippan.cell(row=6, column=22).value = '被害金額'
            ### ws_ippan.cell(row=6, column=23).value = ''
            ### ws_ippan.cell(row=6, column=24).value = ''
            ### ws_ippan.cell(row=6, column=25).value = ''
            ### ws_ippan.cell(row=6, column=26).value = ''
            ### ws_ippan.cell(row=6, column=27).value = '営業停止期間等'
            ### ws_ippan.cell(row=6, column=28).value = ''
            ### ws_ippan.cell(row=6, column=29).value = ''
            ### ws_ippan.cell(row=6, column=30).value = '照会先'
            ### ws_ippan.cell(row=6, column=31).value = ''
            ### ws_ippan.cell(row=6, column=32).value = ''
            ### ws_ippan.cell(row=6, column=33).value = '備考'
            
            ### ws_ippan.cell(row=7, column=1).value = ''
            ### ws_ippan.cell(row=7, column=2).value = ''
            ### ws_ippan.cell(row=7, column=3).value = ''
            ### ws_ippan.cell(row=7, column=4).value = ''
            ### ws_ippan.cell(row=7, column=5).value = ''
            ### ws_ippan.cell(row=7, column=6).value = '都道府県[全角]'
            ### ws_ippan.cell(row=7, column=7).value = '都道府県コード'
            ### ws_ippan.cell(row=7, column=8).value = '市区町村名[全角]'
            ### ws_ippan.cell(row=7, column=9).value = '市区町村コード'
            ### ws_ippan.cell(row=7, column=10).value = '町丁名・大字名[全角]'
            ### ws_ippan.cell(row=7, column=11).value = '水系・沿岸名[全角]'
            ### ws_ippan.cell(row=7, column=12).value = '水系種別[全角]'
            ### ws_ippan.cell(row=7, column=13).value = '河川・海岸名[全角]'
            ### ws_ippan.cell(row=7, column=14).value = '河川種別[全角]'
            ### ws_ippan.cell(row=7, column=15).value = ''
            ### ws_ippan.cell(row=7, column=16).value = ''
            ### ws_ippan.cell(row=7, column=17).value = ''
            ### ws_ippan.cell(row=7, column=18).value = ''
            ### ws_ippan.cell(row=7, column=19).value = ''
            ### ws_ippan.cell(row=7, column=20).value = ''
            ### ws_ippan.cell(row=7, column=21).value = ''
            ### ws_ippan.cell(row=7, column=22).value = '物的被害額(千円)'
            ### ws_ippan.cell(row=7, column=23).value = '営業停止損失額(千円)'
            ### ws_ippan.cell(row=7, column=24).value = ''
            ### ws_ippan.cell(row=7, column=25).value = ''
            ### ws_ippan.cell(row=7, column=26).value = '営業停止損失額合計(千円)'
            ### ws_ippan.cell(row=7, column=27).value = '停止期間'
            ### ws_ippan.cell(row=7, column=28).value = ''
            ### ws_ippan.cell(row=7, column=29).value = '停止数量'
            ### ws_ippan.cell(row=7, column=30).value = '調査担当課名[全角]'
            ### ws_ippan.cell(row=7, column=31).value = '調査担当者名[全角]'
            ### ws_ippan.cell(row=7, column=32).value = '電話番号'
            ### ws_ippan.cell(row=7, column=33).value = '備考'

            ### ws_ippan.cell(row=8, column=1).value = ''
            ### ws_ippan.cell(row=8, column=2).value = '月'
            ### ws_ippan.cell(row=8, column=3).value = '日'
            ### ws_ippan.cell(row=8, column=4).value = '月'
            ### ws_ippan.cell(row=8, column=5).value = '日'
            ### ws_ippan.cell(row=8, column=6).value = ''
            ### ws_ippan.cell(row=8, column=7).value = ''
            ### ws_ippan.cell(row=8, column=8).value = ''
            ### ws_ippan.cell(row=8, column=9).value = ''
            ### ws_ippan.cell(row=8, column=10).value = ''
            ### ws_ippan.cell(row=8, column=11).value = ''
            ### ws_ippan.cell(row=8, column=12).value = ''
            ### ws_ippan.cell(row=8, column=13).value = ''
            ### ws_ippan.cell(row=8, column=14).value = ''
            ### ws_ippan.cell(row=8, column=15).value = ''
            ### ws_ippan.cell(row=8, column=16).value = '1'
            ### ws_ippan.cell(row=8, column=17).value = '2'
            ### ws_ippan.cell(row=8, column=18).value = '3'
            ### ws_ippan.cell(row=8, column=19).value = ''
            ### ws_ippan.cell(row=8, column=20).value = ''
            ### ws_ippan.cell(row=8, column=21).value = ''
            ### ws_ippan.cell(row=8, column=22).value = ''
            ### ws_ippan.cell(row=8, column=23).value = '営業停止に伴う売上減少額'
            ### ws_ippan.cell(row=8, column=24).value = '代替活動費（外注費）'
            ### ws_ippan.cell(row=8, column=25).value = 'その他'
            ### ws_ippan.cell(row=8, column=26).value = ''
            ### ws_ippan.cell(row=8, column=27).value = '日'
            ### ws_ippan.cell(row=8, column=28).value = '時間'
            ### ws_ippan.cell(row=8, column=29).value = ''
            ### ws_ippan.cell(row=8, column=30).value = ''
            ### ws_ippan.cell(row=8, column=31).value = ''
            ### ws_ippan.cell(row=8, column=32).value = ''
            ### ws_ippan.cell(row=8, column=33).value = ''
            
            ws_ippan.protection.enable()
            green_fill = PatternFill(fgColor='CCFFCC', patternType='solid')

            ###################################################################
            ### EXCEL入出力処理(0090)
            ### EXCELのセルに、値を埋め込む。
            ###################################################################
            print_log('[DEBUG] P0900Action.action_S01_download_koeki.handle()関数 STEP 10/13.', 'DEBUG')
            koeki_list = KOEKI.objects.raw("""
                SELECT 
                    KO1.koeki_id AS koeki_id,
                    KO1.koeki_name AS koeki_name,
                    KO1.koeki_file_id AS koeki_file_id, 
                    KF1.upload_file_path AS upload_file_path, 
                    KF1.upload_file_name AS upload_file_name, 
                    KF1.summary_file_path AS summary_file_path, 
                    KF1.summary_file_name AS summary_file_name, 
                    KO1.ken_code AS ken_code, 
                    KE1.ken_name AS ken_name, 
                    KO1.city_code AS city_code, 
                    CT1.city_name AS city_name, 
                    KO1.suikei_code AS suikei_code, 
                    SK1.suikei_name AS suikei_name, 
                    KO1.kasen_code AS kasen_code, 
                    KA1.kasen_name AS kasen_name, 
                    KO1.begin_date AS begin_date, 
                    KO1.end_date AS end_date, 
                    KO1.kasen_kaigan_code AS kasen_kaigan_code, 
                    KK1.kasen_kaigan_name AS kasen_kaigan_name, 
                    KO1.weather_id AS weather_id, 
                    WE1.weather_name AS weather_name, 
                    KO1.cause_1_code AS cause_1_code, 
                    CA1.cause_name AS cause_1_name, 
                    KO1.cause_2_code AS cause_2_code, 
                    CA2.cause_name AS cause_2_name, 
                    KO1.cause_3_code AS cause_3_code, 
                    CA1.cause_name AS cause_3_name, 
                    KO1.business_code AS business_code, 
                    KO1.institution_name AS institution_name, 
                    KO1.damage_property AS damage_property, 
                    KO1.damage_sales AS damage_sales,
                    KO1.damage_alt AS damage_alt, 
                    KO1.damage_other AS damage_other, 
                    KO1.damage_total AS damage_total, 
                    KO1.sus_day AS sus_day, 
                    KO1.sus_hour AS sus_hour, 
                    KO1.sus_amount AS sus_amount, 
                    KO1.ref_dep_name AS ref_dep_name, 
                    KO1.ref_emp_name AS ref_emp_name, 
                    KO1.ref_tel AS ref_tel, 
                    KO1.comment AS comment 
                FROM KOEKI KO1 
                LEFT JOIN KOEKI_FILE KF1 ON KO1.koeki_file_id=KF1.koeki_file_id 
                LEFT JOIN KEN KE1 ON KO1.ken_code=KE1.ken_code 
                LEFT JOIN CITY CT1 ON KO1.city_code=CT1.city_code 
                LEFT JOIN SUIKEI SK1 ON KO1.suikei_code=SK1.suikei_code 
                LEFT JOIN KASEN KA1 ON KO1.kasen_code=KA1.kasen_code 
                LEFT JOIN KASEN_KAIGAN KK1 ON KO1.kasen_kaigan_code=KK1.kasen_kaigan_code 
                LEFT JOIN WEATHER WE1 ON KO1.weather_id=WE1.weather_id 
                LEFT JOIN CAUSE CA1 ON KO1.cause_1_code=CA1.cause_code 
                LEFT JOIN CAUSE CA2 ON KO1.cause_2_code=CA2.cause_code 
                LEFT JOIN CAUSE CA3 ON KO1.cause_3_code=CA3.cause_code 
                WHERE 
                    KO1.ken_code=%s AND KO1.deleted_at IS NULL 
                ORDER BY CAST(KO1.city_code AS INTEGER), CAST(KF1.koeki_file_id AS INTEGER), CAST(KO1.koeki_id AS INTEGER)""", [
                    trigger_list[0].ken_code, 
                ])
    
            if koeki_list:
                for i, koeki in enumerate(koeki_list):
                    if koeki.koeki_id is not None and len(str(koeki.koeki_id)) > 0:
                        ws_koeki.cell(row=9+i, column=1).value = str(koeki.koeki_id)
                        ws_koeki.cell(row=9+i, column=1).fill = green_fill
                    
                    if koeki.begin_month is not None and len(str(koeki.begin_month)) > 0:
                        ws_koeki.cell(row=9+i, column=2).value = str(koeki.begin_month)

                    if koeki.begin_day is not None and len(str(koeki.begin_day)) > 0:
                        ws_koeki.cell(row=9+i, column=3).value = str(koeki.begin_day)
                    
                    if koeki.end_month is not None and len(str(koeki.end_month)) > 0:
                        ws_koeki.cell(row=9+i, column=4).value = str(koeki.end_month)
                        
                    if koeki.end_day is not None and len(str(koeki.end_day)) > 0:
                        ws_koeki.cell(row=9+i, column=5).value = str(koeki.end_day)
                        
                    if koeki.ken_name is not None and koeki.ken_code is not None and len(str(koeki.ken_name)) > 0 and len(str(koeki.ken_code)) > 0:
                        ws_koeki.cell(row=9+i, column=6).value = str(koeki.ken_name) + ":" + str(koeki.ken_code)
                    
                    if koeki.ken_code is not None and len(str(koeki.ken_code)) > 0:
                        ws_koeki.cell(row=9+i, column=7).value = str(koeki.ken_code)

                    if koeki.city_name is not None and koeki.city_code is not None and len(str(koeki.city_name)) > 0 and len(str(koeki.city_code)) > 0:
                        ws_koeki.cell(row=9+i, column=8).value = str(koeki.city_name) + ":" + str(koeki.city_code)

                    if koeki.city_code is not None and len(str(koeki.city_code)) > 0:
                        ws_koeki.cell(row=9+i, column=9).value = str(koeki.city_code)
                    
                    if koeki.koeki_name is not None and len(str(koeki.koeki_name)) > 0:
                        ws_koeki.cell(row=9+i, column=10).value = str(koeki.koeki_name)
                    
                    if koeki.suikei_name is not None and koeki.suikey_code is not None and len(str(koeki.suikei_name)) > 0 and len(str(koeki.suikei_code)) > 0:
                        ws_koeki.cell(row=9+i, column=11).value = str(koeki.suikei_name) + ":" + str(koeki.suikei_code)

                    if koeki.suikei_type_name is not None and koeki.suikei_type_code is not None and len(str(koeki.suikei_type_name)) > 0 and len(str(koeki.suikei_type_code)) > 0
                        ws_koeki.cell(row=9+i, column=12).value = str(koeki.suikei_type_name) + ":" + str(koeki.suikei_type_code)
                    
                    if koeki.kasen_name is not None and koeki.kasen_code is not None and len(str(koeki.kasen_name)) > 0 and len(str(koeki.kasen_code)) > 0:
                        ws_koeki.cell(row=9+i, column=13).value = str(koeki.kasen_name) + ":" + str(koeki.kasen_code)
                        
                    if koeki.kasen_type_name is not None and koeki.kasen_type_code is not None and len(str(koeki.kasen_type_name)) > 0 and len(str(koeki.kasen_type_code)) > 0:
                        ws_koeki.cell(row=9+i, column=14).value = str(koeki.kasen_type_name) + ":" + str(koeki.kasen_type_code)

                    if koeki.kasen_kaigan_name is not None and koeki.kasen_kaigan_code is not None and len(str(koeki.kasen_kaigan_name)) > 0 and len(str(koeki.kasen_kaigan_code)) > 0:
                        ws_koeki.cell(row=9+i, column=15).value = str(koeki.kasen_kaigan_name) + ":" + str(koeki.kasen_kaigan_code)
                    
                    if koeki.cause_1_name is not None and koeki.cause_1_code is not None and len(str(koeki.cause_1_name)) > 0 and len(str(koeki.cause_1_code)) > 0:
                        ws_koeki.cell(row=9+i, column=16).value = str(koeki.cause_1_name) + ":" + str(koeki.cause_1_code)

                    if koeki.cause_2_name is not None and koeki.cause_2_code is not None and len(str(koeki.cause_2_name)) > 0 and len(str(koeki.cause_2_code)) > 0:
                        ws_koeki.cell(row=9+i, column=17).value = str(koeki.cause_2_name) + ":" + str(koeki.cause_2_code)

                    if koeki.cause_3_name is not None and koeki.cause_3_code is not None and len(str(koeki.cause_3_name)) > 0 and len(str(koeki.cause_3_code)) > 0:
                        ws_koeki.cell(row=9+i, column=18).value = str(koeki.cause_3_name) + ":" + str(koeki.cause_3_code)

                    if koeki.weather_name is not None and koeki.weather_id is not None and len(str(koeki.weather_name)) > 0 and len(str(koeki.weather_id)) > 0:
                        ws_koeki.cell(row=9+i, column=19).value = str(koeki.weather_name) + ":" + str(koeki.weather_id)

                    if koeki.business_name is not None and koeki.business_code is not None and len(str(koeki.business_name)) > 0 and len(str(koeki.business_code)) > 0:
                        ws_koeki.cell(row=9+i, column=20).value = str(koeki.business_name) + ":" + str(koeki.business_code)

                    if koeki.institution_name is not None and len(str(koeki.institution_name)) > 0:
                        ws_koeki.cell(row=9+i, column=21).value = str(koeki.institution_name)

                    if koeki.damage_property is not None and len(str(koeki.damage_property)) > 0:
                        ws_koeki.cell(row=9+i, column=22).value = str(koeki.damage_property)

                    if koeki.damage_sales is not None and len(str(koeki.damage_sales)) > 0:
                        ws_koeki.cell(row=9+i, column=23).value = str(koeki.damage_sales)

                    if koeki.damage_alt is not None and len(str(koeki.damage_alt)) > 0:
                        ws_koeki.cell(row=9+i, column=24).value = str(koeki.damage_alt)

                    if koeki.damage_other is not None and len(str(koeki.damage_other)) > 0:
                        ws_koeki.cell(row=9+i, column=25).value = str(koeki.damage_other)

                    if koeki.damage_total is not None and len(str(koeki.damage_total)) > 0:
                        ws_koeki.cell(row=9+i, column=26).value = str(koeki.damage_total)

                    if koeki.sus_day is not None and len(str(koeki.sus_day)) > 0:
                        ws_koeki.cell(row=9+i, column=27).value = str(koeki.sus_day)

                    if koeki.sus_hour is not None and len(str(koeki.sus_hour)) > 0:
                        ws_koeki.cell(row=9+i, column=28).value = str(koeki.sus_hour)

                    if koeki.sus_amount is not None and len(str(koeki.sus_amount)) > 0:
                        ws_koeki.cell(row=9+i, column=29).value = str(koeki.sus_amount)

                    if koeki.ref_dep_name is not None and len(str(koeki.ref_dep_name)) > 0:
                        ws_koeki.cell(row=9+i, column=30).value = str(koeki.ref_dep_name)

                    if koeki.ref_emp_name is not None and len(str(koeki.ref_emp_name)) > 0:
                        ws_koeki.cell(row=9+i, column=31).value = str(koeki.ref_emp_name)

                    if koeki.ref_tel is not None and len(str(koeki.ref_tel)) > 0:
                        ws_koeki.cell(row=9+i, column=32).value = str(koeki.ref_tel)
                    
                    if koeki.comment is not None and len(str(koeki.comment)) > 0:
                        ws_koeki.cell(row=9+i, column=33).value = str(koeki.comment)

            ###################################################################
            ### EXCEL入出力処理(0100)
            ### ダウンロード用のEXCELファイルを保存する。
            ###################################################################
            print_log('[DEBUG] P0900Action.action_S01_download_koeki()関数 STEP 11/13.', 'DEBUG')
            wb.save(download_file_path)

            ###################################################################
            ### DBアクセス処理(0110)
            ### トリガーデータを更新する。
            ###################################################################
            print_log('[DEBUG] P0900Action.action_S01_download_koeki.handle()関数 STEP 12/13.', 'DEBUG')
            connection_cursor = connection.cursor()
            try:
                connection_cursor.execute("""BEGIN""", [])
                
                connection_cursor.execute("""
                    UPDATE TRIGGER SET 
                        status_code=%s, -- status_code
                        success_count=%s, -- success_count
                        failure_count=%s, -- failure_count
                        consumed_at=CURRENT_TIMESTAMP, 
                        integrity_ok=%s, -- integrity_ok
                        integrity_ng=%s  -- integrity_ng
                    WHERE
                        trigger_id=%s -- trigger_id
                    """, [
                        'SUCCESS', ### status_count
                        1, ### success_count
                        0, ### failure_count
                        '\n'.join(get_info_log()), ### integrity_ok
                        '\n'.join(get_warn_log()), ### integrity_ng
                        trigger_list[0].trigger_id, ### trigger_id
                    ])
                ### transaction.commit()    
                connection_cursor.execute("""COMMIT""", [])
            except:
                print_log('[ERROR] P0900Action.action_S01_download_koeki.handle()関数 {}'.format(sys.exc_info()[0]), 'ERROR')
                ### connection_cursor.rollback()
                connection_cursor.execute("""ROLLBACK""", [])
            finally:
                connection_cursor.close()
                
            ###################################################################
            ### 戻り値セット処理(0120)
            ### 戻り値を戻す。
            ###################################################################
            print_log('[DEBUG] P0900Action.action_S01_download_koeki.handle()関数 STEP 13/13.', 'DEBUG')
            print_log('[INFO] P0900Action.action_S01_download_koeki.handle()関数が正常終了しました。', 'INFO')
            return 0
        except:
            print_log('[ERROR] P0900Action.action_S01_download_koeki.handle()関数 {}'.format(sys.exc_info()[0]), 'ERROR')
            print_log('[ERROR] P0900Action.action_S01_download_koeki.handle()関数でエラーが発生しました。', 'ERROR')
            print_log('[ERROR] P0900Action.action_S01_download_koeki.handle()関数が異常終了しました。', 'ERROR')
            return 8
            