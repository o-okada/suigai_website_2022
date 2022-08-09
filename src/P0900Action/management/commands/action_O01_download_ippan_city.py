#!/usr/bin/env python3
# -*- coding: utf-8 -*-
###############################################################################
### ファイル名：P0900Action/management/commands/action_O01_download_ippan_city.py
### C01：ダウンロード
###############################################################################

###############################################################################
### 処理名：インポート処理
###############################################################################
from django.core.management.base import BaseCommand

import sys
from django.contrib.auth.decorators import login_required
from django.db import connection
from django.db import transaction
from django.http import Http404
from django.http import HttpResponse
from django.http import HttpResponseRedirect
from django.shortcuts import render
from django.template import loader
from django.views import generic
from django.views.generic import FormView
from django.views.generic.base import TemplateView

import openpyxl
from openpyxl.comments import Comment
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import PatternFill
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.writer.excel import save_virtual_workbook

from decimal import Decimal

from P0000Common.models import BUILDING                ### 1000: 建物区分
from P0000Common.models import KEN                     ### 1010: 都道府県
from P0000Common.models import CITY                    ### 1020: 市区町村
from P0000Common.models import KASEN_KAIGAN            ### 1030: 水害発生地点工種（河川海岸区分）
from P0000Common.models import SUIKEI                  ### 1040: 水系（水系・沿岸）
from P0000Common.models import SUIKEI_TYPE             ### 1050: 水系種別（水系・沿岸種別）
from P0000Common.models import KASEN                   ### 1060: 河川（河川・海岸）
from P0000Common.models import KASEN_TYPE              ### 1070: 河川種別（河川・海岸種別）
from P0000Common.models import CAUSE                   ### 1080: 水害原因
from P0000Common.models import UNDERGROUND             ### 1090: 地上地下区分
from P0000Common.models import USAGE                   ### 1100: 地下空間の利用形態
from P0000Common.models import FLOOD_SEDIMENT          ### 1110: 浸水土砂区分
from P0000Common.models import GRADIENT                ### 1120: 地盤勾配区分
from P0000Common.models import INDUSTRY                ### 1130: 産業分類

from P0000Common.models import HOUSE_ASSET             ### 2000: 家屋評価額
from P0000Common.models import HOUSE_RATE              ### 2010: 家屋被害率
from P0000Common.models import HOUSE_ALT               ### 2020: 家庭応急対策費_代替活動費
from P0000Common.models import HOUSE_CLEAN             ### 2030: 家庭応急対策費_清掃日数

from P0000Common.models import HOUSEHOLD_ASSET         ### 3000: 家庭用品自動車以外所有額
from P0000Common.models import HOUSEHOLD_RATE          ### 3010: 家庭用品自動車以外被害率

from P0000Common.models import CAR_ASSET               ### 4000: 家庭用品自動車所有額
from P0000Common.models import CAR_RATE                ### 4010: 家庭用品自動車被害率

from P0000Common.models import OFFICE_ASSET            ### 5000: 事業所資産額
from P0000Common.models import OFFICE_RATE             ### 5010: 事業所被害率
from P0000Common.models import OFFICE_SUSPEND          ### 5020: 事業所営業停止日数
from P0000Common.models import OFFICE_STAGNATE         ### 5030: 事業所営業停滞日数
from P0000Common.models import OFFICE_ALT              ### 5040: 事業所応急対策費_代替活動費

from P0000Common.models import FARMER_FISHER_ASSET     ### 6000: 農漁家資産額
from P0000Common.models import FARMER_FISHER_RATE      ### 6010: 農漁家被害率

from P0000Common.models import AREA                    ### 7000: 入力データ_水害区域
from P0000Common.models import WEATHER                 ### 7010: 入力データ_異常気象
from P0000Common.models import SUIGAI                  ### 7020: 入力データ_ヘッダ部分
from P0000Common.models import IPPAN                   ### 7030: 入力データ_一覧表部分
from P0000Common.models import IPPAN_VIEW              ### 7040: ビューデータ_一覧表部分

from P0000Common.models import IPPAN_SUMMARY           ### 8000: 集計データ

from P0000Common.models import ACTION                  ### 10000: アクション
from P0000Common.models import STATUS                  ### 10010: 状態
from P0000Common.models import TRIGGER                 ### 10020: トリガーメッセージ

from P0000Common.common import get_debug_log
from P0000Common.common import get_error_log
from P0000Common.common import get_info_log
from P0000Common.common import get_warn_log
from P0000Common.common import print_log
from P0000Common.common import reset_log

VLOOK_VALUE = [
    'B', 'G', 'L', 'Q', 'V', 'AA', 'AF', 'AK', 'AP', 'AU', 
    'AZ', 'BE', 'BJ', 'BO', 'BT', 'BY', 'CD', 'CI', 'CN', 'CS', 
    'CX', 'DC', 'DH', 'DM', 'DR', 'DW', 'EB', 'EG', 'EL', 'EQ', 
    'EV', 'FA', 'FF', 'FK', 'FP', 'FU', 'FZ', 'GE', 'GJ', 'GO', 
    'GT', 'GY', 'HD', 'HI', 'HN', 'HS', 'HX', 'IC', 'IH', 'IM', 
    'IR', 'IW', 'JB', 'JG', 'JL', 'JQ', 'JV', 'KA', 'KF', 'KK', 
    'KP', 'KU', 'KZ', 'LE', 'LJ', 'LO', 'LT', 'LY', 'MD', 'MI', 
    'MN', 'MS', 'MX', 'NC', 'NH', 'NM', 'NR', 'NW', 'OB', 'OG', 
    'OL', 'OQ', 'OV', 'PA', 'PF', 'PK', 'PP', 'PU', 'PZ', 'QE', 
    'QJ', 'QO', 'QT', 'QY', 'RD', 'RI', 'RN', 'RS', 'RX', 'SC', 
    'SH', 'SM', 'SR', 'SW', 'TB', 'TG', 'TL', 'TQ', 'TV', 'UA', 
    'UF', 'UK', 'UP', 'UU', 'UZ', 'VE', 'VJ', 'VO', 'VT', 'VY', 
    'WD', 'WI', 'WN', 'WS', 'WX', 'XC', 'XH', 'XM', 'XR', 'XW', 
    'YB', 'YG', 'YL', 'YQ', 'YV', 'ZA', 'ZF', 'ZK', 'ZP', 'ZU', 
    'ZZ', 'AAE', 'AAJ', 'AAO', 'AAT', 'AAY', 'ABD', 'ABI', 'ABN', 'ABS', 
    'ABX', 'ACC', 'ACH', 'ACM', 'ACR', 'ACW', 'ADB', 'ADG', 'ADL', 'ADQ', 
    'ADV', 'AEA', 'AEF', 'AEK', 'AEP', 'AEU', 'AEZ', 'AFE', 'AFJ', 'AFO', 
    'AFT', 'AFY', 'AGD', 'AGI', 'AGN', 'AGS', 'AGX', 'AHC', 'AHH', 'AHM', 
    'AHR', 'AHW', 'AIB', 'AIG', 'AIL', 'AIQ', 'AIV', 'AJA', 'AJF', 'AJK', 
    'AJP', 'AJU', 'AJZ', 'AKE', 'AKJ', 'AKO', 'AKT', 'AKY', 'ALD', 'ALI', 
    'ALN', 'ALS', 'ALX', 'AMC', 'AMH', 'AMM', 'AMR', 'AMW', 'ANB', 'ANG', 
    'ANL', 'ANQ', 'ANV', 'AOA', 'AOF', 'AOK', 'AOP', 'AOU', 'AOZ', 'APE', 
    'APJ', 'APO', 'APT', 'APY', 'AQD', 'AQI', 'AQN', 'AQS', 'AQX', 'ARC', 
    'ARH', 'ARM', 'ARR', 'ARW', 'ASB', 'ASG', 'ASL', 'ASQ', 'ASV', 'ATA', 
    'ATF', 'ATK', 'ATP', 'ATU', 'ATZ', 'AUE', 'AUJ', 'AUO', 'AUT', 'AUY', 
    'AVD', 'AVI', 'AVN', 'AVS', 'AVX', 'AWC', 'AWH', 'AWM', 'AWR', 'AWW', 
    'AXB', 'AXG', 'AXL', 'AXQ', 'AXV', 'AYA', 'AYF', 'AYK', 'AYP', 'AYU', 
    'AYZ', 'AZE', 'AZJ', 'AZO', 'AZT', 'AZY'
    ]

###############################################################################
### クラス名： Command
###############################################################################
class Command(BaseCommand):
    
    ###########################################################################
    ### 関数名： handle
    ### チェックOKの場合、トリガーの状態を成功に更新、消費日時をカレントに更新、新たなトリガーを生成する。
    ### チェックNGの場合、トリガーの状態を失敗に更新、消費日時をカレントに更新する。
    ### チェックNGの場合、当該水害IDについて、以降の処理を止めて、手動？で再実行、または、入力データから再登録するイメージである。
    ### 上記は、このアプリの共通の考え方とする。
    ###########################################################################
    def handle(self, *args, **options):
        try:
            ###################################################################
            ### 引数チェック処理(0000)
            ### コマンドラインからの引数をチェックする。
            ###################################################################
            reset_log()
            print_log('[INFO] P0900Action.action_O01_download_ippan_city.handle()関数が開始しました。', 'INFO')
            print_log('[DEBUG] P0900Action.action_O01_download_ippan_city.handle()関数 STEP 1/13.', 'DEBUG')

            ###################################################################
            ### DBアクセス処理(0010)
            ### TRIGGERテーブルからSUIGAI_IDのリストを取得する。
            ### トリガーメッセージにアクションが発行されているかを検索する。
            ### 処理対象の水害IDを取得する。
            ### トリガーメッセージにアクションが発行されていなければ、処理を終了する。
            ### ※ネストを浅くするために、処理対象外の場合、処理を終了させる。
            ###################################################################
            print_log('[DEBUG] P0900Action.action_O01_download_ippan_city.handle()関数 STEP 2/13.', 'DEBUG')
            trigger_list = TRIGGER.objects.raw("""
                SELECT 
                    * 
                FROM TRIGGER 
                WHERE 
                    action_code='O01' AND 
                    consumed_at IS NULL AND 
                    deleted_at IS NULL 
                ORDER BY CAST(ken_code AS INTEGER), CAST(city_code AS INTEGER) LIMIT 1""", [])
            
            if trigger_list is None:
                print_log('[INFO] P0900Action.action_O01_download_ippan_city.handle()関数が正常終了しました。', 'INFO')
                return 0
            
            if len(trigger_list) == 0:
                print_log('[INFO] P0900Action.action_O01_download_ippan_city.handle()関数が正常終了しました。', 'INFO')
                return 0
            
            print_log('[DEBUG] P0900Action.action_O01_download_ippan_city.handle()関数 trigger_list[0].trigger_id = {}'.format(trigger_list[0].trigger_id), 'DEBUG')
            print_log('[DEBUG] P0900Action.action_O01_download_ippan_city.handle()関数 trigger_list[0].suigai_id = {}'.format(trigger_list[0].suigai_id), 'DEBUG')
            print_log('[DEBUG] P0900Action.action_O01_download_ippan_city.handle()関数 trigger_list[0].city_code = {}'.format(trigger_list[0].city_code), 'DEBUG')
            print_log('[DEBUG] P0900Action.action_O01_download_ippan_city.handle()関数 trigger_list[0].ken_code = {}'.format(trigger_list[0].ken_code), 'DEBUG')
            print_log('[DEBUG] P0900Action.action_O01_download_ippan_city.handle()関数 trigger_list[0].download_file_path = {}'.format(trigger_list[0].download_file_path), 'DEBUG')
            print_log('[DEBUG] P0900Action.action_O01_download_ippan_city.handle()関数 trigger_list[0].download_file_name = {}'.format(trigger_list[0].download_file_name), 'DEBUG')
            print_log('[DEBUG] P0900Action.action_O01_download_ippan_city.handle()関数 trigger_list[0].upload_file_path = {}'.format(trigger_list[0].upload_file_path), 'DEBUG')
            print_log('[DEBUG] P0900Action.action_O01_download_ippan_city.handle()関数 trigger_list[0].upload_file_name = {}'.format(trigger_list[0].upload_file_name), 'DEBUG')
    
            ###################################################################
            ### DBアクセス処理(0020)
            ### マスタデータを取得する。
            ###################################################################
            print_log('[DEBUG] P0900Action.action_O01_download_ippan_city.handle()関数 STEP 3/13.', 'DEBUG')
            ### 1000: 建物区分シート
            ### 1010: 都道府県シート
            ### 1020: 市区町村シート、連動プルダウン用、VLOOKUP用
            ### 1030: 水害発生地点工種（河川海岸区分）
            ### 1040: 水系（水系・沿岸）
            ### 1050: 水系種別（水系・沿岸種別）
            ### 1060: 河川（河川・海岸）、連動プルダウン用、VLOOKUP用
            ### 1070: 河川種別（河川・海岸種別）
            ### 1080: 水害原因
            ### 1090: 地上地下区分
            ### 1100: 地下空間の利用形態
            ### 1110: 浸水土砂区分
            ### 1120: 地盤勾配区分
            ### 1130: 産業分類
            ### 7000: 入力データ_水害区域
            ### 7010: 入力データ_異常気象
            ### 7020: 入力データ_ヘッダ部分、水害
            building_list = BUILDING.objects.raw("""SELECT * FROM BUILDING ORDER BY CAST(BUILDING_CODE AS INTEGER)""", [])
            
            cities_list = []
            ken_list = KEN.objects.raw("""SELECT * FROM KEN ORDER BY CAST(KEN_CODE AS INTEGER)""", [])
            if ken_list:
                for i, ken in enumerate(ken_list):
                    cities_list.append(CITY.objects.raw("""SELECT * FROM CITY WHERE KEN_CODE=%s ORDER BY CAST(CITY_CODE AS INTEGER)""", [ken.ken_code,]))
    
            kasen_kaigan_list = KASEN_KAIGAN.objects.raw("""SELECT * FROM KASEN_KAIGAN ORDER BY CAST(KASEN_KAIGAN_CODE AS INTEGER)""", [])
            suikei_list = SUIKEI.objects.raw("""SELECT * FROM SUIKEI ORDER BY CAST(SUIKEI_CODE AS INTEGER)""", [])
            suikei_type_list = SUIKEI_TYPE.objects.raw("""SELECT * FROM SUIKEI_TYPE ORDER BY CAST(SUIKEI_TYPE_CODE AS INTEGER)""", [])
    
            kasens_list = []
            suikei_list = SUIKEI.objects.raw("""SELECT * FROM SUIKEI ORDER BY CAST(SUIKEI_CODE AS INTEGER)""", [])
            if suikei_list:
                for i, suikei in enumerate(suikei_list):
                    kasens_list.append(KASEN.objects.raw("""SELECT * FROM KASEN WHERE SUIKEI_CODE=%s ORDER BY CAST(KASEN_CODE AS INTEGER)""", [suikei.suikei_code,]))
    
            kasen_type_list = KASEN_TYPE.objects.raw("""SELECT * FROM KASEN_TYPE ORDER BY CAST(KASEN_TYPE_CODE AS INTEGER)""", [])
            cause_list = CAUSE.objects.raw("""SELECT * FROM CAUSE ORDER BY CAST(CAUSE_CODE AS INTEGER)""", [])
            underground_list = UNDERGROUND.objects.raw("""SELECT * FROM UNDERGROUND ORDER BY CAST(UNDERGROUND_CODE AS INTEGER)""", [])
            usage_list = USAGE.objects.raw("""SELECT * FROM USAGE ORDER BY CAST(USAGE_CODE AS INTEGER)""", [])
            flood_sediment_list = FLOOD_SEDIMENT.objects.raw("""SELECT * FROM FLOOD_SEDIMENT ORDER BY CAST(FLOOD_SEDIMENT_CODE AS INTEGER)""", [])
            gradient_list = GRADIENT.objects.raw("""SELECT * FROM GRADIENT ORDER BY CAST(GRADIENT_CODE AS INTEGER)""", [])
            industry_list = INDUSTRY.objects.raw("""SELECT * FROM INDUSTRY ORDER BY CAST(INDUSTRY_CODE AS INTEGER)""", [])
            area_list = AREA.objects.raw("""SELECT * FROM AREA ORDER BY CAST(AREA_ID AS INTEGER)""", [])
            weather_list = WEATHER.objects.raw("""SELECT * FROM WEATHER ORDER BY CAST(WEATHER_ID AS INTEGER)""", [])
            suigai_list = SUIGAI.objects.raw("""SELECT * FROM SUIGAI ORDER BY CAST(SUIGAI_ID AS INTEGER)""", [])
            
            ###################################################################
            ### EXCEL入出力処理(0030)
            ### ダウンロード用ファイルパスをセットする。
            ###################################################################
            print_log('[DEBUG] P0900Action.action_O01_download_ippan_city.handle()関数 STEP 4/13.', 'DEBUG')
            template_file_path = 'static/template_ippan_city.xlsx'
            ### download_file_path = 'static/ippan_chosa_' + str(ken_name_request) + '_' + str(city_name_request) + '.xlsx'
            ### download_file_name = 'ippan_chosa_' + str(ken_name_request) + '_' + str(city_name_request) + '.xlsx'
            download_file_path = trigger_list[0].download_file_path
            download_file_name = trigger_list[0].download_file_name
            print_log('[DEBUG] P0900Action.action_O01_download_ippan_city.handle()関数 download_file_path = {}'.format(download_file_path), 'DEBUG')
            print_log('[DEBUG] P0900Action.action_O01_download_ippan_city.handle()関数 download_file_name = {}'.format(download_file_name), 'DEBUG')

            ###################################################################
            ### EXCEL入出力処理(0040)
            ### ワークシートを局所変数にセットする。
            ###################################################################
            print_log('[DEBUG] P0900Action.action_O01_download_ippan_city.handle()関数 STEP 5/13.', 'DEBUG')
            wb = openpyxl.load_workbook(template_file_path, keep_vba=False)
            
            ws_building = wb["BUILDING"]
            ws_ken = wb["KEN"]
            ws_city = wb["CITY"]
            ws_kasen_kaigan = wb["KASEN_KAIGAN"]
            ws_suikei = wb["SUIKEI"]
            ws_suikei_type = wb["SUIKEI_TYPE"]
            ws_kasen = wb["KASEN"]
            ws_kasen_type = wb["KASEN_TYPE"]
            ws_cause = wb["CAUSE"]
            ws_underground = wb["UNDERGROUND"]
            ws_usage = wb["USAGE"]
            ws_flood_sediment = wb["FLOOD_SEDIMENT"]
            ws_gradient = wb["GRADIENT"]
            ws_industry = wb["INDUSTRY"]
            ws_area = wb["AREA"]
            ws_weather = wb["WEATHER"]
            ws_suigai = wb["SUIGAI"]
            
            ws_city_vlook = wb["CITY_VLOOK"]
            ws_kasen_vlook = wb["KASEN_VLOOK"]

            ws_ippan = wb["IPPAN"]

            ###################################################################
            ### EXCEL入出力処理(0050)
            ### 各EXCELシートで枠線を表示しないにセットする。
            ###################################################################
            print_log('[DEBUG] P0900Action.action_O01_download_ippan_city.handle()関数 STEP 6/13.', 'DEBUG')
            ws_building.sheet_view.showGridLines = False
            ws_ken.sheet_view.showGridLines = False
            ws_city.sheet_view.showGridLines = False
            ws_kasen_kaigan.sheet_view.showGridLines = False
            ws_suikei.sheet_view.showGridLines = False
            ws_suikei_type.sheet_view.showGridLines = False
            ws_kasen.sheet_view.showGridLines = False
            ws_kasen_type.sheet_view.showGridLines = False
            ws_cause.sheet_view.showGridLines = False
            ws_underground.sheet_view.showGridLines = False
            ws_usage.sheet_view.showGridLines = False
            ws_flood_sediment.sheet_view.showGridLines = False
            ws_gradient.sheet_view.showGridLines = False
            ws_industry.sheet_view.showGridLines = False
            ws_area.sheet_view.showGridLines = False
            ws_weather.sheet_view.showGridLines = False
            ws_suigai.sheet_view.showGridLines = False
            
            ws_city_vlook.sheet_view.showGridLines = False
            ws_kasen_vlook.sheet_view.showGridLines = False
    
            ws_ippan.sheet_view.showGridLines = False

            ###################################################################
            ### EXCEL入出力処理(0060)
            ### マスタ用のシートに値をセットする。
            ###################################################################
            print_log('[DEBUG] P0900Action.action_O01_download_ippan_city.handle()関数 STEP 7/13.', 'DEBUG')
            ### 1000: 建物区分シート
            print("handle_7_1", flush=True)
            if building_list:
                for i, building in enumerate(building_list):
                    ws_building.cell(row=i+1, column=1).value = building.building_code
                    ws_building.cell(row=i+1, column=2).value = str(building.building_name) + ":" + str(building.building_code)

            ### 1010: 都道府県シート
            print("handle_7_2", flush=True)
            if ken_list:
                for i, ken in enumerate(ken_list):
                    ws_ken.cell(row=i+1, column=1).value = ken.ken_code
                    ws_ken.cell(row=i+1, column=2).value = str(ken.ken_name) + ":" + str(ken.ken_code)
                    ### ws_city_vlook.cell(row=j+1, column=1).value = str(ken.ken_name) + ":" + str(ken.ken_code)

            ### 1020: 市区町村シート
            print("handle_7_3", flush=True)
            cities_list = []
            if ken_list:
                for i, ken in enumerate(ken_list):
                    cities_list.append(CITY.objects.raw("""
                        SELECT 
                            * 
                        FROM CITY 
                        WHERE ken_code=%s 
                        ORDER BY CAST(city_code AS INTEGER)""", [ken.ken_code, ]))
    
            print("handle_7_4", flush=True)
            if cities_list:
                for i, cities in enumerate(cities_list):
                    if cities:
                        for j, city in enumerate(cities):
                            ws_city.cell(row=j+1, column=i*5+1).value = city.city_code
                            ws_city.cell(row=j+1, column=i*5+2).value = str(city.city_name) + ":" + str(city.city_code)
                            ws_city.cell(row=j+1, column=i*5+3).value = city.ken_code
                            ws_city.cell(row=j+1, column=i*5+4).value = city.city_population
                            ws_city.cell(row=j+1, column=i*5+5).value = city.city_area

            ### 1030: 水害発生地点工種（河川海岸区分）
            print("handle_7_5", flush=True)
            if kasen_kaigan_list:
                for i, kasen_kaigan in enumerate(kasen_kaigan_list):
                    ws_kasen_kaigan.cell(row=i+1, column=1).value = kasen_kaigan.kasen_kaigan_code
                    ws_kasen_kaigan.cell(row=i+1, column=2).value = str(kasen_kaigan.kasen_kaigan_name) + ":" + str(kasen_kaigan.kasen_kaigan_code)

            ### 1040: 水系（水系・沿岸）
            print("handle_7_6", flush=True)
            if suikei_list:
                for i, suikei in enumerate(suikei_list):
                    ws_suikei.cell(row=i+1, column=1).value = suikei.suikei_code
                    ws_suikei.cell(row=i+1, column=2).value = str(suikei.suikei_name) + ":" + str(suikei.suikei_code)
                    ws_suikei.cell(row=i+1, column=3).value = suikei.suikei_type_code

            ### 1050: 水系種別（水系・沿岸種別）
            print("handle_7_7", flush=True)
            if suikei_type_list:
                for i, suikei_type in enumerate(suikei_type_list):
                    ws_suikei_type.cell(row=i+1, column=1).value = suikei_type.suikei_type_code
                    ws_suikei_type.cell(row=i+1, column=2).value = str(suikei_type.suikei_type_name) + ":" + str(suikei_type.suikei_type_code)

            ### 1060: 河川（河川・海岸）、連動プルダウン用
            print("handle_7_8", flush=True)
            kasens_list = []
            if suikei_list:
                for i, suikei in enumerate(suikei_list):
                    kasens_list.append(KASEN.objects.raw("""
                        SELECT 
                            * 
                        FROM KASEN 
                        WHERE suikei_code=%s 
                        ORDER BY CAST(kasen_code AS INTEGER)""", [suikei.suikei_code, ]))
    
            print("handle_7_9", flush=True)
            if kasens_list:
                for i, kasens in enumerate(kasens_list):
                    if kasens:
                        for j, kasen in enumerate(kasens):
                            ws_kasen.cell(row=j+1, column=i*5+1).value = kasen.kasen_code
                            ws_kasen.cell(row=j+1, column=i*5+2).value = str(kasen.kasen_name) + ":" + str(kasen.kasen_code)
                            ws_kasen.cell(row=j+1, column=i*5+3).value = kasen.kasen_type_code
                            ws_kasen.cell(row=j+1, column=i*5+4).value = kasen.suikei_code

            ### 1070: 河川種別（河川・海岸種別）
            print("handle_7_10", flush=True)
            if kasen_type_list:
                for i, kasen_type in enumerate(kasen_type_list):
                    ws_kasen_type.cell(row=i+1, column=1).value = kasen_type.kasen_type_code
                    ws_kasen_type.cell(row=i+1, column=2).value = str(kasen_type.kasen_type_name) + ":" + str(kasen_type.kasen_type_code)
            
            ### 1080: 水害原因
            print("handle_7_11", flush=True)
            if cause_list:
                for i, cause in enumerate(cause_list):
                    ws_cause.cell(row=i+1, column=1).value = cause.cause_code
                    ws_cause.cell(row=i+1, column=2).value = str(cause.cause_name) + ":" + str(cause.cause_code)

            ### 1090: 地上地下区分
            print("handle_7_12", flush=True)
            if underground_list:
                for i, underground in enumerate(underground_list):
                    ws_underground.cell(row=i+1, column=1).value = underground.underground_code
                    ws_underground.cell(row=i+1, column=2).value = str(underground.underground_name) + ":" + str(underground.underground_code)

            ### 1100: 地下空間の利用形態
            print("handle_7_13", flush=True)
            if usage_list:
                for i, usage in enumerate(usage_list):
                    ws_usage.cell(row=i+1, column=1).value = usage.usage_code
                    ws_usage.cell(row=i+1, column=2).value = str(usage.usage_name) + ":" + str(usage.usage_code)

            ### 1110: 浸水土砂区分
            print("handle_7_14", flush=True)
            if flood_sediment_list:
                for i, flood_sediment in enumerate(flood_sediment_list):
                    ws_flood_sediment.cell(row=i+1, column=1).value = flood_sediment.flood_sediment_code
                    ws_flood_sediment.cell(row=i+1, column=2).value = str(flood_sediment.flood_sediment_name) + ":" + str(flood_sediment.flood_sediment_code)

            ### 1120: 地盤勾配区分
            print("handle_7_15", flush=True)
            if gradient_list:
                for i, gradient in enumerate(gradient_list):
                    ws_gradient.cell(row=i+1, column=1).value = gradient.gradient_code
                    ws_gradient.cell(row=i+1, column=2).value = str(gradient.gradient_name) + ":" + str(gradient.gradient_code)

            ### 1130: 産業分類
            print("handle_7_16", flush=True)
            if industry_list:
                for i, industry in enumerate(industry_list):
                    ws_industry.cell(row=i+1, column=1).value = industry.industry_code
                    ws_industry.cell(row=i+1, column=2).value = str(industry.industry_name) + ":" + str(industry.industry_code)

            ### 7000: 入力データ_水害区域
            print("handle_7_17", flush=True)
            if area_list:
                for i, area in enumerate(area_list):
                    ws_area.cell(row=i+1, column=1).value = area.area_id
                    ws_area.cell(row=i+1, column=2).value = str(area.area_name) + ":" + str(area.area_id)

            ### 7010: 入力データ_異常気象
            print("handle_7_18", flush=True)
            if weather_list:
                for i, weather in enumerate(weather_list):
                    ws_weather.cell(row=i+1, column=1).value = weather.weather_id
                    ws_weather.cell(row=i+1, column=2).value = str(weather.weather_name) + ":" + str(weather.weather_id)

            ### 7020: 入力データ_ヘッダ部分、水害
            print("handle_7_19", flush=True)
            if suigai_list:
                for i, suigai in enumerate(suigai_list):
                    ws_suigai.cell(row=i+1, column=1).value = suigai.suigai_id
                    ws_suigai.cell(row=i+1, column=2).value = str(suigai.suigai_name) + ":" + str(suigai.suigai_id)
                    ws_suigai.cell(row=i+1, column=3).value = suigai.ken_code
                    ws_suigai.cell(row=i+1, column=4).value = suigai.city_code
                    ws_suigai.cell(row=i+1, column=5).value = suigai.begin_date
                    ws_suigai.cell(row=i+1, column=6).value = suigai.end_date
                    ws_suigai.cell(row=i+1, column=7).value = suigai.cause_1_code
                    ws_suigai.cell(row=i+1, column=8).value = suigai.cause_2_code
                    ws_suigai.cell(row=i+1, column=9).value = suigai.cause_3_code
                    ws_suigai.cell(row=i+1, column=10).value = suigai.area_id
                    ws_suigai.cell(row=i+1, column=11).value = suigai.suikei_code
                    ws_suigai.cell(row=i+1, column=12).value = suigai.kasen_code
                    ws_suigai.cell(row=i+1, column=13).value = suigai.gradient_code
                    ws_suigai.cell(row=i+1, column=14).value = suigai.residential_area
                    ws_suigai.cell(row=i+1, column=15).value = suigai.agricultural_area
                    ws_suigai.cell(row=i+1, column=16).value = suigai.underground_area
                    ws_suigai.cell(row=i+1, column=17).value = suigai.kasen_kaigan_code
                    ws_suigai.cell(row=i+1, column=18).value = suigai.crop_damage
                    ws_suigai.cell(row=i+1, column=19).value = suigai.weather_id

            ###################################################################
            ### EXCEL入出力処理(0070)
            ### VLOOKUP用のシートに値をセットする。
            ###################################################################
            print_log('[DEBUG] P0900Action.action_O01_download_ippan_city.handle()関数 STEP 8/13.', 'DEBUG')
            ### 1020: 市区町村VLOOKUP
            print("handle_8_1", flush=True)
            if ken_list and cities_list:
                for i, ken in enumerate(ken_list):
                    ws_city_vlook.cell(row=i+1, column=1).value = str(ken.ken_name) + ":" + str(ken.ken_code)
        
                for i, cities in enumerate(cities_list):
                    ws_city_vlook.cell(row=i+1, column=2).value = 'CITY!$' + VLOOK_VALUE[i] + '$1:$' + VLOOK_VALUE[i] + '$%d' % len(cities)
    
            ### 1060: 河川（河川・海岸）VLOOKUP
            print("handle_8_2", flush=True)
            if suikei_list and kasens_list:
                for i, suikei in enumerate(suikei_list):
                    ws_kasen_vlook.cell(row=i+1, column=1).value = str(suikei.suikei_name) + ":" + str(suikei.suikei_code)
    
                for i, kasens in enumerate(kasens_list):
                    ws_kasen_vlook.cell(row=i+1, column=2).value = 'KASEN!$' + VLOOK_VALUE[i] + '$1:$' + VLOOK_VALUE[i] + '$%d' % len(kasens)

            ###################################################################
            ### EXCEL入出力処理(0080)
            ### 入力データ用EXCELシートのキャプションに値をセットする。
            ###################################################################
            print_log('[DEBUG] P0900Action.action_O01_download_ippan_city.handle()関数 STEP 9/13.', 'DEBUG')
            ### ws_ippan.cell(row=6, column=1).value = 'NO.'
            ### ws_ippan.cell(row=6, column=2).value = 'ファイル名'
            ### ws_ippan.cell(row=6, column=3).value = ''
            ### ws_ippan.cell(row=6, column=4).value = '都道府県[全角]'
            ### ws_ippan.cell(row=6, column=5).value = '市区町村名[全角]'
            ### ws_ippan.cell(row=6, column=6).value = '水害発生月日'
            ### ws_ippan.cell(row=6, column=7).value = ''
            ### ws_ippan.cell(row=6, column=8).value = '水害終了月日'
            ### ws_ippan.cell(row=6, column=9).value = ''
            ### ws_ippan.cell(row=6, column=10).value = '水害原因'
            ### ws_ippan.cell(row=6, column=11).value = ''
            ### ws_ippan.cell(row=6, column=12).value = ''
            ### ws_ippan.cell(row=6, column=13).value = '水害区域番号'
            ### ws_ippan.cell(row=6, column=14).value = '水系・沿岸名全角'
            ### ws_ippan.cell(row=6, column=15).value = '水系種別全角'
            ### ws_ippan.cell(row=6, column=16).value = '河川・海岸名全角'
            ### ws_ippan.cell(row=6, column=17).value = '河川種別全角'
            ### ws_ippan.cell(row=6, column=18).value = '地盤勾配区分'
            ### ws_ippan.cell(row=6, column=19).value = '町丁名・大字名全角'
            ### ws_ippan.cell(row=6, column=20).value = '名称全角'
            ### ws_ippan.cell(row=6, column=21).value = '地上地下区分'
            ### ws_ippan.cell(row=6, column=22).value = '浸水土砂区分'
            ### ws_ippan.cell(row=6, column=23).value = '被害建物棟数'
            ### ws_ippan.cell(row=6, column=24).value = ''
            ### ws_ippan.cell(row=6, column=25).value = ''
            ### ws_ippan.cell(row=6, column=26).value = ''
            ### ws_ippan.cell(row=6, column=27).value = ''
            ### ws_ippan.cell(row=6, column=28).value = ''
            ### ws_ippan.cell(row=6, column=29).value = '被害建物の延床面積'
            ### ws_ippan.cell(row=6, column=30).value = '被災世帯数'
            ### ws_ippan.cell(row=6, column=31).value = '被災事業所数'
            ### ws_ippan.cell(row=6, column=32).value = '農家・漁家戸数'
            ### ws_ippan.cell(row=6, column=33).value = ''
            ### ws_ippan.cell(row=6, column=34).value = ''
            ### ws_ippan.cell(row=6, column=35).value = ''
            ### ws_ippan.cell(row=6, column=36).value = ''
            ### ws_ippan.cell(row=6, column=37).value = '事業所従業者数'
            ### ws_ippan.cell(row=6, column=38).value = ''
            ### ws_ippan.cell(row=6, column=39).value = ''
            ### ws_ippan.cell(row=6, column=40).value = ''
            ### ws_ippan.cell(row=6, column=41).value = ''
            ### ws_ippan.cell(row=6, column=42).value = '事業所の産業分類'
            ### ws_ippan.cell(row=6, column=43).value = '地下空間の利用形態'
            ### ws_ippan.cell(row=6, column=44).value = '備考1'
            ### ws_ippan.cell(row=6, column=45).value = '備考2'
            ### ws_ippan.cell(row=6, column=46).value = '水害区域面積(宅地)'
            ### ws_ippan.cell(row=6, column=47).value = '水害区域面積(農地)'
            ### ws_ippan.cell(row=6, column=48).value = '水害区域面積(地下)'
            ### ws_ippan.cell(row=6, column=49).value = '工種'
            ### ws_ippan.cell(row=6, column=50).value = '農作物被害額(千円)'
            ### ws_ippan.cell(row=6, column=51).value = '異常気象コード'
            ### ws_ippan.cell(row=6, column=52).value = '調査票シート名'
            ### ws_ippan.cell(row=6, column=53).value = ''
            ### ws_ippan.cell(row=6, column=54).value = 'ファイル最終更新日時'
            ### ws_ippan.cell(row=6, column=55).value = 'ファイルパス'
            ### ws_ippan.cell(row=7, column=1).value = ''
            ### ws_ippan.cell(row=7, column=2).value = ''
            ### ws_ippan.cell(row=7, column=3).value = ''
            ### ws_ippan.cell(row=7, column=4).value = ''
            ### ws_ippan.cell(row=7, column=5).value = ''
            ### ws_ippan.cell(row=7, column=6).value = '月'
            ### ws_ippan.cell(row=7, column=7).value = '日'
            ### ws_ippan.cell(row=7, column=8).value = '月'
            ### ws_ippan.cell(row=7, column=9).value = '日'
            ### ws_ippan.cell(row=7, column=10).value = '1'
            ### ws_ippan.cell(row=7, column=11).value = '2'
            ### ws_ippan.cell(row=7, column=12).value = '3'
            ### ws_ippan.cell(row=7, column=13).value = ''
            ### ws_ippan.cell(row=7, column=14).value = ''
            ### ws_ippan.cell(row=7, column=15).value = ''
            ### ws_ippan.cell(row=7, column=16).value = ''
            ### ws_ippan.cell(row=7, column=17).value = ''
            ### ws_ippan.cell(row=7, column=18).value = ''
            ### ws_ippan.cell(row=7, column=19).value = ''
            ### ws_ippan.cell(row=7, column=20).value = ''
            ### ws_ippan.cell(row=7, column=21).value = ''
            ### ws_ippan.cell(row=7, column=22).value = ''
            ### ws_ippan.cell(row=7, column=23).value = '床下'
            ### ws_ippan.cell(row=7, column=24).value = '1〜49cm'
            ### ws_ippan.cell(row=7, column=25).value = '50〜99cm'
            ### ws_ippan.cell(row=7, column=26).value = '1m以上'
            ### ws_ippan.cell(row=7, column=27).value = '半壊'
            ### ws_ippan.cell(row=7, column=28).value = '全壊'
            ### ws_ippan.cell(row=7, column=29).value = ''
            ### ws_ippan.cell(row=7, column=30).value = ''
            ### ws_ippan.cell(row=7, column=31).value = ''
            ### ws_ippan.cell(row=7, column=32).value = '床下'
            ### ws_ippan.cell(row=7, column=33).value = '1〜49cm'
            ### ws_ippan.cell(row=7, column=34).value = '50〜99cm'
            ### ws_ippan.cell(row=7, column=35).value = '1m以上'
            ### ws_ippan.cell(row=7, column=36).value = '全壊'
            ### ws_ippan.cell(row=7, column=37).value = '床下'
            ### ws_ippan.cell(row=7, column=38).value = '1〜49cm'
            ### ws_ippan.cell(row=7, column=39).value = '50〜99cm'
            ### ws_ippan.cell(row=7, column=40).value = '1m以上'
            ### ws_ippan.cell(row=7, column=41).value = '全壊'
            ### ws_ippan.cell(row=7, column=42).value = ''
            ### ws_ippan.cell(row=7, column=43).value = ''
            ### ws_ippan.cell(row=7, column=44).value = ''
            ### ws_ippan.cell(row=7, column=45).value = ''
            ### ws_ippan.cell(row=7, column=46).value = ''
            ### ws_ippan.cell(row=7, column=47).value = ''
            ### ws_ippan.cell(row=7, column=48).value = ''
            ### ws_ippan.cell(row=7, column=49).value = ''
            ### ws_ippan.cell(row=7, column=50).value = ''
            ### ws_ippan.cell(row=7, column=51).value = ''
            ### ws_ippan.cell(row=7, column=52).value = ''
            ### ws_ippan.cell(row=7, column=53).value = ''
            ### ws_ippan.cell(row=7, column=54).value = ''
            ### ws_ippan.cell(row=7, column=55).value = ''
            
            ws_ippan.protection.enable()
            green_fill = PatternFill(fgColor='CCFFCC', patternType='solid')

            ###################################################################
            ### EXCEL入出力処理(0090)
            ### EXCELのセルに、値を埋め込む。
            ###################################################################
            print_log('[DEBUG] P0900Action.action_O01_download_ippan_city.handle()関数 STEP 10/13.', 'DEBUG')
            ippan_view_list = IPPAN_VIEW.objects.raw("""
                SELECT 
                    IV1.ippan_id AS ippan_id,
                    IV1.ippan_name AS ippan_name,
                    IV1.suigai_id AS suigai_id, 
                    IV1.suigai_name AS suigai_name, 
                    IV1.ken_code AS ken_code, 
                    IV1.ken_name AS ken_name, 
                    IV1.city_code AS city_code, 
                    IV1.city_name AS city_name, 
                    TO_CHAR(timezone('JST', IV1.begin_date::timestamptz), 'mm') AS begin_month, 
                    TO_CHAR(timezone('JST', IV1.begin_date::timestamptz), 'dd') AS begin_day, 
                    TO_CHAR(timezone('JST', IV1.end_date::timestamptz), 'mm') AS end_month, 
                    TO_CHAR(timezone('JST', IV1.end_date::timestamptz), 'dd') AS end_day, 
                    IV1.cause_1_code AS cause_1_code, 
                    IV1.cause_1_name AS cause_1_name, 
                    IV1.cause_2_code AS cause_2_code, 
                    IV1.cause_2_name AS cause_2_name, 
                    IV1.cause_3_code AS cause_3_code, 
                    IV1.cause_3_name AS cause_3_name, 
                    IV1.area_id AS area_id, 
                    IV1.area_name AS area_name, 
                    IV1.suikei_code AS suikei_code, 
                    IV1.suikei_name AS suikei_name, 
                    IV1.suikei_type_code AS suikei_type_code, 
                    IV1.suikei_type_name AS suikei_type_name, 
                    IV1.kasen_code AS kasen_code, 
                    IV1.kasen_name AS kasen_name, 
                    IV1.kasen_type_code AS kasen_type_code, 
                    IV1.kasen_type_name AS kasen_type_name, 
                    IV1.gradient_code AS gradient_code, 
                    IV1.gradient_name AS gradient_name, 
                    CAST(IV1.residential_area AS NUMERIC(20,10)) AS residential_area, 
                    CAST(IV1.agricultural_area AS NUMERIC(20,10)) AS agricultural_area, 
                    CAST(IV1.underground_area AS NUMERIC(20,10)) AS underground_area, 
                    IV1.kasen_kaigan_code AS kasen_kaigan_code, 
                    IV1.kasen_kaigan_name AS kasen_kaigan_name, 
                    CAST(IV1.crop_damage AS NUMERIC(20,10)) AS crop_damage, 
                    IV1.weather_id AS weather_id, 
                    IV1.weather_name AS weather_name, 
                    IV1.upload_file_path AS upload_file_path, 
                    IV1.upload_file_name AS upload_file_name, 
                    IV1.summary_file_path AS summary_file_path, 
                    IV1.summary_file_name AS summary_file_name, 
                    IV1.action_code AS action_code,
                    IV1.status_code AS status_code,
                    TO_CHAR(timezone('JST', IV1.committed_at::timestamptz), 'yyyy/mm/dd HH24:MI') AS committed_at, 
                    TO_CHAR(timezone('JST', IV1.deleted_at::timestamptz), 'yyyy/mm/dd HH24:MI') AS deleted_at, 
                    
                    IV1.building_code AS building_code,
                    IV1.building_name AS building_Name,
                    IV1.underground_code AS underground_code,
                    IV1.underground_name AS underground_name,
                    IV1.flood_sediment_code AS flood_sediment_code,
                    IV1.flood_sediment_name AS flood_sediment_name,
                    
                    CASE WHEN (IV1.building_lv00) IS NULL THEN NULL ELSE CAST(IV1.building_lv00 AS NUMERIC(20,10)) END AS building_lv00,
                    CASE WHEN (IV1.building_lv01_49) IS NULL THEN NULL ELSE CAST(IV1.building_lv01_49 AS NUMERIC(20,10)) END AS building_lv01_49,
                    CASE WHEN (IV1.building_lv50_99) IS NULL THEN NULL ELSE CAST(IV1.building_lv50_99 AS NUMERIC(20,10)) END AS building_lv50_99,
                    CASE WHEN (IV1.building_lv100) IS NULL THEN NULL ELSE CAST(IV1.building_lv100 AS NUMERIC(20,10)) END AS building_lv100,
                    CASE WHEN (IV1.building_half) IS NULL THEN NULL ELSE CAST(IV1.building_half AS NUMERIC(20,10)) END AS building_half,
                    CASE WHEN (IV1.building_full) IS NULL THEN NULL ELSE CAST(IV1.building_full AS NUMERIC(20,10)) END AS building_full,
                    
                    CASE WHEN (IV1.floor_area) IS NULL THEN NULL ELSE CAST(IV1.floor_area AS NUMERIC(20,10)) END AS floor_area,
                    CASE WHEN (IV1.family) IS NULL THEN NULL ELSE CAST(IV1.family AS NUMERIC(20,10)) END AS family,
                    CASE WHEN (IV1.office) IS NULL THEN NULL ELSE CAST(IV1.office AS NUMERIC(20,10)) END AS office,
                    
                    CASE WHEN (IV1.floor_area_lv00) IS NULL THEN NULL ELSE CAST(IV1.floor_area_lv00 AS NUMERIC(20,10)) END AS floor_area_lv00,
                    CASE WHEN (IV1.floor_area_lv01_49) IS NULL THEN NULL ELSE CAST(IV1.floor_area_lv01_49 AS NUMERIC(20,10)) END AS floor_area_lv01_49,
                    CASE WHEN (IV1.floor_area_lv50_99) IS NULL THEN NULL ELSE CAST(IV1.floor_area_lv50_99 AS NUMERIC(20,10)) END AS floor_area_lv50_99,
                    CASE WHEN (IV1.floor_area_lv100) IS NULL THEN NULL ELSE CAST(IV1.floor_area_lv100 AS NUMERIC(20,10)) END AS floor_area_lv100,
                    CASE WHEN (IV1.floor_area_half) IS NULL THEN NULL ELSE CAST(IV1.floor_area_half AS NUMERIC(20,10)) END AS floor_area_half,
                    CASE WHEN (IV1.floor_area_full) IS NULL THEN NULL ELSE CAST(IV1.floor_area_full AS NUMERIC(20,10)) END AS floor_area_full,
                    
                    CASE WHEN (IV1.family_lv00) IS NULL THEN NULL ELSE CAST(IV1.family_lv00 AS NUMERIC(20,10)) END AS family_lv00,
                    CASE WHEN (IV1.family_lv01_49) IS NULL THEN NULL ELSE CAST(IV1.family_lv01_49 AS NUMERIC(20,10)) END AS family_lv01_49,
                    CASE WHEN (IV1.family_lv50_99) IS NULL THEN NULL ELSE CAST(IV1.family_lv50_99 AS NUMERIC(20,10)) END AS family_lv50_99,
                    CASE WHEN (IV1.family_lv100) IS NULL THEN NULL ELSE CAST(IV1.family_lv100 AS NUMERIC(20,10)) END AS family_lv100,
                    CASE WHEN (IV1.family_half) IS NULL THEN NULL ELSE CAST(IV1.family_half AS NUMERIC(20,10)) END AS family_half,
                    CASE WHEN (IV1.family_full) IS NULL THEN NULL ELSE CAST(IV1.family_full AS NUMERIC(20,10)) END AS family_full,
                    
                    CASE WHEN (IV1.office_lv00) IS NULL THEN NULL ELSE CAST(IV1.office_lv00 AS NUMERIC(20,10)) END AS office_lv00,
                    CASE WHEN (IV1.office_lv01_49) IS NULL THEN NULL ELSE CAST(IV1.office_lv01_49 AS NUMERIC(20,10)) END AS office_lv01_49,
                    CASE WHEN (IV1.office_lv50_99) IS NULL THEN NULL ELSE CAST(IV1.office_lv50_99 AS NUMERIC(20,10)) END AS office_lv50_99,
                    CASE WHEN (IV1.office_lv100) IS NULL THEN NULL ELSE CAST(IV1.office_lv100 AS NUMERIC(20,10)) END AS office_lv100,
                    CASE WHEN (IV1.office_half) IS NULL THEN NULL ELSE CAST(IV1.office_half AS NUMERIC(20,10)) END AS office_half,
                    CASE WHEN (IV1.office_full) IS NULL THEN NULL ELSE CAST(IV1.office_full AS NUMERIC(20,10)) END AS office_full,
                    
                    CASE WHEN (IV1.farmer_fisher_lv00) IS NULL THEN NULL ELSE CAST(IV1.farmer_fisher_lv00 AS NUMERIC(20,10)) END AS farmer_fisher_lv00,
                    CASE WHEN (IV1.farmer_fisher_lv01_49) IS NULL THEN NULL ELSE CAST(IV1.farmer_fisher_lv01_49 AS NUMERIC(20,10)) END AS farmer_fisher_lv01_49,
                    CASE WHEN (IV1.farmer_fisher_lv50_99) IS NULL THEN NULL ELSE CAST(IV1.farmer_fisher_lv50_99 AS NUMERIC(20,10)) END AS farmer_fisher_lv50_99,
                    CASE WHEN (IV1.farmer_fisher_lv100) IS NULL THEN NULL ELSE CAST(IV1.farmer_fisher_lv100 AS NUMERIC(20,10)) END AS farmer_fisher_lv100,
                    CASE WHEN (IV1.farmer_fisher_full) IS NULL THEN NULL ELSE CAST(IV1.farmer_fisher_full AS NUMERIC(20,10)) END AS farmer_fisher_full,
                    
                    CASE WHEN (IV1.employee_lv00) IS NULL THEN NULL ELSE CAST(IV1.employee_lv00 AS NUMERIC(20,10)) END AS employee_lv00,
                    CASE WHEN (IV1.employee_lv01_49) IS NULL THEN NULL ELSE CAST(IV1.employee_lv01_49 AS NUMERIC(20,10)) END AS employee_lv01_49,
                    CASE WHEN (IV1.employee_lv50_99) IS NULL THEN NULL ELSE CAST(IV1.employee_lv50_99 AS NUMERIC(20,10)) END AS employee_lv50_99,
                    CASE WHEN (IV1.employee_lv100) IS NULL THEN NULL ELSE CAST(IV1.employee_lv100 AS NUMERIC(20,10)) END AS employee_lv100,
                    CASE WHEN (IV1.employee_full) IS NULL THEN NULL ELSE CAST(IV1.employee_full AS NUMERIC(20,10)) END AS employee_full,
                    
                    IV1.industry_code AS industry_code,
                    IV1.industry_name AS industry_name,
                    IV1.usage_code as usage_code,
                    IV1.usage_name as usage_name,
                    IV1.comment as comment 
                    
                FROM IPPAN_VIEW IV1 
                WHERE 
                    city_code=%s AND deleted_at IS NULL 
                ORDER BY CAST(suigai_id AS INTEGER), CAST(ippan_id AS INTEGER)""", [
                    trigger_list[0].city_code, 
                ])
    
            if ippan_view_list:
                for i, ippan in enumerate(ippan_view_list):
                    if ippan.ippan_id is not None and len(str(ippan.ippan_id)) > 0:
                        ws_ippan.cell(row=8+i, column=1).value = str(ippan.ippan_id)
                        ws_ippan.cell(row=8+i, column=1).fill = green_fill
                        
                    if ippan.upload_file_name is not None and len(str(ippan.upload_file_name)) > 0:
                        ws_ippan.cell(row=8+i, column=2).value = str(ippan.upload_file_name)
                        
                    ws_ippan.cell(row=8+i, column=3).value = ""
                    
                    if ippan.ken_name is not None and ippan.ken_code is not None and len(str(ippan.ken_name)) > 0 and len(str(ippan.ken_code)) > 0:
                        ws_ippan.cell(row=8+i, column=4).value = str(ippan.ken_name) + str(ippan.ken_code)
                        
                    if ippan.city_name is not None and ippan.city_code is not None and len(str(ippan.city_name)) > 0 and len(str(ippan.city_code)) > 0:
                        ws_ippan.cell(row=8+i, column=5).value = str(ippan.city_name) + ":" + str(ippan.city_code)
                    
                    if ippan.begin_month is not None and len(str(ippan.begin_month)) > 0:
                        ws_ippan.cell(row=8+i, column=6).value = str(ippan.begin_month)
                        
                    if ippan.begin_day is not None and len(str(ippan.begin_day)) > 0:
                        ws_ippan.cell(row=8+i, column=7).value = str(ippan.begin_day)
                        
                    if ippan.end_month is not None and len(str(ippan.end_month)) > 0:
                        ws_ippan.cell(row=8+i, column=8).value = str(ippan.end_month)
                        
                    if ippan.end_day is not None and len(str(ippan.end_day)) > 0:
                        ws_ippan.cell(row=8+i, column=9).value = str(ippan.end_day)
                        
                    if ippan.cause_1_name is not None and len(str(ippan.cause_1_name)) > 0:
                        ws_ippan.cell(row=8+i, column=10).value = str(ippan.cause_1_name)
                        
                    if ippan.cause_2_name is not None and len(str(ippan.cause_2_name)) > 0:
                        ws_ippan.cell(row=8+i, column=11).value = str(ippan.cause_2_name)
                        
                    if ippan.cause_3_name is not None and len(str(ippan.cause_3_name)) > 0:
                        ws_ippan.cell(row=8+i, column=12).value = str(ippan.cause_3_name)
                    
                    if ippan.area_name is not None and len(str(ippan.area_name)) > 0:
                        ws_ippan.cell(row=8+i, column=13).value = str(ippan.area_name)
                        
                    if ippan.suikei_name is not None and len(str(ippan.suikei_name)) > 0:
                        ws_ippan.cell(row=8+i, column=14).value = str(ippan.suikei_name)
                        
                    if ippan.suikei_type_name is not None and len(str(ippan.suikei_type_name)) > 0:
                        ws_ippan.cell(row=8+i, column=15).value = str(ippan.suikei_type_name)
                        
                    if ippan.kasen_name is not None and len(str(ippan.kasen_name)) > 0:
                        ws_ippan.cell(row=8+i, column=16).value = str(ippan.kasen_name)
                        
                    if ippan.kasen_type_name is not None and len(str(ippan.kasen_type_name)) > 0:
                        ws_ippan.cell(row=8+i, column=17).value = str(ippan.kasen_type_name)
                        
                    if ippan.gradient_name is not None and len(str(ippan.gradient_name)) > 0:
                        ws_ippan.cell(row=8+i, column=18).value = str(ippan.gradient_name)
                        
                    if ippan.ippan_name is not None and len(str(ippan.ippan_name)) > 0:
                        ws_ippan.cell(row=8+i, column=19).value = str(ippan.ippan_name)
                        
                    if ippan.building_name is not None and len(str(ippan.building_name)) > 0:
                        ws_ippan.cell(row=8+i, column=20).value = str(ippan.building_name)
                        
                    if ippan.underground_name is not None and len(str(ippan.underground_name)) > 0:
                        ws_ippan.cell(row=8+i, column=21).value = str(ippan.underground_name)
                        
                    if ippan.flood_sediment_name is not None and len(str(ippan.flood_sediment_name)) > 0:
                        ws_ippan.cell(row=8+i, column=22).value = str(ippan.flood_sediment_name)
                        
                    if ippan.building_lv00 is not None and len(str(ippan.building_lv00)) > 0:
                        ws_ippan.cell(row=8+i, column=23).value = str(ippan.building_lv00)
                        
                    if ippan.building_lv01_49 is not None and len(str(ippan.building_lv01_49)) > 0:
                        ws_ippan.cell(row=8+i, column=24).value = str(ippan.building_lv01_49)
                        
                    if ippan.building_lv50_99 is not None and len(str(ippan.building_lv50_99)) > 0:
                        ws_ippan.cell(row=8+i, column=25).value = str(ippan.building_lv50_99)
                        
                    if ippan.building_lv100 is not None and len(str(ippan.building_lv100)) > 0:
                        ws_ippan.cell(row=8+i, column=26).value = str(ippan.building_lv100)
                        
                    if ippan.building_half is not None and len(str(ippan.building_half)) > 0:
                        ws_ippan.cell(row=8+i, column=27).value = str(ippan.building_half)
                        
                    if ippan.building_full is not None and len(str(ippan.building_full)) > 0:
                        ws_ippan.cell(row=8+i, column=28).value = str(ippan.building_full)
                        
                    if ippan.floor_area is not None and len(str(ippan.floor_area)) > 0:
                        ws_ippan.cell(row=8+i, column=29).value = str(ippan.floor_area)
                        
                    if ippan.family is not None and len(str(ippan.family)) > 0:
                        ws_ippan.cell(row=8+i, column=30).value = str(ippan.family)
                        
                    if ippan.office is not None and len(str(ippan.office)) > 0:
                        ws_ippan.cell(row=8+i, column=31).value = str(ippan.office)
                        
                    if ippan.farmer_fisher_lv00 is not None and len(str(ippan.farmer_fisher_lv00)) > 0:
                        ws_ippan.cell(row=8+i, column=32).value = str(ippan.farmer_fisher_lv00)
                        
                    if ippan.farmer_fisher_lv01_49 is not None and len(str(ippan.farmer_fisher_lv01_49)) > 0:
                        ws_ippan.cell(row=8+i, column=33).value = str(ippan.farmer_fisher_lv01_49)
                        
                    if ippan.farmer_fisher_lv50_99 is not None and len(str(ippan.farmer_fisher_lv50_99)) > 0:
                        ws_ippan.cell(row=8+i, column=34).value = str(ippan.farmer_fisher_lv50_99)
                        
                    if ippan.farmer_fisher_lv100 is not None and len(str(ippan.farmer_fisher_lv100)) > 0:
                        ws_ippan.cell(row=8+i, column=35).value = str(ippan.farmer_fisher_lv100)
                        
                    if ippan.farmer_fisher_full is not None and len(str(ippan.farmer_fisher_full)) > 0:
                        ws_ippan.cell(row=8+i, column=36).value = str(ippan.farmer_fisher_full)
                        
                    if ippan.employee_lv00 is not None and len(str(ippan.employee_lv00)) > 0:
                        ws_ippan.cell(row=8+i, column=37).value = str(ippan.employee_lv00)
                        
                    if ippan.employee_lv01_49 is not None and len(str(ippan.employee_lv01_49)) > 0:
                        ws_ippan.cell(row=8+i, column=38).value = str(ippan.employee_lv01_49)
                        
                    if ippan.employee_lv50_99 is not None and len(str(ippan.employee_lv50_99)) > 0:
                        ws_ippan.cell(row=8+i, column=39).value = str(ippan.employee_lv50_99)
                        
                    if ippan.employee_lv100 is not None and len(str(ippan.employee_lv100)) > 0:
                        ws_ippan.cell(row=8+i, column=40).value = str(ippan.employee_lv100)
                        
                    if ippan.employee_full is not None and len(str(ippan.employee_full)) > 0:
                        ws_ippan.cell(row=8+i, column=41).value = str(ippan.employee_full)
                        
                    if ippan.industry_name is not None and len(str(ippan.industry_name)) > 0:
                        ws_ippan.cell(row=8+i, column=42).value = str(ippan.industry_name)
                        
                    if ippan.usage_name is not None and len(str(ippan.usage_name)) > 0:
                        ws_ippan.cell(row=8+i, column=43).value = str(ippan.usage_name)
                        
                    if ippan.comment is not None and len(str(ippan.comment)) > 0:
                        ws_ippan.cell(row=8+i, column=44).value = str(ippan.comment)
                    
                    ### TO-DO TODO TO_DO to-do todo to_do
                    if ippan.comment is not None and len(str(ippan.comment)) > 0:
                        ws_ippan.cell(row=8+i, column=45).value = str(ippan.comment)
                    
                    if ippan.residential_area is not None and len(str(ippan.residential_area)) > 0:
                        ws_ippan.cell(row=8+i, column=46).value = str(ippan.residential_area)
                        
                    if ippan.agricultural_area is not None and len(str(ippan.agricultural_area)) > 0:
                        ws_ippan.cell(row=8+i, column=47).value = str(ippan.agricultural_area)
                        
                    if ippan.underground_area is not None and len(str(ippan.underground_area)) > 0:
                        ws_ippan.cell(row=8+i, column=48).value = str(ippan.underground_area)
                        
                    if ippan.kasen_kaigan_name is not None and len(str(ippan.kasen_kaigan_name)) > 0:
                        ws_ippan.cell(row=8+i, column=49).value = str(ippan.kasen_kaigan_name)
                        
                    if ippan.crop_damage is not None and len(str(ippan.crop_damage)) > 0:
                        ws_ippan.cell(row=8+i, column=50).value = str(ippan.crop_damage)
                        
                    if ippan.weather_name is not None and len(str(ippan.weather_name)) > 0:
                        ws_ippan.cell(row=8+i, column=51).value = str(ippan.weather_name)
                        
                    if ippan.suigai_name is not None and len(str(ippan.suigai_name)) > 0:
                        ws_ippan.cell(row=8+i, column=52).value = str(ippan.suigai_name)
                        
                    ws_ippan.cell(row=8+i, column=53).value = ""
                    
                    if ippan.committed_at is not None and len(str(ippan.committed_at)) > 0:
                        ws_ippan.cell(row=8+i, column=54).value = str(ippan.committed_at)
                        
                    if ippan.upload_file_path is not None and len(str(ippan.upload_file_path)) > 0:
                        ws_ippan.cell(row=8+i, column=55).value = str(ippan.upload_file_path)

            ###################################################################
            ### EXCEL入出力処理(0100)
            ### ダウンロード用のEXCELファイルを保存する。
            ###################################################################
            print_log('[DEBUG] P0900Action.action_O01_download_ippan_city.handle()関数 STEP 11/13.', 'DEBUG')
            wb.save(download_file_path)

            ###################################################################
            ### DBアクセス処理(0110)
            ### トリガーデータを更新する。
            ###################################################################
            print_log('[DEBUG] P0900Action.action_O01_download_ippan_city.handle()関数 STEP 12/13.', 'DEBUG')
            connection_cursor = connection.cursor()
            try:
                connection_cursor.execute("""BEGIN""", [])
                
                connection_cursor.execute("""
                    UPDATE TRIGGER SET 
                        status_code=%s, -- status_code
                        success_count=%s, -- success_count
                        failure_count=%s, -- failure_count
                        consumed_at=CURRENT_TIMESTAMP, 
                        integrity_ok=%s, -- integrity_ok
                        integrity_ng=%s  -- integrity_ng
                    WHERE
                        trigger_id=%s -- trigger_id
                    """, [
                        'SUCCESS', ### status_count
                        1, ### success_count
                        0, ### failure_count
                        '\n'.join(get_info_log()), ### integrity_ok
                        '\n'.join(get_warn_log()), ### integrity_ng
                        trigger_list[0].trigger_id, ### trigger_id
                    ])
                ### transaction.commit()    
                connection_cursor.execute("""COMMIT""", [])
            except:
                print_log('[ERROR] P0900Action.action_O01_download_ippan_city.handle()関数 {}'.format(sys.exc_info()[0]), 'ERROR')
                ### connection_cursor.rollback()
                connection_cursor.execute("""ROLLBACK""", [])
            finally:
                connection_cursor.close()
                
            ###################################################################
            ### 戻り値セット処理(0120)
            ### 戻り値を戻す。
            ###################################################################
            print_log('[DEBUG] P0900Action.action_O01_download_ippan_city.handle()関数 STEP 13/13.', 'DEBUG')
            print_log('[INFO] P0900Action.action_O01_download_ippan_city.handle()関数が正常終了しました。', 'INFO')
            return 0
        except:
            print_log('[ERROR] P0900Action.action_O01_download_ippan_city.handle()関数 {}'.format(sys.exc_info()[0]), 'ERROR')
            print_log('[ERROR] P0900Action.action_O01_download_ippan_city.handle()関数でエラーが発生しました。', 'ERROR')
            print_log('[ERROR] P0900Action.action_O01_download_ippan_city.handle()関数が異常終了しました。', 'ERROR')
            return 8
            