#!/usr/bin/env python3
# -*- coding: utf-8 -*-
###############################################################################
### ファイル名：P0900Action/management/commands/action_A04_verify_idb_by_diff_method.py
### A04：差分検証
###############################################################################

###############################################################################
### 処理名：インポート処理
###############################################################################
from django.core.management.base import BaseCommand

import sys
from django.contrib.auth.decorators import login_required
from django.db import connection
from django.db import transaction
from django.http import Http404
from django.http import HttpResponse
from django.http import HttpResponseRedirect
from django.shortcuts import render
from django.template import loader
from django.views import generic
from django.views.generic import FormView
from django.views.generic.base import TemplateView

import openpyxl
from openpyxl.comments import Comment
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import PatternFill
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.writer.excel import save_virtual_workbook

from P0000Common.models import BUILDING                ### 1000: 建物区分
from P0000Common.models import KEN                     ### 1010: 都道府県
from P0000Common.models import CITY                    ### 1020: 市区町村
from P0000Common.models import KASEN_KAIGAN            ### 1030: 水害発生地点工種（河川海岸区分）
from P0000Common.models import SUIKEI                  ### 1040: 水系（水系・沿岸）
from P0000Common.models import SUIKEI_TYPE             ### 1050: 水系種別（水系・沿岸種別）
from P0000Common.models import KASEN                   ### 1060: 河川（河川・海岸）
from P0000Common.models import KASEN_TYPE              ### 1070: 河川種別（河川・海岸種別）
from P0000Common.models import CAUSE                   ### 1080: 水害原因
from P0000Common.models import UNDERGROUND             ### 1090: 地上地下区分
from P0000Common.models import USAGE                   ### 1100: 地下空間の利用形態
from P0000Common.models import FLOOD_SEDIMENT          ### 1110: 浸水土砂区分
from P0000Common.models import GRADIENT                ### 1120: 地盤勾配区分
from P0000Common.models import INDUSTRY                ### 1130: 産業分類

from P0000Common.models import HOUSE_ASSET             ### 2000: 家屋評価額
from P0000Common.models import HOUSE_RATE              ### 2010: 家屋被害率
from P0000Common.models import HOUSE_ALT               ### 2020: 家庭応急対策費_代替活動費
from P0000Common.models import HOUSE_CLEAN             ### 2030: 家庭応急対策費_清掃日数

from P0000Common.models import HOUSEHOLD_ASSET         ### 3000: 家庭用品自動車以外所有額
from P0000Common.models import HOUSEHOLD_RATE          ### 3010: 家庭用品自動車以外被害率

from P0000Common.models import CAR_ASSET               ### 4000: 家庭用品自動車所有額
from P0000Common.models import CAR_RATE                ### 4010: 家庭用品自動車被害率

from P0000Common.models import OFFICE_ASSET            ### 5000: 事業所資産額
from P0000Common.models import OFFICE_RATE             ### 5010: 事業所被害率
from P0000Common.models import OFFICE_SUSPEND          ### 5020: 事業所営業停止日数
from P0000Common.models import OFFICE_STAGNATE         ### 5030: 事業所営業停滞日数
from P0000Common.models import OFFICE_ALT              ### 5040: 事業所応急対策費_代替活動費

from P0000Common.models import FARMER_FISHER_ASSET     ### 6000: 農漁家資産額
from P0000Common.models import FARMER_FISHER_RATE      ### 6010: 農漁家被害率

from P0000Common.models import AREA                    ### 7000: 入力データ_水害区域
from P0000Common.models import WEATHER                 ### 7010: 入力データ_異常気象
from P0000Common.models import SUIGAI                  ### 7020: 入力データ_ヘッダ部分
from P0000Common.models import IPPAN                   ### 7030: 入力データ_一覧表部分
from P0000Common.models import IPPAN_VIEW              ### 7040: ビューデータ_一覧表部分

from P0000Common.models import IPPAN_SUMMARY           ### 8000: 集計データ

from P0000Common.models import ACTION                  ### 10000: アクション
from P0000Common.models import STATUS                  ### 10010: 状態
from P0000Common.models import TRIGGER                 ### 10020: トリガーメッセージ
from P0000Common.models import APPROVAL                ### 10030: 承認メッセージ
from P0000Common.models import FEEDBACK                ### 10040: フィードバックメッセージ

from P0000Common.common import get_debug_log
from P0000Common.common import get_error_log
from P0000Common.common import get_info_log
from P0000Common.common import get_warn_log
from P0000Common.common import print_log
from P0000Common.common import reset_log

###############################################################################
### クラス名： Command
###############################################################################
class Command(BaseCommand):
    
    ###########################################################################
    ### 関数名： handle
    ### チェックOKの場合、トリガーの状態を成功に更新、消費日時をカレントに更新、新たなトリガーを生成する。
    ### チェックNGの場合、トリガーの状態を失敗に更新、消費日時をカレントに更新する。
    ### チェックNGの場合、当該水害IDについて、以降の処理を止めて、手動？で再実行、または、入力データから再登録するイメージである。
    ### 上記は、このアプリの共通の考え方とする。
    ###########################################################################
    def handle(self, *args, **options):
        try:
            ###################################################################
            ### 引数チェック処理(0000)
            ### コマンドラインからの引数をチェックする。
            ###################################################################
            reset_log()
            print_log('[INFO] P0900Action.action_A04_verify_idb_by_diff_method.handle()関数が開始しました。', 'INFO')
            print_log('[DEBUG] P0900Action.action_A04_verify_idb_by_diff_method.handle()関数 STEP 1/12.', 'DEBUG')

            ###################################################################
            ### DBアクセス処理(0010)
            ### TRIGGERテーブルからSUIGAI_IDのリストを取得する。
            ### トリガーメッセージにアクションが発行されているかを検索する。
            ### 処理対象の水害IDを取得する。
            ### トリガーメッセージにアクションが発行されていなければ、処理を終了する。
            ### ※ネストを浅くするために、処理対象外の場合、処理を終了させる。
            ###################################################################
            print_log('[DEBUG] P0900Action.action_A04_verify_idb_by_diff_method.handle()関数 STEP 2/12.', 'DEBUG')
            trigger_list = TRIGGER.objects.raw("""
                SELECT 
                    * 
                FROM TRIGGER 
                WHERE 
                    action_code='A04' AND 
                    consumed_at IS NULL AND 
                    deleted_at IS NULL 
                ORDER BY CAST(trigger_id AS INTEGER) LIMIT 1""", [])
            
            if trigger_list is None:
                print_log('[INFO] P0900Action.action_A04_verify_idb_by_diff_method.handle()関数が正常終了しました。', 'INFO')
                return 0
            
            if len(trigger_list) == 0:
                print_log('[INFO] P0900Action.action_A04_verify_idb_by_diff_method.handle()関数が正常終了しました。', 'INFO')
                return 0

            print_log('[DEBUG] P0900Action.action_A04_verify_idb_by_diff_method.handle()関数 trigger_list[0].trigger_id = {}'.format(trigger_list[0].trigger_id), 'DEBUG')
            print_log('[DEBUG] P0900Action.action_A04_verify_idb_by_diff_method.handle()関数 trigger_list[0].suigai_id = {}'.format(trigger_list[0].suigai_id), 'DEBUG')
            print_log('[DEBUG] P0900Action.action_A04_verify_idb_by_diff_method.handle()関数 trigger_list[0].city_code = {}'.format(trigger_list[0].city_code), 'DEBUG')
            print_log('[DEBUG] P0900Action.action_A04_verify_idb_by_diff_method.handle()関数 trigger_list[0].ken_code = {}'.format(trigger_list[0].ken_code), 'DEBUG')
            print_log('[DEBUG] P0900Action.action_A04_verify_idb_by_diff_method.handle()関数 trigger_list[0].download_file_path = {}'.format(trigger_list[0].download_file_path), 'DEBUG')
            print_log('[DEBUG] P0900Action.action_A04_verify_idb_by_diff_method.handle()関数 trigger_list[0].download_file_name = {}'.format(trigger_list[0].download_file_name), 'DEBUG')
            print_log('[DEBUG] P0900Action.action_A04_verify_idb_by_diff_method.handle()関数 trigger_list[0].upload_file_path = {}'.format(trigger_list[0].upload_file_path), 'DEBUG')
            print_log('[DEBUG] P0900Action.action_A04_verify_idb_by_diff_method.handle()関数 trigger_list[0].upload_file_name = {}'.format(trigger_list[0].upload_file_name), 'DEBUG')

            ###################################################################
            ### DBアクセス処理(0020)
            ### DBから入力データ_ヘッダ部分のデータを取得する。
            ###################################################################
            print_log('[DEBUG] P0900Action.action_A04_verify_idb_by_diff_method.handle()関数 STEP 3/12.', 'DEBUG')
            suigai_list = SUIGAI.objects.raw("""
                SELECT 
                    SG1.suigai_id AS suigai_id,
                    SG1.suigai_name AS suigai_name,
                    SG1.ken_code AS ken_code, 
                    KE1.ken_name AS ken_name,
                    SG1.city_code AS city_code,
                    CI1.city_name AS city_name,
                    TO_CHAR(timezone('JST', SG1.begin_date::timestamptz), 'yyyy/mm/dd') AS begin_date,
                    TO_CHAR(timezone('JST', SG1.end_date::timestamptz), 'yyyy/mm/dd') AS end_date,
                    SG1.cause_1_code AS cause_1_code,
                    CA1.cause_name AS cause_1_name,
                    SG1.cause_2_code AS cause_2_code,
                    CA2.cause_name AS cause_2_name,
                    SG1.cause_3_code AS cause_3_code,
                    CA3.cause_name AS cause_3_name,
                    SG1.area_id AS area_id,
                    AR1.area_name AS area_name,
                    
                    SG1.suikei_code AS suikei_code,
                    SK1.suikei_name AS suikei_name,
                    SKT1.suikei_type_code AS suikei_type_code,
                    SKT1.suikei_type_name AS suikei_type_name,
                    SG1.kasen_code AS kasen_code,
                    KA1.kasen_name AS kasen_name,
                    KAT1.kasen_type_code AS kasen_type_code,
                    KAT1.kasen_type_name AS kasen_type_name,
                    SG1.gradient_code AS gradient_code,
                    GR1.gradient_name AS gradient_name,
                    
                    CASE WHEN (SG1.residential_area) IS NULL THEN NULL ELSE CAST(SG1.residential_area AS NUMERIC(20,10)) END AS residential_area,
                    CASE WHEN (SG1.agricultural_area) IS NULL THEN NULL ELSE CAST(SG1.agricultural_area AS NUMERIC(20,10)) END AS agricultural_area,
                    CASE WHEN (SG1.underground_area) IS NULL THEN NULL ELSE CAST(SG1.underground_area AS NUMERIC(20,10)) END AS underground_area,
                    SG1.kasen_kaigan_code AS kasen_kaigan_code,
                    KK1.kasen_kaigan_name AS kasen_kaigan_name,
                    CASE WHEN (SG1.crop_damage) IS NULL THEN NULL ELSE CAST(SG1.crop_damage AS NUMERIC(20,10)) END AS crop_damage,
                    SG1.weather_id AS weather_id,
                    WE1.weather_name AS weather_name, 
                    
                    TO_CHAR(timezone('JST', SG1.committed_at::timestamptz), 'yyyy/mm/dd HH24:MI') AS committed_at, 
                    TO_CHAR(timezone('JST', SG1.deleted_at::timestamptz), 'yyyy/mm/dd HH24:MI') AS deleted_at, 
                    
                    SG1.upload_file_path AS upload_file_path, 
                    SG1.upload_file_name AS upload_file_name, 
                    SG1.summary_file_path AS summary_file_path, 
                    SG1.summary_file_name AS summary_file_name, 
                    SG1.action_code AS action_code, 
                    AC1.action_name AS action_name, 
                    SG1.status_code AS status_code, 
                    ST1.status_name AS status_name 
                    
                FROM SUIGAI SG1 
                LEFT JOIN KEN KE1 ON SG1.ken_code = KE1.ken_code 
                LEFT JOIN CITY CI1 ON SG1.city_code = CI1.city_code 
                LEFT JOIN CAUSE CA1 ON SG1.cause_1_code = CA1.cause_code 
                LEFT JOIN CAUSE CA2 ON SG1.cause_2_code = CA2.cause_code 
                LEFT JOIN CAUSE CA3 ON SG1.cause_3_code = CA3.cause_code 
                LEFT JOIN AREA AR1 ON SG1.area_id = AR1.area_id 
                LEFT JOIN SUIKEI SK1 ON SG1.suikei_code = SK1.suikei_code 
                LEFT JOIN SUIKEI_TYPE SKT1 ON SK1.suikei_type_code = SKT1.suikei_type_code 
                LEFT JOIN KASEN KA1 ON SG1.kasen_code = KA1.kasen_code 
                LEFT JOIN KASEN_TYPE KAT1 ON KA1.kasen_type_code = KAT1.kasen_type_code 
                LEFT JOIN GRADIENT GR1 ON SG1.gradient_code = GR1.gradient_code 
                LEFT JOIN KASEN_KAIGAN KK1 ON SG1.kasen_kaigan_code = KK1.kasen_kaigan_code 
                LEFT JOIN WEATHER WE1 ON SG1.weather_id = WE1.weather_id 
                LEFT JOIN ACTION AC1 ON SG1.action_code = AC1.action_code 
                LEFT JOIN STATUS ST1 ON SG1.status_code = ST1.status_code 
                WHERE 
                    SG1.suigai_id = %s""", [trigger_list[0].suigai_id, ])

            ###################################################################
            ### DBアクセス処理(0030)
            ### DBから入力データ_一覧表部分のデータを取得する。
            ###################################################################
            print_log('[DEBUG] P0900Action.action_A04_verify_idb_by_diff_method.handle()関数 STEP 4/12.', 'DEBUG')
            ippan_list = IPPAN.objects.raw("""
                SELECT 
                    IV1.ippan_id AS ippan_id, 
                    IV1.ippan_name AS ippan_name, 
                    IV1.suigai_id AS suigai_id, 
                    IV1.suigai_name AS suigai_name, 
                    IV1.ken_code AS ken_code, 
                    IV1.ken_name AS ken_name, 
                    IV1.city_code AS city_code, 
                    IV1.city_name AS city_name, 
                    TO_CHAR(timezone('JST', IV1.begin_date::timestamptz), 'yyyy/mm/dd') AS begin_date, 
                    TO_CHAR(timezone('JST', IV1.end_date::timestamptz), 'yyyy/mm/dd') AS end_date, 
                    IV1.cause_1_code AS cause_1_code, 
                    IV1.cause_1_name AS cause_1_name, 
                    IV1.cause_2_code AS cause_2_code, 
                    IV1.cause_2_name AS cause_2_name, 
                    IV1.cause_3_code AS cause_3_code, 
                    IV1.cause_3_name AS cause_3_name, 
                    IV1.area_id AS area_id, 
                    IV1.area_name AS area_name, 
                    
                    IV1.suikei_code AS suikei_code, 
                    IV1.suikei_name AS suikei_name, 
                    IV1.suikei_type_code AS suikei_type_code, 
                    IV1.suikei_type_name AS suikei_type_name, 
                    IV1.kasen_code AS kasen_code, 
                    IV1.kasen_name AS kasen_name, 
                    IV1.kasen_type_code AS kasen_type_code, 
                    IV1.kasen_type_name AS kasen_type_name, 
                    IV1.gradient_code AS gradient_code, 
                    IV1.gradient_name AS gradient_name, 
                    
                    CASE WHEN (IV1.residential_area) IS NULL THEN NULL ELSE CAST(IV1.residential_area AS NUMERIC(20,10)) END AS residential_area, 
                    CASE WHEN (IV1.agricultural_area) IS NULL THEN NULL ELSE CAST(IV1.agricultural_area AS NUMERIC(20,10)) END AS agricultural_area, 
                    CASE WHEN (IV1.underground_area) IS NULL THEN NULL ELSE CAST(IV1.underground_area AS NUMERIC(20,10)) END AS underground_area, 
                    IV1.kasen_kaigan_code AS kasen_kaigan_code, 
                    IV1.kasen_kaigan_name AS kasen_kaigan_name, 
                    CASE WHEN (IV1.crop_damage) IS NULL THEN NULL ELSE CAST(IV1.crop_damage AS NUMERIC(20,10)) END AS crop_damage, 
                    IV1.weather_id AS weather_id, 
                    IV1.weather_name AS weather_name, 
                    IV1.upload_file_path AS upload_file_path, 
                    IV1.upload_file_name AS upload_file_name, 
                    IV1.summary_file_path AS summary_file_path, 
                    IV1.summary_file_name AS summary_file_name, 
                    
                    IV1.building_code AS building_code, 
                    IV1.building_name AS building_Name, 
                    IV1.underground_code AS underground_code, 
                    IV1.underground_name AS underground_name, 
                    IV1.flood_sediment_code AS flood_sediment_code, 
                    IV1.flood_sediment_name AS flood_sediment_name, 
                    
                    CASE WHEN (IV1.building_lv00) IS NULL THEN NULL ELSE CAST(IV1.building_lv00 AS NUMERIC(20,10)) END AS building_lv00, 
                    CASE WHEN (IV1.building_lv01_49) IS NULL THEN NULL ELSE CAST(IV1.building_lv01_49 AS NUMERIC(20,10)) END AS building_lv01_49, 
                    CASE WHEN (IV1.building_lv50_99) IS NULL THEN NULL ELSE CAST(IV1.building_lv50_99 AS NUMERIC(20,10)) END AS building_lv50_99, 
                    CASE WHEN (IV1.building_lv100) IS NULL THEN NULL ELSE CAST(IV1.building_lv100 AS NUMERIC(20,10)) END AS building_lv100, 
                    CASE WHEN (IV1.building_half) IS NULL THEN NULL ELSE CAST(IV1.building_half AS NUMERIC(20,10)) END AS building_half, 
                    CASE WHEN (IV1.building_full) IS NULL THEN NULL ELSE CAST(IV1.building_full AS NUMERIC(20,10)) END AS building_full, 
                    
                    CASE WHEN (IV1.floor_area) IS NULL THEN NULL ELSE CAST(IV1.floor_area AS NUMERIC(20,10)) END AS floor_area, 
                    CASE WHEN (IV1.family) IS NULL THEN NULL ELSE CAST(IV1.family AS NUMERIC(20,10)) END AS family, 
                    CASE WHEN (IV1.office) IS NULL THEN NULL ELSE CAST(IV1.office AS NUMERIC(20,10)) END AS office, 
                    
                    CASE WHEN (IV1.floor_area_lv00) IS NULL THEN NULL ELSE CAST(IV1.floor_area_lv00 AS NUMERIC(20,10)) END AS floor_area_lv00, 
                    CASE WHEN (IV1.floor_area_lv01_49) IS NULL THEN NULL ELSE CAST(IV1.floor_area_lv01_49 AS NUMERIC(20,10)) END AS floor_area_lv01_49, 
                    CASE WHEN (IV1.floor_area_lv50_99) IS NULL THEN NULL ELSE CAST(IV1.floor_area_lv50_99 AS NUMERIC(20,10)) END AS floor_area_lv50_99, 
                    CASE WHEN (IV1.floor_area_lv100) IS NULL THEN NULL ELSE CAST(IV1.floor_area_lv100 AS NUMERIC(20,10)) END AS floor_area_lv100, 
                    CASE WHEN (IV1.floor_area_half) IS NULL THEN NULL ELSE CAST(IV1.floor_area_half AS NUMERIC(20,10)) END AS floor_area_half, 
                    CASE WHEN (IV1.floor_area_full) IS NULL THEN NULL ELSE CAST(IV1.floor_area_full AS NUMERIC(20,10)) END AS floor_area_full, 
                    
                    CASE WHEN (IV1.family_lv00) IS NULL THEN NULL ELSE CAST(IV1.family_lv00 AS NUMERIC(20,10)) END AS family_lv00, 
                    CASE WHEN (IV1.family_lv01_49) IS NULL THEN NULL ELSE CAST(IV1.family_lv01_49 AS NUMERIC(20,10)) END AS family_lv01_49, 
                    CASE WHEN (IV1.family_lv50_99) IS NULL THEN NULL ELSE CAST(IV1.family_lv50_99 AS NUMERIC(20,10)) END AS family_lv50_99, 
                    CASE WHEN (IV1.family_lv100) IS NULL THEN NULL ELSE CAST(IV1.family_lv100 AS NUMERIC(20,10)) END AS family_lv100, 
                    CASE WHEN (IV1.family_half) IS NULL THEN NULL ELSE CAST(IV1.family_half AS NUMERIC(20,10)) END AS family_half, 
                    CASE WHEN (IV1.family_full) IS NULL THEN NULL ELSE CAST(IV1.family_full AS NUMERIC(20,10)) END AS family_full, 
                    
                    CASE WHEN (IV1.office_lv00) IS NULL THEN NULL ELSE CAST(IV1.office_lv00 AS NUMERIC(20,10)) END AS office_lv00, 
                    CASE WHEN (IV1.office_lv01_49) IS NULL THEN NULL ELSE CAST(IV1.office_lv01_49 AS NUMERIC(20,10)) END AS office_lv01_49, 
                    CASE WHEN (IV1.office_lv50_99) IS NULL THEN NULL ELSE CAST(IV1.office_lv50_99 AS NUMERIC(20,10)) END AS office_lv50_99, 
                    CASE WHEN (IV1.office_lv100) IS NULL THEN NULL ELSE CAST(IV1.office_lv100 AS NUMERIC(20,10)) END AS office_lv100, 
                    CASE WHEN (IV1.office_half) IS NULL THEN NULL ELSE CAST(IV1.office_half AS NUMERIC(20,10)) END AS office_half, 
                    CASE WHEN (IV1.office_full) IS NULL THEN NULL ELSE CAST(IV1.office_full AS NUMERIC(20,10)) END AS office_full, 
                    
                    CASE WHEN (IV1.farmer_fisher_lv00) IS NULL THEN NULL ELSE CAST(IV1.farmer_fisher_lv00 AS NUMERIC(20,10)) END AS farmer_fisher_lv00, 
                    CASE WHEN (IV1.farmer_fisher_lv01_49) IS NULL THEN NULL ELSE CAST(IV1.farmer_fisher_lv01_49 AS NUMERIC(20,10)) END AS farmer_fisher_lv01_49, 
                    CASE WHEN (IV1.farmer_fisher_lv50_99) IS NULL THEN NULL ELSE CAST(IV1.farmer_fisher_lv50_99 AS NUMERIC(20,10)) END AS farmer_fisher_lv50_99, 
                    CASE WHEN (IV1.farmer_fisher_lv100) IS NULL THEN NULL ELSE CAST(IV1.farmer_fisher_lv100 AS NUMERIC(20,10)) END AS farmer_fisher_lv100, 
                    CASE WHEN (IV1.farmer_fisher_full) IS NULL THEN NULL ELSE CAST(IV1.farmer_fisher_full AS NUMERIC(20,10)) END AS farmer_fisher_full, 
                    
                    CASE WHEN (IV1.employee_lv00) IS NULL THEN NULL ELSE CAST(IV1.employee_lv00 AS NUMERIC(20,10)) END AS employee_lv00, 
                    CASE WHEN (IV1.employee_lv01_49) IS NULL THEN NULL ELSE CAST(IV1.employee_lv01_49 AS NUMERIC(20,10)) END AS employee_lv01_49, 
                    CASE WHEN (IV1.employee_lv50_99) IS NULL THEN NULL ELSE CAST(IV1.employee_lv50_99 AS NUMERIC(20,10)) END AS employee_lv50_99, 
                    CASE WHEN (IV1.employee_lv100) IS NULL THEN NULL ELSE CAST(IV1.employee_lv100 AS NUMERIC(20,10)) END AS employee_lv100, 
                    CASE WHEN (IV1.employee_full) IS NULL THEN NULL ELSE CAST(IV1.employee_full AS NUMERIC(20,10)) END AS employee_full, 
                    
                    IV1.industry_code AS industry_code, 
                    IV1.industry_name AS industry_name, 
                    IV1.usage_code AS usage_code, 
                    IV1.usage_name AS usage_name, 
                    IV1.comment AS comment, 
                    TO_CHAR(timezone('JST', IV1.committed_at::timestamptz), 'yyyy/mm/dd HH24:MI') AS committed_at, 
                    TO_CHAR(timezone('JST', IV1.deleted_at::timestamptz), 'yyyy/mm/dd HH24:MI') AS deleted_at 
                    
                FROM IPPAN_VIEW IV1 
                WHERE 
                    IV1.suigai_id=%s
                ORDER BY CAST (IV1.ippan_id AS INTEGER)""", [trigger_list[0].suigai_id, ])

            ###################################################################
            ### EXCELファイル入出力処理(0040)
            ###################################################################
            print_log('[DEBUG] P0900Action.action_A04_verify_idb_by_diff_method.handle()関数 STEP 5/12.', 'DEBUG')
            upload_file_path = trigger_list[0].upload_file_path
            output_file_path = trigger_list[0].upload_file_path + '_output'
            template_file_path = 'static/template_ippan_chosa.xlsx'
            
            wb_input = openpyxl.load_workbook(upload_file_path)
            wb_output = openpyxl.load_workbook(template_file_path, keep_vba=False)
            
            ws_output = wb_output["IPPAN"]
            ws_output.title = suigai_list[0].suigai_name

            ###############################################################
            ### EXCEL入出力処理(0050)
            ### EXCELのヘッダ部のセルに、DBから取得した入力データ_ヘッダ部分の値を埋め込む。
            ###############################################################
            print_log('[DEBUG] P0900Action.action_A04_verify_idb_by_diff_method.handle()関数 STEP 6/12.', 'DEBUG')
            if suigai_list:
                if suigai_list[0].ken_name is not None and suigai_list[0].ken_code is not None and len(str(suigai_list[0].ken_name)) > 0 and len(str(suigai_list[0].ken_code)) > 0:
                    ws_output.cell(row=7, column=2).value = str(suigai_list[0].ken_name) + ":" + str(suigai_list[0].ken_code)

                if suigai_list[0].city_name is not None and suigai_list[0].city_code is not None and len(str(suigai_list[0].city_name)) > 0 and len(str(suigai_list[0].city_code)) > 0:
                    ws_output.cell(row=7, column=3).value = str(suigai_list[0].city_name) + ":" + str(suigai_list[0].city_code)

                if suigai_list[0].begin_date is not None and len(str(suigai_list[0].begin_date)) > 0:
                    ws_output.cell(row=7, column=4).value = str(suigai_list[0].begin_date)

                if suigai_list[0].end_date is not None and len(str(suigai_list[0].end_date)) > 0:
                    ws_output.cell(row=7, column=5).value = str(suigai_list[0].end_date)

                if suigai_list[0].cause_1_name is not None and suigai_list[0].cause_1_code is not None and len(str(suigai_list[0].cause_1_name)) > 0 and len(str(suigai_list[0].cause_1_code)) > 0:
                    ws_output.cell(row=7, column=6).value = str(suigai_list[0].cause_1_name) + ":" + str(suigai_list[0].cause_1_code)

                if suigai_list[0].cause_2_name is not None and suigai_list[0].cause_2_code is not None and len(str(suigai_list[0].cause_2_name)) > 0 and len(str(suigai_list[0].cause_2_code)) > 0:
                    ws_output.cell(row=7, column=7).value = str(suigai_list[0].cause_2_name) + ":" + str(suigai_list[0].cause_2_code)

                if suigai_list[0].cause_3_name is not None and suigai_list[0].suigai_3_code is not None and len(str(suigai_list[0].cause_3_name)) > 0 and len(str(suigai_list[0].cause_3_code)) > 0:
                    ws_output.cell(row=7, column=8).value = str(suigai_list[0].cause_3_name) + ":" + str(suigai_list[0].cause_3_code)

                if suigai_list[0].area_name is not None and suigai_list[0].area_id is not None and len(str(suigai_list[0].area_name)) > 0 and len(str(suigai_list[0].area_id)) > 0:
                    ws_output.cell(row=7, column=9).value = str(suigai_list[0].area_name) + ":" + str(suigai_list[0].area_id)

                if suigai_list[0].suikei_name is not None and suigai_list[0].suikei_code is not None and len(str(suigai_list[0].suikei_name)) > 0 and len(str(suigai_list[0].suikei_code)) > 0:
                    ws_output.cell(row=10, column=2).value = str(suigai_list[0].suikei_name) + ":" + str(suigai_list[0].suikei_code)

                if suigai_list[0].suikei_type_name is not None and suigai_list[0].suikei_type_code is not None and len(str(suigai_list[0].suikei_type_name)) > 0 and len(str(suigai_list[0].suikei_type_code)) > 0:
                    ws_output.cell(row=10, column=3).value = str(suigai_list[0].suikei_type_name) + ":" + str(suigai_list[0].suikei_type_code)

                if suigai_list[0].kasen_name is not None and suigai_list[0].kasen_code is not None and len(str(suigai_list[0].kasen_name)) > 0 and len(str(suigai_list[0].kasen_code)) > 0:
                    ws_output.cell(row=10, column=4).value = str(suigai_list[0].kasen_name) + ":" + str(suigai_list[0].kasen_code)

                if suigai_list[0].kasen_type_name is not None and suigai_list[0].kasen_type_code is not None and len(str(suigai_list[0].kasen_type_name)) > 0 and len(str(suigai_list[0].kasen_type_code)) > 0:
                    ws_output.cell(row=10, column=5).value = str(suigai_list[0].kasen_type_name) + ":" + str(suigai_list[0].kasen_type_code)

                if suigai_list[0].gradient_name is not None and suigai_list[0].gradient_code is not None and len(str(suigai_list[0].gradient_name)) > 0 and len(str(suigai_list[0].gradient_code)) > 0:
                    ws_output.cell(row=10, column=6).value = str(suigai_list[0].gradient_name) + ":" + str(suigai_list[0].gradient_code)

                if suigai_list[0].residential_area is not None and len(str(suigai_list[0].residential_area)) > 0:
                    ws_output.cell(row=14, column=2).value = str(suigai_list[0].residential_area)

                if suigai_list[0].agricultural_area is not None and len(str(suigai_list[0].agricultural_area)) > 0:
                    ws_output.cell(row=14, column=3).value = str(suigai_list[0].agricultural_area)

                if suigai_list[0].underground_area is not None and len(str(suigai_list[0].underground_area)) > 0:
                    ws_output.cell(row=14, column=4).value = str(suigai_list[0].underground_area)

                if suigai_list[0].kasen_kaigan_name is not None and suigai_list[0].kasen_kaigan_code is not None and len(str(suigai_list[0].kasen_kaigan_name)) > 0 and len(str(suigai_list[0].kasen_kaigan_code)) > 0:
                    ws_output.cell(row=14, column=6).value = str(suigai_list[0].kasen_kaigan_name) + ":" + str(suigai_list[0].kasen_kaigan_code)

                if suigai_list[0].crop_damage is not None and len(str(suigai_list[0].crop_damage)) > 0:
                    ws_output.cell(row=14, column=8).value = str(suigai_list[0].crop_damage)

                if suigai_list[0].weather_name is not None and suigai_list[0].weather_id is not None and len(str(suigai_list[0].weather_name)) > 0 and len(str(suigai_list[0].weather_id)) > 0:
                    ws_output.cell(row=14, column=10).value = str(suigai_list[0].weather_name) + ":" + str(suigai_list[0].weather_id)

                if suigai_list[0].suigai_id is not None and len(str(suigai_list[0].suigai_id)) > 0:
                    ws_output.cell(row=3, column=28).value = suigai_list[0].suigai_id
                    
            ###############################################################
            ### EXCEL入出力処理(0060)
            ### EXCELの一覧部のセルに、DBから取得した入力データ_一覧表部分の値を埋め込む。
            ###############################################################
            print_log('[DEBUG] P0900Action.action_A04_verify_idb_by_diff_method.handle()関数 STEP 7/12.', 'DEBUG')
            if ippan_list:
                for i, ippan in enumerate(ippan_list):
                    if ippan.ippan_name is not None and len(str(ippan.ippan_name)) > 0:
                        ws_output.cell(row=i+20, column=2).value = str(ippan.ippan_name)
                    
                    if ippan.building_name is not None and ippan.building_code is not None and len(str(ippan.building_name)) > 0 and len(str(ippan.building_code)) > 0:
                        ws_output.cell(row=i+20, column=3).value = str(ippan.building_name) + ":" + str(ippan.building_code)
                        
                    if ippan.underground_name is not None and ippan.underground_code is not None and len(str(ippan.underground_name)) > 0 and len(str(ippan.underground_code)) > 0:
                        ws_output.cell(row=i+20, column=4).value = str(ippan.underground_name) + ":" + str(ippan.underground_code)
                        
                    if ippan.flood_sediment_name is not None and ippan.flood_sediment_code is not None and len(str(ippan.flood_sediment_name)) > 0 and len(str(ippan.flood_sediment_code)) > 0:
                        ws_output.cell(row=i+20, column=5).value = str(ippan.flood_sediment_name) + ":" + str(ippan.flood_sediment_code)
                        
                    if ippan.building_lv00 is not None and len(str(ippan.building_lv00)) > 0:
                        ws_output.cell(row=i+20, column=6).value = ippan.building_lv00

                    if ippan.building_lv01_49 is not None and len(str(ippan.building_lv01_49)) > 0:                    
                        ws_output.cell(row=i+20, column=7).value = ippan.building_lv01_49
                    
                    if ippan.building_lv50_99 is not None and len(str(ippan.building_lv50_99)) > 0:
                        ws_output.cell(row=i+20, column=8).value = ippan.building_lv50_99

                    if ippan.building_lv100 is not None and len(str(ippan.building_lv100)) > 0:                    
                        ws_output.cell(row=i+20, column=9).value = ippan.building_lv100

                    if ippan.building_half is not None and len(str(ippan.building_half)) > 0:                    
                        ws_output.cell(row=i+20, column=10).value = ippan.building_half

                    if ippan.building_full is not None and len(str(ippan.building_full)) > 0:                    
                        ws_output.cell(row=i+20, column=11).value = ippan.building_full
                    
                    if ippan.floor_area is not None and len(str(ippan.floor_area)) > 0:
                        ws_output.cell(row=i+20, column=12).value = ippan.floor_area
                    
                    if ippan.family is not None and len(str(ippan.family)) > 0:
                        ws_output.cell(row=i+20, column=13).value = ippan.family

                    if ippan.office is not None and len(str(ippan.office))  > 0:                   
                        ws_output.cell(row=i+20, column=14).value = ippan.office

                    if ippan.farmer_fisher_lv00 is not None and len(str(ippan.farmer_fisher_lv00)) > 0:
                        ws_output.cell(row=i+20, column=15).value = ippan.farmer_fisher_lv00

                    if ippan.farmer_fisher_lv01_49 is not None and len(str(ippan.farmer_fisher_lv01_49)) > 0:
                        ws_output.cell(row=i+20, column=16).value = ippan.farmer_fisher_lv01_49
                    
                    if ippan.farmer_fisher_lv50_99 is not None and len(str(ippan.farmer_fisher_lv50_99)) > 0:
                        ws_output.cell(row=i+20, column=17).value = ippan.farmer_fisher_lv50_99

                    if ippan.farmer_fisher_lv100 is not None and len(str(ippan.farmer_fisher_lv100)) > 0:
                        ws_output.cell(row=i+20, column=18).value = ippan.farmer_fisher_lv100

                    if ippan.farmer_fisher_full is not None and len(str(ippan.farmer_fisher_full)) > 0:
                        ws_output.cell(row=i+20, column=19).value = ippan.farmer_fisher_full

                    if ippan.employee_lv00 is not None and len(str(ippan.employee_lv00)) > 0:
                        ws_output.cell(row=i+20, column=20).value = ippan.employee_lv00
                    
                    if ippan.employee_lv01_49 is not None and len(str(ippan.employee_lv01_49)) > 0:
                        ws_output.cell(row=i+20, column=21).value = ippan.employee_lv01_49
                    
                    if ippan.employee_lv50_99 is not None and len(str(ippan.employee_lv50_99)) > 0:
                        ws_output.cell(row=i+20, column=22).value = ippan.employee_lv50_99
                    
                    if ippan.employee_lv100 is not None and len(str(ippan.employee_lv100)) > 0:
                        ws_output.cell(row=i+20, column=23).value = ippan.employee_lv100
                    
                    if ippan.employee_full is not None and len(str(ippan.employee_full)) > 0:
                        ws_output.cell(row=i+20, column=24).value = ippan.employee_full
                    
                    if ippan.industry_name is not None and ippan.industry_code is not None and len(str(ippan.industry_name)) > 0 and len(str(ippan.industry_code)) > 0:
                        ws_output.cell(row=i+20, column=25).value = str(ippan.industry_name) + ":" + str(ippan.industry_code)
                        
                    if ippan.usage_name is not None and ippan.usage_code is not None and len(str(ippan.usage_name)) > 0 and len(str(ippan.usage_code)) > 0:
                        ws_output.cell(row=i+20, column=26).value = str(ippan.usage_name) + ":" + str(ippan.usage_code)
                        
                    if ippan.comment is not None and len(str(ippan.comment)) > 0:
                        ws_output.cell(row=i+20, column=27).value = ippan.comment

                    if ippan.ippan_id is not None and len(str(ippan.ippan_id)) > 0:
                        ws_output.cell(row=i+20, column=28).value = ippan.ippan_id

            wb_output.save(output_file_path)
    
            ###################################################################
            ### EXCELファイル入出力処理(0070)
            ###################################################################
            print_log('[DEBUG] P0900Action.action_A04_verify_idb_by_diff_method.handle()関数 STEP 8/12.', 'DEBUG')
            ws_input = None
            for ws_temp in wb_input.worksheets:
                if str(ws_temp.title) == str(suigai_list[0].suigai_name):
                    ws_input = ws_temp
                    break
                
            if ws_input is None:
                print_log('[ERROR] P0900Action.action_A04_verify_idb_by_diff_method.handle()関数 ws_input is None', 'ERROR')
                print_log('[ERROR] P0900Action.action_A04_verify_idb_by_diff_method.handle()関数でエラーが発生しました。', 'ERROR')
                print_log('[ERROR] P0900Action.action_A04_verify_idb_by_diff_method.handle()関数が異常終了しました。', 'ERROR')
                return 8

            max_row_temp = 1
            for i in range(ws_input.max_row + 1, 1, -1):
                if ws_input.cell(row=i, column=2).value is None:
                    pass
                else:
                    max_row_temp = i
                    break
                    
            ws_max_row_input = max_row_temp
            print_log('[DEBUG] P0900Action.action_A04_verify_idb_by_diff_method.handle()関数 ws_max_row_input = {}'.format(ws_max_row_input), 'DEBUG')

            max_row_temp = 1
            for i in range(ws_output.max_row + 1, 1, -1):
                if ws_output.cell(row=i, column=2).value is None:
                    pass
                else:
                    max_row_temp = i
                    break
                    
            ws_max_row_output = max_row_temp
            print_log('[DEBUG] P0900Action.action_A04_verify_idb_by_diff_method.handle()関数 ws_max_row_output = {}'.format(ws_max_row_output), 'DEBUG')
            
            if ws_max_row_input > ws_max_row_output:
                ws_max_row = ws_max_row_input
            else:
                ws_max_row = ws_max_row_output

            print_log('[DEBUG] P0900Action.action_A04_verify_idb_by_diff_method.handle()関数 ws_max_row = {}'.format(ws_max_row), 'DEBUG')

            if ws_max_row < 20:
                print_log('[ERROR] P0900Action.action_A04_verify_idb_by_diff_method.handle()関数 ws_max_row = {}'.format(ws_max_row), 'ERROR')
                print_log('[ERROR] P0900Action.action_A04_verify_idb_by_diff_method.handle()関数でエラーが発生しました。', 'ERROR')
                print_log('[ERROR] P0900Action.action_A04_verify_idb_by_diff_method.handle()関数が異常終了しました。', 'ERROR')
                return 8

            ###################################################################
            ### EXCELファイル入出力処理(0080)
            ### 検証の心臓部
            ###################################################################
            print_log('[DEBUG] P0900Action.action_A04_verify_idb_by_diff_method.handle()関数 STEP 9/12.', 'DEBUG')
            success_count = 0
            failure_count = 0
            
            ### 7行目
            for i in [2, 3, 4, 5, 6, 7, 8, 9]:
                print_log('[DEBUG] P0900Action.action_A04_verify_idb_by_diff_method.handle()関数 ws_input.cell.value = {}'.format(ws_input.cell(row=7, column=i).value), 'DEBUG')
                print_log('[DEBUG] P0900Action.action_A04_verify_idb_by_diff_method.handle()関数 ws_output.cell.value = {}'.format(ws_output.cell(row=7, column=i).value), 'DEBUG')
                if str(ws_input.cell(row=7, column=i).value) != str(ws_output.cell(row=7, column=i).value):
                    print_log('[WARN] {0} {1} {2} {3} {4} {5}'.format(ws_input.title, ws_output.title, 7, i, ws_input.cell(row=7, column=i).value, ws_output.cell(row=7, column=i).value), 'WARN')
                    failure_count += 1
                else:
                    success_count += 1
            
            ### 10行目
            for i in [2, 3, 4, 5, 6]:
                print_log('[DEBUG] P0900Action.action_A04_verify_idb_by_diff_method.handle()関数 ws_input.cell.value = {}'.format(ws_input.cell(row=10, column=i).value), 'DEBUG')
                print_log('[DEBUG] P0900Action.action_A04_verify_idb_by_diff_method.handle()関数 ws_output.cell.value = {}'.format(ws_output.cell(row=10, column=i).value), 'DEBUG')
                if str(ws_input.cell(row=10, column=i).value) != str(ws_output.cell(row=10, column=i).value):
                    print_log('[WARN] {0} {1} {2} {3} {4} {5}'.format(ws_input.title, ws_output.title, 10, i, ws_input.cell(row=10, column=i).value, ws_output.cell(row=10, column=i).value), 'WARN')
                    failure_count += 1
                else:
                    success_count += 1

            ### 14行目
            for i in [2, 3, 4, 8]:
                print_log('[DEBUG] P0900Action.action_A04_verify_idb_by_diff_method.handle()関数 ws_input.cell.value = {}'.format(ws_input.cell(row=14, column=i).value), 'DEBUG')
                print_log('[DEBUG] P0900Action.action_A04_verify_idb_by_diff_method.handle()関数 ws_output.cell.value = {}'.format(ws_output.cell(row=14, column=i).value), 'DEBUG')
                if ws_input.cell(row=14, column=i).value == None and ws_output.cell(row=14, column=i).value == None:
                    success_count += 1
                elif ws_input.cell(row=14, column=i).value == None and ws_output.cell(row=14, column=i).value != None:
                    print_log('[WARN] {0} {1} {2} {3} {4} {5}'.format(ws_input.title, ws_output.title, 14, i, ws_input.cell(row=14, column=i).value, ws_output.cell(row=14, column=i).value), 'WARN')
                    failure_count += 1
                elif ws_input.cell(row=14, column=i).value != None and ws_output.cell(row=14, column=i).value == None:
                    print_log('[WARN] {0} {1} {2} {3} {4} {5}'.format(ws_input.title, ws_output.title, 14, i, ws_input.cell(row=14, column=i).value, ws_output.cell(row=14, column=i).value), 'WARN')
                    failure_count += 1
                else:
                    if abs(float(ws_input.cell(row=14, column=i).value) - float(ws_output.cell(row=14, column=i).value)) > 1e-5:
                        print_log('[WARN] {0} {1} {2} {3} {4} {5}'.format(ws_input.title, ws_output.title, 14, i, ws_input.cell(row=14, column=i).value, ws_output.cell(row=14, column=i).value), 'WARN')
                        failure_count += 1
                    else:
                        success_count += 1

            for i in [6, 10]:
                print_log('[DEBUG] P0900Action.action_A04_verify_idb_by_diff_method.handle()関数 ws_input.cell.value = {}'.format(ws_input.cell(row=14, column=i).value), 'DEBUG')
                print_log('[DEBUG] P0900Action.action_A04_verify_idb_by_diff_method.handle()関数 ws_output.cell.value = {}'.format(ws_output.cell(row=14, column=i).value), 'DEBUG')
                if str(ws_input.cell(row=14, column=i).value) != str(ws_output.cell(row=14, column=i).value):
                    print_log('[WARN] {0} {1} {2} {3} {4} {5}'.format(ws_input.title, ws_output.title, 14, i, ws_input.cell(row=14, column=i).value, ws_output.cell(row=14, column=i).value), 'WARN')
                    failure_count += 1
                else:
                    success_count += 1

            ### 20行目            
            ### TO-DO TODO TO_DO
            for j in range(20, ws_max_row + 1):
                for k in range(1, 6):
                    print_log('[DEBUG] P0900Action.action_A04_verify_idb_by_diff_method.handle()関数 ws_input.cell.value = {}'.format(ws_input.cell(row=j, column=k).value), 'DEBUG')
                    print_log('[DEBUG] P0900Action.action_A04_verify_idb_by_diff_method.handle()関数 ws_output.cell.value = {}'.format(ws_output.cell(row=j, column=k).value), 'DEBUG')
                    if str(ws_input.cell(row=j, column=k).value) != str(ws_output.cell(row=j, column=k).value):
                        print_log('[WARN] {0} {1} {2} {3} {4} {5}'.format(ws_input.title, ws_output.title, j, k, ws_input.cell(row=j, column=k).value, ws_output.cell(row=j, column=k).value), 'WARN')
                        failure_count += 1
                    else:
                        success_count += 1

                for k in range(6, 26):
                    print_log('[DEBUG] P0900Action.action_A04_verify_idb_by_diff_method.handle()関数 ws_input.cell.value = {}'.format(ws_input.cell(row=j, column=k).value), 'DEBUG')
                    print_log('[DEBUG] P0900Action.action_A04_verify_idb_by_diff_method.handle()関数 ws_output.cell.value = {}'.format(ws_output.cell(row=j, column=k).value), 'DEBUG')
                    if ws_input.cell(row=j, column=k).value == None and ws_output.cell(row=j, column=k).value == None:
                        success_count += 1
                    elif ws_input.cell(row=j, column=k).value == None and ws_output.cell(row=j, column=k).value != None:
                        print_log('[WARN] {0} {1} {2} {3} {4} {5}'.format(ws_input.title, ws_output.title, j, k, ws_input.cell(row=j, column=k).value, ws_output.cell(row=j, column=k).value), 'WARN')
                        failure_count += 1
                    elif ws_input.cell(row=j, column=k).value != None and ws_output.cell(row=j, column=k).value == None:
                        print_log('[WARN] {0} {1} {2} {3} {4} {5}'.format(ws_input.title, ws_output.title, j, k, ws_input.cell(row=j, column=k).value, ws_output.cell(row=j, column=k).value), 'WARN')
                        failure_count += 1
                    else:
                        if abs(float(ws_input.cell(row=j, column=k).value) - float(ws_output.cell(row=j, column=k).value)) > 1e-5:
                            print_log('[WARN] {0} {1} {2} {3} {4} {5}'.format(ws_input.title, ws_output.title, j, k, ws_input.cell(row=j, column=k).value, ws_output.cell(row=j, column=k).value), 'WARN')
                            failure_count += 1
                        else:
                            success_count += 1

                for k in range(26, 28):
                    print_log('[DEBUG] P0900Action.action_A04_verify_idb_by_diff_method.handle()関数 ws_input.cell.value = {}'.format(ws_input.cell(row=j, column=k).value), 'DEBUG')
                    print_log('[DEBUG] P0900Action.action_A04_verify_idb_by_diff_method.handle()関数 ws_output.cell.value = {}'.format(ws_output.cell(row=j, column=k).value), 'DEBUG')
                    if str(ws_input.cell(row=j, column=k).value) != str(ws_output.cell(row=j, column=k).value):
                        print_log('[WARN] {0} {1} {2} {3} {4} {5}'.format(ws_input.title, ws_output.title, j, k, ws_input.cell(row=j, column=k).value, ws_output.cell(row=j, column=k).value), 'WARN')
                        failure_count += 1
                    else:
                        success_count += 1

            ################################################################### 
            ### DBアクセス処理(0090)
            ### 当該トリガーの実行が終了したため、当該トリガーの状態、成功数、失敗数等を更新する。
            ################################################################### 
            print_log('[DEBUG] P0900Action.action_A04_verify_idb_by_diff_method.handle()関数 STEP 10/12.', 'DEBUG')
            ### 成功
            if failure_count == 0:
                print_log('[INFO] success_count = {}'.format(success_count), 'INFO')
                print_log('[INFO] failure_count = {}'.format(failure_count), 'INFO')
            ### 失敗
            else:
                print_log('[WARN] success_count = {}'.format(success_count), 'WARN')
                print_log('[WARN] failure_count = {}'.format(failure_count), 'WARN')
            
            connection_cursor = connection.cursor()
            try:
                connection_cursor.execute("""BEGIN""", [])
                
                ### 成功
                if failure_count == 0:
                    connection_cursor.execute("""
                        UPDATE TRIGGER SET 
                            status_code=%s, -- status_code
                            success_count=%s, -- success_count
                            failure_count=%s, -- failure_count
                            consumed_at=CURRENT_TIMESTAMP, 
                            integrity_ok=%s, -- integrity_ok
                            integrity_ng=%s  -- integrity_ng
                        WHERE 
                            trigger_id=%s -- trigger_id
                        """, [
                            'SUCCESS', ### status_code
                            success_count, ### success_count
                            failure_count, ### failure_count
                            '\n'.join(get_info_log()), ### integrity_ok
                            '\n'.join(get_warn_log()), ### integrity_ng
                            trigger_list[0].trigger_id, ### trigger_id
                        ])
                ### 失敗
                else:
                    connection_cursor.execute("""
                        UPDATE TRIGGER SET 
                            status_code=%s, -- status_code
                            success_count=%s, -- success_count
                            failure_count=%s, -- failure_count
                            consumed_at=CURRENT_TIMESTAMP, 
                            integrity_ok=%s, -- integrity_ok
                            integrity_ng=%s  -- integrity_ng
                        WHERE 
                            trigger_id=%s -- trigger_id
                        """, [
                            'FAILURE', ### status_code
                            success_count, ### success_count
                            failure_count, ### failure_count
                            '\n'.join(get_info_log()), ### integrity_ok
                            '\n'.join(get_warn_log()), ### integrity_ng
                            trigger_list[0].trigger_id, ### trigger_id
                        ])
    
                ###############################################################
                ### DBアクセス処理(0100)
                ### 当該トリガーの実行が終了したため、
                ### (1)成功の場合は、次のトリガーを発行する。
                ### (2)失敗の場合は、次のトリガーを発行しない。
                ############################################################### 
                print_log('[DEBUG] P0900Action.action_A04_verify_idb_by_diff_method.handle()関数 STEP 11/12.', 'DEBUG')
                ### 成功 -> 次のトリガー
                if failure_count == 0:
                    connection_cursor.execute("""
                        INSERT INTO TRIGGER (
                            trigger_id, suigai_id, action_code, status_code, success_count, failure_count, 
                            published_at, consumed_at, deleted_at, integrity_ok, integrity_ng, ken_code, city_code, 
                            download_file_path, download_file_name, upload_file_path, upload_file_name 
                        ) VALUES (
                            (SELECT CASE WHEN (MAX(trigger_id+1)) IS NULL THEN CAST(0 AS INTEGER) ELSE CAST(MAX(trigger_id+1) AS INTEGER) END AS trigger_id FROM TRIGGER), -- trigger_id 
                            %s, -- suigai_id 
                            %s, -- action_code 
                            %s, -- status_code 
                            %s, -- success_count 
                            %s, -- failure_count 
                            CURRENT_TIMESTAMP, -- published_at 
                            %s, -- consumed_at 
                            %s, -- deleted_at 
                            %s, -- integrity_ok 
                            %s, -- integrity_ng 
                            %s, -- ken_code 
                            %s, -- city_code 
                            %s, -- download_file_path 
                            %s, -- download_file_name 
                            %s, -- upload_file_path 
                            %s  -- upload_file_name 
                        )""", [
                            trigger_list[0].suigai_id, ### suigai_id 
                            'A05',  ### action_code 
                            None, ### status_code 
                            None, ### success_count 
                            None, ### failure_count 
                            None, ### consumed_at 
                            None, ### deleted_at 
                            None, ### integrity_ok 
                            None, ### integrity_ng 
                            trigger_list[0].ken_code,  ### ken_code 
                            trigger_list[0].city_code, ### city_code 
                            trigger_list[0].download_file_path, ### download_file_path 
                            trigger_list[0].download_file_name, ### download_file_name 
                            trigger_list[0].upload_file_path, ### upload_file_path 
                            trigger_list[0].upload_file_name, ### upload_file_name
                        ])
                ### transaction.commit()
                connection_cursor.execute("""COMMIT""", [])
            except:
                print_log('[ERROR] P0900Action.action_A04_verify_idb_by_diff_method.handle()関数 {}'.format(sys.exc_info()[0]), 'ERROR')
                ### connection_cursor.rollback()
                connection_cursor.execute("""ROLLBACK""", [])
            finally:
                connection_cursor.close()

            ###################################################################
            ### 戻り値セット処理(0110)
            ### 戻り値を戻す。
            ###################################################################
            print_log('[DEBUG] P0900Action.action_A04_verify_idb_by_diff_method.handle()関数 STEP 12/12.', 'DEBUG')
            print_log('[INFO] P0900Action.action_A04_verify_idb_by_diff_method.handle()関数が正常終了しました。', 'INFO')
            return 0
        except:
            print_log('[ERROR] P0900Action.action_A04_verify_idb_by_diff_method.handle()関数 {}'.format(sys.exc_info()[0]), 'ERROR')
            print_log('[ERROR] P0900Action.action_A04_verify_idb_by_diff_method.handle()関数でエラーが発生しました。', 'ERROR')
            print_log('[ERROR] P0900Action.action_A04_verify_idb_by_diff_method.handle()関数が異常終了しました。', 'ERROR')
            return 8
            