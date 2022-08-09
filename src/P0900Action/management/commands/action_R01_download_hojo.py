#!/usr/bin/env python3
# -*- coding: utf-8 -*-
###############################################################################
### ファイル名：P0900Action/management/commands/action_R01_download_hojo.py
### R01：ダウンロード
###############################################################################

###############################################################################
### 処理名：インポート処理
###############################################################################
from django.core.management.base import BaseCommand

import sys
from django.contrib.auth.decorators import login_required
from django.db import connection
from django.db import transaction
from django.http import Http404
from django.http import HttpResponse
from django.http import HttpResponseRedirect
from django.shortcuts import render
from django.template import loader
from django.views import generic
from django.views.generic import FormView
from django.views.generic.base import TemplateView

import openpyxl
from openpyxl.comments import Comment
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import PatternFill
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.writer.excel import save_virtual_workbook

from decimal import Decimal

from P0000Common.models import BUILDING                ### 1000: 建物区分
from P0000Common.models import KEN                     ### 1010: 都道府県
from P0000Common.models import CITY                    ### 1020: 市区町村
from P0000Common.models import KASEN_KAIGAN            ### 1030: 水害発生地点工種（河川海岸区分）
from P0000Common.models import SUIKEI                  ### 1040: 水系（水系・沿岸）
from P0000Common.models import SUIKEI_TYPE             ### 1050: 水系種別（水系・沿岸種別）
from P0000Common.models import KASEN                   ### 1060: 河川（河川・海岸）
from P0000Common.models import KASEN_TYPE              ### 1070: 河川種別（河川・海岸種別）
from P0000Common.models import CAUSE                   ### 1080: 水害原因
from P0000Common.models import UNDERGROUND             ### 1090: 地上地下区分
from P0000Common.models import USAGE                   ### 1100: 地下空間の利用形態
from P0000Common.models import FLOOD_SEDIMENT          ### 1110: 浸水土砂区分
from P0000Common.models import GRADIENT                ### 1120: 地盤勾配区分
from P0000Common.models import INDUSTRY                ### 1130: 産業分類

from P0000Common.models import HOUSE_ASSET             ### 2000: 家屋評価額
from P0000Common.models import HOUSE_RATE              ### 2010: 家屋被害率
from P0000Common.models import HOUSE_ALT               ### 2020: 家庭応急対策費_代替活動費
from P0000Common.models import HOUSE_CLEAN             ### 2030: 家庭応急対策費_清掃日数

from P0000Common.models import HOUSEHOLD_ASSET         ### 3000: 家庭用品自動車以外所有額
from P0000Common.models import HOUSEHOLD_RATE          ### 3010: 家庭用品自動車以外被害率

from P0000Common.models import CAR_ASSET               ### 4000: 家庭用品自動車所有額
from P0000Common.models import CAR_RATE                ### 4010: 家庭用品自動車被害率

from P0000Common.models import OFFICE_ASSET            ### 5000: 事業所資産額
from P0000Common.models import OFFICE_RATE             ### 5010: 事業所被害率
from P0000Common.models import OFFICE_SUSPEND          ### 5020: 事業所営業停止日数
from P0000Common.models import OFFICE_STAGNATE         ### 5030: 事業所営業停滞日数
from P0000Common.models import OFFICE_ALT              ### 5040: 事業所応急対策費_代替活動費

from P0000Common.models import FARMER_FISHER_ASSET     ### 6000: 農漁家資産額
from P0000Common.models import FARMER_FISHER_RATE      ### 6010: 農漁家被害率

from P0000Common.models import AREA                    ### 7000: 入力データ_水害区域
from P0000Common.models import WEATHER                 ### 7010: 入力データ_異常気象
from P0000Common.models import SUIGAI                  ### 7020: 入力データ_ヘッダ部分
from P0000Common.models import IPPAN                   ### 7030: 入力データ_一覧表部分
from P0000Common.models import IPPAN_VIEW              ### 7040: ビューデータ_一覧表部分

from P0000Common.models import IPPAN_SUMMARY           ### 8000: 集計データ

from P0000Common.models import ACTION                  ### 10000: アクション
from P0000Common.models import STATUS                  ### 10010: 状態
from P0000Common.models import TRIGGER                 ### 10020: トリガーメッセージ

from P0000Common.common import get_debug_log
from P0000Common.common import get_error_log
from P0000Common.common import get_info_log
from P0000Common.common import get_warn_log
from P0000Common.common import print_log
from P0000Common.common import reset_log

VLOOK_VALUE = [
    'B', 'G', 'L', 'Q', 'V', 'AA', 'AF', 'AK', 'AP', 'AU', 
    'AZ', 'BE', 'BJ', 'BO', 'BT', 'BY', 'CD', 'CI', 'CN', 'CS', 
    'CX', 'DC', 'DH', 'DM', 'DR', 'DW', 'EB', 'EG', 'EL', 'EQ', 
    'EV', 'FA', 'FF', 'FK', 'FP', 'FU', 'FZ', 'GE', 'GJ', 'GO', 
    'GT', 'GY', 'HD', 'HI', 'HN', 'HS', 'HX', 'IC', 'IH', 'IM', 
    'IR', 'IW', 'JB', 'JG', 'JL', 'JQ', 'JV', 'KA', 'KF', 'KK', 
    'KP', 'KU', 'KZ', 'LE', 'LJ', 'LO', 'LT', 'LY', 'MD', 'MI', 
    'MN', 'MS', 'MX', 'NC', 'NH', 'NM', 'NR', 'NW', 'OB', 'OG', 
    'OL', 'OQ', 'OV', 'PA', 'PF', 'PK', 'PP', 'PU', 'PZ', 'QE', 
    'QJ', 'QO', 'QT', 'QY', 'RD', 'RI', 'RN', 'RS', 'RX', 'SC', 
    'SH', 'SM', 'SR', 'SW', 'TB', 'TG', 'TL', 'TQ', 'TV', 'UA', 
    'UF', 'UK', 'UP', 'UU', 'UZ', 'VE', 'VJ', 'VO', 'VT', 'VY', 
    'WD', 'WI', 'WN', 'WS', 'WX', 'XC', 'XH', 'XM', 'XR', 'XW', 
    'YB', 'YG', 'YL', 'YQ', 'YV', 'ZA', 'ZF', 'ZK', 'ZP', 'ZU', 
    'ZZ', 'AAE', 'AAJ', 'AAO', 'AAT', 'AAY', 'ABD', 'ABI', 'ABN', 'ABS', 
    'ABX', 'ACC', 'ACH', 'ACM', 'ACR', 'ACW', 'ADB', 'ADG', 'ADL', 'ADQ', 
    'ADV', 'AEA', 'AEF', 'AEK', 'AEP', 'AEU', 'AEZ', 'AFE', 'AFJ', 'AFO', 
    'AFT', 'AFY', 'AGD', 'AGI', 'AGN', 'AGS', 'AGX', 'AHC', 'AHH', 'AHM', 
    'AHR', 'AHW', 'AIB', 'AIG', 'AIL', 'AIQ', 'AIV', 'AJA', 'AJF', 'AJK', 
    'AJP', 'AJU', 'AJZ', 'AKE', 'AKJ', 'AKO', 'AKT', 'AKY', 'ALD', 'ALI', 
    'ALN', 'ALS', 'ALX', 'AMC', 'AMH', 'AMM', 'AMR', 'AMW', 'ANB', 'ANG', 
    'ANL', 'ANQ', 'ANV', 'AOA', 'AOF', 'AOK', 'AOP', 'AOU', 'AOZ', 'APE', 
    'APJ', 'APO', 'APT', 'APY', 'AQD', 'AQI', 'AQN', 'AQS', 'AQX', 'ARC', 
    'ARH', 'ARM', 'ARR', 'ARW', 'ASB', 'ASG', 'ASL', 'ASQ', 'ASV', 'ATA', 
    'ATF', 'ATK', 'ATP', 'ATU', 'ATZ', 'AUE', 'AUJ', 'AUO', 'AUT', 'AUY', 
    'AVD', 'AVI', 'AVN', 'AVS', 'AVX', 'AWC', 'AWH', 'AWM', 'AWR', 'AWW', 
    'AXB', 'AXG', 'AXL', 'AXQ', 'AXV', 'AYA', 'AYF', 'AYK', 'AYP', 'AYU', 
    'AYZ', 'AZE', 'AZJ', 'AZO', 'AZT', 'AZY'
    ]

###############################################################################
### クラス名： Command
###############################################################################
class Command(BaseCommand):
    
    ###########################################################################
    ### 関数名： handle
    ### チェックOKの場合、トリガーの状態を成功に更新、消費日時をカレントに更新、新たなトリガーを生成する。
    ### チェックNGの場合、トリガーの状態を失敗に更新、消費日時をカレントに更新する。
    ### チェックNGの場合、当該水害IDについて、以降の処理を止めて、手動？で再実行、または、入力データから再登録するイメージである。
    ### 上記は、このアプリの共通の考え方とする。
    ###########################################################################
    def handle(self, *args, **options):
        try:
            ###################################################################
            ### 引数チェック処理(0000)
            ### コマンドラインからの引数をチェックする。
            ###################################################################
            reset_log()
            print_log('[INFO] P0900Action.action_R01_download_hojo.handle()関数が開始しました。', 'INFO')
            print_log('[DEBUG] P0900Action.action_R01_download_hojo.handle()関数 STEP 1/13.', 'DEBUG')

            ###################################################################
            ### DBアクセス処理(0010)
            ### TRIGGERテーブルからSUIGAI_IDのリストを取得する。
            ### トリガーメッセージにアクションが発行されているかを検索する。
            ### 処理対象の水害IDを取得する。
            ### トリガーメッセージにアクションが発行されていなければ、処理を終了する。
            ### ※ネストを浅くするために、処理対象外の場合、処理を終了させる。
            ###################################################################
            print_log('[DEBUG] P0900Action.action_R01_download_hojo.handle()関数 STEP 2/13.', 'DEBUG')
            trigger_list = TRIGGER.objects.raw("""
                SELECT 
                    * 
                FROM TRIGGER 
                WHERE 
                    action_code='R01' AND 
                    consumed_at IS NULL AND 
                    deleted_at IS NULL 
                ORDER BY CAST(ken_code AS INTEGER), CAST(city_code AS INTEGER) LIMIT 1""", [])
            
            if trigger_list is None:
                print_log('[INFO] P0900Action.action_R01_download_hojo.handle()関数が正常終了しました。', 'INFO')
                return 0
            
            if len(trigger_list) == 0:
                print_log('[INFO] P0900Action.action_R01_download_hojo.handle()関数が正常終了しました。', 'INFO')
                return 0
            
            print_log('[DEBUG] P0900Action.action_R01_download_hojo.handle()関数 trigger_list[0].trigger_id = {}'.format(trigger_list[0].trigger_id), 'DEBUG')
            print_log('[DEBUG] P0900Action.action_R01_download_hojo.handle()関数 trigger_list[0].suigai_id = {}'.format(trigger_list[0].suigai_id), 'DEBUG')
            print_log('[DEBUG] P0900Action.action_R01_download_hojo.handle()関数 trigger_list[0].city_code = {}'.format(trigger_list[0].city_code), 'DEBUG')
            print_log('[DEBUG] P0900Action.action_R01_download_hojo.handle()関数 trigger_list[0].ken_code = {}'.format(trigger_list[0].ken_code), 'DEBUG')
            print_log('[DEBUG] P0900Action.action_R01_download_hojo.handle()関数 trigger_list[0].download_file_path = {}'.format(trigger_list[0].download_file_path), 'DEBUG')
            print_log('[DEBUG] P0900Action.action_R01_download_hojo.handle()関数 trigger_list[0].download_file_name = {}'.format(trigger_list[0].download_file_name), 'DEBUG')
            print_log('[DEBUG] P0900Action.action_R01_download_hojo.handle()関数 trigger_list[0].upload_file_path = {}'.format(trigger_list[0].upload_file_path), 'DEBUG')
            print_log('[DEBUG] P0900Action.action_R01_download_hojo.handle()関数 trigger_list[0].upload_file_name = {}'.format(trigger_list[0].upload_file_name), 'DEBUG')
    
            ###################################################################
            ### DBアクセス処理(0020)
            ### マスタデータを取得する。
            ###################################################################
            print_log('[DEBUG] P0900Action.action_R01_download_hojo.handle()関数 STEP 3/13.', 'DEBUG')
            ### 1010: 都道府県シート
            ### 1020: 市区町村シート、連動プルダウン用、VLOOKUP用
            ### 1030: 水害発生地点工種（河川海岸区分）
            ### 1040: 水系（水系・沿岸）
            ### 1050: 水系種別（水系・沿岸種別）
            ### 1060: 河川（河川・海岸）、連動プルダウン用、VLOOKUP用
            ### 1070: 河川種別（河川・海岸種別）
            ### 7010: 入力データ_異常気象
            cities_list = []
            ken_list = KEN.objects.raw("""SELECT * FROM KEN ORDER BY CAST(KEN_CODE AS INTEGER)""", [])
            if ken_list:
                for i, ken in enumerate(ken_list):
                    cities_list.append(CITY.objects.raw("""SELECT * FROM CITY WHERE KEN_CODE=%s ORDER BY CAST(CITY_CODE AS INTEGER)""", [ken.ken_code,]))
    
            kasen_kaigan_list = KASEN_KAIGAN.objects.raw("""SELECT * FROM KASEN_KAIGAN ORDER BY CAST(KASEN_KAIGAN_CODE AS INTEGER)""", [])
            suikei_list = SUIKEI.objects.raw("""SELECT * FROM SUIKEI ORDER BY CAST(SUIKEI_CODE AS INTEGER)""", [])
            suikei_type_list = SUIKEI_TYPE.objects.raw("""SELECT * FROM SUIKEI_TYPE ORDER BY CAST(SUIKEI_TYPE_CODE AS INTEGER)""", [])
    
            kasens_list = []
            suikei_list = SUIKEI.objects.raw("""SELECT * FROM SUIKEI ORDER BY CAST(SUIKEI_CODE AS INTEGER)""", [])
            if suikei_list:
                for i, suikei in enumerate(suikei_list):
                    kasens_list.append(KASEN.objects.raw("""SELECT * FROM KASEN WHERE SUIKEI_CODE=%s ORDER BY CAST(KASEN_CODE AS INTEGER)""", [suikei.suikei_code,]))
    
            kasen_type_list = KASEN_TYPE.objects.raw("""SELECT * FROM KASEN_TYPE ORDER BY CAST(KASEN_TYPE_CODE AS INTEGER)""", [])
            weather_list = WEATHER.objects.raw("""SELECT * FROM WEATHER ORDER BY CAST(WEATHER_ID AS INTEGER)""", [])
            
            ###################################################################
            ### EXCEL入出力処理(0030)
            ### ダウンロード用ファイルパスをセットする。
            ###################################################################
            print_log('[DEBUG] P0900Action.action_R01_download_hojo.handle()関数 STEP 4/13.', 'DEBUG')
            template_file_path = 'static/template_hojo.xlsx'
            download_file_path = trigger_list[0].download_file_path
            download_file_name = trigger_list[0].download_file_name
            print_log('[DEBUG] P0900Action.action_R01_download_hojo.handle()関数 download_file_path = {}'.format(download_file_path), 'DEBUG')
            print_log('[DEBUG] P0900Action.action_R01_download_hojo.handle()関数 download_file_name = {}'.format(download_file_name), 'DEBUG')

            ###################################################################
            ### EXCEL入出力処理(0040)
            ### ワークシートを局所変数にセットする。
            ###################################################################
            print_log('[DEBUG] P0900Action.action_R01_download_hojo.handle()関数 STEP 5/13.', 'DEBUG')
            wb = openpyxl.load_workbook(template_file_path, keep_vba=False)
            
            ws_ken = wb["KEN"]
            ws_city = wb["CITY"]
            ws_kasen_kaigan = wb["KASEN_KAIGAN"]
            ws_suikei = wb["SUIKEI"]
            ws_suikei_type = wb["SUIKEI_TYPE"]
            ws_kasen = wb["KASEN"]
            ws_kasen_type = wb["KASEN_TYPE"]
            ws_weather = wb["WEATHER"]
            
            ws_city_vlook = wb["CITY_VLOOK"]
            ws_kasen_vlook = wb["KASEN_VLOOK"]

            ws_hojo = wb["HOJO"]

            ###################################################################
            ### EXCEL入出力処理(0050)
            ### 各EXCELシートで枠線を表示しないにセットする。
            ###################################################################
            print_log('[DEBUG] P0900Action.action_R01_download_hojo.handle()関数 STEP 6/13.', 'DEBUG')
            ws_ken.sheet_view.showGridLines = False
            ws_city.sheet_view.showGridLines = False
            ws_kasen_kaigan.sheet_view.showGridLines = False
            ws_suikei.sheet_view.showGridLines = False
            ws_suikei_type.sheet_view.showGridLines = False
            ws_kasen.sheet_view.showGridLines = False
            ws_kasen_type.sheet_view.showGridLines = False
            ws_weather.sheet_view.showGridLines = False
            
            ws_city_vlook.sheet_view.showGridLines = False
            ws_kasen_vlook.sheet_view.showGridLines = False
    
            ws_hojo.sheet_view.showGridLines = False

            ###################################################################
            ### EXCEL入出力処理(0060)
            ### マスタ用のシートに値をセットする。
            ###################################################################
            print_log('[DEBUG] P0900Action.action_R01_download_hojo.handle()関数 STEP 7/13.', 'DEBUG')
            ### 1010: 都道府県シート
            print("handle_7_2", flush=True)
            if ken_list:
                for i, ken in enumerate(ken_list):
                    ws_ken.cell(row=i+1, column=1).value = ken.ken_code
                    ws_ken.cell(row=i+1, column=2).value = str(ken.ken_name) + ":" + str(ken.ken_code)
                    ### ws_city_vlook.cell(row=j+1, column=1).value = str(ken.ken_name) + ":" + str(ken.ken_code)

            ### 1020: 市区町村シート
            print("handle_7_3", flush=True)
            cities_list = []
            if ken_list:
                for i, ken in enumerate(ken_list):
                    cities_list.append(CITY.objects.raw("""
                        SELECT 
                            * 
                        FROM CITY 
                        WHERE ken_code=%s 
                        ORDER BY CAST(city_code AS INTEGER)""", [ken.ken_code, ]))
    
            print("handle_7_4", flush=True)
            if cities_list:
                for i, cities in enumerate(cities_list):
                    if cities:
                        for j, city in enumerate(cities):
                            ws_city.cell(row=j+1, column=i*5+1).value = city.city_code
                            ws_city.cell(row=j+1, column=i*5+2).value = str(city.city_name) + ":" + str(city.city_code)
                            ws_city.cell(row=j+1, column=i*5+3).value = city.ken_code
                            ws_city.cell(row=j+1, column=i*5+4).value = city.city_population
                            ws_city.cell(row=j+1, column=i*5+5).value = city.city_area

            ### 1030: 水害発生地点工種（河川海岸区分）
            print("handle_7_5", flush=True)
            if kasen_kaigan_list:
                for i, kasen_kaigan in enumerate(kasen_kaigan_list):
                    ws_kasen_kaigan.cell(row=i+1, column=1).value = kasen_kaigan.kasen_kaigan_code
                    ws_kasen_kaigan.cell(row=i+1, column=2).value = str(kasen_kaigan.kasen_kaigan_name) + ":" + str(kasen_kaigan.kasen_kaigan_code)

            ### 1040: 水系（水系・沿岸）
            print("handle_7_6", flush=True)
            if suikei_list:
                for i, suikei in enumerate(suikei_list):
                    ws_suikei.cell(row=i+1, column=1).value = suikei.suikei_code
                    ws_suikei.cell(row=i+1, column=2).value = str(suikei.suikei_name) + ":" + str(suikei.suikei_code)
                    ws_suikei.cell(row=i+1, column=3).value = suikei.suikei_type_code

            ### 1050: 水系種別（水系・沿岸種別）
            print("handle_7_7", flush=True)
            if suikei_type_list:
                for i, suikei_type in enumerate(suikei_type_list):
                    ws_suikei_type.cell(row=i+1, column=1).value = suikei_type.suikei_type_code
                    ws_suikei_type.cell(row=i+1, column=2).value = str(suikei_type.suikei_type_name) + ":" + str(suikei_type.suikei_type_code)

            ### 1060: 河川（河川・海岸）、連動プルダウン用
            print("handle_7_8", flush=True)
            kasens_list = []
            if suikei_list:
                for i, suikei in enumerate(suikei_list):
                    kasens_list.append(KASEN.objects.raw("""
                        SELECT 
                            * 
                        FROM KASEN 
                        WHERE suikei_code=%s 
                        ORDER BY CAST(kasen_code AS INTEGER)""", [suikei.suikei_code, ]))
    
            print("handle_7_9", flush=True)
            if kasens_list:
                for i, kasens in enumerate(kasens_list):
                    if kasens:
                        for j, kasen in enumerate(kasens):
                            ws_kasen.cell(row=j+1, column=i*5+1).value = kasen.kasen_code
                            ws_kasen.cell(row=j+1, column=i*5+2).value = str(kasen.kasen_name) + ":" + str(kasen.kasen_code)
                            ws_kasen.cell(row=j+1, column=i*5+3).value = kasen.kasen_type_code
                            ws_kasen.cell(row=j+1, column=i*5+4).value = kasen.suikei_code

            ### 1070: 河川種別（河川・海岸種別）
            print("handle_7_10", flush=True)
            if kasen_type_list:
                for i, kasen_type in enumerate(kasen_type_list):
                    ws_kasen_type.cell(row=i+1, column=1).value = kasen_type.kasen_type_code
                    ws_kasen_type.cell(row=i+1, column=2).value = str(kasen_type.kasen_type_name) + ":" + str(kasen_type.kasen_type_code)
            
            ### 7010: 入力データ_異常気象
            print("handle_7_18", flush=True)
            if weather_list:
                for i, weather in enumerate(weather_list):
                    ws_weather.cell(row=i+1, column=1).value = weather.weather_id
                    ws_weather.cell(row=i+1, column=2).value = str(weather.weather_name) + ":" + str(weather.weather_id)

            ###################################################################
            ### EXCEL入出力処理(0070)
            ### VLOOKUP用のシートに値をセットする。
            ###################################################################
            print_log('[DEBUG] P0900Action.action_R01_download_hojo.handle()関数 STEP 8/13.', 'DEBUG')
            ### 1020: 市区町村VLOOKUP
            print("handle_8_1", flush=True)
            if ken_list and cities_list:
                for i, ken in enumerate(ken_list):
                    ws_city_vlook.cell(row=i+1, column=1).value = str(ken.ken_name) + ":" + str(ken.ken_code)
        
                for i, cities in enumerate(cities_list):
                    ws_city_vlook.cell(row=i+1, column=2).value = 'CITY!$' + VLOOK_VALUE[i] + '$1:$' + VLOOK_VALUE[i] + '$%d' % len(cities)
    
            ### 1060: 河川（河川・海岸）VLOOKUP
            print("handle_8_2", flush=True)
            if suikei_list and kasens_list:
                for i, suikei in enumerate(suikei_list):
                    ws_kasen_vlook.cell(row=i+1, column=1).value = str(suikei.suikei_name) + ":" + str(suikei.suikei_code)
    
                for i, kasens in enumerate(kasens_list):
                    ws_kasen_vlook.cell(row=i+1, column=2).value = 'KASEN!$' + VLOOK_VALUE[i] + '$1:$' + VLOOK_VALUE[i] + '$%d' % len(kasens)

            ###################################################################
            ### EXCEL入出力処理(0080)
            ### 入力データ用EXCELシートのキャプションに値をセットする。
            ###################################################################
            print_log('[DEBUG] P0900Action.action_R01_download_hojo.handle()関数 STEP 9/13.', 'DEBUG')
            ### ws_ippan.cell(row=6, column=1).value = 'NO.'
            ### ws_ippan.cell(row=6, column=2).value = '水系・沿岸名[全角]'
            ### ws_ippan.cell(row=6, column=3).value = '水系種別[全角]'
            ### ws_ippan.cell(row=6, column=4).value = '河川・海岸名[全角]'
            ### ws_ippan.cell(row=6, column=5).value = '河川種別[全角]'
            ### ws_ippan.cell(row=6, column=6).value = '都道府県名[全角]'
            ### ws_ippan.cell(row=6, column=7).value = '工事番号'
            ### ws_ippan.cell(row=6, column=8).value = ''
            ### ws_ippan.cell(row=6, column=9).value = '工事区分'
            ### ws_ippan.cell(row=6, column=10).value = '市区町村コード'
            ### ws_ippan.cell(row=6, column=11).value = '工種区分'
            ### ws_ippan.cell(row=6, column=12).value = ''
            ### ws_ippan.cell(row=6, column=13).value = '異常気象コード'
            ### ws_ippan.cell(row=6, column=14).value = '水害発生'
            ### ws_ippan.cell(row=6, column=15).value = ''
            ### ws_ippan.cell(row=6, column=16).value = '決定額(千円)'
            ### ws_ippan.cell(row=6, column=17).value = '備考'
            ### ws_ippan.cell(row=7, column=1).value = ''
            ### ws_ippan.cell(row=7, column=2).value = ''
            ### ws_ippan.cell(row=7, column=3).value = ''
            ### ws_ippan.cell(row=7, column=4).value = ''
            ### ws_ippan.cell(row=7, column=5).value = ''
            ### ws_ippan.cell(row=7, column=6).value = ''
            ### ws_ippan.cell(row=7, column=7).value = '工事番号'
            ### ws_ippan.cell(row=7, column=8).value = '枝番'
            ### ws_ippan.cell(row=7, column=9).value = ''
            ### ws_ippan.cell(row=7, column=10).value = ''
            ### ws_ippan.cell(row=7, column=11).value = ''
            ### ws_ippan.cell(row=7, column=12).value = ''
            ### ws_ippan.cell(row=7, column=13).value = ''
            ### ws_ippan.cell(row=7, column=14).value = '月'
            ### ws_ippan.cell(row=7, column=15).value = '日'
            ### ws_ippan.cell(row=7, column=16).value = ''
            ### ws_ippan.cell(row=7, column=17).value = ''
            
            ws_ippan.protection.enable()
            green_fill = PatternFill(fgColor='CCFFCC', patternType='solid')

            ###################################################################
            ### EXCEL入出力処理(0090)
            ### EXCELのセルに、値を埋め込む。
            ###################################################################
            print_log('[DEBUG] P0900Action.action_R01_download_hojo.handle()関数 STEP 10/13.', 'DEBUG')
            hojo_list = HOJO.objects.raw("""
                SELECT 
                    HJ1.hojo_id AS hojo_id, 
                    HJ1.hojo_name AS hojo_name, 
                    HJ1.hojo_file_id AS hojo_file_id, 
                    HF1.hojo_file_name AS hojo_file_name, 
                    HF1.upload_file_path AS upload_file_path, 
                    HF1.upload_file_name AS upload_file_name, 
                    HF1.summary_file_path AS summary_file_path, 
                    HF1.summary_file_name AS summary_file_name, 
                    HJ1.suikei_code AS suikei_code, 
                    SK1.suikei_name AS suikei_name, 
                    HJ1.kasen_code AS kasen_code, 
                    KA1.kasen_name AS kasen_name, 
                    HJ1.ken_code AS ken_code, 
                    KE1.ken_name AS ken_name, 
                    HJ1.city_code AS city_code, 
                    CT1.city_name AS city_name, 
                    HJ1.weather_id AS weather_id, 
                    WE1.weather_name AS weather_name, 
                    TO_CHAR(timezone('JST', HJ1.begin_date::timestamptz), 'mm') AS begin_month, 
                    TO_CHAR(timezone('JST', HJ1.begin_date::timestamptz), 'dd') AS begin_day, 
                    HJ1.kasen_kaigan_code AS kasen_kaigan_code,
                    KK1.kasen_kaigan_name AS kasen_kaigan_name, 
                    HJ1.koji_code AS koji_code, 
                    HJ1.branch_code AS branch_code, 
                    CAST(HJ1.determined AS NUMERIC(20,10)) AS determined, 
                    HJ1.comment as comment, 
                    TO_CHAR(timezone('JST', HJ1.committed_at::timestamptz), 'yyyy/mm/dd HH24:MI') AS committed_at, 
                    TO_CHAR(timezone('JST', HJ1.deleted_at::timestamptz), 'yyyy/mm/dd HH24:MI') AS deleted_at 
                FROM HOJO HJ1 
                LEFT JOIN HOJO_FILE HF1 ON HJ1.hojo_file_id=HF1.hojo_file_id
                LEFT JOIN SUIKEI SK1 ON HJ1.suikei_code=SK1.suikei_code 
                LEFT JOIN KASEN KA1 ON HJ1.kasen_code=KA1.kasen_code
                LEFT JOIN KEN KE1 ON HJ1.ken_code=KE1.ken_code 
                LEFT JOIN CITY CT1 ON HJ1.city_code=CT1.city_code 
                LEFT JOIN WEATHER WE1 ON HJ1.weather_id=WE1.weather_id 
                LEFT JOIN KASEN_KAIGAN KK1 ON HJ1.kasen_kaigan_code=KK1.kasen_kaigan_code 
                WHERE 
                    HJ1.ken_code=%s AND HJ1.deleted_at IS NULL 
                ORDER BY CAST(HJ1.city_code AS INTEGER), CAST(HJ1.hojo_file_id AS INTEGER), CAST(HJ1.hojo_id AS INTEGER)""", [
                    trigger_list[0].ken_code, 
                ])
    
            if hojo_list:
                for i, hojo in enumerate(hojo_list):
                    if hojo.hojo_id is not None and len(str(hojo.hojo_id)) > 0:
                        ws_hojo.cell(row=8+i, column=1).value = str(hojo.hojo_id)
                        ws_hojo.cell(row=8+i, column=1).fill = green_fill
                    
                    if hojo.suikei_name is not None and hojo.suikei_code is not None and len(str(hojo.suikei_name)) > 0 and len(str(hojo.suikei_code)) > 0:
                        ws_hojo.cell(row=8+i, column=2).value = str(hojo.suikei_name) + ":" + str(hojo.suikei_code)

                    if hojo.suikei_type_name is not None and hojo.suikei_type_code is not None and len(str(hojo.suikei_type_name)) > 0 and len(str(hojo.suikei_type_code)) > 0:                        
                        ws_hojo.cell(row=8+i, column=3).value = str(hojo.suikei_type_name) + ":" + str(hojo.suikei_type_code)
                    
                    if hojo.kasen_name is not None and hojo.kasen_code is not None and len(str(hojo.kasen_name)) > 0 and len(str(hojo.kasen_code)) > 0:
                        ws_hojo.cell(row=8+i, column=4).value = str(hojo.kasen_name) + str(hojo.kasen_code)
                        
                    if hojo.kasen_type_name is not None and hojo.kasen_type_code is not None and len(str(hojo.kasen_type_name)) > 0 and len(str(hojo.kasen_type_code)) > 0:
                        ws_hojo.cell(row=8+i, column=5).value = str(hojo.kasen_type_name) + ":" + str(hojo.kasen_type_code)
                        
                    if hojo.ken_name is not None and hojo.ken_code is not None and len(str(hojo.ken_name)) > 0 and len(str(hojo.ken_code)) > 0:
                        ws_hojo.cell(row=8+i, column=6).value = str(hojo.ken_name) + ":" + str(hojo.ken_code)
                    
                    if hojo.koji_code is not None and len(str(hojo.koji_code)) > 0:
                        ws_hojo.cell(row=8+i, column=7).value = str(hojo.koji_code)

                    if hojo.branch_code is not None and len(str(hojo.branch_code)) > 0:
                        ws_hojo.cell(row=8+i, column=8).value = str(hojo.branch_code)

                    ### TO-DO TO_DO TODO to-do to_do todo
                    if hojo.koji_code is not None and len(str(hojo.koji_code)) > 0:
                        ws_hojo.cell(row=8+i, column=9).value = str(hojo.koji_code)
                    
                    if hojo.city_code is not None and len(str(hojo.city_code)) > 0:
                        ws_hojo.cell(row=8+i, column=10).value = str(hojo.city_code)
                    
                    if hojo.kasen_kaigan_code is not None and len(str(hojo.kasen_kaigan_code)) > 0:
                        ws_hojo.cell(row=8+i, column=11).value = str(hojo.kasen_kaigan_code)

                    ws_hojo.cell(row=8+i, column=12).value = ""
                    
                    if hojo.weather_name is not None and hojo.weather_id is not None and len(str(hojo.weather_name)) > 0 and len(str(hojo.weather_id)) > 0:
                        ws_hojo.cell(row=8+i, column=13).value = str(hojo.weather_name) + ":" + str(hojo.weather_id)
                        
                    if hojo.begin_month is not None and len(str(hojo.begin_month)) > 0:
                        ws_hojo.cell(row=8+i, column=14).value = str(hojo.begin_month)

                    if hojo.begin_day is not None and len(str(hojo.begin_day)) > 0:
                        ws_hojo.cell(row=8+i, column=15).value = str(hojo.begin_day)
                    
                    if hojo.determined is not None and len(str(hojo.determined)) > 0:
                        ws_hojo.cell(row=8+i, column=16).value = str(hojo.determined)
                    
                    if hojo.comment is not None and len(str(hojo.comment)) > 0:
                        ws_hojo.cell(row=8+i, column=17).value = str(hojo.comment)

            ###################################################################
            ### EXCEL入出力処理(0100)
            ### ダウンロード用のEXCELファイルを保存する。
            ###################################################################
            print_log('[DEBUG] P0900Action.action_R01_download_hojo()関数 STEP 11/13.', 'DEBUG')
            wb.save(download_file_path)

            ###################################################################
            ### DBアクセス処理(0110)
            ### トリガーデータを更新する。
            ###################################################################
            print_log('[DEBUG] P0900Action.action_R01_download_hojo.handle()関数 STEP 12/13.', 'DEBUG')
            connection_cursor = connection.cursor()
            try:
                connection_cursor.execute("""BEGIN""", [])
                
                connection_cursor.execute("""
                    UPDATE TRIGGER SET 
                        status_code=%s, -- status_code
                        success_count=%s, -- success_count
                        failure_count=%s, -- failure_count
                        consumed_at=CURRENT_TIMESTAMP, 
                        integrity_ok=%s, -- integrity_ok
                        integrity_ng=%s  -- integrity_ng
                    WHERE
                        trigger_id=%s -- trigger_id
                    """, [
                        'SUCCESS', ### status_count
                        1, ### success_count
                        0, ### failure_count
                        '\n'.join(get_info_log()), ### integrity_ok
                        '\n'.join(get_warn_log()), ### integrity_ng
                        trigger_list[0].trigger_id, ### trigger_id
                    ])
                ### transaction.commit()    
                connection_cursor.execute("""COMMIT""", [])
            except:
                print_log('[ERROR] P0900Action.action_R01_download_hojo.handle()関数 {}'.format(sys.exc_info()[0]), 'ERROR')
                ### connection_cursor.rollback()
                connection_cursor.execute("""ROLLBACK""", [])
            finally:
                connection_cursor.close()
                
            ###################################################################
            ### 戻り値セット処理(0120)
            ### 戻り値を戻す。
            ###################################################################
            print_log('[DEBUG] P0900Action.action_R01_download_hojo.handle()関数 STEP 13/13.', 'DEBUG')
            print_log('[INFO] P0900Action.action_R01_download_hojo.handle()関数が正常終了しました。', 'INFO')
            return 0
        except:
            print_log('[ERROR] P0900Action.action_R01_download_hojo.handle()関数 {}'.format(sys.exc_info()[0]), 'ERROR')
            print_log('[ERROR] P0900Action.action_R01_download_hojo.handle()関数でエラーが発生しました。', 'ERROR')
            print_log('[ERROR] P0900Action.action_R01_download_hojo.handle()関数が異常終了しました。', 'ERROR')
            return 8
            