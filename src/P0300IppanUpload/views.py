#!/usr/bin/env python3
# -*- coding: utf-8 -*-
###############################################################################
### ファイル名：P0300IppanUpload/views.py
### 一般資産調査票アップロード
###############################################################################

###############################################################################
### 処理名：インポート処理
###############################################################################
import os
import sys
from datetime import date, datetime, timedelta, timezone
from django.contrib.auth.decorators import login_required
from django.db import connection
from django.db import transaction
from django.db.models import Max
from django.http import Http404
from django.http import HttpResponse
from django.http import HttpResponseRedirect
from django.shortcuts import render
from django.template import loader
from django.views import generic

import hashlib

import openpyxl
from openpyxl.comments import Comment
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import PatternFill
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.writer.excel import save_virtual_workbook

from .forms import IppanUploadForm

from P0000Common.models import BUILDING                ### 1000: 建物区分
from P0000Common.models import KEN                     ### 1010: 都道府県
from P0000Common.models import CITY                    ### 1020: 市区町村
from P0000Common.models import KASEN_KAIGAN            ### 1030: 水害発生地点工種（河川海岸区分）
from P0000Common.models import SUIKEI                  ### 1040: 水系（水系・沿岸）
from P0000Common.models import SUIKEI_TYPE             ### 1050: 水系種別（水系・沿岸種別）
from P0000Common.models import KASEN                   ### 1060: 河川（河川・海岸）
from P0000Common.models import KASEN_TYPE              ### 1070: 河川種別（河川・海岸種別）
from P0000Common.models import CAUSE                   ### 1080: 水害原因
from P0000Common.models import UNDERGROUND             ### 1090: 地上地下区分
from P0000Common.models import USAGE                   ### 1100: 地下空間の利用形態
from P0000Common.models import FLOOD_SEDIMENT          ### 1110: 浸水土砂区分
from P0000Common.models import GRADIENT                ### 1120: 地盤勾配区分
from P0000Common.models import INDUSTRY                ### 1130: 産業分類
from P0000Common.models import KOEKI_INDUSTRY          ### 1140: 公益事業分類

from P0000Common.models import HOUSE_ASSET             ### 2000: 家屋評価額
from P0000Common.models import HOUSE_ALT               ### 2020: 家庭応急対策費_代替活動費
from P0000Common.models import HOUSE_CLEAN             ### 2030: 家庭応急対策費_清掃日数

from P0000Common.models import HOUSEHOLD_ASSET         ### 3000: 家庭用品自動車以外所有額
from P0000Common.models import HOUSEHOLD_RATE          ### 3010: 家庭用品自動車以外被害率

from P0000Common.models import CAR_ASSET               ### 4000: 家庭用品自動車所有額
from P0000Common.models import CAR_RATE                ### 4010: 家庭用品自動車被害率

from P0000Common.models import OFFICE_ASSET            ### 5000: 事業所資産額
from P0000Common.models import OFFICE_RATE             ### 5010: 事業所被害率
from P0000Common.models import OFFICE_SUSPEND          ### 5020: 事業所営業停止日数
from P0000Common.models import OFFICE_STAGNATE         ### 5030: 事業所営業停滞日数
from P0000Common.models import OFFICE_ALT              ### 5040: 事業所応急対策費_代替活動費

from P0000Common.models import FARMER_FISHER_ASSET     ### 6000: 農漁家資産額
from P0000Common.models import FARMER_FISHER_RATE      ### 6010: 農漁家被害率

from P0000Common.models import AREA                    ### 7000: 入力データ_水害区域
from P0000Common.models import WEATHER                 ### 7010: 入力データ_異常気象
from P0000Common.models import SUIGAI                  ### 7020: 入力データ_ヘッダ部分
from P0000Common.models import IPPAN                   ### 7030: 入力データ_一覧表部分
from P0000Common.models import IPPAN_VIEW              ### 7040: ビューデータ_一覧表部分
from P0000Common.models import CHITAN_FILE             ### 7050: 入力データ_公共土木施設調査票_地方単独事業_ファイル部分
from P0000Common.models import CHITAN                  ### 7060: 入力データ_公共土木施設調査票_地方単独事業_一覧表部分
from P0000Common.models import HOJO_FILE               ### 7070: 入力データ_公共土木施設調査票_補助事業_ファイル部分
from P0000Common.models import HOJO                    ### 7080: 入力データ_公共土木施設調査票_補助事業_一覧表部分
from P0000Common.models import KOEKI_FILE              ### 7090: 入力データ_公益事業等調査票_ファイル部分
from P0000Common.models import KOEKI                   ### 7100: 入力データ_公益事業等調査票_一覧表部分

from P0000Common.models import IPPAN_SUMMARY           ### 8000: 集計データ

from P0000Common.models import ACTION                  ### 10000: アクション
from P0000Common.models import STATUS                  ### 10010: 状態
from P0000Common.models import TRIGGER                 ### 10020: トリガーメッセージ

from P0000Common.common import get_debug_log
from P0000Common.common import get_error_log
from P0000Common.common import get_info_log
from P0000Common.common import get_warn_log
from P0000Common.common import print_log
from P0000Common.common import reset_log

###############################################################################
### 処理名：定数定義
### 単体入力の必須チェックを行った結果のメッセージを定義する。
###############################################################################
MESSAGE = []

### 単体入力の必須をチェックする。
MESSAGE.append([0, 'W0000', '必須', '都道府県が入力されていません。', '都道府県を入力してください。'])
MESSAGE.append([1, 'W0001', '必須', '市区町村が入力されていません。', '市区町村を入力してください。'])
MESSAGE.append([2, 'W0002', '必須', '水害発生月日が入力されていません。', '水害発生月日を入力してください。'])
MESSAGE.append([3, 'W0003', '必須', '水害終了月日が入力されていません。', '水害終了月日を入力してください。'])
MESSAGE.append([4, 'W0004', '必須', '水害原因1が入力されていません。', '水害原因1を入力してください。'])
MESSAGE.append([5, 'W0005', '必須', '水害原因2が入力されていません。', '水害原因2を入力してください。'])
MESSAGE.append([6, 'W0006', '必須', '水害原因3が入力されていません。', '水害原因3を入力してください。'])
MESSAGE.append([7, 'W0007', '必須', '水害区域番号が入力されていません。', '水害区域番号を入力してください。'])
MESSAGE.append([8, 'W0008', '必須', '水系・沿岸名が入力されていません。', '水系・沿岸名を入力してください。'])
MESSAGE.append([9, 'W0009', '必須', '水系種別が入力されていません。', '水系種別を入力してください。'])
MESSAGE.append([10, 'W0010', '必須', '河川・海岸名が入力されていません。', '河川・海岸名を入力してください。'])
MESSAGE.append([11, 'W0011', '必須', '河川種別が入力されていません。', '河川種別を入力してください。'])
MESSAGE.append([12, 'W0012', '必須', '地盤勾配区分が入力されていません。', '地盤勾配区分を入力してください。'])
MESSAGE.append([13, 'W0013', '必須', '水害区域面積の宅地が入力されていません。', '水害区域面積の宅地を入力してください。'])
MESSAGE.append([14, 'W0014', '必須', '水害区域面積の農地が入力されていません。', '水害区域面積の農地を入力してください。'])
MESSAGE.append([15, 'W0015', '必須', '水害区域面積の地下が入力されていません。', '水害区域面積の地下を入力してください。'])
MESSAGE.append([16, 'W0016', '必須', '工種が入力されていません。', '工種を入力してください。'])
MESSAGE.append([17, 'W0017', '必須', '農作物被害額が入力されていません。', '農作物被害額を入力してください。'])
MESSAGE.append([18, 'W0018', '必須', '異常気象コードが入力されていません。', '異常気象コードを入力してください。'])
for i in range(19, 50):
    MESSAGE.append([i, '', '', '', ''])

MESSAGE.append([50, 'W0050', '必須', '町丁名・大字名が入力されていません。', '町丁名・大字名を入力してください。'])
MESSAGE.append([51, 'W0051', '必須', '名称が入力されていません。', '名称を入力してください。'])
MESSAGE.append([52, 'W0052', '必須', '地上・地下被害の区分が入力されていません。', '地上・地下被害の区分を入力してください。'])
MESSAGE.append([53, 'W0053', '必須', '浸水土砂被害の区分が入力されていません。', '浸水土砂被害の区分を入力してください。'])
MESSAGE.append([54, 'W0054', '必須', '被害建物棟数, 床下浸水が入力されていません。', '被害建物棟数, 床下浸水を入力してください。'])
MESSAGE.append([55, 'W0055', '必須', '被害建物棟数, 1cm〜49cmが入力されていません。', '被害建物棟数, 1cm〜49cmを入力してください。'])
MESSAGE.append([56, 'W0056', '必須', '被害建物棟数, 50cm〜99cmが入力されていません。', '被害建物棟数, 50cm〜99cmを入力してください。'])
MESSAGE.append([57, 'W0057', '必須', '被害建物棟数, 1m以上が入力されていません。', '被害建物棟数, 1m以上を入力してください。'])
MESSAGE.append([58, 'W0058', '必須', '被害建物棟数, 半壊が入力されていません。', '被害建物棟数, 半壊を入力してください。'])
MESSAGE.append([59, 'W0059', '必須', '被害建物棟数, 全壊・流失が入力されていません。', '被害建物棟数, 全壊・流失を入力してください。'])
MESSAGE.append([60, 'W0060', '必須', '被害建物の延床面積が入力されていません。', '被害建物の延床面積を入力してください。'])
MESSAGE.append([61, 'W0061', '必須', '被災世帯数が入力されていません。', '被災世帯数を入力してください。'])
MESSAGE.append([62, 'W0062', '必須', '被災事業所数が入力されていません。', '被災事業所数を入力してください。'])
MESSAGE.append([63, 'W0063', '必須', '農家・漁家戸数, 床下浸水が入力されていません。', '農家・漁家戸数, 床下浸水を入力してください。'])
MESSAGE.append([64, 'W0064', '必須', '農家・漁家戸数, 1cm〜49cmが入力されていません。', '農家・漁家戸数, 1cm〜49cmを入力してください。'])
MESSAGE.append([65, 'W0065', '必須', '農家・漁家戸数, 50cm〜99cmが入力されていません。', '農家・漁家戸数, 50cm〜99cmを入力してください。'])
MESSAGE.append([66, 'W0066', '必須', '農家・漁家戸数, 1m以上・半壊が入力されていません。', '農家・漁家戸数, 1m以上・半壊を入力してください。'])
MESSAGE.append([67, 'W0067', '必須', '農家・漁家戸数, 全壊・流失が入力されていません。', '農家・漁家戸数, 全壊・流失を入力してください。'])
MESSAGE.append([68, 'W0068', '必須', '事業所従業者数, 床下浸水が入力されていません。', '事業所従業者数, 床下浸水を入力してください。'])
MESSAGE.append([69, 'W0069', '必須', '事業所従業者数, 1cm〜49cmが入力されていません。', '事業所従業者数, 1cm〜49cmを入力してください。'])
MESSAGE.append([70, 'W0070', '必須', '事業所従業者数, 50cm〜99cmが入力されていません。', '事業所従業者数, 50cm〜99cmを入力してください。'])
MESSAGE.append([71, 'W0071', '必須', '事業所従業者数, 1m以上・半壊が入力されていません。', '事業所従業者数, 1m以上・半壊を入力してください。'])
MESSAGE.append([72, 'W0072', '必須', '事業所従業者数, 全壊・流失が入力されていません。', '事業所従業者数, 全壊・流失を入力してください。'])
MESSAGE.append([73, 'W0073', '必須', '事業所の産業区分が入力されていません。', '事業所の産業区分を入力してください。'])
MESSAGE.append([74, 'W0074', '必須', '地下空間の利用形態が入力されていません。', '地下空間の利用形態を入力してください。'])
MESSAGE.append([75, 'W0075', '必須', '備考が入力されていません。', '備考を入力してください。'])
for i in range(76, 100):
    MESSAGE.append([i, '', '', '', ''])

###############################################################################
### 処理名：定数定義
### 単体入力の形式チェックを行った結果のメッセージを定義する。
###############################################################################
MESSAGE.append([100, 'W0100', '形式', '都道府県に無効な文字が入力されています。', '全角文字:半角数字の形式で入力してください。'])
MESSAGE.append([101, 'W0101', '形式', '市区町村に無効な文字が入力されています。', '全角文字:半角数字の形式で入力してください。'])
MESSAGE.append([102, 'W0102', '形式', '水害発生月日に日付として無効な文字が入力されています。', '日付として有効な文字を入力してください。'])
MESSAGE.append([103, 'W0103', '形式', '水害終了月日に日付として無効な文字が入力されています。', '日付として有効な文字を入力してください。'])
MESSAGE.append([104, 'W0104', '形式', '水害原因1に無効な文字が入力されています。', '全角文字:半角数字の形式で入力してください。'])
MESSAGE.append([105, 'W0105', '形式', '水害原因2に無効な文字が入力されています。', '全角文字:半角数字の形式で入力してください。'])
MESSAGE.append([106, 'W0106', '形式', '水害原因3に無効な文字が入力されています。', '全角文字:半角数字の形式で入力してください。'])
MESSAGE.append([107, 'W0107', '形式', '水害区域番号に無効な文字が入力されています。', '全角文字:半角数字の形式で入力してください。'])
MESSAGE.append([108, 'W0108', '形式', '水系・沿岸名に無効な文字が入力されています。', '全角文字:半角数字の形式で入力してください。'])
MESSAGE.append([109, 'W0109', '形式', '水系種別に無効な文字が入力されています。', '全角文字:半角数字の形式で入力してください。'])
MESSAGE.append([110, 'W0110', '形式', '河川・海岸名に無効な文字が入力されています。', '全角文字:半角数字の形式で入力してください。'])
MESSAGE.append([111, 'W0111', '形式', '河川種別に無効な文字が入力されています。', '全角文字:半角数字の形式で入力してください。'])
MESSAGE.append([112, 'W0112', '形式', '地盤勾配区分に無効な文字が入力されています。', '全角文字:半角数字の形式で入力してください。'])
MESSAGE.append([113, 'W0113', '形式', '水害区域面積の宅地に無効な文字が入力されています。', '半角数字の形式で入力してください。'])
MESSAGE.append([114, 'W0114', '形式', '水害区域面積の農地に無効な文字が入力されています。', '半角数字の形式で入力してください。'])
MESSAGE.append([115, 'W0115', '形式', '水害区域面積の地下に無効な文字が入力されています。', '半角数字の形式で入力してください。'])
MESSAGE.append([116, 'W0116', '形式', '工種に全角以外の無効な文字が入力されています。', '全角文字:半角数字の形式で入力してください。'])
MESSAGE.append([117, 'W0117', '形式', '農作物被害額に無効な文字が入力されています。', '半角数字の形式で入力してください。'])
MESSAGE.append([118, 'W0118', '形式', '異常気象コードに無効な文字が入力されています。', '全角文字:半角数字の形式で入力してください。'])
for i in range(119, 150):
    MESSAGE.append([i, '', '', '', ''])

MESSAGE.append([150, 'W0150', '形式', '町丁名・大字名に無効な文字が入力されています。', '全角文字の形式で入力してください。'])
MESSAGE.append([151, 'W0151', '形式', '名称に無効な文字が入力されています。', '全角文字:半角数字の形式で入力してください。'])
MESSAGE.append([152, 'W0152', '形式', '地上・地下被害の区分に無効な文字が入力されています。', '全角文字:半角数字の形式で入力してください。'])
MESSAGE.append([153, 'W0153', '形式', '浸水土砂被害の区分に無効な文字が入力されています。', '全角文字:半角数字の形式で入力してください。'])
MESSAGE.append([154, 'W0154', '形式', '被害建物棟数, 床下浸水に無効な文字が入力されています。', '半角数字の形式で入力してください。'])
MESSAGE.append([155, 'W0155', '形式', '被害建物棟数, 1cm〜49cmに無効な文字が入力されています。', '半角数字の形式で入力してください。'])
MESSAGE.append([156, 'W0156', '形式', '被害建物棟数, 50cm〜99cmに無効な文字が入力されています。', '半角数字の形式で入力してください。'])
MESSAGE.append([157, 'W0157', '形式', '被害建物棟数, 1m以上に無効な文字が入力されています。', '半角数字の形式で入力してください。'])
MESSAGE.append([158, 'W0158', '形式', '被害建物棟数, 半壊に無効な文字が入力されています。', '半角数字の形式で入力してください。'])
MESSAGE.append([159, 'W0159', '形式', '被害建物棟数, 全壊・流失に無効な文字が入力されています。', '半角数字の形式で入力してください。'])
MESSAGE.append([160, 'W0160', '形式', '被害建物の延床面積に無効な文字が入力されています。', '半角数字の形式で入力してください。'])
MESSAGE.append([161, 'W0161', '形式', '被災世帯数に無効な文字が入力されています。', '半角数字の形式で入力してください。'])
MESSAGE.append([162, 'W0162', '形式', '被災事業所数に無効な文字が入力されています。', '半角数字の形式で入力してください。'])
MESSAGE.append([163, 'W0163', '形式', '農家・漁家戸数, 床下浸水に無効な文字が入力されています。', '半角数字の形式で入力してください。'])
MESSAGE.append([164, 'W0164', '形式', '農家・漁家戸数, 1cm〜49cmに無効な文字が入力されています。', '半角数字の形式で入力してください。'])
MESSAGE.append([165, 'W0165', '形式', '農家・漁家戸数, 50cm〜99cmに無効な文字が入力されています。', '半角数字の形式で入力してください。'])
MESSAGE.append([166, 'W0166', '形式', '農家・漁家戸数, 1m以上・半壊に無効な文字が入力されています。', '半角数字の形式で入力してください。'])
MESSAGE.append([167, 'W0167', '形式', '農家・漁家戸数, 全壊・流失に無効な文字が入力されています。', '半角数字の形式で入力してください。'])
MESSAGE.append([168, 'W0168', '形式', '事業所従業者数, 床下浸水に無効な文字が入力されています。', '半角数字の形式で入力してください。'])
MESSAGE.append([169, 'W0169', '形式', '事業所従業者数, 1cm〜49cmに無効な文字が入力されています。', '半角数字の形式で入力してください。'])
MESSAGE.append([170, 'W0170', '形式', '事業所従業者数, 50cm〜99cmに無効な文字が入力されています。', '半角数字の形式で入力してください。'])
MESSAGE.append([171, 'W0171', '形式', '事業所従業者数, 1m以上・半壊に無効な文字が入力されています。', '半角数字の形式で入力してください。'])
MESSAGE.append([172, 'W0172', '形式', '事業所従業者数, 全壊・流失に無効な文字が入力されています。', '半角数字の形式で入力してください。'])
MESSAGE.append([173, 'W0173', '形式', '事業所の産業区分に無効な文字が入力されています。', '全角文字:半角数字の形式で入力してください。'])
MESSAGE.append([174, 'W0174', '形式', '地下空間の利用形態に無効な文字が入力されています。', '全角文字:半角数字の形式で入力してください。'])
MESSAGE.append([175, 'W0175', '形式', '備考に無効な文字が入力されています。', '全角文字の形式で入力してください。'])
for i in range(176, 200):
    MESSAGE.append([i, '', '', '', ''])

###############################################################################
### 処理名：定数定義
### 単体入力の範囲チェックを行った結果のメッセージを定義する。
###############################################################################
MESSAGE.append([200, 'W0200', '範囲', '都道府県に範囲外の無効な値が入力されています。', ''])
MESSAGE.append([201, 'W0201', '範囲', '市区町村に範囲外の無効な値が入力されています。', ''])
MESSAGE.append([202, 'W0202', '範囲', '水害発生月日に範囲外の無効な値が入力されています。', ''])
MESSAGE.append([203, 'W0203', '範囲', '水害終了月日に範囲外の無効な値が入力されています。', ''])
MESSAGE.append([204, 'W0204', '範囲', '水害原因1に範囲外の無効な値が入力されています。', ''])
MESSAGE.append([205, 'W0205', '範囲', '水害原因2に範囲外の無効な値が入力されています。', ''])
MESSAGE.append([206, 'W0206', '範囲', '水害原因3に範囲外の無効な値が入力されています。', ''])
MESSAGE.append([207, 'W0207', '範囲', '水害区域番号に範囲外の無効な値が入力されています。', ''])
MESSAGE.append([208, 'W0208', '範囲', '水系・沿岸名に範囲外の無効な値が入力されています。', ''])
MESSAGE.append([209, 'W0209', '範囲', '水系種別に範囲外の無効な値が入力されています。', ''])
MESSAGE.append([210, 'W0210', '範囲', '河川・海岸名に範囲外の無効な値が入力されています。', ''])
MESSAGE.append([211, 'W0211', '範囲', '河川種別に範囲外の無効な値が入力されています。', ''])
MESSAGE.append([212, 'W0212', '範囲', '地盤勾配区分に範囲外の無効な値が入力されています。', ''])
MESSAGE.append([213, 'W0213', '範囲', '水害区域面積の宅地に範囲外の無効な値が入力されています。', '正の値を入力してください。'])
MESSAGE.append([214, 'W0214', '範囲', '水害区域面積の農地に範囲外の無効な値が入力されています。', '正の値を入力してください。'])
MESSAGE.append([215, 'W0215', '範囲', '水害区域面積の地下に範囲外の無効な値が入力されています。', '正の値を入力してください。'])
MESSAGE.append([216, 'W0216', '範囲', '工種に範囲外の無効な値が入力されています。', ''])
MESSAGE.append([217, 'W0217', '範囲', '農作物被害額に範囲外の無効な値が入力されています。', '正の値を入力してください。'])
MESSAGE.append([218, 'W0218', '範囲', '異常気象コードに範囲外の無効な値が入力されています。', ''])
for i in range(219, 250):
    MESSAGE.append([i, '', '', '', ''])

MESSAGE.append([250, 'W0250', '範囲', '町丁名・大字名に範囲外の無効な値が入力されています。', ''])
MESSAGE.append([251, 'W0251', '範囲', '名称に範囲外の無効な値が入力されています。', ''])
MESSAGE.append([252, 'W0252', '範囲', '地上・地下被害の区分に範囲外の無効な値が入力されています。', ''])
MESSAGE.append([253, 'W0253', '範囲', '浸水土砂被害の区分に範囲外の無効な値が入力されています。', ''])
MESSAGE.append([254, 'W0254', '範囲', '被害建物棟数, 床下浸水に範囲外の無効な値が入力されています。', '正の値を入力してください。'])
MESSAGE.append([255, 'W0255', '範囲', '被害建物棟数, 1cm〜49cmに範囲外の無効な値が入力されています。', '正の値を入力してください。'])
MESSAGE.append([256, 'W0256', '範囲', '被害建物棟数, 50cm〜99cmに範囲外の無効な値が入力されています。', '正の値を入力してください。'])
MESSAGE.append([257, 'W0257', '範囲', '被害建物棟数, 1m以上に範囲外の無効な値が入力されています。', '正の値を入力してください。'])
MESSAGE.append([258, 'W0258', '範囲', '被害建物棟数, 半壊に範囲外の無効な値が入力されています。', '正の値を入力してください。'])
MESSAGE.append([259, 'W0259', '範囲', '被害建物棟数, 全壊・流失に範囲外の無効な値が入力されています。', '正の値を入力してください。'])
MESSAGE.append([260, 'W0260', '範囲', '被害建物の延床面積に範囲外の無効な値が入力されています。', '正の値を入力してください。'])
MESSAGE.append([261, 'W0261', '範囲', '被災世帯数に範囲外の無効な値が入力されています。', '正の値を入力してください。'])
MESSAGE.append([262, 'W0262', '範囲', '被災事業所数に範囲外の無効な値が入力されています。', '正の値を入力してください。'])
MESSAGE.append([263, 'W0263', '範囲', '農家・漁家戸数, 床下浸水に範囲外の無効な値が入力されています。', '正の値を入力してください。'])
MESSAGE.append([264, 'W0264', '範囲', '農家・漁家戸数, 1cm〜49cmに範囲外の無効な値が入力されています。', '正の値を入力してください。'])
MESSAGE.append([265, 'W0265', '範囲', '農家・漁家戸数, 50cm〜99cmに範囲外の無効な値が入力されています。', '正の値を入力してください。'])
MESSAGE.append([266, 'W0266', '範囲', '農家・漁家戸数, 1m以上・半壊に範囲外の無効な値が入力されています。', '正の値を入力してください。'])
MESSAGE.append([267, 'W0267', '範囲', '農家・漁家戸数, 全壊・流失に範囲外の無効な値が入力されています。', '正の値を入力してください。'])
MESSAGE.append([268, 'W0268', '範囲', '事業所従業者数, 床下浸水に範囲外の無効な値が入力されています。', '正の値を入力してください。'])
MESSAGE.append([269, 'W0269', '範囲', '事業所従業者数, 1cm〜49cmに範囲外の無効な値が入力されています。', '正の値を入力してください。'])
MESSAGE.append([270, 'W0270', '範囲', '事業所従業者数, 50cm〜99cmに範囲外の無効な値が入力されています。', '正の値を入力してください。'])
MESSAGE.append([271, 'W0271', '範囲', '事業所従業者数, 1m以上・半壊に範囲外の無効な値が入力されています。', '正の値を入力してください。'])
MESSAGE.append([272, 'W0272', '範囲', '事業所従業者数, 全壊・流失に範囲外の無効な値が入力されています。', '正の値を入力してください。'])
MESSAGE.append([273, 'W0273', '範囲', '事業所の産業区分に範囲外の無効な値が入力されています。', ''])
MESSAGE.append([274, 'W0274', '範囲', '地下空間の利用形態に範囲外の無効な値が入力されています。', ''])
MESSAGE.append([275, 'W0275', '範囲', '備考に範囲外の無効な値が入力されています。', ''])
for i in range(276, 300):
    MESSAGE.append([i, '', '', '', ''])

###############################################################################
### 処理名：定数定義
### 単体入力の相関チェックを行った結果のメッセージを定義する。
###############################################################################
MESSAGE.append([300, 'W0300', '相関', '水害発生月日が水害終了月日より後に入力されています。', '水害発生月日と水害終了月日を正しく入力してください。'])
MESSAGE.append([301, 'W0301', '相関', '水害原因1と工種の関係が正しく入力されていません。', '水害原因が「10:破堤」「20:有堤部溢水」「30:無堤部溢水」「40:内水」の場合、工種は、「1:河川」を入力してください。'])
MESSAGE.append([302, 'W0302', '相関', '水害原因1と工種の関係が正しく入力されていません。', '水害原因が「50:窪地内水」「80:地すべり」「90:急傾斜地崩壊」の場合、工種は、「3:河川海岸以外」を入力してください。'])
MESSAGE.append([303, 'W0303', '相関', '水害原因1と工種の関係が正しく入力されていません。', '水害原因が「93:波浪」の場合、工種は、「2:海岸」を入力してください。'])
MESSAGE.append([304, 'W0304', '相関', '水害原因1と工種の関係が正しく入力されていません。', '水害原因が「60:洗堀・流出」「91:高潮」「92:津波」の場合、工種は、「1:河川」「2:海岸」のいずれかを入力してください。'])
MESSAGE.append([305, 'W0305', '相関', '水害原因1と工種の関係が正しく入力されていません。', '水害原因が「70:土石流」の場合、工種は、「1:河川」「3:河川海岸以外」のいずれかを入力してください。'])
MESSAGE.append([306, 'W0306', '相関', '水害原因2と工種の関係が正しく入力されていません。', '水害原因が「10:破堤」「20:有堤部溢水」「30:無堤部溢水」「40:内水」の場合、工種は、「1:河川」を入力してください。'])
MESSAGE.append([307, 'W0307', '相関', '水害原因2と工種の関係が正しく入力されていません。', '水害原因が「50:窪地内水」「80:地すべり」「90:急傾斜地崩壊」の場合、工種は、「3:河川海岸以外」を入力してください。'])
MESSAGE.append([308, 'W0308', '相関', '水害原因2と工種の関係が正しく入力されていません。', '水害原因が「93:波浪」の場合、工種は、「2:海岸」を入力してください。'])
MESSAGE.append([309, 'W0309', '相関', '水害原因2と工種の関係が正しく入力されていません。', '水害原因が「60:洗堀・流出」「91:高潮」「92:津波」の場合、工種は、「1:河川」「2:海岸」のいずれかを入力してください。'])
MESSAGE.append([310, 'W0310', '相関', '水害原因2と工種の関係が正しく入力されていません。', '水害原因が「70:土石流」の場合、工種は、「1:河川」「3:河川海岸以外」のいずれかを入力してください。'])
MESSAGE.append([311, 'W0311', '相関', '水害原因3と工種の関係が正しく入力されていません。', '水害原因が「10:破堤」「20:有堤部溢水」「30:無堤部溢水」「40:内水」の場合、工種は、「1:河川」を入力してください。'])
MESSAGE.append([312, 'W0312', '相関', '水害原因3と工種の関係が正しく入力されていません。', '水害原因が「50:窪地内水」「80:地すべり」「90:急傾斜地崩壊」の場合、工種は、「3:河川海岸以外」を入力してください。'])
MESSAGE.append([313, 'W0313', '相関', '水害原因3と工種の関係が正しく入力されていません。', '水害原因が「93:波浪」の場合、工種は、「2:海岸」を入力してください。'])
MESSAGE.append([314, 'W0314', '相関', '水害原因3と工種の関係が正しく入力されていません。', '水害原因が「60:洗堀・流出」「91:高潮」「92:津波」の場合、工種は、「1:河川」「2:海岸」のいずれかを入力してください。'])
MESSAGE.append([315, 'W0315', '相関', '水害原因3と工種の関係が正しく入力されていません。', '水害原因が「70:土石流」の場合、工種は、「1:河川」「3:河川海岸以外」のいずれかを入力してください。'])
for i in range(316, 320):
    MESSAGE.append([i, '', '', '', ''])

MESSAGE.append([320, 'W0320', '相関', '水系種別と河川種別の関係が正しく入力されていません。', '水系種別が「1:一級」のときに、河川種別は、「1:直轄」「2:指定」「4:準用」「5:普通」のいずれかを入力してください。'])
MESSAGE.append([321, 'W0321', '相関', '水系種別と河川種別の関係が正しく入力されていません。', '水系種別が「2:二級」のときに、河川種別は、「3:二級」「4:準用」「5:普通」のいずれかを入力してください。'])
MESSAGE.append([322, 'W0322', '相関', '水系種別と河川種別の関係が正しく入力されていません。', '水系種別が「3:準用」のときに、河川種別は、「4:準用」「5:普通」のいずれかを入力してください。'])
MESSAGE.append([323, 'W0323', '相関', '水系種別と河川種別の関係が正しく入力されていません。', '水系種別が「4:普通」のときに、河川種別は、「5:普通」を入力してください。'])
MESSAGE.append([324, 'W0324', '相関', '水系種別と河川種別の関係が正しく入力されていません。', '水系種別が「5:沿岸」のときに、河川種別は、「6:海岸」を入力してください。'])
MESSAGE.append([325, 'W0325', '相関', '水系種別と河川種別の関係が正しく入力されていません。', '水系種別が「6:河川海岸以外」のときに、河川種別は、「7:河川海岸以外」を入力してください。'])
MESSAGE.append([326, 'W0326', '相関', '水系種別と工種の関係が正しく入力されていません。', '水系種別が「1:一級」「2:二級」「3:準用」「4:普通」のときに、工種は、「1:河川」を入力してください。'])
MESSAGE.append([327, 'W0327', '相関', '水系種別と工種の関係が正しく入力されていません。', '水系種別が「5:沿岸」のときに、工種は、「2:海岸」を入力してください。'])
MESSAGE.append([328, 'W0328', '相関', '水系種別と工種の関係が正しく入力されていません。', '水系種別が「6:河川海岸以外」のときに、工種は、「3:河川海岸以外」を入力してください。'])
for i in range(329, 340):
    MESSAGE.append([i, '', '', '', ''])

MESSAGE.append([340, 'W0340', '相関', '水害区域面積の農地と農作物被害額の関係が正しく入力されていません。', '水害区域面積の農地を入力するときには、農作物被害額を入力してください。'])
MESSAGE.append([341, 'W0341', '相関', '水害区域面積の農地と農作物被害額の関係が正しく入力されていません。', '水害区域面積の農地を入力しないときには、農作物被害額を入力しないでください。'])
MESSAGE.append([342, 'W0342', '相関', '水害区域面積の宅地、農地、地下の関係が正しく入力されていません。', '少なくとも、水害区域面積の宅地、農地、地下のいずれかには、値を入力してください。'])
for i in range(343, 360):
    MESSAGE.append([i, '', '', '', ''])

MESSAGE.append([360, 'W0360', '相関', '地上・地下被害の区分と水害区域面積の宅地または農地の関係が正しく入力されていません。', '地上・地下被害の区分が「1:地上のみ」のときに、少なくとも、水害区域面積の宅地、水害区域面積の農地のいずれかを入力してください。'])
MESSAGE.append([361, 'W0361', '相関', '地上・地下被害の区分と水害区域面積の宅地または農地の関係が正しく入力されていません。', '地上・地下被害の区分が「2:地上部分」のときに、少なくとも、水害区域面積の宅地、水害区域面積の農地のいずれかを入力してください。'])
MESSAGE.append([362, 'W0362', '相関', '地上・地下被害の区分と水害区域面積の宅地または農地の関係が正しく入力されていません。', '地上・地下被害の区分が「3:地下部分」のときに、少なくとも、水害区域面積の地下を入力してください。'])
MESSAGE.append([363, 'W0363', '相関', '地上・地下被害の区分と水害区域面積の宅地または農地の関係が正しく入力されていません。', '地上・地下被害の区分が「4:地下のみ」のときに、少なくとも、水害区域面積の地下を入力してください。'])
MESSAGE.append([364, 'W0364', '相関', '地上・地下被害の区分と地下空間の利用形態の関係が正しく入力されていません。', '地上・地下被害の区分が「1:地上のみ」「2:地上部分」のときには、地下空間の利用形態を入力しないでください。'])
MESSAGE.append([365, 'W0365', '相関', '地上・地下被害の区分と地下空間の利用形態の関係が正しく入力されていません。', '地上・地下被害の区分が「3:地下部分」「4:地下のみ」のときには、地下空間の利用形態を入力してください。'])
MESSAGE.append([366, 'W0366', '相関', '被害建物棟数と延床面積の関係が正しく入力されていません。', '被害建物棟数を入力しないときには、延床面積を入力しないでください。'])
MESSAGE.append([367, 'W0367', '相関', '被害建物棟数と延床面積の関係が正しく入力されていません。', '被害建物棟数を入力するときには、延床面積を入力してください。'])
MESSAGE.append([368, 'W0368', '相関', '被災事業所数と事業所の産業区分の関係が正しく入力されていません。', '被災事業所数を入力しないときには、事業所の産業区分を入力しないでください。'])
MESSAGE.append([369, 'W0369', '相関', '被災事業所数と事業所の産業区分の関係が正しく入力されていません。', '被災事業所数を入力するときには、事業所の産業区分を入力してください。'])
MESSAGE.append([370, 'W0370', '相関', '被災事業所数と事業所従業者数の関係が正しく入力されていません。', '被災事業所数を入力しないときには、事業所従業者数を入力しないでください。'])
MESSAGE.append([371, 'W0371', '相関', '被災事業所数と事業所従業者数の関係が正しく入力されていません。', '被災事業所数を入力するときには、事業所従業者数を入力してください。'])
for i in range(372, 400):
    MESSAGE.append([i, '', '', '', ''])

###############################################################################
### 処理名：定数定義
### 単体入力の突合チェックを行った結果のメッセージを定義する。
###############################################################################
MESSAGE.append([400, 'W0400', '突合', '都道府県がデータベースに登録されている都道府県と一致しません。', '正しい都道府県を入力してください。'])
MESSAGE.append([401, 'W0401', '突合', '市区町村がデータベースに登録されている市区町村と一致しません。', '正しい市区町村を入力してください。'])
MESSAGE.append([402, 'W0402', '突合', '水害発生月日がデータベースに登録されている水害発生月日と一致しません。', ''])
MESSAGE.append([403, 'W0403', '突合', '水害終了月日がデータベースに登録されている水害終了月日と一致しません。', ''])
MESSAGE.append([404, 'W0404', '突合', '水害原因1がデータベースに登録されている水害原因1と一致しません。', '正しい水害原因を入力してください。'])
MESSAGE.append([405, 'W0405', '突合', '水害原因2がデータベースに登録されている水害原因2と一致しません。', '正しい水害原因を入力してください。'])
MESSAGE.append([406, 'W0406', '突合', '水害原因3がデータベースに登録されている水害原因3と一致しません。', '正しい水害原因を入力してください。'])
MESSAGE.append([407, 'W0407', '突合', '水害区域番号がデータベースに登録されている水害区域番号と一致しません。', '正しい水害区域番号を入力してください。'])
MESSAGE.append([408, 'W0408', '突合', '水系・沿岸名がデータベースに登録されている水系・沿岸名と一致しません。', '正しい水系・沿岸名を入力してください。'])
MESSAGE.append([409, 'W0409', '突合', '水系種別がデータベースに登録されている水系種別と一致しません。', '正しい水系種別を入力してください。'])
MESSAGE.append([410, 'W0410', '突合', '河川・海岸名がデータベースに登録されている河川・海岸名と一致しません。', '正しい河川・海岸名を入力してください。'])
MESSAGE.append([411, 'W0411', '突合', '河川種別がデータベースに登録されている河川種別と一致しません。', '正しい河川種別を入力してください。'])
MESSAGE.append([412, 'W0412', '突合', '地盤勾配区分がデータベースに登録されている地盤勾配区分と一致しません。', '正しい地盤勾配区分を入力してください。'])
MESSAGE.append([413, 'W0413', '突合', '水害区域面積の宅地がデータベースに登録されている水害区域面積の宅地と一致しません。', ''])
MESSAGE.append([414, 'W0414', '突合', '水害区域面積の農地がデータベースに登録されている水害区域面積の農地と一致しません。', ''])
MESSAGE.append([415, 'W0415', '突合', '水害区域面積の地下がデータベースに登録されている水害区域面積の地下と一致しません。', ''])
MESSAGE.append([416, 'W0416', '突合', '工種がデータベースに登録されている工種と一致しません。', '正しい工種を入力してください。'])
MESSAGE.append([417, 'W0417', '突合', '農作物被害額がデータベースに登録されている農作物被害額と一致しません。', ''])
MESSAGE.append([418, 'W0418', '突合', '異常気象コードがデータベースに登録されている異常気象コードと一致しません。', '正しい異常気象コードを入力してください。'])
for i in range(419, 450):
    MESSAGE.append([i, '', '', '', ''])

MESSAGE.append([450, 'W0450', '突合', '町丁名・大字名がデータベースに登録されている町丁名・大字名と一致しません。', ''])
MESSAGE.append([451, 'W0451', '突合', '名称がデータベースに登録されている名称と一致しません。', '正しい名称を入力してください。'])
MESSAGE.append([452, 'W0452', '突合', '地上・地下被害の区分がデータベースに登録されている地上・地下被害の区分と一致しません。', '正しい地上・地下被害の区分を入力してください。'])
MESSAGE.append([453, 'W0453', '突合', '浸水土砂被害の区分がデータベースに登録されている浸水土砂被害の区分と一致しません。', '正しい浸水土砂被害の区分を入力してください。'])
MESSAGE.append([454, 'W0454', '突合', '被害建物棟数, 床下浸水がデータベースに登録されている被害建物棟数, 床下浸水と一致しません。', ''])
MESSAGE.append([455, 'W0455', '突合', '被害建物棟数, 1cm〜49cmがデータベースに登録されている被害建物棟数, 1cm〜49cmと一致しません。', ''])
MESSAGE.append([456, 'W0456', '突合', '被害建物棟数, 50cm〜99cmがデータベースに登録されている被害建物棟数, 50cm〜99cmと一致しません。', ''])
MESSAGE.append([457, 'W0457', '突合', '被害建物棟数, 1m以上がデータベースに登録されている被害建物棟数, 1m以上と一致しません。', ''])
MESSAGE.append([458, 'W0458', '突合', '被害建物棟数, 半壊がデータベースに登録されている被害建物棟数, 半壊と一致しません。', ''])
MESSAGE.append([459, 'W0459', '突合', '被害建物棟数, 全壊・流失がデータベースに登録されている被害建物棟数, 全壊・流失と一致しません。', ''])
MESSAGE.append([460, 'W0460', '突合', '被害建物の延床面積がデータベースに登録されている被害建物の延床面積と一致しません。', ''])
MESSAGE.append([461, 'W0461', '突合', '被災世帯数がデータベースに登録されている被災世帯数と一致しません。', ''])
MESSAGE.append([462, 'W0462', '突合', '被災事業所数がデータベースに登録されている被災事業所数と一致しません。', ''])
MESSAGE.append([463, 'W0463', '突合', '農家・漁家戸数, 床下浸水がデータベースに登録されている農家・漁家戸数, 床下浸水と一致しません。', ''])
MESSAGE.append([464, 'W0464', '突合', '農家・漁家戸数, 1cm〜49cmがデータベースに登録されている農家・漁家戸数, 1cm〜49cmと一致しません。', ''])
MESSAGE.append([465, 'W0465', '突合', '農家・漁家戸数, 50cm〜99cmがデータベースに登録されている農家・漁家戸数, 50cm〜99cmと一致しません。', ''])
MESSAGE.append([466, 'W0466', '突合', '農家・漁家戸数, 1m以上・半壊がデータベースに登録されている農家・漁家戸数, 1m以上・半壊と一致しません。', ''])
MESSAGE.append([467, 'W0467', '突合', '農家・漁家戸数, 全壊・流失がデータベースに登録されている農家・漁家戸数, 全壊・流失と一致しません。', ''])
MESSAGE.append([468, 'W0468', '突合', '事業所従業者数, 床下浸水がデータベースに登録されている事業所従業者数, 床下浸水と一致しません。', ''])
MESSAGE.append([469, 'W0469', '突合', '事業所従業者数, 1cm〜49cmがデータベースに登録されている事業所従業者数, 1cm〜49cmと一致しません。', ''])
MESSAGE.append([470, 'W0470', '突合', '事業所従業者数, 50cm〜99cmがデータベースに登録されている事業所従業者数, 50cm〜99cmと一致しません。', ''])
MESSAGE.append([471, 'W0471', '突合', '事業所従業者数, 1m以上・半壊がデータベースに登録されている事業所従業者数, 1m以上・半壊と一致しません。', ''])
MESSAGE.append([472, 'W0472', '突合', '事業所従業者数, 全壊・流失がデータベースに登録されている事業所従業者数, 全壊・流失と一致しません。', ''])
MESSAGE.append([473, 'W0473', '突合', '事業所の産業区分がデータベースに登録されている事業所の産業区分と一致しません。', '正しい事業所の産業区分を入力してください。'])
MESSAGE.append([474, 'W0474', '突合', '地下空間の利用形態がデータベースに登録されている地下空間の利用形態と一致しません。', '正しい地下空間の利用形態を入力してください。'])
MESSAGE.append([475, 'W0475', '突合', '備考がデータベースに登録されている備考と一致しません。', ''])
for i in range(476, 500):
    MESSAGE.append([i, '', '', '', ''])

###############################################################################
### 関数名：split_name_code
### (1) 引数がname:codeの場合、[name,code]を返す。
### (2) 引数がnameの場合、[name,'']を返す。
### (3) 引数がcodeの場合、['', code]を返す。
###############################################################################
def split_name_code(data_text):
    try:
        name_code = ['', '']
        if data_text is not None:
            if len(data_text.split(':')) == 0:
                name_code = ['', '']
            elif len(data_text.split(':')) == 1:
                if data_text.isdecimal():
                    name_code = ['', str(data_text)]
                else:
                    name_code = [str(data_text), '']
            elif len(data_text.split(':')) == 2:
                name_code = data_text.split(':')
            else:
                name_code = ['', '']
    except:
        return ['', '']
    
    return name_code

###############################################################################
### 関数名：isdate
### (1) 引数がYYYY-MM-DD形式の場合、Trueを返す。
### (2) 引数がYYYY/MM/DD形式の場合、Trueを返す。
### (3) 引数が上記以外の場合、Falseを返す。
###############################################################################
def isdate(data_text):
    try:
        try:
            datetime.strptime(data_text, '%Y/%m/%d')
            return True
        except ValueError:
            print_log('Incorrect data format, should be YYYY/MM/DD.', 'INFO')
            
        try:
            datetime.strptime(data_text, '%Y-%m-%d')
            return True
        except ValueError:
            print_log('Incorrect data format, should be YYYY-MM-DD', 'INFO')
            
        try:
            datetime.strptime(data_text, '%m/%d')
            return True
        except ValueError:
            print_log('Incorrect data format, should be MM/DD', 'INFO')
            
        try:
            datetime.strptime(data_text, '%m-%d')
            return True
        except ValueError:
            print_log('Incorrect data format, should be MM-DD', 'INFO')
            
        return False
    except:
        return False

###############################################################################
### 関数名：convert_empty_to_none
### (1) 引数がNoneの場合、Noneを返す。
### (2) 引数が''の場合、Noneを返す。
### (3) 引数が上記以外の場合、引数を返す。
###############################################################################
def convert_empty_to_none(arg):
    if arg is None or arg == None:
        return None
    elif arg == '' or arg == "":
        return None
    else:
        return arg

###############################################################################
### 関数名：add_comment
### 背景をセルにセットする。
### コメントをセルにセットする。
###############################################################################
def add_comment(ws_ippan, ws_result, row, column, fill, com_id):
    ws_ippan.cell(row=row, column=column).fill = fill
    ws_result.cell(row=row, column=column).fill = fill
    
    msg = MESSAGE[com_id][3] + MESSAGE[com_id][4]

    if ws_ippan.cell(row=row, column=column).comment is None:
        ws_ippan.cell(row=row, column=column).comment = Comment(msg, '')
    else:
        ws_ippan.cell(row=row, column=column).comment = Comment(str(ws_ippan.cell(row=row, column=column).comment.text) + msg, '')
    if ws_result.cell(row=row, column=column).comment is None:
        ws_result.cell(row=row, column=column).comment = Comment(msg, '')
    else:
        ws_result.cell(row=row, column=column).comment = Comment(str(ws_result.cell(row=row, column=column).comment.text) + msg, '')
    return True    

###############################################################################
### 関数名：index_view
### urlpattern：path('', views.index_view, name='index_view')
### template：P0300IppanUpload/index.html
### (1)GETの場合、一般資産調査票アップロード画面を表示する。
### (2)POSTの場合、アップロードされた一般資産調査票ファイルをチェックして、正常ケースの場合、DBに登録する。
### (3)POSTの場合、アップロードされた一般資産調査票ファイルをチェックして、警告ケースの場合、DBに登録する。
### ※複数EXCELシート対応版
###############################################################################
@login_required(None, login_url='/P0100Login/')
def index_view(request):
    try:
        #######################################################################
        ### 引数チェック処理(0000)
        ### ブラウザからのリクエストと引数をチェックする。
        #######################################################################
        reset_log()
        print_log('[INFO] P0300IppanUpload.index_view()関数が開始しました。', 'INFO')
        print_log('[DEBUG] P0300IppanUpload.index_view()関数 request = {}'.format(request.method), 'DEBUG')
        print_log('[DEBUG] P0300IppanUpload.index_view()関数 STEP 1/35.', 'DEBUG')
        
        #######################################################################
        ### 局所変数セット処理(0010)
        ### チェック結果を格納するために局所変数をセットする。
        #######################################################################
        print_log('[DEBUG] P0300IppanUpload.index_view()関数 STEP 2/35.', 'DEBUG')
        ### 必須チェックの結果を格納する
        require_OK_list = []
        require_NG_list = []
        require_OK_grid = []
        require_NG_grid = []

        ### 形式チェックの結果を格納する
        format_OK_list = []
        format_NG_list = []
        format_OK_grid = []
        format_NG_grid = []

        ### 範囲チェックの結果を格納する
        range_OK_list = []
        range_NG_list = []
        range_OK_grid = []
        range_NG_grid = []

        ### 相関チェックの結果を格納する
        correlate_OK_list = []
        correlate_NG_list = []
        correlate_OK_grid = []
        correlate_NG_grid = []

        ### 突合せチェックの結果を格納する
        compare_OK_list = []
        compare_NG_list = []
        compare_OK_grid = []
        compare_NG_grid = []
    
        #######################################################################
        ### 条件分岐処理(0020)
        ### (1)GETの場合、一般資産調査票アップロード画面を表示して関数を抜ける。
        ### (2)POSTの場合、アップロードされた一般資産調査票ファイルをチェックする。
        ### ※関数の内部のネスト数を浅くするため。
        #######################################################################
        print_log('[DEBUG] P0300IppanUpload.index_view()関数 STEP 3/35.', 'DEBUG')
        if request.method == 'GET':
            form = IppanUploadForm()
            return render(request, 'P0300IppanUpload/index.html', {'form': form})
        
        elif request.method == 'POST':
            form = IppanUploadForm(request.POST, request.FILES)
            
        #######################################################################
        ### フォーム検証処理(0030)
        ### (1)フォームが正しい場合、処理を継続する。
        ### (2)フォームが正しくない場合、ERROR画面を表示して関数を抜ける。
        ### ※関数の内部のネスト数を浅くするため。
        #######################################################################
        print_log('[DEBUG] P0300IppanUpload.index_view()関数 STEP 4/35.', 'DEBUG')
        if form.is_valid():
            pass
        else:
            return HttpResponseRedirect('fail')
    
        #######################################################################
        ### EXCELファイル入出力処理(0040)
        ### (1)局所変数に値をセットする。
        ### (2)アップロードされた一般資産調査票ファイルを保存する。
        #######################################################################
        print_log('[DEBUG] P0300IppanUpload.index_view()関数 STEP 5/35.', 'DEBUG')
        JST = timezone(timedelta(hours=9), 'JST')
        datetime_now_Ym = datetime.now(JST).strftime('%Y%m')
        datetime_now_YmdHMS = datetime.now(JST).strftime('%Y%m%d%H%M%S')
        
        ### upload_file_object = request.FILES['file']
        ### upload_file_path = 'static/repository/' + datetime_now_Ym + '/ippan_chosa_upload_' + datetime_now_YmdHMS + '.xlsx'
        ### upload_file_name = 'ippan_chosa_upload_' + datetime_now_YmdHMS + '.xlsx'
        
        print('1_0')
        upload_file_object = request.FILES['file']
        print('1_1')
        upload_file_name, upload_file_ext = os.path.splitext(request.FILES['file'].name)
        print('1_2')
        upload_file_name = upload_file_name + '_' + datetime_now_YmdHMS + '.xlsx'
        print('1_3')
        upload_file_path = 'static/repository/' + datetime_now_Ym + '/' + upload_file_name
        
        print('2_0')
        with open(upload_file_path, 'wb+') as destination:
            for chunk in upload_file_object.chunks():
                destination.write(chunk)

        ### output_file_path = 'static/repository/'+ datetime_now_Ym +'/ippan_chosa_output_' + datetime_now_YmdHMS + '.xlsx'
        ### output_file_name = 'ippan_chosa_output_' + datetime_now_YmdHMS + '.xlsx'

        print('3_0')
        output_file_name, output_file_ext = os.path.splitext(request.FILES['file'].name)
        print('3_1')
        output_file_name = output_file_name + '_' + datetime_now_YmdHMS + '_output' + '.xlsx'
        print('3_2')
        output_file_path = 'static/repository/' + datetime_now_Ym + '/' + output_file_name
        print('3_3')
        
        print_log('[DEBUG] P0300IppanUpload.index_view()関数 upload_file_object = {}'.format(upload_file_object), 'DEBUG')
        print_log('[DEBUG] P0300IppanUpload.index_view()関数 upload_file_path = {}'.format(upload_file_path), 'DEBUG')
        print_log('[DEBUG] P0300IppanUpload.index_view()関数 upload_file_name = {}'.format(upload_file_name), 'DEBUG')
        print_log('[DEBUG] P0300IppanUpload.index_view()関数 output_file_path = {}'.format(output_file_path), 'DEBUG')
        print_log('[DEBUG] P0300IppanUpload.index_view()関数 output_file_name = {}'.format(output_file_name), 'DEBUG')
                
        #######################################################################
        ### EXCELファイル入出力処理(0050)
        ### (1)アップロードされた一般資産調査票ファイルのワークブックを読み込む。
        ### (2)IPPANワークシートをコピーして、チェック結果を格納するCHECK_RESULTワークシートを追加する。
        ### (3)追加したワークシートを2シート目に移動する。
        ### (4)ワークシートの最大行数を局所変数のmax_rowにセットする。
        ### (5)背景赤色の塗りつぶしを局所変数のfillにセットする。
        ### wb: ワークブック
        ### ws_ippan: IPPANワークシート
        ### ws_result: CHECK_RESULTワークシート
        ### wx_max_row: ワークシートの最大行数
        ### fill: 背景赤色の塗りつぶし
        #######################################################################
        print_log('[DEBUG] P0300IppanUpload.index_view()関数 STEP 6/35.', 'DEBUG')
        wb = openpyxl.load_workbook(upload_file_path)
        ws_ippan = []
        ws_result = []
        ws_title = []

        for ws_temp in wb.worksheets:
            if 'IPPAN' in ws_temp.title:
                ws_ippan.append(ws_temp)
                ws_result.append(wb.copy_worksheet(ws_temp))
                ws_result[-1].title = 'RESULT' + ws_temp.title
                ws_title.append(ws_temp.title)
                
        for ws_temp in wb.worksheets:
            if 'RESULT' in ws_temp.title:
                wb.move_sheet(ws_temp.title, offset=-wb.index(ws_temp))
                wb.move_sheet(ws_temp.title, offset=1)
        
        for ws_temp in wb.worksheets:
            ws_temp.sheet_view.tabSelected = None
            
        wb.active = 1

        #######################################################################
        ### EXCELファイル入出力処理(0060)
        ### (1)EXCELシート毎に最大行を探索する。
        ### (2)局所変数のmax_rowリストに追加する。
        ### ※max_row_tempの初期値の19は、町丁名・大字名、名称等のキャプション部分のEXCEL行番号である。
        #######################################################################
        print_log('[DEBUG] P0300IppanUpload.index_view()関数 STEP 7/35.', 'DEBUG')
        max_row = []
        
        for ws_temp in ws_ippan:
            max_row_temp = 19
            for i in range(ws_temp.max_row + 1, 19, -1):
                if ws_temp.cell(row=i, column=2).value is None:
                    pass
                else:
                    max_row_temp = i
                    break
                    
            max_row.append(max_row_temp)

        #######################################################################
        ### EXCELファイル入出力処理(0070)
        ### EXCELセルの背景赤色を局所変数のfillに設定する。
        #######################################################################
        print_log('[DEBUG] P0300IppanUpload.index_view()関数 STEP 8/35.', 'DEBUG')
        fill = openpyxl.styles.PatternFill(patternType='solid', fgColor='FF0000', bgColor='FF0000')

        #######################################################################
        ### DBアクセス処理(0080)
        ### (1)DBから突合せチェック用のデータを取得する。
        ### (2)突合せチェック用のリストを生成する。
        #######################################################################
        print_log('[DEBUG] P0300IppanUpload.index_view()関数 STEP 9/35.', 'DEBUG')
        ken_list = KEN.objects.raw("""SELECT * FROM KEN ORDER BY CAST(KEN_CODE AS INTEGER)""", [])
        city_list = CITY.objects.raw("""SELECT * FROM CITY ORDER BY CAST(CITY_CODE AS INTEGER)""", [])
        cause_list = CAUSE.objects.raw("""SELECT * FROM CAUSE ORDER BY CAST(CAUSE_CODE AS INTEGER)""", [])
        area_list = AREA.objects.raw("""SELECT * FROM AREA ORDER BY CAST(AREA_ID AS INTEGER)""", [])
        suikei_list = SUIKEI.objects.raw("""SELECT * FROM SUIKEI ORDER BY CAST(SUIKEI_CODE AS INTEGER)""", [])
        suikei_type_list = SUIKEI_TYPE.objects.raw("""SELECT * FROM SUIKEI_TYPE ORDER BY CAST(SUIKEI_TYPE_CODE AS INTEGER)""", [])
        kasen_list = KASEN.objects.raw("""SELECT * FROM KASEN ORDER BY CAST(KASEN_CODE AS INTEGER)""", [])
        kasen_type_list = KASEN_TYPE.objects.raw("""SELECT * FROM KASEN_TYPE ORDER BY CAST(KASEN_TYPE_CODE AS INTEGER)""", [])
        gradient_list = GRADIENT.objects.raw("""SELECT * FROM GRADIENT ORDER BY CAST(GRADIENT_CODE AS INTEGER)""", [])
        kasen_kaigan_list = KASEN_KAIGAN.objects.raw("""SELECT * FROM KASEN_KAIGAN ORDER BY CAST(KASEN_KAIGAN_CODE AS INTEGER)""", [])
        weather_list = WEATHER.objects.raw("""SELECT * FROM WEATHER ORDER BY CAST(WEATHER_ID AS INTEGER)""", [])
        building_list = BUILDING.objects.raw("""SELECT * FROM BUILDING ORDER BY CAST(BUILDING_CODE AS INTEGER)""", [])
        underground_list = UNDERGROUND.objects.raw("""SELECT * FROM UNDERGROUND ORDER BY CAST(UNDERGROUND_CODE AS INTEGER)""", [])
        flood_sediment_list = FLOOD_SEDIMENT.objects.raw("""SELECT * FROM FLOOD_SEDIMENT ORDER BY CAST(FLOOD_SEDIMENT_CODE AS INTEGER)""", [])
        industry_list = INDUSTRY.objects.raw("""SELECT * FROM INDUSTRY ORDER BY CAST(INDUSTRY_CODE AS INTEGER)""", [])
        usage_list = USAGE.objects.raw("""SELECT * FROM USAGE ORDER BY CAST(USAGE_CODE AS INTEGER)""", [])
        
        ken_code_list = [ken.ken_code for ken in ken_list]
        ken_name_list = [ken.ken_name for ken in ken_list]
        ken_name_code_list = [str(ken.ken_name) + ":" + str(ken.ken_code) for ken in ken_list]
        city_code_list = [city.city_code for city in city_list]
        city_name_list = [city.city_name for city in city_list]
        city_name_code_list = [str(city.city_name) + ":" + str(city.city_code) for city in city_list]
        cause_code_list = [cause.cause_code for cause in cause_list]
        cause_name_list = [cause.cause_name for cause in cause_list]
        cause_name_code_list = [str(cause.cause_name) + ":" + str(cause.cause_code) for cause in cause_list]
        area_id_list = [area.area_id for area in area_list]
        area_name_list = [area.area_name for area in area_list]
        area_name_id_list = [str(area.area_name) + ":" + str(area.area_id) for area in area_list]
        suikei_code_list = [suikei.suikei_code for suikei in suikei_list]
        suikei_name_list = [suikei.suikei_name for suikei in suikei_list]
        suikei_name_code_list = [str(suikei.suikei_name) + ":" + str(suikei.suikei_code) for suikei in suikei_list]
        suikei_type_code_list = [suikei_type.suikei_type_code for suikei_type in suikei_type_list]
        suikei_type_name_list = [suikei_type.suikei_type_name for suikei_type in suikei_type_list]
        suikei_type_name_code_list = [str(suikei_type.suikei_type_name) + ":" + str(suikei_type.suikei_type_code) for suikei_type in suikei_type_list]
        kasen_code_list = [kasen.kasen_code for kasen in kasen_list]
        kasen_name_list = [kasen.kasen_name for kasen in kasen_list]
        kasen_name_code_list = [str(kasen.kasen_name) + ":" + str(kasen.kasen_code) for kasen in kasen_list]
        kasen_type_code_list = [kasen_type.kasen_type_code for kasen_type in kasen_type_list]
        kasen_type_name_list = [kasen_type.kasen_type_name for kasen_type in kasen_type_list]
        kasen_type_name_code_list = [str(kasen_type.kasen_type_name) + ":" + str(kasen_type.kasen_type_code) for kasen_type in kasen_type_list]
        gradient_code_list = [gradient.gradient_code for gradient in gradient_list]
        gradient_name_list = [gradient.gradient_name for gradient in gradient_list]
        gradient_name_code_list = [str(gradient.gradient_name) + ":" + str(gradient.gradient_code) for gradient in gradient_list]
        kasen_kaigan_code_list = [kasen_kaigan.kasen_kaigan_code for kasen_kaigan in kasen_kaigan_list]
        kasen_kaigan_name_list = [kasen_kaigan.kasen_kaigan_name for kasen_kaigan in kasen_kaigan_list]
        kasen_kaigan_name_code_list = [str(kasen_kaigan.kasen_kaigan_name) + ":" + str(kasen_kaigan.kasen_kaigan_code) for kasen_kaigan in kasen_kaigan_list]
        weather_id_list = [weather.weather_id for weather in weather_list]
        weather_name_list = [weather.weather_name for weather in weather_list]
        weather_name_id_list = [str(weather.weather_name) + ":" + str(weather.weather_id) for weather in weather_list]
        building_code_list = [building.building_code for building in building_list]
        building_name_list = [building.building_name for building in building_list]
        building_name_code_list = [str(building.building_name) + ":" + str(building.building_code) for building in building_list]
        underground_code_list = [underground.underground_code for underground in underground_list]
        underground_name_list = [underground.underground_name for underground in underground_list]
        underground_name_code_list = [str(underground.underground_name) + ":" + str(underground.underground_code) for underground in underground_list]
        flood_sediment_code_list = [flood_sediment.flood_sediment_code for flood_sediment in flood_sediment_list]
        flood_sediment_name_list = [flood_sediment.flood_sediment_name for flood_sediment in flood_sediment_list]
        flood_sediment_name_code_list = [str(flood_sediment.flood_sediment_name) + ":" + str(flood_sediment.flood_sediment_code) for flood_sediment in flood_sediment_list]
        industry_code_list = [industry.industry_code for industry in industry_list]
        industry_name_list = [industry.industry_name for industry in industry_list]
        industry_name_code_list = [str(industry.industry_name) + ":" + str(industry.industry_code) for industry in industry_list]
        usage_code_list = [usage.usage_code for usage in usage_list]
        usage_name_list = [usage.usage_name for usage in usage_list]
        usage_name_code_list = [str(usage.usage_name) + ":" + str(usage.usage_code) for usage in usage_list]
    
        #######################################################################
        #######################################################################
        ### EXCELセルデータ必須チェック処理（1000）
        ### (1)セル[7:2]からセル[7:9]について、必須項目に値がセットされていることをチェックする。
        ### (2)チェック結果リストにセルの行、列とメッセージを追加する。
        ### (3)IPPANワークシートとRESULTワークシートのセルに背景赤色の塗りつぶしをセットする。
        #######################################################################
        #######################################################################
        print_log('[DEBUG] P0300IppanUpload.index_view()関数 STEP 10/35.', 'DEBUG')
        for i, _ in enumerate(ws_ippan):
            ### セル[7:2]: 都道府県に値がセットされていることをチェックする。
            if ws_ippan[i].cell(row=7, column=2).value is None:
                add_comment(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=7, column=2, fill=fill, com_id=0)
                require_NG_list.append([ws_ippan[i].title, 7, 2, 0])
            else:
                require_OK_list.append([ws_ippan[i].title, 7, 2, 0])
                
            ### セル[7:3]: 市区町村に値がセットされていることをチェックする。
            if ws_ippan[i].cell(row=7, column=3).value is None:
                add_comment(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=7, column=3, fill=fill, com_id=1)
                require_NG_list.append([ws_ippan[i].title, 7, 3, 1])
            else:
                require_OK_list.append([ws_ippan[i].title, 7, 3, 1])
            
            ### セル[7:4]: 水害発生月日に値がセットされていることをチェックする。
            if ws_ippan[i].cell(row=7, column=4).value is None:
                add_comment(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=7, column=4, fill=fill, com_id=2)
                require_NG_list.append([ws_ippan[i].title, 7, 4, 2])
            else:
                require_OK_list.append([ws_ippan[i].title, 7, 4, 2])
    
            ### セル[7:5]: 水害終了月日に値がセットされていることをチェックする。
    
            ### セル[7:6]: 水害原因1に値がセットされていることをチェックする。
            if ws_ippan[i].cell(row=7, column=6).value is None:
                add_comment(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=7, column=6, fill=fill, com_id=4)
                require_NG_list.append([ws_ippan[i].title, 7, 6, 4])
            else:
                require_OK_list.append([ws_ippan[i].title, 7, 6, 4])
    
            ### セル[7:7]: 水害原因2に値がセットされていることをチェックする。
            ### セル[7:8]: 水害原因3に値がセットされていることをチェックする。
                
            ### セル[7:9]: 水害区域番号に値がセットされていることをチェックする。
            ### if ws_ippan[i].cell(row=7, column=9).value is None:
            ###     add_comment(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=7, column=9, fill=fill, com_id=7)
            ###     require_NG_list.append([ws_ippan[i].title, 7, 9, 7])
            ### else:
            ###     require_OK_list.append([ws_ippan[i].title, 7, 9, 7])
    
        #######################################################################
        ### EXCELセルデータ必須チェック処理（1010）
        ### (1)セル[10:2]からセル[10:6]について、必須項目に値がセットされていることをチェックする。
        ### (2)チェック結果リストにセルの行、列とメッセージを追加する。
        ### (3)IPPANワークシートとRESULTワークシートのセルに背景赤色の塗りつぶしをセットする。
        #######################################################################
        print_log('[DEBUG] P0300IppanUpload.index_view()関数 STEP 11/35.', 'DEBUG')
        for i, _ in enumerate(ws_ippan):
            ### セル[10:2]: 水系・沿岸名に値がセットされていることをチェックする。
            if ws_ippan[i].cell(row=10, column=2).value is None:
                add_comment(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=10, column=2, fill=fill, com_id=8)
                require_NG_list.append([ws_ippan[i].title, 10, 2, 8])
            else:
                require_OK_list.append([ws_ippan[i].title, 10, 2, 8])
                
            ### セル[10:3]: 水系種別に値がセットされていることをチェックする。
            if ws_ippan[i].cell(row=10, column=3).value is None:
                add_comment(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=10, column=3, fill=fill, com_id=9)
                require_NG_list.append([ws_ippan[i].title, 10, 3, 9])
            else:
                require_OK_list.append([ws_ippan[i].title, 10, 3, 9])
                
            ### セル[10:4]: 河川・海岸名に値がセットされていることをチェックする。
            if ws_ippan[i].cell(row=10, column=4).value is None:
                add_comment(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=10, column=4, fill=fill, com_id=10)
                require_NG_list.append([ws_ippan[i].title, 10, 4, 10])
            else:
                require_OK_list.append([ws_ippan[i].title, 10, 4, 10])
                
            ### セル[10:5]: 河川種別に値がセットされていることをチェックする。
            if ws_ippan[i].cell(row=10, column=5).value is None:
                add_comment(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=10, column=5, fill=fill, com_id=11)
                require_NG_list.append([ws_ippan[i].title, 10, 5, 11])
            else:
                require_OK_list.append([ws_ippan[i].title, 10, 5, 11])
                
            ### セル[10:6]: 地盤勾配区分に値がセットされていることをチェックする。
            if ws_ippan[i].cell(row=10, column=6).value is None:
                add_comment(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=10, column=6, fill=fill, com_id=12)
                require_NG_list.append([ws_ippan[i].title, 10, 6, 12])
            else:
                require_OK_list.append([ws_ippan[i].title, 10, 6, 12])
    
        #######################################################################
        ### EXCELセルデータ必須チェック処理（1020）
        ### (1)セル[14:2]からセル[14:10]について、必須項目に値がセットされていることをチェックする。
        ### (2)チェック結果リストにセルの行、列とメッセージを追加する。
        ### (3)IPPANワークシートとRESULTワークシートのセルに背景赤色の塗りつぶしをセットする。
        #######################################################################
        print_log('[DEBUG] P0300IppanUpload.index_view()関数 STEP 12/35.', 'DEBUG')
        ### for i, _ in enumerate(ws_ippan):
            ### セル[14:2]: 水害区域面積の宅地に値がセットされていることをチェックする。
            ### セル[14:3]: 水害区域面積の農地に値がセットされていることをチェックする。
            ### セル[14:4]: 水害区域面積の地下に値がセットされていることをチェックする。
            ### セル[14:5]: 
            ### セル[14:6]: 工種に値がセットされていることをチェックする。
            ### セル[14:7]: 
            ### セル[14:8]: 農作物被害額に値がセットされていることをチェックする。
            ### セル[14:9]: 
            ### セル[14:10]: 異常気象コードに値がセットされていることをチェックする。
    
        #######################################################################
        ### EXCELセルデータ必須チェック処理（1030）
        ### (1)セル[20:2]からセル[20:27]について、必須項目に値がセットされていることをチェックする。
        ### (2)チェック結果リストにセルの行、列とメッセージを追加する。
        ### (3)IPPANワークシートとRESULTワークシートのセルに背景赤色の塗りつぶしをセットする。
        ### ※max_rowの20は入力部分の開始EXCEL行番号である。
        ### ※max_rowの20は町丁名・大字名、名称等のキャプション部分の1つ下のEXCEL行番号である。
        #######################################################################
        print_log('[DEBUG] P0300IppanUpload.index_view()関数 STEP 13/35.', 'DEBUG')
        for i, _ in enumerate(ws_ippan):
            if max_row[i] >= 20:
                for j in range(20, max_row[i] + 1):
                    ### セル[20:2]: 町丁名・大字名に値がセットされていることをチェックする。
                    if ws_ippan[i].cell(row=j, column=2).value is None:
                        add_comment(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=j, column=2, fill=fill, com_id=50)
                        require_NG_grid.append([ws_ippan[i].title, j, 2, 50])
                    else:
                        require_OK_grid.append([ws_ippan[i].title, j, 2, 50])
                        
                    ### セル[20:3]: 名称に値がセットされていることをチェックする。
                    if ws_ippan[i].cell(row=j, column=3).value is None:
                        add_comment(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=j, column=3, fill=fill, com_id=51)
                        require_NG_grid.append([ws_ippan[i].title, j, 3, 51])
                    else:
                        require_OK_grid.append([ws_ippan[i].title, j, 3, 51])
                        
                    ### セル[20:4]: 地上・地下被害の区分に値がセットされていることをチェックする。
                    if ws_ippan[i].cell(row=j, column=4).value is None:
                        add_comment(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=j, column=4, fill=fill, com_id=52)
                        require_NG_grid.append([ws_ippan[i].title, j, 4, 52])
                    else:
                        require_OK_grid.append([ws_ippan[i].title, j, 4, 52])
                        
                    ### セル[20:5]: 浸水土砂被害の区分に値がセットされていることをチェックする。
                    if ws_ippan[i].cell(row=j, column=5).value is None:
                        add_comment(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=j, column=5, fill=fill, com_id=53)
                        require_NG_grid.append([ws_ippan[i].title, j, 5, 53])
                    else:
                        require_OK_grid.append([ws_ippan[i].title, j, 5, 53])
                        
                    ### セル[20:6]: 被害建物棟数, 床下浸水に値がセットされていることをチェックする。
                    ### セル[20:7]: 被害建物棟数, 1cm〜49cmに値がセットされていることをチェックする。
                    ### セル[20:8]: 被害建物棟数, 50cm〜99cmに値がセットされていることをチェックする。
                    ### セル[20:9]: 被害建物棟数, 1m以上に値がセットされていることをチェックする。
                    ### セル[20:10]: 被害建物棟数, 半壊に値がセットされていることをチェックする。
                    ### セル[20:11]: 被害建物棟数, 全壊・流失に値がセットされていることをチェックする。
                    ### セル[20:12]: 被害建物の延床面積に値がセットされていることをチェックする。
                    ### セル[20:13]: 被災世帯数に値がセットされていることをチェックする。
                    ### セル[20:14]: 被災事業所数に値がセットされていることをチェックする。
                    ### セル[20:15]: 農家・漁家戸数, 床下浸水に値がセットされていることをチェックする。
                    ### セル[20:16]: 農家・漁家戸数, 1cm〜49cmに値がセットされていることをチェックする。
                    ### セル[20:17]: 農家・漁家戸数, 50cm〜99cmに値がセットされていることをチェックする。
                    ### セル[20:18]: 農家・漁家戸数, 1m以上・半壊に値がセットされていることをチェックする。
                    ### セル[20:19]: 農家・漁家戸数, 全壊・流失に値がセットされていることをチェックする。
                    ### セル[20:20]: 事業所従業者数, 床下浸水に値がセットされていることをチェックする。
                    ### セル[20:21]: 事業所従業者数, 1cm〜49cmに値がセットされていることをチェックする。
                    ### セル[20:22]: 事業所従業者数, 50cm〜99cmに値がセットされていることをチェックする。
                    ### セル[20:23]: 事業所従業者数, 1m以上・半壊に値がセットされていることをチェックする。
                    ### セル[20:24]: 事業所従業者数, 全壊・流失に値がセットされていることをチェックする。
                    ### セル[20:25]: 事業所の産業区分に値がセットされていることをチェックする。
                    ### セル[20:26]: 地下空間の利用形態に値がセットされていることをチェックする。
                    ### セル[20:27]: 備考に値がセットされていることをチェックする。
        
        #######################################################################
        #######################################################################
        ### EXCELセルデータ形式チェック処理（2000）
        ### (1)セル[7:2]からセル[7:9]について形式が正しいことをチェックする。
        ### (2)チェック結果リストにセルの行、列とメッセージを追加する。
        ### (3)IPPANワークシートとRESULTワークシートのセルに背景赤色の塗りつぶしをセットする。
        ### 形式チェックでは、値がセットされている場合のみチェックを行う。
        ### 必須チェックは別途必須チェックで行うためである。
        #######################################################################
        #######################################################################
        print_log('[DEBUG] P0300IppanUpload.index_view()関数 STEP 14/35.', 'DEBUG')
        for i, _ in enumerate(ws_ippan):
            ### セル[7:2]: 都道府県について形式が正しいことをチェックする。
            if ws_ippan[i].cell(row=7, column=2).value is None:
                pass
            else:
                if split_name_code(ws_ippan[i].cell(row=7, column=2).value)[-1].isdecimal() == False:
                    add_comment(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=7, column=2, fill=fill, com_id=100)
                    format_NG_list.append([ws_ippan[i].title, 7, 2, 100])
                else:
                    format_OK_list.append([ws_ippan[i].title, 7, 2, 100])
        
            ### セル[7:3]: 市区町村について形式が正しいことをチェックする。
            if ws_ippan[i].cell(row=7, column=3).value is None:
                pass
            else:
                if split_name_code(ws_ippan[i].cell(row=7, column=3).value)[-1].isdecimal() == False:
                    add_comment(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=7, column=3, fill=fill, com_id=101)
                    format_NG_list.append([ws_ippan[i].title, 7, 3, 101])
                else:
                    format_OK_list.append([ws_ippan[i].title, 7, 3, 101])
      
            ### セル[7:4]: 水害発生月日について形式が正しいことをチェックする。
            if ws_ippan[i].cell(row=7, column=4).value is None:
                pass
            else:
                if isdate(ws_ippan[i].cell(row=7, column=4).value) == False:
                    add_comment(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=7, column=4, fill=fill, com_id=102)
                    format_NG_list.append([ws_ippan[i].title, 7, 4, 102])
                else:
                    format_OK_list.append([ws_ippan[i].title, 7, 4, 102])
        
            ### セル[7:5]: 水害終了月日について形式が正しいことをチェックする。
            if ws_ippan[i].cell(row=7, column=5).value is None:
                pass
            else:
                if isdate(ws_ippan[i].cell(row=7, column=5).value) == False:
                    add_comment(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=7, column=5, fill=fill, com_id=103)
                    format_NG_list.append([ws_ippan[i].title, 7, 5, 103])
                else:
                    format_OK_list.append([ws_ippan[i].title, 7, 5, 103])
                
            ### セル[7:6]: 水害原因1について形式が正しいことをチェックする。
            if ws_ippan[i].cell(row=7, column=6).value is None:
                pass
            else:
                if split_name_code(ws_ippan[i].cell(row=7, column=6).value)[-1].isdecimal() == False:
                    add_comment(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=7, column=6, fill=fill, com_id=104)
                    format_NG_list.append([ws_ippan[i].title, 7, 6, 104])
                else:
                    format_OK_list.append([ws_ippan[i].title, 7, 6, 104])
                
            ### セル[7:7]: 水害原因2について形式が正しいことをチェックする。
            if ws_ippan[i].cell(row=7, column=7).value is None:
                pass
            else:
                if split_name_code(ws_ippan[i].cell(row=7, column=7).value)[-1].isdecimal() == False:
                    add_comment(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=7, column=7, fill=fill, com_id=105)
                    format_NG_list.append([ws_ippan[i].title, 7, 7, 105])
                else:
                    format_OK_list.append([ws_ippan[i].title, 7, 7, 105])
                
            ### セル[7:8]: 水害原因3について形式が正しいことをチェックする。
            if ws_ippan[i].cell(row=7, column=8).value is None:
                pass
            else:
                if split_name_code(ws_ippan[i].cell(row=7, column=8).value)[-1].isdecimal() == False:
                    add_comment(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=7, column=8, fill=fill, com_id=106)
                    format_NG_list.append([ws_ippan[i].title, 7, 8, 106])
                else:
                    format_OK_list.append([ws_ippan[i].title, 7, 8, 106])
                
            ### セル[7:9]: 水害区域番号について形式が正しいことをチェックする。
            if ws_ippan[i].cell(row=7, column=9).value is None:
                pass
            else:
                if split_name_code(ws_ippan[i].cell(row=7, column=9).value)[-1].isdecimal() == False:
                    add_comment(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=7, column=9, fill=fill, com_id=107)
                    format_NG_list.append([ws_ippan[i].title, 7, 9, 107])
                else:
                    format_OK_list.append([ws_ippan[i].title, 7, 9, 107])
    
        #######################################################################
        ### EXCELセルデータ形式チェック処理（2010）
        ### (1)セル[10:2]からセル[10:6]について形式が正しいことをチェックする。
        ### (2)チェック結果リストにセルの行、列とメッセージを追加する。
        ### (3)IPPANワークシートとRESULTワークシートのセルに背景赤色の塗りつぶしをセットする。
        #######################################################################
        print_log('[DEBUG] P0300IppanUpload.index_view()関数 STEP 15/35.', 'DEBUG')
        for i, _ in enumerate(ws_ippan):
            ### セル[10:2]: 水系・沿岸名について形式が正しいことをチェックする。
            if ws_ippan[i].cell(row=10, column=2).value is None:
                pass
            else:
                if split_name_code(ws_ippan[i].cell(row=10, column=2).value)[-1].isdecimal() == False:
                    add_comment(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=10, column=2, fill=fill, com_id=108)
                    format_NG_list.append([ws_ippan[i].title, 10, 2, 108])
                else:
                    format_OK_list.append([ws_ippan[i].title, 10, 2, 108])
                
            ### セル[10:3]: 水系種別について形式が正しいことをチェックする。
            if ws_ippan[i].cell(row=10, column=3).value is None:
                pass
            else:
                if split_name_code(ws_ippan[i].cell(row=10, column=3).value)[-1].isdecimal() == False:
                    add_comment(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=10, column=3, fill=fill, com_id=109)
                    format_NG_list.append([ws_ippan[i].title, 10, 3, 109])
                else:
                    format_OK_list.append([ws_ippan[i].title, 10, 3, 109])
                
            ### セル[10:4]: 河川・海岸名について形式が正しいことをチェックする。
            if ws_ippan[i].cell(row=10, column=4).value is None:
                pass
            else:
                if split_name_code(ws_ippan[i].cell(row=10, column=4).value)[-1].isdecimal() == False:
                    add_comment(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=10, column=4, fill=fill, com_id=110)
                    format_NG_list.append([ws_ippan[i].title, 10, 4, 110])
                else:
                    format_OK_list.append([ws_ippan[i].title, 10, 4, 110])
                
            ### セル[10:5]: 河川種別について形式が正しいことをチェックする。
            if ws_ippan[i].cell(row=10, column=5).value is None:
                pass
            else:
                if split_name_code(ws_ippan[i].cell(row=10, column=5).value)[-1].isdecimal() == False:
                    add_comment(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=10, column=5, fill=fill, com_id=111)
                    format_NG_list.append([ws_ippan[i].title, 10, 5, 111])
                else:
                    format_OK_list.append([ws_ippan[i].title, 10, 5, 111])
                
            ### セル[10:6]: 地盤勾配区分について形式が正しいことをチェックする。
            if ws_ippan[i].cell(row=10, column=6).value is None:
                pass
            else:
                if split_name_code(ws_ippan[i].cell(row=10, column=6).value)[-1].isdecimal() == False:
                    add_comment(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=10, column=6, fill=fill, com_id=112)
                    format_NG_list.append([ws_ippan[i].title, 10, 6, 112])
                else:
                    format_OK_list.append([ws_ippan[i].title, 10, 6, 112])
    
        #######################################################################
        ### EXCELセルデータ形式チェック処理（2020）
        ### (1)セル[14:2]からセル[14:10]について形式が正しいことをチェックする。
        ### (2)チェック結果リストにセルの行、列とメッセージを追加する。
        ### (3)IPPANワークシートとRESULTワークシートのセルに背景赤色の塗りつぶしをセットする。
        #######################################################################
        print_log('[DEBUG] P0300IppanUpload.index_view()関数 STEP 16/35.', 'DEBUG')
        for i, _ in enumerate(ws_ippan):
            ### セル[14:2]: 水害区域面積の宅地について形式が正しいことをチェックする。
            if ws_ippan[i].cell(row=14, column=2).value is None:
                pass
            else:
                if isinstance(ws_ippan[i].cell(row=14, column=2).value, int) == False and \
                    isinstance(ws_ippan[i].cell(row=14, column=2).value, float) == False:
                    add_comment(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=14, column=2, fill=fill, com_id=113)
                    format_NG_list.append([ws_ippan[i].title, 14, 2, 113])
                else:
                    format_OK_list.append([ws_ippan[i].title, 14, 2, 113])
    
            ### セル[14:3]: 水害区域面積の農地について形式が正しいことをチェックする。
            if ws_ippan[i].cell(row=14, column=3).value is None:
                pass
            else:
                if isinstance(ws_ippan[i].cell(row=14, column=3).value, int) == False and \
                    isinstance(ws_ippan[i].cell(row=14, column=3).value, float) == False:
                    add_comment(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=14, column=3, fill=fill, com_id=114)
                    format_NG_list.append([ws_ippan[i].title, 14, 3, 114])
                else:
                    format_OK_list.append([ws_ippan[i].title, 14, 3, 114])
                
            ### セル[14:4]: 水害区域面積の地下について形式が正しいことをチェックする。
            if ws_ippan[i].cell(row=14, column=4).value is None:
                pass
            else:
                if isinstance(ws_ippan[i].cell(row=14, column=4).value, int) == False and \
                    isinstance(ws_ippan[i].cell(row=14, column=4).value, float) == False:
                    add_comment(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=14, column=4, fill=fill, com_id=115)
                    format_NG_list.append([ws_ippan[i].title, 14, 4, 115])
                else:
                    format_OK_list.append([ws_ippan[i].title, 14, 4, 115])
                
            ### セル[14:6]: 工種について形式が正しいことをチェックする。
            if ws_ippan[i].cell(row=14, column=6).value is None:
                pass
            else:
                if split_name_code(ws_ippan[i].cell(row=14, column=6).value)[-1].isdecimal() == False:
                    add_comment(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=14, column=6, fill=fill, com_id=116)
                    format_NG_list.append([ws_ippan[i].title, 14, 6, 116])
                else:
                    format_OK_list.append([ws_ippan[i].title, 14, 6, 116])
                
            ### セル[14:8]: 農作物被害額について形式が正しいことをチェックする。
            if ws_ippan[i].cell(row=14, column=8).value is None:
                pass
            else:
                if isinstance(ws_ippan[i].cell(row=14, column=8).value, int) == False and \
                    isinstance(ws_ippan[i].cell(row=14, column=8).value, float) == False:
                    add_comment(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=14, column=8, fill=fill, com_id=117)
                    format_NG_list.append([ws_ippan[i].title, 14, 8, MESSAGE[117][0], 117])
                else:
                    format_OK_list.append([ws_ippan[i].title, 14, 8, MESSAGE[117][0], 117])
                
            ### セル[14:10]: 異常気象コードについて形式が正しいことをチェックする。
            if ws_ippan[i].cell(row=14, column=10).value is None:
                pass
            else:
                if split_name_code(ws_ippan[i].cell(row=14, column=10).value)[-1].isdecimal() == False:
                    add_comment(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=14, column=10, fill=fill, com_id=118)
                    format_NG_list.append([ws_ippan[i].title, 14, 10, 118])
                else:
                    format_OK_list.append([ws_ippan[i].title, 14, 10, 118])
                
        #######################################################################
        ### EXCELセルデータ形式チェック処理（2030）
        ### (1)セル[20:2]からセル[20:27]について形式が正しいことをチェックする。
        ### (2)チェック結果リストにセルの行、列とメッセージを追加する。
        ### (3)IPPANワークシートとRESULTワークシートのセルに背景赤色の塗りつぶしをセットする。
        #######################################################################
        print_log('[DEBUG] P0300IppanUpload.index_view()関数 STEP 17/35.', 'DEBUG')
        for i, _ in enumerate(ws_ippan):
            if max_row[i] >= 20:
                for j in range(20, max_row[i] + 1):
                    ### セル[20:2]: 町丁名・大字名について形式が正しいことをチェックする。
                        
                    ### セル[20:3]: 名称について形式が正しいことをチェックする。
                    if ws_ippan[i].cell(row=j, column=3).value is None:
                        pass
                    else:
                        if split_name_code(ws_ippan[i].cell(row=j, column=3).value)[-1].isdecimal() == False:
                            add_comment(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=j, column=3, fill=fill, com_id=151)
                            format_NG_grid.append([ws_ippan[i].title, j, 3, 151])
                        else:
                            format_OK_grid.append([ws_ippan[i].title, j, 3, 151])
                        
                    ### セル[20:4]: 地上・地下被害の区分について形式が正しいことをチェックする。
                    if ws_ippan[i].cell(row=j, column=4).value is None:
                        pass
                    else:
                        if split_name_code(ws_ippan[i].cell(row=j, column=4).value)[-1].isdecimal() == False:
                            add_comment(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=j, column=4, fill=fill, com_id=152)
                            format_NG_grid.append([ws_ippan[i].title, j, 4, 152])
                        else:
                            format_OK_grid.append([ws_ippan[i].title, j, 4, 152])
                        
                    ### セル[20:5]: 浸水土砂被害の区分について形式が正しいことをチェックする。
                    if ws_ippan[i].cell(row=j, column=5).value is None:
                        pass
                    else:
                        if split_name_code(ws_ippan[i].cell(row=j, column=5).value)[-1].isdecimal() == False:
                            add_comment(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=j, column=5, fill=fill, com_id=153)
                            format_NG_grid.append([ws_ippan[i].title, j, 5, 153])
                        else:
                            format_OK_grid.append([ws_ippan[i].title, j, 5, 153])
                        
                    ### セル[20:6]: 被害建物棟数, 床下浸水について形式が正しいことをチェックする。
                    if ws_ippan[i].cell(row=j, column=6).value is None:
                        pass
                    else:
                        if isinstance(ws_ippan[i].cell(row=j, column=6).value, int) == False and \
                            isinstance(ws_ippan[i].cell(row=j, column=6).value, float) == False:
                            add_comment(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=j, column=6, fill=fill, com_id=154)
                            format_NG_grid.append([ws_ippan[i].title, j, 6, 154])
                        else:
                            format_OK_grid.append([ws_ippan[i].title, j, 6, 154])
                        
                    ### セル[20:7]: 被害建物棟数, 1cm〜49cmについて形式が正しいことをチェックする。
                    if ws_ippan[i].cell(row=j, column=7).value is None:
                        pass
                    else:
                        if isinstance(ws_ippan[i].cell(row=j, column=7).value, int) == False and \
                            isinstance(ws_ippan[i].cell(row=j, column=7).value, float) == False:
                            add_comment(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=j, column=7, fill=fill, com_id=155)
                            format_NG_grid.append([ws_ippan[i].title, j, 7, 155])
                        else:
                            format_OK_grid.append([ws_ippan[i].title, j, 7, 155])
                        
                    ### セル[20:8]: 被害建物棟数, 50cm〜99cmについて形式が正しいことをチェックする。
                    if ws_ippan[i].cell(row=j, column=8).value is None:
                        pass
                    else:
                        if isinstance(ws_ippan[i].cell(row=j, column=8).value, int) == False and \
                            isinstance(ws_ippan[i].cell(row=j, column=8).value, float) == False:
                            add_comment(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=j, column=8, fill=fill, com_id=156)
                            format_NG_grid.append([ws_ippan[i].title, j, 8, 156])
                        else:
                            format_OK_grid.append([ws_ippan[i].title, j, 8, 156])
                        
                    ### セル[20:9]: 被害建物棟数, 1m以上について形式が正しいことをチェックする。
                    if ws_ippan[i].cell(row=j, column=9).value is None:
                        pass
                    else:
                        if isinstance(ws_ippan[i].cell(row=j, column=9).value, int) == False and \
                            isinstance(ws_ippan[i].cell(row=j, column=9).value, float) == False:
                            add_comment(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=j, column=9, fill=fill, com_id=157)
                            format_NG_grid.append([ws_ippan[i].title, j, 9, 157])
                        else:
                            format_OK_grid.append([ws_ippan[i].title, j, 9, 157])
                        
                    ### セル[20:10]: 被害建物棟数, 半壊について形式が正しいことをチェックする。
                    if ws_ippan[i].cell(row=j, column=10).value is None:
                        pass
                    else:
                        if isinstance(ws_ippan[i].cell(row=j, column=10).value, int) == False and \
                            isinstance(ws_ippan[i].cell(row=j, column=10).value, float) == False:
                            add_comment(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=j, column=10, fill=fill, com_id=158)
                            format_NG_grid.append([ws_ippan[i].title, j, 10, 158])
                        else:
                            format_OK_grid.append([ws_ippan[i].title, j, 10, 158])
                        
                    ### セル[20:11]: 被害建物棟数, 全壊・流失について形式が正しいことをチェックする。
                    if ws_ippan[i].cell(row=j, column=11).value is None:
                        pass
                    else:
                        if isinstance(ws_ippan[i].cell(row=j, column=11).value, int) == False and \
                            isinstance(ws_ippan[i].cell(row=j, column=11).value, float) == False:
                            add_comment(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=j, column=11, fill=fill, com_id=159)
                            format_NG_grid.append([ws_ippan[i].title, j, 11, 159])
                        else:
                            format_OK_grid.append([ws_ippan[i].title, j, 11, 159])
                        
                    ### セル[20:12]: 被害建物の延床面積について形式が正しいことをチェックする。
                    if ws_ippan[i].cell(row=j, column=12).value is None:
                        pass
                    else:
                        if isinstance(ws_ippan[i].cell(row=j, column=12).value, int) == False and \
                            isinstance(ws_ippan[i].cell(row=j, column=12).value, float) == False:
                            add_comment(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=j, column=12, fill=fill, com_id=160)
                            format_NG_grid.append([ws_ippan[i].title, j, 12, 160])
                        else:
                            format_OK_grid.append([ws_ippan[i].title, j, 12, 160])
                        
                    ### セル[20:13]: 被災世帯数について形式が正しいことをチェックする。
                    if ws_ippan[i].cell(row=j, column=13).value is None:
                        pass
                    else:
                        if isinstance(ws_ippan[i].cell(row=j, column=13).value, int) == False and \
                            isinstance(ws_ippan[i].cell(row=j, column=13).value, float) == False:
                            add_comment(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=j, column=13, fill=fill, com_id=161)
                            format_NG_grid.append([ws_ippan[i].title, j, 13, 161])
                        else:
                            format_OK_grid.append([ws_ippan[i].title, j, 13, 161])
                        
                    ### セル[20:14]: 被災事業所数について形式が正しいことをチェックする。
                    if ws_ippan[i].cell(row=j, column=14).value is None:
                        pass
                    else:
                        if isinstance(ws_ippan[i].cell(row=j, column=14).value, int) == False and \
                            isinstance(ws_ippan[i].cell(row=j, column=14).value, float) == False:
                            add_comment(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=j, column=14, fill=fill, com_id=162)
                            format_NG_grid.append([ws_ippan[i].title, j, 14, 162])
                        else:
                            format_OK_grid.append([ws_ippan[i].title, j, 14, 162])
                        
                    ### セル[20:15]: 農家・漁家戸数, 床下浸水について形式が正しいことをチェックする。
                    if ws_ippan[i].cell(row=j, column=15).value is None:
                        pass
                    else:
                        if isinstance(ws_ippan[i].cell(row=j, column=15).value, int) == False and \
                            isinstance(ws_ippan[i].cell(row=j, column=15).value, float) == False:
                            add_comment(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=j, column=15, fill=fill, com_id=163)
                            format_NG_grid.append([ws_ippan[i].title, j, 15, 163])
                        else:
                            format_OK_grid.append([ws_ippan[i].title, j, 15, 163])
                        
                    ### セル[20:16]: 農家・漁家戸数, 1cm〜49cmについて形式が正しいことをチェックする。
                    if ws_ippan[i].cell(row=j, column=16).value is None:
                        pass
                    else:
                        if isinstance(ws_ippan[i].cell(row=j, column=16).value, int) == False and \
                            isinstance(ws_ippan[i].cell(row=j, column=16).value, float) == False:
                            add_comment(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=j, column=16, fill=fill, com_id=164)
                            format_NG_grid.append([ws_ippan[i].title, j, 16, 164])
                        else:
                            format_OK_grid.append([ws_ippan[i].title, j, 16, 164])
                        
                    ### セル[20:17]: 農家・漁家戸数, 50cm〜99cmについて形式が正しいことをチェックする。
                    if ws_ippan[i].cell(row=j, column=17).value is None:
                        pass
                    else:
                        if isinstance(ws_ippan[i].cell(row=j, column=17).value, int) == False and \
                            isinstance(ws_ippan[i].cell(row=j, column=17).value, float) == False:
                            add_comment(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=j, column=17, fill=fill, com_id=165)
                            format_NG_grid.append([ws_ippan[i].title, j, 17, 165])
                        else:
                            format_OK_grid.append([ws_ippan[i].title, j, 17, 165])
                        
                    ### セル[20:18]: 農家・漁家戸数, 1m以上・半壊について形式が正しいことをチェックする。
                    if ws_ippan[i].cell(row=j, column=18).value is None:
                        pass
                    else:
                        if isinstance(ws_ippan[i].cell(row=j, column=18).value, int) == False and \
                            isinstance(ws_ippan[i].cell(row=j, column=18).value, float) == False:
                            add_comment(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=j, column=18, fill=fill, com_id=166)
                            format_NG_grid.append([ws_ippan[i].title, j, 18, 166])
                        else:
                            format_OK_grid.append([ws_ippan[i].title, j, 18, 166])
                        
                    ### セル[20:19]: 農家・漁家戸数, 全壊・流失について形式が正しいことをチェックする。
                    if ws_ippan[i].cell(row=j, column=19).value is None:
                        pass
                    else:
                        if isinstance(ws_ippan[i].cell(row=j, column=19).value, int) == False and \
                            isinstance(ws_ippan[i].cell(row=j, column=19).value, float) == False:
                            add_comment(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=j, column=19, fill=fill, com_id=167)
                            format_NG_grid.append([ws_ippan[i].title, j, 19, 167])
                        else:
                            format_OK_grid.append([ws_ippan[i].title, j, 19, 167])
                        
                    ### セル[20:20]: 事業所従業者数, 床下浸水について形式が正しいことをチェックする。
                    if ws_ippan[i].cell(row=j, column=20).value is None:
                        pass
                    else:
                        if isinstance(ws_ippan[i].cell(row=j, column=20).value, int) == False and \
                            isinstance(ws_ippan[i].cell(row=j, column=20).value, float) == False:
                            add_comment(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=j, column=20, fill=fill, com_id=168)
                            format_NG_grid.append([ws_ippan[i].title, j, 20, 168])
                        else:
                            format_OK_grid.append([ws_ippan[i].title, j, 20, 168])
                        
                    ### セル[20:21]: 事業所従業者数, 1cm〜49cmについて形式が正しいことをチェックする。
                    if ws_ippan[i].cell(row=j, column=21).value is None:
                        pass
                    else:
                        if isinstance(ws_ippan[i].cell(row=j, column=21).value, int) == False and \
                            isinstance(ws_ippan[i].cell(row=j, column=21).value, float) == False:
                            add_comment(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=j, column=21, fill=fill, com_id=169)
                            format_NG_grid.append([ws_ippan[i].title, j, 21, 169])
                        else:
                            format_OK_grid.append([ws_ippan[i].title, j, 21, 169])
                        
                    ### セル[20:22]: 事業所従業者数, 50cm〜99cmについて形式が正しいことをチェックする。
                    if ws_ippan[i].cell(row=j, column=22).value is None:
                        pass
                    else:
                        if isinstance(ws_ippan[i].cell(row=j, column=22).value, int) == False and \
                            isinstance(ws_ippan[i].cell(row=j, column=22).value, float) == False:
                            add_comment(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=j, column=22, fill=fill, com_id=170)
                            format_NG_grid.append([ws_ippan[i].title, j, 22, 170])
                        else:
                            format_OK_grid.append([ws_ippan[i].title, j, 22, 170])
                        
                    ### セル[20:23]: 事業所従業者数, 1m以上・半壊について形式が正しいことをチェックする。
                    if ws_ippan[i].cell(row=j, column=23).value is None:
                        pass
                    else:
                        if isinstance(ws_ippan[i].cell(row=j, column=23).value, int) == False and \
                            isinstance(ws_ippan[i].cell(row=j, column=23).value, float) == False:
                            add_comment(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=j, column=23, fill=fill, com_id=171)
                            format_NG_grid.append([ws_ippan[i].title, j, 23, 171])
                        else:
                            format_OK_grid.append([ws_ippan[i].title, j, 23, 171])
                        
                    ### セル[20:24]: 事業所従業者数, 全壊・流失について形式が正しいことをチェックする。
                    if ws_ippan[i].cell(row=j, column=24).value is None:
                        pass
                    else:
                        if isinstance(ws_ippan[i].cell(row=j, column=24).value, int) == False and \
                            isinstance(ws_ippan[i].cell(row=j, column=24).value, float) == False:
                            add_comment(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=j, column=24, fill=fill, com_id=172)
                            format_NG_grid.append([ws_ippan[i].title, j, 24, 172])
                        else:
                            format_OK_grid.append([ws_ippan[i].title, j, 24, 172])
                        
                    ### セル[20:25]: 事業所の産業区分について形式が正しいことをチェックする。
                    if ws_ippan[i].cell(row=j, column=25).value is None:
                        pass
                    else:
                        if split_name_code(ws_ippan[i].cell(row=j, column=25).value)[-1].isdecimal() == False:
                            add_comment(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=j, column=25, fill=fill, com_id=173)
                            format_NG_grid.append([ws_ippan[i].title, j, 25, 173])
                        else:
                            format_OK_grid.append([ws_ippan[i].title, j, 25, 173])
                        
                    ### セル[20:26]: 地下空間の利用形態について形式が正しいことをチェックする。
                    if ws_ippan[i].cell(row=j, column=26).value is None:
                        pass
                    else:
                        if split_name_code(ws_ippan[i].cell(row=j, column=26).value)[-1].isdecimal() == False:
                            add_comment(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=j, column=26, fill=fill, com_id=174)
                            format_NG_grid.append([ws_ippan[i].title, j, 26, 174])
                        else:
                            format_OK_grid.append([ws_ippan[i].title, j, 26, 174])
                        
                    ### セル[20:27]: 備考について形式が正しいことをチェックする。
    
        #######################################################################
        #######################################################################
        ### EXCELセルデータ範囲チェック処理（3000）
        ### (1)セル[7:2]からセル[7:9]について範囲が正しいことをチェックする。
        ### (2)チェック結果リストにセルの行、列とメッセージを追加する。
        ### (3)IPPANワークシートとRESULTワークシートのセルに背景赤色の塗りつぶしをセットする。
        ### 範囲チェックでは、値がセットされている場合のみチェックを行う。
        ### 必須チェックは別途必須チェックで行うためである。
        ### 範囲チェックでは、形式が正しい場合のみチェックを行う。
        ### 形式チェックは別途形式チェックで行うためである。
        ### 例えば、面積などの数値について、float()で例外とならないように、int、floatの場合のみチェックする。
        ### 範囲チェックでは、数値の下限と上限のチェックを行う。
        ### 範囲チェックでは、DBとの突合チェックに該当するチェックは行わない。
        ### DBとの突合チェックは別途突合チェックで行うためである。
        ### 範囲チェックの例は値が正であることである。
        #######################################################################
        #######################################################################
        print_log('[DEBUG] P0300IppanUpload.index_view()関数 STEP 18/35.', 'DEBUG')
        ### for i, _ in enumerate(ws_ippan):
            ### セル[7:2]: 都道府県について範囲が正しいことをチェックする。
            ### セル[7:3]: 市区町村について範囲が正しいことをチェックする。
            ### セル[7:4]: 水害発生月日について範囲が正しいことをチェックする。
            ### セル[7:5]: 水害終了月日について範囲が正しいことをチェックする。
            ### セル[7:6]: 水害原因1について範囲が正しいことをチェックする。
            ### セル[7:7]: 水害原因2について範囲が正しいことをチェックする。
            ### セル[7:8]: 水害原因3について範囲が正しいことをチェックする。
            ### セル[7:9]: 水害区域番号について範囲が正しいことをチェックする。

        #######################################################################
        ### EXCELセルデータ範囲チェック処理（3010）
        ### (1)セル[10:2]からセル[10:6]について範囲が正しいことをチェックする。
        ### (2)チェック結果リストにセルの行、列とメッセージを追加する。
        ### (3)IPPANワークシートとRESULTワークシートのセルに背景赤色の塗りつぶしをセットする。
        #######################################################################
        print_log('[DEBUG] P0300IppanUpload.index_view()関数 STEP 19/35.', 'DEBUG')
        ### for i, _ in enumerate(ws_ippan):
            ### セル[10:2]: 水系・沿岸名について範囲が正しいことをチェックする。
            ### セル[10:3]: 水系種別について範囲が正しいことをチェックする。
            ### セル[10:4]: 河川・海岸名について範囲が正しいことをチェックする。
            ### セル[10:5]: 河川種別について範囲が正しいことをチェックする。
            ### セル[10:6]: 地盤勾配区分について範囲が正しいことをチェックする。

        #######################################################################
        ### EXCELセルデータ範囲チェック処理（3020）
        ### (1)セル[14:2]からセル[14:10]について範囲が正しいことをチェックする。
        ### (2)チェック結果リストにセルの行、列とメッセージを追加する。
        ### (3)IPPANワークシートとRESULTワークシートのセルに背景赤色の塗りつぶしをセットする。
        #######################################################################
        print_log('[DEBUG] P0300IppanUpload.index_view()関数 STEP 20/35.', 'DEBUG')
        for i, _ in enumerate(ws_ippan):
            ### セル[14:2]: 水害区域面積の宅地について範囲が正しいことをチェックする。
            if ws_ippan[i].cell(row=14, column=2).value is None:
                pass
            else:
                if isinstance(ws_ippan[i].cell(row=14, column=2).value, int) == True or \
                    isinstance(ws_ippan[i].cell(row=14, column=2).value, float) == True:
                    if float(ws_ippan[i].cell(row=14, column=2).value) < 0:
                        add_comment(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=14, column=2, fill=fill, com_id=213)
                        range_NG_list.append([ws_ippan[i].title, 14, 2, 213])
                    else:
                        range_OK_list.append([ws_ippan[i].title, 14, 2, 213])
                
            ### セル[14:3]: 水害区域面積の農地について範囲が正しいことをチェックする。
            if ws_ippan[i].cell(row=14, column=3).value is None:
                pass
            else:
                if isinstance(ws_ippan[i].cell(row=14, column=3).value, int) == True or \
                    isinstance(ws_ippan[i].cell(row=14, column=3).value, float) == True:
                    if float(ws_ippan[i].cell(row=14, column=3).value) < 0:
                        add_comment(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=14, column=3, fill=fill, com_id=214)
                        range_NG_list.append([ws_ippan[i].title, 14, 3, 214])
                    else:
                        range_OK_list.append([ws_ippan[i].title, 14, 3, 214])
    
            ### セル[14:4]: 水害区域面積の地下について範囲が正しいことをチェックする。
            if ws_ippan[i].cell(row=14, column=4).value is None:
                pass
            else:
                if isinstance(ws_ippan[i].cell(row=14, column=4).value, int) == True or \
                    isinstance(ws_ippan[i].cell(row=14, column=4).value, float) == True:
                    if float(ws_ippan[i].cell(row=14, column=4).value) < 0:
                        add_comment(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=14, column=4, fill=fill, com_id=215)
                        range_NG_list.append([ws_ippan[i].title, 14, 4, 215])
                    else:
                        range_OK_list.append([ws_ippan[i].title, 14, 4, 215])
                
            ### セル[14:6]: 工種について範囲が正しいことをチェックする。
    
            ### セル[14:8]: 農作物被害額について範囲が正しいことをチェックする。
            if ws_ippan[i].cell(row=14, column=8).value is None:
                pass
            else:
                if isinstance(ws_ippan[i].cell(row=14, column=8).value, int) == True or \
                    isinstance(ws_ippan[i].cell(row=14, column=8).value, float) == True:
                    if float(ws_ippan[i].cell(row=14, column=8).value) < 0:
                        add_comment(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=14, column=8, fill=fill, com_id=217)
                        range_NG_list.append([ws_ippan[i].title, 14, 8, 217])
                    else:
                        range_OK_list.append([ws_ippan[i].title, 14, 8, 217])
    
            ### セル[14:10]: 異常気象コードについて範囲が正しいことをチェックする。

        #######################################################################
        ### EXCELセルデータ範囲チェック処理（3030）
        ### (1)セル[20:2]からセル[20:27]について範囲が正しいことをチェックする。
        ### (2)チェック結果リストにセルの行、列とメッセージを追加する。
        ### (3)IPPANワークシートとRESULTワークシートのセルに背景赤色の塗りつぶしをセットする。
        #######################################################################
        print_log('[DEBUG] P0300IppanUpload.index_view()関数 STEP 21/35.', 'DEBUG')
        for i, _ in enumerate(ws_ippan):
            if max_row[i] >= 20:
                for j in range(20, max_row[i] + 1):
                    ### セル[20:2]: 町丁名・大字名について範囲が正しいことをチェックする。
                    ### セル[20:3]: 名称について範囲が正しいことをチェックする。
                    ### セル[20:4]: 地上・地下被害の区分について範囲が正しいことをチェックする。
                    ### セル[20:5]: 浸水土砂被害の区分について範囲が正しいことをチェックする。
    
                    ### セル[20:6]: 被害建物棟数, 床下浸水について範囲が正しいことをチェックする。
                    if ws_ippan[i].cell(row=j, column=6).value is None:
                        pass
                    else:
                        if isinstance(ws_ippan[i].cell(row=j, column=6).value, int) == True or \
                            isinstance(ws_ippan[i].cell(row=j, column=6).value, float) == True:
                            if float(ws_ippan[i].cell(row=j, column=6).value) < 0:
                                add_comment(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=j, column=6, fill=fill, com_id=254)
                                range_NG_grid.append([ws_ippan[i].title, j, 6, 254])
                            else:
                                range_OK_grid.append([ws_ippan[i].title, j, 6, 254])
    
                    ### セル[20:7]: 被害建物棟数, 1cm〜49cmについて範囲が正しいことをチェックする。
                    if ws_ippan[i].cell(row=j, column=7).value is None:
                        pass
                    else:
                        if isinstance(ws_ippan[i].cell(row=j, column=7).value, int) == True or \
                            isinstance(ws_ippan[i].cell(row=j, column=7).value, float) == True:
                            if float(ws_ippan[i].cell(row=j, column=7).value) < 0:
                                add_comment(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=j, column=7, fill=fill, com_id=255)
                                range_NG_grid.append([ws_ippan[i].title, j, 7, 255])
                            else:
                                range_OK_grid.append([ws_ippan[i].title, j, 7, 255])
                    
                    ### セル[20:8]: 被害建物棟数, 50cm〜99cmについて範囲が正しいことをチェックする。
                    if ws_ippan[i].cell(row=j, column=8).value is None:
                        pass
                    else:
                        if isinstance(ws_ippan[i].cell(row=j, column=8).value, int) == True or \
                            isinstance(ws_ippan[i].cell(row=j, column=8).value, float) == True:
                            if float(ws_ippan[i].cell(row=j, column=8).value) < 0:
                                add_comment(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=j, column=8, fill=fill, com_id=256)
                                range_NG_grid.append([ws_ippan[i].title, j, 8, 256])
                            else:
                                range_OK_grid.append([ws_ippan[i].title, j, 8, 256])
    
                    ### セル[20:9]: 被害建物棟数, 1m以上について範囲が正しいことをチェックする。
                    if ws_ippan[i].cell(row=j, column=9).value is None:
                        pass
                    else:
                        if isinstance(ws_ippan[i].cell(row=j, column=9).value, int) == True or \
                            isinstance(ws_ippan[i].cell(row=j, column=9).value, float) == True:
                            if float(ws_ippan[i].cell(row=j, column=9).value) < 0:
                                add_comment(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=j, column=9, fill=fill, com_id=257)
                                range_NG_grid.append([ws_ippan[i].title, j, 9, 257])
                            else:
                                range_OK_grid.append([ws_ippan[i].title, j, 9, 257])
    
                    ### セル[20:10]: 被害建物棟数, 半壊について範囲が正しいことをチェックする。
                    if ws_ippan[i].cell(row=j, column=10).value is None:
                        pass
                    else:
                        if isinstance(ws_ippan[i].cell(row=j, column=10).value, int) == True or \
                            isinstance(ws_ippan[i].cell(row=j, column=10).value, float) == True:
                            if float(ws_ippan[i].cell(row=j, column=10).value) < 0:
                                add_comment(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=j, column=10, fill=fill, com_id=258)
                                range_NG_grid.append([ws_ippan[i].title, j, 10, 258])
                            else:
                                range_OK_grid.append([ws_ippan[i].title, j, 10, 258])
    
                    ### セル[20:11]: 被害建物棟数, 全壊・流失について範囲が正しいことをチェックする。
                    if ws_ippan[i].cell(row=j, column=11).value is None:
                        pass
                    else:
                        if isinstance(ws_ippan[i].cell(row=j, column=11).value, int) == True or \
                            isinstance(ws_ippan[i].cell(row=j, column=11).value, float) == True:
                            if float(ws_ippan[i].cell(row=j, column=11).value) < 0:
                                add_comment(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=j, column=11, fill=fill, com_id=259)
                                range_NG_grid.append([ws_ippan[i].title, j, 11, 259])
                            else:
                                range_OK_grid.append([ws_ippan[i].title, j, 11, 259])
    
                    ### セル[20:12]: 被害建物の延床面積について範囲が正しいことをチェックする。
                    if ws_ippan[i].cell(row=j, column=12).value is None:
                        pass
                    else:
                        if isinstance(ws_ippan[i].cell(row=j, column=12).value, int) == True or \
                            isinstance(ws_ippan[i].cell(row=j, column=12).value, float) == True:
                            if float(ws_ippan[i].cell(row=j, column=12).value) < 0:
                                add_comment(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=j, column=12, fill=fill, com_id=260)
                                range_NG_grid.append([ws_ippan[i].title, j, 12, 260])
                            else:
                                range_OK_grid.append([ws_ippan[i].title, j, 12, 260])
    
                    ### セル[20:13]: 被災世帯数について範囲が正しいことをチェックする。
                    if ws_ippan[i].cell(row=j, column=13).value is None:
                        pass
                    else:
                        if isinstance(ws_ippan[i].cell(row=j, column=13).value, int) == True or \
                            isinstance(ws_ippan[i].cell(row=j, column=13).value, float) == True:
                            if float(ws_ippan[i].cell(row=j, column=13).value) < 0:
                                add_comment(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=j, column=13, fill=fill, com_id=261)
                                range_NG_grid.append([ws_ippan[i].title, j, 13, 261])
                            else:
                                range_OK_grid.append([ws_ippan[i].title, j, 13, 261])
                    
                    ### セル[20:14]: 被災事業所数について範囲が正しいことをチェックする。
                    if ws_ippan[i].cell(row=j, column=14).value is None:
                        pass
                    else:
                        if isinstance(ws_ippan[i].cell(row=j, column=14).value, int) == True or \
                            isinstance(ws_ippan[i].cell(row=j, column=14).value, float) == True:
                            if float(ws_ippan[i].cell(row=j, column=14).value) < 0:
                                add_comment(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=j, column=14, fill=fill, com_id=262)
                                range_NG_grid.append([ws_ippan[i].title, j, 14, 262])
                            else:
                                range_OK_grid.append([ws_ippan[i].title, j, 14, 262])
    
                    ### セル[20:15]: 農家・漁家戸数, 床下浸水について範囲が正しいことをチェックする。
                    if ws_ippan[i].cell(row=j, column=15).value is None:
                        pass
                    else:
                        if isinstance(ws_ippan[i].cell(row=j, column=15).value, int) == True or \
                            isinstance(ws_ippan[i].cell(row=j, column=15).value, float) == True:
                            if float(ws_ippan[i].cell(row=j, column=15).value) < 0:
                                add_comment(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=j, column=15, fill=fill, com_id=263)
                                range_NG_grid.append([ws_ippan[i].title, j, 15, 263])
                            else:
                                range_OK_grid.append([ws_ippan[i].title, j, 15, 263])
    
                    ### セル[20:16]: 農家・漁家戸数, 1cm〜49cmについて範囲が正しいことをチェックする。
                    if ws_ippan[i].cell(row=j, column=16).value is None:
                        pass
                    else:
                        if isinstance(ws_ippan[i].cell(row=j, column=16).value, int) == True or \
                            isinstance(ws_ippan[i].cell(row=j, column=16).value, float) == True:
                            if float(ws_ippan[i].cell(row=j, column=16).value) < 0:
                                add_comment(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=j, column=16, fill=fill, com_id=264)
                                range_NG_grid.append([ws_ippan[i].title, j, 16, 264])
                            else:
                                range_OK_grid.append([ws_ippan[i].title, j, 16, 264])
    
                    ### セル[20:17]: 農家・漁家戸数, 50cm〜99cmについて範囲が正しいことをチェックする。
                    if ws_ippan[i].cell(row=j, column=17).value is None:
                        pass
                    else:
                        if isinstance(ws_ippan[i].cell(row=j, column=17).value, int) == True or \
                            isinstance(ws_ippan[i].cell(row=j, column=17).value, float) == True:
                            if float(ws_ippan[i].cell(row=j, column=17).value) < 0:
                                add_comment(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=j, column=17, fill=fill, com_id=265)
                                range_NG_grid.append([ws_ippan[i].title, j, 17, 265])
                            else:
                                range_OK_grid.append([ws_ippan[i].title, j, 17, 265])
    
                    ### セル[20:18]: 農家・漁家戸数, 1m以上・半壊について範囲が正しいことをチェックする。
                    if ws_ippan[i].cell(row=j, column=18).value is None:
                        pass
                    else:
                        if isinstance(ws_ippan[i].cell(row=j, column=18).value, int) == True or \
                            isinstance(ws_ippan[i].cell(row=j, column=18).value, float) == True:
                            if float(ws_ippan[i].cell(row=j, column=18).value) < 0:
                                add_comment(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=j, column=18, fill=fill, com_id=266)
                                range_NG_grid.append([ws_ippan[i].title, j, 18, 266])
                            else:
                                range_OK_grid.append([ws_ippan[i].title, j, 18, 266])
    
                    ### セル[20:19]: 農家・漁家戸数, 全壊・流失について範囲が正しいことをチェックする。
                    if ws_ippan[i].cell(row=j, column=19).value is None:
                        pass
                    else:
                        if isinstance(ws_ippan[i].cell(row=j, column=19).value, int) == True or \
                            isinstance(ws_ippan[i].cell(row=j, column=19).value, float) == True:
                            if float(ws_ippan[i].cell(row=j, column=19).value) < 0:
                                add_comment(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=j, column=19, fill=fill, com_id=267)
                                range_NG_grid.append([ws_ippan[i].title, j, 19, 267])
                            else:
                                range_OK_grid.append([ws_ippan[i].title, j, 19, 267])
    
                    ### セル[20:20]: 事業所従業者数, 床下浸水について範囲が正しいことをチェックする。
                    if ws_ippan[i].cell(row=j, column=20).value is None:
                        pass
                    else:
                        if isinstance(ws_ippan[i].cell(row=j, column=20).value, int) == True or \
                            isinstance(ws_ippan[i].cell(row=j, column=20).value, float) == True:
                            if float(ws_ippan[i].cell(row=j, column=20).value) < 0:
                                add_comment(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=j, column=20, fill=fill, com_id=268)
                                range_NG_grid.append([ws_ippan[i].title, j, 20, 268])
                            else:
                                range_OK_grid.append([ws_ippan[i].title, j, 20, 268])
    
                    ### セル[20:21]: 事業所従業者数, 1cm〜49cmについて範囲が正しいことをチェックする。
                    if ws_ippan[i].cell(row=j, column=21).value is None:
                        pass
                    else:
                        if isinstance(ws_ippan[i].cell(row=j, column=21).value, int) == True or \
                            isinstance(ws_ippan[i].cell(row=j, column=21).value, float) == True:
                            if float(ws_ippan[i].cell(row=j, column=21).value) < 0:
                                add_comment(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=j, column=21, fill=fill, com_id=269)
                                range_NG_grid.append([ws_ippan[i].title, j, 21, 269])
                            else:
                                range_OK_grid.append([ws_ippan[i].title, j, 21, 269])
    
                    ### セル[20:22]: 事業所従業者数, 50cm〜99cmについて範囲が正しいことをチェックする。
                    if ws_ippan[i].cell(row=j, column=22).value is None:
                        pass
                    else:
                        if isinstance(ws_ippan[i].cell(row=j, column=22).value, int) == True or \
                            isinstance(ws_ippan[i].cell(row=j, column=22).value, float) == True:
                            if float(ws_ippan[i].cell(row=j, column=22).value) < 0:
                                add_comment(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=j, column=22, fill=fill, com_id=270)
                                range_NG_grid.append([ws_ippan[i].title, j, 22, 270])
                            else:
                                range_OK_grid.append([ws_ippan[i].title, j, 22, 270])
    
                    ### セル[20:23]: 事業所従業者数, 1m以上・半壊について範囲が正しいことをチェックする。
                    if ws_ippan[i].cell(row=j, column=23).value is None:
                        pass
                    else:
                        if isinstance(ws_ippan[i].cell(row=j, column=23).value, int) == True or \
                            isinstance(ws_ippan[i].cell(row=j, column=23).value, float) == True:
                            if float(ws_ippan[i].cell(row=j, column=23).value) < 0:
                                add_comment(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=j, column=23, fill=fill, com_id=271)
                                range_NG_grid.append([ws_ippan[i].title, j, 23, 271])
                            else:
                                range_OK_grid.append([ws_ippan[i].title, j, 23, 271])
    
                    ### セル[20:24]: 事業所従業者数, 全壊・流失について範囲が正しいことをチェックする。
                    if ws_ippan[i].cell(row=j, column=24).value is None:
                        pass
                    else:
                        if isinstance(ws_ippan[i].cell(row=j, column=24).value, int) == True or \
                            isinstance(ws_ippan[i].cell(row=j, column=24).value, float) == True:
                            if float(ws_ippan[i].cell(row=j, column=24).value) < 0:
                                add_comment(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=j, column=24, fill=fill, com_id=272)
                                range_NG_grid.append([ws_ippan[i].title, j, 24, 272])
                            else:
                                range_OK_grid.append([ws_ippan[i].title, j, 24, 272])
    
                    ### セル[20:25]: 事業所の産業区分について範囲が正しいことをチェックする。
                    ### セル[20:26]: 地下空間の利用形態について範囲が正しいことをチェックする。
                    ### セル[20:27]: 備考について範囲が正しいことをチェックする。
    
        #######################################################################
        #######################################################################
        ### EXCELセルデータ相関チェック処理（4000）
        ### (1)セル[7:2]からセル[7:9]について他項目との相関関係が正しいことをチェックする。
        ### (2)チェック結果リストにセルの行、列とメッセージを追加する。
        ### (3)IPPANワークシートとRESULTワークシートのセルに背景赤色の塗りつぶしをセットする。
        #######################################################################
        #######################################################################
        print_log('[DEBUG] P0300IppanUpload.index_view()関数 STEP 22/35.', 'DEBUG')
        for i, _ in enumerate(ws_ippan):
            ### セル[7:2]: 都道府県が何かの値のときに、相関する市区町村名は正しく選択されているか。
            ### セル[7:3]: 市区町村が何かの値のときに、相関する他項目は正しく選択されているか。
    
            ### セル[7:4]: 水害発生月日 vs セル[7:5]: 水害終了月日
            if ws_ippan[i].cell(row=7, column=4).value is None or \
                ws_ippan[i].cell(row=7, column=5).value is None:
                pass
            else:
                if isdate(ws_ippan[i].cell(row=7, column=4).value) == False or \
                    isdate(ws_ippan[i].cell(row=7, column=5).value) == False:
                    pass
                else:
                    ### 水害発生月日 > 水害終了月日である。
                    if datetime.strptime(ws_ippan[i].cell(row=7, column=4).value, '%Y/%m/%d') > datetime.strptime(ws_ippan[i].cell(row=7, column=5).value, '%Y/%m/%d'): 
                        add_comment(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=7, column=4, fill=fill, com_id=300)
                        add_comment(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=7, column=5, fill=fill, com_id=300)
                        correlate_NG_list.append([ws_ippan[i].title, 7, 4, 300])
                        correlate_NG_list.append([ws_ippan[i].title, 7, 5, 300])
                    else:
                        correlate_OK_list.append([ws_ippan[i].title, 7, 4, 300])
                        correlate_OK_list.append([ws_ippan[i].title, 7, 5, 300])
    
            ### セル[7:5]: 水害終了月日が何かの値のときに、相関する他項目は正しく選択されているか。 参照 セル[7:4]: 水害発生月日 vs セル[7:5]: 水害終了月日
    
            ### セル[7:6]: 水害原因1 vs セル[14:6]: 工種
            if ws_ippan[i].cell(row=7, column=6).value is None or \
                ws_ippan[i].cell(row=14, column=6).value is None:
                pass
            else:
                if split_name_code(ws_ippan[i].cell(row=7, column=6).value)[-1].isdecimal() == False or \
                    split_name_code(ws_ippan[i].cell(row=14, column=6).value)[-1].isdecimal() == False:
                    pass
                else:
                    ### 水害原因が「10:破堤」「20:有堤部溢水」「30:無堤部溢水」「40:内水」の場合、相関する工種は、「1:河川」である。
                    if split_name_code(ws_ippan[i].cell(row=7, column=6).value)[-1] in [10, 20, 30, 40]:
                        if split_name_code(ws_ippan[i].cell(row=14, column=6).value)[-1] == 1:
                            correlate_OK_list.append([ws_ippan[i].title, 7, 6, 301])
                            correlate_OK_list.append([ws_ippan[i].title, 14, 6, 301])
                        else:
                            add_comment(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=7, column=6, fill=fill, com_id=301)
                            add_comment(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=14, column=6, fill=fill, com_id=301)
                            correlate_NG_list.append([ws_ippan[i].title, 7, 6, 301])
                            correlate_NG_list.append([ws_ippan[i].title, 14, 6, 301])
                            
                    ### 水害原因が「50:窪地内水」「80:地すべり」「90:急傾斜地崩壊」の場合、相関する工種は、「3:河川海岸以外」である。
                    elif split_name_code(ws_ippan[i].cell(row=7, column=6).value)[-1] in [50, 80, 90]:
                        if split_name_code(ws_ippan[i].cell(row=14, column=6).value)[-1] == 3:
                            correlate_OK_list.append([ws_ippan[i].title, 7, 6, 302])
                            correlate_OK_list.append([ws_ippan[i].title, 14, 6, 302])
                        else:
                            add_comment(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=7, column=6, fill=fill, com_id=302)
                            add_comment(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=14, column=6, fill=fill, com_id=302)
                            correlate_NG_list.append([ws_ippan[i].title, 7, 6, 302])
                            correlate_NG_list.append([ws_ippan[i].title, 14, 6, 302])
                            
                    ### 水害原因が「93:波浪」の場合、相関する工種は、「2:海岸」である。
                    elif split_name_code(ws_ippan[i].cell(row=7, column=6).value)[-1] in [93]:
                        if split_name_code(ws_ippan[i].cell(row=14, column=6).value)[-1] == 2:
                            correlate_OK_list.append([ws_ippan[i].title, 7, 6, 303])
                            correlate_OK_list.append([ws_ippan[i].title, 14, 6, 303])
                        else:
                            add_comment(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=7, column=6, fill=fill, com_id=303)
                            add_comment(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=14, column=6, fill=fill, com_id=303)
                            correlate_NG_list.append([ws_ippan[i].title, 7, 6, 303])
                            correlate_NG_list.append([ws_ippan[i].title, 14, 6, 303])
                            
                    ### 水害原因が「60:洗堀・流出」「91:高潮」「92:津波」の場合、相関する工種は、「1:河川」「2:海岸」である。
                    elif split_name_code(ws_ippan[i].cell(row=7, column=6).value)[-1] in [60, 91, 92]:
                        if split_name_code(ws_ippan[i].cell(row=14, column=6).value)[-1] == 1 or \
                            split_name_code(ws_ippan[i].cell(row=14, column=6).value)[-1] == 2:
                            correlate_OK_list.append([ws_ippan[i].title, 7, 6, 304])
                            correlate_OK_list.append([ws_ippan[i].title, 14, 6, 304])
                        else:
                            add_comment(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=7, column=6, fill=fill, com_id=304)
                            add_comment(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=14, column=6, fill=fill, com_id=304)
                            correlate_NG_list.append([ws_ippan[i].title, 7, 6, 304])
                            correlate_NG_list.append([ws_ippan[i].title, 14, 6, 304])
                            
                    ### 水害原因が「70:土石流」の場合、相関する工種は、「1:河川」「3:河川海岸以外」である。
                    elif split_name_code(ws_ippan[i].cell(row=7, column=6).value)[-1] in [70]:
                        if split_name_code(ws_ippan[i].cell(row=14, column=6).value)[-1] == 1 or \
                            split_name_code(ws_ippan[i].cell(row=14, column=6).value)[-1] == 3:
                            correlate_OK_list.append([ws_ippan[i].title, 7, 6, 305])
                            correlate_OK_list.append([ws_ippan[i].title, 14, 6, 305])
                        else:
                            add_comment(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=7, column=6, fill=fill, com_id=305)
                            add_comment(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=14, column=6, fill=fill, com_id=305)
                            correlate_NG_list.append([ws_ippan[i].title, 7, 6, 305])
                            correlate_NG_list.append([ws_ippan[i].title, 14, 6, 305])
                            
                    else:
                        pass
    
            ### セル[7:7]: 水害原因2 vs セル[14:6]: 工種
            if ws_ippan[i].cell(row=7, column=7).value is None or \
                ws_ippan[i].cell(row=14, column=6).value is None:
                pass
            else:
                if split_name_code(ws_ippan[i].cell(row=7, column=7).value)[-1].isdecimal() == False or \
                    split_name_code(ws_ippan[i].cell(row=14, column=6).value)[-1].isdecimal() == False:
                    pass
                else:
                    ### 水害原因が「10:破堤」「20:有堤部溢水」「30:無堤部溢水」「40:内水」の場合、相関する工種は、「1:河川」である。
                    if split_name_code(ws_ippan[i].cell(row=7, column=7).value)[-1] in [10, 20, 30, 40]:
                        if split_name_code(ws_ippan[i].cell(row=14, column=6).value)[-1] == 1:
                            correlate_OK_list.append([ws_ippan[i].title, 7, 7, 306])
                            correlate_OK_list.append([ws_ippan[i].title, 14, 6, 306])
                        else:
                            add_comment(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=7, column=7, fill=fill, com_id=306)
                            add_comment(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=14, column=6, fill=fill, com_id=306)
                            correlate_NG_list.append([ws_ippan[i].title, 7, 7, 306])
                            correlate_NG_list.append([ws_ippan[i].title, 14, 6, 306])
                            
                    ### 水害原因が「50:窪地内水」「80:地すべり」「90:急傾斜地崩壊」の場合、相関する工種は、「3:河川海岸以外」である。
                    elif split_name_code(ws_ippan[i].cell(row=7, column=7).value)[-1] in [50, 80, 90]:
                        if split_name_code(ws_ippan[i].cell(row=14, column=6).value)[-1] == 3:
                            correlate_OK_list.append([ws_ippan[i].title, 7, 7, 307])
                            correlate_OK_list.append([ws_ippan[i].title, 14, 6, 307])
                        else:
                            add_comment(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=7, column=7, fill=fill, com_id=307)
                            add_comment(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=14, column=6, fill=fill, com_id=307)
                            correlate_NG_list.append([ws_ippan[i].title, 7, 7, 307])
                            correlate_NG_list.append([ws_ippan[i].title, 14, 6, 307])
                            
                    ### 水害原因が「93:波浪」の場合、相関する工種は、「2:海岸」である。
                    elif split_name_code(ws_ippan[i].cell(row=7, column=7).value)[-1] in [93]:
                        if split_name_code(ws_ippan[i].cell(row=14, column=6).value)[-1] == 2:
                            correlate_OK_list.append([ws_ippan[i].title, 7, 7, 308])
                            correlate_OK_list.append([ws_ippan[i].title, 14, 6, 308])
                        else:
                            add_comment(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=7, column=7, fill=fill, com_id=308)
                            add_comment(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=14, column=6, fill=fill, com_id=308)
                            correlate_NG_list.append([ws_ippan[i].title, 7, 7, 308])
                            correlate_NG_list.append([ws_ippan[i].title, 14, 6, 308])
                            
                    ### 水害原因が「60:洗堀・流出」「91:高潮」「92:津波」の場合、相関する工種は、「1:河川」「2:海岸」である。
                    elif split_name_code(ws_ippan[i].cell(row=7, column=7).value)[-1] in [60, 91, 92]:
                        if split_name_code(ws_ippan[i].cell(row=14, column=6).value)[-1] == 1 or \
                            split_name_code(ws_ippan[i].cell(row=14, column=6).value)[-1] == 2:
                            correlate_OK_list.append([ws_ippan[i].title, 7, 7, 309])
                            correlate_OK_list.append([ws_ippan[i].title, 14, 6, 309])
                        else:
                            add_comment(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=7, column=7, fill=fill, com_id=309)
                            add_comment(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=14, column=6, fill=fill, com_id=309)
                            correlate_NG_list.append([ws_ippan[i].title, 7, 7, 309])
                            correlate_NG_list.append([ws_ippan[i].title, 14, 6, 309])
                            
                    ### 水害原因が「70:土石流」の場合、相関する工種は、「1:河川」「3:河川海岸以外」である。
                    elif split_name_code(ws_ippan[i].cell(row=7, column=7).value)[-1] in [70]:
                        if split_name_code(ws_ippan[i].cell(row=14, column=6).value)[-1] == 1 or \
                            split_name_code(ws_ippan[i].cell(row=14, column=6).value)[-1] == 3:
                            correlate_OK_list.append([ws_ippan[i].title, 7, 7, 310])
                            correlate_OK_list.append([ws_ippan[i].title, 14, 6, 310])
                        else:
                            add_comment(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=7, column=7, fill=fill, com_id=310)
                            add_comment(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=14, column=6, fill=fill, com_id=310)
                            correlate_NG_list.append([ws_ippan[i].title, 7, 7, 310])
                            correlate_NG_list.append([ws_ippan[i].title, 14, 6, 310])
                            
                    else:
                        pass
    
            ### セル[7:8]: 水害原因3 vs セル[14:6]: 工種
            if ws_ippan[i].cell(row=7, column=8).value is None or \
                ws_ippan[i].cell(row=14, column=6).value is None:
                pass
            else:
                if split_name_code(ws_ippan[i].cell(row=7, column=8).value)[-1].isdecimal() == False or \
                    split_name_code(ws_ippan[i].cell(row=14, column=6).value)[-1].isdecimal() == False:
                    pass
                else:
                    ### 水害原因が「10:破堤」「20:有堤部溢水」「30:無堤部溢水」「40:内水」の場合、相関する工種は、「1:河川」である。
                    if split_name_code(ws_ippan[i].cell(row=7, column=8).value)[-1] in [10, 20, 30, 40]:
                        if split_name_code(ws_ippan[i].cell(row=14, column=6).value)[-1] == 1:
                            correlate_OK_list.append([ws_ippan[i].title, 7, 8, 311])
                            correlate_OK_list.append([ws_ippan[i].title, 14, 6, 311])
                        else:
                            add_comment(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=7, column=8, fill=fill, com_id=311)
                            add_comment(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=14, column=6, fill=fill, com_id=311)
                            correlate_NG_list.append([ws_ippan[i].title, 7, 8, 311])
                            correlate_NG_list.append([ws_ippan[i].title, 14, 6, 311])
                            
                    ### 水害原因が「50:窪地内水」「80:地すべり」「90:急傾斜地崩壊」の場合、相関する工種は、「3:河川海岸以外」である。
                    elif split_name_code(ws_ippan[i].cell(row=7, column=8).value)[-1] in [50, 80, 90]:
                        if split_name_code(ws_ippan[i].cell(row=14, column=6).value)[-1] == 3:
                            correlate_OK_list.append([ws_ippan[i].title, 7, 8, 312])
                            correlate_OK_list.append([ws_ippan[i].title, 14, 6, 312])
                        else:
                            add_comment(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=7, column=8, fill=fill, com_id=312)
                            add_comment(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=14, column=6, fill=fill, com_id=312)
                            correlate_NG_list.append([ws_ippan[i].title, 7, 8, 312])
                            correlate_NG_list.append([ws_ippan[i].title, 14, 6, 312])
                            
                    ### 水害原因が「93:波浪」の場合、相関する工種は、「2:海岸」である。
                    elif split_name_code(ws_ippan[i].cell(row=7, column=8).value)[-1] in [93]:
                        if split_name_code(ws_ippan[i].cell(row=14, column=6).value)[-1] == 2:
                            correlate_OK_list.append([ws_ippan[i].title, 7, 8, 313])
                            correlate_OK_list.append([ws_ippan[i].title, 14, 6, 313])
                        else:
                            add_comment(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=7, column=8, fill=fill, com_id=313)
                            add_comment(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=14, column=6, fill=fill, com_id=313)
                            correlate_NG_list.append([ws_ippan[i].title, 7, 8, 313])
                            correlate_NG_list.append([ws_ippan[i].title, 14, 6, 313])
                            
                    ### 水害原因が「60:洗堀・流出」「91:高潮」「92:津波」の場合、相関する工種は、「1:河川」「2:海岸」である。
                    elif split_name_code(ws_ippan[i].cell(row=7, column=8).value)[-1] in [60, 91, 92]:
                        if split_name_code(ws_ippan[i].cell(row=14, column=6).value)[-1] == 1 or \
                            split_name_code(ws_ippan[i].cell(row=14, column=6).value)[-1] == 2:
                            correlate_OK_list.append([ws_ippan[i].title, 7, 8, 314])
                            correlate_OK_list.append([ws_ippan[i].title, 14, 6, 314])
                        else:
                            add_comment(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=7, column=8, fill=fill, com_id=314)
                            add_comment(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=14, column=6, fill=fill, com_id=314)
                            correlate_NG_list.append([ws_ippan[i].title, 7, 8, 314])
                            correlate_NG_list.append([ws_ippan[i].title, 14, 6, 314])
                            
                    ### 水害原因が「70:土石流」の場合、相関する工種は、「1:河川」「3:河川海岸以外」である。
                    elif split_name_code(ws_ippan[i].cell(row=7, column=8).value)[-1] in [70]:
                        if split_name_code(ws_ippan[i].cell(row=14, column=6).value)[-1] == 1 or \
                            split_name_code(ws_ippan[i].cell(row=14, column=6).value)[-1] == 3:
                            correlate_OK_list.append([ws_ippan[i].title, 7, 8, 315])
                            correlate_OK_list.append([ws_ippan[i].title, 14, 6, 315])
                        else:
                            add_comment(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=7, column=8, fill=fill, com_id=315)
                            add_comment(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=14, column=6, fill=fill, com_id=315)
                            correlate_NG_list.append([ws_ippan[i].title, 7, 8, 315])
                            correlate_NG_list.append([ws_ippan[i].title, 14, 6, 315])
                            
                    else:
                        pass
    
            ### セル[7:9]: 水害区域番号が何かの値のときに、相関する他項目は正しく選択されているか。
    
        #######################################################################
        ### EXCELセルデータ相関チェック処理（4010）
        ### (1)セル[10:2]からセル[10:6]について他項目との相関関係が正しいことをチェックする。
        ### (2)チェック結果リストにセルの行、列とメッセージを追加する。
        ### (3)IPPANワークシートとRESULTワークシートのセルに背景赤色の塗りつぶしをセットする。
        #######################################################################
        print_log('[DEBUG] P0300IppanUpload.index_view()関数 STEP 23/35.', 'DEBUG')
        for i, _ in enumerate(ws_ippan):
            ### セル[10:2]: 水系・沿岸名が何かの値のときに、相関する水系種別は正しく選択されているか。
            ### if ws_ippan[i].cell(row=10, column=2).value is None:
            ###     result_correlate_list.append([ws_ippan[i].title, 10, 2, MESSAGE[][0], MESSAGE[][1], MESSAGE[][2], MESSAGE[][3], MESSAGE[][4]])

            ### セル[10:2]: 水系・沿岸名が何かの値のときに、相関する工種は正しく選択されているか。
            ### if ws_ippan[i].cell(row=10, column=2).value is None or \
            ###     ws_ippan[i].cell(row=14, column=6).value is None:
            ###     pass
            ### else:
            ###     if split_name_code(ws_ippan[i].cell(row=10, column=2).value)[-1].isdecimal() == False or \
            ###         split_name_code(ws_ippan[i].cell(row=14, column=6).value)[-1].isdecimal() == False:
            ###         pass
            ###     else:
            ###         ### 水系・沿岸名が「水系」のときに、相関する工種は、「1:河川」である。
            ###         if split_name_code(ws_ippan[i].cell(row=10, column=2).value)[-1] == :
            ###             if split_name_code(ws_ippan[i].cell(row=14, column=6).value)[-1] in [1]:
            ###                 correlate_OK_list.append([ws_ippan[i].title, 14, 6, ])
            ###             else:
            ###                 add_comment(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=14, column=6, fill=fill, com_id=)
            ###                 correlate_NG_list.append([ws_ippan[i].title, 14, 6, ])
            ###         ### 水系・沿岸名が「沿岸」のときに、相関する工種は、「2:海岸」である。
            ###         elif split_name_code(ws_ippan[i].cell(row=10, column=2).value)[-1] == :
            ###             if split_name_code(ws_ippan[i].cell(row=14, column=6).value)[-1] in [2]:
            ###                 correlate_OK_list.append([ws_ippan[i].title, 14, 6, ])
            ###             else:
            ###                 add_comment(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=14, column=6, fill=fill, com_id=)
            ###                 correlate_NG_list.append([ws_ippan[i].title, 14, 6, ])
            ###         ### 水系・沿岸名が「河川海岸以外」のときに、相関する工種は、「3:河川海岸以外」である。
            ###         elif split_name_code(ws_ippan[i].cell(row=10, column=2).value)[-1] == :
            ###             if split_name_code(ws_ippan[i].cell(row=14, column=6).value)[-1] in [3]:
            ###                 correlate_OK_list.append([ws_ippan[i].title, 14, 6, ])
            ###             else:
            ###                 add_comment(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=14, column=6, fill=fill, com_id=)
            ###                 correlate_NG_list.append([ws_ippan[i].title, 14, 6, ])
            ###         else:
            ###             pass
            
            ### セル[10:3]: 水系種別 vs セル[10:5]: 河川種別
            if ws_ippan[i].cell(row=10, column=3).value is None or \
                ws_ippan[i].cell(row=10, column=5).value is None:
                pass
            else:
                if split_name_code(ws_ippan[i].cell(row=10, column=3).value)[-1].isdecimal() == False or \
                    split_name_code(ws_ippan[i].cell(row=10, column=5).value)[-1].isdecimal() == False:
                    pass
                else:
                    ### 水系種別が「1:一級」のときに、相関する河川種別は、「1:直轄」「2:指定」「4:準用」「5:普通」である。
                    if split_name_code(ws_ippan[i].cell(row=10, column=3).value)[-1] == 1:
                        if split_name_code(ws_ippan[i].cell(row=10, column=5).value)[-1] in [1, 2, 4, 5]:
                            correlate_OK_list.append([ws_ippan[i].title, 10, 3, 320])
                            correlate_OK_list.append([ws_ippan[i].title, 10, 5, 320])
                        else:
                            add_comment(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=10, column=3, fill=fill, com_id=320)
                            add_comment(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=10, column=5, fill=fill, com_id=320)
                            correlate_NG_list.append([ws_ippan[i].title, 10, 3, 320])
                            correlate_NG_list.append([ws_ippan[i].title, 10, 5, 320])

                    ### 水系種別が「2:二級」のときに、相関する河川種別は、「3:二級」「4:準用」「5:普通」である。
                    elif split_name_code(ws_ippan[i].cell(row=10, column=3).value)[-1] == 2:
                        if split_name_code(ws_ippan[i].cell(row=10, column=5).value)[-1] in [3, 4, 5]:
                            correlate_OK_list.append([ws_ippan[i].title, 10, 3, 321])
                            correlate_OK_list.append([ws_ippan[i].title, 10, 5, 321])
                        else:
                            add_comment(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=10, column=3, fill=fill, com_id=321)
                            add_comment(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=10, column=5, fill=fill, com_id=321)
                            correlate_NG_list.append([ws_ippan[i].title, 10, 3, 321])
                            correlate_NG_list.append([ws_ippan[i].title, 10, 5, 321])

                    ### 水系種別が「3:準用」のときに、相関する河川種別は、「4:準用」「5:普通」である。
                    elif split_name_code(ws_ippan[i].cell(row=10, column=3).value)[-1] == 3:
                        if split_name_code(ws_ippan[i].cell(row=10, column=5).value)[-1] in [4, 5]:
                            correlate_OK_list.append([ws_ippan[i].title, 10, 3, 322])
                            correlate_OK_list.append([ws_ippan[i].title, 10, 5, 322])
                        else:
                            add_comment(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=10, column=3, fill=fill, com_id=322)
                            add_comment(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=10, column=5, fill=fill, com_id=322)
                            correlate_NG_list.append([ws_ippan[i].title, 10, 3, 322])
                            correlate_NG_list.append([ws_ippan[i].title, 10, 5, 322])

                    ### 水系種別が「4:普通」のときに、相関する河川種別は、「5:普通」である。
                    elif split_name_code(ws_ippan[i].cell(row=10, column=3).value)[-1] == 4:
                        if split_name_code(ws_ippan[i].cell(row=10, column=5).value)[-1] in [5]:
                            correlate_OK_list.append([ws_ippan[i].title, 10, 3, 323])
                            correlate_OK_list.append([ws_ippan[i].title, 10, 5, 323])
                        else:
                            add_comment(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=10, column=3, fill=fill, com_id=323)
                            add_comment(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=10, column=5, fill=fill, com_id=323)
                            correlate_NG_list.append([ws_ippan[i].title, 10, 3, 323])
                            correlate_NG_list.append([ws_ippan[i].title, 10, 5, 323])

                    ### 水系種別が「5:沿岸」のときに、相関する河川種別は、「6:海岸」である。
                    elif split_name_code(ws_ippan[i].cell(row=10, column=3).value)[-1] == 5:
                        if split_name_code(ws_ippan[i].cell(row=10, column=5).value)[-1] in [6]:
                            correlate_OK_list.append([ws_ippan[i].title, 10, 3, 324])
                            correlate_OK_list.append([ws_ippan[i].title, 10, 5, 324])
                        else:
                            add_comment(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=10, column=3, fill=fill, com_id=324)
                            add_comment(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=10, column=5, fill=fill, com_id=324)
                            correlate_NG_list.append([ws_ippan[i].title, 10, 3, 324])
                            correlate_NG_list.append([ws_ippan[i].title, 10, 5, 324])

                    ### 水系種別が「6:河川海岸以外」のときに、相関する河川種別は、「7:河川海岸以外」である。
                    elif split_name_code(ws_ippan[i].cell(row=10, column=3).value)[-1] == 6:
                        if split_name_code(ws_ippan[i].cell(row=10, column=5).value)[-1] in [7]:
                            correlate_OK_list.append([ws_ippan[i].title, 10, 3, 325])
                            correlate_OK_list.append([ws_ippan[i].title, 10, 5, 325])
                        else:
                            add_comment(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=10, column=3, fill=fill, com_id=325)
                            add_comment(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=10, column=5, fill=fill, com_id=325)
                            correlate_NG_list.append([ws_ippan[i].title, 10, 3, 325])
                            correlate_NG_list.append([ws_ippan[i].title, 10, 5, 325])
                            
                    else:
                        pass

            ### セル[10:3]: 水系種別 vs セル[14:6]: 工種
            if ws_ippan[i].cell(row=10, column=3).value is None or \
                ws_ippan[i].cell(row=14, column=6).value is None:
                pass
            else:
                if split_name_code(ws_ippan[i].cell(row=10, column=3).value)[-1].isdecimal() == False or \
                    split_name_code(ws_ippan[i].cell(row=14, column=6).value)[-1].isdecimal() == False:
                    pass
                else:
                    ### 水系種別が「1:一級」「2:二級」「3:準用」「4:普通」のときに、相関する工種は、「1:河川」である。
                    if split_name_code(ws_ippan[i].cell(row=10, column=3).value)[-1] in [1, 2, 3, 4]:
                        if split_name_code(ws_ippan[i].cell(row=14, column=6).value)[-1] in [1]:
                            correlate_OK_list.append([ws_ippan[i].title, 10, 3, 326])
                            correlate_OK_list.append([ws_ippan[i].title, 14, 6, 326])
                        else:
                            add_comment(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=10, column=3, fill=fill, com_id=326)
                            add_comment(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=14, column=6, fill=fill, com_id=326)
                            correlate_NG_list.append([ws_ippan[i].title, 10, 3, 326])
                            correlate_NG_list.append([ws_ippan[i].title, 14, 6, 326])
                
                    ### 水系種別が「5:沿岸」のときに、相関する工種は、「2:海岸」である。
                    elif split_name_code(ws_ippan[i].cell(row=10, column=3).value)[-1] in [5]:
                        if split_name_code(ws_ippan[i].cell(row=14, column=6).value)[-1] in [2]:
                            correlate_OK_list.append([ws_ippan[i].title, 10, 3, 327])
                            correlate_OK_list.append([ws_ippan[i].title, 14, 6, 327])
                        else:
                            add_comment(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=10, column=3, fill=fill, com_id=327)
                            add_comment(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=14, column=6, fill=fill, com_id=327)
                            correlate_NG_list.append([ws_ippan[i].title, 10, 3, 327])
                            correlate_NG_list.append([ws_ippan[i].title, 14, 6, 327])
                            
                    ### 水系種別が「6:河川海岸以外」のときに、相関する工種は、「3:河川海岸以外」である。
                    elif split_name_code(ws_ippan[i].cell(row=10, column=3).value)[-1] in [6]:
                        if split_name_code(ws_ippan[i].cell(row=14, column=6).value)[-1] in [6]:
                            correlate_OK_list.append([ws_ippan[i].title, 10, 3, 328])
                            correlate_OK_list.append([ws_ippan[i].title, 14, 6, 328])
                        else:
                            add_comment(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=10, column=3, fill=fill, com_id=328)
                            add_comment(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=14, column=6, fill=fill, com_id=328)
                            correlate_NG_list.append([ws_ippan[i].title, 10, 3, 328])
                            correlate_NG_list.append([ws_ippan[i].title, 14, 6, 328])

            ### セル[10:4]: 河川・海岸名が何かの値のときに、相関する他項目は正しく選択されているか。
            ### セル[10:5]: 河川種別が何かの値のときに、相関する他項目は正しく選択されているか。
            ### セル[10:6]: 地盤勾配区分が何かの値のときに、相関する他項目は正しく選択されているか。
    
        #######################################################################
        ### EXCELセルデータ相関チェック処理（4020）
        ### (1)セル[14:2]からセル[14:10]について他項目との相関関係が正しいことをチェックする。
        ### (2)チェック結果リストにセルの行、列とメッセージを追加する。
        ### (3)IPPANワークシートとRESULTワークシートのセルに背景赤色の塗りつぶしをセットする。
        #######################################################################
        print_log('[DEBUG] P0300IppanUpload.index_view()関数 STEP 24/35.', 'DEBUG')
        for i, _ in enumerate(ws_ippan):
            ### セル[14:3]: 水害区域面積の農地 vs セル[14:8]: 農作物被害額
            ### セル[14:3]: 水害区域面積の農地がNoneのとき、農作物被害額はNoneが正しい。
            if ws_ippan[i].cell(row=14, column=3).value is None:
                if ws_ippan[i].cell(row=14, column=8).value is None:
                    correlate_OK_list.append([ws_ippan[i].title, 14, 3, 340])
                    correlate_OK_list.append([ws_ippan[i].title, 14, 8, 340])
                else:
                    add_comment(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=14, column=3, fill=fill, com_id=340)
                    add_comment(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=14, column=8, fill=fill, com_id=340)
                    correlate_NG_list.append([ws_ippan[i].title, 14, 3, 340])
                    correlate_NG_list.append([ws_ippan[i].title, 14, 8, 340])
            else:
                if ws_ippan[i].cell(row=14, column=8).value is None:
                    add_comment(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=14, column=3, fill=fill, com_id=341)
                    add_comment(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=14, column=8, fill=fill, com_id=341)
                    correlate_NG_list.append([ws_ippan[i].title, 14, 3, 341])
                    correlate_NG_list.append([ws_ippan[i].title, 14, 8, 341])
                else:
                    correlate_OK_list.append([ws_ippan[i].title, 14, 3, 341])
                    correlate_OK_list.append([ws_ippan[i].title, 14, 8, 341])

            ### セル[14:2]: 水害区域面積の宅地が何かの値のときに、相関する他項目は正しく選択されているか。
            ### セル[14:3]: 水害区域面積の農地が何かの値のときに、相関する他項目は正しく選択されているか。
            ### セル[14:4]: 水害区域面積の地下が何かの値のときに、相関する他項目は正しく選択されているか。
            ### 水害区域面積の宅地、農地、地下のいずれかに入力されているか。
            if ws_ippan[i].cell(row=14, column=2).value is None and \
                ws_ippan[i].cell(row=14, column=3).value is None and \
                ws_ippan[i].cell(row=14, column=4).value is None:
                add_comment(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=14, column=2, fill=fill, com_id=342)
                add_comment(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=14, column=3, fill=fill, com_id=342)
                add_comment(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=14, column=4, fill=fill, com_id=342)
                correlate_NG_list.append([ws_ippan[i].title, 14, 2, 342])
                correlate_NG_list.append([ws_ippan[i].title, 14, 3, 342])
                correlate_NG_list.append([ws_ippan[i].title, 14, 4, 342])
            else:
                correlate_OK_list.append([ws_ippan[i].title, 14, 2, 342])
                correlate_OK_list.append([ws_ippan[i].title, 14, 3, 342])
                correlate_OK_list.append([ws_ippan[i].title, 14, 4, 342])
            
            ### セル[14:6]: 工種が何かの値のときに、相関する他項目は正しく選択されているか。
            ### セル[14:8]: 農作物被害額が何かの値のときに、相関する他項目は正しく選択されているか。
            ### セル[14:10]: 異常気象が何かの値のときに、相関する他項目は正しく選択されているか。
    
        #######################################################################
        ### EXCELセルデータ相関チェック処理（4030）
        ### (1)セル[20:2]からセル[20:27]について他項目との相関関係が正しいことをチェックする。
        ### (2)チェック結果リストにセルの行、列とメッセージを追加する。
        ### (3)IPPANワークシートとRESULTワークシートのセルに背景赤色の塗りつぶしをセットする。
        #######################################################################
        print_log('[DEBUG] P0300IppanUpload.index_view()関数 STEP 25/35.', 'DEBUG')
        for i, _ in enumerate(ws_ippan):
            if max_row[i] >= 20:
                for j in range(20, max_row[i] + 1):
                    ### セル[20:2]: 町丁名・大字名が何かの値のときに、相関する他項目は正しく選択されているか。
                    ### セル[20:3]: 名称が何かの値のときに、相関する他項目は正しく選択されているか。
                        
                    ### セル[20:4]: 地上・地下被害の区分 vs セル[14:2]: 水害区域面積の宅地、またはセル[14:3]: 水害区域面積の農地
                    if ws_ippan[i].cell(row=j, column=4).value is None:
                        pass
                    else:
                        if split_name_code(ws_ippan[i].cell(row=j, column=4).value)[-1].isdecimal() == False:
                            pass
                        else:
                            ### 地上・地下被害の区分が「地上のみ:1」のときに、相関する水害区域面積の宅地または水害区域面積の農地が入力されているか。
                            if split_name_code(ws_ippan[i].cell(row=j, column=4).value)[-1] in [1]:
                                if ws_ippan[i].cell(row=14, column=2).value is None and \
                                    ws_ippan[i].cell(row=14, column=3).value is None:
                                    add_comment(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=j, column=4, fill=fill, com_id=360)
                                    correlate_NG_list.append([ws_ippan[i].title, j, 4, 360])
                                else:
                                    correlate_OK_list.append([ws_ippan[i].title, j, 4, 360])

                            ### 地上・地下被害の区分が「地上部分:2」のときに、相関する水害区域面積の宅地または水害区域面積の農地が入力されているか。
                            elif split_name_code(ws_ippan[i].cell(row=j, column=4).value)[-1] in [2]:
                                if ws_ippan[i].cell(row=14, column=2).value is None and \
                                    ws_ippan[i].cell(row=14, column=3).value is None:
                                    add_comment(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=j, column=4, fill=fill, com_id=361)
                                    correlate_NG_list.append([ws_ippan[i].title, j, 4, 361])
                                else:
                                    correlate_OK_list.append([ws_ippan[i].title, j, 4, 361])

                            ### 地上・地下被害の区分が「地下部分:3」のときに、相関する水害区域面積の地下が入力されているか。
                            elif split_name_code(ws_ippan[i].cell(row=j, column=4).value)[-1] in [3]:
                                if ws_ippan[i].cell(row=14, column=4).value is None:
                                    add_comment(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=j, column=4, fill=fill, com_id=362)
                                    correlate_NG_list.append([ws_ippan[i].title, j, 4, 362])
                                else:
                                    correlate_OK_list.append([ws_ippan[i].title, j, 4, 362])

                            ### 地上・地下被害の区分が「地下のみ:4」のときに、相関する水害区域面積の地下が入力されているか。
                            elif split_name_code(ws_ippan[i].cell(row=j, column=4).value)[-1] in [4]:
                                if ws_ippan[i].cell(row=14, column=4).value is None:
                                    add_comment(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=j, column=4, fill=fill, com_id=363)
                                    correlate_NG_list.append([ws_ippan[i].title, j, 4, 363])
                                else:
                                    correlate_OK_list.append([ws_ippan[i].title, j, 4, 363])
                            
                            else:
                                pass

                    ### セル[20:4]: 地上・地下被害の区分 vs セル[20:26]: 地下空間の利用形態
                    if ws_ippan[i].cell(row=j, column=4).value is None:
                        pass
                    else:
                        if split_name_code(ws_ippan[i].cell(row=j, column=4).value)[-1].isdecimal() == False:
                            pass
                        else:
                            ### 地上・地下被害の区分が「1:」「2:」の場合、相関する地下空間の利用形態は、何らかの値が入力される。
                            if split_name_code(ws_ippan[i].cell(row=j, column=4).value)[-1] in [1, 2]:
                                if ws_ippan[i].cell(row=j, column=27).value is None:
                                    correlate_OK_list.append([ws_ippan[i].title, j, 4, 364])
                                    correlate_OK_list.append([ws_ippan[i].title, j, 27, 364])
                                else:
                                    add_comment(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=j, column=4, fill=fill, com_id=364)
                                    add_comment(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=j, column=27, fill=fill, com_id=364)
                                    correlate_NG_list.append([ws_ippan[i].title, j, 4, 364])
                                    correlate_NG_list.append([ws_ippan[i].title, j, 27, 364])

                            elif split_name_code(ws_ippan[i].cell(row=j, column=4).value)[-1] in [3, 4]:
                                if ws_ippan[i].cell(row=j, column=27).value is None:
                                    add_comment(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=j, column=4, fill=fill, com_id=365)
                                    add_comment(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=j, column=27, fill=fill, com_id=365)
                                    correlate_NG_list.append([ws_ippan[i].title, j, 4, 365])
                                    correlate_NG_list.append([ws_ippan[i].title, j, 27, 365])
                                else:
                                    correlate_OK_list.append([ws_ippan[i].title, j, 4, 365])
                                    correlate_OK_list.append([ws_ippan[i].title, j, 27, 365])
                        
                    ### セル[20:5]: 浸水土砂被害の区分が何かの値のときに、相関する他項目は正しく選択されているか。
                        
                    ### セル[20:6]: 被害建物棟数, 床下浸水 vs セル[20:12]: 延床面積
                    ### セル[20:7]: 被害建物棟数, 1cm〜49cm vs セル[20:12]: 延床面積
                    ### セル[20:8]: 被害建物棟数, 50cm〜99cm vs セル[20:12]: 延床面積
                    ### セル[20:9]: 被害建物棟数, 1m以上 vs セル[20:12]: 延床面積
                    ### セル[20:10]: 被害建物棟数, 半壊 vs セル[20:12]: 延床面積
                    ### セル[20:11]: 被害建物棟数, 全壊・流失 vs セル[20:12]: 延床面積
                    if ws_ippan[i].cell(row=j, column=6).value is None and \
                        ws_ippan[i].cell(row=j, column=7).value is None and \
                        ws_ippan[i].cell(row=j, column=8).value is None and \
                        ws_ippan[i].cell(row=j, column=9).value is None and \
                        ws_ippan[i].cell(row=j, column=10).value is None and \
                        ws_ippan[i].cell(row=j, column=11).value is None:
                        if ws_ippan[i].cell(row=j, column=12).value is None:
                            correlate_OK_list.append([ws_ippan[i].title, j, 6, 366])
                            correlate_OK_list.append([ws_ippan[i].title, j, 7, 366])
                            correlate_OK_list.append([ws_ippan[i].title, j, 8, 366])
                            correlate_OK_list.append([ws_ippan[i].title, j, 9, 366])
                            correlate_OK_list.append([ws_ippan[i].title, j, 10, 366])
                            correlate_OK_list.append([ws_ippan[i].title, j, 11, 366])
                            correlate_OK_list.append([ws_ippan[i].title, j, 12, 366])
                        else:
                            add_comment(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=j, column=6, fill=fill, com_id=366)
                            add_comment(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=j, column=7, fill=fill, com_id=366)
                            add_comment(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=j, column=8, fill=fill, com_id=366)
                            add_comment(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=j, column=9, fill=fill, com_id=366)
                            add_comment(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=j, column=10, fill=fill, com_id=366)
                            add_comment(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=j, column=11, fill=fill, com_id=366)
                            add_comment(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=j, column=12, fill=fill, com_id=366)
                            correlate_NG_list.append([ws_ippan[i].title, j, 6, 366])
                            correlate_NG_list.append([ws_ippan[i].title, j, 7, 366])
                            correlate_NG_list.append([ws_ippan[i].title, j, 8, 366])
                            correlate_NG_list.append([ws_ippan[i].title, j, 9, 366])
                            correlate_NG_list.append([ws_ippan[i].title, j, 10, 366])
                            correlate_NG_list.append([ws_ippan[i].title, j, 11, 366])
                            correlate_NG_list.append([ws_ippan[i].title, j, 12, 366])
                    else:
                        if ws_ippan[i].cell(row=j, column=12).value is None:
                            add_comment(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=j, column=6, fill=fill, com_id=367)
                            add_comment(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=j, column=7, fill=fill, com_id=367)
                            add_comment(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=j, column=8, fill=fill, com_id=367)
                            add_comment(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=j, column=9, fill=fill, com_id=367)
                            add_comment(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=j, column=10, fill=fill, com_id=367)
                            add_comment(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=j, column=11, fill=fill, com_id=367)
                            add_comment(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=j, column=12, fill=fill, com_id=367)
                            correlate_NG_list.append([ws_ippan[i].title, j, 6, 367])
                            correlate_NG_list.append([ws_ippan[i].title, j, 7, 367])
                            correlate_NG_list.append([ws_ippan[i].title, j, 8, 367])
                            correlate_NG_list.append([ws_ippan[i].title, j, 9, 367])
                            correlate_NG_list.append([ws_ippan[i].title, j, 10, 367])
                            correlate_NG_list.append([ws_ippan[i].title, j, 11, 367])
                            correlate_NG_list.append([ws_ippan[i].title, j, 12, 367])
                        else:
                            correlate_OK_list.append([ws_ippan[i].title, j, 6, 367])
                            correlate_OK_list.append([ws_ippan[i].title, j, 7, 367])
                            correlate_OK_list.append([ws_ippan[i].title, j, 8, 367])
                            correlate_OK_list.append([ws_ippan[i].title, j, 9, 367])
                            correlate_OK_list.append([ws_ippan[i].title, j, 10, 367])
                            correlate_OK_list.append([ws_ippan[i].title, j, 11, 367])
                            correlate_OK_list.append([ws_ippan[i].title, j, 12, 367])
                        
                    ### セル[20:12]: 被害建物の延床面積が何かの値のときに、相関する他項目は正しく選択されているか。
                    ### セル[20:13]: 被災世帯数が何かの値のときに、相関する他項目は正しく選択されているか。
                        
                    ### セル[20:14]: 被災事業所数 vs セル[20:25]: 事業所の産業区分
                    if ws_ippan[i].cell(row=j, column=14).value is None:
                        if ws_ippan[i].cell(row=j, column=25).value is None:
                            correlate_OK_list.append([ws_ippan[i].title, j, 14, 368])
                            correlate_OK_list.append([ws_ippan[i].title, j, 25, 368])
                        else:
                            add_comment(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=j, column=14, fill=fill, com_id=368)
                            add_comment(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=j, column=25, fill=fill, com_id=368)
                            correlate_NG_list.append([ws_ippan[i].title, j, 14, 368])
                            correlate_NG_list.append([ws_ippan[i].title, j, 25, 368])
                    else:
                        if ws_ippan[i].cell(row=j, column=25).value is None:
                            add_comment(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=j, column=14, fill=fill, com_id=369)
                            add_comment(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=j, column=25, fill=fill, com_id=369)
                            correlate_NG_list.append([ws_ippan[i].title, j, 14, 369])
                            correlate_NG_list.append([ws_ippan[i].title, j, 25, 369])
                        else:
                            correlate_OK_list.append([ws_ippan[i].title, j, 14, 369])
                            correlate_OK_list.append([ws_ippan[i].title, j, 25, 369])

                    ### セル[20:14]: 被災事業所数 vs セル[20:20]: 事業所従業者数, 床下浸水
                    ### セル[20:14]: 被災事業所数 vs セル[20:21]: 事業所従業者数, 1cm〜49cm
                    ### セル[20:14]: 被災事業所数 vs セル[20:22]: 事業所従業者数, 50cm〜99cm
                    ### セル[20:14]: 被災事業所数 vs セル[20:23]: 事業所従業者数, 1m以上・半壊
                    ### セル[20:14]: 被災事業所数 vs セル[20:24]: 事業所従業者数, 全壊・流失
                    if ws_ippan[i].cell(row=j, column=14).value is None:
                        if ws_ippan[i].cell(row=j, column=20).value is None and \
                            ws_ippan[i].cell(row=j, column=21).value is None and \
                            ws_ippan[i].cell(row=j, column=22).value is None and \
                            ws_ippan[i].cell(row=j, column=23).value is None and \
                            ws_ippan[i].cell(row=j, column=24).value is None:
                            correlate_OK_list.append([ws_ippan[i].title, j, 14, 370])
                            correlate_OK_list.append([ws_ippan[i].title, j, 20, 370])
                            correlate_OK_list.append([ws_ippan[i].title, j, 21, 370])
                            correlate_OK_list.append([ws_ippan[i].title, j, 22, 370])
                            correlate_OK_list.append([ws_ippan[i].title, j, 23, 370])
                            correlate_OK_list.append([ws_ippan[i].title, j, 24, 370])
                        else:
                            add_comment(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=j, column=14, fill=fill, com_id=370)
                            add_comment(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=j, column=20, fill=fill, com_id=370)
                            add_comment(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=j, column=21, fill=fill, com_id=370)
                            add_comment(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=j, column=22, fill=fill, com_id=370)
                            add_comment(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=j, column=23, fill=fill, com_id=370)
                            add_comment(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=j, column=24, fill=fill, com_id=370)
                            correlate_NG_list.append([ws_ippan[i].title, j, 14, 370])
                            correlate_NG_list.append([ws_ippan[i].title, j, 20, 370])
                            correlate_NG_list.append([ws_ippan[i].title, j, 21, 370])
                            correlate_NG_list.append([ws_ippan[i].title, j, 22, 370])
                            correlate_NG_list.append([ws_ippan[i].title, j, 23, 370])
                            correlate_NG_list.append([ws_ippan[i].title, j, 24, 370])
                    else:
                        if ws_ippan[i].cell(row=j, column=20).value is None and \
                            ws_ippan[i].cell(row=j, column=21).value is None and \
                            ws_ippan[i].cell(row=j, column=22).value is None and \
                            ws_ippan[i].cell(row=j, column=23).value is None and \
                            ws_ippan[i].cell(row=j, column=24).value is None:
                            add_comment(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=j, column=14, fill=fill, com_id=371)
                            add_comment(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=j, column=20, fill=fill, com_id=371)
                            add_comment(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=j, column=21, fill=fill, com_id=371)
                            add_comment(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=j, column=22, fill=fill, com_id=371)
                            add_comment(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=j, column=23, fill=fill, com_id=371)
                            add_comment(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=j, column=24, fill=fill, com_id=371)
                            correlate_NG_list.append([ws_ippan[i].title, j, 14, 371])
                            correlate_NG_list.append([ws_ippan[i].title, j, 20, 371])
                            correlate_NG_list.append([ws_ippan[i].title, j, 21, 371])
                            correlate_NG_list.append([ws_ippan[i].title, j, 22, 371])
                            correlate_NG_list.append([ws_ippan[i].title, j, 23, 371])
                            correlate_NG_list.append([ws_ippan[i].title, j, 24, 371])
                        else:
                            correlate_OK_list.append([ws_ippan[i].title, j, 14, 371])
                            correlate_OK_list.append([ws_ippan[i].title, j, 20, 371])
                            correlate_OK_list.append([ws_ippan[i].title, j, 21, 371])
                            correlate_OK_list.append([ws_ippan[i].title, j, 22, 371])
                            correlate_OK_list.append([ws_ippan[i].title, j, 23, 371])
                            correlate_OK_list.append([ws_ippan[i].title, j, 24, 371])
                        
                    ### セル[20:15]: 農家・漁家戸数, 床下浸水が何かの値のときに、相関する他項目は正しく選択されているか。
                    ### セル[20:16]: 農家・漁家戸数, 1cm〜49cmが何かの値のときに、相関する他項目は正しく選択されているか。
                    ### セル[20:17]: 農家・漁家戸数, 50cm〜99cmが何かの値のときに、相関する他項目は正しく選択されているか。
                    ### セル[20:18]: 農家・漁家戸数, 1m以上・半壊が何かの値のときに、相関する他項目は正しく選択されているか。
                    ### セル[20:19]: 農家・漁家戸数, 全壊・流失が何かの値のときに、相関する他項目は正しく選択されているか。
                        
                    ### セル[20:20]: 事業所従業者数, 床下浸水が何かの値のときに、相関する他項目は正しく選択されているか。 参照 セル[20:14]: 被災事業所数 vs セル[20:20]: 事業所従業者数
                    ### セル[20:21]: 事業所従業者数, 1cm〜49cmが何かの値のときに、相関する他項目は正しく選択されているか。 参照 セル[20:14]: 被災事業所数 vs セル[20:20]: 事業所従業者数
                    ### セル[20:22]: 事業所従業者数, 50cm〜99cmが何かの値のときに、相関する他項目は正しく選択されているか。 参照 セル[20:14]: 被災事業所数 vs セル[20:20]: 事業所従業者数
                    ### セル[20:23]: 事業所従業者数, 1m以上・半壊が何かの値のときに、相関する他項目は正しく選択されているか。 参照 セル[20:14]: 被災事業所数 vs セル[20:20]: 事業所従業者数
                    ### セル[20:24]: 事業所従業者数, 全壊・流失が何かの値のときに、相関する他項目は正しく選択されているか。 参照 セル[20:14]: 被災事業所数 vs セル[20:20]: 事業所従業者数
                        
                    ### セル[20:25]: 事業所の産業区分が何かの値のときに、相関する他項目は正しく選択されているか。 参照  
                    ### セル[20:26]: 地下空間の利用形態が何かの値のときに、相関する他項目は正しく選択されているか。 参照 
                    ### セル[20:27]: 備考が何かの値のときに、相関する他項目は正しく選択されているか。
    
        #######################################################################
        #######################################################################
        ### EXCELセルデータ突合チェック処理（5000）
        ### (1)セル[7:2]からセル[7:9]についてデータベースに登録されている値と突合せチェックする。
        ### (2)チェック結果リストにセルの行、列とメッセージを追加する。
        ### (3)IPPANワークシートとRESULTワークシートのセルに背景赤色の塗りつぶしをセットする。
        #######################################################################
        #######################################################################
        print_log('[DEBUG] P0300IppanUpload.index_view()関数 STEP 26/35.', 'DEBUG')
        for i, _ in enumerate(ws_ippan):
            ### セル[7:2]: 都道府県についてデータベースに登録されている値と突合せチェックする。
            if ws_ippan[i].cell(row=7, column=2).value is None:
                pass
            else:
                if ws_ippan[i].cell(row=7, column=2).value not in list(ken_code_list) and \
                    ws_ippan[i].cell(row=7, column=2).value not in list(ken_name_list) and \
                    ws_ippan[i].cell(row=7, column=2).value not in list(ken_name_code_list):
                    add_comment(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=7, column=2, fill=fill, com_id=400)
                    compare_NG_list.append([ws_ippan[i].title, 7, 2, 400])
                else:
                    compare_OK_list.append([ws_ippan[i].title, 7, 2, 400])
                
            ### セル[7:3]: 市区町村についてデータベースに登録されている値と突合せチェックする。
            if ws_ippan[i].cell(row=7, column=3).value is None:
                pass
            else:
                if ws_ippan[i].cell(row=7, column=3).value not in list(city_code_list) and \
                    ws_ippan[i].cell(row=7, column=3).value not in list(city_name_list) and \
                    ws_ippan[i].cell(row=7, column=3).value not in list(city_name_code_list):
                    add_comment(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=7, column=3, fill=fill, com_id=401)
                    compare_NG_list.append([ws_ippan[i].title, 7, 3, 401])
                else:
                    compare_OK_list.append([ws_ippan[i].title, 7, 3, 401])
                
            ### セル[7:4]: 水害発生月日についてデータベースに登録されている値と突合せチェックする。
            ### セル[7:5]: 水害終了月日についてデータベースに登録されている値と突合せチェックする。
                
            ### セル[7:6]: 水害原因1についてデータベースに登録されている値と突合せチェックする。
            if ws_ippan[i].cell(row=7, column=6).value is None:
                pass
            else:
                if ws_ippan[i].cell(row=7, column=6).value not in list(cause_code_list) and \
                    ws_ippan[i].cell(row=7, column=6).value not in list(cause_name_list) and \
                    ws_ippan[i].cell(row=7, column=6).value not in list(cause_name_code_list):
                    add_comment(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=7, column=6, fill=fill, com_id=404)
                    compare_NG_list.append([ws_ippan[i].title, 7, 6, 404])
                else:
                    compare_OK_list.append([ws_ippan[i].title, 7, 6, 404])
                
            ### セル[7:7]: 水害原因2についてデータベースに登録されている値と突合せチェックする。
            if ws_ippan[i].cell(row=7, column=7).value is None:
                pass
            else:
                if ws_ippan[i].cell(row=7, column=7).value not in list(cause_code_list) and \
                    ws_ippan[i].cell(row=7, column=7).value not in list(cause_name_list) and \
                    ws_ippan[i].cell(row=7, column=7).value not in list(cause_name_code_list):
                    add_comment(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=7, column=7, fill=fill, com_id=405)
                    compare_NG_list.append([ws_ippan[i].title, 7, 7, 405])
                else:
                    compare_OK_list.append([ws_ippan[i].title, 7, 7, 405])
                
            ### セル[7:8]: 水害原因3についてデータベースに登録されている値と突合せチェックする。
            if ws_ippan[i].cell(row=7, column=8).value is None:
                pass
            else:
                if ws_ippan[i].cell(row=7, column=8).value not in list(cause_code_list) and \
                    ws_ippan[i].cell(row=7, column=8).value not in list(cause_name_list) and \
                    ws_ippan[i].cell(row=7, column=8).value not in list(cause_name_code_list):
                    add_comment(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=7, column=8, fill=fill, com_id=406)
                    compare_NG_list.append([ws_ippan[i].title, 7, 8, 406])
                else:
                    compare_OK_list.append([ws_ippan[i].title, 7, 8, 406])
                
            ### セル[7:9]: 水害区域番号についてデータベースに登録されている値と突合せチェックする。
            if ws_ippan[i].cell(row=7, column=9).value is None:
                pass
            else:
                if ws_ippan[i].cell(row=7, column=9).value not in list(area_id_list) and \
                    ws_ippan[i].cell(row=7, column=9).value not in list(area_name_list) and \
                    ws_ippan[i].cell(row=7, column=9).value not in list(area_name_id_list):
                    add_comment(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=7, column=9, fill=fill, com_id=407)
                    compare_NG_list.append([ws_ippan[i].title, 7, 9, 407])
                else:
                    compare_OK_list.append([ws_ippan[i].title, 7, 9, 407])
    
        #######################################################################
        ### EXCELセルデータ突合チェック処理（5010）
        ### (1)セル[10:2]からセル[10:6]についてデータベースに登録されている値と突合せチェックする。
        ### (2)チェック結果リストにセルの行、列とメッセージを追加する。
        ### (3)IPPANワークシートとRESULTワークシートのセルに背景赤色の塗りつぶしをセットする。
        ### TO-DO: if == ''はダミーの処理である。突合せチェック処理を記述する。
        #######################################################################
        print_log('[DEBUG] P0300IppanUpload.index_view()関数 STEP 27/35.', 'DEBUG')
        for i, _ in enumerate(ws_ippan):
            ### セル[10:2]: 水系・沿岸名についてデータベースに登録されている値と突合せチェックする。
            if ws_ippan[i].cell(row=10, column=2).value is None:
                pass
            else:
                if ws_ippan[i].cell(row=10, column=2).value not in list(suikei_code_list) and \
                    ws_ippan[i].cell(row=10, column=2).value not in list(suikei_name_list) and \
                    ws_ippan[i].cell(row=10, column=2).value not in list(suikei_name_code_list):
                    add_comment(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=10, column=2, fill=fill, com_id=408)
                    compare_NG_list.append([ws_ippan[i].title, 10, 2, 408])
                else:
                    compare_OK_list.append([ws_ippan[i].title, 10, 2, 408])
                
            ### セル[10:3]: 水系種別についてデータベースに登録されている値と突合せチェックする。
            if ws_ippan[i].cell(row=10, column=3).value is None:
                pass
            else:
                if ws_ippan[i].cell(row=10, column=3).value not in list(suikei_type_code_list) and \
                    ws_ippan[i].cell(row=10, column=3).value not in list(suikei_type_name_list) and \
                    ws_ippan[i].cell(row=10, column=3).value not in list(suikei_type_name_code_list):
                    add_comment(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=10, column=3, fill=fill, com_id=409)
                    compare_NG_list.append([ws_ippan[i].title, 10, 3, 409])
                else:
                    compare_OK_list.append([ws_ippan[i].title, 10, 3, 409])
                
            ### セル[10:4]: 河川・海岸名についてデータベースに登録されている値と突合せチェックする。
            if ws_ippan[i].cell(row=10, column=4).value is None:
                pass
            else:
                if ws_ippan[i].cell(row=10, column=4).value not in list(kasen_code_list) and \
                    ws_ippan[i].cell(row=10, column=4).value not in list(kasen_name_list) and \
                    ws_ippan[i].cell(row=10, column=4).value not in list(kasen_name_code_list):
                    add_comment(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=10, column=4, fill=fill, com_id=410)
                    compare_NG_list.append([ws_ippan[i].title, 10, 4, 410])
                else:
                    compare_OK_list.append([ws_ippan[i].title, 10, 4, 410])
                
            ### セルE10[10:5]: 河川種別についてデータベースに登録されている値と突合せチェックする。
            if ws_ippan[i].cell(row=10, column=5).value is None:
                pass
            else:
                if ws_ippan[i].cell(row=10, column=5).value not in list(kasen_type_code_list) and \
                    ws_ippan[i].cell(row=10, column=5).value not in list(kasen_type_name_list) and \
                    ws_ippan[i].cell(row=10, column=5).value not in list(kasen_type_name_code_list):
                    add_comment(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=10, column=5, fill=fill, com_id=411)
                    compare_NG_list.append([ws_ippan[i].title, 10, 5, 411])
                else:
                    compare_OK_list.append([ws_ippan[i].title, 10, 5, 411])
                
            ### セル[10:6]: 地盤勾配区分についてデータベースに登録されている値と突合せチェックする。
            if ws_ippan[i].cell(row=10, column=6).value is None:
                pass
            else:
                if ws_ippan[i].cell(row=10, column=6).value not in list(gradient_code_list) and \
                    ws_ippan[i].cell(row=10, column=6).value not in list(gradient_name_list) and \
                    ws_ippan[i].cell(row=10, column=6).value not in list(gradient_name_code_list):
                    add_comment(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=10, column=6, fill=fill, com_id=412)
                    compare_NG_list.append([ws_ippan[i].title, 10, 6, 412])
                else:
                    compare_OK_list.append([ws_ippan[i].title, 10, 6, 412])
    
        #######################################################################
        ### EXCELセルデータ突合チェック処理（5020）
        ### (1)セル[14:2]からセル[14:10]についてデータベースに登録されている値と突合せチェックする。
        ### (2)チェック結果リストにセルの行、列とメッセージを追加する。
        ### (3)IPPANワークシートとRESULTワークシートのセルに背景赤色の塗りつぶしをセットする。
        #######################################################################
        print_log('[DEBUG] P0300IppanUpload.index_view()関数 STEP 28/35.', 'DEBUG')
        for i, _ in enumerate(ws_ippan):
            ### セル[14:2]: 水害区域面積の宅地についてデータベースに登録されている値と突合せチェックする。
            ### セル[14:3]: 水害区域面積の農地についてデータベースに登録されている値と突合せチェックする。
            ### セル[14:4]: 水害区域面積の地下についてデータベースに登録されている値と突合せチェックする。
                
            ### セル[14:6]: 工種についてデータベースに登録されている値と突合せチェックする。
            if ws_ippan[i].cell(row=14, column=6).value is None:
                pass
            else:
                if ws_ippan[i].cell(row=14, column=6).value not in list(kasen_kaigan_code_list) and \
                    ws_ippan[i].cell(row=14, column=6).value not in list(kasen_kaigan_name_list) and \
                    ws_ippan[i].cell(row=14, column=6).value not in list(kasen_kaigan_name_code_list):
                    add_comment(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=14, column=2, fill=fill, com_id=416)
                    compare_NG_list.append([ws_ippan[i].title, 14, 6, 416])
                else:
                    compare_OK_list.append([ws_ippan[i].title, 14, 6, 416])
                
            ### セル[14:8]: 農作物被害額についてデータベースに登録されている値と突合せチェックする。
                
            ### セル[14:10]: 異常気象コードについてデータベースに登録されている値と突合せチェックする。
            if ws_ippan[i].cell(row=14, column=10).value is None:
                pass
            else:
                if ws_ippan[i].cell(row=14, column=10).value not in list(weather_id_list) and \
                    ws_ippan[i].cell(row=14, column=10).value not in list(weather_name_list) and \
                    ws_ippan[i].cell(row=14, column=10).value not in list(weather_name_id_list):
                    add_comment(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=14, column=10, fill=fill, com_id=418)
                    compare_NG_list.append([ws_ippan[i].title, 14, 10, 418])
                else:
                    compare_OK_list.append([ws_ippan[i].title, 14, 10, 418])
    
        #######################################################################
        ### EXCELセルデータ突合チェック処理（5030）
        ### (1)セル[20:2]からセル[20:27]についてデータベースに登録されている値と突合せチェックする。
        ### (2)チェック結果リストにセルの行、列とメッセージを追加する。
        ### (3)IPPANワークシートとRESULTワークシートのセルに背景赤色の塗りつぶしをセットする。
        #######################################################################
        print_log('[DEBUG] P0300IppanUpload.index_view()関数 STEP 29/35.', 'DEBUG')
        for i, _ in enumerate(ws_ippan):
            if max_row[i] >= 20:
                for j in range(20, max_row[i] + 1):
                    ### セル[20:2]: 町丁名・大字名についてデータベースに登録されている値と突合せチェックする。
                    
                    ### セル[20:3]: 名称についてデータベースに登録されている値と突合せチェックする。
                    if ws_ippan[i].cell(row=j, column=4).value is None:
                        pass
                    else:
                        if ws_ippan[i].cell(row=j, column=3).value not in list(building_code_list) and \
                            ws_ippan[i].cell(row=j, column=3).value not in list(building_name_list) and \
                            ws_ippan[i].cell(row=j, column=3).value not in list(building_name_code_list):
                            add_comment(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=j, column=3, fill=fill, com_id=451)
                            compare_NG_grid.append([ws_ippan[i].title, j, 3, 451])
                        else:
                            compare_OK_grid.append([ws_ippan[i].title, j, 3, 451])
                        
                    ### セル[20:4]: 地上・地下被害の区分についてデータベースに登録されている値と突合せチェックする。
                    if ws_ippan[i].cell(row=j, column=4).value is None:
                        pass
                    else:
                        if ws_ippan[i].cell(row=j, column=4).value not in list(underground_code_list) and \
                            ws_ippan[i].cell(row=j, column=4).value not in list(underground_name_list) and \
                            ws_ippan[i].cell(row=j, column=4).value not in list(underground_name_code_list):
                            add_comment(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=j, column=4, fill=fill, com_id=452)
                            compare_NG_grid.append([ws_ippan[i].title, j, 4, 452])
                        else:
                            compare_OK_grid.append([ws_ippan[i].title, j, 4, 452])
                        
                    ### セル[20:5]: 浸水土砂被害の区分についてデータベースに登録されている値と突合せチェックする。
                    if ws_ippan[i].cell(row=j, column=5).value is None:
                        pass
                    else:
                        if ws_ippan[i].cell(row=j, column=5).value not in list(flood_sediment_code_list) and \
                            ws_ippan[i].cell(row=j, column=5).value not in list(flood_sediment_name_list) and \
                            ws_ippan[i].cell(row=j, column=5).value not in list(flood_sediment_name_code_list):
                            add_comment(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=j, column=5, fill=fill, com_id=453)
                            compare_NG_grid.append([ws_ippan[i].title, j, 5, 453])
                        else:
                            compare_OK_grid.append([ws_ippan[i].title, j, 5, 453])
                        
                    ### セル[20:6]: 被害建物棟数, 床下浸水についてデータベースに登録されている値と突合せチェックする。
                    ### セル[20:7]: 被害建物棟数, 1cm〜49cmについてデータベースに登録されている値と突合せチェックする。
                    ### セル[20:8]: 被害建物棟数, 50cm〜99cmについてデータベースに登録されている値と突合せチェックする。
                    ### セル[20:9]: 被害建物棟数, 1m以上についてデータベースに登録されている値と突合せチェックする。
                    ### セル[20:10]: 被害建物棟数, 半壊についてデータベースに登録されている値と突合せチェックする。
                    ### セル[20:11]: 被害建物棟数, 全壊・流失についてデータベースに登録されている値と突合せチェックする。
                    ### セル[20:12]: 被害建物の延床面積についてデータベースに登録されている値と突合せチェックする。
                    ### セル[20:13]: 被災世帯数についてデータベースに登録されている値と突合せチェックする。
                    ### セル[20:14]: 被災事業所数についてデータベースに登録されている値と突合せチェックする。
                    ### セル[20:15]: 農家・漁家戸数, 床下浸水についてデータベースに登録されている値と突合せチェックする。
                    ### セル[20:16]: 農家・漁家戸数, 1cm〜49cmについてデータベースに登録されている値と突合せチェックする。
                    ### セル[20:17]: 農家・漁家戸数, 50cm〜99cmについてデータベースに登録されている値と突合せチェックする。
                    ### セル[20:18]: 農家・漁家戸数, 1m以上・半壊についてデータベースに登録されている値と突合せチェックする。
                    ### セル[20:19]: 農家・漁家戸数, 全壊・流失についてデータベースに登録されている値と突合せチェックする。
                    ### セル[20:20]: 事業所従業者数, 床下浸水についてデータベースに登録されている値と突合せチェックする。
                    ### セル[20:21]: 事業所従業者数, 1cm〜49cmについてデータベースに登録されている値と突合せチェックする。
                    ### セル[20:22]: 事業所従業者数, 50cm〜99cmについてデータベースに登録されている値と突合せチェックする。
                    ### セル[20:23]: 事業所従業者数, 1m以上・半壊についてデータベースに登録されている値と突合せチェックする。
                    ### セル[20:24]: 事業所従業者数, 全壊・流失についてデータベースに登録されている値と突合せチェックする。
                        
                    ### セル[20:25]: 事業所の産業区分についてデータベースに登録されている値と突合せチェックする。
                    if ws_ippan[i].cell(row=j, column=25).value is None:
                        pass
                    else:
                        if ws_ippan[i].cell(row=j, column=25).value not in list(industry_code_list) and \
                            ws_ippan[i].cell(row=j, column=25).value not in list(industry_name_list) and \
                            ws_ippan[i].cell(row=j, column=25).value not in list(industry_name_code_list):
                            add_comment(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=j, column=25, fill=fill, com_id=473)
                            compare_NG_grid.append([ws_ippan[i].title, j, 25, 473])
                        else:
                            compare_OK_grid.append([ws_ippan[i].title, j, 25, 473])
                        
                    ### セル[20:26]: 地下空間の利用形態についてデータベースに登録されている値と突合せチェックする。
                    if ws_ippan[i].cell(row=j, column=26).value is None:
                        pass
                    else:
                        if ws_ippan[i].cell(row=j, column=26).value not in list(usage_code_list) and \
                            ws_ippan[i].cell(row=j, column=26).value not in list(usage_name_list) and \
                            ws_ippan[i].cell(row=j, column=26).value not in list(usage_name_code_list):
                            add_comment(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=j, column=26, fill=fill, com_id=474)
                            compare_NG_grid.append([ws_ippan[i].title, j, 26, 474])
                        else:
                            compare_OK_grid.append([ws_ippan[i].title, j, 26, 474])
                        
                    ### セル[20:27]: 備考についてデータベースに登録されている値と突合せチェックする。

        #######################################################################
        ### ファイル入出力処理(6000)
        ### チェック結果ファイルを保存する。
        #######################################################################
        print_log('[DEBUG] P0300IppanUpload.index_view()関数 STEP 30/35.', 'DEBUG')
        wb.save(output_file_path)

        #######################################################################
        ### ログ出力処理(7000)
        #######################################################################
        print_log('[DEBUG] P0300IppanUpload.index_view()関数 STEP 31/35.', 'DEBUG')
        if len(require_NG_list) > 0 or \
            len(require_NG_grid) > 0 or \
            len(format_NG_list) > 0 or \
            len(format_NG_grid) > 0 or \
            len(range_NG_list) > 0 or \
            len(range_NG_grid) > 0 or \
            len(correlate_NG_list) > 0 or \
            len(correlate_NG_grid) > 0 or \
            len(compare_NG_list) > 0 or \
            len(compare_NG_grid) > 0:
            print_log('[DEBUG] P0300IppanUpload.index_view()関数 False', 'DEBUG')
        else:
            print_log('[DEBUG] P0300IppanUpload.index_view()関数 True', 'DEBUG')
            
        print_log('[DEBUG] P0300IppanUpload.index_view()関数 max_row = {}'.format(max_row), 'DEBUG')
        print_log('[DEBUG] P0300IppanUpload.index_view()関数 len(require_NG_list) = {}'.format(len(require_NG_list)), 'DEBUG')
        print_log('[DEBUG] P0300IppanUpload.index_view()関数 len(format_NG_list) = {}'.format(len(format_NG_list)), 'DEBUG')
        print_log('[DEBUG] P0300IppanUpload.index_view()関数 len(range_NG_list) = {}'.format(len(range_NG_list)), 'DEBUG')
        print_log('[DEBUG] P0300IppanUpload.index_view()関数 len(correlate_NG_list) = {}'.format(len(correlate_NG_list)), 'DEBUG')
        print_log('[DEBUG] P0300IppanUpload.index_view()関数 len(compare_NG_list) = {}'.format(len(compare_NG_list)), 'DEBUG')
        print_log('[DEBUG] P0300IppanUpload.index_view()関数 len(require_OK_list) = {}'.format(len(require_OK_list)), 'DEBUG')
        print_log('[DEBUG] P0300IppanUpload.index_view()関数 len(format_OK_list) = {}'.format(len(format_OK_list)), 'DEBUG')
        print_log('[DEBUG] P0300IppanUpload.index_view()関数 len(range_OK_list) = {}'.format(len(range_OK_list)), 'DEBUG')
        print_log('[DEBUG] P0300IppanUpload.index_view()関数 len(correlate_OK_list) = {}'.format(len(correlate_OK_list)), 'DEBUG')
        print_log('[DEBUG] P0300IppanUpload.index_view()関数 len(compare_OK_list) = {}'.format(len(compare_OK_list)), 'DEBUG')
        
        success_str = ''
        if len(require_OK_list) > 0:
            for i in range(len(require_OK_list)):
                ### success_str = success_str+str(require_OK_list[i][0])+','+str(require_OK_list[i][1])+','+str(require_OK_list[i][2])+','+str(require_OK_list[i][3])+','+str(require_OK_list[i][4])+','+str(require_OK_list[i][5])+'\n'        
                success_str = success_str + \
                    str(require_OK_list[i][0]) + ',' + \
                    str(require_OK_list[i][1]) + ',' + \
                    str(require_OK_list[i][2]) + ',' + \
                    str(MESSAGE[require_OK_list[i][3]][0]) + ',' + \
                    str(MESSAGE[require_OK_list[i][3]][1]) + ',' + \
                    str(MESSAGE[require_OK_list[i][3]][2]) + '\n'        
        
        if len(format_OK_list) > 0:
            for i in range(len(format_OK_list)):
                success_str = success_str + \
                    str(format_OK_list[i][0]) + ',' + \
                    str(format_OK_list[i][1]) + ',' + \
                    str(format_OK_list[i][2]) + ',' + \
                    str(MESSAGE[format_OK_list[i][3]][0]) + ',' + \
                    str(MESSAGE[format_OK_list[i][3]][1]) + ',' + \
                    str(MESSAGE[format_OK_list[i][3]][2]) + '\n'        

        if len(range_OK_list) > 0:
            for i in range(len(range_OK_list)):
                success_str = success_str + \
                    str(range_OK_list[i][0]) + ',' + \
                    str(range_OK_list[i][1]) + ',' + \
                    str(range_OK_list[i][2]) + ',' + \
                    str(MESSAGE[range_OK_list[i][3]][0]) + ',' + \
                    str(MESSAGE[range_OK_list[i][3]][1]) + ',' + \
                    str(MESSAGE[range_OK_list[i][3]][2]) + '\n'        

        if len(correlate_OK_list) > 0:
            for i in range(len(correlate_OK_list)):
                success_str = success_str + \
                    str(correlate_OK_list[i][0]) + ',' + \
                    str(correlate_OK_list[i][1]) + ',' + \
                    str(correlate_OK_list[i][2]) + ',' + \
                    str(MESSAGE[correlate_OK_list[i][3]][0]) + ',' + \
                    str(MESSAGE[correlate_OK_list[i][3]][1]) + ',' + \
                    str(MESSAGE[correlate_OK_list[i][3]][2]) + '\n'        

        if len(compare_OK_list) > 0:
            for i in range(len(compare_OK_list)):
                success_str = success_str + \
                    str(compare_OK_list[i][0]) + ',' + \
                    str(compare_OK_list[i][1]) + ',' + \
                    str(compare_OK_list[i][2]) + ',' + \
                    str(MESSAGE[compare_OK_list[i][3]][0]) + ',' + \
                    str(MESSAGE[compare_OK_list[i][3]][1]) + ',' + \
                    str(MESSAGE[compare_OK_list[i][3]][2]) + '\n'        

        failure_str = ''
        if len(require_NG_list) > 0:
            for i in range(len(require_NG_list)):
                failure_str = failure_str + \
                    str(require_NG_list[i][0]) + ',' + \
                    str(require_NG_list[i][1]) + ',' + \
                    str(require_NG_list[i][2]) + ',' + \
                    str(MESSAGE[require_NG_list[i][3]][0]) + ',' + \
                    str(MESSAGE[require_NG_list[i][3]][1]) + ',' + \
                    str(MESSAGE[require_NG_list[i][3]][2]) + '\n'        
        
        if len(format_NG_list) > 0:
            for i in range(len(format_NG_list)):
                failure_str = failure_str + \
                    str(format_NG_list[i][0]) + ',' + \
                    str(format_NG_list[i][1]) + ',' + \
                    str(format_NG_list[i][2]) + ',' + \
                    str(MESSAGE[format_NG_list[i][3]][0]) + ',' + \
                    str(MESSAGE[format_NG_list[i][3]][1]) + ',' + \
                    str(MESSAGE[format_NG_list[i][3]][2]) + '\n'        

        if len(range_NG_list) > 0:
            for i in range(len(range_NG_list)):
                failure_str = failure_str + \
                    str(range_NG_list[i][0]) + ',' + \
                    str(range_NG_list[i][1]) + ',' + \
                    str(range_NG_list[i][2]) + ',' + \
                    str(MESSAGE[range_NG_list[i][3]][0]) + ',' + \
                    str(MESSAGE[range_NG_list[i][3]][1]) + ',' + \
                    str(MESSAGE[range_NG_list[i][3]][2]) + '\n'        

        if len(correlate_NG_list) > 0:
            for i in range(len(correlate_NG_list)):
                failure_str = failure_str + \
                    str(correlate_NG_list[i][0]) + ',' + \
                    str(correlate_NG_list[i][1]) + ',' + \
                    str(correlate_NG_list[i][2]) + ',' + \
                    str(MESSAGE[correlate_NG_list[i][3]][0]) + ',' + \
                    str(MESSAGE[correlate_NG_list[i][3]][1]) + ',' + \
                    str(MESSAGE[correlate_NG_list[i][3]][2]) + '\n'        

        if len(compare_NG_list) > 0:
            for i in range(len(compare_NG_list)):
                failure_str = failure_str + \
                    str(compare_NG_list[i][0]) + ',' + \
                    str(compare_NG_list[i][1]) + ',' + \
                    str(compare_NG_list[i][2]) + ',' + \
                    str(MESSAGE[compare_NG_list[i][3]][0]) + ',' + \
                    str(MESSAGE[compare_NG_list[i][3]][1]) + ',' + \
                    str(MESSAGE[compare_NG_list[i][3]][2]) + '\n'        
            
        #######################################################################
        ### レスポンスセット処理(8000)
        ### テンプレートとコンテキストを設定して、レスポンスをブラウザに戻す。
        ### ※入力チェックでエラーが発見された場合、
        ### ※ネストを浅くするために、処理対象外の場合、終了させる。
        #######################################################################
        print_log('[DEBUG] P0300IppanUpload.index_view()関数 STEP 32/35.', 'DEBUG')
        if len(require_NG_list) > 0 or \
            len(require_NG_grid) > 0 or \
            len(format_NG_list) > 0 or \
            len(format_NG_grid) > 0 or \
            len(range_NG_list) > 0 or \
            len(range_NG_grid) > 0 or \
            len(correlate_NG_list) > 0 or \
            len(correlate_NG_grid) > 0 or \
            len(compare_NG_list) > 0 or \
            len(compare_NG_grid) > 0:

            connection_cursor = connection.cursor()
            try:
                connection_cursor.execute("""BEGIN""", []);
                
                ###############################################################
                ### 市区町村をアップロード処理の単位とする。
                ### 当該市区町村の全データを入れ替える。
                ### 既存の市区町村のSUIGAI、IPPAN、IPPAN_SUMMARYのデータは、削除日時をセットして、削除済の扱いとする。
                ### ※入力チェックでエラーが発見された場合、
                ###############################################################
                print("handle_32_1", flush=True)
                connection_cursor.execute("""
                    UPDATE SUIGAI SET 
                        deleted_at=CURRENT_TIMESTAMP 
                    WHERE suigai_id IN (SELECT suigai_id FROM SUIGAI WHERE city_code=%s AND deleted_at IS NULL)
                    """, [
                        split_name_code(ws_ippan[0].cell(row=7, column=3).value)[-1], 
                    ])
        
                print("handle_32_2", flush=True)
                connection_cursor.execute("""
                    UPDATE IPPAN SET 
                        deleted_at=CURRENT_TIMESTAMP 
                    WHERE ippan_id IN (SELECT ippan_id FROM IPPAN_VIEW WHERE city_code=%s AND deleted_at IS NULL)
                    """, [
                        split_name_code(ws_ippan[0].cell(row=7, column=3).value)[-1], 
                    ])
    
                print("handle_32_3", flush=True)
                connection_cursor.execute("""
                    UPDATE IPPAN_SUMMARY SET 
                        deleted_at=CURRENT_TIMESTAMP 
                    WHERE ippan_id IN (
                    SELECT 
                        IS1.ippan_id AS ippan_id 
                    FROM IPPAN_SUMMARY IS1 
                    LEFT JOIN SUIGAI SG1 ON IS1.suigai_id=SG1.suigai_id 
                    WHERE 
                        SG1.city_code=%s AND 
                        IS1.deleted_at IS NULL)
                    """, [
                        split_name_code(ws_ippan[0].cell(row=7, column=3).value)[-1], 
                    ])
    
                print("handle_32_4", flush=True)
                connection_cursor.execute("""
                    UPDATE TRIGGER SET 
                        deleted_at=CURRENT_TIMESTAMP 
                    WHERE trigger_id IN (SELECT trigger_id FROM TRIGGER WHERE city_code=%s AND deleted_at IS NULL AND action_code IN ('A01','A02','A03','A04','A05','A06','A07','A99'))
                    """, [
                        split_name_code(ws_ippan[0].cell(row=7, column=3).value)[-1], 
                    ])
                
                ###############################################################
                ### トリガーテーブルにWF2アップロードトリガーを実行済、成功として登録する。
                ### ※入力チェックでエラーが発見された場合、
                ###############################################################
                print("handle_32_5", flush=True)
                connection_cursor.execute("""
                    INSERT INTO TRIGGER (
                        trigger_id, suigai_id, action_code, status_code, success_count, failure_count, 
                        published_at, consumed_at, deleted_at, integrity_ok, integrity_ng, ken_code, 
                        city_code, download_file_path, download_file_name, upload_file_path, upload_file_name 
                    ) VALUES (
                        (SELECT CASE WHEN (MAX(trigger_id+1)) IS NULL THEN CAST(0 AS INTEGER) ELSE CAST(MAX(trigger_id+1) AS INTEGER) END AS trigger_id FROM TRIGGER), -- trigger_id 
                        %s, -- suigai_id 
                        %s, -- action_code 
                        %s, -- status_code 
                        %s, -- success_count 
                        %s, -- failure_count 
                        CURRENT_TIMESTAMP,  -- published_at 
                        CURRENT_TIMESTAMP,  -- consumed_at 
                        %s, -- deleted_at 
                        %s, -- integrity_ok 
                        %s, -- integrity_ng 
                        %s, -- ken_code 
                        %s, -- city_code 
                        %s, -- download_file_path 
                        %s, -- download_file_name 
                        %s, -- upload_file_path 
                        %s  -- upload_file_name 
                    )""", [
                        None, ### suigai_id 
                        'A01', ### action_code 
                        'SUCCESS', ### status_code 
                        1,  ### success_count
                        0,  ### failure_count
                        None, ### deleted_at 
                        '\n'.join(get_info_log()), ### integrity_ok 
                        '\n'.join(get_warn_log()), ### integrity_ng 
                        convert_empty_to_none(split_name_code(ws_ippan[0].cell(row=7, column=2).value)[-1]), ### ken_code 
                        convert_empty_to_none(split_name_code(ws_ippan[0].cell(row=7, column=3).value)[-1]), ### city_code 
                        None, ### download_file_path 
                        None, ### download_file_name 
                        upload_file_path, ### upload_file_path 
                        upload_file_name, ### upload_file_name 
                    ])
                    
                ###############################################################
                ### トリガーテーブルにWF3データ検証トリガーを実行済、失敗として登録する。
                ### ※入力チェックでエラーが発見された場合、
                ###############################################################
                print("handle_32_6", flush=True)
                connection_cursor.execute("""
                    INSERT INTO TRIGGER (
                        trigger_id, suigai_id, action_code, status_code, success_count, failure_count, 
                        published_at, consumed_at, deleted_at, integrity_ok, integrity_ng, ken_code, 
                        city_code, download_file_path, download_file_name, upload_file_path, upload_file_name 
                    ) VALUES (
                        (SELECT CASE WHEN (MAX(trigger_id+1)) IS NULL THEN CAST(0 AS INTEGER) ELSE CAST(MAX(trigger_id+1) AS INTEGER) END AS trigger_id FROM TRIGGER), -- trigger_id 
                        %s, -- suigai_id 
                        %s, -- action_code 
                        %s, -- status_code 
                        %s, -- success_count 
                        %s, -- failure_count 
                        CURRENT_TIMESTAMP,  -- published_at 
                        CURRENT_TIMESTAMP,  -- consumed_at 
                        %s, -- deleted_at 
                        %s, -- integrity_ok 
                        %s, -- integrity_ng 
                        %s, -- ken_code 
                        %s, -- city_code 
                        %s, -- download_file_path 
                        %s, -- download_file_name 
                        %s, -- upload_file_path 
                        %s  -- upload_file_name 
                    )""", [
                        None, ### suigai_id 
                        'A02', ### action_code 
                        'FAILURE', ### status_code 
                        0,  ### success_count
                        len(ws_ippan),  ### failure_count
                        None, ### deleted_at 
                        success_str, ### integrity_ok 
                        failure_str, ### integrity_ng 
                        convert_empty_to_none(split_name_code(ws_ippan[0].cell(row=7, column=2).value)[-1]), ### ken_code 
                        convert_empty_to_none(split_name_code(ws_ippan[0].cell(row=7, column=3).value)[-1]), ### city_code 
                        None, ### download_file_path 
                        None, ### download_file_name 
                        upload_file_path, ### upload_file_path 
                        upload_file_name, ### upload_file_name 
                    ])
                connection_cursor.execute("""COMMIT""", []);
            except:
                print_log('[ERROR] P0300IppanUpload.index_view()関数 {}'.format(sys.exc_info()[0]), 'ERROR')
                connection_cursor.execute("""ROLLBACK""", [])
            finally:
                connection_cursor.close()
                
            ### src/P0300IppanUpload/templates/P0300IppanUpload/fail.htmlを使用する。
            template = loader.get_template('P0300IppanUpload/fail.html')
            context = {
                'require_NG_list': require_NG_list,
                'require_NG_grid': require_NG_grid,
                'format_NG_list': format_NG_list,
                'format_NG_grid': format_NG_grid,
                'range_NG_list': range_NG_list,
                'range_NG_grid': range_NG_grid,
                'correlate_NG_list': correlate_NG_list,
                'correlate_NG_grid': correlate_NG_grid,
                'compare_NG_list': compare_NG_list,
                'compare_NG_grid': compare_NG_grid,
                'output_file_path': output_file_path,
            }
            print_log('[INFO] P0300IppanUpload.index_view()関数が正常終了しました。', 'INFO')
            return HttpResponse(template.render(context, request))
        
        #######################################################################
        ### DBアクセス処理(9000)
        ### (1)入力データ_ヘッダ部分テーブルにデータを登録する。
        ### (2)入力データ_一覧票部分テーブルにデータを登録する。
        ### (3)レポジトリテーブルにデータを登録する。
        ### ※入力チェックでエラーが発見されなかった場合、
        #######################################################################
        print_log('[DEBUG] P0300IppanUpload.index_view()関数 STEP 33/35.', 'DEBUG')
        connection_cursor = connection.cursor()
        try:
            connection_cursor.execute("""BEGIN""", [])
            
            ###################################################################
            ### 市区町村をアップロード処理の単位とする。
            ### 当該市区町村の全データを入れ替える。
            ### 既存の市区町村のSUIGAI、IPPAN、IPPAN_SUMMARYのデータは、削除日時をセットして、削除済の扱いとする。
            ### ※入力チェックでエラーが発見されなかった場合、
            ###################################################################
            print("handle_33_1", flush=True)
            connection_cursor.execute("""
                UPDATE SUIGAI SET 
                    deleted_at=CURRENT_TIMESTAMP 
                WHERE suigai_id IN (SELECT suigai_id FROM SUIGAI WHERE city_code=%s AND deleted_at IS NULL)
                """, [
                    split_name_code(ws_ippan[0].cell(row=7, column=3).value)[-1], 
                ])
    
            print("handle_33_2", flush=True)
            connection_cursor.execute("""
                UPDATE IPPAN SET 
                    deleted_at=CURRENT_TIMESTAMP 
                WHERE ippan_id IN (SELECT ippan_id FROM IPPAN_VIEW WHERE city_code=%s AND deleted_at IS NULL)
                """, [
                    split_name_code(ws_ippan[0].cell(row=7, column=3).value)[-1], 
                ])

            print("handle_33_3", flush=True)
            connection_cursor.execute("""
                UPDATE IPPAN_SUMMARY SET 
                    deleted_at=CURRENT_TIMESTAMP 
                WHERE ippan_id IN (
                SELECT 
                    IS1.ippan_id AS ippan_id 
                FROM IPPAN_SUMMARY IS1 
                LEFT JOIN SUIGAI SG1 ON IS1.suigai_id=SG1.suigai_id 
                WHERE 
                    SG1.city_code=%s AND 
                    IS1.deleted_at IS NULL)
                """, [
                    split_name_code(ws_ippan[0].cell(row=7, column=3).value)[-1], 
                ])

            print("handle_33_4", flush=True)
            connection_cursor.execute("""
                UPDATE TRIGGER SET 
                    deleted_at=CURRENT_TIMESTAMP 
                WHERE trigger_id IN (SELECT trigger_id FROM TRIGGER WHERE city_code=%s AND deleted_at IS NULL AND action_code IN ('A01','A02','A03','A04','A05','A06','A07','A99'))
                """, [
                    split_name_code(ws_ippan[0].cell(row=7, column=3).value)[-1], 
                ])

            ###################################################################
            ### ※入力チェックでエラーが発見されなかった場合、
            ###################################################################
            for i, _ in enumerate(ws_ippan):
                print_log('[DEBUG] P0300IppanUpload.index_view()関数 STEP 34_1/35.', 'DEBUG')
                ### suigai_id__max で正しい。
                suigai_id = SUIGAI.objects.all().aggregate(Max('suigai_id'))['suigai_id__max']
                if suigai_id is None:
                    suigai_id = 1
                else:
                    suigai_id = suigai_id + 1
                    
                print_log('[DEBUG] P0300IppanUpload.index_view()関数 suigai_id = {}'.format(suigai_id), 'DEBUG')
                
                ###############################################################
                ### DBアクセス処理(9030)
                ### 一般資産入力データ_ヘッダ部分テーブルにデータを登録する。
                ### ※入力チェックでエラーが発見されなかった場合、
                ###############################################################
                print_log('[DEBUG] P0300IppanUpload.index_view()関数 STEP 34_2/35.', 'DEBUG')
                connection_cursor.execute("""
                    INSERT INTO SUIGAI (
                        suigai_id, suigai_name, ken_code, city_code, begin_date, end_date, 
                        cause_1_code, cause_2_code, cause_3_code, area_id, suikei_code, 
                        kasen_code, gradient_code, residential_area, agricultural_area, underground_area, 
                        kasen_kaigan_code, crop_damage, weather_id, committed_at, deleted_at, upload_file_path, 
                        upload_file_name, summary_file_path, summary_file_name, action_code, status_code 
                    ) VALUES (
                        %s, -- suigai_id 
                        %s, -- suigai_name 
                        %s, -- ken_code 
                        %s, -- city_code 
                        TO_DATE(%s, 'yyyy/mm/dd'), -- begin_date 
                        TO_DATE(%s, 'yyyy/mm/dd'), -- end_date 
                        %s, -- cause_1_code 
                        %s, -- cause_2_code 
                        %s, -- cause_3_code 
                        %s, -- area_id 
                        %s, -- suikei_code 
                        %s, -- kasen_code 
                        %s, -- gradient_code 
                        %s, -- residential_area 
                        %s, -- agricultural_area 
                        %s, -- underground_area 
                        %s, -- kasen_kaigan_code 
                        %s, -- crop_damage 
                        %s, -- weather_id 
                        CURRENT_TIMESTAMP, -- committed_at 
                        %s, -- deleted_at 
                        %s, -- upload_file_path 
                        %s, -- upload_file_name 
                        %s, -- summary_file_path 
                        %s, -- summary_file_name 
                        %s, -- action_code 
                        %s  -- status_code 
                    )""", [
                        suigai_id, ### suigai_id 
                        ws_title[i], ### suigai_name 
                        convert_empty_to_none(split_name_code(ws_ippan[i].cell(row=7, column=2).value)[-1]), ### ken_code 
                        convert_empty_to_none(split_name_code(ws_ippan[i].cell(row=7, column=3).value)[-1]), ### city_code 
                        convert_empty_to_none(ws_ippan[i].cell(row=7, column=4).value), ### begin_date 
                        convert_empty_to_none(ws_ippan[i].cell(row=7, column=5).value), ### end_date 
                        convert_empty_to_none(split_name_code(ws_ippan[i].cell(row=7, column=6).value)[-1]), ### cause_1_code 
                        convert_empty_to_none(split_name_code(ws_ippan[i].cell(row=7, column=7).value)[-1]), ### cause_2_code 
                        convert_empty_to_none(split_name_code(ws_ippan[i].cell(row=7, column=8).value)[-1]), ### cause_3_code 
                        convert_empty_to_none(split_name_code(ws_ippan[i].cell(row=7, column=9).value)[-1]), ### area_id 
                        convert_empty_to_none(split_name_code(ws_ippan[i].cell(row=10, column=2).value)[-1]), ### suikei_code 
                        convert_empty_to_none(split_name_code(ws_ippan[i].cell(row=10, column=4).value)[-1]), ### kasen_code 
                        convert_empty_to_none(split_name_code(ws_ippan[i].cell(row=10, column=6).value)[-1]), ### gradient_code 
                        convert_empty_to_none(ws_ippan[i].cell(row=14, column=2).value), ### residential_area 
                        convert_empty_to_none(ws_ippan[i].cell(row=14, column=3).value), ### agricultural_area 
                        convert_empty_to_none(ws_ippan[i].cell(row=14, column=4).value), ### underground_area 
                        convert_empty_to_none(split_name_code(ws_ippan[i].cell(row=14, column=6).value)[-1]), ### kasen_kaigan_code 
                        convert_empty_to_none(ws_ippan[i].cell(row=14, column=8).value), ### crop_damaga 
                        convert_empty_to_none(split_name_code(ws_ippan[i].cell(row=14, column=10).value)[-1]), ### weather_id 
                        None, ### deleted_at 
                        upload_file_path, ### upload_file_path 
                        upload_file_name, ### upload_file_name 
                        None, ### summary_file_path 
                        None, ### summary_file_name 
                        None,  ### action_code 
                        None,  ### status_code 
                    ])
                    
                ###############################################################
                ### DBアクセス処理(9040)
                ### 一般資産入力データ_一覧表部分テーブルにデータを登録する。
                ### ※入力チェックでエラーが発見されなかった場合、
                ###############################################################
                print_log('[DEBUG] P0300IppanUpload.index_view()関数 STEP 34_3/35.', 'DEBUG')
                print_log('[DEBUG] P0300IppanUpload.index_view()関数 max_row[i] = {}'.format(max_row[i]), 'DEBUG')
                if max_row[i] >= 20:
                    for j in range(20, max_row[i] + 1):
                        print_log('[DEBUG] P0300IppanUpload.index_view()関数 j = {}'.format(j), 'DEBUG')
                        connection_cursor.execute(""" 
                            INSERT INTO IPPAN (
                                ippan_id, ippan_name, suigai_id, building_code, underground_code, flood_sediment_code, 
                                building_lv00, building_lv01_49, building_lv50_99, building_lv100, building_half, building_full, 
                                floor_area, family, office, 
                                farmer_fisher_lv00, farmer_fisher_lv01_49, farmer_fisher_lv50_99, farmer_fisher_lv100, farmer_fisher_full, 
                                employee_lv00, employee_lv01_49, employee_lv50_99, employee_lv100, employee_full, 
                                industry_code, usage_code, comment, deleted_at 
                            ) VALUES (
                                (SELECT CASE WHEN (MAX(ippan_id+1)) IS NULL THEN CAST(0 AS INTEGER) ELSE CAST(MAX(ippan_id+1) AS INTEGER) END AS ippan_id FROM IPPAN), -- ippan_id 
                                %s, -- ippan_name 
                                %s, -- suigai_id 
                                %s, -- building_code 
                                %s, -- underground_code 
                                %s, -- flood_sediment_code 
                                %s, -- building_lv00 
                                %s, -- building_lv01_49 
                                %s, -- building_lv50_99 
                                %s, -- building_lv100 
                                %s, -- building_half 
                                %s, -- building_full 
                                %s, -- floor_area 
                                %s, -- family 
                                %s, -- office 
                                %s, -- farmer_fisher_lv00 
                                %s, -- farmer_fisher_lv01_49 
                                %s, -- farmer_fisher_lv50_99 
                                %s, -- farmer_fisher_lv100 
                                %s, -- farmer_fisher_full 
                                %s, -- employee_lv00 
                                %s, -- employee_lv01_49 
                                %s, -- employee_lv50_99 
                                %s, -- employee_lv100 
                                %s, -- employee_full 
                                %s, -- industry_code 
                                %s, -- usage_code 
                                %s, -- comment 
                                %s  -- deleted_at 
                            ) """, [
                                convert_empty_to_none(ws_ippan[i].cell(row=j, column=2).value), ### ippan_name
                                suigai_id, ### suigai_id
                                convert_empty_to_none(split_name_code(ws_ippan[i].cell(row=j, column=3).value)[-1]), ### building_code
                                convert_empty_to_none(split_name_code(ws_ippan[i].cell(row=j, column=4).value)[-1]), ### underground_code
                                convert_empty_to_none(split_name_code(ws_ippan[i].cell(row=j, column=5).value)[-1]), ### flood_sediment_code
                                convert_empty_to_none(ws_ippan[i].cell(row=j, column=6).value), ### building_lv00
                                convert_empty_to_none(ws_ippan[i].cell(row=j, column=7).value), ### building_lv01_49
                                convert_empty_to_none(ws_ippan[i].cell(row=j, column=8).value), ### building_lv50_99
                                convert_empty_to_none(ws_ippan[i].cell(row=j, column=9).value), ### building_lv100
                                convert_empty_to_none(ws_ippan[i].cell(row=j, column=10).value), ### building_half
                                convert_empty_to_none(ws_ippan[i].cell(row=j, column=11).value), ### building_full
                                convert_empty_to_none(ws_ippan[i].cell(row=j, column=12).value), ### floor_area
                                convert_empty_to_none(ws_ippan[i].cell(row=j, column=13).value), ### family
                                convert_empty_to_none(ws_ippan[i].cell(row=j, column=14).value), ### office
                                convert_empty_to_none(ws_ippan[i].cell(row=j, column=15).value), ### farmer_fisher_lv00
                                convert_empty_to_none(ws_ippan[i].cell(row=j, column=16).value), ### farmer_fisher_lv01_49
                                convert_empty_to_none(ws_ippan[i].cell(row=j, column=17).value), ### farmer_fisher_lv50_99
                                convert_empty_to_none(ws_ippan[i].cell(row=j, column=18).value), ### farmer_fisher_lv100
                                convert_empty_to_none(ws_ippan[i].cell(row=j, column=19).value), ### farmer_fisher_full
                                convert_empty_to_none(ws_ippan[i].cell(row=j, column=20).value), ### employee_lv00
                                convert_empty_to_none(ws_ippan[i].cell(row=j, column=21).value), ### employee_lv01_49
                                convert_empty_to_none(ws_ippan[i].cell(row=j, column=22).value), ### employee_lv50_99
                                convert_empty_to_none(ws_ippan[i].cell(row=j, column=23).value), ### employee_lv100
                                convert_empty_to_none(ws_ippan[i].cell(row=j, column=24).value), ### employee_full
                                convert_empty_to_none(split_name_code(ws_ippan[i].cell(row=j, column=25).value)[-1]), ### industry_code
                                convert_empty_to_none(split_name_code(ws_ippan[i].cell(row=j, column=26).value)[-1]), ### usage_code
                                convert_empty_to_none(ws_ippan[i].cell(row=j, column=27).value), ### comment
                                None, ### deleted_at 
                            ])

                ###############################################################
                ### DBアクセス処理(9060)
                ### (1)トリガーテーブルにWF2アップロードトリガーを実行済、成功として登録する。
                ### (2)トリガーテーブルにWF3データ検証トリガーを実行済、成功として登録する。
                ### (3)トリガーテーブルにWF4差分検証トリガーを未実行＝次回実行対象として登録する。
                ### ※入力チェックでエラーが発見されなかった場合、
                ###############################################################
                print_log('[DEBUG] P0300IppanUpload.index_view()関数 STEP 34_5/35.', 'DEBUG')
                ### トリガーテーブルにA02アップロードトリガーを実行済、成功として登録する。
                connection_cursor.execute("""
                    INSERT INTO TRIGGER (
                        trigger_id, suigai_id, action_code, status_code, success_count, failure_count, 
                        published_at, consumed_at, deleted_at, integrity_ok, integrity_ng, ken_code, 
                        city_code, download_file_path, download_file_name, upload_file_path, upload_file_name 
                    ) VALUES (
                        (SELECT CASE WHEN (MAX(trigger_id+1)) IS NULL THEN CAST(0 AS INTEGER) ELSE CAST(MAX(trigger_id+1) AS INTEGER) END AS trigger_id FROM TRIGGER), -- trigger_id 
                        %s, -- suigai_id 
                        %s, -- action_code 
                        %s, -- status_code 
                        %s, -- success_count 
                        %s, -- failure_count 
                        CURRENT_TIMESTAMP, -- published_at 
                        CURRENT_TIMESTAMP, -- consumed_at 
                        %s, -- deleted_at 
                        %s, -- integrity_ok 
                        %s, -- integrity_ng 
                        %s, -- ken_code 
                        %s, -- city_code 
                        %s, -- download_file_path 
                        %s, -- download_file_path 
                        %s, -- upload_file_path 
                        %s  -- upload_file_name 
                    )""", [
                        suigai_id, ### suigai_id 
                        'A01', ### action_code 
                        'SUCCESS', ### status_code 
                        1, ### success_count
                        0, ### failure_count
                        None, ### deleted_at 
                        '\n'.join(get_info_log()), ### integrity_ok 
                        '\n'.join(get_warn_log()), ### integrity_ng 
                        convert_empty_to_none(split_name_code(ws_ippan[0].cell(row=7, column=2).value)[-1]), ### ken_code 
                        convert_empty_to_none(split_name_code(ws_ippan[0].cell(row=7, column=3).value)[-1]), ### city_code 
                        None, ### download_file_path 
                        None, ### download_file_name 
                        upload_file_path, ### upload_file_path 
                        upload_file_name, ### upload_file_name 
                    ])

                ### トリガーテーブルにA03データ検証トリガーを実行済、成功として登録する。
                connection_cursor.execute("""
                    INSERT INTO TRIGGER (
                        trigger_id, suigai_id, action_code, status_code, success_count, failure_count, 
                        published_at, consumed_at, deleted_at, integrity_ok, integrity_ng, ken_code, 
                        city_code, download_file_path, download_file_name, upload_file_path, upload_file_name 
                    ) VALUES (
                        (SELECT CASE WHEN (MAX(trigger_id+1)) IS NULL THEN CAST(0 AS INTEGER) ELSE CAST(MAX(trigger_id+1) AS INTEGER) END AS trigger_id FROM TRIGGER), -- trigger_id 
                        %s, -- suigai_id 
                        %s, -- action_code 
                        %s, -- status_code 
                        %s, -- success_count 
                        %s, -- failure_count 
                        CURRENT_TIMESTAMP, -- published_at 
                        CURRENT_TIMESTAMP, -- consumed_at 
                        %s, -- deleted_at 
                        %s, -- integrity_ok 
                        %s, -- integrity_ng 
                        %s, -- ken_code 
                        %s, -- city_code 
                        %s, -- download_file_path 
                        %s, -- download_file_name 
                        %s, -- upload_file_path 
                        %s  -- upload_file_name 
                    )""", [
                        suigai_id, ### suigai_id 
                        'A02', ### action_code 
                        'SUCCESS', ### status_code 
                        1, ### success_count
                        0, ### failure_count
                        None, ### deleted_at 
                        success_str, ### integrity_ok 
                        failure_str, ### integrity_ng 
                        convert_empty_to_none(split_name_code(ws_ippan[i].cell(row=7, column=2).value)[-1]), ### ken_code 
                        convert_empty_to_none(split_name_code(ws_ippan[i].cell(row=7, column=3).value)[-1]), ### city_code 
                        None, ### download_file_path 
                        None, ### download_file_name 
                        upload_file_path, ### upload_file_path 
                        upload_file_name, ### upload_file_name 
                    ])
            
                ### トリガーテーブルにA04差分検証トリガーを未実行＝次回実行対象として登録する。
                connection_cursor.execute("""
                    INSERT INTO TRIGGER (
                        trigger_id, suigai_id, action_code, status_code, success_count, failure_count, 
                        published_at, consumed_at, deleted_at, integrity_ok, integrity_ng, ken_code, 
                        city_code, download_file_path, download_file_name, upload_file_path, upload_file_name 
                    ) VALUES (
                        (SELECT CASE WHEN (MAX(trigger_id+1)) IS NULL THEN CAST(0 AS INTEGER) ELSE CAST(MAX(trigger_id+1) AS INTEGER) END AS trigger_id FROM TRIGGER), -- trigger_id 
                        %s, -- suigai_id 
                        %s, -- action_code 
                        %s, -- status_code 
                        %s, -- success_count 
                        %s, -- failure_count 
                        CURRENT_TIMESTAMP, -- published_at 
                        %s, -- consumed_at 
                        %s, -- deleted_at 
                        %s, -- integrity_ok 
                        %s, -- integrity_ng 
                        %s, -- ken_code 
                        %s, -- city_code 
                        %s, -- download_file_path 
                        %s, -- download_file_name 
                        %s, -- upload_file_path 
                        %s  -- upload_file_name 
                    )""", [
                        suigai_id, ### suigai_id 
                        'A03', ### action_code 
                        'RUNNING', ### status_code 
                        None, ### success_count
                        None, ### failure_count
                        None, ### consumed_at 
                        None, ### deleted_at 
                        None, ### integrity_ok 
                        None, ### integrity_ng 
                        convert_empty_to_none(split_name_code(ws_ippan[i].cell(row=7, column=2).value)[-1]), ### ken_code 
                        convert_empty_to_none(split_name_code(ws_ippan[i].cell(row=7, column=3).value)[-1]), ### city_code 
                        None, ### download_file_path 
                        None, ### download_file_name 
                        upload_file_path, ### upload_file_path 
                        upload_file_name, ### upload_file_name 
                    ])
            connection_cursor.execute("""COMMIT""", [])        
        except:
            print_log('[ERROR] P0300IppanUpload.index_view()関数 {}'.format(sys.exc_info()[0]), 'ERROR')
            connection_cursor.execute("""ROLLBACK""", [])
        finally:
            connection_cursor.close()
        
        #######################################################################
        ### レスポンスセット処理(10000)
        ### テンプレートとコンテキストを設定して、レスポンスをブラウザに戻す。
        ### ※入力チェックでエラーが発見されなかった場合、
        #######################################################################
        print_log('[DEBUG] P0300IppanUpload.index_view()関数 STEP 35/35.', 'DEBUG')
        ### src/P0300IppanUpload/templates/P0300IppanUpload/success.htmlを使用する。
        template = loader.get_template('P0300IppanUpload/success.html')
        context = {}
        print_log('[INFO] P0300IppanUpload.index_view()関数が正常終了しました。', 'INFO')
        return HttpResponse(template.render(context, request))
        
    except:
        print_log('[ERROR] P0300IppanUpload.index_view()関数 {}'.format(sys.exc_info()[0]), 'ERROR')
        print_log('[ERROR] P0300IppanUpload.index_view()関数でエラーが発生しました。', 'ERROR')
        print_log('[ERROR] P0300IppanUpload.index_viwe()関数が異常終了しました。', 'ERROR')
        return render(request, 'error.html')
