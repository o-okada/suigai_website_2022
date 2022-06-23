#!/usr/bin/env python3
# -*- coding: utf-8 -*-
###############################################################################
### ファイル名：P0300ExcelUpload/views.py
###############################################################################

###############################################################################
### 処理名：インポート処理
###############################################################################
import sys
from datetime import date, datetime, timedelta, timezone
from django.contrib.auth.decorators import login_required
from django.db import connection
from django.db import transaction
from django.db.models import Max
from django.http import Http404
from django.http import HttpResponse
from django.http import HttpResponseRedirect
from django.shortcuts import render
from django.template import loader
from django.views import generic
import openpyxl
from openpyxl.comments import Comment
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import PatternFill
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.writer.excel import save_virtual_workbook

from .forms import ExcelUploadForm

from P0000Common.models import BUILDING                ### 1000: 建物区分
from P0000Common.models import KEN                     ### 1010: 都道府県
from P0000Common.models import CITY                    ### 1020: 市区町村
from P0000Common.models import KASEN_KAIGAN            ### 1030: 水害発生地点工種（河川海岸区分）
from P0000Common.models import SUIKEI                  ### 1040: 水系（水系・沿岸）
from P0000Common.models import SUIKEI_TYPE             ### 1050: 水系種別（水系・沿岸種別）
from P0000Common.models import KASEN                   ### 1060: 河川（河川・海岸）
from P0000Common.models import KASEN_TYPE              ### 1070: 河川種別（河川・海岸種別）
from P0000Common.models import CAUSE                   ### 1080: 水害原因
from P0000Common.models import UNDERGROUND             ### 1090: 地上地下区分
from P0000Common.models import USAGE                   ### 1100: 地下空間の利用形態
from P0000Common.models import FLOOD_SEDIMENT          ### 1110: 浸水土砂区分
from P0000Common.models import GRADIENT                ### 1120: 地盤勾配区分
from P0000Common.models import INDUSTRY                ### 1130: 産業分類

from P0000Common.models import HOUSE_ASSET             ### 2000: 家屋評価額
from P0000Common.models import HOUSE_ALT               ### 2020: 家庭応急対策費_代替活動費
from P0000Common.models import HOUSE_CLEAN             ### 2030: 家庭応急対策費_清掃日数

from P0000Common.models import HOUSEHOLD_ASSET         ### 3000: 家庭用品自動車以外所有額
from P0000Common.models import HOUSEHOLD_RATE          ### 3010: 家庭用品自動車以外被害率

from P0000Common.models import CAR_ASSET               ### 4000: 家庭用品自動車所有額
from P0000Common.models import CAR_RATE                ### 4010: 家庭用品自動車被害率

from P0000Common.models import OFFICE_ASSET            ### 5000: 事業所資産額
from P0000Common.models import OFFICE_RATE             ### 5010: 事業所被害率
from P0000Common.models import OFFICE_SUSPEND          ### 5020: 事業所営業停止日数
from P0000Common.models import OFFICE_STAGNATE         ### 5030: 事業所営業停滞日数
from P0000Common.models import OFFICE_ALT              ### 5040: 事業所応急対策費_代替活動費

from P0000Common.models import FARMER_FISHER_ASSET     ### 6000: 農漁家資産額
from P0000Common.models import FARMER_FISHER_RATE      ### 6010: 農漁家被害率

from P0000Common.models import AREA                    ### 7000: 一般資産入力データ_水害区域
from P0000Common.models import WEATHER                 ### 7010: 一般資産入力データ_異常気象
from P0000Common.models import SUIGAI                  ### 7020: 一般資産入力データ_ヘッダ部分
from P0000Common.models import IPPAN                   ### 7030: 一般資産入力データ_一覧表部分
from P0000Common.models import IPPAN_VIEW              ### 7040: 一般資産ビューデータ_一覧表部分

from P0000Common.models import IPPAN_SUMMARY           ### 8000: 一般資産集計データ

from P0000Common.models import ACTION                  ### 10000: アクション
from P0000Common.models import STATUS                  ### 10010: 状態
from P0000Common.models import TRIGGER                 ### 10020: トリガーメッセージ
from P0000Common.models import APPROVAL                ### 10030: 承認メッセージ
from P0000Common.models import FEEDBACK                ### 10040: フィードバックメッセージ
from P0000Common.models import REPOSITORY              ### 10050: EXCELファイルレポジトリ
### from P0000Common.models import EXECUTE             ### 10060: 実行管理

from P0000Common.common import print_log

###############################################################################
### 処理名：定数定義
### 単体入力の必須チェックを行った結果のメッセージを定義する。
###############################################################################
MESSAGE = []

### 単体入力の必須をチェックする。
MESSAGE.append([0, 'W0000', '必須', '都道府県が入力されていません。', '都道府県を入力してください。'])
MESSAGE.append([1, 'W0001', '必須', '市区町村が入力されていません。', '市区町村を入力してください。'])
MESSAGE.append([2, 'W0002', '必須', '水害発生月日が入力されていません。', '水害発生月日を入力してください。'])
MESSAGE.append([3, 'W0003', '必須', '水害終了月日が入力されていません。', '水害終了月日を入力してください。'])
MESSAGE.append([4, 'W0004', '必須', '水害原因1が入力されていません。', '水害原因1を入力してください。'])
MESSAGE.append([5, 'W0005', '必須', '水害原因2が入力されていません。', '水害原因2を入力してください。'])
MESSAGE.append([6, 'W0006', '必須', '水害原因3が入力されていません。', '水害原因3を入力してください。'])
MESSAGE.append([7, 'W0007', '必須', '水害区域番号が入力されていません。', '水害区域番号を入力してください。'])
MESSAGE.append([8, 'W0008', '必須', '水系・沿岸名が入力されていません。', '水系・沿岸名を入力してください。'])
MESSAGE.append([9, 'W0009', '必須', '水系種別が入力されていません。', '水系種別を入力してください。'])
MESSAGE.append([10, 'W0010', '必須', '河川・海岸名が入力されていません。', '河川・海岸名を入力してください。'])
MESSAGE.append([11, 'W0011', '必須', '河川種別が入力されていません。', '河川種別を入力してください。'])
MESSAGE.append([12, 'W0012', '必須', '地盤勾配区分が入力されていません。', '地盤勾配区分を入力してください。'])
MESSAGE.append([13, 'W0013', '必須', '水害区域面積の宅地が入力されていません。', '水害区域面積の宅地を入力してください。'])
MESSAGE.append([14, 'W0014', '必須', '水害区域面積の農地が入力されていません。', '水害区域面積の農地を入力してください。'])
MESSAGE.append([15, 'W0015', '必須', '水害区域面積の地下が入力されていません。', '水害区域面積の地下を入力してください。'])
MESSAGE.append([16, 'W0016', '必須', '工種が入力されていません。', '工種を入力してください。'])
MESSAGE.append([17, 'W0017', '必須', '農作物被害額が入力されていません。', '農作物被害額を入力してください。'])
MESSAGE.append([18, 'W0018', '必須', '異常気象コードが入力されていません。', '異常気象コードを入力してください。'])
for i in range(19, 50):
    MESSAGE.append([i, '', '', '', ''])

MESSAGE.append([50, 'W0050', '必須', '町丁名・大字名が入力されていません。', '町丁名・大字名を入力してください。'])
MESSAGE.append([51, 'W0051', '必須', '名称が入力されていません。', '名称を入力してください。'])
MESSAGE.append([52, 'W0052', '必須', '地上・地下被害の区分が入力されていません。', '地上・地下被害の区分を入力してください。'])
MESSAGE.append([53, 'W0053', '必須', '浸水土砂被害の区分が入力されていません。', '浸水土砂被害の区分を入力してください。'])
MESSAGE.append([54, 'W0054', '必須', '被害建物棟数, 床下浸水が入力されていません。', '被害建物棟数, 床下浸水を入力してください。'])
MESSAGE.append([55, 'W0055', '必須', '被害建物棟数, 1cm〜49cmが入力されていません。', '被害建物棟数, 1cm〜49cmを入力してください。'])
MESSAGE.append([56, 'W0056', '必須', '被害建物棟数, 50cm〜99cmが入力されていません。', '被害建物棟数, 50cm〜99cmを入力してください。'])
MESSAGE.append([57, 'W0057', '必須', '被害建物棟数, 1m以上が入力されていません。', '被害建物棟数, 1m以上を入力してください。'])
MESSAGE.append([58, 'W0058', '必須', '被害建物棟数, 半壊が入力されていません。', '被害建物棟数, 半壊を入力してください。'])
MESSAGE.append([59, 'W0059', '必須', '被害建物棟数, 全壊・流失が入力されていません。', '被害建物棟数, 全壊・流失を入力してください。'])
MESSAGE.append([60, 'W0060', '必須', '被害建物の延床面積が入力されていません。', '被害建物の延床面積を入力してください。'])
MESSAGE.append([61, 'W0061', '必須', '被災世帯数が入力されていません。', '被災世帯数を入力してください。'])
MESSAGE.append([62, 'W0062', '必須', '被災事業所数が入力されていません。', '被災事業所数を入力してください。'])
MESSAGE.append([63, 'W0063', '必須', '農家・漁家戸数, 床下浸水が入力されていません。', '農家・漁家戸数, 床下浸水を入力してください。'])
MESSAGE.append([64, 'W0064', '必須', '農家・漁家戸数, 1cm〜49cmが入力されていません。', '農家・漁家戸数, 1cm〜49cmを入力してください。'])
MESSAGE.append([65, 'W0065', '必須', '農家・漁家戸数, 50cm〜99cmが入力されていません。', '農家・漁家戸数, 50cm〜99cmを入力してください。'])
MESSAGE.append([66, 'W0066', '必須', '農家・漁家戸数, 1m以上・半壊が入力されていません。', '農家・漁家戸数, 1m以上・半壊を入力してください。'])
MESSAGE.append([67, 'W0067', '必須', '農家・漁家戸数, 全壊・流失が入力されていません。', '農家・漁家戸数, 全壊・流失を入力してください。'])
MESSAGE.append([68, 'W0068', '必須', '事業所従業者数, 床下浸水が入力されていません。', '事業所従業者数, 床下浸水を入力してください。'])
MESSAGE.append([69, 'W0069', '必須', '事業所従業者数, 1cm〜49cmが入力されていません。', '事業所従業者数, 1cm〜49cmを入力してください。'])
MESSAGE.append([70, 'W0070', '必須', '事業所従業者数, 50cm〜99cmが入力されていません。', '事業所従業者数, 50cm〜99cmを入力してください。'])
MESSAGE.append([71, 'W0071', '必須', '事業所従業者数, 1m以上・半壊が入力されていません。', '事業所従業者数, 1m以上・半壊を入力してください。'])
MESSAGE.append([72, 'W0072', '必須', '事業所従業者数, 全壊・流失が入力されていません。', '事業所従業者数, 全壊・流失を入力してください。'])
MESSAGE.append([73, 'W0073', '必須', '事業所の産業区分が入力されていません。', '事業所の産業区分を入力してください。'])
MESSAGE.append([74, 'W0074', '必須', '地下空間の利用形態が入力されていません。', '地下空間の利用形態を入力してください。'])
MESSAGE.append([75, 'W0075', '必須', '備考が入力されていません。', '備考を入力してください。'])
for i in range(76, 100):
    MESSAGE.append([i, '', '', '', ''])

###############################################################################
### 処理名：定数定義
### 単体入力の形式チェックを行った結果のメッセージを定義する。
###############################################################################
MESSAGE.append([100, 'W0100', '形式', '都道府県に全角以外の無効な文字が入力されています。', '全角文字を入力してください。'])
MESSAGE.append([101, 'W0101', '形式', '市区町村に全角以外の無効な文字が入力されています。', '全角文字を入力してください。'])
MESSAGE.append([102, 'W0102', '形式', '水害発生月日に日付として無効な文字が入力されています。', '日付として有効な文字を入力してください。'])
MESSAGE.append([103, 'W0103', '形式', '水害終了月日に日付として無効な文字が入力されています。', '日付として有効な文字を入力してください。'])
MESSAGE.append([104, 'W0104', '形式', '水害原因1に全角以外の無効な文字が入力されています。', '全角文字を入力してください。'])
MESSAGE.append([105, 'W0105', '形式', '水害原因2に全角以外の無効な文字が入力されています。', '全角文字を入力してください。'])
MESSAGE.append([106, 'W0106', '形式', '水害原因3に全角以外の無効な文字が入力されています。', '全角文字を入力してください。'])
MESSAGE.append([107, 'W0107', '形式', '水害区域番号に半角数字以外の無効な文字が入力されています。', '半角数字を入力してください。'])
MESSAGE.append([108, 'W0108', '形式', '水系・沿岸名に全角以外の無効な文字が入力されています。', '全角文字を入力してください。'])
MESSAGE.append([109, 'W0109', '形式', '水系種別に全角以外の無効な文字が入力されています。', '全角文字を入力してください。'])
MESSAGE.append([110, 'W0110', '形式', '河川・海岸名に全角全角以外の無効な文字が入力されています。', '全角文字を入力してください。'])
MESSAGE.append([111, 'W0111', '形式', '河川種別に全角以外の無効な文字が入力されています。', '全角文字を入力してください。'])
MESSAGE.append([112, 'W0112', '形式', '地盤勾配区分に全角以外の無効な文字が入力されています。', '全角文字を入力してください。'])
MESSAGE.append([113, 'W0113', '形式', '水害区域面積の宅地に半角数字以外の無効な文字が入力されています。', '半角数字を入力してください。'])
MESSAGE.append([114, 'W0114', '形式', '水害区域面積の農地に半角数字以外の無効な文字が入力されています。', '半角数字を入力してください。'])
MESSAGE.append([115, 'W0115', '形式', '水害区域面積の地下に半角数字以外の無効な文字が入力されています。', '半角数字を入力してください。'])
MESSAGE.append([116, 'W0116', '形式', '工種に全角以外の無効な文字が入力されています。', '全角文字を入力してください。'])
MESSAGE.append([117, 'W0117', '形式', '農作物被害額に半角数字以外の無効な文字が入力されています。', '半角数字を入力してください。'])
MESSAGE.append([118, 'W0118', '形式', '異常気象コードに半角数字以外の無効な文字が入力されています。', '半角数字を入力してください。'])
for i in range(119, 150):
    MESSAGE.append([i, '', '', '', ''])

MESSAGE.append([150, 'W0150', '形式', '町丁名・大字名に全角以外の無効な文字が入力されています。', '全角文字を入力してください。'])
MESSAGE.append([151, 'W0151', '形式', '名称に全角以外の無効な文字が入力されています。', '全角文字を入力してください。'])
MESSAGE.append([152, 'W0152', '形式', '地上・地下被害の区分に全角以外の無効な文字が入力されています。', '全角文字を入力してください。'])
MESSAGE.append([153, 'W0153', '形式', '浸水土砂被害の区分に全角以外の無効な文字が入力されています。', '全角文字を入力してください。'])
MESSAGE.append([154, 'W0154', '形式', '被害建物棟数, 床下浸水に半角数字以外の無効な文字が入力されています。', '半角数字を入力してください。'])
MESSAGE.append([155, 'W0155', '形式', '被害建物棟数, 1cm〜49cmに半角数字以外の無効な文字が入力されています。', '半角数字を入力してください。'])
MESSAGE.append([156, 'W0156', '形式', '被害建物棟数, 50cm〜99cmに半角数字以外の無効な文字が入力されています。', '半角数字を入力してください。'])
MESSAGE.append([157, 'W0157', '形式', '被害建物棟数, 1m以上に半角数字以外の無効な文字が入力されています。', '半角数字を入力してください。'])
MESSAGE.append([158, 'W0158', '形式', '被害建物棟数, 半壊に半角数字以外の無効な文字が入力されています。', '半角数字を入力してください。'])
MESSAGE.append([159, 'W0159', '形式', '被害建物棟数, 全壊・流失に半角数字以外の無効な文字が入力されています。', '半角数字を入力してください。'])
MESSAGE.append([160, 'W0160', '形式', '被害建物の延床面積に半角数字以外の無効な文字が入力されています。', '半角数字を入力してください。'])
MESSAGE.append([161, 'W0161', '形式', '被災世帯数に半角数字以外の無効な文字が入力されています。', '半角数字を入力してください。'])
MESSAGE.append([162, 'W0162', '形式', '被災事業所数に半角数字以外の無効な文字が入力されています。', '半角数字を入力してください。'])
MESSAGE.append([163, 'W0163', '形式', '農家・漁家戸数, 床下浸水に半角数字以外の無効な文字が入力されています。', '半角数字を入力してください。'])
MESSAGE.append([164, 'W0164', '形式', '農家・漁家戸数, 1cm〜49cmに半角数字以外の無効な文字が入力されています。', '半角数字を入力してください。'])
MESSAGE.append([165, 'W0165', '形式', '農家・漁家戸数, 50cm〜99cmに半角数字以外の無効な文字が入力されています。', '半角数字を入力してください。'])
MESSAGE.append([166, 'W0166', '形式', '農家・漁家戸数, 1m以上・半壊に半角数字以外の無効な文字が入力されています。', '半角数字を入力してください。'])
MESSAGE.append([167, 'W0167', '形式', '農家・漁家戸数, 全壊・流失に半角数字以外の無効な文字が入力されています。', '半角数字を入力してください。'])
MESSAGE.append([168, 'W0168', '形式', '事業所従業者数, 床下浸水に半角数字以外の無効な文字が入力されています。', '半角数字を入力してください。'])
MESSAGE.append([169, 'W0169', '形式', '事業所従業者数, 1cm〜49cmに半角数字以外の無効な文字が入力されています。', '半角数字を入力してください。'])
MESSAGE.append([170, 'W0170', '形式', '事業所従業者数, 50cm〜99cmに半角数字以外の無効な文字が入力されています。', '半角数字を入力してください。'])
MESSAGE.append([171, 'W0171', '形式', '事業所従業者数, 1m以上・半壊に半角数字以外の無効な文字が入力されています。', '半角数字を入力してください。'])
MESSAGE.append([172, 'W0172', '形式', '事業所従業者数, 全壊・流失に半角数字以外の無効な文字が入力されています。', '半角数字を入力してください。'])
MESSAGE.append([173, 'W0173', '形式', '事業所の産業区分に全角以外の無効な文字が入力されています。', '全角文字を入力してください。'])
MESSAGE.append([174, 'W0174', '形式', '地下空間の利用形態に全角以外の無効な文字が入力されています。', '全角文字を入力してください。'])
MESSAGE.append([175, 'W0175', '形式', '備考に全角以外の無効な文字が入力されています。', '全角文字を入力してください。'])
for i in range(176, 200):
    MESSAGE.append([i, '', '', '', ''])

###############################################################################
### 処理名：定数定義
### 単体入力の範囲チェックを行った結果のメッセージを定義する。
###############################################################################
MESSAGE.append([200, 'W0200', '範囲', '都道府県に入力範囲外の無効な値が入力されています。', ''])
MESSAGE.append([201, 'W0201', '範囲', '市区町村に入力範囲外の無効な値が入力されています。', ''])
MESSAGE.append([202, 'W0202', '範囲', '水害発生月日に入力範囲外の無効な値が入力されています。', ''])
MESSAGE.append([203, 'W0203', '範囲', '水害終了月日に入力範囲外の無効な値が入力されています。', ''])
MESSAGE.append([204, 'W0204', '範囲', '水害原因1に入力範囲外の無効な値が入力されています。', '「10:破堤」「20:有堤部溢水」「30:無堤部溢水」「40:内水」「50:窪地内水」「60:洗堀・流出」「70:土石流」「80:地すべり」「90:急傾斜地崩壊」「91:高潮」「92:津波」「93:波浪」「99:その他」のいずれかを入力してください。'])
MESSAGE.append([205, 'W0205', '範囲', '水害原因2に入力範囲外の無効な値が入力されています。', '「10:破堤」「20:有堤部溢水」「30:無堤部溢水」「40:内水」「50:窪地内水」「60:洗堀・流出」「70:土石流」「80:地すべり」「90:急傾斜地崩壊」「91:高潮」「92:津波」「93:波浪」「99:その他」のいずれかを入力してください。'])
MESSAGE.append([206, 'W0206', '範囲', '水害原因3に入力範囲外の無効な値が入力されています。', '「10:破堤」「20:有堤部溢水」「30:無堤部溢水」「40:内水」「50:窪地内水」「60:洗堀・流出」「70:土石流」「80:地すべり」「90:急傾斜地崩壊」「91:高潮」「92:津波」「93:波浪」「99:その他」のいずれかを入力してください。'])
MESSAGE.append([207, 'W0207', '範囲', '水害区域番号に入力範囲外の無効な値が入力されています。', ''])
MESSAGE.append([208, 'W0208', '範囲', '水系・沿岸名に入力範囲外の無効な値が入力されています。', ''])
MESSAGE.append([209, 'W0209', '範囲', '水系種別に入力範囲外の無効な値が入力されています。', ''])
MESSAGE.append([210, 'W0210', '範囲', '河川・海岸名に入力範囲外の無効な値が入力されています。', ''])
MESSAGE.append([211, 'W0211', '範囲', '河川種別に入力範囲外の無効な値が入力されています。', ''])
MESSAGE.append([212, 'W0212', '範囲', '地盤勾配区分に入力範囲外の無効な値が入力されています。', ''])
MESSAGE.append([213, 'W0213', '範囲', '水害区域面積の宅地に入力択範囲外の無効な値が入力されています。', ''])
MESSAGE.append([214, 'W0214', '範囲', '水害区域面積の農地に入力範囲外の無効な値が入力されています。', ''])
MESSAGE.append([215, 'W0215', '範囲', '水害区域面積の地下に入力範囲外の無効な値が入力されています。', ''])
MESSAGE.append([216, 'W0216', '範囲', '工種に入力範囲外の無効な値が入力されています。', ''])
MESSAGE.append([217, 'W0217', '範囲', '農作物被害額に入力範囲外の無効な値が入力されています。', ''])
MESSAGE.append([218, 'W0218', '範囲', '異常気象コードに入力範囲外の無効な値が入力されています。', ''])
for i in range(219, 250):
    MESSAGE.append([i, '', '', '', ''])

MESSAGE.append([250, 'W0250', '範囲', '町丁名・大字名に入力範囲外の無効な値が入力されています。', 'のいずれかを入力してください。'])
MESSAGE.append([251, 'W0251', '範囲', '名称に入力範囲外の無効な値が入力されています。', 'のいずれかを入力してください。'])
MESSAGE.append([252, 'W0252', '範囲', '地上・地下被害の区分に入力範囲外の無効な値が入力されています。', 'のいずれかを入力してください。'])
MESSAGE.append([253, 'W0253', '範囲', '浸水土砂被害の区分に入力範囲外の無効な値が入力されています。', 'のいずれかを入力してください。'])
MESSAGE.append([254, 'W0254', '範囲', '被害建物棟数, 床下浸水に入力範囲外の無効な値が入力されています。', 'のいずれかを入力してください。'])
MESSAGE.append([255, 'W0255', '範囲', '被害建物棟数, 1cm〜49cmに入力範囲外の無効な値が入力されています。', 'のいずれかを入力してください。'])
MESSAGE.append([256, 'W0256', '範囲', '被害建物棟数, 50cm〜99cmに入力範囲外の無効な値が入力されています。', 'のいずれかを入力してください。'])
MESSAGE.append([257, 'W0257', '範囲', '被害建物棟数, 1m以上に入力範囲外の無効な値が入力されています。', 'のいずれかを入力してください。'])
MESSAGE.append([258, 'W0258', '範囲', '被害建物棟数, 半壊に入力範囲外の無効な値が入力されています。', 'のいずれかを入力してください。'])
MESSAGE.append([259, 'W0259', '範囲', '被害建物棟数, 全壊・流失に入力範囲外の無効な値が入力されています。', 'のいずれかを入力してください。'])
MESSAGE.append([260, 'W0260', '範囲', '被害建物の延床面積に入力範囲外の無効な値が入力されています。', 'のいずれかを入力してください。'])
MESSAGE.append([261, 'W0261', '範囲', '被災世帯数に入力範囲外の無効な値が入力されています。', 'のいずれかを入力してください。'])
MESSAGE.append([262, 'W0262', '範囲', '被災事業所数に入力範囲外の無効な値が入力されています。', 'のいずれかを入力してください。'])
MESSAGE.append([263, 'W0263', '範囲', '農家・漁家戸数, 床下浸水に入力範囲外の無効な値が入力されています。', 'のいずれかを入力してください。'])
MESSAGE.append([264, 'W0264', '範囲', '農家・漁家戸数, 1cm〜49cmに入力範囲外の無効な値が入力されています。', 'のいずれかを入力してください。'])
MESSAGE.append([265, 'W0265', '範囲', '農家・漁家戸数, 50cm〜99cmに入力範囲外の無効な値が入力されています。', 'のいずれかを入力してください。'])
MESSAGE.append([266, 'W0266', '範囲', '農家・漁家戸数, 1m以上・半壊に入力範囲外の無効な値が入力されています。', 'のいずれかを入力してください。'])
MESSAGE.append([267, 'W0267', '範囲', '農家・漁家戸数, 全壊・流失に入力範囲外の無効な値が入力されています。', 'のいずれかを入力してください。'])
MESSAGE.append([268, 'W0268', '範囲', '事業所従業者数, 床下浸水に入力範囲外の無効な値が入力されています。', 'のいずれかを入力してください。'])
MESSAGE.append([269, 'W0269', '範囲', '事業所従業者数, 1cm〜49cmに入力範囲外の無効な値が入力されています。', 'のいずれかを入力してください。'])
MESSAGE.append([270, 'W0270', '範囲', '事業所従業者数, 50cm〜99cmに入力範囲外の無効な値が入力されています。', 'のいずれかを入力してください。'])
MESSAGE.append([271, 'W0271', '範囲', '事業所従業者数, 1m以上・半壊に入力範囲外の無効な値が入力されています。', 'のいずれかを入力してください。'])
MESSAGE.append([272, 'W0272', '範囲', '事業所従業者数, 全壊・流失に入力範囲外の無効な値が入力されています。', 'のいずれかを入力してください。'])
MESSAGE.append([273, 'W0273', '範囲', '事業所の産業区分に入力範囲外の無効な値が入力されています。', 'のいずれかを入力してください。'])
MESSAGE.append([274, 'W0274', '範囲', '地下空間の利用形態に入力範囲外の無効な値が入力されています。', 'のいずれかを入力してください。'])
MESSAGE.append([275, 'W0275', '範囲', '備考に入力範囲外の無効な値が入力されています。', 'のいずれかを入力してください。'])
for i in range(276, 300):
    MESSAGE.append([i, '', '', '', ''])

###############################################################################
### 処理名：定数定義
### 単体入力の相関チェックを行った結果のメッセージを定義する。
###############################################################################
MESSAGE.append([300, 'W0300', '相関', '都道府県名が入力されていません。', '都道府県名が入力されていません。'])
MESSAGE.append([301, 'W0301', '相関', '都道府県名が入力されていません。', '都道府県名が入力されていません。'])
MESSAGE.append([302, 'W0302', '相関', '都道府県名が入力されていません。', '都道府県名が入力されていません。'])
MESSAGE.append([303, 'W0303', '相関', '都道府県名が入力されていません。', '都道府県名が入力されていません。'])
MESSAGE.append([304, 'W0304', '相関', '都道府県名が入力されていません。', '都道府県名が入力されていません。'])
MESSAGE.append([305, 'W0305', '相関', '都道府県名が入力されていません。', '都道府県名が入力されていません。'])
MESSAGE.append([306, 'W0306', '相関', '都道府県名が入力されていません。', '都道府県名が入力されていません。'])
MESSAGE.append([307, 'W0307', '相関', '都道府県名が入力されていません。', '都道府県名が入力されていません。'])
MESSAGE.append([308, 'W0308', '相関', '都道府県名が入力されていません。', '都道府県名が入力されていません。'])
MESSAGE.append([309, 'W0309', '相関', '都道府県名が入力されていません。', '都道府県名が入力されていません。'])
MESSAGE.append([310, 'W0310', '相関', '都道府県名が入力されていません。', '都道府県名が入力されていません。'])
MESSAGE.append([311, 'W0311', '相関', '都道府県名が入力されていません。', '都道府県名が入力されていません。'])
MESSAGE.append([312, 'W0312', '相関', '都道府県名が入力されていません。', '都道府県名が入力されていません。'])
MESSAGE.append([313, 'W0313', '相関', '都道府県名が入力されていません。', '都道府県名が入力されていません。'])
MESSAGE.append([314, 'W0314', '相関', '都道府県名が入力されていません。', '都道府県名が入力されていません。'])
for i in range(315, 350):
    MESSAGE.append([i, '', '', '', ''])

MESSAGE.append([350, 'W0350', '相関', '都道府県名が入力されていません。', '都道府県名が入力されていません。'])
MESSAGE.append([351, 'W0351', '相関', '都道府県名が入力されていません。', '都道府県名が入力されていません。'])
MESSAGE.append([352, 'W0352', '相関', '都道府県名が入力されていません。', '都道府県名が入力されていません。'])
MESSAGE.append([353, 'W0353', '相関', '都道府県名が入力されていません。', '都道府県名が入力されていません。'])
MESSAGE.append([354, 'W0354', '相関', '都道府県名が入力されていません。', '都道府県名が入力されていません。'])
MESSAGE.append([355, 'W0355', '相関', '都道府県名が入力されていません。', '都道府県名が入力されていません。'])
MESSAGE.append([356, 'W0356', '相関', '都道府県名が入力されていません。', '都道府県名が入力されていません。'])
MESSAGE.append([357, 'W0357', '相関', '都道府県名が入力されていません。', '都道府県名が入力されていません。'])
MESSAGE.append([358, 'W0358', '相関', '都道府県名が入力されていません。', '都道府県名が入力されていません。'])
MESSAGE.append([359, 'W0359', '相関', '都道府県名が入力されていません。', '都道府県名が入力されていません。'])
MESSAGE.append([360, 'W0360', '相関', '都道府県名が入力されていません。', '都道府県名が入力されていません。'])
MESSAGE.append([361, 'W0361', '相関', '都道府県名が入力されていません。', '都道府県名が入力されていません。'])
MESSAGE.append([362, 'W0362', '相関', '都道府県名が入力されていません。', '都道府県名が入力されていません。'])
MESSAGE.append([363, 'W0363', '相関', '都道府県名が入力されていません。', '都道府県名が入力されていません。'])
MESSAGE.append([364, 'W0364', '相関', '都道府県名が入力されていません。', '都道府県名が入力されていません。'])
MESSAGE.append([365, 'W0365', '相関', '都道府県名が入力されていません。', '都道府県名が入力されていません。'])
MESSAGE.append([366, 'W0366', '相関', '都道府県名が入力されていません。', '都道府県名が入力されていません。'])
MESSAGE.append([367, 'W0367', '相関', '都道府県名が入力されていません。', '都道府県名が入力されていません。'])
MESSAGE.append([368, 'W0368', '相関', '都道府県名が入力されていません。', '都道府県名が入力されていません。'])
MESSAGE.append([369, 'W0369', '相関', '都道府県名が入力されていません。', '都道府県名が入力されていません。'])
MESSAGE.append([370, 'W0370', '相関', '都道府県名が入力されていません。', '都道府県名が入力されていません。'])
MESSAGE.append([371, 'W0371', '相関', '都道府県名が入力されていません。', '都道府県名が入力されていません。'])
MESSAGE.append([372, 'W0372', '相関', '都道府県名が入力されていません。', '都道府県名が入力されていません。'])
MESSAGE.append([373, 'W0373', '相関', '都道府県名が入力されていません。', '都道府県名が入力されていません。'])
MESSAGE.append([374, 'W0374', '相関', '都道府県名が入力されていません。', '都道府県名が入力されていません。'])
MESSAGE.append([375, 'W0375', '相関', '都道府県名が入力されていません。', '都道府県名が入力されていません。'])
for i in range(376, 400):
    MESSAGE.append([i, '', '', '', ''])

###############################################################################
### 処理名：定数定義
### 単体入力の突合チェックを行った結果のメッセージを定義する。
###############################################################################
MESSAGE.append([400, 'W0400', '突合', '都道府県がデータベースに登録されている都道府県と一致しません。', '正しい都道府県を入力してください。'])
MESSAGE.append([401, 'W0401', '突合', '市区町村がデータベースに登録されている市区町村と一致しません。', '正しい市区町村を入力してください。'])
MESSAGE.append([402, 'W0402', '突合', '水害発生月日がデータベースに登録されている水害発生月日と一致しません。', '正しい水害発生月日を入力してください。'])
MESSAGE.append([403, 'W0403', '突合', '水害終了月日がデータベースに登録されている水害終了月日と一致しません。', '正しい水害終了月日を入力してください。'])
MESSAGE.append([404, 'W0404', '突合', '水害原因1がデータベースに登録されている水害原因1と一致しません。', '正しい水害原因1を入力してください。'])
MESSAGE.append([405, 'W0405', '突合', '水害原因2がデータベースに登録されている水害原因2と一致しません。', '正しい水害原因2を入力してください。'])
MESSAGE.append([406, 'W0406', '突合', '水害原因3がデータベースに登録されている水害原因3と一致しません。', '正しい水害原因3を入力してください。'])
MESSAGE.append([407, 'W0407', '突合', '水害区域番号がデータベースに登録されている水害区域番号と一致しません。', '正しい水害区域番号を入力してください。'])
MESSAGE.append([408, 'W0408', '突合', '水系・沿岸名がデータベースに登録されている水系・沿岸名と一致しません。', '正しい水系・沿岸名を入力してください。'])
MESSAGE.append([409, 'W0409', '突合', '水系種別がデータベースに登録されている水系種別と一致しません。', '正しい水系種別を入力してください。'])
MESSAGE.append([410, 'W0410', '突合', '河川・海岸名がデータベースに登録されている河川・海岸名と一致しません。', '正しい河川・海岸名を入力してください。'])
MESSAGE.append([411, 'W0411', '突合', '河川種別がデータベースに登録されている河川種別と一致しません。', '正しい河川種別を入力してください。'])
MESSAGE.append([412, 'W0412', '突合', '地盤勾配区分がデータベースに登録されている地盤勾配区分と一致しません。', '正しい地盤勾配区分を入力してください。'])
MESSAGE.append([413, 'W0413', '突合', '水害区域面積の宅地がデータベースに登録されている水害区域面積の宅地と一致しません。', '正しい水害区域面積の宅地を入力してください。'])
MESSAGE.append([414, 'W0414', '突合', '水害区域面積の農地がデータベースに登録されている水害区域面積の農地と一致しません。', '正しい水害区域面積の農地を入力してください。'])
MESSAGE.append([415, 'W0415', '突合', '水害区域面積の地下がデータベースに登録されている水害区域面積の地下と一致しません。', '正しい水害区域面積の地下を入力してください。'])
MESSAGE.append([416, 'W0416', '突合', '工種がデータベースに登録されている工種と一致しません。', '正しい工種を入力してください。'])
MESSAGE.append([417, 'W0417', '突合', '農作物被害額がデータベースに登録されている農作物被害額と一致しません。', '正しい農作物被害額を入力してください。'])
MESSAGE.append([418, 'W0418', '突合', '異常気象コードがデータベースに登録されている異常気象コードと一致しません。', '正しい異常気象コードを入力してください。'])
for i in range(419, 450):
    MESSAGE.append([i, '', '', '', ''])

MESSAGE.append([450, 'W0450', '突合', '町丁名・大字名がデータベースに登録されている町丁名・大字名と一致しません。', '正しい町丁名・大字名を入力してください。'])
MESSAGE.append([451, 'W0451', '突合', '名称がデータベースに登録されている名称と一致しません。', '正しい名称を入力してください。'])
MESSAGE.append([452, 'W0452', '突合', '地上・地下被害の区分がデータベースに登録されている地上・地下被害の区分と一致しません。', '正しい地上・地下被害の区分を入力してください。'])
MESSAGE.append([453, 'W0453', '突合', '浸水土砂被害の区分がデータベースに登録されている浸水土砂被害の区分と一致しません。', '正しい浸水土砂被害の区分を入力してください。'])
MESSAGE.append([454, 'W0454', '突合', '被害建物棟数, 床下浸水がデータベースに登録されている被害建物棟数, 床下浸水と一致しません。', '正しい被害建物棟数, 床下浸水を入力してください。'])
MESSAGE.append([455, 'W0455', '突合', '被害建物棟数, 1cm〜49cmがデータベースに登録されている被害建物棟数, 1cm〜49cmと一致しません。', '正しい被害建物棟数, 1cm〜49cmを入力してください。'])
MESSAGE.append([456, 'W0456', '突合', '被害建物棟数, 50cm〜99cmがデータベースに登録されている被害建物棟数, 50cm〜99cmと一致しません。', '正しい被害建物棟数, 50cm〜99cmを入力してください。'])
MESSAGE.append([457, 'W0457', '突合', '被害建物棟数, 1m以上がデータベースに登録されている被害建物棟数, 1m以上と一致しません。', '正しい被害建物棟数, 1m以上を入力してください。'])
MESSAGE.append([458, 'W0458', '突合', '被害建物棟数, 半壊がデータベースに登録されている被害建物棟数, 半壊と一致しません。', '正しい被害建物棟数, 半壊を入力してください。'])
MESSAGE.append([459, 'W0459', '突合', '被害建物棟数, 全壊・流失がデータベースに登録されている被害建物棟数, 全壊・流失と一致しません。', '正しい被害建物棟数, 全壊・流失を入力してください。'])
MESSAGE.append([460, 'W0460', '突合', '被害建物の延床面積がデータベースに登録されている被害建物の延床面積と一致しません。', '正しい被害建物の延床面積を入力してください。'])
MESSAGE.append([461, 'W0461', '突合', '被災世帯数がデータベースに登録されている被災世帯数と一致しません。', '正しい被災世帯数を入力してください。'])
MESSAGE.append([462, 'W0462', '突合', '被災事業所数がデータベースに登録されている被災事業所数と一致しません。', '正しい被災事業所数を入力してください。'])
MESSAGE.append([463, 'W0463', '突合', '農家・漁家戸数, 床下浸水がデータベースに登録されている農家・漁家戸数, 床下浸水と一致しません。', '正しい農家・漁家戸数, 床下浸水を入力してください。'])
MESSAGE.append([464, 'W0464', '突合', '農家・漁家戸数, 1cm〜49cmがデータベースに登録されている農家・漁家戸数, 1cm〜49cmと一致しません。', '正しい農家・漁家戸数, 1cm〜49cmを入力してください。'])
MESSAGE.append([465, 'W0465', '突合', '農家・漁家戸数, 50cm〜99cmがデータベースに登録されている農家・漁家戸数, 50cm〜99cmと一致しません。', '正しい農家・漁家戸数, 50cm〜99cmを入力してください。'])
MESSAGE.append([466, 'W0466', '突合', '農家・漁家戸数, 1m以上・半壊がデータベースに登録されている農家・漁家戸数, 1m以上・半壊と一致しません。', '正しい農家・漁家戸数, 1m以上・半壊を入力してください。'])
MESSAGE.append([467, 'W0467', '突合', '農家・漁家戸数, 全壊・流失がデータベースに登録されている農家・漁家戸数, 全壊・流失と一致しません。', '正しい農家・漁家戸数, 全壊・流失を入力してください。'])
MESSAGE.append([468, 'W0468', '突合', '事業所従業者数, 床下浸水がデータベースに登録されている事業所従業者数, 床下浸水と一致しません。', '正しい事業所従業者数, 床下浸水を入力してください。'])
MESSAGE.append([469, 'W0469', '突合', '事業所従業者数, 1cm〜49cmがデータベースに登録されている事業所従業者数, 1cm〜49cmと一致しません。', '正しい事業所従業者数, 1cm〜49cmを入力してください。'])
MESSAGE.append([470, 'W0470', '突合', '事業所従業者数, 50cm〜99cmがデータベースに登録されている事業所従業者数, 50cm〜99cmと一致しません。', '正しい事業所従業者数, 50cm〜99cmを入力してください。'])
MESSAGE.append([471, 'W0471', '突合', '事業所従業者数, 1m以上・半壊がデータベースに登録されている事業所従業者数, 1m以上・半壊と一致しません。', '正しい事業所従業者数, 1m以上・半壊を入力してください。'])
MESSAGE.append([472, 'W0472', '突合', '事業所従業者数, 全壊・流失がデータベースに登録されている事業所従業者数, 全壊・流失と一致しません。', '正しい事業所従業者数, 全壊・流失を入力してください。'])
MESSAGE.append([473, 'W0473', '突合', '事業所の産業区分がデータベースに登録されている事業所の産業区分と一致しません。', '正しい事業所の産業区分を入力してください。'])
MESSAGE.append([474, 'W0474', '突合', '地下空間の利用形態がデータベースに登録されている地下空間の利用形態と一致しません。', '正しい地下空間の利用形態を入力してください。'])
MESSAGE.append([475, 'W0475', '突合', '備考がデータベースに登録されている備考と一致しません。', '正しい備考を入力してください。'])
for i in range(476, 500):
    MESSAGE.append([i, '', '', '', ''])

###############################################################################
### 関数名：is_zenkaku
###############################################################################
### @login_required(None, login_url='/P0100Login/')
def is_zenkaku(arg):
    try:
        pass
    except:
        raise Http404("[ERROR] is_zenkaku().")
    return True

###############################################################################
### 関数名：is_mmdd
###############################################################################
### @login_required(None, login_url='/P0100Login/')
def is_mmdd(arg):
    try:
        pass
    except:
        raise Http404("[ERROR] is_mmdd().")
    return True

###############################################################################
### 関数名：split_name_code
### (1) 引数がname:codeの場合、[name,code]を返す。
### (2) 引数がnameの場合、[name,'']を返す。
### (3) 引数がcodeの場合、['', code]を返す。
###############################################################################
def split_name_code(arg):
    try:
        name_code = ['', '']
        if arg is not None:
            if len(arg.split(':')) == 0:
                name_code = ['', '']
            elif len(arg.split(':')) == 1:
                if arg.isdecimal():
                    name_code = ['', str(arg)]
                else:
                    name_code = [str(arg), '']
            elif len(arg.split(':')) == 2:
                name_code = arg.split(':')
            else:
                name_code = ['', '']
    except:
        return ['', '']
    
    return name_code

###############################################################################
### 関数名：is_cell_value_date
### (1) 引数がYYYY-MM-DD形式の場合、Trueを返す。
### (2) 引数がYYYY/MM/DD形式の場合、Trueを返す。
### (3) 引数が上記以外の場合、Falseを返す。
###############################################################################
def is_cell_value_date(arg):
    try:
        try:
            _ = datetime.strptime(arg, '%Y/%m/%d')
            return True
        except ValueError:
            pass
        try:
            _ = datetime.strptime(arg, '%Y-%m-%d')
            return True
        except ValueError:
            pass
        return False
    except:
        return False

###############################################################################
### 関数名：none_if_empty
### 関数名：convert_empty_to_none
### (1) 引数がNoneの場合、Noneを返す。
### (2) 引数が''の場合、Noneを返す。
### (3) 引数が上記以外の場合、引数を返す。
###############################################################################
### def none_if_empty(arg):
def convert_empty_to_none(arg):
    if arg is None or arg == None:
        return None
    elif arg == '' or arg == "":
        return None
    else:
        return arg

###############################################################################
### 関数名：add_comment_to_cell
### 
###############################################################################
def add_comment_to_cell(ws_ippan, ws_result, row, column, message):
    if ws_ippan.cell(row=row, column=column).comment is None:
        ws_ippan.cell(row=row, column=column).comment = Comment(message, '')
    else:
        ws_ippan.cell(row=row, column=column).comment = Comment(str(ws_ippan.cell(row=row, column=column).comment.text) + message, '')
    if ws_result.cell(row=row, column=column).comment is None:
        ws_result.cell(row=row, column=column).comment = Comment(message, '')
    else:
        ws_result.cell(row=row, column=column).comment = Comment(str(ws_result.cell(row=row, column=column).comment.text) + message, '')
    return True    

###############################################################################
### 関数名：add_fill_to_cell
### 
###############################################################################
def add_fill_to_cell(ws_ippan, ws_result, row, column, fill):
    ws_ippan.cell(row=row, column=column).fill = fill
    ws_result.cell(row=row, column=column).fill = fill
    return True

###############################################################################
### 関数名：index_view
### (1)GETの場合、EXCELアップロード画面を表示する。
### (2)POSTの場合、アップロードされたEXCELファイルをチェックして、正常ケースの場合、DBに登録する。
### (3)POSTの場合、アップロードされたEXCELファイルをチェックして、警告ケースの場合、DBに登録する。
### ※複数EXCELシート対応版
###############################################################################
@login_required(None, login_url='/P0100Login/')
def index_view(request):
    try:
        #######################################################################
        #######################################################################
        ### 引数チェック処理(0000)
        ### ブラウザからのリクエストと引数をチェックする。
        #######################################################################
        #######################################################################
        print_log('[INFO] ########################################', 'INFO')
        print_log('[INFO] P0300ExcelUpload.index_view()関数が開始しました。', 'INFO')
        print_log('[INFO] P0300ExcelUpload.index_view()関数 request = {}'.format(request.method), 'INFO')
        print_log('[INFO] P0300ExcelUpload.index_view()関数 STEP 1/34.', 'INFO')
        
        #######################################################################
        ### 局所変数セット処理(0010)
        ### チェック結果を格納するために局所変数をセットする。
        ### result_require_list: 必須チェック結果を格納するリスト
        ### result_require_grid: 必須チェック結果を格納するリスト
        ### result_format_list: 形式チェック結果を格納するリスト
        ### result_format_grid: 形式チェック結果を格納するリスト
        ### result_range_list: 範囲チェック結果を格納するリスト
        ### result_range_grid: 範囲チェック結果を格納するリスト
        ### result_correlate_list: 相関チェック結果を格納するリスト
        ### result_correlate_grid: 相関チェック結果を格納するリスト
        ### result_compare_list: 突合チェック結果を格納するリスト
        ### result_compare_grid: 突合チェック結果を格納するリスト
        #######################################################################
        print_log('[INFO] P0300ExcelUpload.index_view()関数 STEP 2/34.', 'INFO')
        result_require_list = []
        result_require_grid = []
        result_format_list = []
        result_format_grid = []
        result_range_list = []
        result_range_grid = []
        result_correlate_list = []
        result_correlate_grid = []
        result_compare_list = []
        result_compare_grid = []
    
        #######################################################################
        ### 条件分岐処理(0020)
        ### (1)GETの場合、EXCELアップロード画面を表示して関数を抜ける。
        ### (2)POSTの場合、アップロードされたEXCELファイルをチェックする。
        ### ※関数の内部のネスト数を浅くするため。
        #######################################################################
        print_log('[INFO] P0300ExcelUpload.index_view()関数 STEP 3/34.', 'INFO')
        if request.method == 'GET':
            form = ExcelUploadForm()
            return render(request, 'P0300ExcelUpload/index.html', {'form': form})
        
        elif request.method == 'POST':
            form = ExcelUploadForm(request.POST, request.FILES)
            
        #######################################################################
        ### フォーム検証処理(0030)
        ### (1)フォームが正しい場合、処理を継続する。
        ### (2)フォームが正しくない場合、ERROR画面を表示して関数を抜ける。
        ### ※関数の内部のネスト数を浅くするため。
        #######################################################################
        print_log('[INFO] P0300ExcelUpload.index_view()関数 STEP 4/34.', 'INFO')
        if form.is_valid():
            pass
        
        else:
            return HttpResponseRedirect('fail')
    
        #######################################################################
        ### EXCELファイル入出力処理(0040)
        ### (1)局所変数に値をセットする。
        ### (2)アップロードされたEXCELファイルを保存する。
        #######################################################################
        print_log('[INFO] P0300ExcelUpload.index_view()関数 STEP 5/34.', 'INFO')
        ### upload_file_object = request.FILES['file']
        ### upload_file_path = 'media/documents/' + upload_file_object.name
        ### result_file_path = 'static/ippan_chosa_result2.xlsx'
        ### with open(upload_file_path, 'wb+') as destination:
        ###     for chunk in upload_file_object.chunks():
        ###         destination.write(chunk)

        JST = timezone(timedelta(hours=9), 'JST')
        datetime_now_strftime = datetime.now(JST).strftime('%Y%m%d%H%M%S')
        
        input_file_object = request.FILES['file']
        print_log('input_file_object = {}'.format(input_file_object), 'INFO')
        
        input_file_path = 'repository/202206/ippan_chosa_input_' + datetime_now_strftime + '.xlsx'
        ### input_file_path = '/static/repository/202206/ippan_chosa_input_' + datetime_now_strftime + '.xlsx'
        print_log('input_file_path = {}'.format(input_file_path), 'INFO')
        
        with open(input_file_path, 'wb+') as destination:
            for chunk in input_file_object.chunks():
                destination.write(chunk)

        output_file_path = 'repository/202206/ippan_chosa_output_' + datetime_now_strftime + '.xlsx'
        ### output_file_path = '/static/repository/202206/ippan_chosa_output_' + datetime_now_strftime + '.xlsx'
        
        print_log('[INFO] P0300ExcelUpload.index_view()関数 input_file_path = {}'.format(input_file_path), 'INFO')
        print_log('[INFO] P0300ExcelUpload.index_view()関数 output_file_path = {}'.format(output_file_path), 'INFO')
                
        #######################################################################
        ### EXCELファイル入出力処理(0050)
        ### (1)アップロードされたEXCELファイルのワークブックを読み込む。
        ### (2)IPPANワークシートをコピーして、チェック結果を格納するCHECK_RESULTワークシートを追加する。
        ### (3)追加したワークシートを2シート目に移動する。
        ### (4)ワークシートの最大行数を局所変数のws_max_rowにセットする。
        ### (5)背景赤色の塗りつぶしを局所変数のfillにセットする。
        ### wb: ワークブック
        ### ws_ippan: IPPANワークシート
        ### ws_result: CHECK_RESULTワークシート
        ### wx_max_row: ワークシートの最大行数
        ### fill: 背景赤色の塗りつぶし
        #######################################################################
        print_log('[INFO] P0300ExcelUpload.index_view()関数 STEP 6/34.', 'INFO')
        ### wb = openpyxl.load_workbook(upload_file_path)
        wb = openpyxl.load_workbook(input_file_path)
        ws_ippan = []
        ws_result = []
        ws_max_row = []

        for ws_temp in wb.worksheets:
            if 'IPPAN' in ws_temp.title:
                ws_ippan.append(ws_temp)

        for i, ws_temp in enumerate(wb.worksheets):
            if 'IPPAN' in ws_temp.title:
                ws_result.append(wb.copy_worksheet(ws_temp))
                ws_result[-1].title = 'RESULT' + ws_temp.title
                
        for ws_temp in wb.worksheets:
            if 'RESULT' in ws_temp.title:
                wb.move_sheet(ws_temp.title, offset=-wb.index(ws_temp))
                wb.move_sheet(ws_temp.title, offset=1)
        
        for ws_temp in wb.worksheets:
            ws_temp.sheet_view.tabSelected = None
            
        wb.active = 1

        #######################################################################
        ### EXCELファイル入出力処理(0060)
        ### (1)EXCELシート毎に最大行を探索する。
        ### (2)局所変数のws_max_rowリストに追加する。
        #######################################################################
        print_log('[INFO] P0300ExcelUpload.index_view()関数 STEP 7/34.', 'INFO')
        for i, ws_temp in enumerate(ws_ippan):
            _max_row_ = 19
            for j in range(ws_temp.max_row + 1, 19, -1):
                if ws_temp.cell(row=j, column=2).value is None:
                    pass
                else:
                    _max_row_ = j
                    break
                    
            ws_max_row.append(_max_row_)

        #######################################################################
        ### EXCELファイル入出力処理(0070)
        ### EXCELセルの背景赤色を局所変数のfillに設定する。
        #######################################################################
        print_log('[INFO] P0300ExcelUpload.index_view()関数 STEP 8/34.', 'INFO')
        fill = openpyxl.styles.PatternFill(patternType='solid', fgColor='FF0000', bgColor='FF0000')

        #######################################################################
        ### DBアクセス処理(0080)
        ### (1)DBから突合せチェック用のデータを取得する。
        ### (2)突合せチェック用のリストを生成する。
        #######################################################################
        print_log('[INFO] P0300ExcelUpload.index_view()関数 STEP 9/34.', 'INFO')
        ken_list = KEN.objects.raw("""SELECT * FROM KEN ORDER BY CAST(KEN_CODE AS INTEGER)""", [])
        city_list = CITY.objects.raw("""SELECT * FROM CITY ORDER BY CAST(CITY_CODE AS INTEGER)""", [])
        cause_list = CAUSE.objects.raw("""SELECT * FROM CAUSE ORDER BY CAST(CAUSE_CODE AS INTEGER)""", [])
        area_list = AREA.objects.raw("""SELECT * FROM AREA ORDER BY CAST(AREA_ID AS INTEGER)""", [])
        suikei_list = SUIKEI.objects.raw("""SELECT * FROM SUIKEI ORDER BY CAST(SUIKEI_CODE AS INTEGER)""", [])
        suikei_type_list = SUIKEI_TYPE.objects.raw("""SELECT * FROM SUIKEI_TYPE ORDER BY CAST(SUIKEI_TYPE_CODE AS INTEGER)""", [])
        kasen_list = KASEN.objects.raw("""SELECT * FROM KASEN ORDER BY CAST(KASEN_CODE AS INTEGER)""", [])
        kasen_type_list = KASEN_TYPE.objects.raw("""SELECT * FROM KASEN_TYPE ORDER BY CAST(KASEN_TYPE_CODE AS INTEGER)""", [])
        gradient_list = GRADIENT.objects.raw("""SELECT * FROM GRADIENT ORDER BY CAST(GRADIENT_CODE AS INTEGER)""", [])
        kasen_kaigan_list = KASEN_KAIGAN.objects.raw("""SELECT * FROM KASEN_KAIGAN ORDER BY CAST(KASEN_KAIGAN_CODE AS INTEGER)""", [])
        weather_list = WEATHER.objects.raw("""SELECT * FROM WEATHER ORDER BY CAST(WEATHER_ID AS INTEGER)""", [])
        building_list = BUILDING.objects.raw("""SELECT * FROM BUILDING ORDER BY CAST(BUILDING_CODE AS INTEGER)""", [])
        underground_list = UNDERGROUND.objects.raw("""SELECT * FROM UNDERGROUND ORDER BY CAST(UNDERGROUND_CODE AS INTEGER)""", [])
        flood_sediment_list = FLOOD_SEDIMENT.objects.raw("""SELECT * FROM FLOOD_SEDIMENT ORDER BY CAST(FLOOD_SEDIMENT_CODE AS INTEGER)""", [])
        industry_list = INDUSTRY.objects.raw("""SELECT * FROM INDUSTRY ORDER BY CAST(INDUSTRY_CODE AS INTEGER)""", [])
        usage_list = USAGE.objects.raw("""SELECT * FROM USAGE ORDER BY CAST(USAGE_CODE AS INTEGER)""", [])
        
        ken_code_list = [ken.ken_code for ken in ken_list]
        ken_name_list = [ken.ken_name for ken in ken_list]
        ken_name_code_list = [str(ken.ken_name) + ":" + str(ken.ken_code) for ken in ken_list]
        city_code_list = [city.city_code for city in city_list]
        city_name_list = [city.city_name for city in city_list]
        city_name_code_list = [str(city.city_name) + ":" + str(city.city_code) for city in city_list]
        cause_code_list = [cause.cause_code for cause in cause_list]
        cause_name_list = [cause.cause_name for cause in cause_list]
        cause_name_code_list = [str(cause.cause_name) + ":" + str(cause.cause_code) for cause in cause_list]
        area_id_list = [area.area_id for area in area_list]
        area_name_list = [area.area_name for area in area_list]
        area_name_id_list = [str(area.area_name) + ":" + str(area.area_id) for area in area_list]
        suikei_code_list = [suikei.suikei_code for suikei in suikei_list]
        suikei_name_list = [suikei.suikei_name for suikei in suikei_list]
        suikei_name_code_list = [str(suikei.suikei_name) + ":" + str(suikei.suikei_code) for suikei in suikei_list]
        suikei_type_code_list = [suikei_type.suikei_type_code for suikei_type in suikei_type_list]
        suikei_type_name_list = [suikei_type.suikei_type_name for suikei_type in suikei_type_list]
        suikei_type_name_code_list = [str(suikei_type.suikei_type_name) + ":" + str(suikei_type.suikei_type_code) for suikei_type in suikei_type_list]
        kasen_code_list = [kasen.kasen_code for kasen in kasen_list]
        kasen_name_list = [kasen.kasen_name for kasen in kasen_list]
        kasen_name_code_list = [str(kasen.kasen_name) + ":" + str(kasen.kasen_code) for kasen in kasen_list]
        kasen_type_code_list = [kasen_type.kasen_type_code for kasen_type in kasen_type_list]
        kasen_type_name_list = [kasen_type.kasen_type_name for kasen_type in kasen_type_list]
        kasen_type_name_code_list = [str(kasen_type.kasen_type_name) + ":" + str(kasen_type.kasen_type_code) for kasen_type in kasen_type_list]
        gradient_code_list = [gradient.gradient_code for gradient in gradient_list]
        gradient_name_list = [gradient.gradient_name for gradient in gradient_list]
        gradient_name_code_list = [str(gradient.gradient_name) + ":" + str(gradient.gradient_code) for gradient in gradient_list]
        kasen_kaigan_code_list = [kasen_kaigan.kasen_kaigan_code for kasen_kaigan in kasen_kaigan_list]
        kasen_kaigan_name_list = [kasen_kaigan.kasen_kaigan_name for kasen_kaigan in kasen_kaigan_list]
        kasen_kaigan_name_code_list = [str(kasen_kaigan.kasen_kaigan_name) + ":" + str(kasen_kaigan.kasen_kaigan_code) for kasen_kaigan in kasen_kaigan_list]
        weather_id_list = [weather.weather_id for weather in weather_list]
        weather_name_list = [weather.weather_name for weather in weather_list]
        weather_name_id_list = [str(weather.weather_name) + ":" + str(weather.weather_id) for weather in weather_list]
        building_code_list = [building.building_code for building in building_list]
        building_name_list = [building.building_name for building in building_list]
        building_name_code_list = [str(building.building_name) + ":" + str(building.building_code) for building in building_list]
        underground_code_list = [underground.underground_code for underground in underground_list]
        underground_name_list = [underground.underground_name for underground in underground_list]
        underground_name_code_list = [str(underground.underground_name) + ":" + str(underground.underground_code) for underground in underground_list]
        flood_sediment_code_list = [flood_sediment.flood_sediment_code for flood_sediment in flood_sediment_list]
        flood_sediment_name_list = [flood_sediment.flood_sediment_name for flood_sediment in flood_sediment_list]
        flood_sediment_name_code_list = [str(flood_sediment.flood_sediment_name) + ":" + str(flood_sediment.flood_sediment_code) for flood_sediment in flood_sediment_list]
        industry_code_list = [industry.industry_code for industry in industry_list]
        industry_name_list = [industry.industry_name for industry in industry_list]
        industry_name_code_list = [str(industry.industry_name) + ":" + str(industry.industry_code) for industry in industry_list]
        usage_code_list = [usage.usage_code for usage in usage_list]
        usage_name_list = [usage.usage_name for usage in usage_list]
        usage_name_code_list = [str(usage.usage_name) + ":" + str(usage.usage_code) for usage in usage_list]
    
        #######################################################################
        #######################################################################
        ### EXCELセルデータ必須チェック処理（1000）
        ### (1)セルB7からセルI7について、必須項目に値がセットされていることをチェックする。
        ### (2)チェック結果リストにセルの行、列とメッセージを追加する。
        ### (3)IPPANワークシートとRESULTワークシートのセルに背景赤色の塗りつぶしをセットする。
        ### TO-DO: ダミーで全項目を必須としている。任意項目はコメントアウトすること。
        #######################################################################
        #######################################################################
        print_log('[INFO] P0300ExcelUpload.index_view()関数 STEP 10/34.', 'INFO')
        for i, _ in enumerate(ws_ippan):
            ### 7行目
            ### セルB7: 都道府県に値がセットされていることをチェックする。
            if ws_ippan[i].cell(row=7, column=2).value is None:
                result_require_list.append([ws_ippan[i].title, 7, 2, MESSAGE[0][0], MESSAGE[0][1], MESSAGE[0][2], MESSAGE[0][3], MESSAGE[0][4]])
                add_fill_to_cell(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=7, column=2, fill=fill)
                add_comment_to_cell(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=7, column=2, message=MESSAGE[0][3]+MESSAGE[0][4])
                
            ### セルC7: 市区町村に値がセットされていることをチェックする。
            if ws_ippan[i].cell(row=7, column=3).value is None:
                result_require_list.append([ws_ippan[i].title, 7, 3, MESSAGE[1][0], MESSAGE[1][1], MESSAGE[1][2], MESSAGE[1][3], MESSAGE[1][4]])
                add_fill_to_cell(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=7, column=3, fill=fill)
                add_comment_to_cell(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=7, column=3, message=MESSAGE[1][3]+MESSAGE[1][4])
            
            ### セルD7: 水害発生月日に値がセットされていることをチェックする。
            if ws_ippan[i].cell(row=7, column=4).value is None:
                result_require_list.append([ws_ippan[i].title, 7, 4, MESSAGE[2][0], MESSAGE[2][1], MESSAGE[2][2], MESSAGE[2][3], MESSAGE[2][4]])
                add_fill_to_cell(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=7, column=4, fill=fill)
                add_comment_to_cell(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=7, column=4, message=MESSAGE[2][3]+MESSAGE[2][4])
    
            ### セルE7: 水害終了月日に値がセットされていることをチェックする。
    
            ### セルF7: 水害原因1に値がセットされていることをチェックする。
            if ws_ippan[i].cell(row=7, column=6).value is None:
                result_require_list.append([ws_ippan[i].title, 7, 6, MESSAGE[4][0], MESSAGE[4][1], MESSAGE[4][2], MESSAGE[4][3], MESSAGE[4][4]])
                add_fill_to_cell(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=7, column=6, fill=fill)
                add_comment_to_cell(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=7, column=6, message=MESSAGE[4][3]+MESSAGE[4][4])
    
            ### セルG7: 水害原因2に値がセットされていることをチェックする。
                
            ### セルH7: 水害原因3に値がセットされていることをチェックする。
                
            ### セルI7: 水害区域番号に値がセットされていることをチェックする。
            if ws_ippan[i].cell(row=7, column=9).value is None:
                result_require_list.append([ws_ippan[i].title, 7, 9, MESSAGE[7][0], MESSAGE[7][1], MESSAGE[7][2], MESSAGE[7][3], MESSAGE[7][4]])
                add_fill_to_cell(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=7, column=9, fill=fill)
                add_comment_to_cell(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=7, column=9, message=MESSAGE[7][3]+MESSAGE[7][4])
    
        #######################################################################
        ### EXCELセルデータ必須チェック処理（1010）
        ### (1)セルB10からセルF10について、必須項目に値がセットされていることをチェックする。
        ### (2)チェック結果リストにセルの行、列とメッセージを追加する。
        ### (3)IPPANワークシートとRESULTワークシートのセルに背景赤色の塗りつぶしをセットする。
        ### TO-DO: ダミーで全項目を必須としている。任意項目はコメントアウトすること。
        #######################################################################
        print_log('[INFO] P0300ExcelUpload.index_view()関数 STEP 11/34.', 'INFO')
        for i, _ in enumerate(ws_ippan):
            ### 10行目
            ### セルB10: 水系・沿岸名に値がセットされていることをチェックする。
            if ws_ippan[i].cell(row=10, column=2).value is None:
                result_require_list.append([ws_ippan[i].title, 10, 2, MESSAGE[8][0], MESSAGE[8][1], MESSAGE[8][2], MESSAGE[8][3], MESSAGE[8][4]])
                add_fill_to_cell(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=10, column=2, fill=fill)
                add_comment_to_cell(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=10, column=2, message=MESSAGE[8][3]+MESSAGE[8][4])
                
            ### セルC10: 水系種別に値がセットされていることをチェックする。
            if ws_ippan[i].cell(row=10, column=3).value is None:
                result_require_list.append([ws_ippan[i].title, 10, 3, MESSAGE[9][0], MESSAGE[9][1], MESSAGE[9][2], MESSAGE[9][3], MESSAGE[9][4]])
                add_fill_to_cell(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=10, column=3, fill=fill)
                add_comment_to_cell(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=10, column=3, message=MESSAGE[9][3]+MESSAGE[9][4])
                
            ### セルD10: 河川・海岸名に値がセットされていることをチェックする。
            if ws_ippan[i].cell(row=10, column=4).value is None:
                result_require_list.append([ws_ippan[i].title, 10, 4, MESSAGE[10][0], MESSAGE[10][1], MESSAGE[10][2], MESSAGE[10][3], MESSAGE[10][4]])
                add_fill_to_cell(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=10, column=4, fill=fill)
                add_comment_to_cell(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=10, column=4, message=MESSAGE[10][3]+MESSAGE[10][4])
                
            ### セルE10: 河川種別に値がセットされていることをチェックする。
            if ws_ippan[i].cell(row=10, column=5).value is None:
                result_require_list.append([ws_ippan[i].title, 10, 5, MESSAGE[11][0], MESSAGE[11][1], MESSAGE[11][2], MESSAGE[11][3], MESSAGE[11][4]])
                add_fill_to_cell(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=10, column=5, fill=fill)
                add_comment_to_cell(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=10, column=5, message=MESSAGE[11][3]+MESSAGE[11][4])
                
            ### セルF10: 地盤勾配区分に値がセットされていることをチェックする。
            if ws_ippan[i].cell(row=10, column=6).value is None:
                result_require_list.append([ws_ippan[i].title, 10, 6, MESSAGE[12][0], MESSAGE[12][1], MESSAGE[12][2], MESSAGE[12][3], MESSAGE[12][4]])
                add_fill_to_cell(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=10, column=6, fill=fill)
                add_comment_to_cell(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=10, column=6, message=MESSAGE[12][3]+MESSAGE[12][4])
    
        #######################################################################
        ### EXCELセルデータ必須チェック処理（1020）
        ### (1)セルB14からセルJ14について、必須項目に値がセットされていることをチェックする。
        ### (2)チェック結果リストにセルの行、列とメッセージを追加する。
        ### (3)IPPANワークシートとRESULTワークシートのセルに背景赤色の塗りつぶしをセットする。
        ### TO-DO: ダミーで全項目を必須としている。任意項目はコメントアウトすること。
        #######################################################################
        print_log('[INFO] P0300ExcelUpload.index_view()関数 STEP 12/34.', 'INFO')
        ### for i, _ in enumerate(ws_ippan):
            ### 14行目
            ### セルB14: 水害区域面積の宅地に値がセットされていることをチェックする。
            ### セルC14: 水害区域面積の農地に値がセットされていることをチェックする。
            ### セルD14: 水害区域面積の地下に値がセットされていることをチェックする。
            ### セルF14: 工種に値がセットされていることをチェックする。
            ### セルH14: 農作物被害額に値がセットされていることをチェックする。
            ### セルJ14: 異常気象コードに値がセットされていることをチェックする。
    
        #######################################################################
        ### EXCELセルデータ必須チェック処理（1030）
        ### (1)セルB20からセルAA20について、必須項目に値がセットされていることをチェックする。
        ### (2)チェック結果リストにセルの行、列とメッセージを追加する。
        ### (3)IPPANワークシートとRESULTワークシートのセルに背景赤色の塗りつぶしをセットする。
        ### TO-DO: ダミーで全項目を必須としている。任意項目はコメントアウトすること。
        #######################################################################
        print_log('[INFO] P0300ExcelUpload.index_view()関数 STEP 13/34.', 'INFO')
        for i, _ in enumerate(ws_ippan):
            if ws_max_row[i] >= 20:
                for j in range(20, ws_max_row[i] + 1):
                    ### セルB20: 町丁名・大字名に値がセットされていることをチェックする。
                    if ws_ippan[i].cell(row=j, column=2).value is None:
                        result_require_grid.append([ws_ippan[i].title, j, 2, MESSAGE[50][0], MESSAGE[50][1], MESSAGE[50][2], MESSAGE[50][3], MESSAGE[50][4]])
                        add_fill_to_cell(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=j, column=2, fill=fill)
                        add_comment_to_cell(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=j, column=2, message=MESSAGE[50][3]+MESSAGE[50][4])
                        
                    ### セルC20: 名称に値がセットされていることをチェックする。
                    if ws_ippan[i].cell(row=j, column=3).value is None:
                        result_require_grid.append([ws_ippan[i].title, j, 3, MESSAGE[51][0], MESSAGE[51][1], MESSAGE[51][2], MESSAGE[51][3], MESSAGE[51][4]])
                        add_fill_to_cell(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=j, column=3, fill=fill)
                        add_comment_to_cell(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=j, column=3, message=MESSAGE[51][3]+MESSAGE[51][4])
                        
                    ### セルD20: 地上・地下被害の区分に値がセットされていることをチェックする。
                    if ws_ippan[i].cell(row=j, column=4).value is None:
                        result_require_grid.append([ws_ippan[i].title, j, 4, MESSAGE[52][0], MESSAGE[52][1], MESSAGE[52][2], MESSAGE[52][3], MESSAGE[52][4]])
                        add_fill_to_cell(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=j, column=4, fill=fill)
                        add_comment_to_cell(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=j, column=4, message=MESSAGE[52][3]+MESSAGE[52][4])
                        
                    ### セルE20: 浸水土砂被害の区分に値がセットされていることをチェックする。
                    if ws_ippan[i].cell(row=j, column=5).value is None:
                        result_require_grid.append([ws_ippan[i].title, j, 5, MESSAGE[53][0], MESSAGE[53][1], MESSAGE[53][2], MESSAGE[53][3], MESSAGE[53][4]])
                        add_fill_to_cell(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=j, column=5, fill=fill)
                        add_comment_to_cell(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=j, column=5, message=MESSAGE[53][3]+MESSAGE[53][4])
                        
                    ### セルF20: 被害建物棟数, 床下浸水に値がセットされていることをチェックする。
                    ### セルG20: 被害建物棟数, 1cm〜49cmに値がセットされていることをチェックする。
                    ### セルH20: 被害建物棟数, 50cm〜99cmに値がセットされていることをチェックする。
                    ### セルI20: 被害建物棟数, 1m以上に値がセットされていることをチェックする。
                    ### セルJ20: 被害建物棟数, 半壊に値がセットされていることをチェックする。
                    ### セルK20: 被害建物棟数, 全壊・流失に値がセットされていることをチェックする。
                    ### セルL20: 被害建物の延床面積に値がセットされていることをチェックする。
                    ### セルM20: 被災世帯数に値がセットされていることをチェックする。
                    ### セルN20: 被災事業所数に値がセットされていることをチェックする。
                    ### セルO20: 農家・漁家戸数, 床下浸水に値がセットされていることをチェックする。
                    ### セルP20: 農家・漁家戸数, 1cm〜49cmに値がセットされていることをチェックする。
                    ### セルQ20: 農家・漁家戸数, 50cm〜99cmに値がセットされていることをチェックする。
                    ### セルR20: 農家・漁家戸数, 1m以上・半壊に値がセットされていることをチェックする。
                    ### セルS20: 農家・漁家戸数, 全壊・流失に値がセットされていることをチェックする。
                    ### セルT20: 事業所従業者数, 床下浸水に値がセットされていることをチェックする。
                    ### セルU20: 事業所従業者数, 1cm〜49cmに値がセットされていることをチェックする。
                    ### セルV20: 事業所従業者数, 50cm〜99cmに値がセットされていることをチェックする。
                    ### セルW20: 事業所従業者数, 1m以上・半壊に値がセットされていることをチェックする。
                    ### セルX20: 事業所従業者数, 全壊・流失に値がセットされていることをチェックする。
                    ### セルY20: 事業所の産業区分に値がセットされていることをチェックする。
                    ### セルZ20: 地下空間の利用形態に値がセットされていることをチェックする。
                    ### セルAA20: 備考に値がセットされていることをチェックする。
        
        #######################################################################
        #######################################################################
        ### EXCELセルデータ形式チェック処理（2000）
        ### (1)セルB7からセルI7について形式が正しいことをチェックする。
        ### (2)チェック結果リストにセルの行、列とメッセージを追加する。
        ### (3)IPPANワークシートとRESULTワークシートのセルに背景赤色の塗りつぶしをセットする。
        ### TO-DO: is_zenkoku関数、is_mmdd関数はダミーである。処理を記述すること。
        ### 形式チェックでは、値がセットされている場合のみチェックを行う。
        ### 必須チェックは別途必須チェックで行うためである。
        #######################################################################
        #######################################################################
        print_log('[INFO] P0300ExcelUpload.index_view()関数 STEP 14/34.', 'INFO')
        for i, _ in enumerate(ws_ippan):
            ### 7行目
            ### セルB7: 都道府県について形式が正しいことをチェックする。
            if ws_ippan[i].cell(row=7, column=2).value is None:
                pass
            else:
                if split_name_code(ws_ippan[i].cell(row=7, column=2).value)[-1].isdecimal() == False:
                    result_format_list.append([ws_ippan[i].title, 7, 2, MESSAGE[100][0], MESSAGE[100][1], MESSAGE[100][2], MESSAGE[100][3], MESSAGE[100][4]])
                    add_fill_to_cell(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=7, column=2, fill=fill)
                    add_comment_to_cell(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=7, column=2, message=MESSAGE[100][3]+MESSAGE[100][4])
        
            ### セルC7: 市区町村について形式が正しいことをチェックする。
            if ws_ippan[i].cell(row=7, column=3).value is None:
                pass
            else:
                if split_name_code(ws_ippan[i].cell(row=7, column=3).value)[-1].isdecimal == False:
                    result_format_list.append([ws_ippan[i].title, 7, 3, MESSAGE[101][0], MESSAGE[101][1], MESSAGE[101][2], MESSAGE[101][3], MESSAGE[101][4]])
                    add_fill_to_cell(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=7, column=3, fill=fill)
                    add_comment_to_cell(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=7, column=3, message=MESSAGE[101][3]+MESSAGE[101][4])
      
            ### セルD7: 水害発生月日について形式が正しいことをチェックする。
            if ws_ippan[i].cell(row=7, column=4).value is None:
                pass
            else:
                if is_cell_value_date(ws_ippan[i].cell(row=7, column=4).value) == False:
                    result_format_list.append([ws_ippan[i].title, 7, 4, MESSAGE[102][0], MESSAGE[102][1], MESSAGE[102][2], MESSAGE[102][3], MESSAGE[102][4]])
                    add_fill_to_cell(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=7, column=4, fill=fill)
                    add_comment_to_cell(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=7, column=4, message=MESSAGE[102][3]+MESSAGE[102][4])
        
            ### セルE7: 水害終了月日について形式が正しいことをチェックする。
            if ws_ippan[i].cell(row=7, column=5).value is None:
                pass
            else:
                if is_cell_value_date(ws_ippan[i].cell(row=7, column=5).value) == False:
                    result_format_list.append([ws_ippan[i].title, 7, 5, MESSAGE[103][0], MESSAGE[103][1], MESSAGE[103][2], MESSAGE[103][3], MESSAGE[103][4]])
                    add_fill_to_cell(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=7, column=5, fill=fill)
                    add_comment_to_cell(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=7, column=5, message=MESSAGE[103][3]+MESSAGE[103][4])
                
            ### セルF7: 水害原因1について形式が正しいことをチェックする。
            if ws_ippan[i].cell(row=7, column=6).value is None:
                pass
            else:
                if split_name_code(ws_ippan[i].cell(row=7, column=6).value)[-1].isdecimal() == False:
                    result_format_list.append([ws_ippan[i].title, 7, 6, MESSAGE[104][0], MESSAGE[104][1], MESSAGE[104][2], MESSAGE[104][3], MESSAGE[104][4]])
                    add_fill_to_cell(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=7, column=6, fill=fill)
                    add_comment_to_cell(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=7, column=6, message=MESSAGE[104][3]+MESSAGE[104][4])
                
            ### セルG7: 水害原因2について形式が正しいことをチェックする。
            if ws_ippan[i].cell(row=7, column=7).value is None:
                pass
            else:
                if split_name_code(ws_ippan[i].cell(row=7, column=7).value)[-1].isdecimal() == False:
                    result_format_list.append([ws_ippan[i].title, 7, 7, MESSAGE[105][0], MESSAGE[105][1], MESSAGE[105][2], MESSAGE[105][3], MESSAGE[105][4]])
                    add_fill_to_cell(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=7, column=7, fill=fill)
                    add_comment_to_cell(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=7, column=7, message=MESSAGE[105][3]+MESSAGE[105][4])
                
            ### セルH7: 水害原因3について形式が正しいことをチェックする。
            if ws_ippan[i].cell(row=7, column=8).value is None:
                pass
            else:
                if split_name_code(ws_ippan[i].cell(row=7, column=8).value)[-1].isdecimal() == False:
                    result_format_list.append([ws_ippan[i].title, 7, 8, MESSAGE[106][0], MESSAGE[106][1], MESSAGE[106][2], MESSAGE[106][3], MESSAGE[106][4]])
                    add_fill_to_cell(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=7, column=8, fill=fill)
                    add_comment_to_cell(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=7, column=8, message=MESSAGE[106][3]+MESSAGE[106][4])
                
            ### セルI7: 水害区域番号について形式が正しいことをチェックする。
            if ws_ippan[i].cell(row=7, column=9).value is None:
                pass
            else:
                if split_name_code(ws_ippan[i].cell(row=7, column=9).value)[-1].isdecimal() == False:
                    result_format_list.append([ws_ippan[i].title, 7, 9, MESSAGE[107][0], MESSAGE[107][1], MESSAGE[107][2], MESSAGE[107][3], MESSAGE[107][4]])
                    add_fill_to_cell(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=7, column=9, fill=fill)
                    add_comment_to_cell(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=7, column=9, message=MESSAGE[107][3]+MESSAGE[107][4])
    
        #######################################################################
        ### EXCELセルデータ形式チェック処理（2010）
        ### (1)セルB10からセルF10について形式が正しいことをチェックする。
        ### (2)チェック結果リストにセルの行、列とメッセージを追加する。
        ### (3)IPPANワークシートとRESULTワークシートのセルに背景赤色の塗りつぶしをセットする。
        ### TO-DO: is_zenkoku関数、is_mmdd関数はダミーである。処理を記述すること。
        #######################################################################
        print_log('[INFO] P0300ExcelUpload.index_view()関数 STEP 15/34.', 'INFO')
        for i, _ in enumerate(ws_ippan):
            ### 10行目
            ### セルB10: 水系・沿岸名について形式が正しいことをチェックする。
            if ws_ippan[i].cell(row=10, column=2).value is None:
                pass
            else:
                if split_name_code(ws_ippan[i].cell(row=10, column=2).value)[-1].isdecimal() == False:
                    result_format_list.append([ws_ippan[i].title, 10, 2, MESSAGE[108][0], MESSAGE[108][1], MESSAGE[108][2], MESSAGE[108][3], MESSAGE[108][4]])
                    add_fill_to_cell(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=10, column=2, fill=fill)
                    add_comment_to_cell(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=10, column=2, message=MESSAGE[108][3]+MESSAGE[108][4])
                
            ### セルC10: 水系種別について形式が正しいことをチェックする。
            if ws_ippan[i].cell(row=10, column=3).value is None:
                pass
            else:
                if split_name_code(ws_ippan[i].cell(row=10, column=3).value)[-1].isdecimal() == False:
                    result_format_list.append([ws_ippan[i].title, 10, 3, MESSAGE[109][0], MESSAGE[109][1], MESSAGE[109][2], MESSAGE[109][3], MESSAGE[109][4]])
                    add_fill_to_cell(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=10, column=3, fill=fill)
                    add_comment_to_cell(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=10, column=3, message=MESSAGE[109][3]+MESSAGE[109][4])
                
            ### セルD10: 河川・海岸名について形式が正しいことをチェックする。
            if ws_ippan[i].cell(row=10, column=4).value is None:
                pass
            else:
                if split_name_code(ws_ippan[i].cell(row=10, column=4).value)[-1].isdecimal() == False:
                    result_format_list.append([ws_ippan[i].title, 10, 4, MESSAGE[110][0], MESSAGE[110][1], MESSAGE[110][2], MESSAGE[110][3], MESSAGE[110][4]])
                    add_fill_to_cell(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=10, column=4, fill=fill)
                    add_comment_to_cell(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=10, column=4, message=MESSAGE[110][3]+MESSAGE[110][4])
                
            ### セルE10: 河川種別について形式が正しいことをチェックする。
            if ws_ippan[i].cell(row=10, column=5).value is None:
                pass
            else:
                if split_name_code(ws_ippan[i].cell(row=10, column=5).value)[-1].isdecimal() == False:
                    result_format_list.append([ws_ippan[i].title, 10, 5, MESSAGE[111][0], MESSAGE[111][1], MESSAGE[111][2], MESSAGE[111][3], MESSAGE[111][4]])
                    add_fill_to_cell(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=10, column=5, fill=fill)
                    add_comment_to_cell(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=10, column=5, message=MESSAGE[111][3]+MESSAGE[111][4])
                
            ### セルF10: 地盤勾配区分について形式が正しいことをチェックする。
            if ws_ippan[i].cell(row=10, column=6).value is None:
                pass
            else:
                if split_name_code(ws_ippan[i].cell(row=10, column=6).value)[-1].isdecimal() == False:
                    result_format_list.append([ws_ippan[i].title, 10, 6, MESSAGE[112][0], MESSAGE[112][1], MESSAGE[112][2], MESSAGE[112][3], MESSAGE[112][4]])
                    add_fill_to_cell(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=10, column=6, fill=fill)
                    add_comment_to_cell(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=10, column=6, message=MESSAGE[112][3]+MESSAGE[112][4])
    
        #######################################################################
        ### EXCELセルデータ形式チェック処理（2020）
        ### (1)セルB14からセルJ14について形式が正しいことをチェックする。
        ### (2)チェック結果リストにセルの行、列とメッセージを追加する。
        ### (3)IPPANワークシートとRESULTワークシートのセルに背景赤色の塗りつぶしをセットする。
        ### TO-DO: is_zenkoku関数、is_mmdd関数はダミーである。処理を記述すること。
        #######################################################################
        print_log('[INFO] P0300ExcelUpload.index_view()関数 STEP 16/34.', 'INFO')
        for i, _ in enumerate(ws_ippan):
            ### 14行目
            ### セルB14: 水害区域面積の宅地について形式が正しいことをチェックする。
            if ws_ippan[i].cell(row=14, column=2).value is None:
                pass
            else:
                if isinstance(ws_ippan[i].cell(row=14, column=2).value, int) == False and \
                    isinstance(ws_ippan[i].cell(row=14, column=2).value, float) == False:
                    result_format_list.append([ws_ippan[i].title, 14, 2, MESSAGE[113][0], MESSAGE[113][1], MESSAGE[113][2], MESSAGE[113][3], MESSAGE[113][4]])
                    add_fill_to_cell(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=14, column=2, fill=fill)
                    add_comment_to_cell(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=14, column=2, message=MESSAGE[113][3]+MESSAGE[113][4])
    
            ### セルC14: 水害区域面積の農地について形式が正しいことをチェックする。
            if ws_ippan[i].cell(row=14, column=3).value is None:
                pass
            else:
                if isinstance(ws_ippan[i].cell(row=14, column=3).value, int) == False and \
                    isinstance(ws_ippan[i].cell(row=14, column=3).value, float) == False:
                    result_format_list.append([ws_ippan[i].title, 14, 3, MESSAGE[114][0], MESSAGE[114][1], MESSAGE[114][2], MESSAGE[114][3], MESSAGE[114][4]])
                    add_fill_to_cell(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=14, column=3, fill=fill)
                    add_comment_to_cell(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=14, column=3, message=MESSAGE[114][3]+MESSAGE[114][4])
                
            ### セルD14: 水害区域面積の地下について形式が正しいことをチェックする。
            if ws_ippan[i].cell(row=14, column=4).value is None:
                pass
            else:
                if isinstance(ws_ippan[i].cell(row=14, column=4).value, int) == False and \
                    isinstance(ws_ippan[i].cell(row=14, column=4).value, float) == False:
                    result_format_list.append([ws_ippan[i].title, 14, 4, MESSAGE[115][0], MESSAGE[115][1], MESSAGE[115][2], MESSAGE[115][3], MESSAGE[115][4]])
                    add_fill_to_cell(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=14, column=4, fill=fill)
                    add_comment_to_cell(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=14, column=4, message=MESSAGE[115][3]+MESSAGE[115][4])
                
            ### セルF14: 工種について形式が正しいことをチェックする。
            if ws_ippan[i].cell(row=14, column=6).value is None:
                pass
            else:
                if split_name_code(ws_ippan[i].cell(row=14, column=6).value)[-1].isdecimal() == False:
                    result_format_list.append([ws_ippan[i].title, 14, 6, MESSAGE[116][0], MESSAGE[116][1], MESSAGE[116][2], MESSAGE[116][3], MESSAGE[116][4]])
                    add_fill_to_cell(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=14, column=6, fill=fill)
                    add_comment_to_cell(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=14, column=6, message=MESSAGE[116][3]+MESSAGE[116][4])
                
            ### セルH14: 農作物被害額について形式が正しいことをチェックする。
            if ws_ippan[i].cell(row=14, column=8).value is None:
                pass
            else:
                if isinstance(ws_ippan[i].cell(row=14, column=8).value, int) == False and \
                    isinstance(ws_ippan[i].cell(row=14, column=8).value, float) == False:
                    result_format_list.append([ws_ippan[i].title, 14, 8, MESSAGE[117][0], MESSAGE[117][1], MESSAGE[117][2], MESSAGE[117][3], MESSAGE[117][4]])
                    add_fill_to_cell(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=14, column=8, fill=fill)
                    add_comment_to_cell(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=14, column=8, message=MESSAGE[117][3]+MESSAGE[117][4])
                
            ### セルJ14: 異常気象コードについて形式が正しいことをチェックする。
            if ws_ippan[i].cell(row=14, column=10).value is None:
                pass
            else:
                if split_name_code(ws_ippan[i].cell(row=14, column=10).value)[-1].isdecimal() == False:
                    result_format_list.append([ws_ippan[i].title, 14, 10, MESSAGE[118][0], MESSAGE[118][1], MESSAGE[118][2], MESSAGE[118][3], MESSAGE[118][4]])
                    add_fill_to_cell(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=14, column=10, fill=fill)
                    add_comment_to_cell(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=14, column=10, message=MESSAGE[118][3]+MESSAGE[118][4])
                
        #######################################################################
        ### EXCELセルデータ形式チェック処理（2030）
        ### (1)セルB20からセルAA20について形式が正しいことをチェックする。
        ### (2)チェック結果リストにセルの行、列とメッセージを追加する。
        ### (3)IPPANワークシートとRESULTワークシートのセルに背景赤色の塗りつぶしをセットする。
        ### TO-DO: is_zenkoku関数、is_mmdd関数はダミーである。処理を記述すること。
        #######################################################################
        print_log('[INFO] P0300ExcelUpload.index_view()関数 STEP 17/34.', 'INFO')
        for i, _ in enumerate(ws_ippan):
            if ws_max_row[i] >= 20:
                for j in range(20, ws_max_row[i] + 1):
                    ### セルB20: 町丁名・大字名について形式が正しいことをチェックする。
                        
                    ### セルC20: 名称について形式が正しいことをチェックする。
                    if ws_ippan[i].cell(row=j, column=3).value is None:
                        pass
                    else:
                        if split_name_code(ws_ippan[i].cell(row=j, column=3).value)[-1].isdecimal() == False:
                            result_format_grid.append([ws_ippan[i].title, j, 3, MESSAGE[151][0], MESSAGE[151][1], MESSAGE[151][2], MESSAGE[151][3], MESSAGE[151][4]])
                            add_fill_to_cell(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=j, column=3, fill=fill)
                            add_comment_to_cell(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=j, column=3, message=MESSAGE[151][3]+MESSAGE[151][4])
                        
                    ### セルD20: 地上・地下被害の区分について形式が正しいことをチェックする。
                    if ws_ippan[i].cell(row=j, column=4).value is None:
                        pass
                    else:
                        if split_name_code(ws_ippan[i].cell(row=j, column=4).value)[-1].isdecimal() == False:
                            result_format_grid.append([ws_ippan[i].title, j, 4, MESSAGE[152][0], MESSAGE[152][1], MESSAGE[152][2], MESSAGE[152][3], MESSAGE[152][4]])
                            add_fill_to_cell(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=j, column=4, fill=fill)
                            add_comment_to_cell(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=j, column=4, message=MESSAGE[152][3]+MESSAGE[152][4])
                        
                    ### セルE20: 浸水土砂被害の区分について形式が正しいことをチェックする。
                    if ws_ippan[i].cell(row=j, column=5).value is None:
                        pass
                    else:
                        if split_name_code(ws_ippan[i].cell(row=j, column=5).value)[-1].isdecimal() == False:
                            result_format_grid.append([ws_ippan[i].title, j, 5, MESSAGE[153][0], MESSAGE[153][1], MESSAGE[153][2], MESSAGE[153][3], MESSAGE[153][4]])
                            add_fill_to_cell(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=j, column=5, fill=fill)
                            add_comment_to_cell(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=j, column=5, message=MESSAGE[153][3]+MESSAGE[153][4])
                        
                    ### セルF20: 被害建物棟数, 床下浸水について形式が正しいことをチェックする。
                    if ws_ippan[i].cell(row=j, column=6).value is None:
                        pass
                    else:
                        if isinstance(ws_ippan[i].cell(row=j, column=6).value, int) == False and \
                            isinstance(ws_ippan[i].cell(row=j, column=6).value, float) == False:
                            result_format_grid.append([ws_ippan[i].title, j, 6, MESSAGE[154][0], MESSAGE[154][1], MESSAGE[154][2], MESSAGE[154][3], MESSAGE[154][4]])
                            add_fill_to_cell(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=j, column=6, fill=fill)
                            add_comment_to_cell(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=j, column=6, message=MESSAGE[154][3]+MESSAGE[154][4])
                        
                    ### セルG20: 被害建物棟数, 1cm〜49cmについて形式が正しいことをチェックする。
                    if ws_ippan[i].cell(row=j, column=7).value is None:
                        pass
                    else:
                        if isinstance(ws_ippan[i].cell(row=j, column=7).value, int) == False and \
                            isinstance(ws_ippan[i].cell(row=j, column=7).value, float) == False:
                            result_format_grid.append([ws_ippan[i].title, j, 7, MESSAGE[155][0], MESSAGE[155][1], MESSAGE[155][2], MESSAGE[155][3], MESSAGE[155][4]])
                            add_fill_to_cell(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=j, column=7, fill=fill)
                            add_comment_to_cell(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=j, column=7, message=MESSAGE[155][3]+MESSAGE[155][4])
                        
                    ### セルH20: 被害建物棟数, 50cm〜99cmについて形式が正しいことをチェックする。
                    if ws_ippan[i].cell(row=j, column=8).value is None:
                        pass
                    else:
                        if isinstance(ws_ippan[i].cell(row=j, column=8).value, int) == False and \
                            isinstance(ws_ippan[i].cell(row=j, column=8).value, float) == False:
                            result_format_grid.append([ws_ippan[i].title, j, 8, MESSAGE[156][0], MESSAGE[156][1], MESSAGE[156][2], MESSAGE[156][3], MESSAGE[156][4]])
                            add_fill_to_cell(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=j, column=8, fill=fill)
                            add_comment_to_cell(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=j, column=8, message=MESSAGE[156][3]+MESSAGE[156][4])
                        
                    ### セルI20: 被害建物棟数, 1m以上について形式が正しいことをチェックする。
                    if ws_ippan[i].cell(row=j, column=9).value is None:
                        pass
                    else:
                        if isinstance(ws_ippan[i].cell(row=j, column=9).value, int) == False and \
                            isinstance(ws_ippan[i].cell(row=j, column=9).value, float) == False:
                            result_format_grid.append([ws_ippan[i].title, j, 9, MESSAGE[157][0], MESSAGE[157][1], MESSAGE[157][2], MESSAGE[157][3], MESSAGE[157][4]])
                            add_fill_to_cell(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=j, column=9, fill=fill)
                            add_comment_to_cell(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=j, column=9, message=MESSAGE[157][3]+MESSAGE[157][4])
                        
                    ### セルJ20: 被害建物棟数, 半壊について形式が正しいことをチェックする。
                    if ws_ippan[i].cell(row=j, column=10).value is None:
                        pass
                    else:
                        if isinstance(ws_ippan[i].cell(row=j, column=10).value, int) == False and \
                            isinstance(ws_ippan[i].cell(row=j, column=10).value, float) == False:
                            result_format_grid.append([ws_ippan[i].title, j, 10, MESSAGE[158][0], MESSAGE[158][1], MESSAGE[158][2], MESSAGE[158][3], MESSAGE[158][4]])
                            add_fill_to_cell(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=j, column=10, fill=fill)
                            add_comment_to_cell(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=j, column=10, message=MESSAGE[158][3]+MESSAGE[158][4])
                        
                    ### セルK20: 被害建物棟数, 全壊・流失について形式が正しいことをチェックする。
                    if ws_ippan[i].cell(row=j, column=11).value is None:
                        pass
                    else:
                        if isinstance(ws_ippan[i].cell(row=j, column=11).value, int) == False and \
                            isinstance(ws_ippan[i].cell(row=j, column=11).value, float) == False:
                            result_format_grid.append([ws_ippan[i].title, j, 11, MESSAGE[159][0], MESSAGE[159][1], MESSAGE[159][2], MESSAGE[159][3], MESSAGE[159][4]])
                            add_fill_to_cell(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=j, column=11, fill=fill)
                            add_comment_to_cell(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=j, column=11, message=MESSAGE[159][3]+MESSAGE[159][4])
                        
                    ### セルL20: 被害建物の延床面積について形式が正しいことをチェックする。
                    if ws_ippan[i].cell(row=j, column=12).value is None:
                        pass
                    else:
                        if isinstance(ws_ippan[i].cell(row=j, column=12).value, int) == False and \
                            isinstance(ws_ippan[i].cell(row=j, column=12).value, float) == False:
                            result_format_grid.append([ws_ippan[i].title, j, 12, MESSAGE[160][0], MESSAGE[160][1], MESSAGE[160][2], MESSAGE[160][3], MESSAGE[160][4]])
                            add_fill_to_cell(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=j, column=12, fill=fill)
                            add_comment_to_cell(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=j, column=12, message=MESSAGE[160][3]+MESSAGE[160][4])
                        
                    ### セルM20: 被災世帯数について形式が正しいことをチェックする。
                    if ws_ippan[i].cell(row=j, column=13).value is None:
                        pass
                    else:
                        if isinstance(ws_ippan[i].cell(row=j, column=13).value, int) == False and \
                            isinstance(ws_ippan[i].cell(row=j, column=13).value, float) == False:
                            result_format_grid.append([ws_ippan[i].title, j, 13, MESSAGE[161][0], MESSAGE[161][1], MESSAGE[161][2], MESSAGE[161][3], MESSAGE[161][4]])
                            add_fill_to_cell(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=j, column=13, fill=fill)
                            add_comment_to_cell(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=j, column=13, message=MESSAGE[161][3]+MESSAGE[161][4])
                        
                    ### セルN20: 被災事業所数について形式が正しいことをチェックする。
                    if ws_ippan[i].cell(row=j, column=14).value is None:
                        pass
                    else:
                        if isinstance(ws_ippan[i].cell(row=j, column=14).value, int) == False and \
                            isinstance(ws_ippan[i].cell(row=j, column=14).value, float) == False:
                            result_format_grid.append([ws_ippan[i].title, j, 14, MESSAGE[162][0], MESSAGE[162][1], MESSAGE[162][2], MESSAGE[162][3], MESSAGE[162][4]])
                            add_fill_to_cell(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=j, column=14, fill=fill)
                            add_comment_to_cell(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=j, column=14, message=MESSAGE[162][3]+MESSAGE[162][4])
                        
                    ### セルO20: 農家・漁家戸数, 床下浸水について形式が正しいことをチェックする。
                    if ws_ippan[i].cell(row=j, column=15).value is None:
                        pass
                    else:
                        if isinstance(ws_ippan[i].cell(row=j, column=15).value, int) == False and \
                            isinstance(ws_ippan[i].cell(row=j, column=15).value, float) == False:
                            result_format_grid.append([ws_ippan[i].title, j, 15, MESSAGE[163][0], MESSAGE[163][1], MESSAGE[163][2], MESSAGE[163][3], MESSAGE[163][4]])
                            add_fill_to_cell(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=j, column=15, fill=fill)
                            add_comment_to_cell(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=j, column=15, message=MESSAGE[163][3]+MESSAGE[163][4])
                        
                    ### セルP20: 農家・漁家戸数, 1cm〜49cmについて形式が正しいことをチェックする。
                    if ws_ippan[i].cell(row=j, column=16).value is None:
                        pass
                    else:
                        if isinstance(ws_ippan[i].cell(row=j, column=16).value, int) == False and \
                            isinstance(ws_ippan[i].cell(row=j, column=16).value, float) == False:
                            result_format_grid.append([ws_ippan[i].title, j, 16, MESSAGE[164][0], MESSAGE[164][1], MESSAGE[164][2], MESSAGE[164][3], MESSAGE[164][4]])
                            add_fill_to_cell(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=j, column=16, fill=fill)
                            add_comment_to_cell(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=j, column=16, message=MESSAGE[164][3]+MESSAGE[164][4])
                        
                    ### セルQ20: 農家・漁家戸数, 50cm〜99cmについて形式が正しいことをチェックする。
                    if ws_ippan[i].cell(row=j, column=17).value is None:
                        pass
                    else:
                        if isinstance(ws_ippan[i].cell(row=j, column=17).value, int) == False and \
                            isinstance(ws_ippan[i].cell(row=j, column=17).value, float) == False:
                            result_format_grid.append([ws_ippan[i].title, j, 17, MESSAGE[165][0], MESSAGE[165][1], MESSAGE[165][2], MESSAGE[165][3], MESSAGE[165][4]])
                            add_fill_to_cell(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=j, column=17, fill=fill)
                            add_comment_to_cell(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=j, column=17, message=MESSAGE[165][3]+MESSAGE[165][4])
                        
                    ### セルR20: 農家・漁家戸数, 1m以上・半壊について形式が正しいことをチェックする。
                    if ws_ippan[i].cell(row=j, column=18).value is None:
                        pass
                    else:
                        if isinstance(ws_ippan[i].cell(row=j, column=18).value, int) == False and \
                            isinstance(ws_ippan[i].cell(row=j, column=18).value, float) == False:
                            result_format_grid.append([ws_ippan[i].title, j, 18, MESSAGE[166][0], MESSAGE[166][1], MESSAGE[166][2], MESSAGE[166][3], MESSAGE[166][4]])
                            add_fill_to_cell(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=j, column=18, fill=fill)
                            add_comment_to_cell(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=j, column=18, message=MESSAGE[166][3]+MESSAGE[166][4])
                        
                    ### セルS20: 農家・漁家戸数, 全壊・流失について形式が正しいことをチェックする。
                    if ws_ippan[i].cell(row=j, column=19).value is None:
                        pass
                    else:
                        if isinstance(ws_ippan[i].cell(row=j, column=19).value, int) == False and \
                            isinstance(ws_ippan[i].cell(row=j, column=19).value, float) == False:
                            result_format_grid.append([ws_ippan[i].title, j, 19, MESSAGE[167][0], MESSAGE[167][1], MESSAGE[167][2], MESSAGE[167][3], MESSAGE[167][4]])
                            add_fill_to_cell(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=j, column=19, fill=fill)
                            add_comment_to_cell(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=j, column=19, message=MESSAGE[167][3]+MESSAGE[167][4])
                        
                    ### セルT20: 事業所従業者数, 床下浸水について形式が正しいことをチェックする。
                    if ws_ippan[i].cell(row=j, column=20).value is None:
                        pass
                    else:
                        if isinstance(ws_ippan[i].cell(row=j, column=20).value, int) == False and \
                            isinstance(ws_ippan[i].cell(row=j, column=20).value, float) == False:
                            result_format_grid.append([ws_ippan[i].title, j, 20, MESSAGE[168][0], MESSAGE[168][1], MESSAGE[168][2], MESSAGE[168][3], MESSAGE[168][4]])
                            add_fill_to_cell(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=j, column=20, fill=fill)
                            add_comment_to_cell(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=j, column=20, message=MESSAGE[168][3]+MESSAGE[168][4])
                        
                    ### セルU20: 事業所従業者数, 1cm〜49cmについて形式が正しいことをチェックする。
                    if ws_ippan[i].cell(row=j, column=21).value is None:
                        pass
                    else:
                        if isinstance(ws_ippan[i].cell(row=j, column=21).value, int) == False and \
                            isinstance(ws_ippan[i].cell(row=j, column=21).value, float) == False:
                            result_format_grid.append([ws_ippan[i].title, j, 21, MESSAGE[169][0], MESSAGE[169][1], MESSAGE[169][2], MESSAGE[169][3], MESSAGE[169][4]])
                            add_fill_to_cell(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=j, column=21, fill=fill)
                            add_comment_to_cell(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=j, column=21, message=MESSAGE[169][3]+MESSAGE[169][4])
                        
                    ### セルV20: 事業所従業者数, 50cm〜99cmについて形式が正しいことをチェックする。
                    if ws_ippan[i].cell(row=j, column=22).value is None:
                        pass
                    else:
                        if isinstance(ws_ippan[i].cell(row=j, column=22).value, int) == False and \
                            isinstance(ws_ippan[i].cell(row=j, column=22).value, float) == False:
                            result_format_grid.append([ws_ippan[i].title, j, 22, MESSAGE[170][0], MESSAGE[170][1], MESSAGE[170][2], MESSAGE[170][3], MESSAGE[170][4]])
                            add_fill_to_cell(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=j, column=22, fill=fill)
                            add_comment_to_cell(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=j, column=22, message=MESSAGE[170][3]+MESSAGE[170][4])
                        
                    ### セルW20: 事業所従業者数, 1m以上・半壊について形式が正しいことをチェックする。
                    if ws_ippan[i].cell(row=j, column=23).value is None:
                        pass
                    else:
                        if isinstance(ws_ippan[i].cell(row=j, column=23).value, int) == False and \
                            isinstance(ws_ippan[i].cell(row=j, column=23).value, float) == False:
                            result_format_grid.append([ws_ippan[i].title, j, 23, MESSAGE[171][0], MESSAGE[171][1], MESSAGE[171][2], MESSAGE[171][3], MESSAGE[171][4]])
                            add_fill_to_cell(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=j, column=23, fill=fill)
                            add_comment_to_cell(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=j, column=23, message=MESSAGE[171][3]+MESSAGE[171][4])
                        
                    ### セルX20: 事業所従業者数, 全壊・流失について形式が正しいことをチェックする。
                    if ws_ippan[i].cell(row=j, column=24).value is None:
                        pass
                    else:
                        if isinstance(ws_ippan[i].cell(row=j, column=24).value, int) == False and \
                            isinstance(ws_ippan[i].cell(row=j, column=24).value, float) == False:
                            result_format_grid.append([ws_ippan[i].title, j, 24, MESSAGE[172][0], MESSAGE[172][1], MESSAGE[172][2], MESSAGE[172][3], MESSAGE[172][4]])
                            add_fill_to_cell(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=j, column=24, fill=fill)
                            add_comment_to_cell(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=j, column=24, message=MESSAGE[172][3]+MESSAGE[172][4])
                        
                    ### セルY20: 事業所の産業区分について形式が正しいことをチェックする。
                    if ws_ippan[i].cell(row=j, column=25).value is None:
                        pass
                    else:
                        if split_name_code(ws_ippan[i].cell(row=j, column=25).value)[-1].isdecimal() == False:
                            result_format_grid.append([ws_ippan[i].title, j, 25, MESSAGE[173][0], MESSAGE[173][1], MESSAGE[173][2], MESSAGE[173][3], MESSAGE[173][4]])
                            add_fill_to_cell(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=j, column=25, fill=fill)
                            add_comment_to_cell(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=j, column=25, message=MESSAGE[173][3]+MESSAGE[173][4])
                        
                    ### セルZ20: 地下空間の利用形態について形式が正しいことをチェックする。
                    if ws_ippan[i].cell(row=j, column=26).value is None:
                        pass
                    else:
                        if split_name_code(ws_ippan[i].cell(row=j, column=26).value)[-1].isdecimal() == False:
                            result_format_grid.append([ws_ippan[i].title, j, 26, MESSAGE[174][0], MESSAGE[174][1], MESSAGE[174][2], MESSAGE[174][3], MESSAGE[174][4]])
                            add_fill_to_cell(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=j, column=26, fill=fill)
                            add_comment_to_cell(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=j, column=26, message=MESSAGE[174][3]+MESSAGE[174][4])
                        
                    ### セルAA20: 備考について形式が正しいことをチェックする。
    
        #######################################################################
        #######################################################################
        ### EXCELセルデータ範囲チェック処理（3000）
        ### (1)セルB7からセルI7について範囲が正しいことをチェックする。
        ### (2)チェック結果リストにセルの行、列とメッセージを追加する。
        ### (3)IPPANワークシートとRESULTワークシートのセルに背景赤色の塗りつぶしをセットする。
        ### 範囲チェックでは、値がセットされている場合のみチェックを行う。
        ### 必須チェックは別途必須チェックで行うためである。
        ### 範囲チェックでは、形式が正しい場合のみチェックを行う。
        ### 形式チェックは別途形式チェックで行うためである。
        ### 例えば、面積などの数値について、float()で例外とならないように、int、floatの場合のみチェックする。
        ### 範囲チェックでは、数値の下限と上限のチェックを行う。
        ### 範囲チェックでは、DBとの突合チェックに該当するチェックは行わない。
        ### DBとの突合チェックは別途突合チェックで行うためである。
        #######################################################################
        #######################################################################
        print_log('[INFO] P0300ExcelUpload.index_view()関数 STEP 18/34.', 'INFO')
        ### for i, _ in enumerate(ws_ippan):
            ### 7行目
            ### セルB7: 都道府県について範囲が正しいことをチェックする。
            ### セルC7: 市区町村について範囲が正しいことをチェックする。
            ### セルD7: 水害発生月日について範囲が正しいことをチェックする。
            ### セルE7: 水害終了月日について範囲が正しいことをチェックする。
            ### セルF7: 水害原因1について範囲が正しいことをチェックする。
            ### セルG7: 水害原因2について範囲が正しいことをチェックする。
            ### セルH7: 水害原因3について範囲が正しいことをチェックする。
            ### セルI7: 水害区域番号について範囲が正しいことをチェックする。

        #######################################################################
        ### EXCELセルデータ範囲チェック処理（3010）
        ### (1)セルB10からセルF10について範囲が正しいことをチェックする。
        ### (2)チェック結果リストにセルの行、列とメッセージを追加する。
        ### (3)IPPANワークシートとRESULTワークシートのセルに背景赤色の塗りつぶしをセットする。
        #######################################################################
        print_log('[INFO] P0300ExcelUpload.index_view()関数 STEP 19/34.', 'INFO')
        ### for i, _ in enumerate(ws_ippan):
            ### 10行目
            ### セルB10: 水系・沿岸名について範囲が正しいことをチェックする。
            ### セルC10: 水系種別について範囲が正しいことをチェックする。
            ### セルD10: 河川・海岸名について範囲が正しいことをチェックする。
            ### セルE10: 河川種別について範囲が正しいことをチェックする。
            ### セルF10: 地盤勾配区分について範囲が正しいことをチェックする。

        #######################################################################
        ### EXCELセルデータ範囲チェック処理（3020）
        ### (1)セルB14からセルJ14について範囲が正しいことをチェックする。
        ### (2)チェック結果リストにセルの行、列とメッセージを追加する。
        ### (3)IPPANワークシートとRESULTワークシートのセルに背景赤色の塗りつぶしをセットする。
        #######################################################################
        print_log('[INFO] P0300ExcelUpload.index_view()関数 STEP 20/34.', 'INFO')
        for i, _ in enumerate(ws_ippan):
            ### 14行目
            ### セルB14: 水害区域面積の宅地について範囲が正しいことをチェックする。
            if ws_ippan[i].cell(row=14, column=2).value is None:
                pass
            else:
                if isinstance(ws_ippan[i].cell(row=14, column=2).value, int) == True or \
                    isinstance(ws_ippan[i].cell(row=14, column=2).value, float) == True:
                    if float(ws_ippan[i].cell(row=14, column=2).value) < 0:
                        result_range_list.append([ws_ippan[i].title, 14, 2, MESSAGE[213][0], MESSAGE[213][1], MESSAGE[213][2], MESSAGE[213][3], MESSAGE[213][4]])
                        add_fill_to_cell(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=14, column=2, fill=fill)
                        add_comment_to_cell(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=14, column=2, message=MESSAGE[213][3]+MESSAGE[213][4])
                
            ### セルC14: 水害区域面積の農地について範囲が正しいことをチェックする。
            if ws_ippan[i].cell(row=14, column=3).value is None:
                pass
            else:
                if isinstance(ws_ippan[i].cell(row=14, column=3).value, int) == True or \
                    isinstance(ws_ippan[i].cell(row=14, column=3).value, float) == True:
                    if float(ws_ippan[i].cell(row=14, column=3).value) < 0:
                        result_range_list.append([ws_ippan[i].title, 14, 3, MESSAGE[214][0], MESSAGE[214][1], MESSAGE[214][2], MESSAGE[214][3], MESSAGE[214][4]])
                        add_fill_to_cell(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=14, column=3, fill=fill)
                        add_comment_to_cell(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=14, column=3, message=MESSAGE[214][3]+MESSAGE[214][4])
    
            ### セルD14: 水害区域面積の地下について範囲が正しいことをチェックする。
            if ws_ippan[i].cell(row=14, column=4).value is None:
                pass
            else:
                if isinstance(ws_ippan[i].cell(row=14, column=4).value, int) == True or \
                    isinstance(ws_ippan[i].cell(row=14, column=4).value, float) == True:
                    if float(ws_ippan[i].cell(row=14, column=4).value) < 0:
                        result_range_list.append([ws_ippan[i].title, 14, 4, MESSAGE[215][0], MESSAGE[215][1], MESSAGE[215][2], MESSAGE[215][3], MESSAGE[215][4]])
                        add_fill_to_cell(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=14, column=4, fill=fill)
                        add_comment_to_cell(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=14, column=4, message=MESSAGE[215][3]+MESSAGE[215][4])
                
            ### セルF14: 工種について範囲が正しいことをチェックする。
    
            ### セルH14: 農作物被害額について範囲が正しいことをチェックする。
            if ws_ippan[i].cell(row=14, column=8).value is None:
                pass
            else:
                if isinstance(ws_ippan[i].cell(row=14, column=8).value, int) == True or \
                    isinstance(ws_ippan[i].cell(row=14, column=8).value, float) == True:
                    if float(ws_ippan[i].cell(row=14, column=8).value) < 0:
                        result_range_list.append([ws_ippan[i].title, 14, 8, MESSAGE[217][0], MESSAGE[217][1], MESSAGE[217][2], MESSAGE[217][3], MESSAGE[217][4]])
                        add_fill_to_cell(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=14, column=8, fill=fill)
                        add_comment_to_cell(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=14, column=8, message=MESSAGE[217][3]+MESSAGE[217][4])
    
            ### セルJ14: 異常気象コードについて範囲が正しいことをチェックする。

        #######################################################################
        ### EXCELセルデータ範囲チェック処理（3030）
        ### (1)セルB20からセルAA20について範囲が正しいことをチェックする。
        ### (2)チェック結果リストにセルの行、列とメッセージを追加する。
        ### (3)IPPANワークシートとRESULTワークシートのセルに背景赤色の塗りつぶしをセットする。
        #######################################################################
        print_log('[INFO] P0300ExcelUpload.index_view()関数 STEP 21/34.', 'INFO')
        for i, _ in enumerate(ws_ippan):
            if ws_max_row[i] >= 20:
                for j in range(20, ws_max_row[i] + 1):
                    ### セルB20: 町丁名・大字名について範囲が正しいことをチェックする。
                    ### セルC20: 名称について範囲が正しいことをチェックする。
                    ### セルD20: 地上・地下被害の区分について範囲が正しいことをチェックする。
                    ### セルE20: 浸水土砂被害の区分について範囲が正しいことをチェックする。
    
                    ### セルF20: 被害建物棟数, 床下浸水について範囲が正しいことをチェックする。
                    if ws_ippan[i].cell(row=j, column=6).value is None:
                        pass
                    else:
                        if isinstance(ws_ippan[i].cell(row=j, column=6).value, int) == True or \
                            isinstance(ws_ippan[i].cell(row=j, column=6).value, float) == True:
                            if float(ws_ippan[i].cell(row=j, column=6).value) < 0:
                                result_range_grid.append([ws_ippan[i].title, j, 6, MESSAGE[254][0], MESSAGE[254][1], MESSAGE[254][2], MESSAGE[254][3], MESSAGE[254][4]])
                                add_fill_to_cell(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=j, column=6, fill=fill)
                                add_comment_to_cell(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=j, column=6, message=MESSAGE[254][3]+MESSAGE[254][4])
    
                    ### セルG20: 被害建物棟数, 1cm〜49cmについて範囲が正しいことをチェックする。
                    if ws_ippan[i].cell(row=j, column=7).value is None:
                        pass
                    else:
                        if isinstance(ws_ippan[i].cell(row=j, column=7).value, int) == True or \
                            isinstance(ws_ippan[i].cell(row=j, column=7).value, float) == True:
                            if float(ws_ippan[i].cell(row=j, column=7).value) < 0:
                                result_range_grid.append([ws_ippan[i].title, j, 7, MESSAGE[255][0], MESSAGE[255][1], MESSAGE[255][2], MESSAGE[255][3], MESSAGE[255][4]])
                                add_fill_to_cell(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=j, column=7, fill=fill)
                                add_comment_to_cell(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=j, column=7, message=MESSAGE[255][3]+MESSAGE[255][4])
                    
                    ### セルH20: 被害建物棟数, 50cm〜99cmについて範囲が正しいことをチェックする。
                    if ws_ippan[i].cell(row=j, column=8).value is None:
                        pass
                    else:
                        if isinstance(ws_ippan[i].cell(row=j, column=8).value, int) == True or \
                            isinstance(ws_ippan[i].cell(row=j, column=8).value, float) == True:
                            if float(ws_ippan[i].cell(row=j, column=8).value) < 0:
                                result_range_grid.append([ws_ippan[i].title, j, 8, MESSAGE[256][0], MESSAGE[256][1], MESSAGE[256][2], MESSAGE[256][3], MESSAGE[256][4]])
                                add_fill_to_cell(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=j, column=8, fill=fill)
                                add_comment_to_cell(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=j, column=8, message=MESSAGE[256][3]+MESSAGE[256][4])
    
                    ### セルI20: 被害建物棟数, 1m以上について範囲が正しいことをチェックする。
                    if ws_ippan[i].cell(row=j, column=9).value is None:
                        pass
                    else:
                        if isinstance(ws_ippan[i].cell(row=j, column=9).value, int) == True or \
                            isinstance(ws_ippan[i].cell(row=j, column=9).value, float) == True:
                            if float(ws_ippan[i].cell(row=j, column=9).value) < 0:
                                result_range_grid.append([ws_ippan[i].title, j, 9, MESSAGE[257][0], MESSAGE[257][1], MESSAGE[257][2], MESSAGE[257][3], MESSAGE[257][4]])
                                add_fill_to_cell(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=j, column=9, fill=fill)
                                add_comment_to_cell(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=j, column=9, message=MESSAGE[257][3]+MESSAGE[257][4])
    
                    ### セルJ20: 被害建物棟数, 半壊について範囲が正しいことをチェックする。
                    if ws_ippan[i].cell(row=j, column=10).value is None:
                        pass
                    else:
                        if isinstance(ws_ippan[i].cell(row=j, column=10).value, int) == True or \
                            isinstance(ws_ippan[i].cell(row=j, column=10).value, float) == True:
                            if float(ws_ippan[i].cell(row=j, column=10).value) < 0:
                                result_range_grid.append([ws_ippan[i].title, j, 10, MESSAGE[258][0], MESSAGE[258][1], MESSAGE[258][2], MESSAGE[258][3], MESSAGE[258][4]])
                                add_fill_to_cell(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=j, column=10, fill=fill)
                                add_comment_to_cell(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=j, column=10, message=MESSAGE[258][3]+MESSAGE[258][4])
    
                    ### セルK20: 被害建物棟数, 全壊・流失について範囲が正しいことをチェックする。
                    if ws_ippan[i].cell(row=j, column=11).value is None:
                        pass
                    else:
                        if isinstance(ws_ippan[i].cell(row=j, column=11).value, int) == True or \
                            isinstance(ws_ippan[i].cell(row=j, column=11).value, float) == True:
                            if float(ws_ippan[i].cell(row=j, column=11).value) < 0:
                                result_range_grid.append([ws_ippan[i].title, j, 11, MESSAGE[259][0], MESSAGE[259][1], MESSAGE[259][2], MESSAGE[259][3], MESSAGE[259][4]])
                                add_fill_to_cell(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=j, column=11, fill=fill)
                                add_comment_to_cell(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=j, column=11, message=MESSAGE[259][3]+MESSAGE[259][4])
    
                    ### セルL20: 被害建物の延床面積について範囲が正しいことをチェックする。
                    if ws_ippan[i].cell(row=j, column=12).value is None:
                        pass
                    else:
                        if isinstance(ws_ippan[i].cell(row=j, column=12).value, int) == True or \
                            isinstance(ws_ippan[i].cell(row=j, column=12).value, float) == True:
                            if float(ws_ippan[i].cell(row=j, column=12).value) < 0:
                                result_range_grid.append([ws_ippan[i].title, j, 12, MESSAGE[260][0], MESSAGE[260][1], MESSAGE[260][2], MESSAGE[260][3], MESSAGE[260][4]])
                                add_fill_to_cell(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=j, column=12, fill=fill)
                                add_comment_to_cell(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=j, column=12, message=MESSAGE[260][3]+MESSAGE[260][4])
    
                    ### セルM20: 被災世帯数について範囲が正しいことをチェックする。
                    if ws_ippan[i].cell(row=j, column=13).value is None:
                        pass
                    else:
                        if isinstance(ws_ippan[i].cell(row=j, column=13).value, int) == True or \
                            isinstance(ws_ippan[i].cell(row=j, column=13).value, float) == True:
                            if float(ws_ippan[i].cell(row=j, column=13).value) < 0:
                                result_range_grid.append([ws_ippan[i].title, j, 13, MESSAGE[261][0], MESSAGE[261][1], MESSAGE[261][2], MESSAGE[261][3], MESSAGE[261][4]])
                                add_fill_to_cell(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=j, column=13, fill=fill)
                                add_comment_to_cell(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=j, column=13, message=MESSAGE[261][3]+MESSAGE[261][4])
                    
                    ### セルN20: 被災事業所数について範囲が正しいことをチェックする。
                    if ws_ippan[i].cell(row=j, column=14).value is None:
                        pass
                    else:
                        if isinstance(ws_ippan[i].cell(row=j, column=14).value, int) == True or \
                            isinstance(ws_ippan[i].cell(row=j, column=14).value, float) == True:
                            if float(ws_ippan[i].cell(row=j, column=14).value) < 0:
                                result_range_grid.append([ws_ippan[i].title, j, 14, MESSAGE[262][0], MESSAGE[262][1], MESSAGE[262][2], MESSAGE[262][3], MESSAGE[262][4]])
                                add_fill_to_cell(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=j, column=14, fill=fill)
                                add_comment_to_cell(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=j, column=14, message=MESSAGE[262][3]+MESSAGE[262][4])
    
                    ### セルO20: 農家・漁家戸数, 床下浸水について範囲が正しいことをチェックする。
                    if ws_ippan[i].cell(row=j, column=15).value is None:
                        pass
                    else:
                        if isinstance(ws_ippan[i].cell(row=j, column=15).value, int) == True or \
                            isinstance(ws_ippan[i].cell(row=j, column=15).value, float) == True:
                            if float(ws_ippan[i].cell(row=j, column=15).value) < 0:
                                result_range_grid.append([ws_ippan[i].title, j, 15, MESSAGE[263][0], MESSAGE[263][1], MESSAGE[263][2], MESSAGE[263][3], MESSAGE[263][4]])
                                add_fill_to_cell(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=j, column=15, fill=fill)
                                add_comment_to_cell(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=j, column=15, message=MESSAGE[263][3]+MESSAGE[263][4])
    
                    ### セルP20: 農家・漁家戸数, 1cm〜49cmについて範囲が正しいことをチェックする。
                    if ws_ippan[i].cell(row=j, column=16).value is None:
                        pass
                    else:
                        if isinstance(ws_ippan[i].cell(row=j, column=16).value, int) == True or \
                            isinstance(ws_ippan[i].cell(row=j, column=16).value, float) == True:
                            if float(ws_ippan[i].cell(row=j, column=16).value) < 0:
                                result_range_grid.append([ws_ippan[i].title, j, 16, MESSAGE[264][0], MESSAGE[264][1], MESSAGE[264][2], MESSAGE[264][3], MESSAGE[264][4]])
                                add_fill_to_cell(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=j, column=16, fill=fill)
                                add_comment_to_cell(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=j, column=16, message=MESSAGE[264][3]+MESSAGE[264][4])
    
                    ### セルQ20: 農家・漁家戸数, 50cm〜99cmについて範囲が正しいことをチェックする。
                    if ws_ippan[i].cell(row=j, column=17).value is None:
                        pass
                    else:
                        if isinstance(ws_ippan[i].cell(row=j, column=17).value, int) == True or \
                            isinstance(ws_ippan[i].cell(row=j, column=17).value, float) == True:
                            if float(ws_ippan[i].cell(row=j, column=17).value) < 0:
                                result_range_grid.append([ws_ippan[i].title, j, 17, MESSAGE[265][0], MESSAGE[265][1], MESSAGE[265][2], MESSAGE[265][3], MESSAGE[265][4]])
                                add_fill_to_cell(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=j, column=17, fill=fill)
                                add_comment_to_cell(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=j, column=17, message=MESSAGE[265][3]+MESSAGE[265][4])
    
                    ### セルR20: 農家・漁家戸数, 1m以上・半壊について範囲が正しいことをチェックする。
                    if ws_ippan[i].cell(row=j, column=18).value is None:
                        pass
                    else:
                        if isinstance(ws_ippan[i].cell(row=j, column=18).value, int) == True or \
                            isinstance(ws_ippan[i].cell(row=j, column=18).value, float) == True:
                            if float(ws_ippan[i].cell(row=j, column=18).value) < 0:
                                result_range_grid.append([ws_ippan[i].title, j, 18, MESSAGE[266][0], MESSAGE[266][1], MESSAGE[266][2], MESSAGE[266][3], MESSAGE[266][4]])
                                add_fill_to_cell(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=j, column=18, fill=fill)
                                add_comment_to_cell(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=j, column=18, message=MESSAGE[266][3]+MESSAGE[266][4])
    
                    ### セルS20: 農家・漁家戸数, 全壊・流失について範囲が正しいことをチェックする。
                    if ws_ippan[i].cell(row=j, column=19).value is None:
                        pass
                    else:
                        if isinstance(ws_ippan[i].cell(row=j, column=19).value, int) == True or \
                            isinstance(ws_ippan[i].cell(row=j, column=19).value, float) == True:
                            if float(ws_ippan[i].cell(row=j, column=19).value) < 0:
                                result_range_grid.append([ws_ippan[i].title, j, 19, MESSAGE[267][0], MESSAGE[267][1], MESSAGE[267][2], MESSAGE[267][3], MESSAGE[267][4]])
                                add_fill_to_cell(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=j, column=19, fill=fill)
                                add_comment_to_cell(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=j, column=19, message=MESSAGE[267][3]+MESSAGE[267][4])
    
                    ### セルT20: 事業所従業者数, 床下浸水について範囲が正しいことをチェックする。
                    if ws_ippan[i].cell(row=j, column=20).value is None:
                        pass
                    else:
                        if isinstance(ws_ippan[i].cell(row=j, column=20).value, int) == True or \
                            isinstance(ws_ippan[i].cell(row=j, column=20).value, float) == True:
                            if float(ws_ippan[i].cell(row=j, column=20).value) < 0:
                                result_range_grid.append([ws_ippan[i].title, j, 20, MESSAGE[268][0], MESSAGE[268][1], MESSAGE[268][2], MESSAGE[268][3], MESSAGE[268][4]])
                                add_fill_to_cell(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=j, column=20, fill=fill)
                                add_comment_to_cell(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=j, column=20, message=MESSAGE[268][3]+MESSAGE[268][4])
    
                    ### セルU20: 事業所従業者数, 1cm〜49cmについて範囲が正しいことをチェックする。
                    if ws_ippan[i].cell(row=j, column=21).value is None:
                        pass
                    else:
                        if isinstance(ws_ippan[i].cell(row=j, column=21).value, int) == True or \
                            isinstance(ws_ippan[i].cell(row=j, column=21).value, float) == True:
                            if float(ws_ippan[i].cell(row=j, column=21).value) < 0:
                                result_range_grid.append([ws_ippan[i].title, j, 21, MESSAGE[269][0], MESSAGE[269][1], MESSAGE[269][2], MESSAGE[269][3], MESSAGE[269][4]])
                                add_fill_to_cell(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=j, column=21, fill=fill)
                                add_comment_to_cell(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=j, column=21, message=MESSAGE[269][3]+MESSAGE[269][4])
    
                    ### セルV20: 事業所従業者数, 50cm〜99cmについて範囲が正しいことをチェックする。
                    if ws_ippan[i].cell(row=j, column=22).value is None:
                        pass
                    else:
                        if isinstance(ws_ippan[i].cell(row=j, column=22).value, int) == True or \
                            isinstance(ws_ippan[i].cell(row=j, column=22).value, float) == True:
                            if float(ws_ippan[i].cell(row=j, column=22).value) < 0:
                                result_range_grid.append([ws_ippan[i].title, j, 22, MESSAGE[270][0], MESSAGE[270][1], MESSAGE[270][2], MESSAGE[270][3], MESSAGE[270][4]])
                                add_fill_to_cell(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=j, column=22, fill=fill)
                                add_comment_to_cell(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=j, column=22, message=MESSAGE[270][3]+MESSAGE[270][4])
    
                    ### セルW20: 事業所従業者数, 1m以上・半壊について範囲が正しいことをチェックする。
                    if ws_ippan[i].cell(row=j, column=23).value is None:
                        pass
                    else:
                        if isinstance(ws_ippan[i].cell(row=j, column=23).value, int) == True or \
                            isinstance(ws_ippan[i].cell(row=j, column=23).value, float) == True:
                            if float(ws_ippan[i].cell(row=j, column=23).value) < 0:
                                result_range_grid.append([ws_ippan[i].title, j, 23, MESSAGE[271][0], MESSAGE[271][1], MESSAGE[271][2], MESSAGE[271][3], MESSAGE[271][4]])
                                add_fill_to_cell(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=j, column=23, fill=fill)
                                add_comment_to_cell(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=j, column=23, message=MESSAGE[271][3]+MESSAGE[271][4])
    
                    ### セルX20: 事業所従業者数, 全壊・流失について範囲が正しいことをチェックする。
                    if ws_ippan[i].cell(row=j, column=24).value is None:
                        pass
                    else:
                        if isinstance(ws_ippan[i].cell(row=j, column=24).value, int) == True or \
                            isinstance(ws_ippan[i].cell(row=j, column=24).value, float) == True:
                            if float(ws_ippan[i].cell(row=j, column=24).value) < 0:
                                result_range_grid.append([ws_ippan[i].title, j, 24, MESSAGE[272][0], MESSAGE[272][1], MESSAGE[272][2], MESSAGE[272][3], MESSAGE[272][4]])
                                add_fill_to_cell(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=j, column=24, fill=fill)
                                add_comment_to_cell(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=j, column=24, message=MESSAGE[272][3]+MESSAGE[272][4])
    
                    ### セルY20: 事業所の産業区分について範囲が正しいことをチェックする。
                    ### セルZ20: 地下空間の利用形態について範囲が正しいことをチェックする。
                    ### セルAA20: 備考について範囲が正しいことをチェックする。
    
        #######################################################################
        #######################################################################
        ### EXCELセルデータ相関チェック処理（4000）
        ### (1)セルB7からセルI7について他項目との相関関係が正しいことをチェックする。
        ### (2)チェック結果リストにセルの行、列とメッセージを追加する。
        ### (3)IPPANワークシートとRESULTワークシートのセルに背景赤色の塗りつぶしをセットする。
        #######################################################################
        #######################################################################
        print_log('[INFO] P0300ExcelUpload.index_view()関数 STEP 22/34.', 'INFO')
        ### for i, _ in enumerate(ws_ippan):
            ### 7行目
            ### セルB7: 都道府県について他項目との相関関係が正しいことをチェックする。
            ### if ws_ippan[i].cell(row=7, column=2).value is None:
            ###     result_correlate_list.append([ws_ippan[i].title, 7, 2, MESSAGE[][0], MESSAGE[][1], MESSAGE[][2], MESSAGE[][3], MESSAGE[][4]])
            ### 都道府県名に対して無効な市区町村名が入力されていないか。
            
            ### セルC7: 市区町村について他項目との相関関係が正しいことをチェックする。
            ### if ws_ippan[i].cell(row=7, column=3).value is None:
            ###     result_correlate_list.append([ws_ippan[i].title, 7, 3, MESSAGE[][0], MESSAGE[][1], MESSAGE[][2], MESSAGE[][3], MESSAGE[][4]])
            ### 都道府県名に対して無効な市区町村名が入力されていないか。
    
            ### セルD7: 水害発生月日について他項目との相関関係が正しいことをチェックする。
            ### if ws_ippan[i].cell(row=7, column=4).value is None:
            ###     result_correlate_list.append([ws_ippan[i].title, 7, 4, MESSAGE[][0], MESSAGE[][1], MESSAGE[][2], MESSAGE[][3], MESSAGE[][4]])
            ### 水害発生月日に対して無効な水害終了月日が入力されていないか。
    
            ### セルE7: 水害終了月日について他項目との相関関係が正しいことをチェックする。
            ### if ws_ippan[i].cell(row=7, column=5).value is None:
            ###     result_correlate_list.append([ws_ippan[i].title, 7, 5, MESSAGE[][0], MESSAGE[][1], MESSAGE[][2], MESSAGE[][3], MESSAGE[][4]])
            ### 水害発生月日に対して無効な水害終了月日が入力されていないか。
    
            ### セルF7: 水害原因1について他項目との相関関係が正しいことをチェックする。
            ### if ws_ippan[i].cell(row=7, column=6).value is None:
            ###     result_correlate_list.append([ws_ippan[i].title, 7, 6, MESSAGE[][0], MESSAGE[][1], MESSAGE[][2], MESSAGE[][3], MESSAGE[][4]])
    
            ### セルG7: 水害原因2について他項目との相関関係が正しいことをチェックする。
            ### if ws_ippan[i].cell(row=7, column=7).value is None:
            ###     result_correlate_list.append([ws_ippan[i].title, 7, 7, MESSAGE[][0], MESSAGE[][1], MESSAGE[][2], MESSAGE[][3], MESSAGE[][4]])
    
            ### セルH7: 水害原因3について他項目との相関関係が正しいことをチェックする。
            ### if ws_ippan[i].cell(row=7, column=8).value is None:
            ###     result_correlate_list.append([ws_ippan[i].title, 7, 8, MESSAGE[][0], MESSAGE[][1], MESSAGE[][2], MESSAGE[][3], MESSAGE[][4]])
    
            ### セルI7: 水害区域番号について他項目との相関関係が正しいことをチェックする。
            ### if ws_ippan[i].cell(row=7, column=9).value is None:
            ###     result_correlate_list.append([ws_ippan[i].title, 7, 9, MESSAGE[][0], MESSAGE[][1], MESSAGE[][2], MESSAGE[][3], MESSAGE[][4]])
    
        #######################################################################
        ### EXCELセルデータ相関チェック処理（4010）
        ### (1)セルB10からセルF10について他項目との相関関係が正しいことをチェックする。
        ### (2)チェック結果リストにセルの行、列とメッセージを追加する。
        ### (3)IPPANワークシートとRESULTワークシートのセルに背景赤色の塗りつぶしをセットする。
        #######################################################################
        print_log('[INFO] P0300ExcelUpload.index_view()関数 STEP 23/34.', 'INFO')
        for i, _ in enumerate(ws_ippan):
            ### 10行目
            ### セルB10: 水系・沿岸名について他項目との相関関係が正しいことをチェックする。
            ### if ws_ippan[i].cell(row=10, column=2).value is None:
            ###     result_correlate_list.append([ws_ippan[i].title, 10, 2, MESSAGE[][0], MESSAGE[][1], MESSAGE[][2], MESSAGE[][3], MESSAGE[][4]])
            ### 水系種別に対して無効な水系・沿岸名が入力されていないか。
            ### 水系・沿岸名に無名水系以外の水系に対して、無効な水系の文字が含まれていないか。
            
            ### セルC10: 水系種別について他項目との相関関係が正しいことをチェックする。
            ### if ws_ippan[i].cell(row=10, column=3).value is None:
            ###     result_correlate_list.append([ws_ippan[i].title, 10, 3, MESSAGE[][0], MESSAGE[][1], MESSAGE[][2], MESSAGE[][3], MESSAGE[][4]])
    
            ### セルD10: 河川・海岸名について他項目との相関関係が正しいことをチェックする。
            ### if ws_ippan[i].cell(row=10, column=4).value is None:
            ###     result_correlate_list.append([ws_ippan[i].title, 10, 4, MESSAGE[][0], MESSAGE[][1], MESSAGE[][2], MESSAGE[][3], MESSAGE[][4]])
    
            ### セルE10: 河川種別について他項目との相関関係が正しいことをチェックする。
            ### if ws_ippan[i].cell(row=10, column=5).value is None:
            ###     result_correlate_list.append([ws_ippan[i].title, 10, 5, MESSAGE[][0], MESSAGE[][1], MESSAGE[][2], MESSAGE[][3], MESSAGE[][4]])
            
            ### セルC10: 水系種別について他項目との相関関係が正しいことをチェックする。
            ### 水系種別が「1:一級」のときに、河川種別に選択範囲外の不正な文字が入力されていないか。
            ### 正しい選択範囲は、「1:直轄」「2:指定」「4:準用」「5:普通」である。
            if ws_ippan[i].cell(row=10, column=3).value == '1' or ws_ippan[i].cell(row=10, column=3).value == '一級':
                if ws_ippan[i].cell(row=10, column=5).value == '1' or ws_ippan[i].cell(row=10, column=5).value == '直轄' or \
                    ws_ippan[i].cell(row=10, column=5).value == '2' or ws_ippan[i].cell(row=10, column=5).value == '指定' or \
                    ws_ippan[i].cell(row=10, column=5).value == '4' or ws_ippan[i].cell(row=10, column=5).value == '準用' or \
                    ws_ippan[i].cell(row=10, column=5).value == '5' or ws_ippan[i].cell(row=10, column=5).value == '普通':
                    pass
                else:
                    pass
                    ### result_correlate_list.append([ws_ippan[i].title, 10, 5, MESSAGE[300][0], MESSAGE[300][1], MESSAGE[300][2], MESSAGE[300][3], MESSAGE[300][4]])
                    ### ws_ippan[i].cell(row=10, column=5).fill = fill
                    ### ws_result[i].cell(row=10, column=5).fill = fill
                    ### print('水系種別が「1:一級」のときに、河川種別に選択範囲外の不正な文字が入力されています。', flush=True)
                    
            ### セルC10: 水系種別について他項目との相関関係が正しいことをチェックする。
            ### 水系種別が「2:二級」のときに、河川種別に選択範囲外の不正な文字が入力されていないか。
            ### 正しい選択範囲は、「3:二級」「4:準用」「5:普通」である。
            if ws_ippan[i].cell(row=10, column=3).value == '2' or ws_ippan[i].cell(row=10, column=3).value == '二級':
                if ws_ippan[i].cell(row=10, column=5).value == '3' or ws_ippan[i].cell(row=10, column=5).value == '二級' or \
                    ws_ippan[i].cell(row=10, column=5).value == '4' or ws_ippan[i].cell(row=10, column=5).value == '準用' or \
                    ws_ippan[i].cell(row=10, column=5).value == '5' or ws_ippan[i].cell(row=10, column=5).value == '普通':
                    pass
                else:
                    pass
                    ### result_correlate_list.append([ws_ippan[i].title, 10, 5, MESSAGE[301][0], MESSAGE[301][1], MESSAGE[301][2], MESSAGE[301][3], MESSAGE[301][4]])
                    ### ws_ippan[i].cell(row=10, column=5).fill = fill
                    ### ws_result[i].cell(row=10, column=5).fill = fill
                    ### print('水系種別が「2:二級」のときに、河川種別に選択範囲外の不正な文字が入力されています。', flush=True)
        
            ### セルC10: 水系種別について他項目との相関関係が正しいことをチェックする。
            ### 水系種別が「3:準用」のときに、河川種別に選択範囲外の不正な文字が入力されていないか。
            ### 正しい選択範囲は、「4:準用」「5:普通」である。
            if ws_ippan[i].cell(row=10, column=3).value == '3' or ws_ippan[i].cell(row=10, column=3).value == '準用':
                if ws_ippan[i].cell(row=10, column=5).value == '4' or ws_ippan[i].cell(row=10, column=5).value == '準用' or \
                    ws_ippan[i].cell(row=10, column=5).value == '5' or ws_ippan[i].cell(row=10, column=5).value == '普通':
                    pass
                else:
                    pass
                    ### result_correlate_list.append([ws_ippan[i].title, 10, 5, MESSAGE[302][0], MESSAGE[302][1], MESSAGE[302][2], MESSAGE[302][3], MESSAGE[302][4]])
                    ### ws_ippan[i].cell(row=10, column=5).fill = fill
                    ### ws_result[i].cell(row=10, column=5).fill = fill
                    ### print('水系種別が「3:準用」のときに、河川種別に選択範囲外の不正な文字が入力されています。', flush=True)
        
            ### セルC10: 水系種別について他項目との相関関係が正しいことをチェックする。
            ### 水系種別が「4:普通」のときに、河川種別に選択範囲外の不正な文字が入力されていないか。
            ### 正しい選択範囲は、「5:普通」である。
            if ws_ippan[i].cell(row=10, column=3).value == '4' or ws_ippan[i].cell(row=10, column=3).value == '普通':
                if ws_ippan[i].cell(row=10, column=5).value == '5' or ws_ippan[i].cell(row=10, column=5).value == '普通':
                    pass
                else:
                    pass
                    ### result_correlate_list.append([ws_ippan[i].title, 10, 5, MESSAGE[303][0], MESSAGE[303][1], MESSAGE[303][2], MESSAGE[303][3], MESSAGE[303][4]])
                    ### ws_ippan[i].cell(row=10, column=5).fill = fill
                    ### ws_result[i].cell(row=10, column=5).fill = fill
                    ### print('水系種別が「4:普通」のときに、河川種別に選択範囲外の不正な文字が入力されています。', flush=True)
        
            ### セルC10: 水系種別について他項目との相関関係が正しいことをチェックする。
            ### 水系種別が「5:沿岸」のときに、河川種別に選択範囲外の不正な文字が入力されていないか。
            ### 正しい選択範囲は、「6:海岸」である。
            if ws_ippan[i].cell(row=10, column=3).value == '5' or ws_ippan[i].cell(row=10, column=3).value == '沿岸':
                if ws_ippan[i].cell(row=10, column=5).value == '6' or ws_ippan[i].cell(row=10, column=5).value == '海岸':
                    pass
                else:
                    pass
                    ### result_correlate_list.append([ws_ippan[i].title, 10, 5, MESSAGE[304][0], MESSAGE[304][1], MESSAGE[304][2], MESSAGE[304][3], MESSAGE[304][4]])
                    ### ws_ippan[i].cell(row=10, column=5).fill = fill
                    ### ws_result[i].cell(row=10, column=5).fill = fill
                    ### print('水系種別が「5:沿岸」のときに、河川種別に選択範囲外の不正な文字が入力されています。', flush=True)
        
            ### セルC10: 水系種別について他項目との相関関係が正しいことをチェックする。
            ### 水系種別が「6:河川海岸以外」のときに、河川種別に選択範囲外の不正な文字が入力されていないか。
            ### 正しい選択範囲は、「7:河川海岸以外」である。
            if ws_ippan[i].cell(row=10, column=3).value == '6' or ws_ippan[i].cell(row=10, column=3).value == '河川海岸以外':
                if ws_ippan[i].cell(row=10, column=5).value == '7' or ws_ippan[i].cell(row=10, column=5).value == '河川海岸以外':
                    pass
                else:
                    pass
                    ### result_correlate_list.append([ws_ippan[i].title, 10, 5, MESSAGE[305][0], MESSAGE[305][1], MESSAGE[305][2], MESSAGE[305][3], MESSAGE[305][4]])
                    ### ws_ippan[i].cell(row=10, column=5).fill = fill
                    ### ws_result[i].cell(row=10, column=5).fill = fill
                    ### print('水系種別が「6:河川海岸以外」のときに、河川種別に選択範囲外の不正な文字が入力されています。', flush=True)
                    
            ### セルF10: 地盤勾配区分について他項目との相関関係が正しいことをチェックする。
            ### if ws_ippan[i].cell(row=10, column=6).value is None:
            ###     result_correlate_list.append([ws_ippan[i].title, 10, 6, MESSAGE[][0], MESSAGE[][1], MESSAGE[][2], MESSAGE[][3], MESSAGE[][4]])
    
        #######################################################################
        ### EXCELセルデータ相関チェック処理（4020）
        ### (1)セルB14からセルJ14について他項目との相関関係が正しいことをチェックする。
        ### (2)チェック結果リストにセルの行、列とメッセージを追加する。
        ### (3)IPPANワークシートとRESULTワークシートのセルに背景赤色の塗りつぶしをセットする。
        #######################################################################
        print_log('[INFO] P0300ExcelUpload.index_view()関数 STEP 24/34.', 'INFO')
        for i, _ in enumerate(ws_ippan):
            ### 14行目
            ### セルB14: 水害区域面積の宅地について他項目との相関関係が正しいことをチェックする。
            ### if ws_ippan[i].cell(row=14, column=2).value is None:
            ###     result_correlate_list.append([ws_ippan[i].title, 14, 2, MESSAGE[][0], MESSAGE[][1], MESSAGE[][2], MESSAGE[][3], MESSAGE[][4]])
            
            ### セルC14: 水害区域面積の農地について他項目との相関関係が正しいことをチェックする。
            ### if ws_ippan[i].cell(row=14, column=3).value is None:
            ###     result_correlate_list.append([ws_ippan[i].title, 14, 3, MESSAGE[][0], MESSAGE[][1], MESSAGE[][2], MESSAGE[][3], MESSAGE[][4]])
            
            ### セルD14: 水害区域面積の地下について他項目との相関関係が正しいことをチェックする。
            ### if ws_ippan[i].cell(row=14, column=4).value is None:
            ###     result_correlate_list.append([ws_ippan[i].title, 14, 4, MESSAGE[][0], MESSAGE[][1], MESSAGE[][2], MESSAGE[][3], MESSAGE[][4]])
            
            ### 地上・地下被害の区分が「1」のときに、水害区域面積の宅地または水害区域面積の農地が入力されているか。
            ### 地上・地下被害の区分が「2上」のときに、水害区域面積の宅地または水害区域面積の農地が入力されているか。
            ### 地上・地下被害の区分が「2下」のときに、水害区域面積の地下が入力されているか。
            ### 地上・地下被害の区分が「3」のときに、水害区域面積の地下が入力されているか。
            ### 水害区域面積の宅地、農地、地下のいずれかに入力されているか。
            
            ### セルF14: 工種について他項目との相関関係が正しいことをチェックする。
            ### if ws_ippan[i].cell(row=14, column=6).value is None:
            ###     result_correlate_list.append([ws_ippan[i].title, 14, 6, MESSAGE[][0], MESSAGE[][1], MESSAGE[][2], MESSAGE[][3], MESSAGE[][4]])
            
            ### セルF14: 工種について他項目との相関関係が正しいことをチェックする。
            ### 水系種別が「1:一級」「2:二級」「3:準用」「4:普通」のときに、工種に選択範囲外の不正な文字が入力されていないか。
            ### 正しい選択範囲は、「1:河川」である。
            if ws_ippan[i].cell(row=10, column=3).value == '1' or ws_ippan[i].cell(row=10, column=3).value == '一級' or \
                ws_ippan[i].cell(row=10, column=3).value == '2' or ws_ippan[i].cell(row=10, column=3).value == '二級' or \
                ws_ippan[i].cell(row=10, column=3).value == '3' or ws_ippan[i].cell(row=10, column=3).value == '準用' or \
                ws_ippan[i].cell(row=10, column=3).value == '4' or ws_ippan[i].cell(row=10, column=3).value == '普通':
                if ws_ippan[i].cell(row=14, column=6).value == '1' or ws_ippan[i].cell(row=14, column=6).value == '河川':
                    pass
                else:
                    pass
                    ### result_correlate_list.append([ws_ippan[i].title, 14, 6, MESSAGE[306][0], MESSAGE[306][1], MESSAGE[306][2], MESSAGE[306][3], MESSAGE[306][4]])
                    ### ws_ippan[i].cell(row=14, column=6).fill = fill
                    ### ws_result[i].cell(row=14, column=6).fill = fill
                    ### print('水系種別が「1:一級」「2:二級」「3:準用」「4:普通」のときに、工種に選択範囲外の不正な文字が入力されています。', flush=True)
        
            ### セルF14: 工種について他項目との相関関係が正しいことをチェックする。
            ### 水系種別が「5:沿岸」のときに、工種に選択範囲外の不正な文字が入力されていないか。
            ### 正しい選択範囲は、「2:海岸」である。
            if ws_ippan[i].cell(row=10, column=3).value == '5' or ws_ippan[i].cell(row=10, column=3).value == '沿岸':
                if ws_ippan[i].cell(row=14, column=6).value == '2' or ws_ippan[i].cell(row=14, column=6).value == '海岸':
                    pass
                else:
                    pass
                    ### result_correlate_list.append([ws_ippan[i].title, 14, 6, MESSAGE[307][0], MESSAGE[307][1], MESSAGE[307][2], MESSAGE[307][3], MESSAGE[307][4]])
                    ### ws_ippan[i].cell(row=14, column=6).fill = fill
                    ### ws_result[i].cell(row=14, column=6).fill = fill
                    ### print('水系種別が「5:沿岸」のときに、工種に選択範囲外の不正な文字が入力されています。', flush=True)
        
            ### セルF14: 工種について他項目との相関関係が正しいことをチェックする。
            ### 水系種別が「6:河川海岸以外」のときに、工種に選択範囲外の不正な文字が入力されていないか。
            ### 正しい選択範囲は、「3:河川海岸以外」である。
            if ws_ippan[i].cell(row=10, column=3).value == '6' or ws_ippan[i].cell(row=10, column=3).value == '河川海岸以外':
                if ws_ippan[i].cell(row=14, column=6).value == '3' or ws_ippan[i].cell(row=14, column=6).value == '河川海岸以外':
                    pass
                else:
                    pass
                    ### result_correlate_list.append([ws_ippan[i].title, 14, 6, MESSAGE[308][0], MESSAGE[308][1], MESSAGE[308][2], MESSAGE[308][3], MESSAGE[308][4]])
                    ### ws_ippan[i].cell(row=14, column=6).fill = fill
                    ### ws_result[i].cell(row=14, column=6).fill = fill
                    ### print('水系種別が「6:河川海岸以外」のときに、工種に選択範囲外の不正な文字が入力されています。', flush=True)
        
            ### セルF14: 工種について他項目との相関関係が正しいことをチェックする。
            ### 水害原因1が「10:破堤」「20:有堤部溢水」「30:無堤部溢水」「40:内水」のときに、工種に選択範囲外の不正な文字が入力されていないか。
            ### 正しい選択範囲は、「1:河川」である。
            if ws_ippan[i].cell(row=7, column=6).value == '10' or ws_ippan[i].cell(row=7, column=6).value == '破堤' or \
                ws_ippan[i].cell(row=7, column=6).value == '20' or ws_ippan[i].cell(row=7, column=6).value == '有堤部溢水' or \
                ws_ippan[i].cell(row=7, column=6).value == '30' or ws_ippan[i].cell(row=7, column=6).value == '無堤部溢水' or \
                ws_ippan[i].cell(row=7, column=6).value == '40' or ws_ippan[i].cell(row=7, column=6).value == '内水':
                if ws_ippan[i].cell(row=14, column=6).value == '1' or ws_ippan[i].cell(row=14, column=6).value == '河川':
                    pass
                else:
                    pass
                    ### result_correlate_list.append([ws_ippan[i].title, 14, 6, MESSAGE[309][0], MESSAGE[309][1], MESSAGE[309][2], MESSAGE[309][3], MESSAGE[309][4]])
                    ### ws_ippan[i].cell(row=14, column=6).fill = fill
                    ### ws_result[i].cell(row=14, column=6).fill = fill
                    ### print('水害原因1が「10:破堤」「20:有堤部溢水」「30:無堤部溢水」「40:内水」のときに、工種に選択範囲外の不正な文字が入力されています。', flush=True)
        
            ### セルF14: 工種について他項目との相関関係が正しいことをチェックする。
            ### 水害原因1が「50:窪地内水」「80:地すべり」「90:急傾斜地崩壊水」のときに、工種に選択範囲外の不正な文字が入力されていないか。
            ### 正しい選択範囲は、「3:河川海岸以外」である。
            if ws_ippan[i].cell(row=7, column=6).value == '50' or ws_ippan[i].cell(row=7, column=6).value == '窪地内水' or \
                ws_ippan[i].cell(row=7, column=6).value == '80' or ws_ippan[i].cell(row=7, column=6).value == '地すべり' or \
                ws_ippan[i].cell(row=7, column=6).value == '90' or ws_ippan[i].cell(row=7, column=6).value == '急傾斜地崩壊水':
                if ws_ippan[i].cell(row=14, column=6).value == '3' or ws_ippan[i].cell(row=14, column=6).value == '河川海岸以外':
                    pass
                else:
                    pass
                    ### result_correlate_list.append([ws_ippan[i].title, 14, 6, MESSAGE[310][0], MESSAGE[310][1], MESSAGE[310][2], MESSAGE[310][3], MESSAGE[310][4]])
                    ### ws_ippan[i].cell(row=14, column=6).fill = fill
                    ### ws_result[i].cell(row=14, column=6).fill = fill
                    ### print('水害原因1が「50:窪地内水」「80:地すべり」「90:急傾斜地崩壊水」のときに、工種に選択範囲外の不正な文字が入力されています。', flush=True)
        
            ### セルF14: 工種について他項目との相関関係が正しいことをチェックする。
            ### 水害原因1が「93:波浪」のときに、工種に選択範囲外の不正な文字が入力されていないか。
            ### 正しい選択範囲は、「2:海岸」である。
            if ws_ippan[i].cell(row=7, column=6).value == '93' or ws_ippan[i].cell(row=7, column=6).value == '波浪':
                if ws_ippan[i].cell(row=14, column=6).value == '2' or ws_ippan[i].cell(row=14, column=6).value == '海岸':
                    pass
                else:
                    pass
                    ### result_correlate_list.append([ws_ippan[i].title, 14, 6, MESSAGE[311][0], MESSAGE[311][1], MESSAGE[311][2], MESSAGE[311][3], MESSAGE[311][4]])
                    ### ws_ippan[i].cell(row=14, column=6).fill = fill
                    ### ws_result[i].cell(row=14, column=6).fill = fill
                    ### print('水害原因1が「93:波浪」のときに、工種に選択範囲外の不正な文字が入力されています。', flush=True)
        
            ### セルF14: 工種について他項目との相関関係が正しいことをチェックする。
            ### 水害原因1が「60:洗堀・流失」「91:高潮」「92:津波」のときに、工種に選択範囲外の不正な文字が入力されていないか。
            ### 正しい選択範囲は、「1:河川」「2:海岸」である。
            if ws_ippan[i].cell(row=7, column=6).value == '60' or ws_ippan[i].cell(row=7, column=6).value == '洗堀・流失' or \
                ws_ippan[i].cell(row=7, column=6).value == '91' or ws_ippan[i].cell(row=7, column=6).value == '高潮' or \
                ws_ippan[i].cell(row=7, column=6).value == '92' or ws_ippan[i].cell(row=7, column=6).value == '津波':
                if ws_ippan[i].cell(row=14, column=6).value == '1' or ws_ippan[i].cell(row=14, column=6).value == '河川' or \
                    ws_ippan[i].cell(row=14, column=6).value == '2' or ws_ippan[i].cell(row=14, column=6).value == '海岸':
                    pass
                else:
                    pass
                    ### result_correlate_list.append([ws_ippan[i].title, 14, 6, MESSAGE[312][0], MESSAGE[312][1], MESSAGE[312][2], MESSAGE[312][3], MESSAGE[312][4]])
                    ### ws_ippan[i].cell(row=14, column=6).fill = fill
                    ### ws_result[i].cell(row=14, column=6).fill = fill
                    ### print('水害原因1が「60:洗堀・流失」「91:高潮」「92:津波」のときに、工種に選択範囲外の不正な文字が入力されています。', flush=True)
        
            ### セルF14: 工種について他項目との相関関係が正しいことをチェックする。
            ### 水害原因1が「70:土石流」のときに、工種に選択範囲外の不正な文字が入力されていないか。
            ### 正しい選択範囲は、「1:河川」「3:河川海岸以外」である。
            if ws_ippan[i].cell(row=7, column=6).value == '70' or ws_ippan[i].cell(row=7, column=6).value == '土石流':
                if ws_ippan[i].cell(row=14, column=6).value == '1' or ws_ippan[i].cell(row=14, column=6).value == '河川' or \
                    ws_ippan[i].cell(row=14, column=6).value == '3' or ws_ippan[i].cell(row=14, column=6).value == '河川海岸以外':
                    pass
                else:
                    pass
                    ### result_correlate_list.append([ws_ippan[i].title, 14, 6, MESSAGE[313][0], MESSAGE[313][1], MESSAGE[313][2], MESSAGE[313][3], MESSAGE[313][4]])
                    ### ws_ippan[i].cell(row=14, column=6).fill = fill
                    ### ws_result[i].cell(row=14, column=6).fill = fill
                    ### print('水害原因1が「70:土石流」のときに、工種に選択範囲外の不正な文字が入力されています。', flush=True)
        
            ### セルH14: 農作物被害額について他項目との相関関係が正しいことをチェックする。
            ### if ws_ippan[i].cell(row=14, column=8).value is None:
            ###     result_correlate_list.append([ws_ippan[i].title, 14, 8])
            
            ### セルH14: 農作物被害額について他項目との相関関係が正しいことをチェックする。
            ### 水害区域面積の農地が入力されているときに、農作物被害額が入力されているか。
            if ws_ippan[i].cell(row=14, column=3).value is None:
                pass
            else:
                if ws_ippan[i].cell(row=14, column=8).value is None:
                    pass
                    ### result_correlate_list.append([ws_ippan[i].title, 14, 8, MESSAGE[314][0], MESSAGE[314][1], MESSAGE[314][2], MESSAGE[314][3], MESSAGE[314][4]])
                    ### ws_ippan[i].cell(row=14, column=8).fill = fill
                    ### ws_result[i].cell(row=14, column=8).fill = fill
                    ### print('水害区域面積の農地が入力されているときに、農作物被害額が入力されていません。', flush=True)
            
            ### セルJ14: 異常気象コードについて他項目との相関関係が正しいことをチェックする。
            ### if ws_ippan[i].cell(row=14, column=10).value is None:
            ###     result_correlate_list.append([ws_ippan[i].title, 14, 10, MESSAGE[][0], MESSAGE[][1], MESSAGE[][2], MESSAGE[][3], MESSAGE[][4]])
    
        #######################################################################
        ### EXCELセルデータ相関チェック処理（4030）
        ### (1)セルB20からセルAA20について他項目との相関関係が正しいことをチェックする。
        ### (2)チェック結果リストにセルの行、列とメッセージを追加する。
        ### (3)IPPANワークシートとRESULTワークシートのセルに背景赤色の塗りつぶしをセットする。
        ### TO-DO: if == ''はダミーの処理である。相関チェック処理を記述する。
        #######################################################################
        print_log('[INFO] P0300ExcelUpload.index_view()関数 STEP 25/34.', 'INFO')
        for i, _ in enumerate(ws_ippan):
            if ws_max_row[i] >= 20:
                for j in range(20, ws_max_row[i] + 1):
                    ### セルB20: 町丁名・大字名について他項目との相関関係が正しいことをチェックする。
                    if ws_ippan[i].cell(row=j, column=2).value is None:
                        pass
                        ### result_correlate_grid.append([ws_ippan[i].title, j, 2, MESSAGE[350][0], MESSAGE[350][1], MESSAGE[350][2], MESSAGE[350][3], MESSAGE[350][4]])
                        ### ws_ippan[i].cell(row=j, column=2).fill = fill
                        ### ws_result[i].cell(row=j, column=2).fill = fill
                        
                    ### セルC20: 名称について他項目との相関関係が正しいことをチェックする。
                    if ws_ippan[i].cell(row=j, column=3).value is None:
                        pass
                        ### result_correlate_grid.append([ws_ippan[i].title, j, 3, MESSAGE[351][0], MESSAGE[351][1], MESSAGE[351][2], MESSAGE[351][3], MESSAGE[351][4]])
                        ### ws_ippan[i].cell(row=j, column=3).fill = fill
                        ### ws_result[i].cell(row=j, column=3).fill = fill
                        
                    ### セルD20: 地上・地下被害の区分について他項目との相関関係が正しいことをチェックする。
                    if ws_ippan[i].cell(row=j, column=4).value is None:
                        pass
                        ### result_correlate_grid.append([ws_ippan[i].title, j, 4, MESSAGE[352][0], MESSAGE[352][1], MESSAGE[352][2], MESSAGE[352][3], MESSAGE[352][4]])
                        ### ws_ippan[i].cell(row=j, column=4).fill = fill
                        ### ws_result[i].cell(row=j, column=4).fill = fill
                        
                    ### セルE20: 浸水土砂被害の区分について他項目との相関関係が正しいことをチェックする。
                    if ws_ippan[i].cell(row=j, column=5).value is None:
                        pass
                        ### result_correlate_grid.append([ws_ippan[i].title, j, 5, MESSAGE[353][0], MESSAGE[353][1], MESSAGE[353][2], MESSAGE[353][3], MESSAGE[353][4]])
                        ### ws_ippan[i].cell(row=j, column=5).fill = fill
                        ### ws_result[i].cell(row=j, column=5).fill = fill
                        
                    ### セルF20: 被害建物棟数, 床下浸水について他項目との相関関係が正しいことをチェックする。
                    if ws_ippan[i].cell(row=j, column=6).value is None:
                        pass
                        ### result_correlate_grid.append([ws_ippan[i].title, j, 6, MESSAGE[354][0], MESSAGE[354][1], MESSAGE[354][2], MESSAGE[354][3], MESSAGE[354][4]])
                        ### ws_ippan[i].cell(row=j, column=6).fill = fill
                        ### ws_result[i].cell(row=j, column=6).fill = fill
                        
                    ### セルG20: 被害建物棟数, 1cm〜49cmについて他項目との相関関係が正しいことをチェックする。
                    if ws_ippan[i].cell(row=j, column=7).value is None:
                        pass
                        ### result_correlate_grid.append([ws_ippan[i].title, j, 7, MESSAGE[355][0], MESSAGE[355][1], MESSAGE[355][2], MESSAGE[355][3], MESSAGE[355][4]])
                        ### ws_ippan[i].cell(row=j, column=7).fill = fill
                        ### ws_result[i].cell(row=j, column=7).fill = fill
                        
                    ### セルH20: 被害建物棟数, 50cm〜99cmについて他項目との相関関係が正しいことをチェックする。
                    if ws_ippan[i].cell(row=j, column=8).value is None:
                        pass
                        ### result_correlate_grid.append([ws_ippan[i].title, j, 8, MESSAGE[356][0], MESSAGE[356][1], MESSAGE[356][2], MESSAGE[356][3], MESSAGE[356][4]])
                        ### ws_ippan[i].cell(row=j, column=8).fill = fill
                        ### ws_result[i].cell(row=j, column=8).fill = fill
                        
                    ### セルI20: 被害建物棟数, 1m以上について他項目との相関関係が正しいことをチェックする。
                    if ws_ippan[i].cell(row=j, column=9).value is None:
                        pass
                        ### result_correlate_grid.append([ws_ippan[i].title, j, 9, MESSAGE[357][0], MESSAGE[357][1], MESSAGE[357][2], MESSAGE[357][3], MESSAGE[357][4]])
                        ### ws_ippan[i].cell(row=j, column=9).fill = fill
                        ### ws_result[i].cell(row=j, column=9).fill = fill
                        
                    ### セルJ20: 被害建物棟数, 半壊について他項目との相関関係が正しいことをチェックする。
                    if ws_ippan[i].cell(row=j, column=10).value is None:
                        pass
                        ### result_correlate_grid.append([ws_ippan[i].title, j, 10, MESSAGE[358][0], MESSAGE[358][1], MESSAGE[358][2], MESSAGE[358][3], MESSAGE[358][4]])
                        ### ws_ippan[i].cell(row=j, column=10).fill = fill
                        ### ws_result[i].cell(row=j, column=10).fill = fill
                        
                    ### セルK20: 被害建物棟数, 全壊・流失について他項目との相関関係が正しいことをチェックする。
                    if ws_ippan[i].cell(row=j, column=11).value is None:
                        pass
                        ### result_correlate_grid.append([ws_ippan[i].title, j, 11, MESSAGE[359][0], MESSAGE[359][1], MESSAGE[359][2], MESSAGE[359][3], MESSAGE[359][4]])
                        ### ws_ippan[i].cell(row=j, column=11).fill = fill
                        ### ws_result[i].cell(row=j, column=11).fill = fill
                        
                    ### セルL20: 被害建物の延床面積について他項目との相関関係が正しいことをチェックする。
                    if ws_ippan[i].cell(row=j, column=12).value is None:
                        pass
                        ### result_correlate_grid.append([ws_ippan[i].title, j, 12, MESSAGE[360][0], MESSAGE[360][1], MESSAGE[360][2], MESSAGE[360][3], MESSAGE[360][4]])
                        ### ws_ippan[i].cell(row=j, column=12).fill = fill
                        ### ws_result[i].cell(row=j, column=12).fill = fill
                        
                    ### セルM20: 被災世帯数について他項目との相関関係が正しいことをチェックする。
                    if ws_ippan[i].cell(row=j, column=13).value is None:
                        pass
                        ### result_correlate_grid.append([ws_ippan[i].title, j, 13, MESSAGE[361][0], MESSAGE[361][1], MESSAGE[361][2], MESSAGE[361][3], MESSAGE[361][4]])
                        ### ws_ippan[i].cell(row=j, column=13).fill = fill
                        ### ws_result[i].cell(row=j, column=13).fill = fill
                        
                    ### セルN20: 被災事業所数について他項目との相関関係が正しいことをチェックする。
                    if ws_ippan[i].cell(row=j, column=14).value is None:
                        pass
                        ### result_correlate_grid.append([ws_ippan[i].title, j, 14, MESSAGE[362][0], MESSAGE[362][1], MESSAGE[362][2], MESSAGE[362][3], MESSAGE[362][4]])
                        ### ws_ippan[i].cell(row=j, column=14).fill = fill
                        ### ws_result[i].cell(row=j, column=14).fill = fill
                        
                    ### セルO20: 農家・漁家戸数, 床下浸水について他項目との相関関係が正しいことをチェックする。
                    if ws_ippan[i].cell(row=j, column=15).value is None:
                        pass
                        ### result_correlate_grid.append([ws_ippan[i].title, j, 15, MESSAGE[363][0], MESSAGE[363][1], MESSAGE[363][2], MESSAGE[363][3], MESSAGE[363][4]])
                        ### ws_ippan[i].cell(row=j, column=15).fill = fill
                        ### ws_result[i].cell(row=j, column=15).fill = fill
                        
                    ### セルP20: 農家・漁家戸数, 1cm〜49cmについて他項目との相関関係が正しいことをチェックする。
                    if ws_ippan[i].cell(row=j, column=16).value is None:
                        pass
                        ### result_correlate_grid.append([ws_ippan[i].title, j, 16, MESSAGE[364][0], MESSAGE[364][1], MESSAGE[364][2], MESSAGE[364][3], MESSAGE[364][4]])
                        ### ws_ippan[i].cell(row=j, column=16).fill = fill
                        ### ws_result[i].cell(row=j, column=16).fill = fill
                        
                    ### セルQ20: 農家・漁家戸数, 50cm〜99cmについて他項目との相関関係が正しいことをチェックする。
                    if ws_ippan[i].cell(row=j, column=17).value is None:
                        pass
                        ### result_correlate_grid.append([ws_ippan[i].title, j, 17, MESSAGE[365][0], MESSAGE[365][1], MESSAGE[365][2], MESSAGE[365][3], MESSAGE[365][4]])
                        ### ws_ippan[i].cell(row=j, column=17).fill = fill
                        ### ws_result[i].cell(row=j, column=17).fill = fill
                        
                    ### セルR20: 農家・漁家戸数, 1m以上・半壊について他項目との相関関係が正しいことをチェックする。
                    if ws_ippan[i].cell(row=j, column=18).value is None:
                        pass
                        ### result_correlate_grid.append([ws_ippan[i].title, j, 18, MESSAGE[366][0], MESSAGE[366][1], MESSAGE[366][2], MESSAGE[366][3], MESSAGE[366][4]])
                        ### ws_ippan[i].cell(row=j, column=18).fill = fill
                        ### ws_result[i].cell(row=j, column=18).fill = fill
                        
                    ### セルS20: 農家・漁家戸数, 全壊・流失について他項目との相関関係が正しいことをチェックする。
                    if ws_ippan[i].cell(row=j, column=19).value is None:
                        pass
                        ### result_correlate_grid.append([ws_ippan[i].title, j, 19, MESSAGE[367][0], MESSAGE[367][1], MESSAGE[367][2], MESSAGE[367][3], MESSAGE[367][4]])
                        ### ws_ippan[i].cell(row=j, column=19).fill = fill
                        ### ws_result[i].cell(row=j, column=19).fill = fill
                        
                    ### セルT20: 事業所従業者数, 床下浸水について他項目との相関関係が正しいことをチェックする。
                    if ws_ippan[i].cell(row=j, column=20).value is None:
                        pass
                        ### result_correlate_grid.append([ws_ippan[i].title, j, 20, MESSAGE[368][0], MESSAGE[368][1], MESSAGE[368][2], MESSAGE[368][3], MESSAGE[368][4]])
                        ### ws_ippan[i].cell(row=j, column=20).fill = fill
                        ### ws_result[i].cell(row=j, column=20).fill = fill
                        
                    ### セルU20: 事業所従業者数, 1cm〜49cmについて他項目との相関関係が正しいことをチェックする。
                    if ws_ippan[i].cell(row=j, column=21).value is None:
                        pass
                        ### result_correlate_grid.append([ws_ippan[i].title, j, 21, MESSAGE[369][0], MESSAGE[369][1], MESSAGE[369][2], MESSAGE[369][3], MESSAGE[369][4]])
                        ### ws_ippan[i].cell(row=j, column=21).fill = fill
                        ### ws_result[i].cell(row=j, column=21).fill = fill
                        
                    ### セルV20: 事業所従業者数, 50cm〜99cmについて他項目との相関関係が正しいことをチェックする。
                    if ws_ippan[i].cell(row=j, column=22).value is None:
                        pass
                        ### result_correlate_grid.append([ws_ippan[i].title, j, 22, MESSAGE[370][0], MESSAGE[370][1], MESSAGE[370][2], MESSAGE[370][3], MESSAGE[370][4]])
                        ### ws_ippan[i].cell(row=j, column=22).fill = fill
                        ### ws_result[i].cell(row=j, column=22).fill = fill
                        
                    ### セルW20: 事業所従業者数, 1m以上・半壊について他項目との相関関係が正しいことをチェックする。
                    if ws_ippan[i].cell(row=j, column=23).value is None:
                        pass
                        ### result_correlate_grid.append([ws_ippan[i].title, j, 23, MESSAGE[371][0], MESSAGE[371][1], MESSAGE[371][2], MESSAGE[371][3], MESSAGE[371][4]])
                        ### ws_ippan[i].cell(row=j, column=23).fill = fill
                        ### ws_result[i].cell(row=j, column=23).fill = fill
                        
                    ### セルX20: 事業所従業者数, 全壊・流失について他項目との相関関係が正しいことをチェックする。
                    if ws_ippan[i].cell(row=j, column=24).value is None:
                        pass
                        ### result_correlate_grid.append([ws_ippan[i].title, j, 24, MESSAGE[372][0], MESSAGE[372][1], MESSAGE[372][2], MESSAGE[372][3], MESSAGE[372][4]])
                        ### ws_ippan[i].cell(row=j, column=24).fill = fill
                        ### ws_result[i].cell(row=j, column=24).fill = fill
                        
                    ### セルY20: 事業所の産業区分について他項目との相関関係が正しいことをチェックする。
                    if ws_ippan[i].cell(row=j, column=25).value is None:
                        pass
                        ### result_correlate_grid.append([ws_ippan[i].title, j, 25, MESSAGE[373][0], MESSAGE[373][1], MESSAGE[373][2], MESSAGE[373][3], MESSAGE[373][4]])
                        ### ws_ippan[i].cell(row=j, column=25).fill = fill
                        ### ws_result[i].cell(row=j, column=25).fill = fill
                        
                    ### セルZ20: 地下空間の利用形態について他項目との相関関係が正しいことをチェックする。
                    if ws_ippan[i].cell(row=j, column=26).value is None:
                        pass
                        ### result_correlate_grid.append([ws_ippan[i].title, j, 26, MESSAGE[374][0], MESSAGE[374][1], MESSAGE[374][2], MESSAGE[374][3], MESSAGE[374][4]])
                        ### ws_ippan[i].cell(row=j, column=26).fill = fill
                        ### ws_result[i].cell(row=j, column=26).fill = fill
                        
                    ### セルAA20: 備考について他項目との相関関係が正しいことをチェックする。
                    if ws_ippan[i].cell(row=j, column=27).value is None:
                        pass
                        ### result_correlate_grid.append([ws_ippan[i].title, j, 27, MESSAGE[375][0], MESSAGE[375][1], MESSAGE[375][2], MESSAGE[375][3], MESSAGE[375][4]])
                        ### ws_ippan[i].cell(row=j, column=27).fill = fill
                        ### ws_result[i].cell(row=j, column=27).fill = fill
    
        #######################################################################
        #######################################################################
        ### EXCELセルデータ突合チェック処理（5000）
        ### (1)セルB7からセルI7についてデータベースに登録されている値と突合せチェックする。
        ### (2)チェック結果リストにセルの行、列とメッセージを追加する。
        ### (3)IPPANワークシートとRESULTワークシートのセルに背景赤色の塗りつぶしをセットする。
        #######################################################################
        #######################################################################
        print_log('[INFO] P0300ExcelUpload.index_view()関数 STEP 26/34.', 'INFO')
        for i, _ in enumerate(ws_ippan):
            ### 7行目
            ### セルB7: 都道府県についてデータベースに登録されている値と突合せチェックする。
            if ws_ippan[i].cell(row=7, column=2).value is None:
                pass
            else:
                if ws_ippan[i].cell(row=7, column=2).value not in list(ken_code_list) and \
                    ws_ippan[i].cell(row=7, column=2).value not in list(ken_name_list) and \
                    ws_ippan[i].cell(row=7, column=2).value not in list(ken_name_code_list):
                    result_compare_list.append([ws_ippan[i].title, 7, 2, MESSAGE[400][0], MESSAGE[400][1], MESSAGE[400][2], MESSAGE[400][3]])
                    add_fill_to_cell(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=7, column=2, fill=fill)
                    add_comment_to_cell(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=7, column=2, message=MESSAGE[400][3]+MESSAGE[400][4])
                
            ### セルC7: 市区町村についてデータベースに登録されている値と突合せチェックする。
            if ws_ippan[i].cell(row=7, column=3).value is None:
                pass
            else:
                if ws_ippan[i].cell(row=7, column=3).value not in list(city_code_list) and \
                    ws_ippan[i].cell(row=7, column=3).value not in list(city_name_list) and \
                    ws_ippan[i].cell(row=7, column=3).value not in list(city_name_code_list):
                    result_compare_list.append([ws_ippan[i].title, 7, 3, MESSAGE[401][0], MESSAGE[401][1], MESSAGE[401][2], MESSAGE[401][3]])
                    add_fill_to_cell(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=7, column=3, fill=fill)
                    add_comment_to_cell(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=7, column=3, message=MESSAGE[401][3]+MESSAGE[401][4])
                
            ### セルD7: 水害発生月日についてデータベースに登録されている値と突合せチェックする。
            ### セルE7: 水害終了月日についてデータベースに登録されている値と突合せチェックする。
                
            ### セルF7: 水害原因1についてデータベースに登録されている値と突合せチェックする。
            if ws_ippan[i].cell(row=7, column=6).value is None:
                pass
            else:
                if ws_ippan[i].cell(row=7, column=6).value not in list(cause_code_list) and \
                    ws_ippan[i].cell(row=7, column=6).value not in list(cause_name_list) and \
                    ws_ippan[i].cell(row=7, column=6).value not in list(cause_name_code_list):
                    result_compare_list.append([ws_ippan[i].title, 7, 6, MESSAGE[404][0], MESSAGE[404][1], MESSAGE[404][2], MESSAGE[404][3]])
                    add_fill_to_cell(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=7, column=6, fill=fill)
                    add_comment_to_cell(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=7, column=6, message=MESSAGE[404][3]+MESSAGE[404][4])
                
            ### セルG7: 水害原因2についてデータベースに登録されている値と突合せチェックする。
            if ws_ippan[i].cell(row=7, column=7).value is None:
                pass
            else:
                if ws_ippan[i].cell(row=7, column=7).value not in list(cause_code_list) and \
                    ws_ippan[i].cell(row=7, column=7).value not in list(cause_name_list) and \
                    ws_ippan[i].cell(row=7, column=7).value not in list(cause_name_code_list):
                    result_compare_list.append([ws_ippan[i].title, 7, 7, MESSAGE[405][0], MESSAGE[405][1], MESSAGE[405][2], MESSAGE[405][3]])
                    add_fill_to_cell(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=7, column=7, fill=fill)
                    add_comment_to_cell(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=7, column=7, message=MESSAGE[405][3]+MESSAGE[405][4])
                
            ### セルH7: 水害原因3についてデータベースに登録されている値と突合せチェックする。
            if ws_ippan[i].cell(row=7, column=8).value is None:
                pass
            else:
                if ws_ippan[i].cell(row=7, column=8).value not in list(cause_code_list) and \
                    ws_ippan[i].cell(row=7, column=8).value not in list(cause_name_list) and \
                    ws_ippan[i].cell(row=7, column=8).value not in list(cause_name_code_list):
                    result_compare_list.append([ws_ippan[i].title, 7, 8, MESSAGE[406][0], MESSAGE[406][1], MESSAGE[406][2], MESSAGE[406][3]])
                    add_fill_to_cell(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=7, column=8, fill=fill)
                    add_comment_to_cell(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=7, column=8, message=MESSAGE[406][3]+MESSAGE[406][4])
                
            ### セルI7: 水害区域番号についてデータベースに登録されている値と突合せチェックする。
            if ws_ippan[i].cell(row=7, column=9).value is None:
                pass
            else:
                if ws_ippan[i].cell(row=7, column=9).value not in list(area_id_list) and \
                    ws_ippan[i].cell(row=7, column=9).value not in list(area_name_list) and \
                    ws_ippan[i].cell(row=7, column=9).value not in list(area_name_id_list):
                    result_compare_list.append([ws_ippan[i].title, 7, 9, MESSAGE[407][0], MESSAGE[407][1], MESSAGE[407][2], MESSAGE[407][3]])
                    add_fill_to_cell(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=7, column=9, fill=fill)
                    add_comment_to_cell(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=7, column=9, message=MESSAGE[407][3]+MESSAGE[407][4])
    
        #######################################################################
        ### EXCELセルデータ突合チェック処理（5010）
        ### (1)セルB10からセルF10についてデータベースに登録されている値と突合せチェックする。
        ### (2)チェック結果リストにセルの行、列とメッセージを追加する。
        ### (3)IPPANワークシートとRESULTワークシートのセルに背景赤色の塗りつぶしをセットする。
        ### TO-DO: if == ''はダミーの処理である。突合せチェック処理を記述する。
        #######################################################################
        print_log('[INFO] P0300ExcelUpload.index_view()関数 STEP 27/34.', 'INFO')
        for i, _ in enumerate(ws_ippan):
            ### 10行目
            ### セルB10: 水系・沿岸名についてデータベースに登録されている値と突合せチェックする。
            if ws_ippan[i].cell(row=10, column=2).value is None:
                pass
            else:
                if ws_ippan[i].cell(row=10, column=2).value not in list(suikei_code_list) and \
                    ws_ippan[i].cell(row=10, column=2).value not in list(suikei_name_list) and \
                    ws_ippan[i].cell(row=10, column=2).value not in list(suikei_name_code_list):
                    result_compare_list.append([ws_ippan[i].title, 10, 2, MESSAGE[408][0], MESSAGE[408][1], MESSAGE[408][2], MESSAGE[408][3]])
                    add_fill_to_cell(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=10, column=2, fill=fill)
                    add_comment_to_cell(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=10, column=2, message=MESSAGE[408][3]+MESSAGE[408][4])
                
            ### セルC10: 水系種別についてデータベースに登録されている値と突合せチェックする。
            if ws_ippan[i].cell(row=10, column=3).value is None:
                pass
            else:
                if ws_ippan[i].cell(row=10, column=3).value not in list(suikei_type_code_list) and \
                    ws_ippan[i].cell(row=10, column=3).value not in list(suikei_type_name_list) and \
                    ws_ippan[i].cell(row=10, column=3).value not in list(suikei_type_name_code_list):
                    result_compare_list.append([ws_ippan[i].title, 10, 3, MESSAGE[409][0], MESSAGE[409][1], MESSAGE[409][2], MESSAGE[409][3]])
                    add_fill_to_cell(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=10, column=3, fill=fill)
                    add_comment_to_cell(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=10, column=3, message=MESSAGE[409][3]+MESSAGE[409][4])
                
            ### セルD10: 河川・海岸名についてデータベースに登録されている値と突合せチェックする。
            if ws_ippan[i].cell(row=10, column=4).value is None:
                pass
            else:
                if ws_ippan[i].cell(row=10, column=4).value not in list(kasen_code_list) and \
                    ws_ippan[i].cell(row=10, column=4).value not in list(kasen_name_list) and \
                    ws_ippan[i].cell(row=10, column=4).value not in list(kasen_name_code_list):
                    result_compare_list.append([ws_ippan[i].title, 10, 4, MESSAGE[410][0], MESSAGE[410][1], MESSAGE[410][2], MESSAGE[410][3]])
                    add_fill_to_cell(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=10, column=4, fill=fill)
                    add_comment_to_cell(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=10, column=4, message=MESSAGE[410][3]+MESSAGE[410][4])
                
            ### セルE10: 河川種別についてデータベースに登録されている値と突合せチェックする。
            if ws_ippan[i].cell(row=10, column=5).value is None:
                pass
            else:
                if ws_ippan[i].cell(row=10, column=5).value not in list(kasen_type_code_list) and \
                    ws_ippan[i].cell(row=10, column=5).value not in list(kasen_type_name_list) and \
                    ws_ippan[i].cell(row=10, column=5).value not in list(kasen_type_name_code_list):
                    result_compare_list.append([ws_ippan[i].title, 10, 5, MESSAGE[411][0], MESSAGE[411][1], MESSAGE[411][2], MESSAGE[411][3]])
                    add_fill_to_cell(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=10, column=5, fill=fill)
                    add_comment_to_cell(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=10, column=5, message=MESSAGE[412][3]+MESSAGE[412][4])
                
            ### セルF10: 地盤勾配区分についてデータベースに登録されている値と突合せチェックする。
            if ws_ippan[i].cell(row=10, column=6).value is None:
                pass
            else:
                if ws_ippan[i].cell(row=10, column=6).value not in list(gradient_code_list) and \
                    ws_ippan[i].cell(row=10, column=6).value not in list(gradient_name_list) and \
                    ws_ippan[i].cell(row=10, column=6).value not in list(gradient_name_code_list):
                    result_compare_list.append([ws_ippan[i].title, 10, 6, MESSAGE[412][0], MESSAGE[412][1], MESSAGE[412][2], MESSAGE[412][3]])
                    add_fill_to_cell(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=10, column=6, fill=fill)
                    add_comment_to_cell(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=10, column=6, message=MESSAGE[412][3]+MESSAGE[412][4])
    
        #######################################################################
        ### EXCELセルデータ突合チェック処理（5020）
        ### (1)セルB14からセルJ14についてデータベースに登録されている値と突合せチェックする。
        ### (2)チェック結果リストにセルの行、列とメッセージを追加する。
        ### (3)IPPANワークシートとRESULTワークシートのセルに背景赤色の塗りつぶしをセットする。
        ### TO-DO: if == ''はダミーの処理である。突合せチェック処理を記述する。
        #######################################################################
        print_log('[INFO] P0300ExcelUpload.index_view()関数 STEP 28/34.', 'INFO')
        for i, _ in enumerate(ws_ippan):
            ### 14行目
            ### セルB14: 水害区域面積の宅地についてデータベースに登録されている値と突合せチェックする。
            ### セルC14: 水害区域面積の農地についてデータベースに登録されている値と突合せチェックする。
            ### セルD14: 水害区域面積の地下についてデータベースに登録されている値と突合せチェックする。
                
            ### セルF14: 工種についてデータベースに登録されている値と突合せチェックする。
            if ws_ippan[i].cell(row=14, column=6).value is None:
                pass
            else:
                if ws_ippan[i].cell(row=14, column=6).value not in list(kasen_kaigan_code_list) and \
                    ws_ippan[i].cell(row=14, column=6).value not in list(kasen_kaigan_name_list) and \
                    ws_ippan[i].cell(row=14, column=6).value not in list(kasen_kaigan_name_code_list):
                    result_compare_list.append([ws_ippan[i].title, 14, 6, MESSAGE[416][0], MESSAGE[416][1], MESSAGE[416][2], MESSAGE[416][3]])
                    add_fill_to_cell(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=14, column=6, fill=fill)
                    add_comment_to_cell(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=14, column=2, message=MESSAGE[416][3]+MESSAGE[416][4])
                
            ### セルH14: 農作物被害額についてデータベースに登録されている値と突合せチェックする。
                
            ### セルJ14: 異常気象コードについてデータベースに登録されている値と突合せチェックする。
            if ws_ippan[i].cell(row=14, column=10).value is None:
                pass
            else:
                if ws_ippan[i].cell(row=14, column=10).value not in list(weather_id_list) and \
                    ws_ippan[i].cell(row=14, column=10).value not in list(weather_name_list) and \
                    ws_ippan[i].cell(row=14, column=10).value not in list(weather_name_id_list):
                    result_compare_list.append([ws_ippan[i].title, 14, 10, MESSAGE[418][0], MESSAGE[418][1], MESSAGE[418][2], MESSAGE[418][3]])
                    add_fill_to_cell(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=14, column=10, fill=fill)
                    add_comment_to_cell(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=14, column=10, message=MESSAGE[418][3]+MESSAGE[418][4])
    
        #######################################################################
        ### EXCELセルデータ突合チェック処理（5030）
        ### (1)セルB20からセルAA20についてデータベースに登録されている値と突合せチェックする。
        ### (2)チェック結果リストにセルの行、列とメッセージを追加する。
        ### (3)IPPANワークシートとRESULTワークシートのセルに背景赤色の塗りつぶしをセットする。
        ### TO-DO: if == ''はダミーの処理である。突合せチェック処理を記述する。
        #######################################################################
        print_log('[INFO] P0300ExcelUpload.index_view()関数 STEP 29/34.', 'INFO')
        for i, _ in enumerate(ws_ippan):
            if ws_max_row[i] >= 20:
                for j in range(20, ws_max_row[i] + 1):
                    ### セルB20: 町丁名・大字名についてデータベースに登録されている値と突合せチェックする。
                    
                    ### セルC20: 名称についてデータベースに登録されている値と突合せチェックする。
                    if ws_ippan[i].cell(row=j, column=4).value is None:
                        pass
                    else:
                        if ws_ippan[i].cell(row=j, column=3).value not in list(building_code_list) and \
                            ws_ippan[i].cell(row=j, column=3).value not in list(building_name_list) and \
                            ws_ippan[i].cell(row=j, column=3).value not in list(building_name_code_list):
                            result_compare_grid.append([ws_ippan[i].title, j, 3, MESSAGE[451][0], MESSAGE[451][1], MESSAGE[451][2], MESSAGE[451][3]])
                            add_fill_to_cell(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=j, column=3, fill=fill)
                            add_comment_to_cell(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=j, column=3, message=MESSAGE[451][3]+MESSAGE[451][4])
                        
                    ### セルD20: 地上・地下被害の区分についてデータベースに登録されている値と突合せチェックする。
                    if ws_ippan[i].cell(row=j, column=4).value is None:
                        pass
                    else:
                        if ws_ippan[i].cell(row=j, column=4).value not in list(underground_code_list) and \
                            ws_ippan[i].cell(row=j, column=4).value not in list(underground_name_list) and \
                            ws_ippan[i].cell(row=j, column=4).value not in list(underground_name_code_list):
                            result_compare_grid.append([ws_ippan[i].title, j, 4, MESSAGE[452][0], MESSAGE[452][1], MESSAGE[452][2], MESSAGE[452][3]])
                            add_fill_to_cell(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=j, column=4, fill=fill)
                            add_comment_to_cell(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=j, column=4, message=MESSAGE[452][3]+MESSAGE[452][4])
                        
                    ### セルE20: 浸水土砂被害の区分についてデータベースに登録されている値と突合せチェックする。
                    if ws_ippan[i].cell(row=j, column=5).value is None:
                        pass
                    else:
                        if ws_ippan[i].cell(row=j, column=5).value not in list(flood_sediment_code_list) and \
                            ws_ippan[i].cell(row=j, column=5).value not in list(flood_sediment_name_list) and \
                            ws_ippan[i].cell(row=j, column=5).value not in list(flood_sediment_name_code_list):
                            result_compare_grid.append([ws_ippan[i].title, j, 5, MESSAGE[453][0], MESSAGE[453][1], MESSAGE[453][2], MESSAGE[453][3]])
                            add_fill_to_cell(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=j, column=5, fill=fill)
                            add_comment_to_cell(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=j, column=5, message=MESSAGE[453][3]+MESSAGE[453][4])
                        
                    ### セルF20: 被害建物棟数, 床下浸水についてデータベースに登録されている値と突合せチェックする。
                    ### セルG20: 被害建物棟数, 1cm〜49cmについてデータベースに登録されている値と突合せチェックする。
                    ### セルH20: 被害建物棟数, 50cm〜99cmについてデータベースに登録されている値と突合せチェックする。
                    ### セルI20: 被害建物棟数, 1m以上についてデータベースに登録されている値と突合せチェックする。
                    ### セルJ20: 被害建物棟数, 半壊についてデータベースに登録されている値と突合せチェックする。
                    ### セルK20: 被害建物棟数, 全壊・流失についてデータベースに登録されている値と突合せチェックする。
                    ### セルL20: 被害建物の延床面積についてデータベースに登録されている値と突合せチェックする。
                    ### セルM20: 被災世帯数についてデータベースに登録されている値と突合せチェックする。
                    ### セルN20: 被災事業所数についてデータベースに登録されている値と突合せチェックする。
                    ### セルO20: 農家・漁家戸数, 床下浸水についてデータベースに登録されている値と突合せチェックする。
                    ### セルP20: 農家・漁家戸数, 1cm〜49cmについてデータベースに登録されている値と突合せチェックする。
                    ### セルQ20: 農家・漁家戸数, 50cm〜99cmについてデータベースに登録されている値と突合せチェックする。
                    ### セルR20: 農家・漁家戸数, 1m以上・半壊についてデータベースに登録されている値と突合せチェックする。
                    ### セルS20: 農家・漁家戸数, 全壊・流失についてデータベースに登録されている値と突合せチェックする。
                    ### セルT20: 事業所従業者数, 床下浸水についてデータベースに登録されている値と突合せチェックする。
                    ### セルU20: 事業所従業者数, 1cm〜49cmについてデータベースに登録されている値と突合せチェックする。
                    ### セルV20: 事業所従業者数, 50cm〜99cmについてデータベースに登録されている値と突合せチェックする。
                    ### セルW20: 事業所従業者数, 1m以上・半壊についてデータベースに登録されている値と突合せチェックする。
                    ### セルX20: 事業所従業者数, 全壊・流失についてデータベースに登録されている値と突合せチェックする。
                        
                    ### セルY20: 事業所の産業区分についてデータベースに登録されている値と突合せチェックする。
                    if ws_ippan[i].cell(row=j, column=25).value is None:
                        pass
                    else:
                        if ws_ippan[i].cell(row=j, column=25).value not in list(industry_code_list) and \
                            ws_ippan[i].cell(row=j, column=25).value not in list(industry_name_list) and \
                            ws_ippan[i].cell(row=j, column=25).value not in list(industry_name_code_list):
                            result_compare_grid.append([ws_ippan[i].title, j, 25, MESSAGE[473][0], MESSAGE[473][1], MESSAGE[473][2], MESSAGE[473][3]])
                            add_fill_to_cell(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=j, column=25, fill=fill)
                            add_comment_to_cell(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=j, column=25, message=MESSAGE[473][3]+MESSAGE[473][4])
                        
                    ### セルZ20: 地下空間の利用形態についてデータベースに登録されている値と突合せチェックする。
                    if ws_ippan[i].cell(row=j, column=26).value is None:
                        pass
                    else:
                        if ws_ippan[i].cell(row=j, column=26).value not in list(usage_code_list) and \
                            ws_ippan[i].cell(row=j, column=26).value not in list(usage_name_list) and \
                            ws_ippan[i].cell(row=j, column=26).value not in list(usage_name_code_list):
                            result_compare_grid.append([ws_ippan[i].title, j, 26, MESSAGE[474][0], MESSAGE[474][1], MESSAGE[474][2], MESSAGE[474][3]])
                            add_fill_to_cell(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=j, column=26, fill=fill)
                            add_comment_to_cell(ws_ippan=ws_ippan[i], ws_result=ws_result[i], row=j, column=26, message=MESSAGE[474][3]+MESSAGE[474][4])
                        
                    ### セルAA20: 備考についてデータベースに登録されている値と突合せチェックする。

        #######################################################################
        #######################################################################
        ### ファイル入出力処理(6000)
        ### チェック結果ファイルを保存する。
        #######################################################################
        #######################################################################
        print_log('[INFO] P0300ExcelUpload.index_view()関数 STEP 30/34.', 'INFO')
        wb.save(output_file_path)

        #######################################################################
        #######################################################################
        ### ログ出力処理(7000)
        #######################################################################
        #######################################################################
        print_log('[INFO] P0300ExcelUpload.index_view()関数 STEP 31/34.', 'INFO')
        if len(result_require_list) > 0 or len(result_require_grid) > 0 or \
            len(result_format_list) > 0 or len(result_format_grid) > 0 or \
            len(result_range_list) > 0 or len(result_range_grid) > 0 or \
            len(result_correlate_list) > 0 or len(result_correlate_grid) > 0 or \
            len(result_compare_list) > 0 or len(result_compare_grid) > 0:
            
            print_log('False', 'INFO')

        else:
            print_log('True', 'INFO')
            
        print_log('ws_max_row = {}'.format(ws_max_row), 'INFO')
        print_log('len(result_require_list) = {}'.format(len(result_require_list)), 'INFO')
        print_log('len(result_format_list) = {}'.format(len(result_format_list)), 'INFO')
        print_log('len(result_range_list) = {}'.format(len(result_range_list)), 'INFO')
        print_log('len(result_correlate_list) = {}'.format(len(result_correlate_list)), 'INFO')
        print_log('len(result_compare_list) = {}'.format(len(result_compare_list)), 'INFO')
            
        #######################################################################
        #######################################################################
        ### レスポンスセット処理(8000)
        ### テンプレートとコンテキストを設定して、レスポンスをブラウザに戻す。
        ### ※入力チェックでエラーが発見された場合、
        ### ※ネストを浅くするために、処理対象外の場合、終了させる。
        #######################################################################
        #######################################################################
        print_log('[INFO] P0300ExcelUpload.index_view()関数 STEP 32/34.', 'INFO')
        if len(result_require_list) > 0 or len(result_require_grid) > 0 or \
            len(result_format_list) > 0 or len(result_format_grid) > 0 or \
            len(result_range_list) > 0 or len(result_range_grid) > 0 or \
            len(result_correlate_list) > 0 or len(result_correlate_grid) > 0 or \
            len(result_compare_list) > 0 or len(result_compare_grid) > 0:
            
            ### src/P0300ExcelUpload/templates/P0300ExcelUpload/fail.htmlを使用する。
            ### 上記はテンプレートファイルの場所がわからなくなることがあるためのメモである。
            template = loader.get_template('P0300ExcelUpload/fail.html')
            context = {
                'result_require_list': result_require_list,
                'result_require_grid': result_require_grid,
                'result_format_list': result_format_list,
                'result_format_grid': result_format_grid,
                'result_range_list': result_range_list,
                'result_range_grid': result_range_grid,
                'result_correlate_list': result_correlate_list,
                'result_correlate_grid': result_correlate_grid,
                'result_compare_list': result_compare_list,
                'result_compare_grid': result_compare_grid,
                'excel_id': 1,
            }
            print_log('[INFO] P0300ExcelUpload.index_view()関数が正常終了しました。', 'INFO')
            return HttpResponse(template.render(context, request))
        
        #######################################################################
        #######################################################################
        ### DBアクセス処理(9000)
        ### (1)一般資産入力データ_ヘッダ部分テーブルにデータを登録する。
        ### (2)一般資産入力データ_一覧票部分テーブルにデータを登録する。
        ### (3)レポジトリテーブルにデータを登録する。
        ### ※入力チェックでエラーが発見されなかった場合、
        #######################################################################
        #######################################################################
        print_log('[INFO] P0300ExcelUpload.index_view()関数 STEP 33/34.', 'INFO')
        connection_cursor = connection.cursor()
        try:
            ###############################################################
            ### DBアクセス処理(9010)
            ### 一般資産入力データ_ヘッダ部分テーブルのデータを削除する。
            ### 一般資産入力データ_一覧表部分テーブルのデータを削除する。
            ### ※入力チェックでエラーが発見されなかった場合、
            ### ※二重登録防止のため、同じ市区町村のデータは削除する。
            ###############################################################
            print_log('[INFO] P0300ExcelUpload.index_view()関数 STEP 33_1/34.', 'INFO')
            del_suigai_list = SUIGAI.objects.raw("""
                SELECT 
                    SUIGAI_ID 
                FROM SUIGAI 
                WHERE CITY_CODE=%s""", [split_name_code(ws_ippan[0].cell(row=7, column=3).value)[-1],])
            del_suigai_id_list = [del_suigai.suigai_id for del_suigai in del_suigai_list]
            del_suigai_id_str = ",".join([str(i) for i in del_suigai_id_list])
            print_log('del_suigai_id_str = {}'.format(del_suigai_id_str), 'INFO')
            
            del_ippan_list = IPPAN_VIEW.objects.raw("""
                SELECT 
                    IPPAN_ID 
                FROM IPPAN_VIEW 
                WHERE CITY_CODE=%s""", [split_name_code(ws_ippan[0].cell(row=7, column=3).value)[-1],])
            del_ippan_id_list = [del_ippan.ippan_id for del_ippan in del_ippan_list]
            del_ippan_id_str = ",".join([str(i) for i in del_ippan_id_list])
            print_log('del_ippan_id_str = {}'.format(del_ippan_id_str), 'INFO')
            
            ### connection_cursor.execute("""
            ###     DELETE 
            ###     FROM SUIGAI 
            ###     WHERE SUIGAI_ID IN (%s)""", [
            ###     del_suigai_id_str,])

            ### connection_cursor.execute("""
            ###     DELETE 
            ###     FROM IPPAN 
            ###     WHERE IPPAN_ID IN (%s)""", [
            ###     del_ippan_id_str,])

            connection_cursor.execute("""
                DELETE FROM SUIGAI WHERE SUIGAI_ID IN (SELECT SUIGAI_ID FROM SUIGAI WHERE CITY_CODE=%s)""", [
                split_name_code(ws_ippan[0].cell(row=7, column=3).value)[-1],])
    
            connection_cursor.execute("""
                DELETE FROM IPPAN WHERE IPPAN_ID IN (SELECT IPPAN_ID FROM IPPAN_VIEW WHERE CITY_CODE=%s)""", [
                split_name_code(ws_ippan[0].cell(row=7, column=3).value)[-1],])

            ###############################################################
            ### DBアクセス処理(9020)
            ###############################################################
            for i, _ in enumerate(ws_ippan):
                print_log('[INFO] P0300ExcelUpload.index_view()関数 STEP 33_2/34.', 'INFO')
                ### suigai_id__max で正しい。
                suigai_id = SUIGAI.objects.all().aggregate(Max('suigai_id'))['suigai_id__max']
                
                ### 一般資産入力データ_ヘッダ部分テーブルにレコードが１件も存在しない場合、
                if suigai_id is None:
                    suigai_id = 0
                ### 一般資産入力データ_ヘッダ部分テーブルにレコードが存在する場合、
                else:
                    suigai_id = suigai_id + 1
                    
                print_log('suigai_id = {}'.format(suigai_id), 'INFO')
                
                ###############################################################
                ### DBアクセス処理(9030)
                ### 一般資産入力データ_ヘッダ部分テーブルにデータを登録する。
                ### ※入力チェックでエラーが発見されなかった場合、
                ###############################################################
                print_log('[INFO] P0300ExcelUpload.index_view()関数 STEP 31_3/34.', 'INFO')
                connection_cursor.execute("""
                    INSERT INTO SUIGAI (
                        suigai_id, 
                        suigai_name, 
                        ken_code, 
                        city_code, 
                        cause_1_code, 
                        cause_2_code, 
                        cause_3_code, 
                        area_id, 
                        suikei_code, 
                        kasen_code, 
                        gradient_code, 
                        residential_area, 
                        agricultural_area, 
                        underground_area, 
                        kasen_kaigan_code, 
                        crop_damage, 
                        weather_id 
                    ) VALUES (
                        %s, -- suigai_id 
                        %s, -- suigai_name 
                        %s, -- ken_code 
                        %s, -- city_code 
                        %s, -- cause_1_code 
                        %s, -- cause_2_code 
                        %s, -- cause_3_code 
                        %s, -- area_id 
                        %s, -- suikei_code 
                        %s, -- kasen_code 
                        %s, -- gradient_code 
                        %s, -- residential_area 
                        %s, -- agricultural_area 
                        %s, -- underground_area 
                        %s, -- kasen_kaigan_code 
                        %s, -- crop_damage 
                        %s  -- weather_id 
                    )""", [
                        suigai_id,                                                                               ### suigai_id
                        'suigai_name',                                                                           ### suigai_name
                        convert_empty_to_none(split_name_code(ws_ippan[i].cell(row=7, column=2).value)[-1]),     ### ken_code
                        convert_empty_to_none(split_name_code(ws_ippan[i].cell(row=7, column=3).value)[-1]),     ### city_code
                        convert_empty_to_none(split_name_code(ws_ippan[i].cell(row=7, column=6).value)[-1]),     ### cause_1_code
                        convert_empty_to_none(split_name_code(ws_ippan[i].cell(row=7, column=7).value)[-1]),     ### cause_2_code
                        convert_empty_to_none(split_name_code(ws_ippan[i].cell(row=7, column=8).value)[-1]),     ### cause_3_code
                        convert_empty_to_none(split_name_code(ws_ippan[i].cell(row=7, column=9).value)[-1]),     ### area_id
                        convert_empty_to_none(split_name_code(ws_ippan[i].cell(row=10, column=2).value)[-1]),    ### suikei_code
                        convert_empty_to_none(split_name_code(ws_ippan[i].cell(row=10, column=4).value)[-1]),    ### kasen_code
                        convert_empty_to_none(split_name_code(ws_ippan[i].cell(row=10, column=6).value)[-1]),    ### gradient_code
                        convert_empty_to_none(ws_ippan[i].cell(row=14, column=2).value),                         ### residential_area
                        convert_empty_to_none(ws_ippan[i].cell(row=14, column=3).value),                         ### agricultural_area
                        convert_empty_to_none(ws_ippan[i].cell(row=14, column=4).value),                         ### underground_area
                        convert_empty_to_none(split_name_code(ws_ippan[i].cell(row=14, column=6).value)[-1]),    ### kasen_kaigan_code
                        convert_empty_to_none(ws_ippan[i].cell(row=14, column=8).value),                         ### crop_damaga
                        convert_empty_to_none(split_name_code(ws_ippan[i].cell(row=14, column=10).value)[-1])    ### weather_id
                    ])
                    
                ###############################################################
                ### DBアクセス処理(9040)
                ### 一般資産入力データ_一覧表部分テーブルにデータを登録する。
                ### ※入力チェックでエラーが発見されなかった場合、
                ###############################################################
                print_log('[INFO] P0300ExcelUpload.index_view()関数 STEP 33_4/34.', 'INFO')
                if ws_max_row[i] >= 20:
                    for j in range(20, ws_max_row[i] + 1):
                        connection_cursor.execute(""" 
                            INSERT INTO IPPAN (
                                ippan_id, 
                                ippan_name, 
                                suigai_id, 
                                building_code, 
                                underground_code, 
                                flood_sediment_code, 
                                building_lv00, 
                                building_lv01_49, 
                                building_lv50_99, 
                                building_lv100, 
                                building_half, 
                                building_full, 
                                floor_area, 
                                family, 
                                office, 
                                farmer_fisher_lv00, 
                                farmer_fisher_lv01_49, 
                                farmer_fisher_lv50_99, 
                                farmer_fisher_lv100, 
                                farmer_fisher_full, 
                                employee_lv00, 
                                employee_lv01_49, 
                                employee_lv50_99, 
                                employee_lv100, 
                                employee_full, 
                                industry_code, 
                                usage_code, 
                                comment 
                            ) VALUES (
                                (SELECT MAX(ippan_id+1) FROM IPPAN), -- ippan_id 
                                %s, -- ippan_name 
                                %s, -- suigai_id 
                                %s, -- building_code 
                                %s, -- underground_code 
                                %s, -- flood_sediment_code 
                                %s, -- building_lv00 
                                %s, -- building_lv01_49 
                                %s, -- building_lv50_99 
                                %s, -- building_lv100 
                                %s, -- building_half 
                                %s, -- building_full 
                                %s, -- floor_area 
                                %s, -- family 
                                %s, -- office 
                                %s, -- farmer_fisher_lv00 
                                %s, -- farmer_fisher_lv01_49 
                                %s, -- farmer_fisher_lv50_99 
                                %s, -- farmer_fisher_lv100 
                                %s, -- farmer_fisher_full 
                                %s, -- employee_lv00 
                                %s, -- employee_lv01_49 
                                %s, -- employee_lv50_99 
                                %s, -- employee_lv100 
                                %s, -- employee_full 
                                %s, -- industry_code 
                                %s, -- usage_code 
                                %s  -- comment 
                            ) """, [
                                convert_empty_to_none(ws_ippan[i].cell(row=j, column=2).value),                            ### ippan_name
                                suigai_id,                                                                                 ### suigai_id
                                convert_empty_to_none(split_name_code(ws_ippan[i].cell(row=j, column=3).value)[-1]),       ### building_code
                                convert_empty_to_none(split_name_code(ws_ippan[i].cell(row=j, column=4).value)[-1]),       ### underground_code
                                convert_empty_to_none(split_name_code(ws_ippan[i].cell(row=j, column=5).value)[-1]),       ### flood_sediment_code
                                convert_empty_to_none(ws_ippan[i].cell(row=j, column=6).value),                            ### building_lv00
                                convert_empty_to_none(ws_ippan[i].cell(row=j, column=7).value),                            ### building_lv01_49
                                convert_empty_to_none(ws_ippan[i].cell(row=j, column=8).value),                            ### building_lv50_99
                                convert_empty_to_none(ws_ippan[i].cell(row=j, column=9).value),                            ### building_lv100
                                convert_empty_to_none(ws_ippan[i].cell(row=j, column=10).value),                           ### building_half
                                convert_empty_to_none(ws_ippan[i].cell(row=j, column=11).value),                           ### building_full
                                convert_empty_to_none(ws_ippan[i].cell(row=j, column=12).value),                           ### floor_area
                                convert_empty_to_none(ws_ippan[i].cell(row=j, column=13).value),                           ### family
                                convert_empty_to_none(ws_ippan[i].cell(row=j, column=14).value),                           ### office
                                convert_empty_to_none(ws_ippan[i].cell(row=j, column=15).value),                           ### farmer_fisher_lv00
                                convert_empty_to_none(ws_ippan[i].cell(row=j, column=16).value),                           ### farmer_fisher_lv01_49
                                convert_empty_to_none(ws_ippan[i].cell(row=j, column=17).value),                           ### farmer_fisher_lv50_99
                                convert_empty_to_none(ws_ippan[i].cell(row=j, column=18).value),                           ### farmer_fisher_lv100
                                convert_empty_to_none(ws_ippan[i].cell(row=j, column=19).value),                           ### farmer_fisher_full
                                convert_empty_to_none(ws_ippan[i].cell(row=j, column=20).value),                           ### employee_lv00
                                convert_empty_to_none(ws_ippan[i].cell(row=j, column=21).value),                           ### employee_lv01_49
                                convert_empty_to_none(ws_ippan[i].cell(row=j, column=22).value),                           ### employee_lv50_99
                                convert_empty_to_none(ws_ippan[i].cell(row=j, column=23).value),                           ### employee_lv100
                                convert_empty_to_none(ws_ippan[i].cell(row=j, column=24).value),                           ### employee_full
                                convert_empty_to_none(split_name_code(ws_ippan[i].cell(row=j, column=25).value)[-1]),      ### industry_code
                                convert_empty_to_none(split_name_code(ws_ippan[i].cell(row=j, column=26).value)[-1]),      ### usage_code
                                convert_empty_to_none(ws_ippan[i].cell(row=j, column=27).value)                            ### comment
                            ])

                ###############################################################
                ### DBアクセス処理(9050)
                ### レポジトリテーブルにデータを登録する。
                ### ※入力チェックでエラーが発見されなかった場合、
                ###############################################################
                print_log('[INFO] P0300ExcelUpload.index_view()関数 STEP 33_5/34.', 'INFO')
                repository_id = REPOSITORY.objects.all().aggregate(Max('repository_id'))['repository_id__max']
                ### 一般資産入力データ_ヘッダ部分テーブルにレコードが１件も存在しない場合、
                if repository_id is None:
                    repository_id = 0
                ### 一般資産入力データ_ヘッダ部分テーブルにレコードが存在する場合、
                else:
                    repository_id = repository_id + 1
                    
                print_log('repository_id = {}'.format(repository_id), 'INFO')

                connection_cursor.execute("""
                    INSERT INTO REPOSITORY (
                        repository_id, 
                        suigai_id, 
                        action_code, 
                        status_code, 
                        created_at, 
                        updated_at, 
                        input_file_path 
                    ) VALUES (
                        %s,                -- repository_id 
                        %s,                -- suigai_id 
                        %s,                -- action_code 
                        %s,                -- status_code 
                        CURRENT_TIMESTAMP, -- created_at 
                        CURRENT_TIMESTAMP, -- updated_at 
                        %s                 -- input_file_path 
                    )""", [
                        repository_id,     ### repository_id 
                        suigai_id,         ### suigai_id 
                        '3',               ### action_code 
                        '3',               ### status_code 
                        input_file_path    ### input_file_path 
                    ])

                ###############################################################
                ### DBアクセス処理(9060)
                ### トリガーテーブルにデータを登録する。
                ### ※入力チェックでエラーが発見されなかった場合、
                ###############################################################
                print_log('[INFO] P0300ExcelUpload.index_view()関数 STEP 33_6/34.', 'INFO')
                connection_cursor.execute("""
                    INSERT INTO TRIGGER (
                        trigger_id, 
                        suigai_id, 
                        repository_id, 
                        action_code, 
                        status_code, 
                        published_at, 
                        consumed_at, 
                        success_count, 
                        failure_count 
                    ) VALUES (
                        (SELECT MAX(trigger_id) + 1 FROM TRIGGER), 
                        %s,                 -- suigai_id 
                        %s,                 -- repository_id 
                        %s,                 -- action_code 
                        %s,                 -- status_code 
                        CURRENT_TIMESTAMP,  -- published_at 
                        %s,                 -- consumed_at 
                        %s,                 -- success_count 
                        %s                  -- failure_count 
                    )""", [
                        suigai_id,          ### suigai_id 
                        repository_id,      ### repository_id  
                        5,                  ### action_code 
                        None,               ### status_code 
                        None,               ### consumed_at 
                        0,                  ### success_count
                        0                   ### failure_count
                    ])
            
            print_log('[INFO] P0300ExcelUpload.index_view()関数 STEP 33_7/34.', 'INFO')
            transaction.commit()
                    
        except:
            connection_cursor.rollback()
        finally:
            connection_cursor.close()
        
        #######################################################################
        #######################################################################
        ### レスポンスセット処理(10000)
        ### テンプレートとコンテキストを設定して、レスポンスをブラウザに戻す。
        ### ※入力チェックでエラーが発見されなかった場合、
        #######################################################################
        #######################################################################
        print_log('[INFO] P0300ExcelUpload.index_view()関数 STEP 34/34.', 'INFO')
        ### src/P0300ExcelUpload/templates/P0300ExcelUpload/success.htmlを使用する。
        ### 上記はテンプレートファイルの場所がわからなくなることがあるためのメモである。
        template = loader.get_template('P0300ExcelUpload/success.html')
        context = {}
        print_log('[INFO] P0300ExcelUpload.index_view()関数が正常終了しました。', 'INFO')
        return HttpResponse(template.render(context, request))
        
    except:
        print_log(sys.exc_info()[0], 'ERROR')
        print_log('[ERROR] P0300ExcelUpload.index_view()関数でエラーが発生しました。', 'ERROR')
        print_log('[ERROR] P0300ExcelUpload.index_viwe()関数が異常終了しました。', 'ERROR')
        return render(request, 'error.html')

###############################################################################
### 関数名：ippan_chosa_result_view
### 
###############################################################################
### @login_required(None, login_url='/P0100Login/')
def ippan_chosa_result_view(request, excel_id):
    try:
        #######################################################################
        ### 引数チェック処理(0000)
        ### ブラウザからのリクエストと引数をチェックする。
        #######################################################################
        print_log('[INFO] ########################################', 'INFO')
        print_log('[INFO] P0300ExcelUpload.ippan_chosa_result_view()関数が開始しました。', 'INFO')
        print_log('[INFO] P0300ExcelUpload.ippan_chosa_result_view()関数 request = {}'.format(request.method), 'INFO')
        print_log('[INFO] P0300ExcelUpload.ippan_chosa_result_view()関数 STEP 1/1.', 'INFO')
        
        result_file_path = 'static/ippan_chosa_result2.xlsx'
        wb = openpyxl.load_workbook(result_file_path)
        
        #######################################################################
        ### レスポンスセット処理(0010)
        ### コンテキストを設定して、レスポンスをブラウザに戻す。
        #######################################################################
        print_log('[INFO] P0300ExcelUpload.ippan_chosa_result_view()関数が正常終了しました。', 'INFO')
        response = HttpResponse(content=save_virtual_workbook(wb), content_type='application/vnd.ms-excel')
        response['Content-Disposition'] = 'attachment; filename="ippan_chosa_result2.xlsx"'
        return response
        
    except:
        print_log(sys.exc_info()[0], 'ERROR')
        print_log('[ERROR] P0300ExcelUpload.ippan_chosa_result_view()関数でエラーが発生しました。', 'ERROR')
        print_log('[ERROR] P0300ExcelUpload.ippan_chosa_result_view()関数が異常終了しました。', 'ERROR')
        return render(request, 'error.html')
    